"""
Find library conversion options Huwa to Passim

This version created by Erwin R. Komen
Date: 5/jan/2023

Usage:
    python huwalibnew.py -i "d:/data files/passim/data/lib_huwa_new.xlsx" -o "d:/data files/passim/data/lib_huwa_new.json"

"""

import sys, getopt, os.path, importlib
import os, sys
import csv, json
import requests
import openpyxl, json
from openpyxl.utils.cell import get_column_letter
from openpyxl.cell import Cell
from openpyxl import Workbook

# My own stuff
import utils

# Make available error handling
errHandle = utils.ErrHandle()

# ----------------------------------------------------------------------------------
# Name :    main
# Goal :    Main body of the function
# History:
# 14/jan/2019    ERK Created
# ----------------------------------------------------------------------------------
def main(prgName, argv) :
  flInput = ''        # input file name: JSON with country definitions
  flOutput = ''       # output file name
  bAddRownum = False  # Add row number

  try:
    sSyntax = prgName + '  [-a] -i <passim huwa library Excel> -o <output JSON file>'
    # get all the arguments
    try:
      # Get arguments and options
      opts, args = getopt.getopt(argv, "hai:o:", ["-ifile=", "-ofile"])
    except getopt.GetoptError:
      print(sSyntax)
      sys.exit(2)
    # Walk all the arguments
    for opt, arg in opts:
      if opt in ("-h", "--help"):
        print(sSyntax)
        sys.exit(0)
      elif opt in ("-i", "--ifile"):
        flInput = arg
      elif opt in ("-o", "--ofile"):
        flOutput = arg
      elif opt in ("-a", "--addrownum"):
        bAddRownum = True

    # Check if all arguments are there
    if (flInput == '' or flOutput == ''):
      errHandle.DoError(sSyntax)

    # Continue with the program
    errHandle.Status('Input is "' + flInput + '"')
    errHandle.Status('Output is "' + flOutput + '"')

    # Call the function that does the job
    oArgs = {'input': flInput,
             'addrownum': bAddRownum,
             'output': flOutput}
    if (not passim_huwa_lib_new(oArgs)) :
      errHandle.DoError("Could not complete")
      return False
    
      # All went fine  
    errHandle.Status("Ready")
  except:
    # act
    errHandle.DoError("main")
    return False

# ----------------------------------------------------------------------------------
# Name :    passim_huwa_lib_new
# Goal :    Convert a passim-to-huwa-library correspondence Excel into JSON
# History:
# 05/jan/2023    ERK Created
# ----------------------------------------------------------------------------------
def passim_huwa_lib_new(oArgs):
    """Create library definitions"""

    # Defaults
    flInput = ""
    flOutput = ""
    bAddRowNumber = False
    bDebug = False
    count = 0

    lExpected = ['lcountry', 'lcity', 'library', 'toevoegen', 'opmerkingen', 'approved']
    lField = ['huwacountry', 'huwacity', 'huwalibrary', 'toevoegen', 'convert', 'approved']

    try:
        # Recover the arguments
        if "input" in oArgs: flInput = oArgs["input"]
        if "output" in oArgs: flOutput = oArgs["output"]
        if "addrownum" in oArgs: bAddRowNumber = oArgs['addrownum']

        # Check input file
        if not os.path.isfile(flInput):
            errHandle.Status("Please specify an input FILE")
            return False

        # Prepare entries
        lEntry = []

        # Load the Excel
        wb = openpyxl.load_workbook(flInput, read_only=True)
        ws = wb['Data']

        # Iterate
        bFirst = True
        lHeader = []
        lData = []
        col_lib = None
        col_remark = None
        row_num = 1
        dic_huwapassim = {}
        dic_huwaonly = {}
        for row in ws.iter_rows(min_row=1, min_col=1, max_col=len(lExpected)):
            if bFirst:
                # Expect header
                for idx, cell in enumerate(row):
                    sValue = cell.value.strip("\t")                    
                    sKey = ""
                    for idx, item in enumerate(lExpected):
                        if item == sValue:
                            sKey = lField[idx]
                            break
                    # Check if it's okay
                    if sKey == "":
                        # Cannot read this
                        msg = "Don't understand column header [{}]".format(sValue)
                        errHandle.Status(msg)
                        return False
                    lHeader.append(sKey)
                    if col_lib is None and "library" in sKey.lower():
                        col_lib = idx
                    # Figure out which row contains a remark
                    if col_remark is None and "opmerkingen" in sKey.lower():
                        col_remark = idx
                bFirst = False
                # Safety precautions
                if col_lib is None:
                    col_lib = 1
                if col_remark is None:
                    col_remark = 1
            elif not row[col_remark].value is None and row[col_remark].value != "":
                # Transform the row into an object with keys
                oRow = {}
                for idx, key in enumerate(lHeader):
                    cell = row[idx]
                    cv = cell.value
                    oRow[key] = cv

                # Process the information in the 'convert' cell
                sConvert = oRow.get("convert")
                if not sConvert is None:
                    # We should have a 'passim' section in here
                    oPassim = {}
                    # We have something to convert
                    lPart = [x.strip() for x in sConvert.split(";")]
                    for sItem in lPart:
                        if "city:" in sItem:
                            # It must be a city: strip this string
                            sItem = sItem.replace("city:", "").strip()
                            # Add the item to passim 
                            oPassim['city'] = sItem
                        elif "country:" in sItem:
                            # It must be a country name correction: strip this string
                            sItem = sItem.replace("city:", "").strip()
                            # Add the item to passim 
                            oPassim['country'] = sItem
                        else:
                            # Strip the quotation marks
                            sItem = sItem.strip('"')
                            # Add the item to passim 
                            oPassim['library'] = sItem
                    # Create a correspondence object
                    oCorr = dict(huwa={}, passim=oPassim)
                    oCorr['huwa']['country'] = oRow['huwacountry']
                    oCorr['huwa']['city'] = oRow['huwacity']
                    oCorr['huwa']['library'] = oRow['huwalibrary']
                    # Add this object to the list
                    lData.append(oCorr)

            # For debugging purposes
            row_num += 1
        # Close the workbook
        wb.close()

        # Write the output
        with open(flOutput, "w", encoding="utf-8") as f:
            json.dump(lData, f, indent=2)

        # Return positively
        return True
    except:
        sMsg = errHandle.get_error_message()
        errHandle.DoError("passim_huwa_lib_new")
        return False



# ----------------------------------------------------------------------------------
# Goal :  If user calls this as main, then follow up on it
# ----------------------------------------------------------------------------------
if __name__ == "__main__":
  # Call the main function with two arguments: program name + remainder
  main(sys.argv[0], sys.argv[1:])
