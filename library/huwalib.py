"""
Find library names in cities in particular countries.
Combine the output into a grand list

This version created by Erwin R. Komen
Date: 14/jan/2019

Usage:
    python library.py -i "d:/data files/passim/data/LLTA_Names.xml" -o "d:/data files/passim/data/authors.json"

"""

import sys, getopt, os.path, importlib
import os, sys
import csv, json
import requests
import openpyxl, json
from openpyxl.utils.cell import get_column_letter
from openpyxl.cell import Cell
from openpyxl import Workbook

# My own stuff
import utils

# Make available error handling
errHandle = utils.ErrHandle()

# ----------------------------------------------------------------------------------
# Name :    main
# Goal :    Main body of the function
# History:
# 14/jan/2019    ERK Created
# ----------------------------------------------------------------------------------
def main(prgName, argv) :
  flInput = ''        # input file name: JSON with country definitions
  flOutput = ''       # output file name
  bAddRownum = False  # Add row number

  try:
    sSyntax = prgName + '  [-a] -i <passim huwa library Excel> -o <output JSON file>'
    # get all the arguments
    try:
      # Get arguments and options
      opts, args = getopt.getopt(argv, "hai:o:", ["-ifile=", "-ofile"])
    except getopt.GetoptError:
      print(sSyntax)
      sys.exit(2)
    # Walk all the arguments
    for opt, arg in opts:
      if opt in ("-h", "--help"):
        print(sSyntax)
        sys.exit(0)
      elif opt in ("-i", "--ifile"):
        flInput = arg
      elif opt in ("-o", "--ofile"):
        flOutput = arg
      elif opt in ("-a", "--addrownum"):
        bAddRownum = True

    # Check if all arguments are there
    if (flInput == '' or flOutput == ''):
      errHandle.DoError(sSyntax)

    # Continue with the program
    errHandle.Status('Input is "' + flInput + '"')
    errHandle.Status('Output is "' + flOutput + '"')

    # Call the function that does the job
    oArgs = {'input': flInput,
             'addrownum': bAddRownum,
             'output': flOutput}
    if (not passim_huwa_libraries(oArgs)) :
      errHandle.DoError("Could not complete")
      return False
    
      # All went fine  
    errHandle.Status("Ready")
  except:
    # act
    errHandle.DoError("main")
    return False

# ----------------------------------------------------------------------------------
# Name :    passim_huwa_libraries
# Goal :    Convert the passim-to-huwa-library correspondence Excel into JSON
# History:
# 07/jul/2022    ERK Created
# ----------------------------------------------------------------------------------
def passim_huwa_libraries(oArgs):
    """Create library definitions"""

    # Defaults
    flInput = ""
    flOutput = ""
    bAddRowNumber = False
    bDebug = False
    count = 0

    lExpected = ['id', 'country', 'city', 'library', 'libtype', 'HUWA id']
    lField = ['passim_id', 'country', 'city', 'library', 'libtype', 'huwa_id']

    try:
        # Recover the arguments
        if "input" in oArgs: flInput = oArgs["input"]
        if "output" in oArgs: flOutput = oArgs["output"]
        if "addrownum" in oArgs: bAddRowNumber = oArgs['addrownum']

        # Check input file
        if not os.path.isfile(flInput):
            errHandle.Status("Please specify an input FILE")
            return False

        # Prepare entries
        lEntry = []

        # Load the Excel
        wb = openpyxl.load_workbook(flInput, read_only=True)
        ws = wb.get_active_sheet()

        # Iterate
        bFirst = True
        lHeader = []
        lData = []
        col_lib = None
        row_num = 1
        dic_huwapassim = {}
        dic_huwaonly = {}
        for row in ws.iter_rows(min_row=1, min_col=1, max_col=len(lExpected)):
            if bFirst:
                # Expect header
                for idx, cell in enumerate(row):
                    sValue = cell.value.strip("\t")                    
                    sKey = ""
                    for idx, item in enumerate(lExpected):
                        if item == sValue:
                            sKey = lField[idx]
                            break
                    # Check if it's okay
                    if sKey == "":
                        # Cannot read this
                        msg = "Don't understand column header [{}]".format(sValue)
                        errHandle.Status(msg)
                        return False
                    lHeader.append(sKey)
                    if col_lib is None and "library" in sKey.lower():
                        col_lib = idx
                bFirst = False
                if col_lib is None:
                    col_lib = 1
            elif row[col_lib].value != None:
                oRow = {}
                for idx, key in enumerate(lHeader):
                    cell = row[idx]
                    cv = cell.value
                    oRow[key] = cv

                # Add the correspondence - provided there are the id's
                passim_id = oRow.get('passim_id')
                huwa_id = oRow.get("huwa_id")

                if passim_id is None or passim_id == "":
                    if not huwa_id is None:
                        # Should add this to the huwa only dictionary
                        dic_huwaonly[str(huwa_id)] = oRow
                elif not huwa_id is None:
                    if isinstance(passim_id, str):
                        passim_id = int(passim_id)
                    # Double check for multiple HUWA id's
                    if isinstance(huwa_id, str) and "," in huwa_id:
                        huwa_ids = json.loads("[{}]".format(huwa_id))
                        oRow['huwa_id'] = huwa_ids
                        for huwa_id in huwa_ids:
                            dic_huwapassim[str(huwa_id)] = passim_id
                    else:
                        oRow['huwa_id'] = [ huwa_id ]
                        dic_huwapassim[str(huwa_id)] = passim_id

                if bAddRowNumber:
                    # Also add the row number (as string)
                    oRow['row_number'] = "{}".format(row_num)
                lData.append(oRow)
            else:
                x = 2
            if bDebug:
                # Show the library value
                errHandle.Status("Row {} has library: [{}]".format(row_num, row[col_lib].value))
            # For debugging purposes
            row_num += 1
        # Close the workbook
        wb.close()

        oData = dict(huwapassim=dic_huwapassim, huwaonly=dic_huwaonly, libraries=lData)

        # Write the output
        with open(flOutput, "w", encoding="utf-8") as f:
            json.dump(oData, f, indent=2)

        # Return positively
        return True
    except:
        sMsg = errHandle.get_error_message()
        errHandle.DoError("passim_huwa_libraries")
        return False



# ----------------------------------------------------------------------------------
# Goal :  If user calls this as main, then follow up on it
# ----------------------------------------------------------------------------------
if __name__ == "__main__":
  # Call the main function with two arguments: program name + remainder
  main(sys.argv[0], sys.argv[1:])
