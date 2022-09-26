"""
Read 'new' libraries from a HUWA json file 

This version created by Erwin R. Komen
Date: 26/sep/2022

Usage:
    python libhuwa.py -i "d:/Data Files/TG/Passim/Data/huwa/passim_huwa_manu_new.json" 
                      -o "d:/Data Files/TG/Passim/Data/huwa/lib_huwa_new.json"

"""

import os, sys, getopt, os.path, importlib
import json
import openpyxl
from openpyxl import Workbook
from openpyxl.cell import Cell
# This part depends on teh specific version of OpenPyXL
try:
    from openpyxl.utils.cell import get_column_letter
except:
    from openpyxl.utils import get_column_letter

# My own stuff
import utils

# Make available error handling
errHandle = utils.ErrHandle()


# ----------------------------------------------------------------------------------
# Name :    main
# Goal :    Main body of the function
# History:
# 14/jan/2019    ERK Created
# ----------------------------------------------------------------------------------
def main(prgName, argv) :
  flInput = ''        # input file name: JSON with country definitions
  flOutput = ''       # output file name

  try:
    sSyntax = prgName + ' -i <input json> -o <output json>'
    # get all the arguments
    try:
      # Get arguments and options
      opts, args = getopt.getopt(argv, "hi:o:", ["-ifile=", "-ofile"])
    except getopt.GetoptError:
      print(sSyntax)
      sys.exit(2)
    # Walk all the arguments
    for opt, arg in opts:
      if opt in ("-h", "--help"):
        print(sSyntax)
        sys.exit(0)
      elif opt in ("-i", "--ifile"):
        flInput = arg
      elif opt in ("-o", "--ofile"):
        flOutput = arg
    # Check if all arguments are there
    if (flInput == '' or flOutput == ''):
      errHandle.DoError(sSyntax)

    # Continue with the program
    errHandle.Status('Input is "' + flInput + '"')
    errHandle.Status('Output is "' + flOutput + '"')

    # Call the function that does the job
    oArgs = {'input': flInput,
             'output': flOutput}
    if (not huwa_libraries(oArgs)) :
      errHandle.DoError("Could not complete")
      return False
    
      # All went fine  
    errHandle.Status("Ready")
  except:
    # act
    errHandle.DoError("main")
    return False

# ----------------------------------------------------------------------------------
# Name :    huwa_libraries
# Goal :    Extract new PASSIM libraries from HUWA JSON
# History:
# 26/sep/2022    ERK Created
# ----------------------------------------------------------------------------------
def huwa_libraries(oArgs):
    """Create library definitions"""

    # Defaults
    flInput = ""
    flOutput = ""
    lst_library = []
    data = None

    count = 0

    try:
        # Recover the arguments
        if "input" in oArgs: flInput = oArgs["input"]
        if "output" in oArgs: flOutput = oArgs["output"]

        # Check input file
        if not os.path.isfile(flInput):
            errHandle.Status("Please specify an input FILE")
            return False

        # Check output file name
        if flOutput.strip() == "":
            errHandle.Status("Specify a valid output file")
            return False

        # Make an Excel output file name
        flOutExcel = flOutput.replace(".json", ".xlsx")

        # Read the file
        with open(flInput, "r") as f:
            data = json.load(f)

        if not data is None:
            for oItem in data:
                library_id = oItem.get("library_id")
                library = oItem.get('library')
                if library_id is None and not library is None: # and library != "":
                    lcity = oItem.get('lcity')
                    lcountry = oItem.get('lcountry')
                    bFound = False
                    for oLib in lst_library:
                        if oLib['library'] == library and oLib['lcountry'] == lcountry and oLib['lcity'] == lcity:
                            bFound = True
                            break
                    if not bFound:
                        oNewLib = dict(library=library, lcity=lcity, lcountry=lcountry)
                        lst_library.append(oNewLib)


        # Write the list
        with open(flOutput, "w", encoding="utf-8") as f:
            json.dump(lst_library, f, indent=2)

        # Store the output in excel
        data_to_excel(lst_library, flOutExcel) 

        # Return positively
        return True
    except:
        sMsg = errHandle.get_error_message()
        errHandle.DoError("huwa_libraries")
        return False

def data_to_excel(lst_data, flOutput):
    """Store the data in the output file as Excel"""

    try:
        # Start workbook
        wb = openpyxl.Workbook()

        # Getting the active sheet is version-dependant...
        try:
            ws = wb.get_active_sheet()
        except:
            ws = wb.active
        # GIve a good title to the active sheet
        ws.title="Data"

        # Set the headers
        headers = ['lcountry', 'lcity', 'library']
        for col_num in range(len(headers)):
            c = ws.cell(row=1, column=col_num+1)
            c.value = headers[col_num]
            c.font = openpyxl.styles.Font(bold=True)
            # Set width to a fixed size
            ws.column_dimensions[get_column_letter(col_num+1)].width = 5.0        

        row_num = 1
        lCsv = []
        for row in lst_data:
            # Keep track of the EXCEL row we are in
            row_num += 1
            # Walk the elements in the data row
            for idx, k in enumerate(headers):
                c = ws.cell(row=row_num, column=idx+1)
                c.value = row[k]
                c.alignment = openpyxl.styles.Alignment(wrap_text=False)

        # Save the result in the file
        wb.save(flOutput)
        
        return True
    except:
        sMsg = get_error_message()
        return False



# ----------------------------------------------------------------------------------
# Goal :  If user calls this as main, then follow up on it
# ----------------------------------------------------------------------------------
if __name__ == "__main__":
  # Call the main function with two arguments: program name + remainder
  main(sys.argv[0], sys.argv[1:])
