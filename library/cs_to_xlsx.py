"""
Convert corpussearch output to an excel with results divided in columns

This version created by Erwin R. Komen
Date: 11/nov/2020
"""

import sys, getopt, os.path, importlib
import re
import openpyxl
from openpyxl import Workbook
from openpyxl.cell import Cell
# This part depends on teh specific version of OpenPyXL
try:
    from openpyxl.utils.cell import get_column_letter
except:
    from openpyxl.utils import get_column_letter
# =========

POS_DICT = {"NP-OB1": "O", "VMFIN": "Aux", "VVINF": "V", "VAFIN": "Aux", "VVPP": "V"}


# ================== General helper functions =============================
def get_error_message():
    arInfo = sys.exc_info()
    if len(arInfo) == 3:
        sMsg = str(arInfo[1])
        if arInfo[2] != None:
            sMsg += " at line " + str(arInfo[2].tb_lineno)
        return sMsg
    else:
        return ""

def Status( msg):
    # Just print the message
    print(msg, file=sys.stderr)

def DoError(msg, bExit = False):
    # get the message
    sErr = get_error_message()
    # Print the error message for the user
    print("Error: {}\nSystem:{}".format(msg, sErr), file=sys.stderr)
    # Is this a fatal error that requires exiting?
    if (bExit):
        sys.exit(2)
    # Otherwise: return the string that has been made
    return msg


# ----------------------------------------------------------------------------------
# Name :    main
# Goal :    Main body of the function
# History:
# 31/jan/2019    ERK Created
# ----------------------------------------------------------------------------------
def main(prgName, argv) :
    flInput = ''        # input file name: XML with author definitions
    flOutput = ''       # output file name

    try:
        sSyntax = prgName + ' -i <CorpusSearch results file> -o <Excel output file name>'
        # get all the arguments
        try:
            # Get arguments and options
            opts, args = getopt.getopt(argv, "hi:o:", ["-ifile=", "-ofile"])
        except getopt.GetoptError:
            print(sSyntax)
            sys.exit(2)
        # Walk all the arguments
        for opt, arg in opts:
            if opt in ("-h", "--help"):
                print(sSyntax)
                sys.exit(0)
            elif opt in ("-i", "--ifile"):
                flInput = arg
            elif opt in ("-o", "--ofile"):
                flOutput = arg
        # Check if all arguments are there
        if (flInput == '' or flOutput == ''):
            DoError(sSyntax)

        # Continue with the program
        Status('Input is "' + flInput + '"')
        Status('Output is "' + flOutput + '"')

        oArgs = dict(input=flInput, output=flOutput)

        # Call the function that actually reads the CorpusSearch file
        if not read_corpussearchresults(oArgs):
            DoError("Could not complete reading CorpusSearch results", True)

        Status("Ready")
    except:
        DoError("main")


# ----------------------------------------------------------------------------------
# Name :    read_authors
# Goal :    Read the authors from the XML intput file
# History:
# 31/jan/2019    ERK Created
# ----------------------------------------------------------------------------------
def read_corpussearchresults(oArgs):
    """Read the CorpusSearch results and convert into Excel"""

    # Defaults
    flInput = ""
    flOutput = ""

    try:
        # Recover the arguments
        if "input" in oArgs: flInput = oArgs["input"]
        if "output" in oArgs: flOutput = oArgs["output"]

        # Check input file
        if not os.path.isfile(flInput):
            errHandle.Status("Please specify an input FILE")
            return False
        
        # Read the text file into an array
        with open(flInput, "r", encoding="utf-8") as fd:
            lst_input = fd.readlines()

        # Initialize flags
        bInSentence = False
        bNeedStructure = False
        bInStructure = False
        bNeedId = False
        bHaveId = False

        # Initialize storages
        lst_structure = []
        lst_sentence = []
        lst_result = []

        # Walk all lines
        for line in lst_input:
            # Debugging
            # Status(line)

            # Look for the start of a new item to process
            if '/~*' in line:
                # Start processing from here
                bInSentence = True
                bInStructure = False
                bHaveId = False
                bNeedStructure = True
                bNeedId = False
                lst_structure = []
                lst_sentence = []
            elif '*~/' in line:
                bInSentence = False
                bNeedStructure = True
            elif bNeedStructure and '/*' in line:
                bInStructure = True
            elif bNeedStructure and '*/' in line:
                bInStructure = False
                bNeedId = True
            else:
                # See where we are and what we can do with this string
                if bInSentence:
                    # Add line to sentence
                    lst_sentence.append(line)
                elif bInStructure:
                    # Add line to structure
                    lst_structure.append(line)

                # elif bNeedId and "(ID " in line:
                elif bNeedId and re.match(r'.*\(([0-9]+\s+)?ID\s+.*',line):
                    # Extract the ID from this line
                    match = re.match(r'.*(\(([0-9]+\s+)?ID\s)([a-zA-Z0-9_\+\.\:\,]+)(\)).*', line)
                    if match:
                        count = len(match.groups())
                        groupnum = 3 if count > 3 else 2
                        str_id = match.group(groupnum)

                        # When we have the ID, we can process this result
                        structure = " ".join(lst_structure)
                        order = get_order(structure)
                        oResult = dict(
                            sentence=" ".join(lst_sentence),
                            structure=structure,
                            id=str_id,
                            order=order
                            )
                        lst_result.append(oResult)

        # Store the output in excel
        data_to_excel(lst_result, flOutput) 

        # Return positively
        return True
    except:
        sMsg = get_error_message()
        DoError("read_corpussearchresults")
        return False

def get_order(structure):
    """Given a corpussearch output structure, determine the order"""

    order = None
    try:
        # Get the part after the colon
        if ':' in structure:
            structure = structure.split(":")[1].strip()
        # Put into an array
        lst_part = structure.split(",")
        # Divide into more meaningful items
        lst_item = []
        for str_item in lst_part:
            num_str_array = str_item.strip().split(" ")
            lst_item.append(dict(position=int(num_str_array[0]), label=num_str_array[1]))

        # Sort the [lst_item] by position
        lst_sorted = sorted(lst_item, key=lambda x: x['position'])
        lst_label = []
        for lst_item in lst_sorted:
            item = lst_item['label']
            if item in POS_DICT:
                lst_label.append(POS_DICT[item])

        # Get the order
        order = "-".join(lst_label)
    except:
        sMsg = get_error_message()
        DoError("read_corpussearchresults")
        order = None

    # Return the order
    return order

def data_to_excel(lst_data, flOutput):
    """Store the data in the output file as Excel"""

    try:
        # Start workbook
        wb = openpyxl.Workbook()

        # Getting the active sheet is version-dependant...
        try:
            ws = wb.get_active_sheet()
        except:
            ws = wb.active
        # GIve a good title to the active sheet
        ws.title="Data"

        # Set the headers
        headers = ['sentence', 'structure', 'id', 'order']
        for col_num in range(len(headers)):
            c = ws.cell(row=1, column=col_num+1)
            c.value = headers[col_num]
            c.font = openpyxl.styles.Font(bold=True)
            # Set width to a fixed size
            ws.column_dimensions[get_column_letter(col_num+1)].width = 5.0        

        row_num = 1
        lCsv = []
        for row in lst_data:
            # Keep track of the EXCEL row we are in
            row_num += 1
            # Walk the elements in the data row
            for idx, k in enumerate(headers):
                c = ws.cell(row=row_num, column=idx+1)
                c.value = row[k]
                c.alignment = openpyxl.styles.Alignment(wrap_text=False)

        # Save the result in the file
        wb.save(flOutput)
        
        return True
    except:
        sMsg = get_error_message()
        return False


# ----------------------------------------------------------------------------------
# Goal :  If user calls this as main, then follow up on it
# ----------------------------------------------------------------------------------
if __name__ == "__main__":
    # Call the main function with two arguments: program name + remainder
    main(sys.argv[0], sys.argv[1:])
