"""Models for the DCT app: dynamic comparison tables

"""
import enum
from django.apps.config import AppConfig
from django.apps import apps
from django.db import models, transaction
from django.contrib.auth.models import User, Group
from django.db.models import Q
from django.db.models.functions import Lower
from django.db.models.query import QuerySet 
from django.urls import reverse
import os

from markdown import markdown
import json, copy

import openpyxl
from openpyxl.utils.cell import get_column_letter
from openpyxl.cell import Cell
from openpyxl import Workbook

# Take from my own app
from passim.utils import ErrHandle
from passim.settings import TIME_ZONE, MEDIA_ROOT
from passim.basic.models import UserSearch
from passim.basic.views import base64_decode, base64_encode
from passim.seeker.models import Author, Keyword, get_current_datetime, get_crpp_date, build_abbr_list, COLLECTION_SCOPE, \
    Collection, Manuscript, Profile, CollectionSuper, Signature, SermonDescrKeyword, \
    SermonDescr, EqualGold, Feast, Project2
from passim.reader.excel import ManuscriptUploadExcel

STANDARD_LENGTH=255
ABBR_LENGTH = 5

SETLIST_TYPE = "dct.setlisttype"
SAVEDITEM_TYPE = "dct.saveditemtype"
SELITEM_TYPE = "dct.selitemtype"
IMPORT_TYPE = "dct.importtype"
IMPORT_STATUS = "dct.importstatus"

def get_passimcode(super_id, super_code):
    code = super_code if super_code and super_code != "" else "(nocode_{})".format(super_id)
    return code

def get_goldsig_dct(super_id):
    """Get the best signature according to DCT rules"""

    sBack = ""
    first = None
    editype_preferences = ['gr', 'cl', 'ot']
    for editype in editype_preferences:
        siglist = Signature.objects.filter(gold__equal_id=super_id, editype=editype).order_by('code').values('code')
        if len(siglist) > 0:
            code = siglist[0]['code']
            ellipsis = "" if len(siglist) == 1 else "..."
            sBack = "{}: {}{}".format(editype, code, ellipsis)
            break
    return sBack

def get_goldsiglist_dct(super_id):
    """Get the list of signature according to DCT rules"""

    lBack = []
    first = None
    editype_preferences = ['gr', 'cl', 'ot']
    siglist = []
    for editype in editype_preferences:
        siglist = Signature.objects.filter(gold__equal_id=super_id, editype=editype).order_by('code').values('code')
        for item in siglist:
            sSig = "{}: {}".format(editype, item['code'])
            lBack.append(sSig)
    return lBack

def get_list_matches(oPMlist, oSsgList):
    """Calculate the number of matches between the two lists"""

    matches = 0
    for oPm in oPMlist['ssglist']:
        ssg = oPm['super']
        # Check how manu times this SSG appears in [oSsgList]
        for oSsg in oSsgList['ssglist']:
            if ssg == oSsg['super']:
                matches += 1
    return matches

def import_path(instance, filename):
    # def import_path(sType, instance, filename):
    """Upload Excel file to the right place, and remove old file if existing
    
    This function is used within the model ImportSet
    NOTE: this must be the relative path w.r.t. MEDIA_ROOT
    """

    oErr = ErrHandle()
    sBack = ""
    sSubdir = "import"
    try:
        # Adapt the filename for storage
        # sAdapted = "{}_{:08d}_{}".format(sType, instance.id, filename.replace(" ", "_"))
        sAdapted = "{:08d}_{}".format(instance.id, filename.replace(" ", "_"))

        # The stuff that we return
        sBack = os.path.join(sSubdir, sAdapted)

        # Add the subdir (defined above)
        fullsubdir = os.path.abspath(os.path.join(MEDIA_ROOT, sSubdir))
        if not os.path.exists(fullsubdir):
            os.makedirs(fullsubdir)

        # Add the actual filename to form an absolute path
        sAbsPath = os.path.abspath(os.path.join(fullsubdir, sAdapted))

        # Also get the bare file name
        sBare = os.path.basename(filename)
        # Store it in the item
        instance.name = sBare

        if os.path.exists(sAbsPath):
            # Remove it
            try:
                os.remove(sAbsPath)
            except:
                oErr.Status("Could not remove file now: {}".format(sAbsPath))

    except:
        msg = oErr.get_error_message()
        oErr.DoError("import_path")
    return sBack

def excel_import_path(instance, filename):
    return import_path(instance, filename)


# ====================== Models needed to work on DCTs ===============================================



class ResearchSet(models.Model):
    """One research set
    
    A research set is a user-curated collection of source-lists.
    It can be used as the basis for creating a DCT, a dynamic comparison table
    """

    # [1] obligatory name
    name = models.CharField("Name", max_length=STANDARD_LENGTH)
    # [1] a research set belongs to a particular user's profilee
    profile = models.ForeignKey(Profile, on_delete=models.CASCADE, related_name="profile_researchsets")

    # [0-1] Optional notes for this set
    notes = models.TextField("Notes", blank=True, null=True)

    # [1] A list of all the DCT parameters for this DCT
    contents = models.TextField("Contents", default="[]")

    # [1] The scope of this collection: who can view it?
    #     E.g: private, team, global - default is 'private'
    scope = models.CharField("Scope", choices=build_abbr_list(COLLECTION_SCOPE), default="priv", max_length=5)

    # [1] And a date: the date of saving this manuscript
    created = models.DateTimeField(default=get_current_datetime)
    saved = models.DateTimeField(null=True, blank=True)

    def __str__(self):
        """Just return the name"""
        return self.name

    def save(self, force_insert = False, force_update = False, using = None, update_fields = None):
        # Adapt the save date
        self.saved = get_current_datetime()
        response = super(ResearchSet, self).save(force_insert, force_update, using, update_fields)

        # If there is no DCT definition yet, create a default one
        SetDef.check_default(self, self.profile.user.username)

        # Return the response when saving
        return response

    def adapt_contents(manu=None, coll=None):
        """Update research set and setlist that involve manuscript [manu] or collection [coll]"""

        oErr = ErrHandle()
        try:
            rset_list = []
            qs = None
            # Decide which one to take
            if not manu is None:
                qs = SetList.objects.filter(manuscript_id=manu.id)
            elif not coll is None:
                qs = SetList.objects.filter(collection_id=coll.id)
            if not qs is None:
                # Look for any setlist having this one
                for setlist in qs:
                    # Add the researchset to the list
                    rset = setlist.researchset
                    if not rset.id in rset_list:
                        rset_list.append(rset.id)
                    # Update this setlist
                    setlist.calculate_contents()
                # Update research sets
                for rset in ResearchSet.objects.filter(id__in=rset_list):
                    # Update with force
                    rset.update_ssglists(True)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("ResearchSet/adapt_contents")
        return None

    def adapt_from_setlist(self, setlist):
        """Adapt, starting from this setlist"""

        oErr = ErrHandle()
        try:
            # Double check argument
            if not setlist is None:
                # Update this setlist
                setlist.calculate_contents()

                # Update all items in the current researchset
                self.update_ssglists()
        except:
            msg = oErr.get_error_message()
            oErr.DoError("ResearchSet/adapt_from_setlist")
        return None

    def adapt_order(self):
        """Re-calculate the order and adapt where needed"""

        qs = self.researchset_setlists.all().order_by("order")
        order = 1
        with transaction.atomic():
            # Walk all the SetList objects
            for obj in qs:
                # Check if the order is as it should be
                if obj.order != order:
                    #No: adapt the order
                    obj.order = order
                    # And save it
                    obj.save()
                # Keep track of how the order should be
                order += 1
        return None

    def add_list(self, obj, setlisttype):
        """Add a list of SSGs (either Manuscript or Collection) to the research set"""

        oErr = ErrHandle()
        bBack = True
        try:
            # Get the current size of the research set
            iCount = self.researchset_setlists.count()
            order = iCount + 1
            setlist = None

            # Action depends on the type of addition
            if setlisttype == "manu":   # Add a manuscript
                setlist = SetList.objects.filter(researchset=self, setlisttype=setlisttype, manuscript=obj).first()
                if setlist is None:
                    setlist = SetList.objects.create(
                        researchset=self, order=order, setlisttype=setlisttype,
                        manuscript=obj, name="Added via add_list")
                pass
            elif setlisttype == "ssgd":   # Add a hc or pd
                setlist = SetList.objects.filter(researchset=self, setlisttype=setlisttype, collection=obj).first()
                if setlist is None:
                    setlist = SetList.objects.create(
                        researchset=self, order=order, setlisttype=setlisttype,
                        collection=obj, name="Added via add_list")

            # If need be, calculate the contents
            if not setlist is None:
                self.update_ssglists()

        except:
            msg = oErr.get_error_message()
            oErr.DoError("ResearchSet/add_list")
            bBack = False
        # Return the status
        return bBack

    def calculate_matches(self, ssglists):
        """Calculate the number of pm-matches for each list from ssglists"""

        oErr = ErrHandle()
        lBack = []
        try:
            # Preparation: create a list of SSG ids per list
            for oItem in ssglists:
                oItem['ssgid'] = [x['super'] for x in oItem['ssglist']]
                oItem['unique_matches'] = 0

            # Calculate the number of matches for each SSglist
            for idx_list in range(len(ssglists)):
                # Take this as the possible pivot list
                lPivot = ssglists[idx_list]['ssgid']

                # Walk all lists that are not this pivot and count matches
                lUnique = []
                oMatches = {}
                if 'matchset' in ssglists[idx_list]['title']:
                    oMatches = ssglists[idx_list]['title']['matchset']
                for idx, setlist in enumerate(ssglists):
                    if idx != idx_list:
                        # Get this ssglist
                        lSsgId = setlist['ssgid']

                        # Start calculating the number of matches this list has with the currently suggested PM
                        sKey = str(ssglists[idx]['title']['order'])
                        lMatches = []

                        # Consider all ssg id's in this list
                        for ssg_id in lSsgId:
                            # Global unique matches
                            if ssg_id in lPivot and not ssg_id in lUnique:
                                lUnique.append(ssg_id)
                            # Calculation with respect to the currently suggested PM
                            if ssg_id in lPivot: # and not ssg_id in lMatches:
                                lMatches.append(ssg_id)
                        # Keep track of the matches (for sorting)
                        oMatches[sKey] = len(lMatches)
                # Store the between-list matches
                ssglists[idx_list]['title']['matchset'] = oMatches

                # Store the number of matches for this list
                unique_matches = len(lUnique)
                ssglists[idx_list]['unique_matches'] = unique_matches

            # What we return
            lBack = ssglists
        except:
            msg = oErr.get_error_message()
            oErr.DoError("ResearchSet/calculate_matches")
            lBack = None
        # Return the status
        return lBack

    def calculate_pm(self):
        """Calculate the Pivot Manuscript/item for this research set"""

        oErr = ErrHandle()
        iBack = -1
        try:
            # Get the lists that we have made
            ssglists = self.get_ssglists()

            # Calculate the matches
            ssglists = self.calculate_matches(ssglists)

            # Figure out what the first best list is
            idx_pm = -1
            max_matches = -1
            min_order = len(ssglists) + 2
            min_year_start = 3000
            min_year_finish = 3000
            for idx, oListItem in enumerate(ssglists):
                unique_matches = oListItem['unique_matches']
                year_start = oListItem['title']['yearstart']
                year_finish = oListItem['title']['yearfinish']
                order = oListItem['title']['order']

                # Check which is the best so far
                bTakeThis = False
                bTakeThis = (unique_matches > max_matches)
                if not bTakeThis and unique_matches == max_matches:
                    bTakeThis = (year_start < min_year_start)
                    if not bTakeThis and year_start == min_year_start:
                        bTakeThis = (year_finish < min_year_finish)
                        if not bTakeThis and year_finish == min_year_finish:
                            bTakeThis = (order < min_order)

                # Adapt if this is the one
                if bTakeThis:
                    max_matches = unique_matches
                    min_year_start = year_start
                    min_year_finish = year_finish
                    min_order = order
                    idx_pm = idx

            # What we return is the 'order' value of the best matching list
            iBack = ssglists[idx_pm]['title']['order']

        except:
            msg = oErr.get_error_message()
            oErr.DoError("ResearchSet/calculate_pm")
            iBack = -1
        # Return the PM that we have found
        return iBack

    def get_created(self):
        """REturn the creation date in a readable form"""

        # sDate = self.created.strftime("%d/%b/%Y %H:%M")
        sDate = get_crpp_date(self.created, True)
        return sDate

    def get_notes_html(self):
        """Convert the markdown notes"""

        sNotes = "-"
        if self.notes != None:
            sNotes = markdown(self.notes)
        return sNotes

    def get_saved(self):
        """REturn the saved date in a readable form"""

        # sDate = self.saved.strftime("%d/%b/%Y %H:%M")
        sDate = get_crpp_date(self.saved, True)
        return sDate

    def get_size_markdown(self):
        """Get a markdown representation of the size of this set"""

        size = 0
        lHtml = []
        size = self.researchset_setlists.count()
        if size > 0:
            # Create a display for this topic
            lHtml.append("<span class='badge signature gr'>{}</span>".format(size))
        else:
            lHtml.append("0")
        sBack = ", ".join(lHtml)
        return sBack

    def get_ssglists(self, recalculate=False):
        """Get the set of lists for this particular ResearchSet"""

        oErr = ErrHandle()
        lBack = None
        try:
            if self.contents != "" and self.contents[0] == "[":
                lst_contents = json.loads(self.contents)
            else:
                lst_contents = []
            if not recalculate and len(lst_contents) > 0:
                # Should the unique_matches be re-calculated?
                if not 'unique_matches' in lst_contents[0]:
                    # Yes, re-calculate
                    lst_contents = self.calculate_matches(lst_contents)
                    # And make sure this gets saved!
                    self.contents = json.dumps(lst_contents)
                    self.save()
                lBack = lst_contents
            else:
                # Re-calculate the lists
                self.update_ssglists()
                # Get the result
                lBack = json.loads(self.contents)

        except:
            msg = oErr.get_error_message()
            oErr.DoError("ResearchSet/get_ssglists")
        return lBack

    def update_ssglists(self, force = False):
        """Re-calculate the set of lists for this particular ResearchSet"""

        oErr = ErrHandle()
        bResult = True
        lst_ssglists = []
        try:
            oPMlist = None
            # Double check and remove setlists of collection or manuscript that has been removed
            delete_setlist = []
            for oItem in self.researchset_setlists.all().values("manuscript__id", "collection__id", "id"):
                if oItem['manuscript__id'] is None and oItem['collection__id'] is None:
                    delete_setlist.append(oItem['id'])
            if len(delete_setlist) > 0:
                SetList.objects.filter(id__in=delete_setlist).delete()

            # Get the lists of SSGs for each list in the set
            for idx, setlist in enumerate(self.researchset_setlists.all().order_by('order')):
                # Check for the contents
                if force or setlist.contents == "" or len(setlist.contents) < 3 or setlist.contents[0] == "[":
                    setlist.calculate_contents()

                # Retrieve the SSG-list from the contents
                oSsgList = json.loads(setlist.contents)

                # If this is not the *first* setlist, calculate the number of matches with the first
                if idx == 0:
                    oPMlist = copy.copy(oSsgList)
                else:
                    # Calculate the matches
                    oSsgList['title']['matches'] = get_list_matches(oPMlist, oSsgList)
                # Always pass on the default order
                oSsgList['title']['order'] = idx + 1

                # Add the list object to the list
                lst_ssglists.append(oSsgList)

            # Calculate the unique_matches for each list
            lst_ssglists = self.calculate_matches(lst_ssglists)

            # Put it in the ResearchSet and save it
            self.contents = json.dumps(lst_ssglists)
            self.save()

            # All related SetDef items should be warned
            # with transaction.atomic():
            for obj in SetDef.objects.filter(researchset=self.id):
                contents = json.loads(obj.contents)
                contents['recalc'] = True
                obj.contents = json.dumps(contents)
                obj.save()

            # Return this list of lists
            bResult = True
        except:
            msg = oErr.get_error_message()
            oErr.DoError("ResearchSet/update_ssglists")
            bResult = False
        return bResult


class SetList(models.Model):
    """A setlist belongs to a research set"""

    # [1] A setlist just belongs to a research set
    researchset = models.ForeignKey(ResearchSet, on_delete=models.CASCADE, related_name="researchset_setlists")
    # [1] The lists must be ordered (and they can be re-ordered by the user)
    order = models.IntegerField("Order", default=0)
    # [1] Each setlist must be of a particular type
    setlisttype = models.CharField("Setlist type", choices=build_abbr_list(SETLIST_TYPE), max_length=5)

    # [0-1] A user-defined name, if this is a SSGD type
    name = models.CharField("Setlist name", max_length=STANDARD_LENGTH, blank=True, null=True)

    # [1] For convenience and faster operation: keep a JSON list of the SSGs in this setlist
    contents = models.TextField("Contents", default="{}")

    # Depending on the type of setlist, there is a pointer to the actual list of SSGs
    # [0-1] Manuscript pointer
    manuscript = models.ForeignKey(Manuscript, blank=True, null=True, on_delete=models.SET_NULL, related_name="manuscript_setlists")
    # [0-1] Collection pointer (that can be HC, public or personal collection
    collection = models.ForeignKey(Collection, blank=True, null=True, on_delete=models.SET_NULL, related_name="collection_setlists")

    def __str__(self):
        """Combine the name and the order"""
        sBack = "{}: {}".format(self.researchset.name, self.order)
        return sBack

    def adapt_rset(self, rset_type = None):
        oErr = ErrHandle()
        try:
            # Adapt the research set which I am part of
            rset = self.researchset
            if not rset is None:
                rset_type = "-" if rset_type is None else rset_type
                # Show what happens
                oErr.Status("adapt_rset on setlist id={} rset_type={}".format(self.id, rset_type))
                # Adapt from the setlist
                rset.adapt_from_setlist(self)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("adapt_rset")
        return None

    def calculate_contents(self):
        oErr = ErrHandle()
        bResult = True
        try:
            oSsgList = {}
            # Only calculate contents, if there is any
            if not self.collection is None or not self.manuscript is None:
                # Add the name object for this list
                oSsgList['title'] = self.get_title_object()
                # Get the list of SSGs for this list
                oSsgList['ssglist'] = self.get_ssg_list()
            # Add this contents and save myself
            self.contents = json.dumps(oSsgList)
            self.save()
        except:
            msg = oErr.get_error_message()
            oErr.DoError("SetList/calculate_contents")
        return bResult

    def get_ssg_list(self):
        """Create a list of SSGs,depending on the type I am"""

        oErr = ErrHandle()
        lBack = None
        collection_types = ['hist', 'ssgd']
        try:
            if self.setlisttype == "manu":
                # Create a list of SSG's from the manuscript
                # WRONG ARGUMENT: lBack = self.ssgs_manuscript(self.manuscript)
                lBack = self.ssgs_manuscript()
            elif self.setlisttype in collection_types:
                # Wrong argument: lBack = self.ssgs_collection(self.collection)
                lBack = self.ssgs_collection()
        except:
            msg = oErr.get_error_message()
            oErr.DoError("SetList/get_ssg_list")
        return lBack

    def get_title_object(self):
        """Depending on the setlisttype, this provides the title of the setlist
        
        The title is stored in an object, to facilitate template rendering
        """

        # issue #713: changed into 0-0
        oBack = {"main": "", "size": 0, "yearstart": 0, "yearfinish": 0, "matches": 0}
        oErr = ErrHandle()
        try:
            if self.setlisttype == "manu" and not self.manuscript is None:          # SSGs via Manuscript > sermons > SSG links
                # This is a manuscript
                oBack = self.manuscript.get_full_name_html(field1="top", field2="middle", field3="main")
                oBack['size'] = self.manuscript.get_sermon_count()
                oBack['url'] = reverse('manuscript_details', kwargs={'pk': self.manuscript.id})
                oBack['yearstart'] = self.manuscript.yearstart
                oBack['yearfinish'] = self.manuscript.yearfinish
            elif self.setlisttype == "hist" and not self.collection is None:        # Historical collection (of SSGs)
                oBack['top'] = "hc"
                oBack['main'] = self.collection.name
                oBack['size'] = self.collection.freqsuper()
                oBack['url'] = reverse('collhist_details', kwargs={'pk': self.collection.id})   # Historical collection
            elif self.setlisttype == "ssgd" and not self.collection is None:        # Personal/public dataset (of SSGs!!!)
                # Personal collection
                oBack['top'] = "pd"
                if self.name == None or self.name == "":
                    oBack['main'] = self.collection.name
                else:
                    oBack['main'] = self.name
                oBack['size'] = self.collection.freqsuper()
                oBack['url'] = reverse('collpriv_details', kwargs={'pk': self.collection.id})
            else:
                # No idea what this is
                oBack['top'] = "UNKNOWN"
                oBack['main'] = self.setlisttype
        except:
            msg = oErr.get_error_message()
            oErr.DoError("SetList/get_title_object")
        # Return the result
        return oBack

    def ssgs_collection(self, bDebug = False):
        """Get the ordered list of SSGs in the [super] type collection"""

        oErr = ErrHandle()
        lBack = None
        try:
            coll = self.collection
            # Get the list of SSGs inside this collection
            # Per issue #402 additional info needs to be there:
            #   author, 
            #   PASSIM code, 
            #   incipit, explicit, HC (if the SSG is part of a historical collection), 
            #   number of SGs contained in the equality set, 
            #   number of links to other SSGs
            # HC/PD specific:
            #   name
            #   description
            #   size
            # HC-specific:
            #   literature
            qs = CollectionSuper.objects.filter(collection=coll).order_by('order').values(
                'order', 'collection__name', 'collection__descrip', 'collection__settype', 
                'super', 'super__code', 'super__author__name',
                'super__incipit', 'super__explicit', 'super__sgcount', 'super__ssgcount')
            lBack = []
            with transaction.atomic():
                for obj in qs:
                    # Get the order number
                    order = obj['order']
                    # Get the collection-specific information
                    name = obj['collection__name']
                    descr = obj['collection__descrip']
                    settype = obj['collection__settype']
                    # Get the SSG
                    super = obj['super']
                    code = obj['super__code']
                    authorname = obj['super__author__name']
                    incipit = obj['super__incipit']
                    explicit = obj['super__explicit']
                    sgcount = obj['super__sgcount']
                    ssgcount = obj['super__ssgcount']
                    # Get the name(s) of the HC
                    lst_hc = CollectionSuper.objects.filter(super=super, collection__settype="hc").values(
                        'collection__id', 'collection__name')
                    hcs = ", ".join( [x['collection__name'] for x in lst_hc])
                    # ======== DEBUGGING ============
                    if hcs != "":
                        iStop = 1
                    # ===============================
                    # Get a URL for this ssg
                    url = reverse('equalgold_details', kwargs={'pk': super})
                    # Treat signatures for this SSG
                    sigbest = get_goldsig_dct(super)
                    if sigbest == "":
                        sigbest = "ssg_{}".format(super)
                    # Signatures for this SSG: get the full list
                    siglist = get_goldsiglist_dct(super)
                    # Put into object
                    oItem = dict(super=super, sig=sigbest, siglist=siglist,
                                 name=name, descr=descr, type=settype,
                                 order=order, code=code, url=url, author=authorname,
                                 incipit=incipit, explicit=explicit, sgcount=sgcount, 
                                 ssgcount=ssgcount, hcs=hcs)
                    # Add to list
                    lBack.append(oItem)

            # Debugging: print the list
            if bDebug:
                print("Collection id={}".format(coll.id))
                for oItem in lBack:
                    order = oItem.get("order")
                    super = oItem.get("super")
                    sig = oItem.get("sig")
                    code = get_passimcode(super, code)
                    print("sermon {}: ssg={} sig=[{}]".format(order, code, sig))

        except:
            msg = oErr.get_error_message()
            oErr.DoError("ssgs_collection")
        return lBack

    def ssgs_manuscript(self, bDebug = False):
        """Get the ordered list of SSGs related to a manuscript"""

        oErr = ErrHandle()
        lBack = None
        try:
            manu = self.manuscript
            # Get a list of SSGs to which the sermons in this manuscript point
            # Per issue #402 additional info needs to be there:
            #   author, 
            #   PASSIM code, 
            #   incipit, explicit, HC (if the SSG is part of a historical collection), 
            #   number of SGs contained in the equality set, 
            #   number of links to other SSGs
            # Per issue #402, phase 2, even more info needs to be added:
            #   for MANUSCRIPT source-lists, there should be an option for the user 
            #       to select additional SERMON manifestation details fields to be shown 
            #       (options: attributed author, section title, lectio, title, incipit, explicit, 
            #       postscriptum, feast, bible reference, cod. notes, notes, keywords)
            qs = manu.sermondescr_super.all().order_by('sermon__msitem__order').values(
                'sermon', 'sermon__msitem__order', 'super', 'super__code', 'super__author__name',
                'super__incipit', 'super__explicit', 'super__sgcount', 'super__ssgcount',
                'sermon__author__name', 'sermon__sectiontitle', 'sermon__quote', 'sermon__title', 
                'sermon__incipit', 'sermon__explicit', 'sermon__postscriptum', 'sermon__feast__name', 
                'sermon__bibleref', 'sermon__additional', 'sermon__note')
            # NOTE: the 'keywords' for issue #402 are a bit more cumbersome to collect...
            lBack = []
            with transaction.atomic():
                for obj in qs:
                    # Get the order number
                    order = obj['sermon__msitem__order']
                    # Get the SSG and all its characteristics
                    super = obj['super']
                    code = obj['super__code']
                    authorname = obj['super__author__name']
                    incipit = obj['super__incipit']
                    explicit = obj['super__explicit']
                    sgcount = obj['super__sgcount']
                    ssgcount = obj['super__ssgcount']
                    # Sermon characteristics
                    sermon = obj['sermon']
                    srm_author = obj['sermon__author__name']
                    srm_sectiontitle = obj['sermon__sectiontitle']
                    srm_lectio = obj['sermon__quote']
                    srm_title = obj['sermon__title']
                    srm_incipit = obj['sermon__incipit']
                    srm_explicit = obj['sermon__explicit']
                    srm_postscriptum = obj['sermon__postscriptum']
                    srm_feast = obj['sermon__feast__name']
                    srm_bibleref = obj['sermon__bibleref']
                    srm_codnotes = obj['sermon__additional']
                    srm_notes = obj['sermon__note']

                    # Get the name(s) of the HC
                    lst_hc = CollectionSuper.objects.filter(super=super, collection__settype="hc").values(
                        'collection__id', 'collection__name')
                    hcs = ", ".join( [x['collection__name'] for x in lst_hc])
                    # ======== DEBUGGING ============
                    if hcs != "":
                        iStop = 1
                    # ===============================

                    # TODO: Get the keywords
                    lst_kw = SermonDescrKeyword.objects.filter(sermon=sermon).values(
                        'keyword__name')
                    kws = ", ".join( [x['keyword__name'] for x in lst_kw])

                    # Get a URL for this ssg
                    url = reverse('equalgold_details', kwargs={'pk': super})
                    # Treat signatures for this SSG: get the best for showing
                    sigbest = get_goldsig_dct(super)
                    if sigbest == "":
                        sigbest = "ssg_{}".format(super)
                    # Signatures for this SSG: get the full list
                    siglist = get_goldsiglist_dct(super)
                    # Put into object
                    oItem = dict(super=super, sig=sigbest, siglist=siglist, hcs=hcs, kws=kws,
                                 order=order, code=code, url=url, author=authorname, type='ms',
                                 incipit=incipit, explicit=explicit, sgcount=sgcount, ssgcount=ssgcount,
                                 srm_author=srm_author, srm_sectiontitle=srm_sectiontitle, srm_lectio=srm_lectio,
                                 srm_title=srm_title, srm_incipit=srm_incipit, srm_explicit=srm_explicit,
                                 srm_postscriptum=srm_postscriptum, srm_feast=srm_feast, 
                                 srm_bibleref=srm_bibleref, srm_codnotes=srm_codnotes, srm_notes=srm_notes)
                    # Add to list
                    lBack.append(oItem)

            # Debugging: print the list
            if bDebug:
                print("Manuscript id={}".format(manu.id))
                for oItem in lBack:
                    order = oItem.get("order")
                    super = oItem.get("super")
                    sig = oItem.get("sig")
                    code = get_passimcode(super, code)
                    print("sermon {}: ssg={} sig=[{}]".format(order, code, sig))

        except:
            msg = oErr.get_error_message()
            oErr.DoError("ssgs_manuscript")
        return lBack
    

class SetDef(models.Model):
    """THe definition of a DCT"""

    # [1] obligatory name
    name = models.CharField("Name", max_length=STANDARD_LENGTH)
    # [1] A setlist just belongs to a research set
    researchset = models.ForeignKey(ResearchSet, on_delete=models.CASCADE, related_name="researchset_setdefs")
    # [1] order number of this DCT (assigned automatically)
    order = models.IntegerField("Order", default = 0)

    # [0-1] Optional notes for this definition
    notes = models.TextField("Notes", blank=True, null=True)

    # [1] A list of all the DCT parameters for this DCT
    contents = models.TextField("Contents", default="{}")

    # [1] And a date: the date of saving this manuscript
    created = models.DateTimeField(default=get_current_datetime)
    saved = models.DateTimeField(null=True, blank=True)

    def __str__(self):
        """Just return the name"""
        return self.name

    def save(self, force_insert = False, force_update = False, using = None, update_fields = None):
        # Adapt the save date
        self.saved = get_current_datetime()

        # Need any name or notes fixed?
        if self.name == None or self.name == "":
            # Calculate the proper order
            last = SetDef.objects.filter(researchset=self.researchset).order_by("-order", "-saved").first()
            if last == None:
                order = 1
            else:
                order = last.order + 1
            self.order = order
            # Automatically assign a name
            default_name = "{}_DCT_{:06}".format(self.researchset.profile.user.username, order)
            self.name = default_name
            self.notes = "Automatically created"

        # Check if the order is specified
        if self.order is None or self.order <= 0:
            # Specify the order
            self.order = SetDef.objects.filter(researchset__profile=self.researchset.profile).count() + 1

        # Initial saving
        response = super(SetDef, self).save(force_insert, force_update, using, update_fields)

        # Return the response when saving
        return response

    def check_default(rset, username):
        """Create a default SetDef for this research set"""

        # Are there any setdefs?
        if SetDef.objects.filter(researchset=rset).count() == 0:
            # Make sure to create at least one (with overall default values)
            obj = SetDef.objects.create(researchset=rset)
        return True

    def get_created(self):
        """REturn the creation date in a readable form"""

        # sDate = self.created.strftime("%d/%b/%Y %H:%M")
        sDate = get_crpp_date(self.created, True)
        return sDate

    def get_contents(self, pivot_id=None, recalculate=False):
        """Get the set of lists for this particular DCT"""

        oErr = ErrHandle()
        oBack = None
        try:
            # Get the SSGlists from the research set
            ssglists = self.researchset.get_ssglists(recalculate)

            # Possibly get (stored) parameters from myself
            params = {}
            if self.contents != "" and self.contents[0] == "{":
                params = json.loads(self.contents)
            
            # Return the parameters and the lists
            oBack = dict(params=params, ssglists=ssglists)

        except:
            msg = oErr.get_error_message()
            oErr.DoError("SetDef/get_contents")
        return oBack

    def get_notes_html(self):
        """Convert the markdown notes"""

        sNotes = "-"
        if self.notes != None:
            sNotes = markdown(self.notes)
        return sNotes

    def get_saved(self):
        """REturn the saved date in a readable form"""

        # sDate = self.saved.strftime("%d/%b/%Y %H:%M")
        sDate = get_crpp_date(self.saved, True)
        return sDate

    def get_setlist(self, pivot_id=None):
        """Get the set of lists for this particular DCT"""

        oErr = ErrHandle()
        oBack = None
        try:
            # Get to the research set
            id_list = [x['id'] for x in self.researchset.researchset_setlists.all().order_by('order').values("id")]
            lst_rset = []
            if pivot_id != None and pivot_id in id_list:
                lst_rset.append(pivot_id)
            # Add the remaining ID's
            for id in id_list:
                if not id in lst_rset: lst_rset.append(id)

            # Check the number of lists in the research set
            if lst_rset == None or len(lst_rset) < 2:
                oErr.Status("Not enough SSG-lists to compare")
                return None

            # We have enough lists: Get the lists of SSGs for each 
            lst_ssglists = []
            for setlist_id in lst_rset:
                # Get the actual setlist
                setlist = SetList.objects.filter(id=setlist_id).first()
                if setlist != None:
                    # Create an empty SSG-list
                    oSsgList = {}
                    # Add the object itself
                    oSsgList['obj'] = setlist
                    # Add the name object for this list
                    oSsgList['title'] = setlist.get_title_object()
                    # Get the list of SSGs for this list
                    oSsgList['ssglist'] = setlist.get_ssg_list()
                    # Add the list object to the list
                    lst_ssglists.append(oSsgList)
            
            # Return this list of lists
            oBack = dict(ssglists=lst_ssglists)
            # Prepare and create an appropriate table = list of rows
            rows = []
            # Create header row
            oRow = []
            oRow.append('Gr/Cl/Ot')
            for oSsgList in lst_ssglists:
                # Add the title *object*
                oRow.append(oSsgList['title'])
            rows.append(oRow)

            # Start out with the pivot: the *first* one in 'ssglist'
            lst_pivot = lst_ssglists[0]
            for oPivot in lst_pivot['ssglist']:
                # Create a row based on this pivot
                oRow = []
                # (1) row header
                oRow.append(oPivot['sig'])
                # (2) pivot SSG number
                oRow.append(oPivot['order'])
                # (3) SSG number in all other manuscripts
                ssg_id = oPivot['super']
                for lst_this in lst_ssglists[1:]:
                    bFound = False
                    order = ""
                    for oItem in lst_this['ssglist']:
                        if ssg_id == oItem['super']:
                            # Found it!
                            order = oItem['order']
                            bFound = True
                            break
                    oRow.append(order)
                # (4) add the row to the list
                rows.append(oRow)
            # Make sure we return the right information
            oBack['setlist'] = rows
        except:
            msg = oErr.get_error_message()
            oErr.DoError("SetDef/get_setlist")
        return oBack

    def get_view_link(self):
        """Get the HTML code to go to and view a DCT"""

        lHtml = []
        oErr = ErrHandle()
        try:
            if not self.contents == "":
                # Create the search
                # rl = "{}?usersearch={}".format(self.usersearch.view, self.usersearch.id)
                url = reverse('setdef_details', kwargs={'pk': self.id})
                name = self.name
                sBack = "<span  class='badge jumbo-1'><a href='{}'  title='Go to this DCT'>{}</a></span>".format(url, name)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("SetDef/get_view_link")
        return sBack

    def hidden_warning(self):
        sBack = ""
        oErr = ErrHandle()
        try:
            # Get the hidden contents
            if self.contents != "" and self.contents[0] == "{":
                params = json.loads(self.contents)
                # Find hidden rows
                hidden_rows = params.get("hidden_rows", [])
                size = len(hidden_rows)
                if size > 0:
                    sTitle = "Remove hidden rows: (a) Expand, (b) Save"
                    sBack = '<span title="{}"><b>Warning</b>: this DCT has <code>{}</code> hidden rows.</span>'.format(sTitle, size)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("SetDef/hidden_warning")
        return sBack        

    def update_order(profile):
        oErr = ErrHandle()
        bOkay = True
        try:
            # Something has happened
            qs = SetDef.objects.filter(researchset__profile=profile).order_by('order', 'id')
            with transaction.atomic():
                order = 1
                for obj in qs:
                    if obj.order != order:
                        obj.order = order
                        obj.save()
                    order += 1
        except:
            msg = oErr.get_error_message()
            oErr.DoError("Setdef/update_order")
            bOkay = False
        return bOkay


# ====================== Personal Research Environment models ========================================

class SaveGroup(models.Model):
    """A group that holds a number of SavedItems together"""

    # [1] a saved item belongs to a particular user's profile
    profile = models.ForeignKey(Profile, on_delete=models.CASCADE, related_name="profile_savegroups")
    # [1] obligatory name for this group
    name = models.CharField("Name", max_length=STANDARD_LENGTH)

    # [1] And a date: the date of saving this manuscript
    created = models.DateTimeField(default=get_current_datetime)
    saved = models.DateTimeField(default=get_current_datetime)

    def __str__(self):
        sBack = ""
        oErr = ErrHandle()
        try:
            sBack = "{}: {}".format(self.profile.user.username, self.name)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("SaveGroup")
        return sBack

    def get_created(self):
        """REturn the created date in a readable form"""

        sDate = get_crpp_date(self.created, True)
        return sDate

    def get_saved(self):
        """REturn the saved date in a readable form"""

        # sDate = self.saved.strftime("%d/%b/%Y %H:%M")
        sDate = get_crpp_date(self.saved, True)
        return sDate

    def get_size_markdown(self):
        """Get a markdown representation of the size of this group"""

        size = 0
        lHtml = []
        size = self.group_saveditems.count()
        if size > 0:
            # Create a display for this topic
            lHtml.append("<span class='badge signature gr'>{}</span>".format(size))
        else:
            lHtml.append("0")
        sBack = ", ".join(lHtml)
        return sBack

    def save(self, force_insert = False, force_update = False, using = None, update_fields = None):
        # Adapt the save date
        self.saved = get_current_datetime()
        if self.name is None or self.name == "":
            iLast = 1
            obj = SaveGroup.objects.all().order_by("-id").first()
            if not obj is None:
                iLast = obj.id + 1
            self.name = "Group #{}".format(iLast)
        response = super(SaveGroup, self).save(force_insert, force_update, using, update_fields)

        # Return the response when saving
        return response


class SavedItem(models.Model):
    """A saved item can be a M/S/SSG or HC or PD"""

    # [1] a saved item belongs to a particular user's profile
    profile = models.ForeignKey(Profile, on_delete=models.CASCADE, related_name="profile_saveditems")
    # [1] The saved items may be ordered (and they can be re-ordered by the user)
    order = models.IntegerField("Order", default=0)
    # [1] Each saved item must be of a particular type
    #     Possibilities: manu, serm, ssg, hist, pd
    sitemtype = models.CharField("Saved item type", choices=build_abbr_list(SAVEDITEM_TYPE), max_length=5)

    # [0-1] A SavedItem can optionally belong to a [SaveGroup]
    group = models.ForeignKey(SaveGroup, blank=True, null=True, on_delete=models.SET_NULL, related_name="group_saveditems")

    # Depending on the type of SavedItem, there is a pointer to the actual item
    # [0-1] Manuscript pointer
    manuscript = models.ForeignKey(Manuscript, blank=True, null=True, on_delete=models.SET_NULL, related_name="manuscript_saveditems")
    # [0-1] Sermon pointer
    sermon = models.ForeignKey(SermonDescr, blank=True, null=True, on_delete=models.SET_NULL, related_name="sermon_saveditems")
    # [0-1] SSG pointer
    equal = models.ForeignKey(EqualGold, blank=True, null=True, on_delete=models.SET_NULL, related_name="equal_saveditems")
    # [0-1] Collection pointer (that can be HC, public or personal collection
    collection = models.ForeignKey(Collection, blank=True, null=True, on_delete=models.SET_NULL, related_name="collection_saveditems")

    def __str__(self):
        sBack = ""
        oErr = ErrHandle()
        try:
            sBack = "{}: {}-{}".format(self.profile.user.username, self.order, self.sitemtype)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("SavedItem")
        return sBack

    def get_saveditem(item, profile, sitemtype):
        """If this is a saved item for the indicated user, get that item"""

        obj = None
        oErr = ErrHandle()
        try:
            if not profile is None:
                # Find out if this object exists
                if sitemtype == "manu":
                    obj = SavedItem.objects.filter(profile=profile, sitemtype=sitemtype, manuscript=item).first()
                elif sitemtype == "serm":
                    obj = SavedItem.objects.filter(profile=profile, sitemtype=sitemtype, sermon=item).first()
                elif sitemtype == "ssg":
                    obj = SavedItem.objects.filter(profile=profile, sitemtype=sitemtype, equal=item).first()
                elif sitemtype == "hc" or sitemtype == "pd":
                    obj = SavedItem.objects.filter(profile=profile, sitemtype=sitemtype, collection=item).first()
        except:
            msg = oErr.get_error_message()
            oErr.DoError("SavedItem/get_saveditem")
        return obj

    def get_saveditem_button(item, profile, sitemtype):
        """Provide a button to either turn this into a saved item or remove it as saved item"""

        sBack = ""
        oErr = ErrHandle()
        try:
            obj = SavedItem.get_saveditem(item, profile, sitemtype)
            html = []
            if obj is None:
                # make it possible to turn this into a saved item
                html.append("<a class='btn btn-xs jumbo-3' onclick='ru.dct.add_saveditem();'>")
                html.append('<span class="glyphicon glyphicon-plus"></span>')
                html.append('<span>Add to your saved items</span>')
                html.append('</a>')
            else:
                # make it possible to remove this as saved item
                html.append("<a class='btn btn-xs jumbo-4' onclick=''>")
                html.append('<span class="glyphicon glyphicon-minus"></span>')
                html.append('<span>Remove from your saved items</span>')
                html.append('</a>')
            # Combine into string
            sBack = "\n".join(html)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("SavedItem/get_saveditem_button")
        return sBack

    def save(self, force_insert = False, force_update = False, using = None, update_fields = None):
        # Check if the order is specified
        if self.order is None or self.order <= 0:
            # Specify the order
            self.order = SavedItem.objects.filter(profile=self.profile).count() + 1
        response = super(SavedItem, self).save(force_insert, force_update, using, update_fields)
        # Return the regular save response
        return response

    def update_order(profile):
        oErr = ErrHandle()
        bOkay = True
        try:
            # Something has happened
            qs = SavedItem.objects.filter(profile=profile).order_by('order', 'id')
            with transaction.atomic():
                order = 1
                for obj in qs:
                    if obj.order != order:
                        obj.order = order
                        obj.save()
                    order += 1
        except:
            msg = oErr.get_error_message()
            oErr.DoError("SavedItem/update_saveditems")
            bOkay = False
        return bOkay
    

class SavedSearch(models.Model):
    """A saved search links to the basic UserSearch"""

    # [1] obligatory name
    name = models.CharField("Name", max_length=STANDARD_LENGTH)
    # [1] a saved item belongs to a particular user's profile
    profile = models.ForeignKey(Profile, on_delete=models.CASCADE, related_name="profile_savedsearches")
    # [1] The saved items may be ordered (and they can be re-ordered by the user)
    order = models.IntegerField("Order", default=0)

    # [1] The usersearch that this saved search points to 
    usersearch = models.ForeignKey(UserSearch, on_delete=models.CASCADE, related_name="usersearch_savedsearches")

    def __str__(self):
        sBack = "{}: {}".format(self.profile.user.username, self.usersearch.id)
        return sBack

    def get_view_name(self):
        """Get the name of the view, without slashes"""

        sBack = "-"
        if not self.usersearch is None:
            sBack = self.usersearch.view
            sBack = sBack.replace("/list", "").replace("/search", "")
            sBack = sBack.replace("/", "")
        return sBack

    def get_view_link(self):
        """Get the HTML code to actually click and perform a saved search"""

        lHtml = []
        oErr = ErrHandle()
        try:
            if not self.usersearch is None:
                # Create the search
                url = "{}?usersearch={}".format(self.usersearch.view, self.usersearch.id)
                name = self.get_view_name()
                sBack = "<span  class='badge jumbo-1'><a href='{}'  title='Execute this saved search'>{}</a></span>".format(url, name)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("SavedSearch/get_view_link")
        return sBack

    def save(self, force_insert = False, force_update = False, using = None, update_fields = None):
        # Check if the order is specified
        if self.order is None or self.order <= 0:
            # Specify the order
            self.order = SavedSearch.objects.filter(profile=self.profile).count() + 1
        response = super(SavedSearch, self).save(force_insert, force_update, using, update_fields)
        # Return the regular save response
        return response

    def update_order(profile):
        oErr = ErrHandle()
        bOkay = True
        try:
            # Something has happened
            qs = SavedSearch.objects.filter(profile=profile).order_by('order', 'id')
            with transaction.atomic():
                order = 1
                for obj in qs:
                    if obj.order != order:
                        obj.order = order
                        obj.save()
                    order += 1
        except:
            msg = oErr.get_error_message()
            oErr.DoError("SavedSearch/update_order")
            bOkay = False
        return bOkay


class SavedVis(models.Model):
    """A saved visualization links to a particular URL plus a set of parameters"""

    # [1] obligatory name
    name = models.CharField("Name", max_length=STANDARD_LENGTH)
    # [1] a saved item belongs to a particular user's profile
    profile = models.ForeignKey(Profile, on_delete=models.CASCADE, related_name="profile_savedvisualizations")
    # [1] The saved items may be ordered (and they can be re-ordered by the user)
    order = models.IntegerField("Order", default=0)

    # [1] The URL that this saved visualization points to 
    visurl = models.CharField("Visualisation URL", max_length=STANDARD_LENGTH)
    # [0-1] The parameters (JSON) belonging to this search
    options = models.TextField("Parameters", blank=True, null=True)

    # [1] And a date: the date when this visualization was created
    created = models.DateTimeField(default=get_current_datetime)

    def __str__(self):
        sBack = "{}: {}".format(self.profile.user.username, self.name)
        return sBack

    def get_view_name(self):
        """Get the name of the view, without slashes"""

        sBack = "-"
        if not self.usersearch is None:
            sBack = self.usersearch.view
            sBack = sBack.replace("/list", "").replace("/search", "")
            sBack = sBack.replace("/", "")
        return sBack

    def get_view_link(self):
        """Get the HTML code to actually click and perform a saved visualization
        
        Note: the 'name' shown here should be the *type* of visualization
              (e.g: AF overlap, AF transmission, DCT)
        """

        lHtml = []
        pass_over = ['visurl', 'vistype']
        oErr = ErrHandle()
        try:
            # The visualization type is filed away in 'options'
            if not self.options is None:
                options = json.loads(self.options)
                # Get the [vistype] parameter
                name = options.get("vistype", "unknown")
                # Get the URL to the visualization
                visurl = options.get("visurl")
                # Add the id of the savedvis
                url = "{}?savedvis={}".format(visurl, self.id)

                ## Find all parameters
                #param_list = []
                #for k,v in options.items():
                #    if not k in pass_over:
                #        param_list.append("{}={}".format(k,v))
                ## Encode the parameters
                #params = base64_encode( "&".join(param_list))
                ## Combine into url
                #url = "{}?params={}".format( visurl, params)

                # Last chance...
                if not url is None and url != "":
                    sBack = "<span  class='badge jumbo-1'><a href='{}'  title='Execute this saved visualization'>{}</a></span>".format(url, name)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("SavedVis/get_view_link")
        return sBack

    def save(self, force_insert = False, force_update = False, using = None, update_fields = None):
        # Check if the order is specified
        if self.order is None or self.order <= 0:
            # Specify the order
            self.order = SavedVis.objects.filter(profile=self.profile).count() + 1
        response = super(SavedVis, self).save(force_insert, force_update, using, update_fields)
        # Return the regular save response
        return response

    def update_order(profile):
        oErr = ErrHandle()
        bOkay = True
        try:
            # Something has happened
            qs = SavedVis.objects.filter(profile=profile).order_by('order', 'id')
            with transaction.atomic():
                order = 1
                for obj in qs:
                    if obj.order != order:
                        obj.order = order
                        obj.save()
                    order += 1
        except:
            msg = oErr.get_error_message()
            oErr.DoError("SavedVis/update_order")
            bOkay = False
        return bOkay


class SelectItem(models.Model):
    """A selected item can be a M/S/SSG or HC or PD"""

    # [1] a saved item belongs to a particular user's profile
    profile = models.ForeignKey(Profile, on_delete=models.CASCADE, related_name="profile_selectitems")
    # [1] The saved items may be ordered (and they can be re-ordered by the user)
    order = models.IntegerField("Order", default=0)
    # [1] Each saved item must be of a particular type
    #     Possibilities: manu, serm, ssg, hist, pd
    selitemtype = models.CharField("Select item type", choices=build_abbr_list(SELITEM_TYPE), max_length=5)

    # Depending on the type of SelectItem, there is a pointer to the actual item
    # [0-1] Manuscript pointer
    manuscript = models.ForeignKey(Manuscript, blank=True, null=True, on_delete=models.SET_NULL, related_name="manuscript_selectitems")
    # [0-1] Sermon pointer
    sermon = models.ForeignKey(SermonDescr, blank=True, null=True, on_delete=models.SET_NULL, related_name="sermon_selectitems")
    # [0-1] SSG pointer
    equal = models.ForeignKey(EqualGold, blank=True, null=True, on_delete=models.SET_NULL, related_name="equal_selectitems")
    # [0-1] Collection pointer (that can be HC, public or personal collection
    collection = models.ForeignKey(Collection, blank=True, null=True, on_delete=models.SET_NULL, related_name="collection_selectitems")
    # [0-1] Pointer to SavedItem
    saveditem = models.ForeignKey(SavedItem, blank=True, null=True, on_delete=models.SET_NULL, related_name="saveditem_selectitems")

    def __str__(self):
        sBack = "{}: {}-{}".format(self.profile.user.username, self.order, self.selitemtype)
        return sBack

    def get_item(self):
        """Return the item to which this one points"""

        obj = None
        oErr = ErrHandle()
        try:
            # Find out if this object exists
            if self.selitemtype == "manu":
                obj = self.manuscript
            elif self.selitemtype == "serm":
                obj = self.sermon
            elif self.selitemtype == "ssg":
                obj = self.equal
            elif self.selitemtype == "hc" or self.selitemtype == "pd":
                obj = self.collection
            elif self.selitemtype == "svdi":
                obj = self.SavedItem
        except:
            msg = oErr.get_error_message()
            oErr.DoError("SelectItem/get_item")
        return obj

    def get_selectitem(item, profile, selitemtype):
        """If this is a selected item for the indicated user, get that item"""

        obj = None
        oErr = ErrHandle()
        try:
            if not profile is None:
                # Find out if this object exists
                if selitemtype == "manu":
                    obj = SelectItem.objects.filter(profile=profile, selitemtype=selitemtype, manuscript=item).first()
                elif selitemtype == "serm":
                    obj = SelectItem.objects.filter(profile=profile, selitemtype=selitemtype, sermon=item).first()
                elif selitemtype == "ssg":
                    obj = SelectItem.objects.filter(profile=profile, selitemtype=selitemtype, equal=item).first()
                elif selitemtype == "hc" or selitemtype == "pd":
                    obj = SelectItem.objects.filter(profile=profile, selitemtype=selitemtype, collection=item).first()
                elif selitemtype == "svdi":
                    obj = SelectItem.objects.filter(profile=profile, selitemtype=selitemtype, saveditem=item).first()
        except:
            msg = oErr.get_error_message()
            oErr.DoError("SelectItem/get_selectitem")
        return obj

    def get_selectcount(profile, selitemtype):
        """Get the amount of selected items for this particular user / selitemtype"""

        iCount = 0
        oErr = ErrHandle()
        try:
            if not profile is None:
                # Find out if this object exists
                iCount = SelectItem.objects.filter(profile=profile, selitemtype=selitemtype).count()
        except:
            msg = oErr.get_error_message()
            oErr.DoError("SelectItem/get_selectcount")
        return iCount


# ====================== Models to work with excel imports and curation ========================================


class ImportSet(models.Model):
    """The user's desire to import a particular Excel
    
    The Excel may be a definition of Manuscript or of an Authority File (EqualGold)
    """

    # [1] An import-set item belongs to a particular user's profile
    profile = models.ForeignKey(Profile, on_delete=models.CASCADE, related_name="profile_importsetitems")
    # [1] The import-set items may be ordered (and they can be re-ordered by the user)
    order = models.IntegerField("Order", default=0)
    # [1] Each import-set item must be of a particular type
    #     Possibilities: manu, serm, ssg, hist, pd
    importtype = models.CharField("Import type", choices=build_abbr_list(IMPORT_TYPE), max_length=5)

    # [0-1] Optional notes for this set
    notes = models.TextField("Notes", blank=True, null=True)

    # [0-1] Report on errors etc for this ImportSet
    report = models.TextField("Report", blank=True, null=True)

    # [0-1] Each importSet item contains a FileField that allows uploading an Excel file
    excel = models.FileField("Excel file", null=True, blank=True, upload_to=import_path)
    # [0-1] Name of the file as the user uploaded it
    name = models.CharField("Name", blank=True, null=True, max_length=STANDARD_LENGTH)

    # Depending on the type of SelectItem, there is a pointer to the actual item
    # [0-1] Manuscript pointer
    manuscript = models.ForeignKey(Manuscript, blank=True, null=True, on_delete=models.SET_NULL, related_name="manuscript_importsetitems")
    # [0-1] SSG pointer
    equal = models.ForeignKey(EqualGold, blank=True, null=True, on_delete=models.SET_NULL, related_name="equal_importsetitems")

    # [1] Each importset item has a status, defining where it is on the acceptance scale
    #     Scale: cre[ated], ch[an]g[ed], sub[mitted], rej[ected], acc[epted]
    status = models.CharField("Import status", choices=build_abbr_list(IMPORT_STATUS), max_length=5, default="cre")

    # [1] And a date: the date of saving this manuscript
    created = models.DateTimeField(default=get_current_datetime)
    saved = models.DateTimeField(default=get_current_datetime)

    # Many-to-many stuff
    projects = models.ManyToManyField(Project2, through="ImportSetProject", related_name="projects_importset")

    def __str__(self):
        sBack = "{}: {}-{}".format(self.profile.user.username, self.order, self.importtype)
        return sBack

    def adapt_order(self):
        """Re-calculate the order and adapt where needed"""

        qs = self.profile.profile_importsetitems.all().order_by("order")
        order = 1
        with transaction.atomic():
            # Walk all the SetList objects
            for obj in qs:
                # Check if the order is as it should be
                if obj.order != order:
                    #No: adapt the order
                    obj.order = order
                    # And save it
                    obj.save()
                # Keep track of how the order should be
                order += 1
        return None

    def do_import(self):
        bResult = True
        oErr = ErrHandle()
        lst_err = []
        try:
            # Actually perform the import
            if self.status == "acc":
                # Yes, we may perform the import
                username = self.profile.user.username

                # What if this is a manuscript upload?
                if self.importtype == "manu":
                    oResult = {'status': 'ok', 'count': 0, 'sermons': 0, 'msg': "", 'user': username}
                    kwargs = {'profile': self.profile, 'username': username, 'team_group': "",
                              'projects': self.projects.all()}
                    # Indicate that a NEW one should be created, if already existing
                    manucreate = True

                    bResult = ManuscriptUploadExcel.upload_one_excel(
                        self.excel.path, self.name, lst_err, oResult, kwargs, manucreate)

                    # What if the result is positive?
                    if bResult:
                        # We have a positive result: Add the link to the manuscript
                        self.manuscript = oResult.get("obj")
                        if self.manuscript is None:
                            oErr.Status("ImportSet/do_import: successful import, but no return manuscript")
                        else:
                            # All is in order, so save the results
                            self.save()

                elif self.importtype == "ssg":
                    # This is an Authority File
                    # TODO: add code to import an AF
                    oErr.Status("do_import: cannot yet process Authority Files")

        except:
            msg = oErr.get_error_message()
            oErr.DoError("ImportSet/do_import")
            bResult = False

        return bResult

    def do_submit(self):
        """Submit the ImportSet"""

        sBack = ""
        oErr = ErrHandle()
        try:
            # Create a review item if it doesn't exist yet
            obj = ImportReview.objects.filter(importset=self).first()
            if obj is None:
                # Create one
                obj = ImportReview.objects.create(importset=self)
            # Make sure to [re]set the ImportReview status
            if obj.status != "cre":
                # It has not been just created: set to change
                obj.status = "chg"
                obj.save()

            # Set my own status to "submitted"
            self.status = "sub"
            self.save()

        except:
            msg = oErr.get_error_message()
            oErr.DoError("ImportSet/do_verify")
        # Return the results of verification
        return sBack

    def do_verify(self):
        """Verify the Excel"""

        def check_for_string(sKey, is_error=False, max_length=None, exclude=None, cls=None, allowed=None, obligatory=False):
            """Check whether an item is a string"""

            oErr= ErrHandle()
            try:
                for idx, sRef in enumerate(oSermList[sKey]):
                    if obligatory and sRef is None:
                        html_err.append("Sermon: expecting a value for `{}` at row **{}**".format(sKey, idx+2))
                    elif not sRef is None:
                        # Note: string or int - both are allowed here
                        if not isinstance(sRef, str) and not isinstance(sRef, int):
                            if is_error:
                                html_err.append("Sermon: expecting string value for `{}` at row **{}**".format(sKey, idx+2))
                            else:
                                html_wrn.append("Sermon: expecting string value for `{}` at row **{}**".format(sKey, idx+2))
                        else:
                            # Make sure it becomes string, if it is not
                            if not isinstance(sRef, str):
                                sRef = str(sRef)
                            if not max_length is None:
                                if len(sRef) > max_length:
                                    html_wrn.append("Sermon: string in column `{}` too large at row **{}**".format(sKey, idx+2))
                            if not exclude is None:
                                if any( ele in exclude for ele in sRef):
                                    html_err.append("Sermon: `{}` may not contain '{}' at row **{}**".format(sKey, exclude, idx+2))
                            if not cls is None:
                                obj = cls.objects.filter(name__iexact=sRef).first()
                                if obj is None:
                                    html_wrn.append("Sermon: unknown `{}` item [{}] at row **{}**".format(sKey, sRef, idx+2))
                            if not allowed is None:
                                if not sRef in allowed:
                                    html_err.append("Sermon `{}` must be one of {} at row **{}**".format(sKey, allowed, idx+2))

            except:
                msg = oErr.get_error_message()
                oErr.DoError("check_for_string")

        def check_for_list(sKey, cls=None, field=None, obligatory=False):
            """Check whether an item is a stringified list"""

            oErr = ErrHandle()
            html_main = []
            try:
                for idx, sItem in enumerate(oSermList[sKey]):
                    if not sItem is None:
                        if not isinstance(sItem, str):
                            html_err.append("Sermon: unintelligable `{}` at row **{}**".format(sKey, idx+2))
                        elif not cls is None:
                            # Try to parse it
                            try:
                                lst_item = json.loads(sItem)
                                for sItem in lst_item:
                                    if field is None:
                                        obj = cls.objects.filter(name__iexact=sItem).first()
                                    else:
                                        obj = cls.objects.filter(**{"{}__iexact".format(field): sItem}).first()
                                    if obj is None:
                                        html_list = html_err if obligatory else html_wrn
                                        sMainPart = "Sermon: unknown `{}` item [{}]".format(sKey, sItem)
                                        if not sMainPart in html_main:
                                            html_main.append(sMainPart)
                                            html_list.append("{} row **{}**".format(sMainPart,idx+2))

                            except:
                                # Not a legitimate json string
                                html_err.append("Sermon: {} must be a legitimate JSON string at row **{}**".format(
                                            sKey, idx+2))
            except:
                msg = oErr.get_error_message()
                oErr.DoError("check_for_list")
                

        sBack = ""
        oErr = ErrHandle()
        html_err = []
        html_wrn = []
        lst_column = ["Order", "Parent", "FirstChild", "Next", "Type", "External ids", "Status", "Locus", 
                      "Attributed author", "Section title", "Lectio", "Title", "Incipit", "Explicit", "Postscriptum", 
                      "Feast", "Bible reference(s)", "Cod. notes", "Note", "Keywords", "Keywords (user)", 
                      "Gryson/Clavis (manual)", "Gryson/Clavis (auto)", "Personal Datasets", "Literature", "SSG links"]
        oSermList = {}
        try:
            # Get the path to this Excel
            excel_path = os.path.abspath(os.path.join(MEDIA_ROOT, self.excel.name))
            # Load the Excel file
            wb = openpyxl.load_workbook(excel_path, read_only=True)
            sheetnames = wb.sheetnames
            ws_manu = None
            ws_sermo = None
            ws_info = None
            lst_ws = []
            for sname in sheetnames:
                if "manuscript" in sname.lower():
                    ws_manu = wb[sname]
                    lst_ws.append(ws_manu)
                elif "sermons" in sname.lower():
                    ws_sermo = wb[sname]
                    lst_ws.append(ws_sermo)
                elif "info" in sname.lower():
                    ws_info = wb[sname]
            # Check if we have the correct number of worksheets
            if len(wb.sheetnames) >3:
                html_err.append( "The Excel should contain just one sheet 'Manuscript' and one sheet 'Sermons' (and optionally 'Info')")
            elif len(lst_ws) != 2:
                # Are there more sheets?
                if len(lst_ws) > 2:
                    html_err.append( "The Excel should contain just one sheet 'Manuscript' and one sheet 'Sermons'")
                else:
                    # It is less than 2
                    if ws_manu is None:
                        # There is no manuscript sheet
                        html_err.append( "The Excel doesn't contain a sheet called 'Manuscript'")
                    elif ws_sermo is None:
                        # There is no manuscript sheet
                        html_err.append( "The Excel doesn't contain a sheet called 'Sermons'")
                    else:
                        html_err.append( "The Excel sheet's names are unintelligable. They should be: 'Manuscript', 'Sermons'")

            # Verify the information on the manuscript sheet
            if len(html_err) == 0 and not ws_manu is None:
                # Read in the first two columns
                row_no = 1
                oValues = {}
                while ws_manu.cell(row=row_no, column=1).value:
                    # First row gets special treatment
                    k = ws_manu.cell(row=row_no, column=1).value
                    v = ws_manu.cell(row=row_no, column=2).value
                    if row_no == 1:
                        if k != "Field" or v != "Value":
                            html_err.append("Manuscript columns should be [Field], [Value]")
                            break
                    else:
                        # Make sure we use case insensitivity
                        k = k.lower()
                        # Check that key is not there yet
                        if k in oValues:
                            html_err.append("Manuscript field is used multiple times: [{}]".format(k))
                            break
                        else:
                            oValues[k] = v

                    row_no += 1

                # Check at least for shelf mark, country, city, library
                shelfmark = oValues.get("shelf mark")
                country = oValues.get("country")
                city = oValues.get("city")
                library = oValues.get("library")
                if shelfmark is None:
                    html_wrn.append("Manuscript has no shelf mark specified")
                if country is None:
                    html_wrn.append("Manuscript has no country specified")
                if city is None:
                    html_wrn.append("Manuscript has no city specified")
                if library is None:
                    html_wrn.append("Manuscript has no library specified")

                # They should all four be supplied
                if shelfmark is None or country is None or city is None or library is None:
                    html_err.append("Manuscript lacks one of: shelf-mark, country, city, library")

            # Verify the information on the sermons sheet
            if len(html_err) == 0 and not ws_sermo is None:
                # Check that all the 26 column names are there
                row_no = 1
                for idx, sColumn in enumerate(lst_column):
                    col_no = idx+1
                    v = ws_sermo.cell(row=row_no, column=col_no).value
                    if not isinstance(v, str):
                        # Note: allow columns 24-26 to not be present.
                        #       those are: Personal Datasets, Literature, SSG links
                        if v is None and col_no < 24:
                            html_err.append("Sermon column number **{}** must be string".format(col_no))
                    else:
                        if v.lower() != sColumn.lower():
                            html_err.append("Sermons column number **{}** expect title `{}`, but excel uses title `{}`".format(
                            col_no, sColumn, v))

                # If we have column name errors
                if len(html_err) > 0:
                    # Provide a warning message with the right column names
                    sColumns = "`{}`".format("`, `".join(lst_column))
                    html_wrn.append("The sheet [Sermons] must have these column names: {}".format(sColumns))

                # Create dictionary with lists
                for col_name in lst_column:
                    oSermList[col_name] = []

                # Iterate over the rows
                row_last = -1
                row_num = 1
                for row in ws_sermo.iter_rows():
                    row_values = [x.value for x in row]
                    v = None if len(row_values) ==0 else row_values[0]
                    if row_num > 1 and not v is None and v!= "":
                        # Possibly append row values if needed
                        while len(row_values) < len(lst_column):
                            row_values.append(None)

                        if row_num > row_last:
                            row_last = row_num
                        # Add all items to their individual lists
                        for idx, col_name in enumerate(lst_column):
                            v = row_values[idx]
                            oSermList[col_name].append(v)
                    # Go to the next row
                    row_num += 1

                # The lists for Order, Parent, First,Next must be there
                lst_mustbe = ['Order', 'Parent', 'FirstChild', 'Next']
                for sListName in lst_mustbe:
                    if len(oSermList[sListName]) == 0:
                        html_err.append("Sermon sheet misses values for column [{}]".format(sListName))

                # The first four columns may only contain numbers: [order, parent, first, next]
                order_prev = 0
                lst_order = []
                for row_no in range(2, row_last+1):
                    idx = row_no - 2
                    order = oSermList['Order'][idx]
                    lst_order.append(order)
                    if not isinstance(order,int):
                        # Order must always be there and it must be an integer
                        html_err.append("Sermon order must be specified and must be integer (row={})".format(row_no))
                        break
                    else:
                        if order > order_prev:
                            order_prev = order
                        else:
                            html_err.append("Sermon order at row {} must be higher than previous row".format(row_no))
                            break

                    # Getting here means there is some ligit data
                    parent = oSermList['Parent'][idx]
                    if not parent is None:
                        # Check parent is among previous ones
                        if isinstance(parent, int):
                            if not parent in lst_order:
                                html_err.append("Sermon parent wrong at row {}".format(row_no))
                        else:
                            html_err.append("Sermon parent at row {} must be integer".format(row_no))

                    # Review use of firstchild
                    firstchild =  oSermList['FirstChild'][idx]
                    if not firstchild is None:
                        if isinstance(firstchild, int):
                            if not firstchild > order:
                                html_err.append("Sermon firstchild must be higher than current row order at row {}".format(row_no))
                            elif not firstchild in oSermList['Order']:
                                html_err.append("Sermon firstchild must be part of 'Order' column at row {}".format(row_no))
                        else:
                            html_err.append("Sermon firstchild at row {} must be integer".format(row_no))

                    # Consider next sibling
                    nextsib =  oSermList['Next'][idx]
                    if not nextsib is None:
                        if isinstance(nextsib, int):
                            if not nextsib > order:
                                html_err.append("Sermon next must be higher than current row order at row {}".format(row_no))
                            elif not nextsib in oSermList['Order']:
                                html_err.append("Sermon next must be part of 'Order' column at row {}".format(row_no))
                        else:
                            html_err.append("Sermon next at row {} must be integer".format(row_no))
                    # Continue to the next row

                # Check column 'Type'
                type_allowed = ['Structural', 'Plain']
                check_for_string("Type", allowed=['Structural', 'Plain'])

                # Column 'Status' is irrelevant, as it will be set itself

                # Column locus: check type and length
                check_for_string("Locus", is_error=True, max_length=15, obligatory=False)

                # Check attributed author(s): do they occur in the database?
                check_for_string("Attributed author", cls=Author)

                # Check for proper string in: section title, lectio, title
                check_for_string("Section title")
                check_for_string("Lectio")
                check_for_string("Title")

                # Check the incipit/explicit/postscriptum
                check_for_string("Incipit", exclude="[]")
                check_for_string("Explicit", exclude="[]")
                check_for_string("Postscriptum", exclude="[]")

                # A FEAST must be a stringified JSON list of strings
                check_for_list("Feast", cls=Feast, obligatory=True)

                # Check whether Bible ref is a string
                check_for_string("Bible reference(s)")
                check_for_string("Cod. notes")
                check_for_string("Note")

                # Keywords must be stringified JSON list of strings
                check_for_list("Keywords", cls=Keyword, obligatory=True)
                check_for_list("Keywords (user)", cls=Keyword)

                # Check signatures
                check_for_list("Gryson/Clavis (manual)", cls=Signature, field="code")
                check_for_list("Gryson/Clavis (auto)", cls=Signature, field="code")

                # Personal datasets must already exist, I guess
                check_for_list("Personal Datasets", cls=Collection)

                # Literature must be a stringified JSON list of strings
                check_for_list("Literature")

                # SSG links must be JSON lists of strings, pointing to an SSG via their PASSIM code
                check_for_list("SSG links", cls=EqualGold, field="code")


            # Combine into reports
            sWarning = "  \n".join(html_wrn)
            sError = "  \n".join(html_err)
            # Do we have a warning/error report?
            if sError == "":
                # There are no errors - maybe only warnings?
                if sWarning == "":
                    sReport = "Excel file has been verified"
                else:
                    sReport = "### WARNINGS\n{}".format(sWarning)
                self.report = sReport
                self.status = "ver"
                self.save()
            else:
                # There are errors: collect and show them
                sReport = "### ERRORS\n{}".format(sError)
                if sWarning != "":
                    sReport = "{}\n### WARNINGS\n{}".format(sReport, sWarning)

                # set the status to REJECTED
                self.report = sReport
                self.status = "rej"
                self.save()

            sBack = sReport

            # Make sure to close the excel
            wb.close()
            # Double check the status of this file
            sStatus = "closed" if self.excel.closed else "open"
            # oErr.Status("Importset/do_verify finishes Workbook as: {}".format(sStatus))

        except:
            msg = oErr.get_error_message()
            oErr.DoError("ImportSet/do_verify")
        # Return the results of verification
        return sBack

    def get_created(self):
        """REturn the created date in a readable form"""

        sDate = get_crpp_date(self.created, True)
        return sDate

    def get_filename(self):
        sBack = str(self.excel)
        return sBack

    def get_importmode(self):
        sBack = ""
        bResult = True
        oErr = ErrHandle()
        
        try:
            mode = ""
            if self.status in ['chg', 'rej']: # 'cre', 
                mode = "verify"
            elif self.status in ['ver']:
                mode = "submit"
            elif self.status in ['sub']:
                mode = "review"
            elif self.status in ['acc']:
                mode = "accepted"
            sBack = mode
        except:
            msg = oErr.get_error_message()
            oErr.DoError("ImportSet/get_importmode")
            bResult = False
        return sBack

    def get_name(self):
        sBack = ""
        if not self.name is None:
            sBack = self.name
        return sBack

    def get_name_html(self):
        """Get the name as well as a link to download the Excel file"""

        sBack = ""
        oErr = ErrHandle()
        try:
            if not self.name is None:
                sBack = '<span class="badge jumbo-1">{}</span>'.format(self.name)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("ImportSet/get_name_html")
        return sBack

    def get_notes_html(self):
        """Convert the markdown notes"""

        sNotes = "-"
        if self.notes != None:
            sNotes = markdown(self.notes)
        return sNotes    

    def get_owner(self):
        sBack = self.profile.user.username
        return sBack

    def get_projects_html(self):
        """Get the list of projects to which the import will be assigned"""

        sBack = "(none)"
        oErr = ErrHandle()
        try:
            # Get the queryset
            qs = self.projects.all()
            if qs.count() > 0:
                html = []
                for obj in qs:
                    if obj.__class__.__name__ == "Project2":
                        project = obj
                    else:
                        project = obj.project
                    url = reverse('project2_details', kwargs={'pk': project.id})
                    html.append("<span class='project'><a href='{}'>{}</a></span>".format(url, project.name))
                sBack = ",".join(html)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("get_projects")

        return sBack

    def get_report_html(self):
        """Convert the markdown report"""

        sReport = "-"
        if self.report != None:
            sReport = markdown(self.report)
        return sReport    

    def get_result(self):
        """Get a button to go to the imported result"""

        sBack = ""
        oErr = ErrHandle()
        try:
            if self.importtype == "manu" and not self.manuscript is None:
                url = reverse('manuscript_details', kwargs={'pk': self.manuscript.id})
                #sBack = '<span class="badge signature ot"><a href="{}"></a></span>'.format(url)
                sBack = '<a role="button" class="btn btn-xs jumbo-1" href="{}">Manuscript</a>'.format(url)
            elif self.importtype == "ssg" and not self.equal is None:
                url = reverse('equalgold_details', kwargs={'pk': self.equal.id})
                # sBack = '<span class="badge signature gr"><a href="{}"></a></span>'.format(url)
                sBack = '<a role="button" class="btn btn-xs jumbo-1" href="{}">Authority File</a>'.format(url)

        except:
            msg = oErr.get_error_message()
            oErr.DoError("ImportReview/get_submission")
        return sBack

    def get_saved(self):
        """REturn the saved date in a readable form"""

        # sDate = self.saved.strftime("%d/%b/%Y %H:%M")
        sDate = get_crpp_date(self.saved, True)
        return sDate

    def get_status(self, html=False):
        sStatus = self.get_status_display()
        if html:
            sStatus = '<span class="badge signature ot">{}</span>'.format(sStatus)
            # Check if there is a result
            sResult = self.get_result()
            if sResult != "":
                # There is a result, so add it to the status
                sStatus = '{}<span>&nbsp;</span>{}'.format(sStatus, sResult)

        return sStatus

    def get_type(self):
        return self.get_importtype_display()

    def save(self, force_insert = False, force_update = False, using = None, update_fields = None):
        response = None
        oErr = ErrHandle()
        try:
            # Just monitor the excel status
            if not self.excel is None:
                sStatus = "closed" if self.excel.closed else "open"
                # oErr.Status("ImportSet excel file status = {}".format(sStatus))
                if sStatus == "open":
                    iDoubleCheck = 1

            # Check if the order is specified
            if self.order is None or self.order <= 0:
                # Specify the order
                self.order = ImportSet.objects.filter(profile=self.profile).count() + 1
            # Adapt the save date
            self.saved = get_current_datetime()
            
            # If the status is 'cre'..
            if self.status == "cre":
                # Check if an excel file has been specified
                if self.importtype in ['manu', 'ssg'] and not self.excel is None and not self.excel.file is None:
                    # Move on to the status 'chg'
                    self.status = "chg"

            ## Possibly first close the file
            #if not self.excel is None:
            #    self.excel.close()

            response = super(ImportSet, self).save(force_insert, force_update, using, update_fields)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("ImportSet/save")
        # Return the response when saving
        return response

    def update_order(profile):
        oErr = ErrHandle()
        bOkay = True
        try:
            # Something has happened
            qs = ImportSet.objects.filter(profile=profile).order_by('order', 'id')
            with transaction.atomic():
                order = 1
                for obj in qs:
                    if obj.order != order:
                        obj.order = order
                        obj.save()
                    order += 1
        except:
            msg = oErr.get_error_message()
            oErr.DoError("ImportSet/update_order")
            bOkay = False
        return bOkay


class ImportReview(models.Model):
    """Review of one importset by a moderator"""

    # [1] Link to the importset
    importset = models.ForeignKey(ImportSet, on_delete=models.CASCADE, related_name="importset_reviews")
    # [0-1] Link to the moderator - as soon as it is assigned
    moderator = models.ForeignKey(Profile, on_delete=models.SET_NULL, blank=True, null=True, related_name="moderator_reviews")
    # [1] The import-set items may be ordered (and they can be re-ordered by the user)
    order = models.IntegerField("Order", default=0)

    # [0-1] Optional notes for this review
    notes = models.TextField("Notes", blank=True, null=True)

    # [1] Each review item has a status, defining what the suggestion is
    #     Scale: cre[ated], ch[an]g[ed], sub[mitted], rej[ected], acc[epted]
    status = models.CharField("Review status", choices=build_abbr_list(IMPORT_STATUS), max_length=5, default="cre")

    # [1] And a date: the date of saving this manuscript
    created = models.DateTimeField(default=get_current_datetime)
    saved = models.DateTimeField(default=get_current_datetime)

    def __str__(self):
        sBack = "{}: {} (id={})".format(self.moderator.user.username, self.order, self.id)
        return sBack

    def adapt_order(self):
        """Re-calculate the order and adapt where needed"""

        # Do we have a moderator?
        if self.moderator is None:
            qs = ImportReview.objects.all().order_by("moderator", "order", "id")
        else:
            # We DO have a moderator
            qs = self.moderator.moderator_reviews.all().order_by("order", "id")
        order = 1
        with transaction.atomic():
            # Walk all the SetList objects
            for obj in qs:
                # Check if the order is as it should be
                if obj.order != order:
                    #No: adapt the order
                    obj.order = order
                    # And save it
                    obj.save()
                # Keep track of how the order should be
                order += 1
        return None

    def do_process(self, profile, verdict):
        """Process action of moderator to accept or reject the Importset"""

        result = ""
        oErr = ErrHandle()
        try:
            # Get tot he importset
            importset = self.importset
            if not importset is None:
                if verdict == "rej":
                    # Reject the submission
                    importset.status = "rej"    # It is now rejected
                    importset.save()
                    # Also change my own status
                    self.status = "rej"
                    self.moderator = profile
                    self.save()
                elif verdict == "acc":
                    # Accept the submission
                    importset.status = "acc"
                    importset.save()

                    # First: try to import the Excel
                    bResult = importset.do_import()
                    if not bResult:
                        # Something has gone wrong
                        importset.status = "rej"
                        importset.notes = "### ERROR\nCould not perform the import\n\n{}".format(importset.notes)
                        importset.save()

                    # And change my own status
                    self.status = "acc"
                    self.moderator = profile
                    self.save()

        except:
            msg = oErr.get_error_message()
            oErr.DoError("ImportReview/do_process")

        return result

    def get_created(self):
        """REturn the created date in a readable form"""

        sDate = get_crpp_date(self.created, True)
        return sDate

    def get_moderator(self):
        sBack = ""
        if not self.moderator is None:
            sBack = self.moderator.user.username
        return sBack

    def get_notes_html(self):
        """Convert the markdown notes"""

        sNotes = "-"
        if self.notes != None:
            sNotes = markdown(self.notes)
        return sNotes    

    def get_owner(self):
        sBack = self.importset.profile.user.username
        return sBack

    def get_saved(self):
        """REturn the saved date in a readable form"""

        # sDate = self.saved.strftime("%d/%b/%Y %H:%M")
        sDate = get_crpp_date(self.saved, True)
        return sDate

    def get_status(self, html=False):
        sStatus = self.get_status_display()
        if html:
            sStatus = '<span class="badge signature gr">{}</span>'.format(sStatus)
        return sStatus

    def get_submission(self):
        """Get a button to go to this submission"""

        sBack = ""
        oErr = ErrHandle()
        try:
            if not self.importset is None:
                url = reverse('importset_details', kwargs={'pk': self.importset.id})
                sBack = '<span class="badge signature ot"><a href="{}"></a></span>'.format(url)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("ImportReview/get_submission")
        return sBack

    def save(self, force_insert = False, force_update = False, using = None, update_fields = None):
        response = None
        oErr = ErrHandle()
        try:
            # Adapt the save date
            self.saved = get_current_datetime()

            # Check if the order is specified
            if self.order is None or self.order <= 0:
                # Specify the order
                if self.moderator is None:
                    # Order the ones without moderator
                    self.order = ImportReview.objects.filter(moderator__isnull=True).count() + 1
                else:
                    # This particular moderator
                    self.order = ImportReview.objects.filter(moderator=self.moderator).count() + 1

            # If needed, set the review status

            # Perform the normal logic
            response = super(ImportReview, self).save(force_insert, force_update, using, update_fields)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("ImportReview/save")
        # Return the response when saving
        return response

    def update_order(profile):
        oErr = ErrHandle()
        bOkay = True
        try:
            # Something has happened
            qs = ImportReview.objects.filter(moderator=profile).order_by('order', 'id')
            with transaction.atomic():
                order = 1
                for obj in qs:
                    if obj.order != order:
                        obj.order = order
                        obj.save()
                    order += 1
        except:
            msg = oErr.get_error_message()
            oErr.DoError("ImportReview/update_order")
            bOkay = False
        return bOkay


class ImportSetProject(models.Model):
    """Project associated with the import set"""

    # [1] Obligatory importset
    importset = models.ForeignKey(ImportSet, on_delete=models.CASCADE, related_name="importset_projects")
    # [1] Obligatory project
    project = models.ForeignKey(Project2, on_delete=models.CASCADE, related_name="importset_projects")
    # [1] And a date: the date of saving this manuscript
    created = models.DateTimeField(default=get_current_datetime)

    def __str__(self):
        sName = self.importset.name
        if sName is None or sName == "":
            sName = "id{}".format(self.importset.id)
        sBack = "{}-{}".format(sName, self.project.name)
        return sBack




