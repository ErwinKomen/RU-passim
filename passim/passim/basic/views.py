"""
Definition of views for the BASIC app.
"""

from django.apps import apps
from django.contrib.auth.models import User, Group
# from django.urls import reverse
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.core.files.base import File
from django.db import transaction
from django.db.models import Q, Prefetch, Count, F
from django.db.models.fields.files import FieldFile
from django.db.models.functions import Lower
from django.db.models.query import QuerySet 
from django.forms.models import model_to_dict
from django.http import HttpRequest, HttpResponse, HttpResponseRedirect, JsonResponse, FileResponse
from django.shortcuts import get_object_or_404, render, redirect
from django.template.loader import render_to_string
from django.urls import reverse
from django.utils import timezone
from django.views.generic.detail import DetailView
from django.views.generic.base import RedirectView
from django.views.generic import ListView, View

import json
import fnmatch
import os
import re
import base64
import csv
import openpyxl
from openpyxl.utils.cell import get_column_letter
from io import StringIO


from datetime import datetime

# provide error handling
from .utils import ErrHandle

from passim.basic.models import UserSearch
from passim.cms.view_utils import cms, cms_translate


# Some constants that can be used
paginateSize = 20
paginateSelect = 15
paginateValues =  (100, 50, 20 )    # See issue #603. Old: (100, 50, 20, 10, 5, 2, 1, )
PaginateMax = 100                   # Don't allow larger than 100

# Global debugging 
bDebug = False

# General functions serving the list and details views

def get_application_name():
    """Try to get the name of this application"""

    # Walk through all the installed apps
    for app in apps.get_app_configs():
        # Check if this is a site-package
        if "site-package" not in app.path:
            # Get the name of this app
            name = app.name
            # Take the first part before the dot
            project_name = name.split(".")[0]
            return project_name
    return "unknown"
# Provide application-specific information
PROJECT_NAME = get_application_name()
app_uploader = "{}_uploader".format(PROJECT_NAME.lower())
app_user = "{}_user".format(PROJECT_NAME.lower())
app_editor = "{}_editor".format(PROJECT_NAME.lower())
app_userplus = "{}_userplus".format(PROJECT_NAME.lower())
app_moderator = "{}_moderator".format(PROJECT_NAME.lower())
app_developer = "{}_developer".format(PROJECT_NAME.lower())

def user_is_authenticated(request):
    # Is this user authenticated?
    username = request.user.username
    user = User.objects.filter(username=username).first()
    response = False 
    if user != None:
        try:
            response = user.is_authenticated()
        except:
            response = user.is_authenticated
    return response

def user_is_ingroup(request, sGroup):
    # Is this user part of the indicated group?
    #username = request.user.username
    #user = User.objects.filter(username=username).first()
    user = request.user

    # Only look at group if the user is known
    if user == None:
        glist = []
    else:
        glist = [x.name for x in user.groups.all()]

        # Only needed for debugging
        if bDebug:
            ErrHandle().Status("User [{}] is in groups: {}".format(user, glist))
    # Evaluate the list
    bIsInGroup = (sGroup in glist)
    return bIsInGroup

def user_is_superuser(request):
    bFound = False
    # Is this user part of the indicated group?
    username = request.user.username
    if username != "":
        user = User.objects.filter(username=username).first()
        if user != None:
            bFound = user.is_superuser
    return bFound

def is_ajax(request):
    return request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest'

def get_breadcrumbs(request, name, is_menu, lst_crumb=[], **kwargs):
    """Process one visit and return updated breadcrumbs"""

    # Initialisations
    p_list = []
    p_list.append({'name': 'Home', 'url': reverse('home')})
    # Find out who this is
    username = "anonymous" if request.user == None else request.user.username
    if username != "anonymous" and request.user.username != "":
        # Add the visit
        currenturl = request.get_full_path()
        # Visit.add(username, name, currenturl, is_menu, **kwargs)
        # Set the full path, dependent on the arguments we get
        for crumb in lst_crumb:
            if len(crumb) == 2:
                p_list.append(dict(name=crumb[0], url=crumb[1]))
            else:
                pass
        # Also add the final one
        p_list.append(dict(name=name, url=currenturl))
    # Return the breadcrumbs
    return p_list

def action_model_changes(form, instance):
    changes = {}
    exclude = []
    oErr = ErrHandle()
    try:
        field_values = model_to_dict(instance)
        changed_fields = form.changed_data
        if hasattr(form, 'exclude'):
            exclude = form.exclude
        for item in changed_fields: 
            if item in field_values:
                if isinstance(field_values[item], File):
                    changes[item] =  str(field_values[item])
                else:
                    changes[item] = field_values[item]
            elif item not in exclude:
                # It is a form field
                try:
                    representation = form.cleaned_data[item]
                    if isinstance(representation, QuerySet):
                        # This is a list
                        rep_list = []
                        for rep in representation:
                            rep_str = str(rep)
                            rep_list.append(rep_str)
                        representation = json.dumps(rep_list)
                    elif isinstance(representation, str) or isinstance(representation, int):
                        representation = representation
                    elif isinstance(representation, object):
                        try:
                            representation = representation.__str__()
                        except:
                            representation = str(representation)
                    changes[item] = representation
                except:
                    changes[item] = "(unavailable)"
    except:
        msg = oErr.get_error_message()
        oErr.DoError("action_model_changes")
    return changes

def has_string_value(field, obj):
    response = (field != None and field in obj and obj[field] != None and obj[field] != "" and \
                ( isinstance(obj[field], str) or isinstance(obj[field], int) ) )
    
    return response

def has_list_value(field, obj):
    response = (field != None and field in obj and obj[field] != None and len(obj[field]) > 0)
    return response

def has_Q_value(field, obj):
    response = (field != None and field in obj and obj[field] != None and obj[field] != "" and \
                not isinstance(obj[field], str) and not isinstance(obj[field], int) )
    
    return response

def isempty(value):
    response = True
    if value != None:
        if isinstance(value, str):
            response = (value == "")
        elif isinstance(value, int):
            response = False
        else:
            response = (len(value) == 0)
    return response

def get_number(s_input):
    """Get the first consecutive number from the string"""

    if isinstance(s_input, int):
        iBack = s_input
    else:
        temp = re.findall(r'\d+', s_input)
        if len(temp) == 0:
            iBack = -1
        else:
            iBack = int(temp[0])
    return iBack

def has_obj_value(field, obj, model_name=None):
    if field == None:
        response = False
    elif field in obj and isinstance(obj[field], str):
        response = (obj[field] != "")
    else:
        response = (field != None and field in obj and obj[field] != None )
    if response and model_name != None:
        response = (obj[field].__class__.__name__ == model_name)
    return response

def adapt_search(val, regex_function=None):
    # First trim
    val = val.strip()
    # Double check whether we don't have a starting ^ and trailing $ yet
    if len(val) > 0:
        if "#" in val:
            val = r'(^|(.*\b))' + val.replace('#', r'((\b.*)|$)')
        else:
            val = fnmatch.translate(val)
            if val[0] != '^':
                val = "^{}".format(val)
            if val[-1] != "$":
                val = "{}$".format(val)

        # Is there a regex function?
        if regex_function != None:
            val = regex_function(val)

    ## Then add start and en matter 
    #val = '^' + fnmatch.translate(val) + '$'
    return val

def make_search_list(filters, oFields, search_list, qd, lstExclude):
    """Using the information in oFields and search_list, produce a revised filters array and a lstQ for a Queryset"""

    def enable_filter(filter_id, head_id=None): 
        full_filter_id = "filter_{}".format(filter_id)
        for item in filters:
            include_id = item.get('include_id', '')
            # EK: this is too loose:
            #     if filter_id in item['id'] or full_filter_id == include_id:
            if full_filter_id == item['id'] or full_filter_id == include_id:
                item['enabled'] = True
                # Break from my loop
                break

            # =============================
            # EK: This is the code that got replaced by the above
            ## (1) first create two strings in order to compare with one another:
            ##           the selected filters 
            ##           and the hidden columns (e.g. title and sectiontitle)
            #temp_filter_id = str("filter_" + filter_id)
            #temp_item_id = str(item['id'])                  
            ## (2) If the selected filter(s) match(es) one or two of the hidden columns,
            ##        then enabled should be given a True
            #if temp_filter_id == temp_item_id:    
            #    item['enabled'] = True            
            # =============================

        # Check if this one has a head
        if head_id != None and head_id != "":
            for item in filters:
                if head_id in item['id']:
                    item['enabled'] = True
                    # Break from this sub-loop
                    break
        return True 

    def get_value(obj, field, default=None):
        if field in obj:
            sBack = obj[field]
        else:
            sBack = default
        return sBack

    oErr = ErrHandle()

    try:
        # (1) Create default lstQ
        lstQ = []
        dictQ = {}

        # (2) Reset the filters in the list we get
        for item in filters: item['enabled'] = False
    
        # (3) Walk all sections
        for part in search_list:
            head_id = get_value(part, 'section')

            # (4) Walk the list of defined searches
            for search_item in part['filterlist']:
                keyS = get_value(search_item, "keyS")
                keyId = get_value(search_item, "keyId")
                keyFk = get_value(search_item, "keyFk")
                keyList = get_value(search_item, "keyList")
                infield = get_value(search_item, "infield")
                dbfield = get_value(search_item, "dbfield")
                fkfield = get_value(search_item, "fkfield")
                keyType = get_value(search_item, "keyType")
                filter_type = get_value(search_item, "filter")
                code_function = get_value(search_item, "code")
                regex_function = get_value(search_item, "regex")
                full_filter_id = "filter_{}".format(filter_type)
                s_q = ""
                arFkField = []
                if fkfield != None:
                    arFkField = fkfield.split("|")
               
                # Main differentiation: fkfield or dbfield
                if fkfield:
                    # Check for keyS
                    if has_string_value(keyS, oFields):
                        # Check for ID field
                        if has_string_value(keyId, oFields):
                            val = oFields[keyId]
                            if not isinstance(val, int): 
                                try:
                                    val = val.id
                                except:
                                    pass
                            enable_filter(filter_type, head_id)
                            s_q = Q(**{"{}__id".format(fkfield): val})
                        elif has_obj_value(fkfield, oFields):
                            val = oFields[fkfield]
                            enable_filter(filter_type, head_id)
                            s_q = Q(**{fkfield: val})
                        else:
                            val = oFields[keyS]
                            enable_filter(filter_type, head_id)
                            # We are dealing with a foreign key (or multiple)
                            if len(arFkField) > 1:
                                iStop = 1
                            # we are dealing with a foreign key, so we should use keyFk
                            s_q = None
                            if "*" in val or "#" in val:
                                val = adapt_search(val, regex_function)
                            for fkfield in arFkField:
                                if "*" in val or "#" in val:
                                    s_q_add = Q(**{"{}__{}__iregex".format(fkfield, keyFk): val})
                                else:
                                    s_q_add = Q(**{"{}__{}__iexact".format(fkfield, keyFk): val})
                                if s_q == None:
                                    s_q = s_q_add
                                else:
                                    s_q |= s_q_add
                    elif has_obj_value(fkfield, oFields):
                        val = oFields[fkfield]
                        enable_filter(filter_type, head_id)
                        s_q = Q(**{fkfield: val})
                        external = get_value(search_item, "external")
                        if has_string_value(external, oFields):
                            qd[external] = getattr(val, "name")
                elif dbfield:
                    # We are dealing with a plain direct field for the model
                    # OR: it is also possible we are dealing with a m2m field -- that gets the same treatment
                    if keyType == "has":
                        # Check the count or the availability for the db field
                        val = oFields[filter_type]
                        if val == "yes" or val == "no":
                            enable_filter(filter_type, head_id)
                            if val == "yes":
                                s_q = Q(**{"{}__gt".format(dbfield): 0})
                            else:
                                s_q = Q(**{"{}".format(dbfield): 0})
                    elif keyType == "exists" and code_function != None:
                        # Check the count or the availability for the db field
                        val = code_function( oFields[keyS])
                        if val == "yes" or val == "no":
                            enable_filter(filter_type, head_id)
                            if val == "yes":
                                s_q = Q(**{"{}__exact".format(dbfield): ""})
                                if lstExclude == None: lstExclude = []
                                lstExclude.append(s_q)
                                s_q = ""
                            else:
                                s_q = Q(**{"{}__exact".format(dbfield): ""})
                    elif keyType == "fieldchoice" and has_obj_value(keyS, oFields, "FieldChoice"):
                        if infield == None: infield = "abbr"
                        
                        val = getattr(oFields[keyS], infield)
                        s_q = Q(**{"{}".format(dbfield): val})
                        enable_filter(filter_type, head_id)
                    # Check for keyS
                    elif has_string_value(keyS, oFields):
                        # Check for ID field
                        if has_string_value(keyId, oFields):
                            val = oFields[keyId]
                            enable_filter(filter_type, head_id)
                            s_q = Q(**{"{}__id".format(dbfield): val})
                        elif has_obj_value(keyFk, oFields):
                            val = oFields[keyFk]
                            enable_filter(filter_type, head_id)
                            s_q = Q(**{dbfield: val})
                        else:
                            val = oFields[keyS]
                            enable_filter(filter_type, head_id)
                            if isinstance(val, int):
                                s_q = Q(**{"{}".format(dbfield): val})
                            elif "*" in val or "#" in val:
                                val = adapt_search(val, regex_function)
                                s_q = Q(**{"{}__iregex".format(dbfield): val})
                            elif "$" in dbfield:
                                val = adapt_search(val, regex_function)
                            else:
                                s_q = Q(**{"{}__iexact".format(dbfield): val})
                    elif has_Q_value(keyS, oFields) and len(oFields[keyS]) > 0:
                        # See issue #484: if "$" in dbfield:
                        s_q = oFields[keyS]
                        enable_filter(filter_type, head_id)

                elif keyS != "" and has_string_value(keyS, oFields):
                    enable_filter(filter_type, head_id)

                # Check for list of specific signatures
                if has_list_value(keyList, oFields):
                    s_q_lst = ""
                    enable_filter(filter_type, head_id)
                    # Check if this is a Q-expression already
                    if isinstance(oFields[keyList], Q):
                        s_q = oFields[keyList]
                    else:
                        if infield == None: infield = "id"
                        code_list = [getattr(x, infield) for x in oFields[keyList]]
                        if fkfield:
                            # Now we need to look at the id's
                            if len(arFkField) > 1:
                                # THere are more foreign keys: combine in logical or
                                s_q_lst = ""
                                for fkfield in arFkField:
                                    if s_q_lst == "":
                                        s_q_lst = Q(**{"{}__{}__in".format(fkfield, infield): code_list})
                                    else:
                                        s_q_lst |= Q(**{"{}__{}__in".format(fkfield, infield): code_list})
                            else:
                                # Just one foreign key
                                s_q_lst = Q(**{"{}__{}__in".format(fkfield, infield): code_list})
                        elif keyType == "fieldchoice":
                            s_q_lst = Q(**{"{}__in".format(dbfield): code_list})
                        elif dbfield:
                            s_q_lst = Q(**{"{}__in".format(infield): code_list})
                        s_q = s_q_lst if s_q == "" else s_q | s_q_lst

                # Possibly add the result to the list
                if s_q != "": 
                    dictQ[full_filter_id] = s_q
                    # lstQ.append(s_q)
        # Combine the query parts in the appropriate order
        qfilter = qd.get("qfilter")
        if qfilter is None or qfilter == "" or qfilter == "[]":
            for k,v in dictQ.items():
                lstQ.append(v)
        else:
            combi_q = None
            lFilters = json.loads(qfilter)
            if len(lFilters) > 0:
                lst_name = []
                for qf in lFilters:
                    sName = qf.get("name")
                    if not sName in lst_name:
                        lst_name.append(sName)
                        operator = qf.get("operator")
                        s_q = dictQ.get(sName)
                        if not s_q is None:
                            if operator == "start" or combi_q is None:
                                combi_q = s_q
                            elif operator == "and":
                                combi_q = combi_q & (s_q)
                            elif operator == "nand":
                                combi_q = combi_q & (~ s_q)
                            elif operator == "or":
                                combi_q = combi_q | (s_q)
                            elif operator == "nor":
                                combi_q = combi_q | (~ s_q)
                # Now set the lstQ
                if not combi_q is None:
                    lstQ.append(combi_q)
                # Now treat any other filters in dictQ
                for k,v in dictQ.items():
                    if not k in lst_name:
                        lstQ.append(v)

    except:
        msg = oErr.get_error_message()
        oErr.DoError("make_search_list")
        lstQ = []

    # Return what we have created
    return filters, lstQ, qd, lstExclude

def make_ordering(qs, qd, order_default, order_cols, order_heads):

    oErr = ErrHandle()

    try:
        bAscending = True
        sType = 'str'
        order = []
        colnum = ""
        # reset 'used' feature for all heads
        for item in order_heads: item['used'] = None

        # Set the default sort order numbers
        for item in order_heads:
            sorting = ""
            if "order" in item:
                sorting = item['order']
                if "=" in sorting: sorting = sorting.split("=")[1]
            item['sorting'] = sorting
            item['direction'] = ""
            item['priority'] = ""

        # Check out the 'o' parameter...
        if 'o' in qd and qd['o'] != "":
            # Initializations
            order = []
            colnum = qd['o']

            # Get the current 'o' parameter value and turn it into a list of column sortables
            sort_list = [int(x) for x in qd['o'].split(".")]

            # Walk through and implement the sort list
            priority = 1
            for iOrderCol in sort_list:

                bAscending = (iOrderCol>0)
                iOrderCol = abs(iOrderCol)

                # Set the column that it is in use
                order_heads[iOrderCol-1]['used'] = 1
                # Set priority and direction
                if len(sort_list) > 1:
                    order_heads[iOrderCol-1]['priority'] = priority
                    priority += 1
                order_heads[iOrderCol-1]['direction'] = "up" if bAscending else "down"

                # Get the type
                sType = order_heads[iOrderCol-1]['type']
                for order_item in order_cols[iOrderCol-1].split(";"):
                    if order_item != "":
                        if sType == 'str':
                            if bAscending:
                                order.append(Lower(order_item).asc(nulls_last=True))
                            else:
                                order.append(Lower(order_item).desc(nulls_last=True))
                        else:
                            if bAscending:
                                order.append(F(order_item).asc(nulls_last=True))
                            else:
                                order.append(F(order_item).desc(nulls_last=True))
 
            # Adapt the 'sorting' parameter for all heads that need it
            for order_head in order_heads:
                # Get the current default sorting (the column number)
                sorting_default = item['sorting']
                # Is this one sortable?
                if 'order' in order_head and '=' in order_head['order']:
                    # Get the column number
                    col_num = int(order_head['order'].split("=")[1])
                    col_num_neg = -1 * col_num
                    order_combined = [str(x) for x in sort_list]
                    # Is this column in the sort_list or not?
                    if col_num in sort_list or col_num_neg in sort_list:
                        # This column is in the sort list: suggest the negation of what is there
                        for idx, order_one in enumerate(order_combined):
                            if abs(int(order_one)) == col_num:
                                order_combined[idx] = str(-1 * int(order_one))
                                break
                    else:
                        # This colum is not in the sort list: just combine
                        order_combined.append(str(col_num))
                    order_head['sorting'] = ".".join(order_combined)
        else:
            orderings = []
            for idx, order_item in enumerate(order_default):
                #if idx == 0 and order_item[0] == "-":
                #    bAscending = False
                #    order_item = order_item[1:]
                # Get the type
                sType = order_heads[idx]['type']
                if ";" in order_item:
                    for sub_item in order_item.split(";"):
                        orderings.append(dict(type=sType, item=sub_item))
                else:
                    orderings.append(dict(type=sType, item=order_item))
            for item in orderings:
                sType = item['type']
                order_item = item['item']
                if order_item != "":
                    if order_item[0] == "-":
                        bAscending = False
                        order_item = order_item[1:]
                    else:
                        bAscending = True
                    # USED TO BE:
                    #if sType == "int":
                    #    order.append(order_item)
                    #else:
                    #    order.append(Lower(order_item))
                    if bAscending:
                        if sType == "str":
                            order.append(Lower(order_item).asc(nulls_last=True))
                        else:
                            order.append(F(order_item).asc(nulls_last=True))
                    else:
                        if sType == "str":
                            order.append(Lower(order_item).desc(nulls_last=True))
                        else:
                            order.append(F(order_item).desc(nulls_last=True))


           #  order.append(Lower(order_cols[0]))
        if sType == 'str':
            if len(order) > 0:
                qs = qs.order_by(*order)
        else:
            qs = qs.order_by(*order)
        ## Possibly reverse the order
        #if not bAscending:
        #    qs = qs.reverse()
    except:
        msg = oErr.get_error_message()
        # Show what the order_default and the order_heads is (as strings)
        oErr.Status("order_heads = {}\norder_default = {}".format(
            json.dumps(order_heads), json.dumps(order_default)))
        oErr.DoError("basic/view/make_ordering")
        lstQ = []

    return qs, order_heads, colnum

def make_ordering_original(qs, qd, order_default, order_cols, order_heads):

    oErr = ErrHandle()

    try:
        bAscending = True
        sType = 'str'
        order = []
        colnum = ""
        # reset 'used' feature for all heads
        for item in order_heads: item['used'] = None
        if 'o' in qd and qd['o'] != "":
            colnum = qd['o']
            if '=' in colnum:
                colnum = colnum.split('=')[1]
            if colnum != "":
                order = []
                iOrderCol = int(colnum)
                bAscending = (iOrderCol>0)
                iOrderCol = abs(iOrderCol)

                # Set the column that it is in use
                order_heads[iOrderCol-1]['used'] = 1
                # Get the type
                sType = order_heads[iOrderCol-1]['type']
                for order_item in order_cols[iOrderCol-1].split(";"):
                    if order_item != "":
                        if sType == 'str':
                            order.append(Lower(order_item).asc(nulls_last=True))
                        else:
                            order.append(F(order_item).asc(nulls_last=True))
                if bAscending:
                    order_heads[iOrderCol-1]['order'] = 'o=-{}'.format(iOrderCol)
                else:
                    # order = "-" + order
                    order_heads[iOrderCol-1]['order'] = 'o={}'.format(iOrderCol)

                # Reset the sort order to ascending for all others
                for idx, item in enumerate(order_heads):
                    if idx != iOrderCol - 1:
                        # Reset this sort order
                        order_heads[idx]['order'] = order_heads[idx]['order'].replace("-", "")
        else:
            orderings = []
            for idx, order_item in enumerate(order_default):
                if idx == 0 and order_item[0] == "-":
                    bAscending = False
                    order_item = order_item[1:]
                # Get the type
                sType = order_heads[idx]['type']
                if ";" in order_item:
                    for sub_item in order_item.split(";"):
                        orderings.append(dict(type=sType, item=sub_item))
                else:
                    orderings.append(dict(type=sType, item=order_item))
            for item in orderings:
                sType = item['type']
                order_item = item['item']
                if order_item != "":
                    if sType == "int":
                        order.append(order_item)
                    else:
                        order.append(Lower(order_item))

           #  order.append(Lower(order_cols[0]))
        if sType == 'str':
            if len(order) > 0:
                qs = qs.order_by(*order)
        else:
            qs = qs.order_by(*order)
        # Possibly reverse the order
        if not bAscending:
            qs = qs.reverse()
    except:
        msg = oErr.get_error_message()
        oErr.DoError("make_ordering_original")
        lstQ = []

    return qs, order_heads, colnum

def add_rel_item(rel_item, value, resizable=False, title=None, align=None, link=None, nowrap=True, main=None, draggable=None):
    oAdd = dict(value=value)
    if resizable: oAdd['initial'] = 'small'
    if title != None: oAdd['title'] = title
    if align != None: oAdd['align'] = align
    if link != None: oAdd['link'] = link
    if main != None: oAdd['main'] = main
    if nowrap != None: oAdd['nowrap'] = nowrap
    if draggable != None: oAdd['draggable'] = draggable
    rel_item.append(oAdd)
    return True

def base64_encode(sInput):
    message_bytes = sInput.encode("utf8")
    base64_bytes = base64.b64encode(message_bytes)
    sOutput = base64_bytes.decode("utf8")
    return sOutput

def base64_decode(sInput):
    base64_bytes = sInput.encode('utf8')
    message_bytes = base64.b64decode(base64_bytes)
    sOutput = message_bytes.decode('utf8')
    return sOutput

def get_current_datetime():
    """Get the current time"""
    return timezone.now()

def treat_bom(sHtml):
    """REmove the BOM marker except at the beginning of the string"""

    # Check if it is in the beginning
    bStartsWithBom = sHtml.startswith(u'\ufeff')
    # Remove everywhere
    sHtml = sHtml.replace(u'\ufeff', '')
    # Return what we have
    return sHtml

def adapt_m2m(cls, instance, field1, qs, field2, extra = [], extrargs = {}, qfilter = {}, 
              related_is_through = False, userplus = None, added=None, deleted=None):
    """Adapt the 'field' of 'instance' to contain only the items in 'qs'
    
    The lists [added] and [deleted] (if specified) will contain links to the elements that have been added and deleted
    If [deleted] is specified, then the items will not be deleted by adapt_m2m(). Caller needs to do this.
    """

    errHandle = ErrHandle()
    try:
        # Get current associations
        lstQ = [Q(**{field1: instance})]
        for k,v in qfilter.items(): lstQ.append(Q(**{k: v}))
        through_qs = cls.objects.filter(*lstQ)
        if related_is_through:
            related_qs = through_qs
        else:
            related_qs = [getattr(x, field2) for x in through_qs]
        # make sure all items in [qs] are associated
        if userplus == None or userplus:
            for obj in qs:
                if obj not in related_qs:
                    # Add the association
                    args = {field1: instance}
                    if related_is_through:
                        args[field2] = getattr(obj, field2)
                    else:
                        args[field2] = obj
                    for item in extra:
                        # Copy the field with this name from [obj] to 
                        args[item] = getattr(obj, item)
                    for k,v in extrargs.items():
                        args[k] = v
                    # cls.objects.create(**{field1: instance, field2: obj})
                    new = cls.objects.create(**args)
                    if added != None:
                        added.append(new)

        # Remove from [cls] all associations that are not in [qs]
        # NOTE: do not allow userplus to delete
        for item in through_qs:
            if related_is_through:
                obj = item
            else:
                obj = getattr(item, field2)
            if obj not in qs:
                if deleted == None:
                    # Remove this item
                    item.delete()
                else:
                    deleted.append(item)
        # Return okay
        return True
    except:
        msg = errHandle.get_error_message()
        return False

def adapt_m2o(cls, instance, field, qs, link_to_obj = None, **kwargs):
    """Adapt the instances of [cls] pointing to [instance] with [field] to only include [qs] """

    errHandle = ErrHandle()
    try:
        # Get all the [cls] items currently linking to [instance]
        lstQ = [Q(**{field: instance})]
        linked_qs = cls.objects.filter(*lstQ)
        if link_to_obj != None:
            linked_through = [getattr(x, link_to_obj) for x in linked_qs]
        # make sure all items in [qs] are linked to [instance]
        for obj in qs:
            if (obj not in linked_qs) and (link_to_obj == None or obj not in linked_through):
                # Create new object
                oNew = cls()
                setattr(oNew, field, instance)
                # Copy the local fields
                for lfield in obj._meta.local_fields:
                    fname = lfield.name
                    if fname != "id" and fname != field:
                        # Copy the field value
                        setattr(oNew, fname, getattr(obj, fname))
                for k, v in kwargs.items():
                    setattr(oNew, k, v)
                # Need to add an object link?
                if link_to_obj != None:
                    setattr(oNew, link_to_obj, obj)
                oNew.save()
        # Remove links that are not in [qs]
        for obj in linked_qs:
            if obj not in qs:
                # Remove this item
                obj.delete()
        # Return okay
        return True
    except:
        msg = errHandle.get_error_message()
        return False

def adapt_m2o_sig(instance, qs):
    """Adapt the instances of [SermonSignature] pointing to [instance] to only include [qs] 
    
    Note: convert SermonSignature into (Gold) Signature
    """

    errHandle = ErrHandle()
    try:
        # Get all the [SermonSignature] items currently linking to [instance]
        linked_qs = SermonSignature.objects.filter(sermon=instance)
        # make sure all items in [qs] are linked to [instance]
        bRedo = False
        for obj in qs:
            # Get the SermonSignature equivalent for Gold signature [obj]
            sermsig = instance.get_sermonsig(obj)
            if sermsig not in linked_qs:
                # Indicate that we need to re-query
                bRedo = True
        # Do we need to re-query?
        if bRedo: 
            # Yes we do...
            linked_qs = SermonSignature.objects.filter(sermon=instance)
        # Remove links that are not in [qs]
        for obj in linked_qs:
            # Get the gold-signature equivalent of this sermon signature
            gsig = obj.get_goldsig()
            # Check if the gold-sermon equivalent is in [qs]
            if gsig not in qs:
                # Remove this item
                obj.delete()
        # Return okay
        return True
    except:
        msg = errHandle.get_error_message()
        return False

def csv_to_excel(sCsvData, response):
    """Convert CSV data to an Excel worksheet"""

    # Start workbook
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title="Data"

    # Start accessing the string data 
    f = StringIO(sCsvData)
    reader = csv.reader(f, delimiter=",")

    # Read the header cells and make a header row in the worksheet
    headers = next(reader)
    for col_num in range(len(headers)):
        c = ws.cell(row=1, column=col_num+1)
        c.value = headers[col_num]
        c.font = openpyxl.styles.Font(bold=True)
        # Set width to a fixed size
        ws.column_dimensions[get_column_letter(col_num+1)].width = 5.0        

    row_num = 1
    lCsv = []
    for row in reader:
        # Keep track of the EXCEL row we are in
        row_num += 1
        # Walk the elements in the data row
        # oRow = {}
        for idx, cell in enumerate(row):
            c = ws.cell(row=row_num, column=idx+1)
            c.value = row[idx]
            c.alignment = openpyxl.styles.Alignment(wrap_text=False)
    # Save the result in the response
    wb.save(response)
    return response




# The views that are defined by 'basic'

class BasicList(ListView):
    """Basic listview
    
    This listview inherits the standard listview and adds a few automatic matters
    """

    paginate_by = paginateSize # 15
    entrycount = 0
    qd = None
    bFilter = False
    basketview = False
    template_name = 'basic/basic_list.html'
    template_help = 'basic/filter_help.html'
    bHasParameters = False
    bUseFilter = False
    new_button = True
    sel_button = False
    sel_count = ""
    initial = None
    listform = None
    has_select2 = False
    plural_name = ""
    sg_name = ""
    basic_name = ""
    basic_name_prefix = ""
    basic_edit = ""
    basic_details = ""
    basic_add = ""
    basic_filter = None
    add_text = "Add a new"
    prefix = ""
    order_default = []
    order_cols = []
    order_heads = []
    filters = []
    searches = []
    downloads = []
    custombuttons = []
    selectbuttons = []
    list_fields = []
    uploads = []
    delete_line = False
    none_on_empty = False
    use_team_group = False
    admin_editable = False
    permission = True
    usersearch_id = ""
    redirectpage = ""
    lst_typeaheads = []
    sort_order = ""
    col_wrap = ""
    sel_mode = ""
    param_list = []
    qfilter = []
    qs = None
    page_function = "ru.basic.search_paged_start"

    def initializations(self):
        return None

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super(BasicList, self).get_context_data(**kwargs)

        oErr = ErrHandle()

        # self.before_context()

        # Get parameters for the search
        if self.initial == None:
            initial = self.request.POST if self.request.POST else self.request.GET
        else:
            initial = self.initial

        # Need to load the correct form
        if self.listform:
            prefix = "" if self.prefix == "any" else self.prefix
            if self.use_team_group:
                frm = self.listform(initial, prefix=self.prefix, username=self.request.user.username, team_group=app_editor, userplus=app_userplus)
            else:
                frm = self.listform(initial, prefix=self.prefix)
            if frm.is_valid():
                context['{}Form'.format(self.prefix)] = frm
                # Get any possible typeahead parameters
                lst_form_ta = getattr(frm, "typeaheads", None)
                if lst_form_ta != None:
                    for item in lst_form_ta:
                        self.lst_typeaheads.append(item)

            if self.has_select2:
                context['has_select2'] = True
            context['listForm'] = frm

        # Determine the count 
        context['entrycount'] = self.entrycount # self.get_queryset().count()

        # Make sure the paginate-values are available
        context['paginateValues'] = paginateValues

        if 'paginate_by' in initial:
            context['paginateSize'] = int(self.get_paginate_size()) #  int(initial['paginate_by'])
        else:
            context['paginateSize'] = paginateSize

        # Need to pass on a pagination function
        if self.page_function:
            context['page_function'] = self.page_function

        # Set the page number if needed
        if 'page_obj' in context and 'page' in initial and initial['page'] != "":
            # context['page_obj'].number = initial['page']
            page_num = int(initial['page'])
            context['page_obj'] = context['paginator'].page( page_num)
            # Make sure to adapt the object_list
            context['object_list'] = context['page_obj']
            self.param_list.append("page={}".format(page_num))
        # Make sure the parameter list becomes available
        params = ""
        if len(self.param_list) > 0:
            params = base64_encode( "&".join(self.param_list))
        context['params'] = params

        # Set the title of the application
        if self.plural_name =="":
            self.plural_name = str(self.model._meta.verbose_name_plural)
        context['title'] = self.plural_name
        if self.basic_name == "":
            if self.basic_name_prefix == "":
                self.basic_name = str(self.model._meta.model_name)
            else:
                self.basic_name = "{}{}".format(self.basic_name_prefix, self.prefix)
        context['titlesg'] = self.sg_name if self.sg_name != "" else self.basic_name.capitalize()
        context['basic_name'] = self.basic_name
        if self.basic_add:
            basic_add = reverse(self.basic_add)
        else:
            basic_add = reverse("{}_details".format(self.basic_name))
        context['basic_add'] = basic_add
        context['basic_list'] = reverse("{}_list".format(self.basic_name))
        context['basic_edit'] = self.basic_edit if self.basic_edit != "" else "{}_edit".format(self.basic_name)
        context['basic_details'] = self.basic_details if self.basic_details != "" else "{}_details".format(self.basic_name)

        # Make sure we get selection mode right early enough...
        selmode = self.qd.get("s", None)
        if selmode != None:
            self.sel_mode = selmode.strip()

        # Make sure to transform the 'object_list'  into a 'result_list'
        context['result_list'] = self.get_result_list(context['object_list'], context)

        context['sortOrder'] = self.sort_order
        context['colWrap'] = self.col_wrap
        context['selMode'] = self.sel_mode

        context['new_button'] = self.new_button
        context['sel_button'] = self.sel_button
        context['sel_count'] = self.sel_count
        context['add_text'] = self.add_text

        context['admin_editable'] = self.admin_editable

        # Adapt possible downloads
        if len(self.downloads) > 0:
            for item in self.downloads:
                if 'url' in item and item['url'] != "" and "/" not in item['url']:
                    item['url'] = reverse(item['url'])
            context['downloads'] = self.downloads

        # Specify possible upload
        if len(self.uploads) > 0:
            for item in self.uploads:
                if 'url' in item and item['url'] != "" and "/" not in item['url']:
                    item['url'] = reverse(item['url'])
            context['uploads'] = self.uploads

        # Custom buttons
        if len(self.custombuttons) > 0:
            for item in self.custombuttons:
                if 'template_name' in item:
                    # get the code of the template
                    pass
            context['custombuttons'] = self.custombuttons

        # Selection buttons
        if len(self.selectbuttons) > 0:
            context['selectbuttons'] = self.selectbuttons

        # Delete button per line?
        if self.delete_line:
            context['delete_line'] = True

        # Make sure we pass on the ordered heads
        context['order_heads'] = self.order_heads
        context['has_filter'] = self.bFilter
        fsections = []
        # Initialize the adapted filters
        for filteritem in self.filters:
            filteritem['fitems'] = []
            filteritem['count'] = 0
            filteritem['hasvalue'] = False
        # Adapt filters with the information from searches
        for section in self.searches:
            oFsection = {}
            bHasValue = False
            section_name = section['section']
            if section_name != "" and section_name not in fsections:          
                oFsection = dict(name=section_name, has_value=False)
                # fsections.append(dict(name=section_name))
            # Copy the relevant search filter
            for item in section['filterlist']:
                bHasItemValue = False
                # Find the corresponding item in the filters
                id = "filter_{}".format(item['filter'])
                for filteritem in self.filters:
                    if id == filteritem['id']:
                        try:
                            # Build a new [fitem]
                            fitem = {}
                            fitem['search'] = item
                            fitem['has_keylist'] = False
                            # Add possible fields
                            if 'keyS' in item and item['keyS'] in frm.cleaned_data: 
                                fitem['keyS'] = frm[item['keyS']]
                                if fitem['keyS'].value(): 
                                    bHasValue = True ; bHasItemValue = True
                            if 'keyList' in item and item['keyList'] in frm.cleaned_data: 
                                if frm.fields[item['keyList']].initial or frm.cleaned_data[item['keyList']].count() > 0: 
                                    bHasValue = True ; bHasItemValue = True
                                fitem['keyList'] = frm[item['keyList']]
                                fitem['has_keylist'] = True
                            if 'keyS' in item and item['keyS'] in frm.cleaned_data: 
                                if 'dbfield' in item and item['dbfield'] in frm.cleaned_data and item['keyS'] != item['dbfield']:
                                    fitem['dbfield'] = frm[item['dbfield']]
                                    if fitem['dbfield'].value(): 
                                        bHasValue = True ; bHasItemValue = True
                                elif 'fkfield' in item and item['fkfield'] in frm.cleaned_data:
                                    fitem['fkfield'] = frm[item['fkfield']]                                    
                                    if fitem['fkfield'].value(): bHasValue = True ; bHasItemValue = True
                                else:
                                    # There is a keyS without corresponding fkfield or dbfield
                                    pass
                            # Append the [fitem] to the [fitems]                            
                            filteritem['fitems'].append(fitem)
                            filteritem['count'] = len(filteritem['fitems'])
                            filteritem['help'] = ""
                            # Possibly add help
                            if 'help' in item:
                                filteritem['helptext'] = self.get_helptext(item['help']) 
                                filteritem['help'] = item['help']
                            # Make sure we indicate that there is a value
                            if bHasItemValue: filteritem['hasvalue'] = True                            

                             # If this is a hidden one, then set 'hasvalue' at the *FIRST* appropriate one TH: hier gaat het niet goed
                            if bHasItemValue:                                 
                                filterinclude = filteritem.get('include_id', "")
                                if filteritem.get('head_id', '') == "hidden":
                                    # Find the first appropriate one including me
                                    for fi in self.filters:
                                        if fi.get("include_id", "") == id:
                                            # We have the first one
                                            fi['hasvalue'] = True
                                            # fi['enabled'] = True
                                            # Now leave the for-loop
                                            break
                                elif filterinclude != "":
                                    # Make sure to enable the include, even though it may not have a value
                                    for fi in self.filters:
                                        if fi.get('id', '') == filterinclude:
                                            fi['hasvalue'] = True
                            break
                        except:
                            sMsg = oErr.get_error_message()
                            break
            if bHasValue: 
                oFsection['has_value'] = True
            if oFsection != None: fsections.append(oFsection)

        # Make it available
        context['filters'] = self.filters
        context['fsections'] = fsections
        context['list_fields'] = self.list_fields
        context['qfilter'] = self.qfilter

        # Add any typeaheads that should be initialized
        context['typeaheads'] = json.dumps( self.lst_typeaheads)

        # Get help for filtering
        context['filterhelp_contents'] = self.get_filterhelp()

        # Check if user may upload
        context['is_authenticated'] = user_is_authenticated(self.request)
        context['authenticated'] = context['is_authenticated'] 
        context['is_app_uploader'] = user_is_ingroup(self.request, app_uploader)
        context['is_app_user'] = user_is_ingroup(self.request, app_user)
        context['is_app_editor'] = user_is_ingroup(self.request, app_editor)
        context['is_app_userplus'] = user_is_ingroup(self.request, app_userplus)
        context['is_app_moderator'] = user_is_superuser(self.request) or user_is_ingroup(self.request, app_moderator)

        # Process this visit and get the new breadcrumbs object
        prevpage = reverse('home')
        context['prevpage'] = prevpage
        context['breadcrumbs'] = get_breadcrumbs(self.request, self.plural_name, True)

        context['usebasket'] = self.basketview

        context['permission'] = self.permission

        # Fill in the selection information
        selectitem_info = ""
        if not self.sel_button is None and self.sel_button != "":
            # Prepare selected item handling
            selectitem_info = self.get_selectitem_info(None, context)

        context['sel_info'] = selectitem_info

        # Add the search url
        if self.usersearch_id == "":
            context['usersearch'] = ""
        else:
            context['usersearch'] = "{}://{}{}?usersearch={}".format(
                self.request.scheme, self.request.get_host(), self.request.path, self.usersearch_id)
        context['usersearch_id'] = self.usersearch_id

        # Allow others to add to context
        context = self.add_to_context(context, initial)
        # x = context['is_app_editor'] and context['new_button']

        # Return the calculated context
        return context

    def add_to_context(self, context, initial):
        return context

    def get_result_list(self, obj_list, context):
        result_list = []
        # Walk all items in the object list
        for obj in obj_list:
            # Transform this object into a list of objects that can be shown
            result = dict(id=obj.id)
            fields = []
            for head in self.order_heads:
                fobj = dict(value="")
                fname = None
                if 'field' in head:
                    # This is a field that can be shown
                    fname = head['field']
                    default = "" if 'default' not in head else head['default']
                    value = getattr(obj, fname, default)
                    if not value is None:
                        fobj['value'] = value
                elif 'custom' in head:
                    # The user should provide a way to determine the value for this field
                    fvalue, ftitle = self.get_field_value(obj, head['custom'])
                    if not fvalue is None:
                        fobj['value']= fvalue
                    if ftitle != None:
                        fobj['title'] = ftitle
                    fname = head['custom']
                classes = []
                if fname != None: classes.append("{}-{}".format(self.basic_name, fname))
                if 'linkdetails' in head and head['linkdetails']: fobj['linkdetails'] = True
                if 'main' in head and head['main']:
                    fobj['styles'] = "width: 100%;"
                    fobj['main'] = True
                    if self.delete_line:
                        classes.append("ms editable")
                elif 'options' in head and len(head['options']) > 0:
                    options = head['options']
                    if 'delete' in options:
                        fobj['delete'] = True
                    else:
                        fobj['styles'] = "min-width: {}px;".format(50 * len(options))
                        if not 'allowwrap' in head or not head['allowwrap']:
                            classes.append("tdnowrap")
                else:
                    if 'width' in head and len(head['width']) > 0:
                        fobj['styles'] = "width: {};".format(head['width'])
                    else:
                        fobj['styles'] = "width: 100px;"
                    if not 'allowwrap' in head or not head['allowwrap']:
                        classes.append("tdnowrap")
                if 'align' in head and head['align'] != "":
                    fobj['align'] = head['align'] 
                fobj['classes'] = " ".join(classes)
                if 'colwrap' in head:
                    fobj['colwrap'] = True
                if 'autohide' in head:
                    fobj['autohide'] = head['autohide']
                fields.append(fobj)
            # Make the list of field-values available
            result['fields'] = fields
            admindetails = "admin:seeker_{}_change".format(self.basic_name)
            try:
                result['admindetails'] = reverse(admindetails, args=[obj.id])
            except:
                pass

            # Fill in the selection information
            selectitem_info = ""
            if not self.sel_button is None and self.sel_button != "":
                # Prepare selected item handling
                selectitem_info = self.get_selectitem_info(obj, context)

            result['sel_info'] = selectitem_info

            # Add to the list of results
            result_list.append(result)
        return result_list

    def get_helptext(self, name):
        return ""

    def get_filterhelp(self):
        sBack = "(no help available)"
        if self.template_help != None and self.template_help != "":
            sBack = render_to_string(self.template_help)
        return sBack

    def get_template_names(self):
        names = [ self.template_name ]
        return names

    def get_field_value(self, instance, custom):
        return "", ""

    def get_paginate_by(self, queryset):
        """
        Paginate by specified value in default class property value.
        """
        paginate_by = self.get_paginate_size()
        return paginate_by

    def get_paginate_size(self):
        """Get the correct size of the pages"""
        initial = self.request.POST if self.request.POST else self.request.GET
        page_size = initial.get('paginate_by', self.paginate_by)
        # Double check the value that we have received
        iNumber = get_number(page_size)
        if iNumber < 0 or iNumber > PaginateMax:
            # Just take the default number
            page_size = self.paginate_by
        else:
            # Yes, take this number
            page_size = str(iNumber)
        return page_size

    def get_basketqueryset(self):
        """User-specific function to get a queryset based on a basket"""
        return None

    def adapt_search(self, fields):
        return fields, None, None

    def get_queryset(self, request = None):
        """Calculate the queryset that should be used"""

        qs = None
        oErr = ErrHandle()
        try:
            if request == None: request = self.request
            get = self.qd

            # Immediately take care of the rangeslider stuff
            lst_remove = []
            for k,v in self.qd.items():
                if "-rangeslider" in k: lst_remove.append(k)
            for item in lst_remove: self.qd.pop(item)

            # username=self.request.user.username
            username=request.user.username
            team_group=app_editor

            self.bFilter = False
            self.bHasParameters = (len(get) > 0)
            bHasListFilters = False
            if self.bHasParameters:
                # y = [x for x in get ]
                bHasListFilters = len([x for x in get if self.prefix in x and get[x] != ""]) > 0
                if not bHasListFilters:
                    self.basketview = ("usebasket" in get and get['usebasket'] == "True")
            # At least get the qFilter
            qfilter = get.get("qfilter")
            if qfilter is None or qfilter == "":
                qfilter = []
            else:
                qfilter = json.loads(qfilter)
            self.qFilter = qfilter

            # Initial setting of qs
            qs = self.model.objects.none()

            # Get the queryset and the filters
            if self.basketview:
                self.basketview = True
                # We should show the contents of the basket
                # (1) Reset the filters
                for item in self.filters: item['enabled'] = False
                # (2) Indicate we have no filters
                self.bFilter = False
                # (3) Set the queryset -- this is listview-specific
                qs = self.get_basketqueryset()

                # Do the ordering of the results
                order = self.order_default
                qs, self.order_heads, colnum = make_ordering(qs, self.qd, order, self.order_cols, self.order_heads)
            elif self.bHasParameters or self.bUseFilter:
                self.basketview = False
                lstQ = []
                # Indicate we have no filters
                self.bFilter = False

                # Read the form with the information
                prefix = self.prefix
                if prefix == "any": prefix = ""
                if self.use_team_group:
                    thisForm = self.listform(self.qd, prefix=prefix, username=username, team_group=team_group)
                else:
                    thisForm = self.listform(self.qd, prefix=prefix)

                if thisForm.is_valid():
                    # Process the criteria for this form
                    oFields = thisForm.cleaned_data

                    # Set the param_list variable
                    # self.param_list = []
                    param_list = []
                    lookfor = "{}-".format(prefix)
                    for k,v in self.qd.items():
                        if lookfor in k and not isempty(v):
                            # self.param_list.append("{}={}".format(k,v))
                            param_list.append("{}={}".format(k,v))

                    if 'qfilter' in self.qd:
                        self.qfilter = self.qd.get("qfilter")

                    # Store the paramlist - but only if this is not a repetition
                    if "usersearch" in self.qd:
                        # Make sure we have the user search number
                        self.usersearch_id = self.qd.get("usersearch")
                    else:
                        # oSearch = UserSearch.add_search(request.path, self.param_list, request.user.username, self.qfilter)
                        oSearch = UserSearch.add_search(request.path, param_list, request.user.username, self.qfilter)
                        if oSearch != None:
                            self.usersearch_id = oSearch.id
                    # Make sure to add the usersearch into the paramlist
                    self.param_list.append("usersearch={}".format(self.usersearch_id))
                
                    # Allow user to adapt the list of search fields
                    oFields, lstExclude, qAlternative = self.adapt_search(oFields)

                    self.filters, lstQ, self.initial, lstExclude = make_search_list(self.filters, oFields, self.searches, self.qd, lstExclude)
                    # qs = self.model.objects.filter(manuitems__itemsermons__goldsermons__goldsignatures__code__in = "AN Mt h 42")

                    # Combine exclude filters with logical or
                    exclude = None
                    if not lstExclude is None and len(lstExclude) > 0:
                        exclude = lstExclude[0]
                        for expr in lstExclude[1:]:
                            exclude |= expr

                    # Calculate the final qs
                    if len(lstQ) == 0 and not self.none_on_empty:
                        if lstExclude:
                            if qAlternative:
                                # qs = self.model.objects.filter(qAlternative).exclude(*lstExclude).distinct()
                                qs = self.model.objects.filter(qAlternative).exclude(exclude).distinct()
                            else:
                                # qs = self.model.objects.exclude(*lstExclude)
                                qs = self.model.objects.exclude(exclude)
                        else:
                            if qAlternative:
                                qs = self.model.objects.filter(qAlternative).distinct()
                            else:
                                # Just show everything
                                qs = self.model.objects.all()
                    else:
                        # There is a filter, so build it up
                        filter = lstQ[0]
                        for item in lstQ[1:]:
                            filter = filter & item
                        if qAlternative:
                            filter = ( ( qAlternative ) & filter )

                        # Check if excluding is needed
                        if lstExclude:
                            # qs = self.model.objects.filter(filter).exclude(*lstExclude).distinct()
                            qs = self.model.objects.filter(filter).exclude(exclude).distinct()
                        else:
                            qs = self.model.objects.filter(filter).distinct()

                    # Only set the [bFilter] value if there is an overt specified filter
                    for filter in self.filters:
                        if filter['enabled'] and ('head_id' not in filter or filter['head_id'] != 'filter_other'):
                            self.bFilter = True
                            break
                        # OLD self.bFilter = True
                elif not self.none_on_empty:
                    # Provide an error message for the LOG...
                    print('Form error WARNING: {}'.format(thisForm.errors))

                    # Just show everything
                    qs = self.model.objects.all().distinct()

                # Do the ordering of the results
                order = self.order_default
                qs, self.order_heads, colnum = make_ordering(qs, self.qd, order, self.order_cols, self.order_heads)

                # Adapt order_heads 'autohide' if a column has a filter set
                for oOrderHead in self.order_heads:
                    if 'filter' in oOrderHead:
                        sFilterId = oOrderHead['filter'] 
                        # Initial: on
                        oOrderHead['autohide'] = "on"
                        # Look for the correct filter
                        for oFilter in self.filters:                        
                            if oFilter['id'] == sFilterId:
                                # We found the filter - is it being used? 
                                if oFilter['enabled']:                      
                                    # It is used, so make sure to switch OFF the autohide
                                    oOrderHead['autohide'] = "off"                                 

            else:
                # No filter and no basket: show all
                self.basketview = False
                if self.basic_filter:
                    qs = self.model.objects.filter(self.basic_filter).distinct()
                else:
                    qs = self.model.objects.all().distinct()
                order = self.order_default
                qs, tmp_heads, colnum = make_ordering(qs, self.qd, order, self.order_cols, self.order_heads)
            self.sort_order = colnum

            # Process column wrapping...
            for oHead in self.order_heads:
                if 'colwrap' in oHead:
                    del oHead['colwrap']
            colwrap = self.qd.get("w", None)
            if colwrap != None:
                self.col_wrap = colwrap.strip()
                if colwrap != "" and colwrap[0] == "[":
                    # Process the column wrapping
                    lColWrap = json.loads(colwrap)
                    for idx, oHead in enumerate(self.order_heads):
                        # if idx+1 in lColWrap:
                        if idx in lColWrap:
                            # Indicate that this column must be hidden
                            oHead['colwrap'] = True

            # Determine the length
            self.entrycount = 0 if qs is None else qs.count()   # len(qs)
        
            # Allow doing something additionally with the queryset
            self.view_queryset(qs)

            # Return the resulting filtered and sorted queryset
            self.qs = qs
        except:
            msg = oErr.get_error_message()
            oErr.DoError("BasicList/get_queryset")

        return qs

    def get_selectitem_info(self, instance, context):
        return ""

    def view_queryset(self, qs):
        return None

    def get(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            # Do not allow to get a good response
            response = redirect(reverse('nlogin'))
        else:
            # FIrst do my own initializations
            self.initializations()

            # Get the parameters passed on with the GET or the POST request
            get = request.GET if request.method == "GET" else request.POST
            get = get.copy()
            # If this is a 'usersearch' then replace the parameters
            usersearch_id = get.get("usersearch")
            if usersearch_id != None and usersearch_id != "":
                get = UserSearch.load_parameters(usersearch_id, get)
            self.qd = get
            self.param_list = []

            # Then check if we have a redirect or not
            if self.redirectpage == "":
                # We can continue with the normal 'get()'
                response = super(BasicList, self).get(request, *args, **kwargs)
            else:
                response = redirect(self.redirectpage)
        # REturn the appropriate response
        return response

    def post(self, request, *args, **kwargs):
        return self.get(request, *args, **kwargs)
    

class BasicDetails(DetailView):
    """Extension of the normal DetailView class for PASSIM"""

    template_name = ""      # Template for GET
    template_post = ""      # Template for POST
    formset_objects = []    # List of formsets to be processed
    form_objects = []       # List of form objects
    afternewurl = ""        # URL to move to after adding a new item
    prefix = ""             # The prefix for the one (!) form we use
    previous = None         # Start with empty previous page
    title = ""              # The title to be passed on with the context
    titlesg = None          # Alternative title in singular
    rtype = "json"          # JSON response (alternative: html)
    prefix_type = ""        # Whether the adapt the prefix or not ('simple')
    mForm = None            # Model form
    basic_name = None
    basic_name_prefix = ""
    basic_add = ""
    add_text = "Add a new"
    permission = "read"     # Permission can be: (nothing), "read", "readonly", "write"
    new_button = False
    do_not_save = False
    no_delete = False
    afterdelurl = None
    listview = None
    listviewtitle = None
    has_select2 = False
    backbutton = True
    bNeedReload = False     # Needed to signal a Ctrl+F5 reload for JS
    custombuttons = []
    selectbuttons = []
    sel_button = None
    newRedirect = False     # Redirect the page name to a correct one after creating
    initRedirect = False    # Perform redirect right after initializations
    use_team_group = False
    redirectpage = ""       # Where to redirect to
    add = False             # Are we adding a new record or editing an existing one?
    is_basic = True         # Is this a basic details/edit view?
    history_button = False  # Show history button for this view
    comment_button = False  # Show user comment button for this view
    comment_count = None
    lst_typeahead = []

    def get(self, request, pk=None, *args, **kwargs):
        # Initialisation
        data = {'status': 'ok', 'html': '', 'statuscode': ''}
        # always do this initialisation to get the object
        self.initializations(request, pk)
        if not request.user.is_authenticated:
            # Do not allow to get a good response
            if self.rtype == "json":
                data['html'] = "(No authorization)"
                data['status'] = "error"

                response = JsonResponse(data)
            else:
                response = redirect(reverse('nlogin'))
        else:
            context = self.get_context_data(object=self.object)

            if self.is_basic and self.template_name == "":
                if self.rtype == "json":
                    self.template_name = "basic/basic_edit.html"
                else:
                    self.template_name = "basic/basic_details.html"
            # Possibly indicate form errors
            # NOTE: errors is a dictionary itself...
            if 'errors' in context and len(context['errors']) > 0:
                data['status'] = "error"
                data['msg'] = context['errors']

            if self.rtype == "json":
                # We render to the _name 
                sHtml = render_to_string(self.template_name, context, request)
                sHtml = sHtml.replace("\ufeff", "")
                data['html'] = sHtml
                # Set any possible typeaheads
                data['typeaheads'] = self.lst_typeahead
                response = JsonResponse(data)
            elif self.redirectpage != "":
                return redirect(self.redirectpage)
            else:
                # Set any possible typeaheads
                context['typeaheads'] = json.dumps(self.lst_typeahead)
                # This takes self.template_name...
                sHtml = render_to_string(self.template_name, context, request)
                sHtml = sHtml.replace("\ufeff", "")
                response = HttpResponse(sHtml)
                # response = self.render_to_response(context)

        # Return the response
        return response

    def post(self, request, pk=None, *args, **kwargs):
        # Initialisation
        data = {'status': 'ok', 'html': '', 'statuscode': ''}
        # always do this initialisation to get the object
        self.initializations(request, pk)
        # Make sure only POSTS get through that are authorized
        if request.user.is_authenticated:
            # Check for initredirect
            if self.initRedirect and self.redirectpage != "":
                # Redirect to this page
                return redirect(self.redirectpage)
            # Get the context and perform some standard matters
            context = self.get_context_data(object=self.object)
            # Check if 'afternewurl' needs adding
            if 'afternewurl' in context:
                data['afternewurl'] = context['afternewurl']
            if hasattr(self, "redirect_to"):
                data['afternewurl'] = getattr(self,"redirect_to")
            # Check if 'afterdelurl' needs adding
            if 'afterdelurl' in context:
                data['afterdelurl'] = context['afterdelurl']
            # Possibly indicate form errors
            # NOTE: errors is a dictionary itself...
            if 'errors' in context and len(context['errors']) > 0:
                data['status'] = "error"
                data['msg'] = context['errors']

            if self.is_basic and self.template_name == "":
                if self.rtype == "json":
                    self.template_name = "basic/basic_edit.html"
                else:
                    self.template_name = "basic/basic_details.html"

            if self.rtype == "json":
                if self.template_post == "": self.template_post = self.template_name
                response = render_to_string(self.template_post, context, request)
                response = response.replace("\ufeff", "")
                data['html'] = response
                # Set any possible typeaheads
                data['typeaheads'] = self.lst_typeahead
                response = JsonResponse(data)
            elif self.newRedirect and self.redirectpage != "":
                # Redirect to this page
                return redirect(self.redirectpage)
            else:
                # Set any possible typeaheads
                context['typeaheads'] = json.dumps(self.lst_typeahead)
                # This takes self.template_name...
                response = self.render_to_response(context)
        else:
            data['html'] = "(No authorization)"
            data['status'] = "error"
            response = JsonResponse(data)

        # Return the response
        return response

    def initializations(self, request, pk):
        # Store the previous page
        # self.previous = get_previous_page(request)

        self.lst_typeahead = []

        # Copy any pk
        self.pk = pk
        self.add = pk is None
        # Get the parameters
        if request.POST:
            self.qd = request.POST
        else:
            self.qd = request.GET

        # Check for action
        if 'action' in self.qd:
            self.action = self.qd['action']

        # Find out what the Main Model instance is, if any
        if self.add:
            self.object = None
        else:
            # Get the instance of the Main Model object
            # NOTE: if the object doesn't exist, we will NOT get an error here
            self.object = self.get_object()

        # Possibly perform custom initializations
        self.custom_init(self.object)
        
    def custom_init(self, instance):
        pass

    def before_delete(self, instance):
        """Anything that needs doing before deleting [instance] """
        return True, "" 

    def after_new(self, form, instance):
        """Action to be performed after adding a new item"""
        return True, "" 

    def before_save(self, form, instance):
        """Action to be performed after saving an item preliminarily, and before saving completely"""
        return True, "" 

    def after_save(self, form, instance):
        """Actions to be performed after saving"""
        return True, "" 

    def add_to_context(self, context, instance):
        """Add to the existing context"""
        return context

    def may_edit(self, context = {}):
        """Is this a user who may edit?"""

        bResult = False
        oErr = ErrHandle()
        try:
            # First step: authentication
            if user_is_authenticated(self.request):
                # Second step: app_user
                if context.get("is_app_user", False):
                    bEditor = context.get("is_app_editor", False)
                    bModerator = context.get("is_app_moderator", False)
                    bResult = (bEditor or bModerator)
            # Otherwise: no permissions!
        except:
            oErr.DoError("BasicPart/userpermissions")
        return bResult

    def process_formset(self, prefix, request, formset):
        return None

    def get_formset_queryset(self, prefix):
        return None

    def get_form_kwargs(self, prefix):
        return None

    def get_history(self, instance):
        """Get the history of this element"""
        return ""

    def get_context_data(self, **kwargs):
        # Get the current context
        context = super(BasicDetails, self).get_context_data(**kwargs)

        oErr = ErrHandle()

        # Check this user: is he allowed to UPLOAD data?
        context['authenticated'] = user_is_authenticated(self.request)
        context['is_app_uploader'] = user_is_ingroup(self.request, app_uploader)
        context['is_app_editor'] = user_is_ingroup(self.request, app_editor)
        context['is_app_user'] = user_is_ingroup(self.request, app_user)
        context['is_app_userplus'] = user_is_ingroup(self.request, app_userplus)
        context['is_app_moderator'] = user_is_superuser(self.request) or user_is_ingroup(self.request, app_moderator)
        # context['prevpage'] = get_previous_page(self.request) # self.previous
        context['afternewurl'] = ""

        context['topleftbuttons'] = ""
        context['history_button'] = self.history_button
        context['comment_button'] = self.comment_button
        context['comment_count'] = self.comment_count
        context['no_delete'] = self.no_delete

        if context['authenticated'] and self.permission != "readonly":
            if self.permission != "write":
                self.permission = "read"
            if context['is_app_editor']:
                self.permission = "write"
        context['permission'] = self.permission

        if self.has_select2:
            context['has_select2'] = True

        # Possibly define where a listview is
        classname = self.model._meta.model_name
        if self.basic_name == None or self.basic_name == "":
            if self.basic_name_prefix == "":
                self.basic_name = classname
            else:
                self.basic_name = "{}{}".format(self.basic_name_prefix, self.prefix)
        basic_name = self.basic_name
        if self.listview != None:
            context['listview'] = self.listview
        else:
            listviewname = "{}_list".format(basic_name)
            try:
                context['listview'] = reverse(listviewname)
            except:
                context['listview'] = reverse('home')

        if self.basic_add:
            basic_add = reverse(self.basic_add)
        else:
            basic_add = reverse("{}_details".format(basic_name))
        context['basic_add'] = basic_add

        context['new_button'] = self.new_button
        context['add_text'] = self.add_text
        context['backbutton'] = self.backbutton

        # Selection buttons
        if len(self.selectbuttons) > 0:
            context['selectbuttons'] = self.selectbuttons

        if self.is_basic and context.get('afterdelurl') == None :
            if self.afterdelurl != None:
                context['afterdelurl'] = self.afterdelurl
            else:
                context['afterdelurl'] = context['listview']

        # Custom buttons
        if len(self.custombuttons) > 0:
            for item in self.custombuttons:
                if 'template_name' in item:
                    # get the code of the template
                    pass
            context['custombuttons'] = self.custombuttons

        # Get the parameters passed on with the GET or the POST request
        get = self.request.GET if self.request.method == "GET" else self.request.POST
        initial = get.copy()
        self.qd = initial

        # Possibly first get params
        params = ""
        if "params" in dict(self.qd):
            params = base64_decode( "".join(self.qd.pop("params")))
        context['params'] = params

        # Now see if anything is left
        self.bHasFormInfo = (len(self.qd) > 0)

        # Set the title of the application
        context['title'] = self.title

        # Get the instance
        instance = self.object

        try:
            # Prepare form
            frm = self.prepare_form(instance, context, initial)

            if frm:

                if instance == None:
                    instance = frm.instance
                    self.object = instance

                # Walk all the formset objects
                bFormsetChanged = False
                for formsetObj in self.formset_objects:
                    formsetClass = formsetObj['formsetClass']
                    prefix  = formsetObj['prefix']
                    formset = None
                    form_kwargs = self.get_form_kwargs(prefix)
                    formsetObj['may_evaluate'] = True

                    if 'noinit' in formsetObj and formsetObj['noinit'] and not self.add:
                        # Only process actual changes!!
                        if self.request.method == "POST" and self.request.POST:

                            #if self.add:
                            #    # Saving a NEW item
                            #    if 'initial' in formsetObj:
                            #        formset = formsetClass(self.request.POST, self.request.FILES, prefix=prefix, initial=formsetObj['initial'], form_kwargs = form_kwargs)
                            #    else:
                            #        formset = formsetClass(self.request.POST, self.request.FILES, prefix=prefix, form_kwargs = form_kwargs)
                            #else:
                            #    # Get a formset including any stuff from POST
                            #    formset = formsetClass(self.request.POST, prefix=prefix, instance=instance)

                            # Double check to see if information on this formset is available or not
                            formsetObj['may_evaluate'] = (self.qd.get("{}-TOTAL_FORMS".format(prefix), None) != None)
                            bypass = False
                            if bypass or formsetObj['may_evaluate']:

                                # Get a formset including any stuff from POST
                                formset = formsetClass(self.request.POST, prefix=prefix, instance=instance)
                                # Process this formset
                                self.process_formset(prefix, self.request, formset)
                        
                                # Process all the correct forms in the formset
                                for subform in formset:
                                    if subform.is_valid() and not hasattr(subform, 'do_not_save'):
                                        # DO the actual saving - but that will only work if all is *actually* valid
                                        try:
                                            subform.save()
                                            
                                            # Log the SAVE action
                                            details = {'id': instance.id}
                                            details["savetype"] = "add_sub" # if bNew else "change"
                                            details["form"] = subform.__class__.__name__
                                            details['model'] = subform.instance.__class__.__name__
                                            if subform.changed_data != None and len(subform.changed_data) > 0:
                                                details['changes'] = action_model_changes(subform, subform.instance)
                                            self.action_add(instance, details, "add")

                                            # Signal that the *FORM* needs refreshing, because the formset changed
                                            bFormsetChanged = True
                                        except:
                                            if hasattr(subform, "warning"):
                                                context['errors'] = getattr(subform, "warning")
                                            else:
                                                msg = oErr.get_error_message()
                                                oErr.DoError("BasicDetails/get_context_data")
                                                context['errors'] = {'subform':  msg }

                                if formset.is_valid():
                                    # Load an explicitly empty formset
                                    formset = formsetClass(initial=[], prefix=prefix, form_kwargs=form_kwargs)
                                else:
                                    # Retain the original formset, that now contains the error specifications per form
                                    # But: do *NOT* add an additional form to it
                                    pass

                        else:
                            # All other cases: Load an explicitly empty formset
                            formset = formsetClass(initial=[], prefix=prefix, form_kwargs=form_kwargs)
                    else:
                        # show the data belonging to the current [obj]
                        qs = self.get_formset_queryset(prefix)
                        if qs == None:
                            formset = formsetClass(prefix=prefix, instance=instance, form_kwargs=form_kwargs)
                        else:
                            formset = formsetClass(prefix=prefix, instance=instance, queryset=qs, form_kwargs=form_kwargs)

                    # Only continue if we have a formset
                    if formset != None:
                        # Process all the forms in the formset
                        ordered_forms = self.process_formset(prefix, self.request, formset)
                        if ordered_forms:
                            context[prefix + "_ordered"] = ordered_forms
                        # Store the instance
                        formsetObj['formsetinstance'] = formset
                        # Add the formset to the context
                        context[prefix + "_formset"] = formset
                        # Get any possible typeahead parameters
                        lst_formset_ta = getattr(formset.form, "typeaheads", None)
                        if lst_formset_ta != None:
                            for item in lst_formset_ta:
                                self.lst_typeahead.append(item)

                # Essential formset information
                for formsetObj in self.formset_objects:
                    #if formsetObj['may_evaluate']:
                    prefix = formsetObj['prefix']
                    if 'fields' in formsetObj: context["{}_fields".format(prefix)] = formsetObj['fields']
                    if 'linkfield' in formsetObj: context["{}_linkfield".format(prefix)] = formsetObj['linkfield']

                # Check if the formset made any changes to the form
                if bFormsetChanged:
                    # OLD: 
                    frm = self.prepare_form(instance, context)

                # Put the form and the formset in the context
                context['{}Form'.format(self.prefix)] = frm
                context['basic_form'] = frm
                context['instance'] = instance
                context['options'] = json.dumps({"isnew": (instance == None)})

                # Possibly define the admin detailsview
                if instance:
                    admindetails = "admin:seeker_{}_change".format(classname)
                    try:
                        context['admindetails'] = reverse(admindetails, args=[instance.id])
                    except:
                        pass
                context['modelname'] = self.model._meta.object_name
       
                # Make sure we have a url for editing
                if instance and instance.id:
                    # There is a details and edit url
                    context['editview'] = reverse("{}_edit".format(basic_name), kwargs={'pk': instance.id})
                    context['detailsview'] = reverse("{}_details".format(basic_name), kwargs={'pk': instance.id})
                # Make sure we have an url for new
                context['addview'] = reverse("{}_details".format(basic_name))
            
            # Determine title as shown in the template
            context['titlesg'] = self.titlesg if self.titlesg else self.title if self.title != "" else basic_name.capitalize()
            
            # Determine breadcrumbs and previous page
            if self.is_basic:
                title = self.title if self.title != "" else basic_name
                if self.rtype == "json":
                    # This is the EditView
                    context['breadcrumbs'] = get_breadcrumbs(self.request, "{} edit".format(title), False)
                    prevpage = reverse('home')
                    context['prevpage'] = prevpage
                else:
                    # This is DetailsView
                    # Process this visit and get the new breadcrumbs object
                    prevpage = context['listview']
                    context['prevpage'] = prevpage
                    crumbs = []
                    if self.listviewtitle == None:
                        crumbs.append([title + " list", prevpage])
                    else:
                        crumbs.append([ self.listviewtitle, prevpage])
                    current_name = title if instance else "{} (new)".format(title)
                    context['breadcrumbs'] = get_breadcrumbs(self.request, current_name, True, crumbs)

            # Possibly add to context by the calling function
            if instance.id:
                context = self.add_to_context(context, instance)
                if self.history_button:
                    # Retrieve history
                    context['history_contents'] = self.get_history(instance)

            # Perform CMS operations on the mainitems
            if 'mainitems' in context:
                view_name = "{}_details".format(basic_name)
                context['mainitems'] = cms_translate(view_name, context['mainitems'])

            # fill in the form values
            if frm and 'mainitems' in context:
                for mobj in context['mainitems']:
                    bResult = self.process_mainitem(frm, mobj)

            # Check for mainsections
            if frm and 'mainsections' in context:
                for section in context['mainsections']:
                    for mobj in section['fields']:
                        bResult = self.process_mainitem(frm, mobj)

            # Define where to go to after deletion
            if 'afterdelurl' not in context or context['afterdelurl'] == "":
                context['afterdelurl'] = get_previous_page(self.request)

            # CHeck if reloading is needed
            if self.bNeedReload:
                self.bNeedReload = False
                context['needreload'] = True
        except:
            msg = oErr.get_error_message()
            oErr.DoError("BasicDetails, get_context_data")

        # Return the calculated context
        return context

    def process_mainitem(self, frm, mobj):
        """Process one 'mainitem' or one 'field' from 'mainsections'"""

        bResult = True
        oErr = ErrHandle()
        try:
            # Check for possible form field information
            if 'field_key' in mobj: 
                mobj['field_abbr'] = "{}-{}".format(frm.prefix, mobj['field_key'])
                mobj['field_key'] = frm[mobj['field_key']]
            if 'field_view' in mobj: mobj['field_view'] = frm[mobj['field_view']]
            if 'field_ta' in mobj: mobj['field_ta'] = frm[mobj['field_ta']]
            if 'field_list' in mobj: mobj['field_list'] = frm[mobj['field_list']]

            # Calculate view-mode versus any-mode
            #  'field_key' in mainitem or 'field_list' in mainitem and permission == "write"  or  is_app_userplus and mainitem.maywrite
            if self.permission == "write":       # or app_userplus and 'maywrite' in mobj and mobj['maywrite']:
                mobj['allowing'] = "edit"
            else:
                mobj['allowing'] = "view"
            if ('field_key' in mobj or 'field_list' in mobj) and (mobj['allowing'] == "edit"):
                mobj['allowing_key_list'] = "edit"
            else:
                mobj['allowing_key_list'] = "view"
        except:
            msg = oErr.get_error_message()
            oErr.DoError("process_mainitem")
            bResult = False
        return bResult

    def get_abs_uri(self, sName, obj):
        sBack =  "{}{}".format(self.request.get_host(), reverse(sName, kwargs={'pk': obj.id}))
        return sBack

    def action_add(self, instance, details, actiontype):
        """User can fill this in to his/her liking"""

        # Example: 
        #   Action.add(self.request.user.username, instance.__class__.__name__, "delete", json.dumps(details))
        pass

    def prepare_form(self, instance, context, initial=[]):
        # Initialisations
        bNew = False
        mForm = self.mForm
        oErr = ErrHandle()
        frm = None
        username=self.request.user.username
        team_group=app_editor
        userplus = app_userplus

        try:
            # Determine the prefix
            if self.prefix_type == "":
                id = "n" if instance == None else instance.id
                prefix = "{}-{}".format(self.prefix, id)
            else:
                prefix = self.prefix

            # Check if this is a POST or a GET request
            if self.request.method == "POST" and not self.do_not_save:
                # Determine what the action is (if specified)
                action = ""
                if 'action' in initial: action = initial['action']
                if action == "delete":
                    bContinue = False
                    # The user wants to delete this item
                    try:
                        bResult, msg = self.before_delete(instance)
                        if bResult:
                            # Log the DELETE action
                            details = {'id': instance.id}
                            self.action_add(instance, details, "delete")
                        
                            # Remove this sermongold instance
                            instance.delete()
                        elif msg in ["ok", "okay"]:
                            # Do not delete, and this is okay, by purpose
                            bContinue = True
                        else:
                            # Removing is not possible
                            context['errors'] = {'delete': msg }
                    except:
                        msg = oErr.get_error_message()
                        # Create an errors object
                        context['errors'] = {'delete':  msg }

                    if bContinue:
                        context['afterdelurl'] = reverse("{}_details".format(self.basic_name), kwargs={'pk': instance.id})
                    else:
                        if 'afterdelurl' not in context or context['afterdelurl'] == "":
                            context['afterdelurl'] = get_previous_page(self.request, True)

                    # Make sure we are returning JSON
                    self.rtype = "json"

                    # Possibly add to context by the calling function
                    if instance.id:
                        context = self.add_to_context(context, instance)

                    # No need to retern a form anymore - we have been deleting
                    if not bContinue:
                        return None
            
                # All other actions just mean: edit or new and send back
                # Make instance available
                context['object'] = instance
                self.object = instance

                # Do we actually have an [mForm]??
                if not mForm is None:
                    # Do we have an existing object or are we creating?
                    if instance == None:
                        # Saving a new item
                        if self.use_team_group:
                            frm = mForm(initial, prefix=prefix, username=username, team_group=team_group, userplus=userplus)
                        else:
                            frm = mForm(initial, prefix=prefix)
                        bNew = True
                        self.add = True
                    elif len(initial) == 0:
                        # Create a completely new form, on the basis of the [instance] only
                        if self.use_team_group:
                            frm = mForm(prefix=prefix, instance=instance, username=username, team_group=team_group, userplus=userplus)
                        else:
                            frm = mForm(prefix=prefix, instance=instance)
                    else:
                        # Editing an existing one
                        if self.use_team_group:
                            frm = mForm(initial, self.request.FILES, prefix=prefix, instance=instance, username=username, team_group=team_group, userplus=userplus)
                        else:
                            frm = mForm(initial, self.request.FILES, prefix=prefix, instance=instance)
                    # Both cases: validation and saving
                    if frm.is_valid():
                        # The form is valid - do a preliminary saving
                        obj = frm.save(commit=False)
                        # Any checks go here...
                        bResult, msg = self.before_save(form=frm, instance=obj)
                        if bResult:
                            # Now save it for real
                            obj.save()

                            # Make sure the form is actually saved completely
                            # Issue #426: put it up here
                            frm.save()
                            instance = obj

                            # Log the SAVE action
                            details = {'id': obj.id}
                            details["savetype"] = "new" if bNew else "change"
                            details["form"] = frm.__class__.__name__
                            if frm.changed_data != None and len(frm.changed_data) > 0:
                                details['changes'] = action_model_changes(frm, obj)
                            self.action_add(obj, details, "save")

                            # Issue #426: comment this
                            ## Make sure the form is actually saved completely
                            #frm.save()
                            #instance = obj
                    
                            # Any action(s) after saving
                            bResult, msg = self.after_save(frm, obj)
                        else:

                            # EK: working on this.
                            #     I've now put this exclusively in EqualGoldEdit's method after_save()
                            #
                            ## ADDED Take over any data from [instance] to [frm.data]
                            ##       Provided these fields are in the form's [initial_fields]
                            #if instance != None and hasattr(frm, "initial_fields"):

                            #    # Walk the fields that need to be taken from the instance
                            #    for key in frm.initial_fields:
                            #        value = getattr(instance, key)

                            #        key_prf = '{}-{}'.format(frm.prefix, key)
                            #        if isinstance(value, str) or isinstance(value, int):
                            #            frm.data[key_prf] = value
                            #        elif isinstance(value, object):
                            #            frm.data[key_prf] = str(value.id)
                    
                            if not msg is None:
                                context['errors'] = {'save': msg }
                    elif frm.errors:
                        # We need to pass on to the user that there are errors
                        context['errors'] = frm.errors
                        oErr.Status("BasicDetails/prepare_form form is not valid: {}".format(frm.errors))

                    # Check if this is a new one
                    if bNew:
                        if self.is_basic:
                            self.afternewurl = context['listview']
                            if self.rtype == "html":
                                # Make sure we do a page redirect
                                self.newRedirect = True
                                self.redirectpage = reverse("{}_details".format(self.basic_name), kwargs={'pk': instance.id})
                        # Any code that should be added when creating a new [SermonGold] instance
                        bResult, msg = self.after_new(frm, instance)
                        if not bResult:
                            # Removing is not possible
                            context['errors'] = {'new': msg }
                        # Check if an 'afternewurl' is specified
                        if self.afternewurl != "":
                            context['afternewurl'] = self.afternewurl
                
            else:
                if mForm != None:
                    # Check if this is asking for a new form
                    if instance == None:
                        # Get the form for the sermon
                        if self.use_team_group:
                            frm = mForm(prefix=prefix, username=username, team_group=team_group, userplus=userplus)
                        else:
                            frm = mForm(prefix=prefix)
                    else:
                        # Get the form for the sermon
                        if self.use_team_group:
                            frm = mForm(instance=instance, prefix=prefix, username=username, team_group=team_group, userplus=userplus)
                        else:
                            frm = mForm(instance=instance, prefix=prefix)
                    if frm.is_valid():
                        iOkay = 1
                # Walk all the form objects
                for formObj in self.form_objects:
                    formClass = formObj['form']
                    prefix = formObj['prefix']
                    # This is only for *NEW* forms (right now)
                    form = formClass(prefix=prefix)
                    context[prefix + "Form"] = form
                    if not 'forminstance' in formObj:
                        # Create a new instance
                        formObj['forminstance'] = formObj['form'](self.request.POST, prefix=formObj['prefix'])
                    # Get any possible typeahead parameters
                    lst_form_ta = getattr(formObj['forminstance'], "typeaheads", None)
                    if lst_form_ta != None:
                        for item in lst_form_ta:
                            self.lst_typeahead.append(item)

            # Get any possible typeahead parameters
            if frm != None:
                lst_form_ta = getattr(frm, "typeaheads", None)
                if lst_form_ta != None:
                    for item in lst_form_ta:
                        self.lst_typeahead.append(item)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("BasicDetails/prepare_form")

        # Return the form we made
        return frm
    

class BasicPart(View):
    """This is my own versatile handling view.

    Note: this version works with <pk> and not with <object_id>
    """

    # Initialisations
    arErr = []              # errors   
    template_name = None    # The template to be used
    template_err_view = None
    permission = True
    form_validated = True   # Used for POST form validation
    savedate = None         # When saving information, the savedate is returned in the context
    add = False             # Are we adding a new record or editing an existing one?
    obj = None              # The instance of the MainModel
    action = ""             # The action to be undertaken
    method = None           # GET or POST
    rtype = "json"          # JSON response (alternative: html)    
    MainModel = None        # The model that is mainly used for this form
    form_objects = []       # List of forms to be processed
    formset_objects = []    # List of formsets to be processed
    previous = None         # Return to this
    downloadname = None     # Name used for downloading
    bDebug = False          # Debugging information
    redirectpage = ""       # Where to redirect to
    data = {'status': 'ok', 'html': ''}       # Create data to be returned    
    
    def post(self, request, pk=None):
        self.method = "POST"
        # A POST request means we are trying to SAVE something
        self.initializations(request, pk)
        # Initialize typeahead list
        lst_typeahead = []

        # Explicitly set the status to OK
        self.data['status'] = "ok"
        
        if self.checkAuthentication(request):
            # Build the context
            context = dict(object_id = pk, savedate=None)
            context['authenticated'] = user_is_authenticated(request)
            context['is_app_uploader'] = user_is_ingroup(request, app_uploader)
            context['is_app_editor'] = user_is_ingroup(request, app_editor)

            write_permission = self.userpermissions("w")

            # Action depends on 'action' value
            if self.action == "":
                if self.bDebug: self.oErr.Status("ResearchPart: action=(empty)")
                # Walk all the forms for preparation of the formObj contents
                for formObj in self.form_objects:
                    # Are we SAVING a NEW item?
                    if self.add:
                        # We are saving a NEW item
                        formObj['forminstance'] = formObj['form'](request.POST, prefix=formObj['prefix'])
                        formObj['action'] = "new"
                    else:
                        # We are saving an EXISTING item
                        # Determine the instance to be passed on
                        instance = self.get_instance(formObj['prefix'])
                        # Make the instance available in the form-object
                        formObj['instance'] = instance
                        # Get an instance of the form
                        formObj['forminstance'] = formObj['form'](request.POST, prefix=formObj['prefix'], instance=instance)
                        formObj['action'] = "change"

                # Initially we are assuming this just is a review
                context['savedate']="reviewed at {}".format(get_current_datetime().strftime("%X"))

                # Iterate again
                for formObj in self.form_objects:
                    prefix = formObj['prefix']
                    # Adapt if it is not readonly
                    if not formObj['readonly']:
                        # Check validity of form
                        if formObj['forminstance'].is_valid() and self.is_custom_valid(prefix, formObj['forminstance']):
                            # Save it preliminarily
                            instance = formObj['forminstance'].save(commit=False)
                            # The instance must be made available (even though it is only 'preliminary')
                            formObj['instance'] = instance
                            # Perform actions to this form BEFORE FINAL saving
                            bNeedSaving = formObj['forminstance'].has_changed()
                            if self.before_save(prefix, request, instance=instance, form=formObj['forminstance']): bNeedSaving = True
                            if formObj['forminstance'].instance.id == None: bNeedSaving = True
                            if bNeedSaving:
                                # Perform the saving
                                instance.save()
                                # Log the SAVE action
                                details = {'id': instance.id}
                                if formObj['forminstance'].changed_data != None:
                                    details['changes'] = action_model_changes(formObj['forminstance'], instance)
                                if 'action' in formObj: details['savetype'] = formObj['action']
                                # Action.add(request.user.username, self.MainModel.__name__, instance.id, "save", json.dumps(details))
                                self.action_add(instance, details, "save")
                                # Set the context
                                context['savedate']="saved at {}".format(get_current_datetime().strftime("%X"))
                                # Put the instance in the form object
                                formObj['instance'] = instance
                                # Store the instance id in the data
                                self.data[prefix + '_instanceid'] = instance.id
                                # Any action after saving this form
                                self.after_save(prefix, instance=instance, form=formObj['forminstance'])
                            # Also get the cleaned data from the form
                            formObj['cleaned_data'] = formObj['forminstance'].cleaned_data
                        else:
                            self.arErr.append(formObj['forminstance'].errors)
                            self.form_validated = False
                            formObj['cleaned_data'] = None
                    else:
                        # Form is readonly

                        # Check validity of form
                        if formObj['forminstance'].is_valid() and self.is_custom_valid(prefix, formObj['forminstance']):
                            # At least get the cleaned data from the form
                            formObj['cleaned_data'] = formObj['forminstance'].cleaned_data

                            # x = json.dumps(sorted(self.qd.items(), key=lambda kv: kv[0]), indent=2)
                    # Add instance to the context object
                    context[prefix + "Form"] = formObj['forminstance']
                    # Get any possible typeahead parameters
                    lst_form_ta = getattr(formObj['forminstance'], "typeaheads", None)
                    if lst_form_ta != None:
                        for item in lst_form_ta:
                            lst_typeahead.append(item)
                # Walk all the formset objects
                for formsetObj in self.formset_objects:
                    prefix  = formsetObj['prefix']
                    if self.can_process_formset(prefix):
                        formsetClass = formsetObj['formsetClass']
                        form_kwargs = self.get_form_kwargs(prefix)
                        if self.add:
                            # Saving a NEW item
                            if 'initial' in formsetObj:
                                formset = formsetClass(request.POST, request.FILES, prefix=prefix, initial=formsetObj['initial'], form_kwargs = form_kwargs)
                            else:
                                formset = formsetClass(request.POST, request.FILES, prefix=prefix, form_kwargs = form_kwargs)
                        else:
                            # Saving an EXISTING item
                            instance = self.get_instance(prefix)
                            qs = self.get_queryset(prefix)
                            if qs == None:
                                formset = formsetClass(request.POST, request.FILES, prefix=prefix, instance=instance, form_kwargs = form_kwargs)
                            else:
                                formset = formsetClass(request.POST, request.FILES, prefix=prefix, instance=instance, queryset=qs, form_kwargs = form_kwargs)
                        # Process all the forms in the formset
                        self.process_formset(prefix, request, formset)
                        # Store the instance
                        formsetObj['formsetinstance'] = formset
                        # Make sure we know what we are dealing with
                        itemtype = "form_{}".format(prefix)
                        # Adapt the formset contents only, when it is NOT READONLY
                        if not formsetObj['readonly']:
                            # Is the formset valid?
                            if formset.is_valid():
                                # Possibly handle the clean() for this formset
                                if 'clean' in formsetObj:
                                    # Call the clean function
                                    self.clean(formset, prefix)
                                has_deletions = False
                                if len(self.arErr) == 0:
                                    # Make sure all changes are saved in one database-go
                                    with transaction.atomic():
                                        # Walk all the forms in the formset
                                        for form in formset:
                                            # At least check for validity
                                            if form.is_valid() and self.is_custom_valid(prefix, form):
                                                # Should we delete?
                                                if 'DELETE' in form.cleaned_data and form.cleaned_data['DELETE']:
                                                    # Check if deletion should be done
                                                    if self.before_delete(prefix, form.instance):
                                                        # Log the delete action
                                                        details = {'id': form.instance.id}
                                                        # Action.add(request.user.username, itemtype, form.instance.id, "delete", json.dumps(details))
                                                        self.action_add(form.instance, details, "delete")
                                                        # Delete this one
                                                        form.instance.delete()
                                                        # NOTE: the template knows this one is deleted by looking at form.DELETE
                                                        has_deletions = True
                                                else:
                                                    # Check if anything has changed so far
                                                    has_changed = form.has_changed()
                                                    # Save it preliminarily
                                                    sub_instance = form.save(commit=False)
                                                    # Any actions before saving
                                                    if self.before_save(prefix, request, sub_instance, form):
                                                        has_changed = True
                                                    # Save this construction
                                                    if has_changed and len(self.arErr) == 0: 
                                                        # Save the instance
                                                        sub_instance.save()
                                                        # Adapt the last save time
                                                        context['savedate']="saved at {}".format(get_current_datetime().strftime("%X"))
                                                        # Log the delete action
                                                        details = {'id': sub_instance.id}
                                                        if form.changed_data != None:
                                                            details['changes'] = action_model_changes(form, sub_instance)
                                                        # Action.add(request.user.username, itemtype,sub_instance.id, "save", json.dumps(details))
                                                        self.action_add(sub_instance, details, "save")
                                                        # Store the instance id in the data
                                                        self.data[prefix + '_instanceid'] = sub_instance.id
                                                        # Any action after saving this form
                                                        self.after_save(prefix, sub_instance)
                                            else:
                                                if len(form.errors) > 0:
                                                    self.arErr.append(form.errors)
                                
                                    # Rebuild the formset if it contains deleted forms
                                    if has_deletions or not has_deletions:
                                        # Or: ALWAYS
                                        if qs == None:
                                            formset = formsetClass(prefix=prefix, instance=instance, form_kwargs=form_kwargs)
                                        else:
                                            formset = formsetClass(prefix=prefix, instance=instance, queryset=qs, form_kwargs=form_kwargs)
                                        formsetObj['formsetinstance'] = formset
                            else:
                                # Iterate over all errors
                                for idx, err_this in enumerate(formset.errors):
                                    if '__all__' in err_this:
                                        self.arErr.append(err_this['__all__'][0])
                                    elif err_this != {}:
                                        # There is an error in item # [idx+1], field 
                                        problem = err_this 
                                        for k,v in err_this.items():
                                            fieldName = k
                                            errmsg = "Item #{} has an error at field [{}]: {}".format(idx+1, k, v[0])
                                            self.arErr.append(errmsg)

                            # self.arErr.append(formset.errors)
                    else:
                        formset = []
                    # Get any possible typeahead parameters
                    lst_formset_ta = getattr(formset.form, "typeaheads", None)
                    if lst_formset_ta != None:
                        for item in lst_formset_ta:
                            lst_typeahead.append(item)
                    # Add the formset to the context
                    context[prefix + "_formset"] = formset
            elif self.action == "download" and write_permission:
                # We are being asked to download something
                if self.dtype != "":
                    plain_type = ["xlsx", "csv", "excel"]
                    # Initialise return status
                    oBack = {'status': 'ok'}
                    sType = "csv" if (self.dtype == "xlsx") else self.dtype

                    # Get the data
                    sData = ""
                    if not self.dtype in ["excel"]:
                        sData = self.get_data('', self.dtype)
                    # Decode the data and compress it using gzip
                    bUtf8 = (self.dtype != "db")
                    bUsePlain = (self.dtype in plain_type)

                    # Create name for download
                    # sDbName = "{}_{}_{}_QC{}_Dbase.{}{}".format(sCrpName, sLng, sPartDir, self.qcTarget, self.dtype, sGz)
                    if self.downloadname is None:
                        downloadname = self.MainModel.__name__
                    else:
                        downloadname = self.downloadname
                    obj_id = "n" if self.obj == None else self.obj.id
                    extension = self.dtype
                    if self.dtype == "excel":
                        extension = "xlsx"
                    elif self.dtype == "tei" or self.dtype == "xml-tei":
                        extension = "xml"
                    sDbName = "passim_{}_{}.{}".format(downloadname, obj_id, extension)
                    sContentType = ""
                    if self.dtype == "csv":
                        sContentType = "text/tab-separated-values"
                    elif self.dtype == "json":
                        sContentType = "application/json"
                    elif self.dtype == "xlsx" or self.dtype == "excel":
                        sContentType = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    elif self.dtype == "hist-svg":
                        sContentType = "application/svg"
                        sData = self.qd['downloaddata']
                        # Set the filename correctly
                        sDbName = "passim_{}_{}.svg".format(downloadname, obj_id)
                    elif self.dtype == "hist-png":
                        sContentType = "image/png"
                        # Read the base64 encoded part
                        sData = self.qd['downloaddata']
                        arPart = sData.split(";")
                        if len(arPart) > 1:
                            dSecond = arPart[1]
                            # Strip off preceding base64 part
                            sData = dSecond.replace("base64,", "")
                            # Convert string to bytestring
                            sData = sData.encode()
                            # Decode base64 into binary
                            sData = base64.decodestring(sData)
                            # Set the filename correctly
                            sDbName = "passim_{}_{}.png".format(downloadname, obj_id)

                    # Excel needs additional conversion
                    if self.dtype in ["excel"]:
                        # Convert 'compressed_content' to an Excel worksheet
                        response = HttpResponse(content_type=sContentType)
                        response['Content-Disposition'] = 'attachment; filename="{}"'.format(sDbName)    
                        response = self.get_data('', self.dtype, response)
                    elif self.dtype == "xlsx":
                        # Convert 'compressed_content' to an Excel worksheet
                        response = HttpResponse(content_type=sContentType)
                        response['Content-Disposition'] = 'attachment; filename="{}"'.format(sDbName)    
                        response = csv_to_excel(sData, response)
                    else:
                        response = HttpResponse(sData, content_type=sContentType)
                        response['Content-Disposition'] = 'attachment; filename="{}"'.format(sDbName)    

                    # Continue for all formats
                        
                    # return gzip_middleware.process_response(request, response)
                    return response
            elif self.action == "delete" and write_permission:
                # The user requests this to be deleted
                if self.before_delete():
                    # Log the delete action
                    details = {'id': self.obj.id}
                    # Action.add(request.user.username, self.MainModel.__name__, self.obj.id, "delete", json.dumps(details))
                    self.action_add(self.obj, details, "delete")
                    # We have permission to delete the instance
                    self.obj.delete()
                    context['deleted'] = True

            # Allow user to add to the context
            context = self.add_to_context(context)
            
            # Possibly add data from [context.data]
            if 'data' in context:
                for k,v in context['data'].items():
                    self.data[k] = v

            # First look at redirect page
            self.data['redirecturl'] = ""
            if self.redirectpage != "":
                self.data['redirecturl'] = self.redirectpage
            # Check if 'afternewurl' needs adding
            # NOTE: this should only be used after a *NEW* instance has been made -hence the self.add check
            if 'afternewurl' in context and self.add:
                self.data['afternewurl'] = context['afternewurl']
            else:
                self.data['afternewurl'] = ""
            if 'afterdelurl' in context:
                self.data['afterdelurl'] = context['afterdelurl']

            # Make sure we have a list of any errors
            error_list = [str(item) for item in self.arErr]
            context['error_list'] = error_list
            context['errors'] = json.dumps( self.arErr)
            if len(self.arErr) > 0:
                # Indicate that we have errors
                self.data['has_errors'] = True
                self.data['status'] = "error"
            else:
                self.data['has_errors'] = False
            # Standard: add request user to context
            context['requestuser'] = request.user

            # Set any possible typeaheads
            self.data['typeaheads'] = lst_typeahead

            # Get the HTML response
            if len(self.arErr) > 0:
                if self.template_err_view != None:
                     # Create a list of errors
                    self.data['err_view'] = render_to_string(self.template_err_view, context, request)
                else:
                    self.data['error_list'] = error_list
                    self.data['errors'] = self.arErr
                self.data['html'] = ''
                # We may not redirect if there is an error!
                self.data['redirecturl'] = ''
            elif self.action == "delete":
                self.data['html'] = "deleted" 
            elif self.template_name != None:
                # In this case reset the errors - they should be shown within the template
                sHtml = render_to_string(self.template_name, context, request)
                sHtml = treat_bom(sHtml)
                self.data['html'] = sHtml
            else:
                if not 'html' in self.data:
                    self.data['html'] = 'no template_name specified'
                else:
                    # No need to do anything, because the data already contains html
                    pass

            # At any rate: empty the error basket
            self.arErr = []
            error_list = []

        else:
            self.data['html'] = "Please log in before continuing"

        # Return the information
        return JsonResponse(self.data)
        
    def get(self, request, pk=None): 
        self.data['status'] = 'ok'
        self.method = "GET"
        # Perform the initializations that need to be made anyway
        self.initializations(request, pk)
        # Initialize typeahead list
        lst_typeahead = []

        oErr = ErrHandle()
        try:
            response = JsonResponse(self.data)
            # Continue if authorized
            if self.checkAuthentication(request):
                context = dict(object_id = pk, savedate=None)
                context['prevpage'] = self.previous
                context['authenticated'] = user_is_authenticated(request)
                context['permission'] = self.permission
                context['is_app_uploader'] = user_is_ingroup(request, app_uploader)
                context['is_app_editor'] = user_is_ingroup(request, app_editor)
                # Walk all the form objects
                for formObj in self.form_objects:        
                    # Used to populate a NEW research project
                    # - CREATE a NEW research form, populating it with any initial data in the request
                    initial = dict(request.GET.items())
                    if self.add:
                        # Create a new form
                        formObj['forminstance'] = formObj['form'](initial=initial, prefix=formObj['prefix'])
                    else:
                        # Used to show EXISTING information
                        instance = self.get_instance(formObj['prefix'])
                        # We should show the data belonging to the current Research [obj]
                        formObj['forminstance'] = formObj['form'](instance=instance, prefix=formObj['prefix'])
                    # Add instance to the context object
                    context[formObj['prefix'] + "Form"] = formObj['forminstance']
                    # Get any possible typeahead parameters
                    lst_form_ta = getattr(formObj['forminstance'], "typeaheads", None)
                    if lst_form_ta != None:
                        for item in lst_form_ta:
                            lst_typeahead.append(item)
                # Walk all the formset objects
                for formsetObj in self.formset_objects:
                    formsetClass = formsetObj['formsetClass']
                    prefix  = formsetObj['prefix']
                    form_kwargs = self.get_form_kwargs(prefix)
                    if self.add:
                        # - CREATE a NEW formset, populating it with any initial data in the request
                        initial = dict(request.GET.items())
                        # Saving a NEW item
                        formset = formsetClass(initial=initial, prefix=prefix, form_kwargs=form_kwargs)
                    else:
                        # Possibly initial (default) values
                        if 'initial' in formsetObj:
                            initial = formsetObj['initial']
                        else:
                            initial = None
                        # show the data belonging to the current [obj]
                        instance = self.get_instance(prefix)
                        qs = self.get_queryset(prefix)
                        if qs == None:
                            formset = formsetClass(prefix=prefix, instance=instance, form_kwargs=form_kwargs)
                        else:
                            formset = formsetClass(prefix=prefix, instance=instance, queryset=qs, initial=initial, form_kwargs=form_kwargs)
                    # Get any possible typeahead parameters
                    lst_formset_ta = getattr(formset.form, "typeaheads", None)
                    if lst_formset_ta != None:
                        for item in lst_formset_ta:
                            lst_typeahead.append(item)
                    # Process all the forms in the formset
                    ordered_forms = self.process_formset(prefix, request, formset)
                    if ordered_forms:
                        context[prefix + "_ordered"] = ordered_forms
                    # Store the instance
                    formsetObj['formsetinstance'] = formset
                    # Add the formset to the context
                    context[prefix + "_formset"] = formset
                # Allow user to add to the context
                context = self.add_to_context(context)
                # Make sure we have a list of any errors
                error_list = [str(item) for item in self.arErr]
                context['error_list'] = error_list
                context['errors'] = self.arErr
                # Standard: add request user to context
                context['requestuser'] = request.user

                # Set any possible typeaheads
                self.data['typeaheads'] = json.dumps(lst_typeahead)
            
                # Get the HTML response
                if self.template_name is None:
                    self.data['html'] = ""
                else:
                    sHtml = render_to_string(self.template_name, context, request)
                    sHtml = treat_bom(sHtml)
                    self.data['html'] = sHtml
            else:
                self.data['html'] = "Please log in before continuing"

            # Determine the response type
            if self.rtype == "json":
                response = JsonResponse(self.data)
            else:
                # This takes self.template_name...
                sHtml = self.data['html']
                response = HttpResponse(sHtml)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("BasicPart/get")

        # Return the information
        return response

    def userpermissions(self, sType = "w"):
        """Basic check for valid user permissions"""

        bResult = False
        oErr = ErrHandle()
        try:
            # First step: authentication
            if user_is_authenticated(self.request):
                # Second step: app_user
                is_moderator = user_is_ingroup(self.request, app_moderator)
                is_app_user = is_moderator or user_is_ingroup(self.request, app_user)
                is_app_editor = is_moderator or user_is_ingroup(self.request, app_editor)
                if is_app_user or is_app_editor:
                    # Any more checking needed?
                    if sType == "w":
                        bResult = is_app_editor
                    else:
                        bResult = True
            # Otherwise: no permissions!
        except:
            oErr.DoError("BasicPart/userpermissions")
        return bResult

    def action_add(self, instance, details, actiontype):
        """User can fill this in to his/her liking"""

        # Example: 
        #   Action.add(self.request.user.username, instance.__class__.__name__, "delete", json.dumps(details))
        pass
      
    def checkAuthentication(self,request):
        # first check for authentication
        if not request.user.is_authenticated:
            # Simply redirect to the home page
            self.data['html'] = "Please log in to work on this project"
            return False
        else:
            return True

    def rebuild_formset(self, prefix, formset):
        return formset

    def initializations(self, request, object_id):
        # Store the previous page
        #self.previous = get_previous_page(request)
        # Clear errors
        self.arErr = []
        # COpy the request
        self.request = request
        # Copy any object id
        self.object_id = object_id
        self.add = object_id is None
        # Get the parameters
        if request.POST:
            self.qd = request.POST.copy()
        else:
            self.qd = request.GET.copy()

        # Immediately take care of the rangeslider stuff
        lst_remove = []
        for k,v in self.qd.items():
            if "-rangeslider" in k: lst_remove.append(k)
        for item in lst_remove: self.qd.pop(item)
        #lst_remove = []
        #dictionary = {}
        #for k,v in self.qd.items():
        #    if "-rangeslider" not in k: 
        #        dictionary[k] = v
        #self.qd = dictionary

        # Check for action
        if 'action' in self.qd:
            self.action = self.qd['action']

        # Find out what the Main Model instance is, if any
        if self.add:
            self.obj = None
        elif self.MainModel != None:
            # Get the instance of the Main Model object
            self.obj =  self.MainModel.objects.filter(pk=object_id).first()
            # NOTE: if the object doesn't exist, we will NOT get an error here
        # ALWAYS: perform some custom initialisations
        self.custom_init()

    def get_instance(self, prefix):
        return self.obj

    def is_custom_valid(self, prefix, form):
        return True

    def get_queryset(self, prefix):
        return None

    def get_form_kwargs(self, prefix):
        return None

    def get_data(self, prefix, dtype, response=None):
        return ""

    def before_save(self, prefix, request, instance=None, form=None):
        return False

    def before_delete(self, prefix=None, instance=None):
        return True

    def after_save(self, prefix, instance=None, form=None):
        return True

    def add_to_context(self, context):
        return context

    def process_formset(self, prefix, request, formset):
        return None

    def can_process_formset(self, prefix):
        return True

    def custom_init(self):
        pass    
           

