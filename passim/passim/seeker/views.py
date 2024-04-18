"""
Definition of views for the SEEKER app.
"""

from django.apps import apps
from django.contrib import admin
from django.contrib.auth import login, authenticate
from django.contrib.auth.models import Group
from django.urls import reverse
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.core.validators import validate_email
from django.db import transaction
from django.db.models import Q, Prefetch, Count, F, Sum
from django.db.models.functions import Lower
from django.db.models.query import QuerySet 
from django.forms import formset_factory, modelformset_factory, inlineformset_factory, ValidationError
from django.forms.models import model_to_dict
from django.http import HttpRequest, HttpResponse, HttpResponseRedirect, JsonResponse, FileResponse
from django.middleware.csrf import get_token
from django.shortcuts import get_object_or_404, render, redirect
from django.template.loader import render_to_string
from django.template import Context
from django.views.generic.detail import DetailView
from django.views.generic.base import RedirectView
from django.views.generic import ListView, View
from django.views.decorators.csrf import csrf_exempt
from lxml import etree as ET

# General imports
from datetime import datetime
import operator 
from operator import itemgetter
from functools import reduce
from pyzotero import zotero
from time import sleep 
import fnmatch
import sys, os
import base64
import copy
import json
import csv, re
import requests
import openpyxl
from threading import Thread
from openpyxl.utils.cell import get_column_letter
import sqlite3
from io import StringIO
from itertools import chain
from unidecode import unidecode



# ======== imports for PDF creation ==========
import io  
from markdown import markdown 
from reportlab.lib.units import inch 
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle 
from reportlab.lib.units import inch 
from reportlab.pdfgen import canvas 
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, ListFlowable, ListItem, Frame, PageBreak   
from reportlab.rl_config import defaultPageSize  

# ======= imports from my own application ======
from passim.settings import APP_PREFIX, MEDIA_DIR, WRITABLE_DIR
from passim.utils import ErrHandle
from passim.seeker.forms import SearchCollectionForm, SearchManuForm, LibrarySearchForm, SignUpForm, \
    AuthorSearchForm, UploadFileForm, UploadFilesForm, ManuscriptForm, SermonForm, SermonGoldForm, CommentForm, \
    SelectGoldForm, SermonGoldSameForm, SermonGoldSignatureForm, AuthorEditForm, BibRangeForm, FeastForm, \
    SermonGoldEditionForm, SermonGoldFtextlinkForm, SermonDescrGoldForm, SermonDescrSuperForm, SearchUrlForm, \
    SermonDescrSignatureForm, SermonGoldKeywordForm, SermonGoldLitrefForm, EqualGoldLinkForm, EqualGoldForm, \
    ReportEditForm, SourceEditForm, ManuscriptProvForm, LocationForm, LocationRelForm, OriginForm, \
    LibraryForm, ManuscriptExtForm, ManuscriptLitrefForm, ManuscriptLinkForm, SermonDescrKeywordForm, KeywordForm, \
    ManuscriptKeywordForm, DaterangeForm, ProjectForm, SermonDescrCollectionForm, CollectionForm, \
    SuperSermonGoldForm, SermonGoldCollectionForm, ManuscriptCollectionForm, CollectionLitrefForm, \
    SuperSermonGoldCollectionForm, ProfileForm, UserKeywordForm, ProvenanceForm, ProvenanceManForm, \
    TemplateForm, TemplateImportForm, ManuReconForm,  ManuscriptProjectForm, \
    CodicoForm, CodicoProvForm, ProvenanceCodForm, OriginCodForm, CodicoOriginForm, OnlineSourceForm, \
    UserForm, SermonDescrLinkForm, CommentResponseForm, DaterangeHistCollForm
from passim.seeker.models import get_crpp_date, get_current_datetime, process_lib_entries, get_searchable, get_now_time, \
    add_gold2equal, add_equal2equal, add_ssg_equal2equal, get_helptext, get_spec_col_num, \
    FieldChoice, Information, Country, City, Author, Manuscript, \
    User, Group, Origin, SermonDescr, MsItem, SermonHead, SermonGold, SermonDescrKeyword, SermonDescrEqual, Nickname, NewsItem, \
    SourceInfo, SermonGoldSame, SermonGoldKeyword, EqualGoldKeyword, Signature, Ftextlink, ManuscriptExt, \
    ManuscriptKeyword, Action, EqualGold, EqualGoldLink, Location, LocationName, LocationIdentifier, LocationRelation, LocationType, \
    ProvenanceMan, Provenance, Daterange, CollOverlap, BibRange, Feast, Comment, CommentRead, CommentResponse, SermonEqualDist, \
    Basket, BasketMan, BasketGold, BasketSuper, Litref, LitrefMan, LitrefCol, LitrefSG, EdirefSG, Report, SermonDescrGold, \
    Visit, Profile, Keyword, SermonSignature, Status, Library, Collection, CollectionSerm, \
    CollectionMan, CollectionSuper, CollectionGold, UserKeyword, Template, ManuscriptLink, \
    EqualGoldExternal, SermonGoldExternal, SermonDescrExternal, ManuscriptExternal, \
    ManuscriptCorpus, ManuscriptCorpusLock, EqualGoldCorpus, ProjectApprover, ProjectEditor, \
    Codico, ProvenanceCod, OriginCod, CodicoKeyword, Reconstruction, Free, SermonDescrLink, \
    Project2, ManuscriptProject, CollectionProject, EqualGoldProject, SermonDescrProject, OnlineSources, DaterangeHistColl, AltPages, \
    choice_value, get_reverse_spec, LINK_EQUAL, LINK_PRT, LINK_REL, LINK_BIDIR, LINK_BIDIR_MANU, \
    LINK_PARTIAL, STYPE_IMPORTED, STYPE_EDITED, STYPE_MANUAL, LINK_UNSPECIFIED
from passim.reader.views import reader_uploads, get_huwa_opera_literature, read_transcription, scan_transcriptions, sync_transcriptions
from passim.bible.models import Reference
from passim.dct.models import ResearchSet, SetList, SavedItem, SavedSearch, SelectItem
from passim.approve.views import approval_parse_changes, approval_parse_formset, approval_pending, approval_pending_list, \
    approval_parse_adding, approval_parse_removing, approval_parse_deleting, addapproval_pending
from passim.seeker.adaptations import listview_adaptations, adapt_codicocopy, add_codico_to_manuscript

# ======= from RU-Basic ========================
from passim.basic.views import BasicPart, BasicList, BasicDetails, make_search_list, add_rel_item, adapt_search, is_ajax

# ======= from RU-cms ==========================
from passim.seeker.views_utils import passim_action_add, passim_get_history


# Some constants that can be used
paginateSize = 20
paginateSelect = 15
paginateValues = (100, 50, 20, 10, 5, 2, 1, )

# Global debugging 
bDebug = False

cnrs_url = "http://medium-avance.irht.cnrs.fr"

def get_application_name():
    """Try to get the name of this application"""

    # Walk through all the installed apps
    for app in apps.get_app_configs():
        # Check if this is a site-package
        if "site-package" not in app.path:
            # Get the name of this app
            name = app.name
            # Take the first part before the dot
            project_name = name.split(".")[0]
            return project_name
    return "unknown"
# Provide application-specific information
PROJECT_NAME = get_application_name()
app_uploader = "{}_uploader".format(PROJECT_NAME.lower())
app_editor = "{}_editor".format(PROJECT_NAME.lower())
app_userplus = "{}_userplus".format(PROJECT_NAME.lower())
app_developer = "{}_developer".format(PROJECT_NAME.lower())
app_moderator = "{}_moderator".format(PROJECT_NAME.lower())
enrich_editor = "enrich_editor"
stemma_editor = "stemma_editor"
stemma_user = "stemma_user"

def get_usercomments(type, instance, profile):
    """Get a HTML list of comments made by this user and possible users in the same group"""

    html = []
    lstQ = []
    if not username_is_ingroup(profile.user, app_editor):
        # App editors have permission to see *all* comments from all users
        lstQ.append(Q(profile=profile))
    if type != "":
        lstQ.append(Q(otype=type))

    # Calculate the list
    qs = instance.comments.filter(*lstQ).order_by("-created")

    # REturn the list
    return qs

def get_application_context(request, context):
    context['is_app_uploader'] = user_is_ingroup(request, app_uploader)
    context['is_app_editor'] = user_is_ingroup(request, app_editor)
    context['is_enrich_editor'] = user_is_ingroup(request, enrich_editor)
    context['is_app_moderator'] = user_is_superuser(request) or user_is_ingroup(request, app_moderator)
    context['is_stemma_editor'] = user_is_ingroup(request, stemma_editor) or context['is_app_moderator']
    context['is_stemma_user'] = user_is_ingroup(request, stemma_user)
    return context

def treat_bom(sHtml):
    """REmove the BOM marker except at the beginning of the string"""

    # Check if it is in the beginning
    bStartsWithBom = sHtml.startswith(u'\ufeff')
    # Remove everywhere
    sHtml = sHtml.replace(u'\ufeff', '')
    # Return what we have
    return sHtml

def adapt_m2m(cls, instance, field1, qs, field2, extra = [], extrargs = {}, qfilter = {}, 
              related_is_through = False, userplus = None, added=None, deleted=None):
    """Adapt the 'field' of 'instance' to contain only the items in 'qs'
    
    The lists [added] and [deleted] (if specified) will contain links to the elements that have been added and deleted
    If [deleted] is specified, then the items will not be deleted by adapt_m2m(). Caller needs to do this.
    """

    errHandle = ErrHandle()
    try:
        # Get current associations
        lstQ = [Q(**{field1: instance})]
        for k,v in qfilter.items(): lstQ.append(Q(**{k: v}))
        through_qs = cls.objects.filter(*lstQ)
        if related_is_through:
            related_qs = through_qs
        else:
            related_qs = [getattr(x, field2) for x in through_qs]
        # make sure all items in [qs] are associated
        if userplus == None or userplus:
            for obj in qs:
                if obj not in related_qs:
                    # Add the association
                    args = {field1: instance}
                    if related_is_through:
                        args[field2] = getattr(obj, field2)
                    else:
                        args[field2] = obj
                    for item in extra:
                        # Copy the field with this name from [obj] to 
                        args[item] = getattr(obj, item)
                    for k,v in extrargs.items():
                        args[k] = v
                    # cls.objects.create(**{field1: instance, field2: obj})
                    new = cls.objects.create(**args)
                    if added != None:
                        added.append(new)

        # Remove from [cls] all associations that are not in [qs]
        # NOTE: do not allow userplus to delete
        for item in through_qs:
            if related_is_through:
                obj = item
            else:
                obj = getattr(item, field2)
            if obj not in qs:
                if deleted == None:
                    # Remove this item
                    item.delete()
                else:
                    deleted.append(item)
        # Return okay
        return True
    except:
        msg = errHandle.get_error_message()
        return False

def adapt_m2o(cls, instance, field, qs, link_to_obj = None, **kwargs):
    """Adapt the instances of [cls] pointing to [instance] with [field] to only include [qs] """

    errHandle = ErrHandle()
    try:
        # Get all the [cls] items currently linking to [instance]
        lstQ = [Q(**{field: instance})]
        linked_qs = cls.objects.filter(*lstQ)
        if link_to_obj != None:
            linked_through = [getattr(x, link_to_obj) for x in linked_qs]
        # make sure all items in [qs] are linked to [instance]
        for obj in qs:
            if (obj not in linked_qs) and (link_to_obj == None or obj not in linked_through):
                # Create new object
                oNew = cls()
                setattr(oNew, field, instance)
                # Copy the local fields
                for lfield in obj._meta.local_fields:
                    fname = lfield.name
                    if fname != "id" and fname != field:
                        # Copy the field value
                        setattr(oNew, fname, getattr(obj, fname))
                for k, v in kwargs.items():
                    setattr(oNew, k, v)
                # Need to add an object link?
                if link_to_obj != None:
                    setattr(oNew, link_to_obj, obj)
                oNew.save()
        # Remove links that are not in [qs]
        for obj in linked_qs:
            if obj not in qs:
                # Remove this item
                obj.delete()
        # Return okay
        return True
    except:
        msg = errHandle.get_error_message()
        return False

def adapt_m2o_sig(instance, qs):
    """Adapt the instances of [SermonSignature] pointing to [instance] to only include [qs] 
    
    Note: convert SermonSignature into (Gold) Signature
    """

    errHandle = ErrHandle()
    try:
        # Get all the [SermonSignature] items currently linking to [instance]
        linked_qs = SermonSignature.objects.filter(sermon=instance)
        # make sure all items in [qs] are linked to [instance]
        bRedo = False
        for obj in qs:
            # Get the SermonSignature equivalent for Gold signature [obj]
            sermsig = instance.get_sermonsig(obj)
            if sermsig not in linked_qs:
                # Indicate that we need to re-query
                bRedo = True
        # Do we need to re-query?
        if bRedo: 
            # Yes we do...
            linked_qs = SermonSignature.objects.filter(sermon=instance)
        # Remove links that are not in [qs]
        for obj in linked_qs:
            # Get the gold-signature equivalent of this sermon signature
            gsig = obj.get_goldsig()
            # Check if the gold-sermon equivalent is in [qs]
            if gsig not in qs:
                # Remove this item
                obj.delete()
        # Return okay
        return True
    except:
        msg = errHandle.get_error_message()
        return False

def project_dependant_delete(request, to_be_deleted):
    """Delete items from the linktable, provided the user has the right to"""

    oErr = ErrHandle()
    bBack = True
    try:
        # Find out who this is
        profile = Profile.get_user_profile(request.user.username)
        # Get the editing rights for this person
        # project_id = [x['id'] for x in profile.projects.all().values("id")]
        project_id = [x.id for x in profile.get_myprojects()]

        # CHeck all deletables
        delete = []
        for obj in to_be_deleted:
            # Get the project id of the deletables
            obj_id = obj.id
            prj_id = obj.project.id
            if prj_id in project_id:
                # The user may delete this project relation
                # Therefore: delete the OBJ that holde this relation!
                delete.append(obj_id)
        # Is anything left?
        if len(delete) > 0:
            # Get the class of the deletables
            cls = to_be_deleted[0].__class__
            # Delete all that need to be deleted
            cls.objects.filter(id__in=delete).delete()

    except:
        msg = oErr.get_error_message()
        oErr.DoError("project_dependant_delete")
        bBack = False
    return bBack

def get_non_editable_projects(profile, projects):
    """Get the number of projects that I do not have editing rights for"""

    oErr = ErrHandle()
    iCount = 0
    try:
        id_list = []
        current_project_ids = [x['id'] for x in projects.values('id')]
        # profile_project_ids = [x['id'] for x in profile.projects.all().values('id')]
        profile_project_ids = [x.id for x in profile.get_myprojects()]
        # Walk all the projects I need to evaluate
        for prj_id in current_project_ids:
            if not prj_id in profile_project_ids:
                # I have*NO*  editing rights for this one
                id_list.append(prj_id)
        iCount = len(id_list)
    except:
        msg = oErr.get_error_message()
        oErr.DoError("get_non_editable_projects")
        iCount = 0

    return iCount

def evaluate_projlist(profile, instance, projlist, sText):
    bBack = True
    msg = ""
    oErr = ErrHandle()
    try:
        if projlist is None or len(projlist) == 0:
            # Check how many projects the user does *NOT* have rights for
            non_editable_projects = get_non_editable_projects(profile, instance.projects.all())
            if non_editable_projects == 0:
                # The user has not selected a project (yet): try default assignment
                # user_projects = profile.projects.all()
                user_projects = profile.get_defaults()
                if user_projects.count() != 1:
                    # We cannot assign the default project
                    bBack = False
                    msg = "Make sure to assign this {} to one project before saving it".format(sText)
    except:
        msg = oErr.get_error_message()
        oErr.DoError("evaluate_projlist")
        bBack = False
    return bBack, msg

def may_edit_project(request, profile, instance):
    """Check if the user is allowed to edit this project"""

    bBack = False
    # Is the user an editor?
    if user_is_ingroup(request, app_editor):
        # Get the projects this user has authority for
        user_projects = profile.get_project_ids()
        if len(user_projects) > 0:
            # True: the user may indeed edit *some* projects
            bBack = True

            # The following is now superfluous
            use_superfluous = False
            if use_superfluous:
                # Get the projects associated with [instance']
                project_ids = [x['id'] for x in instance.projects.all().values('id')]
                # See if there is any match
                for project_id in user_projects:
                    if project_id in project_ids:
                        bBack = True
                        break
    return bBack

def is_empty_form(form):
    """Check if the indicated form has any cleaned_data"""

    if "cleaned_data" not in form:
        form.is_valid()
    cleaned = form.cleaned_data
    return (len(cleaned) == 0)

def user_is_authenticated(request):
    # Is this user authenticated?
    username = request.user.username
    user = User.objects.filter(username=username).first()
    response = False if user == None else user.is_authenticated
    return response

def user_is_ingroup(request, sGroup):
    # Is this user part of the indicated group?
    user = User.objects.filter(username=request.user.username).first()
    response = username_is_ingroup(user, sGroup)
    return response

def username_is_ingroup(user, sGroup):
    # glist = user.groups.values_list('name', flat=True)

    # Only look at group if the user is known
    if user == None:
        glist = []
    else:
        glist = [x.name for x in user.groups.all()]

        # Only needed for debugging
        if bDebug:
            ErrHandle().Status("User [{}] is in groups: {}".format(user, glist))
    # Evaluate the list
    bIsInGroup = (sGroup in glist)
    return bIsInGroup

def user_is_superuser(request):
    bFound = False
    # Is this user part of the indicated group?
    username = request.user.username
    if username != "":
        user = User.objects.filter(username=username).first()
        if user != None:
            bFound = user.is_superuser
    return bFound

def add_visit(request, name, is_menu):
    """Add the visit to the current path"""

    username = "anonymous" if request.user == None else request.user.username
    if username != "anonymous":
        Visit.add(username, name, request.path, is_menu)

def action_model_changes(form, instance):
    field_values = model_to_dict(instance)
    changed_fields = form.changed_data
    exclude = []
    if hasattr(form, 'exclude'):
        exclude = form.exclude
    changes = {}
    for item in changed_fields: 
        if item in field_values:
            changes[item] = field_values[item]
        elif item not in exclude:
            # It is a form field
            try:
                representation = form.cleaned_data[item]
                if isinstance(representation, QuerySet):
                    # This is a list
                    rep_list = []
                    for rep in representation:
                        rep_str = str(rep)
                        rep_list.append(rep_str)
                    representation = json.dumps(rep_list)
                elif isinstance(representation, str) or isinstance(representation, int):
                    representation = representation
                elif isinstance(representation, object):
                    representation = str(representation)
                changes[item] = representation
            except:
                changes[item] = "(unavailable)"
    return changes

def has_string_value(field, obj):
    response = (field != None and field in obj and obj[field] != None and obj[field] != "")
    return response

def has_list_value(field, obj):
    response = (field != None and field in obj and obj[field] != None and len(obj[field]) > 0)
    return response

def has_obj_value(field, obj):
    response = (field != None and field in obj and obj[field] != None)
    return response

def make_ordering(qs, qd, order_default, order_cols, order_heads):

    oErr = ErrHandle()

    try:
        bAscending = True
        sType = 'str'
        order = []
        colnum = ""
        # reset 'used' feature for all heads
        for item in order_heads: item['used'] = None
        if 'o' in qd and qd['o'] != "":
            colnum = qd['o']
            if '=' in colnum:
                colnum = colnum.split('=')[1]
            if colnum != "":
                order = []
                iOrderCol = int(colnum)
                bAscending = (iOrderCol>0)
                iOrderCol = abs(iOrderCol)

                # Set the column that it is in use
                order_heads[iOrderCol-1]['used'] = 1
                # Get the type
                sType = order_heads[iOrderCol-1]['type']
                for order_item in order_cols[iOrderCol-1].split(";"):
                    if order_item != "":
                        if sType == 'str':
                            order.append(Lower(order_item).asc(nulls_last=True))
                        else:
                            order.append(F(order_item).asc(nulls_last=True))
                if bAscending:
                    order_heads[iOrderCol-1]['order'] = 'o=-{}'.format(iOrderCol)
                else:
                    # order = "-" + order
                    order_heads[iOrderCol-1]['order'] = 'o={}'.format(iOrderCol)

                # Reset the sort order to ascending for all others
                for idx, item in enumerate(order_heads):
                    if idx != iOrderCol - 1:
                        # Reset this sort order
                        order_heads[idx]['order'] = order_heads[idx]['order'].replace("-", "")
        else:
            for order_item in order_default[0].split(";"):
                if order_item != "":
                    order.append(Lower(order_item))
           #  order.append(Lower(order_cols[0]))
        if sType == 'str':
            if len(order) > 0:
                qs = qs.order_by(*order)
        else:
            qs = qs.order_by(*order)
        # Possibly reverse the order
        if not bAscending:
            qs = qs.reverse()
    except:
        msg = oErr.get_error_message()
        oErr.DoError("seeker/views/make_ordering")
        lstQ = []

    return qs, order_heads, colnum

def process_visit(request, name, is_menu, **kwargs):
    """Process one visit and return updated breadcrumbs"""

    username = "anonymous" if request.user == None else request.user.username
    if username != "anonymous" and request.user.username != "":
        # Add the visit
        Visit.add(username, name, request.get_full_path(), is_menu, **kwargs)
        # Get the updated path list
        p_list = Profile.get_stack(username)
    else:
        p_list = []
        p_list.append({'name': 'Home', 'url': reverse('home')})
    # Return the breadcrumbs
    # return json.dumps(p_list)
    return p_list

def get_breadcrumbs(request, name, is_menu, lst_crumb=[], **kwargs):
    """Process one visit and return updated breadcrumbs"""

    # Initialisations
    p_list = []
    p_list.append({'name': 'Home', 'url': reverse('home')})
    # Find out who this is
    username = "anonymous" if request.user == None else request.user.username
    if username != "anonymous" and request.user.username != "":
        # Add the visit
        currenturl = request.get_full_path()
        Visit.add(username, name, currenturl, is_menu, **kwargs)
        # Set the full path, dependent on the arguments we get
        for crumb in lst_crumb:
            if len(crumb) == 2:
                p_list.append(dict(name=crumb[0], url=crumb[1]))
            else:
                pass
        # Also add the final one
        p_list.append(dict(name=name, url=currenturl))
    # Return the breadcrumbs
    return p_list

def get_previous_page(request, top=False):
    """Find the previous page for this user"""

    username = "anonymous" if request.user == None else request.user.username
    if username != "anonymous" and request.user.username != "":
        # Get the current path list
        p_list = Profile.get_stack(username)
        p_item = []
        if len(p_list) < 2:
            prevpage = request.META.get('HTTP_REFERER') 
        elif top:
            p_item = p_list[len(p_list)-1]
            prevpage = p_item['url']
        else:
            p_item = p_list[len(p_list)-2]
            prevpage = p_item['url']
        # Possibly add arguments
        if 'kwargs' in p_item:
            # First strip off any arguments (anything after ?) in the url
            if "?" in prevpage:
                prevpage = prevpage.split("?")[0]
            bFirst = True
            for k,v in p_item['kwargs'].items():
                if bFirst:
                    addsign = "?"
                    bFirst = False
                else:
                    addsign = "&"
                prevpage = "{}{}{}={}".format(prevpage, addsign, k, v)
    else:
        prevpage = request.META.get('HTTP_REFERER') 
    # Return the path
    return prevpage

def adapt_regex_incexp(value):
    """Widen searching for incipit and explicit
    
    e=ae, j=i, u=v, k=c
    """

    oTranslation = str.maketrans(dict(j="[ji]", i="[ji]", u="[uv]", v="[uv]", k="[kc]", c="[kc]"))

    if value != None and len(value) > 0:
        # Make changes:
        value = value.replace("ae", "e").replace("e", "a?e").translate(oTranslation)

    return value

def get_saveditem_html(request, instance, profile, htmltype="button", sitemtype=None):
    """Get an indication whether this is a saved item, or allow to add it"""

    oErr = ErrHandle()
    sBack = ""
    try:
        if not sitemtype is None:
            obj = SavedItem.get_saveditem(instance, profile, sitemtype)
            context = {}
            context['profile'] = profile
            context['saveditem'] = obj
            context['item'] = instance
            context['sitemtype'] = sitemtype
            context['sitemaction'] = "add" if obj is None else "remove"

            template_name = "dct/sitem_button.html"
            if htmltype == "form":
                template_name = "dct/sitem_form.html"

            sBack = render_to_string(template_name, context, request)
    except:
        msg = oErr.get_error_message()
        oErr.DoError("get_saveditem_html")
    return sBack

def get_selectitem_info(request, instance, profile, selitemtype=None, context={}):
    """Get an indication whether this is a select item, or allow to add it"""

    oErr = ErrHandle()
    sBack = ""
    try:
        if not selitemtype is None:
            html = []
            # context = {}
            context['profile'] = profile
            context['selitemtype'] = selitemtype

            # Is this the main form or not?
            if instance is None:
                # This is the main form
                template_name = "dct/selitem_main.html"
                html.append( render_to_string(template_name, context, request))

            else:

                obj = SelectItem.get_selectitem(instance, profile, selitemtype)
                context['selitem'] = obj
                context['item'] = instance
                context['selitemaction'] = "add" if obj is None else "remove"

                template_name = "dct/selitem_button.html"
                html.append( render_to_string(template_name, context, request))

                template_name = "dct/selitem_form.html"
                html.append( render_to_string(template_name, context, request))

            sBack = "\n".join(html)
    except:
        msg = oErr.get_error_message()
        oErr.DoError("get_selectitem_info")
    return sBack





# ================= STANDARD views =====================================

def home(request, errortype=None):
    """Renders the home page."""

    assert isinstance(request, HttpRequest)
    # Specify the template
    template_name = 'index.html'

    bOverrideSync = False
    bDebug = False
    oErr = ErrHandle()

    try:
        # Define the initial context
        context =  {'title': 'RU-passim',
                    'year':get_current_datetime().year,
                    'pfx': APP_PREFIX,
                    'site_url': admin.site.site_url}

        context = get_application_context(request, context)
        # NOTE: the context now contains items like 'is_app_editor'

        # Process this visit
        context['breadcrumbs'] = get_breadcrumbs(request, "Home", True)

        # See if this is the result of a particular error
        if errortype != None:
            if errortype == "404":
                context['is_404'] = True

        # Check the newsitems for validity
        if bDebug: 
            # ========== DEBUG ============
            print("- Calling check_until")
        NewsItem.check_until()

        # Create the list of news-items
        lstQ = []
        lstQ.append(Q(status='val'))
        newsitem_list = NewsItem.objects.filter(*lstQ).order_by('-created', '-saved')
        context['newsitem_list'] = newsitem_list


        # Gather the statistics
        if bDebug: 
            # ========== DEBUG ============
            print("counting for statistics")
        context['count_sermon'] = SermonDescr.objects.exclude(mtype="tem").count()
        context['count_manu'] = Manuscript.objects.exclude(mtype="tem").count()

        # Gather pie-chart data
        if bDebug: 
            # ========== DEBUG ============
            print("Fetching pie chart data")
        context['pie_data'] = get_pie_data()

        # Possibly start getting new Stemmatology results
        if bOverrideSync and user_is_superuser(request):
            scan_transcriptions()
        else:
            thread = Thread(target=scan_transcriptions)
            thread.start()

        # Check if the user's / profile's information is up-to-date
        user = request.user
        if not request is None and not user is None and not user.id is None:
            profile = user.user_profiles.first()
            if not profile is None:
                # Validate information from this user
                bNeeds, msg = profile.needs_updating()
                if bNeeds:
                    # Provide popup for user
                    context['profile_msg'] = msg
                    context['profile_url'] = reverse('profile_details', kwargs={'pk': profile.id})

        # Calculate the response
        response = render(request, template_name, context)
    except:
        msg = oErr.get_error_message()
        oErr.DoError("home")
        response = "Home error: {}".format(msg)

    # Render and return the page
    return response

def view_404(request, *args, **kwargs):
    # return home(request, "404")
    sBack = "Error 404 when loading: {}".format(request.path)
    return sBack

def contact(request):
    """Renders the contact page."""
    assert isinstance(request, HttpRequest)
    context =  {'title':'Contact',
                'message':'Shari Boodts',
                'year':get_current_datetime().year,
                'pfx': APP_PREFIX,
                'site_url': admin.site.site_url}

    context = get_application_context(request, context)
    # context['is_app_uploader'] = user_is_ingroup(request, app_uploader)

    # Process this visit
    context['breadcrumbs'] = get_breadcrumbs(request, "Contact", True)

    return render(request,'contact.html', context)

def more(request):
    """Renders the more page."""
    assert isinstance(request, HttpRequest)
    context =  {'title':'More',
                'year':get_current_datetime().year,
                'pfx': APP_PREFIX,
                'site_url': admin.site.site_url}
    context = get_application_context(request, context)
    # context['is_app_uploader'] = user_is_ingroup(request, app_uploader)

    # Process this visit
    context['breadcrumbs'] = get_breadcrumbs(request, "More", True)

    return render(request,'more.html', context)

def technical(request):
    """Renders the technical information page."""
    assert isinstance(request, HttpRequest)
    # Specify the template
    template_name = 'technical.html'
    context =  {'title':'Technical',
                'year':get_current_datetime().year,
                'pfx': APP_PREFIX,
                'site_url': admin.site.site_url}
    context = get_application_context(request, context)
    #context['is_app_uploader'] = user_is_ingroup(request, app_uploader)
    #context['is_app_editor'] = user_is_ingroup(request, app_editor)
    #context['is_enrich_editor'] = user_is_ingroup(request, enrich_editor)
    #context['is_app_moderator'] = user_is_superuser(request) or user_is_ingroup(request, app_moderator)

    # Process this visit
    context['breadcrumbs'] = get_breadcrumbs(request, "Technical", True)

    return render(request,template_name, context)

def guide(request):
    """Renders the user-manual (guide) page."""
    assert isinstance(request, HttpRequest)
    # Specify the template
    template_name = 'guide.html'
    context =  {'title':'User manual',
                'year':get_current_datetime().year,
                'pfx': APP_PREFIX,
                'site_url': admin.site.site_url}
    context = get_application_context(request, context)
    #context['is_app_uploader'] = user_is_ingroup(request, app_uploader)
    #context['is_app_editor'] = user_is_ingroup(request, app_editor)
    #context['is_enrich_editor'] = user_is_ingroup(request, enrich_editor)
    #context['is_app_moderator'] = user_is_superuser(request) or user_is_ingroup(request, app_moderator)

    # Process this visit
    context['breadcrumbs'] = get_breadcrumbs(request, "Guide", True)

    return render(request,template_name, context)

def bibliography(request):
    """Renders the more page."""
    assert isinstance(request, HttpRequest)
    context =  {'title':'Bibliography',
                'year':datetime.now().year,
                'pfx': APP_PREFIX,
                'site_url': admin.site.site_url}
    context = get_application_context(request, context)
    # context['is_app_uploader'] = user_is_ingroup(request, app_uploader)

    # Add the edition references (abreviated and full)



    # Create empty list for the editions
    edition_list = []
    
    # Retrieve all records from the Zotero database (user for now)
    zot = zotero.Zotero('5802673', 'user', 'oVBhIJH5elqA8zxrJGwInwWd')
    
    # Store only the records from the Edition collection (key: from URL 7HQU3AY8)
    zot_editions = zot.collection_items('7HQU3AY8')
   
    for item in zot_editions:
        creators_ed = item['data']['creators']
        creator_list_ed = []
        for creator in creators_ed:
            first_name = creator['firstName']
            last_name = creator['lastName']
            creator_list_ed.append(first_name + " " + last_name)
        edition_list.append({'abbr': item['data']['extra'] + ", " + item['data']['pages'], 'full': item['data']['title'], 'creators': ", ".join(creator_list_ed)})
   
    context['edition_list'] = edition_list    
    
    # Add the literature references from zotero, data extra 
        
    # Create empty list for the literature
    
    reference_list = [] 
    
    # Retrieve all records from the Zotero database (user for now)
    zot = zotero.Zotero('5802673', 'user', 'oVBhIJH5elqA8zxrJGwInwWd')
    
    # Store only the records in the Literature collection,  (key: from URL FEPVSGVX)
    zot_literature = zot.collection_items('FEPVSGVX')
        
    
    for item in zot_literature:
        creators = item['data']['creators']
        creator_list = []
        for creator in creators:
            first_name = creator['firstName']
            last_name = creator['lastName']
            creator_list.append(first_name + " " + last_name)
        reference_list.append({'abbr':creator['lastName'] + ", " + item['data']['extra'] + " " + item['data']['volume'] + " (" + item['data']['date']+ ")"+", "+item['data']['pages'], 'full': item['data']['title'], 'creators': ", ".join(creator_list)})
    context['reference_list'] = reference_list

    # Process this visit
    context['breadcrumbs'] = get_breadcrumbs(request, "Bibliography", True)

    return render(request,'bibliography.html', context)

def about(request):
    """Renders the about page."""
    assert isinstance(request, HttpRequest)
    context =  {'title':'About',
                'message':'Radboud University passim utility.',
                'year':get_current_datetime().year,
                'pfx': APP_PREFIX,
                'site_url': admin.site.site_url}
    context = get_application_context(request, context)
    # context['is_app_uploader'] = user_is_ingroup(request, app_uploader)

    # Calculate statistics
    sites = {}
    people = {}
    for obj in SourceInfo.objects.all().order_by('url', 'collector'):
        if obj.url == None or obj.url == "":
            # Look at PEOPLE
            collector = obj.collector
            if collector != None and collector != "":
                if not collector in people:
                    people[collector] = obj.sourcemanuscripts.filter(mtype='man').count()
                else:
                    people[collector] += obj.sourcemanuscripts.filter(mtype='man').count()
        elif obj.url != None and obj.url != "":
            # Look at SITES
            collector = obj.url
            if collector != None and collector != "":
                if not collector in sites:
                    sites[collector] = obj.sourcemanuscripts.filter(mtype='man').count()
                else:
                    sites[collector] += obj.sourcemanuscripts.filter(mtype='man').count()
    context['sites'] = [{"url": k, "count": v} for k,v in sites.items()]
    people_lst = [{"count": v, "person": k} for k,v in people.items()]
    people_lst = sorted(people_lst, key = lambda x: x['count'], reverse=True)
    context['people'] = people_lst

    # Process this visit
    context['breadcrumbs'] = get_breadcrumbs(request, "About", True)

    return render(request,'about.html', context)

def short(request):
    """Renders the page with the short instructions."""

    assert isinstance(request, HttpRequest)
    template = 'short.html'
    context = {'title': 'Short overview',
               'message': 'Radboud University passim short intro',
               'year': get_current_datetime().year}
    context = get_application_context(request, context)
    # context['is_app_uploader'] = user_is_ingroup(request, app_uploader)
    return render(request, template, context)

def nlogin(request):
    """Renders the not-logged-in page."""
    assert isinstance(request, HttpRequest)
    context =  {    'title':'Not logged in', 
                    'message':'Radboud University passim utility.',
                    'year':get_current_datetime().year,}
    context = get_application_context(request, context)
    # context['is_app_uploader'] = user_is_ingroup(request, app_uploader)
    return render(request,'nlogin.html', context)

# ================ OTHER VIEW HELP FUNCTIONS ============================

def sync_passim(request):
    """-"""
    assert isinstance(request, HttpRequest)

    # Gather info
    context = {'title': 'SyncPassim',
               'message': 'Radboud University PASSIM'
               }
    template_name = 'seeker/syncpassim.html'
    context = get_application_context(request, context)
    #context['is_app_uploader'] = user_is_ingroup(request, app_uploader)
    #context['is_app_editor'] = user_is_ingroup(request, app_editor)
    #context['is_enrich_editor'] = user_is_ingroup(request, enrich_editor)
    #context['is_app_moderator'] = user_is_superuser(request) or user_is_ingroup(request, app_moderator)
    #context['is_superuser'] = user_is_superuser(request)

    # Add the information in the 'context' of the web page
    return render(request, template_name, context)

def sync_start(request):
    """Synchronize information"""

    oErr = ErrHandle()
    data = {'status': 'starting'}
    try:
        # Get the user
        username = request.user.username
        # Authentication
        if not user_is_ingroup(request, app_editor):
            return redirect('home')

        # Get the synchronization type
        get = request.GET
        synctype = ""
        force = False
        if 'synctype' in get:
            synctype = get['synctype']
        if 'force' in get:
            force = get['force']
            force = (force == "true" or force == "1" )

        if synctype == '':
            # Formulate a response
            data['status'] = 'no sync type specified'

        else:
            # Remove previous status objects for this combination of user/type
            qs = Status.objects.filter(user=username, type=synctype)
            qs.delete()

            # Create a status object for this combination of synctype/user
            oStatus = Status(user=username, type=synctype, status="preparing")
            oStatus.save()

            # Formulate a response
            data['status'] = 'done'

            if synctype == "entries":
                # Use the synchronisation object that contains all relevant information
                oStatus.set("loading")

                # Update the models with the new information
                oResult = process_lib_entries(oStatus)
                if oResult == None or oResult['result'] == False:
                    data.status = 'error'
                elif oResult != None:
                    data['count'] = oResult

            elif synctype == "zotero":
                # Use the synchronisation object that contains all relevant information
                oStatus.set("loading")

                # Update the models with the new information
                oResult, msg = Litref.sync_zotero(force=force, oStatus=oStatus)

                if oResult != None and 'status' in oResult:
                    data['count'] = oResult
                else:
                    data['status'] = 'error {}'.format(msg)

            elif synctype == "codico":
                # Use the synchronisation object that contains all relevant information
                oStatus.set("loading")

                # Perform the adaptation
                bResult, msg = adapt_codicocopy(oStatus=oStatus)
                
                if bResult:
                    data['count'] = 1
                else:
                    data['status'] = "error {}".format(msg) 

            elif synctype == "stemma":
                # Use the synchronisation object that contains all relevant information
                oStatus.set("loading")

                # Start up synchronisation
                bResult, msg = sync_transcriptions(oStatus=oStatus)
                
                if bResult:
                    data['count'] = 1
                else:
                    data['status'] = "error {}".format(msg) 

    except:
        oErr.DoError("sync_start error")
        data['status'] = "error"

    # Return this response
    return JsonResponse(data)

def sync_progress(request):
    """Get the progress on the /crpp synchronisation process"""

    oErr = ErrHandle()
    data = {'status': 'preparing'}

    try:
        # Get the user
        username = request.user.username
        # Get the synchronization type
        get = request.GET
        synctype = ""
        if 'synctype' in get:
            synctype = get['synctype']

        if synctype == '':
            # Formulate a response
            data['status'] = 'error'
            data['msg'] = "no sync type specified" 

        else:
            # Formulate a response
            data['status'] = 'UNKNOWN'

            # Get the appropriate status object
            # sleep(1)
            oStatus = Status.objects.filter(user=username, type=synctype).first()

            # Check what we received
            if oStatus == None:
                # There is no status object for this type
                data['status'] = 'error'
                data['msg'] = "Cannot find status for {}/{}".format(
                    username, synctype)
            else:
                # Get the last status information
                data['status'] = oStatus.status
                data['msg'] = oStatus.msg
                data['count'] = oStatus.count

        # Return this response
        return JsonResponse(data)
    except:
        oErr.DoError("sync_start error")
        data = {'status': 'error'}

    # Return this response
    return JsonResponse(data)

def search_generic(s_view, cls, form, qd, username=None, team_group=None):
    """Generic queryset generation for searching"""

    qs = None
    oErr = ErrHandle()
    bFilter = False
    bUsingMan = False
    genForm = None
    oFields = {}
    try:
        bHasFormset = (len(qd) > 0)

        if bHasFormset:
            # Get the formset from the input
            lstQ = []
            prefix = s_view.prefix
            filters = s_view.filters
            searches = s_view.searches
            lstExclude = []

            if s_view.use_team_group:
                genForm = form(qd, prefix=prefix, username=username, team_group=team_group)
            else:
                genForm = form(qd, prefix=prefix)

            if genForm.is_valid():

                # Process the criteria from this form 
                oFields = genForm.cleaned_data

                # Adapt the search for empty passim codes
                if 'codetype' in oFields:
                    codetype = oFields['codetype']
                    if codetype == "non":
                        lstExclude = []
                        lstExclude.append(Q(equal__isnull=False))
                    elif codetype == "spe":
                        lstExclude = []
                        lstExclude.append(Q(equal__isnull=True))
                    # Reset the codetype
                    oFields['codetype'] = ""  
                    
                # Adapt search for soperator/scount
                if 'soperator' in oFields:
                    if not 'scount' in oFields or oFields['soperator'] == "-":
                        oFields.pop('soperator') 

                # Adapt search for mtype, if that is not specified
                if 'mtype' in oFields and oFields['mtype'] == "":
                    # Make sure we only select MAN and not TEM (template)
                    oFields['mtype'] = "man"
                    bUsingMan = True
                                 
                # Create the search based on the specification in searches
                filters, lstQ, qd, lstExclude = make_search_list(filters, oFields, searches, qd, lstExclude)

                # Calculate the final qs
                if len(lstQ) == 0 or (bUsingMan and len(lstQ) == 1):
                    # No filter: Just show everything
                    # qs = cls.objects.all()
                    # No filter: do *NOT* show anything
                    qs = cls.objects.none()
                else:
                    # There is a filter: apply it
                    qs = cls.objects.filter(*lstQ).distinct()
                    bFilter = True
            else:
                # TODO: communicate the error to the user???

                # Just show NOTHING
                qs = cls.objects.none()

        else:
            # Just show everything
            qs = cls.objects.all().distinct()
    except:
        msg = oErr.get_error_message()
        oErr.DoError("search_generic")
        qs = None
        bFilter = False
    # Return the resulting filtered and sorted queryset
    return filters, bFilter, qs, qd, oFields

def search_collection(request):
    """Search for a collection"""

    # Set defaults
    template_name = "seeker/collection.html"

    # Get a link to a form
    searchForm = SearchCollectionForm()

    # Other initialisations
    currentuser = request.user
    authenticated = currentuser.is_authenticated

    # Create context and add to it
    context = dict(title="Search collection",
                   authenticated=authenticated,
                   searchForm=searchForm)
    context['is_app_uploader'] = user_is_ingroup(request, app_uploader)

    # Create and show the result
    return render(request, template_name, context)

def login_as_user(request, user_id):
    assert isinstance(request, HttpRequest)

    # Find out who I am
    supername = request.user.username
    super = User.objects.filter(username__iexact=supername).first()
    if super == None:
        return nlogin(request)

    # Make sure that I am superuser
    if super.is_staff and super.is_superuser:
        user = User.objects.filter(username__iexact=user_id).first()
        if user != None:
            # Perform the login
            login(request, user)
            return HttpResponseRedirect(reverse("home"))

    return home(request)

def signup(request):
    """Provide basic sign up and validation of it """

    if request.method == 'POST':
        form = SignUpForm(request.POST)
        if form.is_valid():
            # Save the form
            form.save()
            # Create the user
            username = form.cleaned_data.get('username')
            raw_password = form.cleaned_data.get('password1')
            affiliation = form.cleaned_data.get('affiliation')

            # also make sure that the user gets into the STAFF,
            #      otherwise he/she may not see the admin pages
            user = authenticate(username=username, 
                                password=raw_password,
                                is_staff=True)
            user.is_staff = True
            user.save()
            # Add user to the "passim_user" group
            gQs = Group.objects.filter(name="passim_user")
            if gQs.count() > 0:
                g = gQs[0]
                g.user_set.add(user)

            # Add a profile for this new user
            profile = Profile.objects.filter(user=user).first()
            if profile == None:
                # There is no profile yet, so make it
                profile = Profile.objects.create(user=user, affiliation=affiliation)

            # Log in as the user
            login(request, user)
            return redirect('home')
    else:
        form = SignUpForm()
    return render(request, 'signup.html', {'form': form})

def redo_zotero(request):
    oErr = ErrHandle()
    data = {'status': 'preparing'}

    try:
        if request.method == 'GET':
            oBack, msg = Litref.sync_zotero(True)
        if oBack != None and 'status' in oBack:
            data['status'] = "ok"
            data['count'] = oBack
        else:
            data['status'] = "error"
            data['msg'] = msg
            data['html'] = msg
    except:
        oErr.DoError("redo_zotero error")
        data = {'status': 'error'}

    # Return this response
    return JsonResponse(data)

def do_clavis(request):
    # Create a regular expression
    r_number = re.compile( r'^[[]?(\d+)[]]?')

    # Walk the whole signature list
    with transaction.atomic():
        for obj in Signature.objects.all():
            # Check the type
            if obj.editype == "cl":
                # This is clavis -- get the code
                code = obj.code
                if r_number.match(code):
                    # This is only a number
                    obj.code = "CPPM I {}".format(code)
                    obj.save()
                else:
                    # Break it up into spaced items
                    arCode = code.split(" ")
                    if len(arCode) > 1:
                        if arCode[1] != "I":
                            if len(arCode) > 2:
                                iStop = True
                            arCode[0] = arCode[0] + " I"
                            obj.code = " ".join(arCode)
                            obj.save()
                    
    # Walk the whole gold-signature list
    with transaction.atomic():
        for obj in SermonSignature.objects.all():
            # Check the type
            if obj.editype == "cl":
                # This is clavis -- get the code
                code = obj.code
                if r_number.match(code):
                    # This is only a number
                    obj.code = "CPPM I {}".format(code)
                    obj.save()
                else:
                    # Break it up into spaced items
                    arCode = code.split(" ")
                    if len(arCode) > 1:
                        if arCode[1] != "I":
                            if len(arCode) > 2:
                                iStop = True
                            arCode[0] = arCode[0] + " I"
                            obj.code = " ".join(arCode)
                            obj.save()

    # Return an appropriate page
    return home(request)

def do_stype(request):
    """Add stype on the appropriate places"""

    oErr = ErrHandle()
    phases = []
    try:
        assert isinstance(request, HttpRequest)
        # Specify the template
        template_name = 'tools.html'
        # Define the initial context
        context =  {'title':'RU-passim-tools',
                    'year':get_current_datetime().year,
                    'pfx': APP_PREFIX,
                    'site_url': admin.site.site_url}
        context['is_app_uploader'] = user_is_ingroup(request, app_uploader)
        context['is_app_editor'] = user_is_ingroup(request, app_editor)

        # Only passim uploaders can do this
        if not context['is_app_uploader']: return reverse('home')

        # Indicate the necessary tools sub-part
        context['tools_part'] = "Repair Stype definitions"

        # Process this visit
        context['breadcrumbs'] = get_breadcrumbs(request, "Stype", True)

        # Create list to be returned
        result_list = []

        # Phase 1: Manuscript
        if '1' in phases:
            with transaction.atomic():
                added = 0
                for item in Manuscript.objects.all():
                    if item.stype == "-":
                        item.stype = STYPE_IMPORTED
                        item.save()
                        added += 1
                result_list.append({'part': 'Manuscript changed stype', 'result': added})

        # Phase 2: SermonDescr
        if '2' in phases:
            with transaction.atomic():
                added = 0
                for item in SermonDescr.objects.all():
                    if item.stype == "-":
                        item.stype = STYPE_IMPORTED
                        item.save()
                        added += 1
                result_list.append({'part': 'SermonDescr changed stype', 'result': added})

        # Phase 3: SermonGold
        if '3' in phases:
            with transaction.atomic():
                added = 0
                for item in SermonGold.objects.all():
                    if item.stype == "-":
                        item.stype = STYPE_IMPORTED
                        item.save()
                        added += 1
                result_list.append({'part': 'SermonGold changed stype', 'result': added})

        # Phase 4: EqualGold
        if '4' in phases:
            with transaction.atomic():
                added = 0
                for item in EqualGold.objects.all():
                    if item.stype == "-":
                        item.stype = STYPE_IMPORTED
                        item.save()
                        added += 1
                    else:
                        # Check i it is in Action
                        if item.sgcount > 1 or Action.objects.filter(itemtype='EqualGold', itemid=item.id).exists():
                            item.stype = STYPE_EDITED
                            item.save()
                            added += 1
                        elif item.equalgold_src.all().exists():
                            item.stype = STYPE_EDITED
                            item.save()
                            added += 1
                result_list.append({'part': 'EqualGold changed stype', 'result': added})

        context['result_list'] = result_list
    
        # Render and return the page
        return render(request, template_name, context)
    except:
        msg = oErr.get_error_message()
        oErr.DoError("goldtogold")
        return reverse('home')

def do_locations(request):
    """Add stype on the appropriate places"""

    oErr = ErrHandle()
    try:
        assert isinstance(request, HttpRequest)
        # Specify the template
        template_name = 'tools.html'
        # Define the initial context
        context =  {'title':'RU-passim-tools',
                    'year':get_current_datetime().year,
                    'pfx': APP_PREFIX,
                    'site_url': admin.site.site_url}
        context['is_app_uploader'] = user_is_ingroup(request, app_uploader)
        context['is_app_editor'] = user_is_ingroup(request, app_editor)

        # Only passim uploaders can do this
        if not context['is_app_uploader']: return reverse('home')

        # Indicate the necessary tools sub-part
        context['tools_part'] = "Establish location definitions"

        # Process this visit
        context['breadcrumbs'] = get_breadcrumbs(request, "Locations", True)

        # Create list to be returned
        result_list = []

        # Phase 1: Country
        name = "country"
        loctype_country = LocationType.objects.filter(name=name).first()
        idname = "idPaysEtab"
        added = 0
        if Location.objects.filter(loctype=loctype_country).count() == 0:
            with transaction.atomic():
                # locationtype = city
                for item in Country.objects.all():
                    # Get the details of this item
                    name = item.name
                    nameFR = item.nameFR
                    idvalue = item.idPaysEtab
                    # Create a location with these parameters
                    obj = Location(name=name, loctype=loctype_country)
                    obj.save()
                    added += 1
                    # Add the French name and identifier
                    if nameFR != None and nameFR != "":
                        lname = LocationName(name=nameFR, language="fre", location=obj)
                        lname.save()
                    if idvalue != None:
                        lidt = LocationIdentifier(idname=idname, idvalue=idvalue, location=obj)
                        lidt.save()

        # Phase 2: City
        name = "city"
        loctype_city = LocationType.objects.filter(name=name).first()
        idname = "idVilleEtab"
        if Location.objects.filter(loctype=loctype_city).count() == 0:
            last_country = None
            last_loc_country = None
            with transaction.atomic():
                # locationtype = city
                for item in City.objects.all().order_by('country__name'):
                    # Get the details of this item
                    name = item.name
                    idvalue = item.idVilleEtab
                    country = item.country
                    # Create a location with these parameters
                    obj = Location(name=name, loctype=loctype_city)
                    obj.save()
                    added += 1
                    # Add the French identifier
                    if idvalue != None and idvalue >= 0:
                        lidt = LocationIdentifier(idname=idname, idvalue=idvalue, location=obj)
                        lidt.save()
                    # Tie the city to the correct country object through a LocationRelation
                    if country != None:
                        if country != last_country:
                            loc = Location.objects.filter(name=country.name, loctype=loctype_country).first()
                            if loc != None:
                                last_loc_country = loc
                                last_country = country
                        # Implement relation
                        loc_rel = LocationRelation(container=last_loc_country, contained=obj)
                        loc_rel.save()

        # Phase 3: Library
        name = "library"
        loctype_lib = LocationType.objects.filter(name=name).first()
        qs = Library.objects.all().order_by('country__name', 'city__name')
        with transaction.atomic():
            for library in qs:
                if library.location == None:
                    city = library.city
                    country = library.country
                    if country != None:
                        countries = Location.objects.filter(name__iexact=country.name)
                    if city != None:
                    
                        # Find the city
                        lstQ = []
                        lstQ.append(Q(name__iexact=city.name))
                        if country != None:
                            lstQ.append(Q(relations_location__in=countries))
                        obj = Location.objects.filter(*lstQ).first()
                        # qs = Location.objects.filter(name__iexact=city.name).filter(relations_location__in=countries)
                        if obj == None:
                            # Cannot find the location of this library...
                            iStop = 1
                        else:
                            # Set the location in Library
                            library.location = obj
                            library.save()
                            added += 1

        # Wrapping it up
        context['result_list'] = result_list

        # Render and return the page
        return render(request, template_name, context)
    except:
        msg = oErr.get_error_message()
        oErr.DoError("do_locations")
        return reverse('home')

def do_provenance(request):
    """Add stype on the appropriate places"""

    oErr = ErrHandle()
    try:
        assert isinstance(request, HttpRequest)
        # Specify the template
        template_name = 'tools.html'
        # Define the initial context
        context =  {'title':'RU-passim-tools',
                    'year':get_current_datetime().year,
                    'pfx': APP_PREFIX,
                    'site_url': admin.site.site_url}
        context['is_app_uploader'] = user_is_ingroup(request, app_uploader)
        context['is_app_editor'] = user_is_ingroup(request, app_editor)

        # Only passim uploaders can do this
        if not context['is_app_uploader']: return reverse('home')

        # Indicate the necessary tools sub-part
        context['tools_part'] = "Tweak Manuscript-Provenance connections"

        # Process this visit
        context['breadcrumbs'] = get_breadcrumbs(request, "Provenance", True)

        # Create list to be returned
        result_list = []

        # Get a list of all Manuscript-Provenance links
        qs = ProvenanceMan.objects.all().order_by('manuscript', 'provenance')
        lst_del = []
        man_id = -1
        prov_id = -1
        for item in qs:
            if item.manuscript.id != man_id:
                # This is a new manuscript
                prov_id = -1
                man_id = item.manuscript.id
            if item.provenance.id == prov_id or item.provenance.id == 1:
                # Set for removal
                lst_del.append(item.id)
            # Take over the prov id
            prov_id = item.provenance.id
        # Remove them
        ProvenanceMan.objects.filter(id__in=lst_del).delete()


        # Wrapping it up
        context['result_list'] = lst_del

        # Render and return the page
        return render(request, template_name, context)
    except:
        msg = oErr.get_error_message()
        oErr.DoError("do_provenance")
        return reverse('home')

def do_mext(request):
    """Copy all 'url' fields from Manuscript instances to separate ManuscriptExt instances and link them to the Manuscript"""

    oErr = ErrHandle()
    try:
        assert isinstance(request, HttpRequest)
        # Specify the template
        template_name = 'tools.html'
        # Define the initial context
        context =  {'title':'RU-passim-tools',
                    'year':get_current_datetime().year,
                    'pfx': APP_PREFIX,
                    'site_url': admin.site.site_url}
        context['is_app_uploader'] = user_is_ingroup(request, app_uploader)
        context['is_app_editor'] = user_is_ingroup(request, app_editor)

        # Only passim uploaders can do this
        if not context['is_app_uploader']: return reverse('home')

        # Indicate the necessary tools sub-part
        context['tools_part'] = "Copy Manuscript links to externals"

        # Process this visit
        context['breadcrumbs'] = get_breadcrumbs(request, "Mext", True)

        # Create list to be returned
        result_list = []
        added_list = []

        # Get a list of all Manuscripts that are not referred to via a ManuscriptExt
        man_ids = [x.manuscript.id for x in ManuscriptExt.objects.all() ]

        qs = Manuscript.objects.exclude(id__in=man_ids)
        with transaction.atomic():
            for item in qs:
                # Get any possible URL here
                url = item.url
                if url != None:
                    # There is an actual URL: Create a new ManuscriptExt instance
                    mext = ManuscriptExt(url=url, manuscript=item)
                    mext.save()
                    added_list.append(mext.id)

        # Wrapping it up
        result_list.append({'part': 'All additions', 'result': json.dumps(added_list)})
        context['result_list'] = result_list

        # Render and return the page
        return render(request, template_name, context)
    except:
        msg = oErr.get_error_message()
        oErr.DoError("do_mext")
        return reverse('home')

def do_goldsearch(request):
    """Re-calculate the srchincipit and srchexplicit fields for gold sermons"""

    oErr = ErrHandle()
    try:
        assert isinstance(request, HttpRequest)
        # Specify the template
        template_name = 'tools.html'
        # Define the initial context
        context =  {'title':'RU-passim-tools',
                    'year':get_current_datetime().year,
                    'pfx': APP_PREFIX,
                    'site_url': admin.site.site_url}
        context['is_app_uploader'] = user_is_ingroup(request, app_uploader)
        context['is_app_editor'] = user_is_ingroup(request, app_editor)

        # Only passim uploaders can do this
        if not context['is_app_uploader']: return reverse('home')

        # Indicate the necessary tools sub-part
        context['tools_part'] = "Re-create Gold sermon searching (incipit/explicit)"

        # Process this visit
        context['breadcrumbs'] = get_breadcrumbs(request, "Sermons", True)

        # Start up processing
        added = 0
        with transaction.atomic():
            for item in SermonGold.objects.all():
                srchincipit = item.srchincipit
                srchexplicit = item.srchexplicit
                # Double check the equal field
                if item.equal == None:
                    geq = EqualGold.objects.create()
                    item.equal = geq
                item.save()
                if item.srchincipit != srchincipit or item.srchexplicit != srchexplicit:
                    added += 1
   
        # Create list to be returned
        result_list = []
        result_list.append({'part': 'Number of changed gold-sermons', 'result': added})

        context['result_list'] = result_list
    
        # Render and return the page
        return render(request, template_name, context)
    except:
        msg = oErr.get_error_message()
        oErr.DoError("do_goldsearch")
        return reverse('home')

def do_sermons(request):
    """Remove duplicate sermons from manuscripts and/or do other SermonDescr initialisations"""

    oErr = ErrHandle()
    try:
        assert isinstance(request, HttpRequest)
        # Specify the template
        template_name = 'tools.html'
        # Define the initial context
        context =  {'title':'RU-passim-tools',
                    'year':get_current_datetime().year,
                    'pfx': APP_PREFIX,
                    'site_url': admin.site.site_url}
        context['is_app_uploader'] = user_is_ingroup(request, app_uploader)
        context['is_app_editor'] = user_is_ingroup(request, app_editor)

        # Only passim uploaders can do this
        if not context['is_app_uploader']: return reverse('home')

        # Indicate the necessary tools sub-part
        context['tools_part'] = "Repair manuscript-sermons"

        # Process this visit
        context['breadcrumbs'] = get_breadcrumbs(request, "Sermons", True)
    
        # Start up processing
        added = 0
        lst_total = []
        lst_total.append("<table><thead><tr><th>Manuscript</th><th>Sermon</th></tr>")
        lst_total.append("<tbody>")

        # Step #1: walk all manuscripts
        qs_m = Manuscript.objects.all().order_by('id')
        for manu in qs_m:
            # Get all the sermons for this manuscript in appropriate order (reverse ID)
            sermon_lst = SermonDescr.objects.filter(msitem__manu=manu).order_by('-id').values('id', 'title', 'author', 'nickname', 'locus', 'incipit', 'explicit', 'note', 'additional', 'order')
            remove_lst = []
            if manu.id == 1245:
                iStop = 1
            for idx, sermon_obj in enumerate(sermon_lst):
                # Check if duplicates are there, and if so put them in the remove list
                start = idx + 1
                for check in sermon_lst[start:]:
                    # compare all relevant elements
                    bEqual = True
                    for attr in check:
                        if attr != 'id':
                            if check[attr] != sermon_obj[attr]:
                                bEqual = False
                                break
                    if bEqual:
                        id = check['id']
                        if id not in remove_lst:
                            remove_lst.append(id)
                            lst_total.append("<tr><td>{}</td><td>{}</td></tr>".format(manu.id, check['id']))
                            added += 1
            # Remove duplicates for this manuscript
            if len(remove_lst) > 0:
                SermonDescr.objects.filter(id__in=remove_lst).delete()

        # Step #2: tidy up sermon fields
        qs_s = SermonDescr.objects.all().order_by('id')
        with transaction.atomic():
            for sermo in qs_s:
                bChange = False
                # Check if the sub title equals the title
                if sermo.subtitle != None and sermo.title != None and sermo.subtitle == sermo.title:
                    sermo.subtitle = ""
                    bChange = True
                # Save if needed
                if bChange:
                    sermo.save()

        # Step #3: init latin
        SermonDescr.init_latin()


        lst_total.append("</tbody></table>")

        # Create list to be returned
        result_list = []
        result_list.append({'part': 'Number of removed sermons', 'result': added})
        result_list.append({'part': 'All changes', 'result': "\n".join(lst_total)})

        context['result_list'] = result_list
    
        # Render and return the page
        return render(request, template_name, context)
    except:
        msg = oErr.get_error_message()
        oErr.DoError("do_sermons")
        return reverse('home')

def do_goldtogold(request):
    """Perform gold-to-gold relation repair -- NEW method that uses EqualGold"""

    oErr = ErrHandle()
    try:
        assert isinstance(request, HttpRequest)
        # Specify the template
        template_name = 'tools.html'
        # Define the initial context
        context =  {'title':'RU-passim-tools',
                    'year':get_current_datetime().year,
                    'pfx': APP_PREFIX,
                    'site_url': admin.site.site_url}
        context['is_app_uploader'] = user_is_ingroup(request, app_uploader)
        context['is_app_editor'] = user_is_ingroup(request, app_editor)

        # Only passim uploaders can do this
        if not context['is_app_uploader']: return reverse('home')

        # Indicate the necessary tools sub-part
        context['tools_part'] = "Repair gold-to-gold links"

        # Process this visit
        context['breadcrumbs'] = get_breadcrumbs(request, "GoldToGold", True)

        added = 0
        lst_total = []
        lst_total.append("<table><thead><tr><th>item</th><th>src</th><th>dst</th><th>linktype</th><th>addtype</th><th>Path</th></tr>")
        lst_total.append("<tbody>")

        method = "goldspread" 

        # Step #1: remove all unnecessary links
        oErr.Status("{} step #1".format(method))
        qs = EqualGoldLink.objects.all()
        lst_delete = []
        for relation in qs:
            if relation.src == relation.dst:
                lst_delete.append(relation.id)
        oErr.Status("Step 1: removing {} links".format(len(lst_delete)))
        if len(lst_delete) > 0:
            EqualGoldLink.objects.filter(Q(id__in=lst_delete)).delete()

        # Step #2: create groups of equals
        oErr.Status("{} step #2".format(method))
        lst_group = []      # List of groups, where each group is a list of equal-related gold-sermons
        qs_eqs = SermonGoldSame.objects.filter(linktype=LINK_EQUAL ).order_by('src')
        for idx, relation in enumerate(qs_eqs):
            src = relation.src
            dst = relation.dst
            # Find out in which group this one fits
            bGroup = False
            for group in lst_group:
                # Find a connection within this group
                for obj in group:
                    id = obj.id
                    if id == src.id or id == dst.id:
                        # We found the group
                        bGroup = True
                        break
                if bGroup:
                    # Add them if needed
                    if relation.src not in group: group.append(relation.src)
                    if relation.dst not in group: group.append(relation.dst)
                    # And then break from the larger one
                    break
            # Did we fit this into a group?
            if not bGroup:
                # Create a new group with two members
                group = [relation.src, relation.dst]
                lst_group.append(group)

        # Make sure the members of the groups get into the same EqualGold, if that is not the case yet
        # Step #2b: add equal groups
        for group in lst_group:
            if len(group) > 0:
                # Check the first item from the group
                first = group[0]
                if first.equal == None:
                    # Create a new equality group and add them all to it
                    eqg = EqualGold()
                    eqg.save()
                    with transaction.atomic():
                        for item in group:
                            item.equal = eqg
                            item.save()
                            added += 1

        # Step #3: add individual equals
        oErr.Status("{} step #3".format(method))
        # Visit all SermonGold instances and create eqg's for those that don't have one yet
        for gold in SermonGold.objects.all():
            if gold.equal == None:
                eqg = EqualGold()
                eqg.save()
                gold.equal = eqg
                gold.save()
                added += 1

        # -- or can a SermonGold be left without an equality group, if he is not equal to anything (yet)?

        # Step #4 (new): copy 'partial' and other near links to linke between groups
        oErr.Status("{} step #4a".format(method))
        for linktype in LINK_PRT:
            lst_prt_add = []    # List of partially equals relations to be added
            # Get all links of the indicated type
            qs_prt = EqualGoldLink.objects.filter(linktype=linktype).order_by('src__id')
            # Walk these links
            for obj_prt in qs_prt:
                # Get the equal groups of the link
                src_eqg = obj_prt.src
                dst_eqg = obj_prt.dst
                # Translate the link to one between equal-groups
                oLink = {'src': src_eqg, 'dst': dst_eqg}
                if oLink not in lst_prt_add: lst_prt_add.append(oLink)
                # And the reverse link
                oLink = {'src': dst_eqg, 'dst': src_eqg}
                if oLink not in lst_prt_add: lst_prt_add.append(oLink)
            # Add all the relations in lst_prt_add
            with transaction.atomic():
                for idx, item in enumerate(lst_prt_add):
                    # Make sure it doesn't yet exist
                    obj = EqualGoldLink.objects.filter(linktype=linktype, src=item['src'], dst=item['dst']).first()
                    if obj == None:
                        obj = EqualGoldLink(linktype=linktype, src=item['src'], dst=item['dst'])
                        obj.save()
                        added += 1
                        lst_total.append("<tr><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td></tr>".format( 
                            (idx+1), item['src'].id, item['dst'].id, linktype, "add", "" ))
        # y = [o for o in lst_prt_add if o['src'].id == 708]
        lst_total.append("</tbody></table>")

        # Create list to be returned
        result_list = []
        result_list.append({'part': 'Number of added relations', 'result': added})
        # result_list.append({'part': 'All additions', 'result': json.dumps(lst_total)})
        result_list.append({'part': 'All additions', 'result': "\n".join(lst_total)})

        context['result_list'] = result_list
    
        # Render and return the page
        return render(request, template_name, context)
    except:
        msg = oErr.get_error_message()
        oErr.DoError("goldtogold")
        return reverse('home')

def do_ssgmigrate(request):
    """Migration of super sermon gold"""

    oErr = ErrHandle()
    try:
        assert isinstance(request, HttpRequest)
        # Specify the template
        template_name = 'tools.html'
        # Define the initial context
        context =  {'title':'RU-passim-tools',
                    'year':get_current_datetime().year,
                    'pfx': APP_PREFIX,
                    'site_url': admin.site.site_url}
        context['is_app_uploader'] = user_is_ingroup(request, app_uploader)
        context['is_app_editor'] = user_is_ingroup(request, app_editor)

        # Only passim uploaders can do this
        if not context['is_app_uploader']: return reverse('home')

        # Indicate the necessary tools sub-part
        context['tools_part'] = "Migrate Authority file"

        added = 0
        lst_total = []

        # Walk all the EqualGold items
        qs = EqualGold.objects.all()
        number = qs.count()
        for idx, obj in enumerate(qs):
            print("Authority file item {} of {}".format(idx+1, number))
            # Need treatment?
            if not obj.author or not obj.number:
                # Check how many SG items there are within this EqualGold (=ssg)
                qs_sg = obj.equal_goldsermons.all()
                if qs_sg.count() == 1:
                    # There is just ONE (1) GS in the set
                    sg = qs_sg.first()
                    # (1) Get the author
                    author = sg.author
                    if author != None:
                        # There is an author!!
                        auth_num = author.get_number()
                        if auth_num > 0:
                            # Check the highest sermon number for this author
                            iNumber = EqualGold.sermon_number(author)
                            # Now we have both an author and a number...
                            obj.author = author
                            obj.number = iNumber
                            obj.code = EqualGold.passim_code(auth_num, iNumber)
                            obj.incipit = sg.incipit
                            obj.srchincipit = sg.srchincipit
                            obj.explicit = sg.explicit
                            obj.srchexplicit = sg.srchexplicit
                            obj.save()
                            added += 1
                            lst_total.append(obj.code)

                # Double check on author
                if obj.author == None:
                    # Set the author to 'undecided'
                    author = Author.get_undecided()
                    obj.author = author

            # Double check to see if something has changed
            if obj.code == "DETERMINE":
                obj.code = "ZZZ_DETERMINE"
                obj.save()

        # Create list to be returned
        result_list = []
        result_list.append({'part': 'Number of added Authority files', 'result': added})
        # result_list.append({'part': 'All additions', 'result': json.dumps(lst_total)})
        result_list.append({'part': 'All additions', 'result': "\n".join(lst_total)})

        context['result_list'] = result_list
    
        # Render and return the page
        return render(request, template_name, context)
    except:
        msg = oErr.get_error_message()
        oErr.DoError("ssgmigrate")
        return reverse('home')

def do_huwa(request):
    """Analyse Huwa SQLite database"""

    oErr = ErrHandle()
    bCreateExcel = True
    excel_output = "huwa_tables.xlsx"
    try:
        assert isinstance(request, HttpRequest)
        # Specify the template
        template_name = 'tools_huwa.html'
        # Define the initial context
        context =  {'title':    'RU-passim-tools',
                    'year':     get_current_datetime().year,
                    'pfx':      APP_PREFIX,
                    'site_url': admin.site.site_url}
        context['is_app_uploader'] = user_is_ingroup(request, app_uploader)
        context['is_app_editor'] = user_is_ingroup(request, app_editor)

        # Only passim uploaders can do this
        if not context['is_app_uploader']: return reverse('home')

        # Indicate the necessary tools sub-part
        context['tools_part'] = "Huwa Analysis"

        count_tbl = 0
        lst_total = []

        if bCreateExcel:
            # Set the excel output to the Media directory
            excel_output = os.path.abspath(os.path.join(WRITABLE_DIR, excel_output))

        # Connect to the Huwa database
        huwa_db = os.path.abspath(os.path.join(MEDIA_DIR, "passim", "huwa_database_for_PASSIM.db"))
        with sqlite3.connect(huwa_db) as db:
            table_info = {}
            standard_fields = ['erstdatum', 'aenddatum', 'aenderer', 'bemerkungen', 'ersteller']

            cursor = db.cursor()
            db_results = cursor.execute("SELECT name FROM sqlite_master WHERE type='table';").fetchall()
            tables = [x[0] for x in db_results]

            count_tbl = len(tables)

            # Walk all tables
            for table_name in tables:
                oInfo = {}
                # Find out what fields this table has
                db_results = cursor.execute("PRAGMA table_info('{}')".format(table_name)).fetchall()
                fields = []
                for field in db_results:
                    field_name = field[1]
                    fields.append(field_name)

                    field_info = dict(type=field[2],
                                      not_null=(field[3] == 1),
                                      default=field[4])
                    oInfo[field_name] = field_info
                oInfo['fields'] = fields
                oInfo['links'] = []

                # Read the table
                table_contents = cursor.execute("SELECT * FROM {}".format(table_name)).fetchall()
                oInfo['contents'] = table_contents

                table_info[table_name] = oInfo

            # Close the database again
            cursor.close()

        # Walk the tables again, looking for foreign keys
        for table_name in tables:
            oTable = table_info[table_name]
            # Check if any fields could be FKs
            for field in oTable['fields']:
                oInfo = oTable[field]
                bFK = False
                if field in tables and oInfo['type'] == "integer":
                    bFK = True
                # Mark that this is a FK to another table
                oInfo['FK'] = bFK
                if bFK:
                    # Add this to the list of "pointing to me" in the other table
                    table_info[field]['links'].append(table_name)

        # Create list to be returned
        result_list = []

        # If needed: start creating Excel
        if bCreateExcel:
            # Start workbook
            wb = openpyxl.Workbook()
            ws = wb.get_active_sheet()
            ws.title="Intro"
            c = ws.cell(row=1, column=1)
            c.value = "Automatically generated list of HUWA database tables"


        # Create a result to be shown
        table_num = 0
        for table_name in tables:
            table_num += 1
            oTable = table_info[table_name]
            field_reg = []
            field_fk = []
            links = []
            for field in oTable['fields']:
                oInfo = oTable[field]
                if oInfo['FK']:
                    field_fk.append("<span class='badge signature gr'>{}</span>".format(field))
                elif not field in standard_fields:
                    field_reg.append("<code>{}</code>".format(field))
            field_fk = ", ".join(field_fk)
            field_reg = ", ".join(field_reg)

            for link in oTable['links']:
                links.append("<span class='keyword'>{}</span>".format(link))
            links_to_me = ", ".join(links)

            lst_total = []
            lst_total.append("<table><thead><tr><th>Field type</th><th>Description</th></tr>")
            lst_total.append("<tbody>")
            # Regular fields
            lst_total.append("<tr><td valign='top' class='tdnowrap'>Regular:</td><td valign='top'>{}</td></tr>".format(field_reg))
            # FK fields
            lst_total.append("<tr><td valign='top' class='tdnowrap'>Foreign Key:</td><td valign='top'>{}</td></tr>".format(field_fk))
            # Other tables linking to me
            lst_total.append("<tr><td valign='top' class='tdnowrap'>Links to me:</td><td valign='top'>{}</td></tr>".format(links_to_me))
            lst_total.append("</tbody></table>")

            result_list.append({'part': table_name, 'result': "\n".join(lst_total)})

            # What about showing some Excel output?
            if bCreateExcel:
                # Debugging: show where we are
                oErr.Status("Doing table {} of {}: {}".format(table_num, count_tbl, table_name))

                # Add a Sheet for this Table
                ws = wb.create_sheet(table_name)
                row_num = 1

                # Set the column header names in bold
                for col_num in range(len(oTable['fields'])):
                    c = ws.cell(row=1, column=col_num+1)
                    c.value = oTable['fields'][col_num]
                    c.font = openpyxl.styles.Font(bold=True)
                    # Set width to a fixed size
                    ws.column_dimensions[get_column_letter(col_num+1)].width = 5.0   
                
                # Walk through the contents of this table
                for table_row in table_info[table_name]['contents']:
                    # Keep track of the EXCEL row we are in
                    row_num += 1
                    # Prepare the elements
                    table_str = []
                    for item in table_row:
                        if isinstance(item, int):
                            table_str.append(item)
                        elif isinstance(item, str):
                            table_str.append("'{}'".format(item.replace("'", '"')))
                        else:
                            table_str.append(item)
                    # WRite all the elements of one row (faster)
                    ws.append(table_str)
                      
                    
        if bCreateExcel:
            # Debugging: show where we are
            oErr.Status("Saving to: {}".format(excel_output))
            # Save the excel
            wb.save(excel_output)   

        context['result_list'] = result_list
    
        if bCreateExcel:
            # Debugging: show where we are
            oErr.Status("Returning render")
        # Render and return the page
        return render(request, template_name, context)
    except:
        if bCreateExcel:
            # Debugging: show where we are
            oErr.Status("Error being caught")
        msg = oErr.get_error_message()
        oErr.DoError("huwa")
        return reverse('home')

def get_old_edi(edi):
    """Split pages and year from edition, keep stripped edition and the pages
    The number of matches increase when the year in the name of the edition
    and in the short reference from Zotero is not used.
    """
 
    pages = None
    result = None
    status = "ok"
    # Split up edition and pages part
    arResult = edi.name.split(",")
    # In case of no pages, split string to get rid of year 
    if len(arResult) ==  1:
        result_tmp1 = arResult[0].split("(")
        result = result_tmp1[0].strip()
    # In case there are pages
    elif len(arResult) > 1: 
        # Split string to get rid of year 
        result_tmp1 = arResult[0].split("(")
        result = result_tmp1[0].strip()
        pages = arResult[1].strip()
        if not re.match(r'^[\d\-]+$', pages): 
            status = pages
            pages = None

    return result, pages, status

def get_short_edit(short):
    # Strip off the year of the short reference in Litref, keep only first part (abbr and seriesnumber)

    result = None
    arResult = short.split("(")
    if len(arResult) > 1:
        result = arResult[0].strip()
    elif len(arResult) == 1:
        result = arResult[0].strip()
    return result

def do_create_pdf_edi(request):
    """This definition creates the input for the pdf with all all used edition (full) references."""
        
    # Store title, and pageinfo (for at the bottom of the page) 
    Title = "Edition references used in PASSIM:"
    pageinfo = "Editions PASSIM"
             
    # Store name of the pdf file 
    filename = "Edi_ref_PASSIM.pdf"
            
    # Calculate the final qs for the edition litrefs
    ediref_ids = [x['reference'] for x in EdirefSG.objects.all().values('reference')]
       
    # Sort and filter all editions
    qs = Litref.objects.filter(id__in=ediref_ids).order_by('short')

    # Create a list of objects
    pdf_list = []
    for obj in qs:
        item = {}
        item = obj.full
        pdf_list.append(item)
    
    # Call create_pdf_passim function with arguments  
    response = create_pdf_passim(Title, pageinfo, filename, pdf_list)
  
    # And return the pdf
    return response

def do_create_pdf_lit(request):
    """"This definition creates the input for the pdf with all used literature (full) references."""
         
    # Store title, and pageinfo (for at the bottom of the page) 
    Title = "Literature references used in PASSIM:"
    pageinfo = "Literature PASSIM"
       
    # Store name of the file    
    filename = "Lit_ref_PASSIM.pdf"
    
    # Calculate the final qs for the manuscript litrefs
    litref_ids_man = [x['reference'] for x in LitrefMan.objects.all().values('reference')]
    
    # Calculate the final qs for the Gold sermon litrefs
    litref_ids_sg = [x['reference'] for x in LitrefSG.objects.all().values('reference')]

    # Combine the two qs into one and filter
    litref_ids = chain(litref_ids_man, litref_ids_sg)
    
    # Hier worden short en full opgehaald?
    qs = Litref.objects.filter(id__in=litref_ids).order_by('short') 
    
    # Create a list of objects 
    pdf_list = []
    for obj in qs:
        item = {}
        item = obj.full
        pdf_list.append(item)
        
    # Call create_pdf_passim function with arguments  
    response  = create_pdf_passim(Title, pageinfo, filename, pdf_list)
       
    # And return the pdf
    return response

def create_pdf_passim(Title, pageinfo, filename, pdf_list):
    """This definition creates a pdf for all passim requests."""
     
    # Define sizes of the pages in the pdf
    PAGE_HEIGHT=defaultPageSize[1]; PAGE_WIDTH=defaultPageSize[0]
    
    # Store text and current date for information on date of the download
    today = datetime.today()
    today.strftime('%Y-%m-%d')
       
    # Set buffer   
    buffer = io.BytesIO()

    # Set the first page
    def myFirstPage(canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica-Bold',22)
        canvas.drawCentredString(PAGE_WIDTH/2.0, PAGE_HEIGHT-80, Title)
        canvas.setFont('Helvetica',10)
        canvas.drawString(75,730, "Downloaded on: ")
        canvas.drawString(150,730, today.strftime('%d-%m-%Y'))
        canvas.setFont('Helvetica',9)
        canvas.drawString(inch, 0.75 * inch, "Page 1 / %s" % pageinfo)
        canvas.restoreState()
    
    # Set the second and later pages
    def myLaterPages(canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica',9)
        canvas.drawString(inch, 0.75 * inch, "Page %d %s" % (doc.page, pageinfo))
        canvas.restoreState()

    # Create the HttpResponse object with the appropriate PDF headers. 
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="{}"'.format(filename)
               
    # Define style of the pdf
    styles = getSampleStyleSheet()
        
    doc = SimpleDocTemplate(buffer)
    Story = [Spacer(1,1.05*inch)]
    style = styles["Normal"]
    # Dit tzt afhankelijk maken van lit en edi, niet van manuscript TH: hier nog met Erwin over hebben

    for line in pdf_list:
        line_2 = markdown(line)
        line_3 = line_2.replace("<em>", "<i>")
        line_4 = line_3.replace("</em>", "</i>")
        line_5 = markdown(line_4)
          
        lit_ref = (line_5) *1 
        p = Paragraph(lit_ref, style, '-')
        Story.append(p)
        Story.append(Spacer(1,0.2*inch))
    doc.build(Story, onFirstPage=myFirstPage, onLaterPages=myLaterPages)
    
    # Write the buffer to the PDF
    response.write(buffer.getvalue())
    # Close the buffer cleanly, and we're done.
    buffer.close()
       
    # And return the pdf
    return response

def do_create_pdf_manu(request):
    """This definition creates the input for the pdf with all manuscripts in the passim database."""
  
    # Store title, and pageinfo (for at the bottom of each page) 
    Title = "Manuscripts listed in PASSIM:"
    pageinfo = "Manuscripts PASSIM"
             
    # Store name of the pdf file 
    filename = "Manu_list_PASSIM.pdf"

    # Which method needs to be used
    method = "msitem"   # "sermondescr"

    # Calculate the qs for the manuscripts       
    qs = Manuscript.objects.all()
    
    # Create temporary list and add relevant fields to the list
    pdf_list_temp = []
    for obj in qs:
        # Count all (SermonDescr) items for each manuscript
        count = obj.get_sermon_count()
        
        # Handle empty origin fields
        origin = None if obj.origin == None else obj.origin.name
        
        # Retrieve all provenances for each manuscript
        qs_prov = obj.manuscripts_provenances.all()
        # Issue #289: innovation below turned back to the above
        # qs_prov = obj.manuprovenances.all()
        
        # Iterate through the queryset, place name and location 
        # for each provenance together 
        prov_texts = []
        for prov in qs_prov:
            prov = prov.provenance
            # Issue #289: innovation turned back to the above
            prov_text = prov.name
            if prov.location:
                prov_text = "{} ({})".format(prov_text, prov.location.name)
            prov_texts.append(prov_text)
        # Join all provenances together
        provenance = ", ".join(prov_texts)
        
        # Store all relevant items in a dictionary and append to the list TH: let op, link lib and city
        name = "" if obj.name == None else obj.name         
        idno = "" if obj.idno == None else obj.idno
        yearstart = "" if obj.yearstart == None else obj.yearstart
        yearfinish = "" if obj.yearfinish == None else obj.yearfinish
        libname = "" if obj.library == None else obj.library.name
        city = "" if obj.library == None or obj.library.city == None else obj.library.city.name
        
        item = dict(name=name, idno=idno, yearstart=yearstart, yearfinish=yearfinish,
                stype=obj.get_stype_display(), libname=libname, city=city, origin=origin, provenance=provenance,
                count=count)
        pdf_list_temp.append(item)
             
    # Sort the list on city and yearstart
    pdf_list_sorted = sorted(pdf_list_temp, key=itemgetter('city', 'yearstart')) 
        
    # Create five strings and add to pdf_list (to be processed in create_pdf_passim)
    pdf_list=[]
    for dict_item in pdf_list_sorted:
       
       # The first string contains the name of the city, library and id code of the manuscript     
       string_1 = dict_item['city'] + ", "+ dict_item['libname'] + ", "+ dict_item['idno']
       
       # The second string contains the start and end year of the manuscript and the number of items of the manuscript
       string_2 = 'Date: ' + str(dict_item['yearstart']) + "-" + str(dict_item['yearfinish']) + ', items: ' + str(dict_item['count'])
       
       # The third string contains the status of the manuscript: manual, imported or edited
       string_3 = 'Status: ' + dict_item['stype']
       
       # The fourth string contains the origin of the manuscript
       if dict_item['origin'] == None:
          origin = ""
       else:
          origin = dict_item['origin']
              
       string_4 = 'Origin: ' + origin
        
       # The fifth string contains the provenances of the manuscript
       if dict_item['provenance'] == None:
          provenance = ""
       else:
          provenance = dict_item['provenance']

       string_5 = 'Provenances: '+ provenance
       
       # The strings are combined into one with markddown line breaks so that each string is placed on a new line in the pdf
       combined = string_1 + "<br />" + string_2 + "<br />" + string_3 + "<br />" + string_4 + "<br />" + string_5
       
       # The new combined strings are placed in a new list, to be used in the create_pdf_passim function.        
       pdf_list.append(combined)

    # Call create_pdf_passim function with arguments  
    response = create_pdf_passim(Title, pageinfo, filename, pdf_list)
  
    # And return the pdf
    return response

def user_is_in_team(request):
    bResult = False
    username = request.user.username
    team_group = app_editor
    # Validate
    if username and team_group and username != "" and team_group != "":
        # First filter on owner
        owner = Profile.get_user_profile(username)
        # Now check for permissions
        bResult = (owner.user.groups.filter(name=team_group).first() != None)
    return bResult

def get_userkeyword_interface(request, context):
    """Get HTML stuff providing a user-keyword edit interface for non-editing users"""

    template_name = "seeker/userkeyword_select.html"
    sBack = ""
    oErr = ErrHandle()
    try:

        pass
    except:
        msg = oErr.get_error_message()
        oErr.DoError("get_userkeyword_interface")
    return sBack





# ============== NOTE: superseded by the READER app ===================
def import_ead(request):
    """Import one or more XML files that each contain one or more EAD items from Archives Et Manuscripts"""

    # HIER DUS CODE TOEVOEGEN
    # TH: dit werkt dus van uit de browser denk ik
    # is wel anders want geen losse bestanden zoals bij de zwitserse site
    # Moet ik dit niet apart testen? Wat hieronder staat is nogal...weinig
    # ook bij Manuscript.codex kijken en andee delen genoemd in mail van Erwin read_ecodex in models.py
    # hoe zit dit samen?
    
    # Initialisations
    # NOTE: do ***not*** add a breakpoint until *AFTER* form.is_valid
    arErr = []
    error_list = []
    transactions = []
    data = {'status': 'ok', 'html': ''}
    template_name = 'seeker/import_manuscripts.html' # Adapt because of multiple descriptions in 1 xml?
    obj = None
    data_file = ""
    bClean = False
    username = request.user.username

    # Check if the user is authenticated and if it is POST
    if request.user.is_authenticated and request.method == 'POST' and user_is_ingroup(request, app_uploader):
    
        # Remove previous status object for this user
        Status.objects.filter(user=username).delete()
        
        # Create a status object # TH: type goed aangepast?
        oStatus = Status(user=username, type="ead", status="preparing")
        oStatus.save()

        form = UploadFilesForm(request.POST, request.FILES)
        lResults = []
        if form.is_valid():
            # NOTE: from here a breakpoint may be inserted!
            print('import_ead: valid form') # TH: import_ aangepast import ead_am?

            # Create a SourceInfo object for this extraction
            source = SourceInfo(url="https://ccfr.bnf.fr/", collector=username) # TH: aanpassen, klopt niet, ccfr
            source.save()
            file_list = []

            # Get the contents of the imported file
            files = request.FILES.getlist('files_field')
            if files != None:
                for data_file in files:
                    filename = data_file.name
                    file_list.append(filename)

                    # Set the status
                    oStatus.set("reading", msg="file={}".format(filename))

                    # Get the source file
                    if data_file == None or data_file == "":
                        arErr.append("No source file specified for the selected project")
                    else:
                        # Check the extension
                        arFile = filename.split(".")
                        extension = arFile[len(arFile)-1]

                        # Further processing depends on the extension
                        oResult = None
                        if extension == "xml":
                            # This is an XML file
                            oResult = Manuscript.read_ecodex(username, data_file, filename, arErr, source=source) # TH:aanpassen , models.py

                        # Determine a status code
                        statuscode = "error" if oResult == None or oResult['status'] == "error" else "completed"
                        if oResult == None:
                            arErr.append("There was an error. No manuscripts have been added")
                        else:
                            lResults.append(oResult)
            # Adapt the 'source' to tell what we did
            source.code = "Imported using the [import_ead??] function on these XML files: {}".format(", ".join(file_list)) # TH: aanpassen
            source.save()
            # Indicate we are ready
            oStatus.set("ready")
            # Get a list of errors
            error_list = [str(item) for item in arErr]

            # Create the context
            context = dict(
                statuscode=statuscode,
                results=lResults,
                error_list=error_list
                )

            if len(arErr) == 0:
                # Get the HTML response
                data['html'] = render_to_string(template_name, context, request)
            else:
                data['html'] = "Please log in before continuing"


        else:
            data['html'] = 'invalid form: {}'.format(form.errors)
            data['status'] = "error"
    else:
        data['html'] = 'Only use POST and make sure you are logged in and authorized for uploading'
        data['status'] = "error"
 
    # Return the information
    return JsonResponse(data)

    pass

def import_ecodex(request):
    """Import one or more XML files that each contain one manuscript definition from e-codices, from Switzerland"""

    # Initialisations
    # NOTE: do ***not*** add a breakpoint until *AFTER* form.is_valid
    arErr = []
    error_list = []
    transactions = []
    data = {'status': 'ok', 'html': ''}
    template_name = 'seeker/import_manuscripts.html'
    obj = None
    data_file = ""
    bClean = False
    username = request.user.username

    # Check if the user is authenticated and if it is POST
    if request.user.is_authenticated and request.method == 'POST' and user_is_ingroup(request, app_uploader):

        # Remove previous status object for this user
        Status.objects.filter(user=username).delete()
        # Create a status object
        oStatus = Status(user=username, type="ecodex", status="preparing")
        oStatus.save()

        def add_manu(lst_manual, lst_read, status="", msg="", user="", name="", url="", yearstart="", yearfinish="",
                     library="", idno="", filename=""):
            oInfo = {}
            oInfo['status'] = status
            oInfo['msg'] = msg
            oInfo['user'] = user
            oInfo['name'] = name
            oInfo['url'] = url
            oInfo['yearstart'] = yearstart
            oInfo['yearfinish'] = yearfinish
            oInfo['library'] = library
            oInfo['idno'] = idno
            oInfo['filename'] = filename
            if status == "error":
                lst_manual.append(oInfo)
            else:
                lst_read.append(oInfo)
            return True

        form = UploadFilesForm(request.POST, request.FILES)
        lResults = []
        if form.is_valid():
            # NOTE: from here a breakpoint may be inserted!
            print('import_ecodex: valid form')
            oErr = ErrHandle()
            try:
                # The list of headers to be shown
                lHeader = ['status', 'msg', 'name', 'yearstart', 'yearfinish', 'library', 'idno', 'filename', 'url']

                # Create a SourceInfo object for this extraction
                source = SourceInfo(url="http://e-codices.unifr.ch", collector=username)
                source.save()
                file_list = []

                # Get the contents of the imported file
                files = request.FILES.getlist('files_field')
                if files != None:
                    for data_file in files:
                        filename = data_file.name
                        file_list.append(filename)

                        # Set the status
                        oStatus.set("reading", msg="file={}".format(filename))

                        # Get the source file
                        if data_file == None or data_file == "":
                            arErr.append("No source file specified for the selected project")
                        else:
                            # Check the extension
                            arFile = filename.split(".")
                            extension = arFile[len(arFile)-1]

                            lst_manual = []
                            lst_read = []

                            # Further processing depends on the extension
                            oResult = None
                            if extension == "xml":
                                # This is an XML file
                                oResult = Manuscript.read_ecodex(username, data_file, filename, arErr, source=source)

                                if oResult == None or oResult['status'] == "error":
                                    # Process results
                                    add_manu(lst_manual, lst_read, status=oResult['status'], msg=oResult['msg'], user=oResult['user'],
                                                 filename=oResult['filename'])
                                else:
                                    # Get the results from oResult
                                    obj = oResult['obj']
                                    # Process results
                                    add_manu(lst_manual, lst_read, status=oResult['status'], user=oResult['user'],
                                                 name=oResult['name'], yearstart=obj.yearstart,
                                                 yearfinish=obj.yearfinish,library=obj.library.name,
                                                 idno=obj.idno,filename=oResult['filename'])

                            elif extension == "txt":
                                # Set the status
                                oStatus.set("reading list", msg="file={}".format(filename))
                                # (1) Read the TXT file
                                lines = []
                                bFirst = True
                                for line in data_file:
                                    # Get a good view of the line
                                    sLine = line.decode("utf-8").strip()
                                    if bFirst:
                                        if "\ufeff" in sLine:
                                            sLine = sLine.replace("\ufeff", "")
                                        bFirst = False
                                    lines.append(sLine)
                                # (2) Walk through the list of XML file names
                                for idx, xml_url in enumerate(lines):
                                    xml_url = xml_url.strip()
                                    if xml_url != "" and ".xml" in xml_url:
                                        # Set the status
                                        oStatus.set("reading XML", msg="{}: file={}".format(idx, xml_url))
                                        # (3) Download the file from the internet and save it 
                                        bOkay, sResult = download_file(xml_url)
                                        if bOkay:
                                            # We have the filename
                                            xml_file = sResult
                                            name = xml_url.split("/")[-1]
                                            # (4) Read the e-codex manuscript
                                            oResult = Manuscript.read_ecodex(username, xml_file, name, arErr, source=source)
                                            # (5) Check before continuing
                                            if oResult == None or oResult['status'] == "error":
                                                msg = "unknown"  
                                                if 'msg' in oResult: 
                                                    msg = oResult['msg']
                                                elif 'status' in oResult:
                                                    msg = oResult['status']
                                                arErr.append("Import-ecodex: file {} has not been loaded ({})".format(xml_url, msg))
                                                # Process results
                                                add_manu(lst_manual, lst_read, status="error", msg=msg, user=oResult['user'],
                                                             filename=oResult['filename'])
                                            else:
                                                # Get the results from oResult
                                                obj = oResult['obj']
                                                # Process results
                                                add_manu(lst_manual, lst_read, status=oResult['status'], user=oResult['user'],
                                                             name=oResult['name'], yearstart=obj.yearstart,
                                                             yearfinish=obj.yearfinish,library=obj.library.name,
                                                             idno=obj.idno,filename=oResult['filename'])

                                        else:
                                            aErr.append("Import-ecodex: failed to download file {}".format(xml_url))

                            # Create a report and add it to what we return
                            oContents = {'headers': lHeader, 'list': lst_manual, 'read': lst_read}
                            oReport = Report.make(username, "iecod", json.dumps(oContents))
                                
                            # Determine a status code
                            statuscode = "error" if oResult == None or oResult['status'] == "error" else "completed"
                            if oResult == None:
                                arErr.append("There was an error. No manuscripts have been added")
                            else:
                                lResults.append(oResult)

                # Adapt the 'source' to tell what we did TH: waar staat import_ecodex?
                source.code = "Imported using the [import_ecodex] function on these XML files: {}".format(", ".join(file_list))
                source.save()
                # Indicate we are ready
                oStatus.set("ready")
                # Get a list of errors
                error_list = [str(item) for item in arErr]

                # Create the context
                context = dict(
                    statuscode=statuscode,
                    results=lResults,
                    error_list=error_list
                    )

                if len(arErr) == 0:
                    # Get the HTML response
                    data['html'] = render_to_string(template_name, context, request)
                else:
                    data['html'] = "Please log in before continuing"
            except:
                msg = oErr.get_error_message()
                oErr.DoError("import_ecodex")
                data['html'] = msg
                data['status'] = "error"

        else:
            data['html'] = 'invalid form: {}'.format(form.errors)
            data['status'] = "error"
    else:
        data['html'] = 'Only use POST and make sure you are logged in and authorized for uploading'
        data['status'] = "error"
 
    # Return the information
    return JsonResponse(data)

def search_ecodex(request):
    arErr = []
    error_list = []
    data = {'status': 'ok', 'html': ''}
    username = request.user.username
    errHandle = ErrHandle()

    # Check if the user is authenticated and if it is POST
    if request.user.is_authenticated and request.method == 'POST' and user_is_ingroup(request, app_editor):
        # Create a regular expression
        r_href = re.compile(r'(href=[\'"]?)([^\'" >]+)')
        # Get the parameters
        if request.POST:
            qd = request.POST
        else:
            qd = request.GET
        if 'search_url' in qd:
            # Get the parameter
            s_url = qd['search_url']
            # Make a search request and get the result into a string
            try:
                r = requests.get(s_url)
            except:
                sMsg = errHandle.get_error_message()
                errHandle.DoError("Request problem")
                return False
            if r.status_code == 200:
                # Get the text into a list
                l_text = r.text.split('\n')
                lHtml = []
                iMatches = 0
                lHtml.append("<table><thead><tr><th>#</th><th>Part</th><th>Text</th><th>Version</th><th>XML</th></tr></thead>")
                lHtml.append("<tbody>")
                # Walk the text
                for line in l_text:
                    # Check if this line contains information
                    if ">Description<" in line:
                        # Extract the information from this line
                        match = r_href.search(line)
                        if match:
                            iMatches += 1
                            # Get the HREF 
                            href = match.group(2)
                            # Split into parts
                            parts = href.split("/")
                            # Find the 'description' part
                            for idx,part in enumerate(parts):
                                if part == "description":
                                    # Take array from one further
                                    idx += 1
                                    mparts = parts[idx:]
                                    break
                            # Reconstruct:
                            xml_name = "{}-{}".format(mparts[0], mparts[1])
                            version = ""
                            if len(mparts) > 2 and mparts[2] != "":
                                xml_name = "{}_{}".format(xml_name, mparts[2])
                                version = mparts[2]
                            xml_url = "{}/{}.xml".format("https://www.e-codices.unifr.ch/xml/tei_published", xml_name)

                            # Add to the HTML table
                            lHtml.append("<tr><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td></tr>".format(iMatches,mparts[0], mparts[1],version,xml_url))
                # We had all lines, now we need to return it in a good way
                lHtml.append("</tbody></table>")
                data['html'] = "\n".join(lHtml)
                data['status'] = "ok"
            else:
                data['html'] = "Website returns status code {}".format(r.status_code)
                data['status'] = "error"
        else:
            data['html'] = "Did not receive parameter search_url"
            data['status'] = "error"
    else:
        data['html'] = 'Only use POST and make sure you are logged in and authorized for uploading'
        data['status'] = "error"
 
    # Return the information
    return JsonResponse(data)

# ============== NOTE: end of READER app superseded material ==========

def import_gold(request):
    """Import one or more Excel files that each contain one Gold-Sermon definition in Excel"""

    # Initialisations
    # NOTE: do ***not*** add a breakpoint until *AFTER* form.is_valid
    arErr = []
    error_list = []
    transactions = []
    data = {'status': 'ok', 'html': ''}
    template_name = 'seeker/import_gold.html'
    obj = None
    data_file = ""
    bClean = False
    username = request.user.username

    if request.user.is_authenticated and request.method == 'POST' and user_is_ingroup(request, app_uploader):

        # Remove previous status object for this user
        Status.objects.filter(user=username).delete()
        # Create a status object
        oStatus = Status(user=username, type="gold", status="preparing")
        oStatus.save()


        form = UploadFilesForm(request.POST, request.FILES)
        lResults = []
        if form.is_valid():
            # NOTE: from here a breakpoint may be inserted!
            print('import_gold: valid form')

            # Initialise a list of files to be processed
            file_list = []

            # Get the contents of the imported file
            files = request.FILES.getlist('files_field')
            if files != None:
                for data_file in files:
                    filename = data_file.name
                    file_list.append(filename)

                    # Set the status
                    oStatus.set("reading", msg="file={}".format(filename))

                    # Get the source file
                    if data_file == None or data_file == "":
                        arErr.append("No source file specified for the selected project")
                    else:
                        # Check the extension
                        arFile = filename.split(".")
                        extension = arFile[len(arFile)-1]

                        # Further processing depends on the extension
                        oResult = None
                        if extension == "xlsx":
                            # This is an Excel XLSX file
                            oResult = SermonGold.read_gold(username, data_file, filename, arErr, oStatus)
                            # Note: the [oResult] also contains a 'report' part

                        # Determine a status code
                        statuscode = "error" if oResult == None or oResult['status'] == "error" else "completed"
                        if oResult == None:
                            arErr.append("There was an error. No golden sermons have been added")
                        elif statuscode == "error" and 'msg' in oResult:
                            arErr.append(oResult['msg'])
                        else:
                            lResults.append(oResult)

            # Indicate we are ready
            oStatus.set("ready")
            # Get a list of errors
            error_list = [str(item) for item in arErr]

            # Create the context
            context = dict(
                statuscode=statuscode,
                results=lResults,
                error_list=error_list
                )

            if statuscode == "error" and len(arErr) == 0:
                data['html'] = "Please log in before continuing"
            else:
                # Get the HTML response
                data['html'] = render_to_string(template_name, context, request)


        else:
            data['html'] = 'invalid form: {}'.format(form.errors)
            data['status'] = "error"
    else:
        data['html'] = 'Only use POST and make sure you are logged in and authorized for uploading'
        data['status'] = "error"

    # Return the information
    return JsonResponse(data)

def get_cnrs_manuscripts(city, library):
    """Get the manuscripts held in the library"""

    oErr = ErrHandle()
    sBack = ""
    try:
        # Get the code of the city
        obj = City.objects.filter(name__iexact=city.name).first()
        if obj != None:
            # Get the code of the city
            idVille = obj.idVilleEtab
            # Build the query
            url = "{}/Manuscrits/manuscritforetablissement".format(cnrs_url)
            data = {"idEtab": library, "idVille": idVille}
            try:
                r = requests.post(url, data=data)
            except:
                sMsg = errHandle.get_error_message()
                errHandle.DoError("Request problem")
                return "Request problem: {}".format(sMsg)
            # Decypher the response
            if r.status_code == 200:
                # Return positively
                # reply = demjson.decode(r.text.replace("\t", " "))
                sText = r.text.replace("\t", "")
                # Take extra care that the interpretation doesn't go wrong:
                try:
                    reply = json.loads(sText)
                    if reply != None and "items" in reply:
                        results = []
                        for item in reply['items']:
                            if item['name'] != "":
                                results.append(item['name'])
                        #data = json.dumps(results)
                        # Interpret the results
                        lst_manu = []
                        for item in results:
                            lst_manu.append("<span class='manuscript'>{}</span>".format(item))
                        sBack = "\n".join(lst_manu)
                except:
                    # THis is not proper JSON, so cannot be decoded
                    oErr.Status("Improper JSON code from cnrs_manuscripts: {}".format(sText))
                    sBack = ""
    except:
        msg = oErr.get_error_message()
        sBack = "Error: {}".format(msg)
        oErr.DoError("get_cnrs_manuscripts")
    return sBack

def get_manuscripts(request):
    """Get a list of manuscripts"""

    data = 'fail'
    errHandle = ErrHandle()
    # Only allow AJAX calls with POST
    if is_ajax(request) and request.method == "POST":
        get = request.POST
        # Get parameters city and library
        city = get.get("city", "")
        lib = get.get("library", "")

        url = "{}/Manuscrits/manuscritforetablissement".format(cnrs_url)
        data = "idEtab={}&idVille={}".format(lib, city)
        data = {"idEtab": lib, "idVille": city}
        #data = []
        #data.append({'name': 'idEtab', 'value': lib})
        #data.append({'name': 'idVille', 'value': city})
        try:
            r = requests.post(url, data=data)
        except:
            sMsg = errHandle.get_error_message()
            errHandle.DoError("Request problem")
            return False
        if r.status_code == 200:
            # Return positively
            # reply = demjson.decode(r.text.replace("\t", " "))
            sText = r.text.replace("\t", "")
            reply = json.loads(sText)
            if reply != None and "items" in reply:
                results = []
                for item in reply['items']:
                    if item['name'] != "":
                        results.append(item['name'])
                data = json.dumps(results)
    else:
        data = "Request is not ajax"
    # Prepare and return data
    mimetype = "application/json"
    return HttpResponse(data, mimetype)

def download_file(url):
    """Download a file from the indicated URL"""

    bResult = True
    sResult = ""
    errHandle = ErrHandle()
    # Get the filename from the url
    name = url.split("/")[-1]
    # Set the output directory
    outdir = os.path.abspath(os.path.join(MEDIA_DIR, "e-codices"))
    if not os.path.exists(outdir):
        os.mkdir(outdir)
    # Create a filename where we can store it
    filename = os.path.abspath(os.path.join(outdir, name))
    try:
        r = requests.get(url)
    except:
        sMsg = errHandle.get_error_message()
        errHandle.DoError("Request problem")
        return False, sMsg
    if r.status_code == 200:
        # Read the response
        sText = r.text
        # Write away
        with open(filename, "w", encoding="utf-8") as f:
            f.write(sText)
        sResult = filename
    else:
        bResult = False
        sResult = "download_file received status {} for {}".format(r.status_code, url)
    # Return the result
    return bResult, sResult

@csrf_exempt
def get_countries(request):
    """Get a list of countries for autocomplete"""

    data = 'fail'
    method = "useLocation"
    if is_ajax(request):
        oErr = ErrHandle()
        try:
            sName = request.GET.get('country', '')
            if sName == "": sName = request.GET.get('country_ta', "")
            lstQ = []
            lstQ.append(Q(name__icontains=sName))
            if method == "useLocation":
                loctype = LocationType.find("country")
                lstQ.append(Q(loctype=loctype))
                countries = Location.objects.filter(*lstQ).order_by('name')
            else:
                countries = Country.objects.filter(*lstQ).order_by('name')
            results = []
            for co in countries:
                co_json = {'name': co.name, 'id': co.id }
                results.append(co_json)
            data = json.dumps(results)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("get_countries")
    else:
        data = "Request is not ajax"
    mimetype = "application/json"
    return HttpResponse(data, mimetype)

@csrf_exempt
def get_cities(request):
    """Get a list of cities for autocomplete"""

    data = 'fail'
    method = "useLocation"
    if is_ajax(request):
        oErr = ErrHandle()
        try:
            # Get the user-specified 'country' and 'city' strings
            country = request.GET.get('country', "")
            if country == "": country = request.GET.get('country_ta', "")
            city = request.GET.get("city", "")
            if city == "": city = request.GET.get('city_ta', "")

            # build the query
            lstQ = []
            if method == "useLocation":
                # Start as broad as possible: country
                qs_loc = None
                if country != "":
                    loctype_country = LocationType.find("country")
                    lstQ.append(Q(name=country))
                    lstQ.append(Q(loctype=loctype_country))
                    qs_country = Location.objects.filter(*lstQ)
                    # Fine-tune on city...
                    loctype_city = LocationType.find("city")
                    lstQ = []
                    lstQ.append(Q(name__icontains=city))
                    lstQ.append(Q(loctype=loctype_city))
                    lstQ.append(Q(relations_location__in=qs_country))
                    cities = Location.objects.filter(*lstQ)
                else:
                    loctype_city = LocationType.find("city")
                    lstQ.append(Q(name__icontains=city))
                    lstQ.append(Q(loctype=loctype_city))
                    cities = Location.objects.filter(*lstQ)
            elif method == "slowLocation":
                # First of all: city...
                loctype_city = LocationType.find("city")
                lstQ.append(Q(name__icontains=city))
                lstQ.append(Q(loctype=loctype_city))
                # Do we have a *country* specification?
                if country != "":
                    loctype_country = LocationType.find("country")
                    lstQ.append(Q(relations_location__name=country))
                    lstQ.append(Q(relations_location__loctype=loctype_country))
                # Combine everything
                cities = Location.objects.filter(*lstQ).order_by('name')
            else:
                if country != "":
                    lstQ.append(Q(country__name__icontains=country))
                lstQ.append(Q(name__icontains=city))
                cities = City.objects.filter(*lstQ).order_by('name')
            results = []
            for co in cities:
                co_json = {'name': co.name, 'id': co.id }
                results.append(co_json)
            data = json.dumps(results)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("get_cities")
    else:
        data = "Request is not ajax"
    mimetype = "application/json"
    return HttpResponse(data, mimetype)
    
@csrf_exempt
def get_libraries(request):
    """Get a list of libraries for autocomplete"""

    data = 'fail'
    if is_ajax(request):
        oErr = ErrHandle()
        try:
            # Get the user-specified 'country' and 'city' strings
            country = request.GET.get('country', "")
            if country == "": country = request.GET.get('country_ta', "")
            city = request.GET.get("city", "")
            if city == "": city = request.GET.get('city_ta', "")
            lib = request.GET.get("library", "")
            if lib == "": lib = request.GET.get('libname_ta', "")

            # build the query
            lstQ = []
            # Start as broad as possible: country
            qs_loc = None
            if country != "":
                loctype_country = LocationType.find("country")
                lstQ.append(Q(name=country))
                lstQ.append(Q(loctype=loctype_country))
                qs_country = Location.objects.filter(*lstQ)
                # What about city?
                if city == "":
                    qs_loc = qs_country
                else:
                    loctype_city = LocationType.find("city")
                    lstQ = []
                    lstQ.append(Q(name__icontains=city))
                    lstQ.append(Q(loctype=loctype_city))
                    lstQ.append(Q(relations_location__in=qs_country))
                    qs_loc = Location.objects.filter(*lstQ)
            elif city != "":
                loctype_city = LocationType.find("city")
                lstQ.append(Q(name__icontains=city))
                lstQ.append(Q(loctype=loctype_city))
                qs_loc = Location.objects.filter(*lstQ)

            # Start out with the idea to look for a library by name:
            lstQ = []
            if lib != "": lstQ.append(Q(name__icontains=lib))
            if qs_loc != None: lstQ.append(Q(location__in=qs_loc))

            # Combine everything
            libraries = Library.objects.filter(*lstQ).order_by('name').values('name','id') 
            results = []
            for co in libraries:
                co_json = {'name': co['name'], 'id': co['id'] }
                results.append(co_json)
            data = json.dumps(results)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("get_libraries")
    else:
        data = "Request is not ajax"
    mimetype = "application/json"
    return HttpResponse(data, mimetype)

@csrf_exempt
def get_origins(request):
    """Get a list of origin names for autocomplete"""

    data = 'fail'
    if is_ajax(request):
        sName = request.GET.get('name', '')
        lstQ = []
        lstQ.append(Q(name__icontains=sName))
        origins = Origin.objects.filter(*lstQ).order_by('name')
        results = []
        for co in origins:
            co_json = {'name': co.name, 'id': co.id }
            results.append(co_json)
        data = json.dumps(results)
    else:
        data = "Request is not ajax"
    mimetype = "application/json"
    return HttpResponse(data, mimetype)

@csrf_exempt
def get_locations(request):
    """Get a list of location names for autocomplete"""

    data = 'fail'
    if is_ajax(request):
        oErr = ErrHandle()
        try:
            sName = request.GET.get('name', '')
            lstQ = []
            lstQ.append(Q(name__icontains=sName))
            locations = Location.objects.filter(*lstQ).order_by('name').values('name', 'loctype__name', 'id')
            results = []
            for co in locations:
                # name = "{} ({})".format(co['name'], co['loctype__name'])
                name = co['name']
                co_json = {'name': name, 'id': co['id'], 'loctype': co['loctype__name'] }
                results.append(co_json)
            data = json.dumps(results)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("get_locations")
    else:
        data = "Request is not ajax"
    mimetype = "application/json"
    return HttpResponse(data, mimetype)

@csrf_exempt
def get_litref(request):
    """Get ONE particular short representation of a litref"""
    
    data = 'fail'
    if is_ajax(request):
        oErr = ErrHandle()
        try:
            sId = request.GET.get('id', '')
            co_json = {'id': sId}
            lstQ = []
            lstQ.append(Q(id=sId))
            litref = Litref.objects.filter(Q(id=sId)).first()
            if litref:
                short = litref.get_short()
                co_json['name'] = short
            data = json.dumps(co_json)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("get_litref")
    else:
        data = "Request is not ajax"
    mimetype = "application/json"
    return HttpResponse(data, mimetype)

@csrf_exempt
def get_litrefs(request):
    """Get a list of literature references for autocomplete"""
    
    data = 'fail'
    if is_ajax(request):
        oErr = ErrHandle()
        try:
            sName = request.GET.get('name', '')
            lstQ = []
            lstQ.append(Q(full__icontains=sName)|Q(short__icontains=sName))
            litrefs = Litref.objects.filter(*lstQ).order_by('short').values('full', 'short', 'id')
            results = [] 
            for co in litrefs:
                name = "{} {}".format(co['full'], co['short'])
                co_json = {'name': name, 'id': co['id'] }
                results.append(co_json)
            data = json.dumps(results)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("get_litrefs")
    else:
        data = "Request is not ajax"
    mimetype = "application/json"
    return HttpResponse(data, mimetype)

@csrf_exempt
def get_sg(request):
    """Get ONE particular short representation of a SG"""
    
    data = 'fail'
    if is_ajax(request):
        oErr = ErrHandle()
        try:
            sId = request.GET.get('id', '')
            co_json = {'id': sId}
            lstQ = []
            lstQ.append(Q(id=sId))
            sg = SermonGold.objects.filter(Q(id=sId)).first()
            if sg:
                short = sg.get_label()
                co_json['name'] = short
            data = json.dumps(co_json)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("get_sg")
    else:
        data = "Request is not ajax"
    mimetype = "application/json"
    return HttpResponse(data, mimetype)

@csrf_exempt
def get_sglink(request):
    """Get ONE particular short representation of a *link* to a SG"""
    
    data = 'fail'
    if is_ajax(request):
        oErr = ErrHandle()
        try:
            sId = request.GET.get('id', '')
            co_json = {'id': sId}
            lstQ = []
            lstQ.append(Q(id=sId))
            sg = SermonDescrGold.objects.filter(Q(id=sId)).first()
            if sg:
                short = sg.get_label()
                co_json['name'] = short
            data = json.dumps(co_json)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("get_sglink")
    else:
        data = "Request is not ajax"
    mimetype = "application/json"
    return HttpResponse(data, mimetype)

@csrf_exempt
def get_ssglink(request):
    """Get ONE particular short representation of a *link* from sermondescr to a SSG"""
    
    data = 'fail'
    if is_ajax(request):
        oErr = ErrHandle()
        try:
            sId = request.GET.get('id', '')
            co_json = {'id': sId}
            lstQ = []
            lstQ.append(Q(id=sId))
            ssg = SermonDescrEqual.objects.filter(Q(id=sId)).first()
            if ssg:
                short = ssg.get_label()
                co_json['name'] = short
            data = json.dumps(co_json)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("get_ssglink")
    else:
        data = "Request is not ajax"
    mimetype = "application/json"
    return HttpResponse(data, mimetype)

@csrf_exempt
def get_ssg2ssg(request):
    """Get ONE particular short representation of a *link* from SSG to a SSG"""
    
    data = 'fail'
    if is_ajax(request):
        oErr = ErrHandle()
        try:
            sId = request.GET.get('id', '')
            co_json = {'id': sId}
            # oErr.Status("get_ssg2ssg id={}".format(sId))
            lstQ = []
            lstQ.append(Q(id=sId))
            ssg = EqualGoldLink.objects.filter(Q(id=sId)).first()
            if ssg:
                short = ssg.get_label()
                co_json['name'] = short
            data = json.dumps(co_json)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("get_ssg2ssg")
    else:
        data = "Request is not ajax"
    mimetype = "application/json"
    return HttpResponse(data, mimetype)

@csrf_exempt
def get_ssg(request):
    """Get ONE particular short representation of a SSG"""
    
    data = 'fail'
    if is_ajax(request):
        oErr = ErrHandle()
        try:
            sId = request.GET.get('id', '')
            co_json = {'id': sId}
            lstQ = []
            lstQ.append(Q(id=sId))
            ssg = EqualGold.objects.filter(Q(id=sId)).first()
            if ssg:
                short = ssg.get_short()
                co_json['name'] = short
            data = json.dumps(co_json)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("get_ssg")
    else:
        data = "Request is not ajax"
    mimetype = "application/json"
    return HttpResponse(data, mimetype)

@csrf_exempt
def get_ssgdist(request):
    """Get ONE particular short representation of a SSG"""
    
    data = 'fail'
    if is_ajax(request):
        oErr = ErrHandle()
        try:
            sId = request.GET.get('id', '')
            co_json = {'id': sId}
            lstQ = []
            lstQ.append(Q(id=sId))
            dist = SermonEqualDist.objects.filter(Q(id=sId)).first()
            if dist != None:
                ssg = EqualGold.objects.filter(Q(id=dist.super.id)).first()
                if ssg:
                    short = ssg.get_short()
                    co_json['name'] = short
                    data = json.dumps(co_json)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("get_ssgdist")
    else:
        data = "Request is not ajax"
    mimetype = "application/json"
    return HttpResponse(data, mimetype)

@csrf_exempt
def get_manuidnos(request):
    """Get a list of manuscript identifiers for autocomplete"""

    oErr = ErrHandle()
    try:
        data = 'fail'
        if is_ajax(request):
            idno = request.GET.get("name", "")
            lstQ = []
            lstQ.append(Q(idno__icontains=idno))
            items = Manuscript.objects.filter(*lstQ).order_by("idno").distinct()
            results = []
            for co in items:
                co_json = {'name': co.idno, 'id': co.id }
                results.append(co_json)
            data = json.dumps(results)
        else:
            data = "Request is not ajax"
    except:
        msg = oErr.get_error_message()
        data = "error: {}".format(msg)
    mimetype = "application/json"
    return HttpResponse(data, mimetype)

@csrf_exempt
def get_authors(request):
    """Get a list of authors for autocomplete"""

    data = 'fail'
    if is_ajax(request):
        author = request.GET.get("name", "")
        lstQ = []
        lstQ.append(Q(name__icontains=author)|Q(abbr__icontains=author))
        authors = Author.objects.filter(*lstQ).order_by('name')
        results = []
        for co in authors:
            co_json = {'name': co.name, 'id': co.id }
            results.append(co_json)
        data = json.dumps(results)
    else:
        data = "Request is not ajax"
    mimetype = "application/json"
    return HttpResponse(data, mimetype)

@csrf_exempt
def get_nicknames(request):
    """Get a list of nicknames for autocomplete"""

    oErr = ErrHandle()
    try:
        data = 'fail'
        if is_ajax(request):
            author = request.GET.get("name", "")
            lstQ = []
            lstQ.append(Q(name__icontains=author))
            authors = Nickname.objects.filter(*lstQ).order_by('name')
            results = []
            for co in authors:
                co_json = {'name': co.name, 'id': co.id }
                results.append(co_json)
            data = json.dumps(results)
        else:
            data = "Request is not ajax"
    except:
        msg = oErr.get_error_message()
        data = "error: {}".format(msg)
    mimetype = "application/json"
    return HttpResponse(data, mimetype)

@csrf_exempt
def get_gldincipits(request):
    """Get a list of Gold-sermon incipits for autocomplete"""

    oErr = ErrHandle()
    try:
        data = 'fail'
        if is_ajax(request):
            author = request.GET.get("name", "")
            lstQ = []
            lstQ.append(Q(srchincipit__icontains=author))
            items = SermonGold.objects.filter(*lstQ).values("srchincipit").distinct().all().order_by('srchincipit')
            # items = SermonGold.objects.order_by("incipit").distinct()
            # items = SermonGold.objects.filter(*lstQ).order_by('incipit').distinct()
            results = []
            for idx, co in enumerate(items):
                val = co['srchincipit']
                co_json = {'name': val, 'id': idx }
                results.append(co_json)
            data = json.dumps(results)
        else:
            data = "Request is not ajax"
    except:
        msg = oErr.get_error_message()
        data = "error: {}".format(msg)
    mimetype = "application/json"
    return HttpResponse(data, mimetype)

@csrf_exempt
def get_srmincipits(request):
    """Get a list of manifestation-sermon incipits for autocomplete"""

    oErr = ErrHandle()
    try:
        data = 'fail'
        if is_ajax(request):
            author = request.GET.get("name", "")
            lstQ = []
            lstQ.append(Q(srchincipit__icontains=author))
            items = SermonDescr.objects.filter(*lstQ).values("srchincipit").distinct().all().order_by('srchincipit')
            results = []
            for idx, co in enumerate(items):
                val = co['srchincipit']
                co_json = {'name': val, 'id': idx }
                results.append(co_json)
            data = json.dumps(results)
        else:
            data = "Request is not ajax"
    except:
        msg = oErr.get_error_message()
        data = "error: {}".format(msg)
    mimetype = "application/json"
    return HttpResponse(data, mimetype)

@csrf_exempt
def get_gldexplicits(request):
    """Get a list of Gold-sermon explicits for autocomplete"""

    oErr = ErrHandle()
    try:
        data = 'fail'
        if is_ajax(request):
            author = request.GET.get("name", "")
            lstQ = []
            lstQ.append(Q(srchexplicit__icontains=author))
            items = SermonGold.objects.filter(*lstQ).values("srchexplicit").distinct().all().order_by('srchexplicit')
            results = []
            for idx, co in enumerate(items):
                val = co['srchexplicit']
                co_json = {'name': val, 'id': idx }
                results.append(co_json)
            data = json.dumps(results)
        else:
            data = "Request is not ajax"
    except:
        msg = oErr.get_error_message()
        data = "error: {}".format(msg)
    mimetype = "application/json"
    return HttpResponse(data, mimetype)

@csrf_exempt
def get_srmexplicits(request):
    """Get a list of Manifestation-sermon explicits for autocomplete"""

    oErr = ErrHandle()
    try:
        data = 'fail'
        if is_ajax(request):
            author = request.GET.get("name", "")
            lstQ = []
            lstQ.append(Q(srchexplicit__icontains=author))
            items = SermonDescr.objects.filter(*lstQ).values("srchexplicit").distinct().all().order_by('srchexplicit')
            results = []
            for idx, co in enumerate(items):
                val = co['srchexplicit']
                co_json = {'name': val, 'id': idx }
                results.append(co_json)
            data = json.dumps(results)
        else:
            data = "Request is not ajax"
    except:
        msg = oErr.get_error_message()
        data = "error: {}".format(msg)
    mimetype = "application/json"
    return HttpResponse(data, mimetype)

@csrf_exempt
def get_gldsignatures(request):
    """Get a list of signature codes (SermonDescr) for autocomplete"""

    oErr = ErrHandle()
    try:
        data = 'fail'
        if is_ajax(request):
            # Get the editype, if that is specified
            editype = request.GET.get("type", "")
            # Get the complete code line, which could use semicolon-separation
            codeline = request.GET.get("name", "")
            codelist = codeline.split(";")
            codename = "" if len(codelist) == 0 else codelist[-1].strip()
            lstQ = []
            lstQ.append(Q(code__icontains=codename))
            if editype != "":
                lstQ.append(Q(editype=editype))
            items = Signature.objects.filter(*lstQ).order_by("code").distinct()
            items = list(items.values("code", "id"))

            # OLD METHOD items = items.values_list('code', "id")
            results = []
            for co in items:
                # OLD METHOD co_json = {'name': co.code, 'id': co.id }
                co_json = {'name': co["code"], 'id': co["id"] }
                results.append(co_json)
            data = json.dumps(results)
        else:
            data = "Request is not ajax"
    except:
        msg = oErr.get_error_message()
        data = "error: {}".format(msg)
    mimetype = "application/json"
    return HttpResponse(data, mimetype)

@csrf_exempt
def get_srmsignatures(request):
    """Get a list of signature codes (for SermonDescr) for autocomplete"""

    oErr = ErrHandle()
    try:
        data = 'fail'
        if is_ajax(request):
            codename = request.GET.get("name", "")
            editype = request.GET.get("type", "")
            lstQ = []
            lstQ.append(Q(code__icontains=codename))
            if editype != "":
                lstQ.append(Q(editype=editype))
            items = SermonSignature.objects.filter(*lstQ).order_by("code").distinct()
            results = []
            for co in items:
                co_json = {'name': co.code, 'id': co.id }
                results.append(co_json)
            data = json.dumps(results)
        else:
            data = "Request is not ajax"
    except:
        msg = oErr.get_error_message()
        data = "error: {}".format(msg)
    mimetype = "application/json"
    return HttpResponse(data, mimetype)

@csrf_exempt
def get_sermosig(request):
    """Get the correct GOLD signature, given a SERMON signature"""
    
    data = 'fail'
    if is_ajax(request):
        oErr = ErrHandle()
        try:
            sId = request.GET.get('id', '')
            co_json = {'id': sId}
            sermosig = SermonSignature.objects.filter(Q(id=sId)).first()
            if sermosig:
                short = sermosig.gsig.code
                co_json['name'] = short
            data = json.dumps(co_json)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("get_sermosig")
    else:
        data = "Request is not ajax"
    mimetype = "application/json"
    return HttpResponse(data, mimetype)

@csrf_exempt
def get_editions(request):
    """Get a list of edition codes for autocomplete"""

    oErr = ErrHandle()
    try:
        data = 'fail'
        if is_ajax(request):
            author = request.GET.get("name", "")
            lstQ = []
            lstQ.append(Q(name__icontains=author))
            items = Edition.objects.filter(*lstQ).order_by("name").distinct()
            results = []
            for co in items:
                co_json = {'name': co.name, 'id': co.id }
                results.append(co_json)
            data = json.dumps(results)
        else:
            data = "Request is not ajax"
    except:
        msg = oErr.get_error_message()
        data = "error: {}".format(msg)
    mimetype = "application/json"
    return HttpResponse(data, mimetype)

@csrf_exempt
def get_keywords(request):
    """Get a list of keywords for autocomplete"""

    oErr = ErrHandle()
    try:
        data = 'fail'
        if is_ajax(request):
            # Get the complete code line, which could use semicolon-separation
            kwline = request.GET.get("name", "")
            kwlist = kwline.split(";")
            kw = "" if len(kwlist) == 0 else kwlist[-1].strip()
            lstQ = []
            lstQ.append(Q(name__icontains=kw))
            items = Keyword.objects.filter(*lstQ).order_by("name").distinct()
            results = []
            for co in items:
                co_json = {'name': co.name, 'id': co.id }
                results.append(co_json)
            data = json.dumps(results)
        else:
            data = "Request is not ajax"
    except:
        msg = oErr.get_error_message()
        data = "error: {}".format(msg)
    mimetype = "application/json"
    return HttpResponse(data, mimetype)

@csrf_exempt
def get_collections(request):
    """Get a list of collections for autocomplete"""

    oErr = ErrHandle()
    try:
        data = 'fail'
        if is_ajax(request):
            # Get the complete code line, which could use semicolon-separation
            coll_line = request.GET.get("name", "")
            coll_list = coll_line.split(";")
            col = "" if len(coll_list) == 0 else coll_list[-1].strip()
            lstQ = []
            lstQ.append(Q(name__icontains=col))
            items = Collection.objects.filter(*lstQ).order_by("name").distinct()
            results = []
            for co in items:
                co_json = {'name': co.name, 'id': co.id }
                results.append(co_json)
            data = json.dumps(results)
        else:
            data = "Request is not ajax"
    except:
        msg = oErr.get_error_message()
        data = "error: {}".format(msg)
    mimetype = "application/json"
    return HttpResponse(data, mimetype)

@csrf_exempt
def get_gold(request, pk=None):
    """Get details of one particular gold sermon"""

    oErr = ErrHandle()
    data = {'status': 'fail'}
    fields = ['author', 'incipit', 'explicit', 'critlinks', 'bibliography' ]
    try:
        if is_ajax(request) and user_is_authenticated(request):
            # Get the id of the gold sermon
            qd = request.GET if request.method == "GET" else request.POST
            goldid = qd.get("goldid", "")
            bFound = False
            if goldid == "" and pk != None:
                obj = SermonGold.objects.filter(id=pk).first()
                bFound = True
            else:
                signature = Signature.objects.filter(id=goldid).first()
                if signature==None:
                    data['msg'] = "Signature not found"
                    data['status'] = "error"
                else:
                    obj = signature.gold
                    bFound = True
            if obj == None:
                data['msg'] = "Gold not found"
                data['status'] = "error"
            elif bFound:
                # Copy all relevant information
                info = {}
                d = model_to_dict(obj)
                for field in fields:
                    info[field] = d[field]  
                # Add the authorname
                authorname = ""
                if obj.author != None:
                    authorname = obj.author.name                  
                info['authorname'] = authorname

                # Copy keywords
                info['keywords'] = [x['id'] for x in obj.keywords.all().values('id')]

                # Copy signatures
                info['signatures'] = [x['id'] for x in obj.goldsignatures.all().order_by('-editype').values('id')]

                data['data'] = info

                data['status'] = "ok"
        else:
            data['msg'] = "Request is not ajax"
            data['status'] = "error"
    except: 
        msg = oErr.get_error_message()
        data['msg'] = msg
        data['status'] = "error"
    mimetype = "application/json"
    return HttpResponse(json.dumps(data), mimetype)

@csrf_exempt
def import_authors(request):
    """Import a CSV file or a JSON file that contains author names"""


    # Initialisations
    # NOTE: do ***not*** add a breakpoint until *AFTER* form.is_valid
    arErr = []
    error_list = []
    transactions = []
    data = {'status': 'ok', 'html': ''}
    template_name = 'seeker/import_authors.html'
    obj = None
    data_file = ""
    bClean = False
    username = request.user.username

    # Check if the user is authenticated and if it is POST
    if not request.user.is_authenticated  or request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            # NOTE: from here a breakpoint may be inserted!
            print('valid form')
            # Get the contents of the imported file
            data_file = request.FILES['file_source']
            filename = data_file.name

            # Get the source file
            if data_file == None or data_file == "":
                arErr.append("No source file specified for the selected project")
            else:
                # Check the extension
                arFile = filename.split(".")
                extension = arFile[len(arFile)-1]

                # Further processing depends on the extension
                if extension == "json":
                    # This is a JSON file
                    oResult = Author.read_json(username, data_file, arErr)
                else:
                    # Read the list of authors as CSV
                    oResult = Author.read_csv(username, data_file, arErr)

                # Determine a status code
                statuscode = "error" if oResult == None or oResult['status'] == "error" else "completed"
                if oResult == None:
                    arErr.append("There was an error. No authors have been added")

            # Get a list of errors
            error_list = [str(item) for item in arErr]

            # Create the context
            context = dict(
                statuscode=statuscode,
                results=oResult,
                error_list=error_list
                )

            if len(arErr) == 0:
                # Get the HTML response
                data['html'] = render_to_string(template_name, context, request)
            else:
                data['html'] = "Please log in before continuing"


        else:
            data['html'] = 'invalid form: {}'.format(form.errors)
            data['status'] = "error"
    else:
        data['html'] = 'Only use POST and make sure you are logged in'
        data['status'] = "error"
 
    # Return the information
    return JsonResponse(data)

def get_pie_data():
    """Fetch data for a particular type of pie-chart for the home page
    
    Current types: 'sermo', 'super', 'manu'
    """

    oErr = ErrHandle()
    red = 0
    orange = 0
    green = 0
    combidata = {}
    ptypes = ['sermo', 'super', 'manu']
    try:
        # Get the values for app, edi, imp, man
        stype_list = FieldChoice.objects.filter(field="seeker.stype").values('id', 'abbr')
        oStype = {}
        for oItem in stype_list:
            oStype[oItem['abbr']] = oItem['id']
        for ptype in ptypes:
            qs = None
            url_red = ""
            url_ora = ""
            url_gre = ""
            if ptype == "sermo":
                qs = SermonDescr.objects.exclude(Q(mtype='tem') & Q(msitem__isnull=True)).order_by('stype').values('stype')
                url_red = "{}?sermo-stypelist={}&sermo-stypelist={}".format(reverse('sermon_list'), oStype['imp'], oStype['man'])
                url_ora = "{}?sermo-stypelist={}".format(reverse('sermon_list'), oStype['edi'])
                url_gre = "{}?sermo-stypelist={}".format(reverse('sermon_list'), oStype['app'])
            elif ptype == "super":
                qs = EqualGold.objects.filter(moved__isnull=True,atype='acc').order_by('stype').values('stype')
                url_red = "{}?ssg-stypelist={}&ssg-stypelist={}".format(reverse('equalgold_list'), oStype['imp'], oStype['man'])
                url_ora = "{}?ssg-stypelist={}".format(reverse('equalgold_list'), oStype['edi'])
                url_gre = "{}?ssg-stypelist={}".format(reverse('equalgold_list'), oStype['app'])
            elif ptype == "manu":
                qs = Manuscript.objects.exclude(Q(mtype='tem')).order_by('stype').values('stype')
                url_red = "{}?manu-stypelist={}&manu-stypelist={}".format(reverse('manuscript_list'), oStype['imp'], oStype['man'])
                url_ora = "{}?manu-stypelist={}".format(reverse('manuscript_list'), oStype['edi'])
                url_gre = "{}?manu-stypelist={}".format(reverse('manuscript_list'), oStype['app'])
            # Calculate the different stype values
            if qs != None:
                app = sum(x['stype'] == "app" for x in qs)  # Approved
                edi = sum(x['stype'] == "edi" for x in qs)  # Edited
                imp = sum(x['stype'] == "imp" for x in qs)  # Imported
                man = sum(x['stype'] == "man" for x in qs)  # Manually created
                und = sum(x['stype'] == "-" for x in qs)    # Undefined
                red = imp + und + man
                orange = edi
                green = app
            total = red + green + orange
            # Create a list of data
            data = []
            data.append({'name': 'Initial', 'value': red, 'total': total, 'url': url_red})
            data.append({'name': 'Edited', 'value': orange, 'total': total, 'url': url_ora})
            data.append({'name': 'Approved', 'value': green, 'total': total, 'url': url_gre})
            combidata[ptype] = data
    except:
        msg = oErr.get_error_message()
        combidata['msg'] = msg
        combidata['status'] = "error"
    return combidata 



class LocationListView(BasicList):
    """Listview of locations"""

    model = Location
    listform = LocationForm
    paginate_by = 15
    prefix = "loc"
    sg_name = "Location"     
    plural_name = "Locations"
    has_select2 = True
    order_cols = ['name', 'loctype__name', '', '']
    order_default = order_cols
    order_heads = [{'name': 'Name',         'order': 'o=1', 'type': 'str', 'custom': 'location', 'linkdetails': True, 'main': True},
                   {'name': 'Type',         'order': 'o=2', 'type': 'str', 'custom': 'loctype',  'linkdetails': True},
                   {'name': 'Part of...',   'order': '',    'type': 'str', 'custom': 'partof'},
                   {'name': '',             'order': '',    'type': 'str', 'custom': 'manulink' }]
    filters = [ {"name": "Name",    "id": "filter_location",    "enabled": False},
                {"name": "Type",    "id": "filter_loctype",     "enabled": False},
               ]
    searches = [
        {'section': '', 'filterlist': [
            {'filter': 'location', 'dbfield': 'name',       'keyS': 'location_ta', 'keyList': 'locchooser', 'infield': 'name' },
            {'filter': 'loctype',  'fkfield': 'loctype',    'keyList': 'loctypechooser', 'infield': 'name' }]}
        ]

    def get_field_value(self, instance, custom):
        sBack = ""
        sTitle = ""
        html = []
        if custom == "location":
            html.append(instance.name)
        elif custom == "loctype":
            sLocType = "-"
            if instance.loctype != None:
                sLocType = instance.loctype.name
            html.append(sLocType)
        elif custom == "partof":
            sName = instance.get_partof_html()
            if sName == "": sName = "<i>(part of nothing)</i>"
            html.append(sName)
        elif custom == "manulink":
            # This is currently unused
            pass
        # Combine the HTML code
        sBack = "\n".join(html)
        return sBack, sTitle


class LocationEdit(BasicDetails):
    model = Location
    mForm = LocationForm
    prefix = "loc"
    title = "Location"
    history_button = True
    mainitems = []

    def add_to_context(self, context, instance):
        """Add to the existing context"""

        # Define the main items to show and edit
        context['mainitems'] = [
            {'type': 'plain', 'label': "Name:",                     'value': instance.name,         'field_key': "name"},
            {'type': 'line',  'label': "Location type:",            'value': instance.loctype.name, 'field_key': 'loctype'},
            {'type': 'line',  'label': "This location is part of:", 'value': instance.get_partof_html(),   
             'field_list': "locationlist"}
            ]

        # Signal that we have select2
        context['has_select2'] = True

        # Return the context we have made
        return context

    def after_save(self, form, instance):
        """This is for processing items from the list of available ones"""

        msg = ""
        bResult = True
        oErr = ErrHandle()
        
        try:
            # Process many-to-many changes: Add and remove relations in accordance with the new set passed on by the user
            if getattr(form, 'cleaned_data') != None:
                # (1) 'locations'
                locationlist = form.cleaned_data['locationlist']
                adapt_m2m(LocationRelation, instance, "contained", locationlist, "container")
            
        except:
            msg = oErr.get_error_message()
            bResult = False
        return bResult, msg

    def action_add(self, instance, details, actiontype):
        """User can fill this in to his/her liking"""
        passim_action_add(self, instance, details, actiontype)

    def get_history(self, instance):
        return passim_get_history(instance)


class LocationDetails(LocationEdit):
    """Like Location Edit, but then html output"""
    rtype = "html"
    

class OriginListView(BasicList):
    """Listview of origins"""

    model = Origin
    listform = OriginForm
    prefix = "prj"
    sg_name = "Origin"     
    plural_name = "Origins"
    has_select2 = True
    paginate_by = 15
    page_function = "ru.passim.seeker.search_paged_start"
    order_cols = ['name', 'location', 'note', '']
    order_default = order_cols
    order_heads = [{'name': 'Name',     'order': 'o=1', 'type': 'str', 'custom': 'origin', 'main': True, 'linkdetails': True},
                   {'name': 'Location', 'order': 'o=2', 'type': 'str', 'custom': 'location'},
                   {'name': 'Note',     'order': 'o=3', 'type': 'str', 'custom': 'note'},
                   {'name': '',         'order': '',    'type': 'str', 'custom': 'manulink' }]
    filters = [ 
        {"name": "Location",        "id": "filter_location",    "enabled": False},
        {"name": "Shelfmark",       "id": "filter_manuid",      "enabled": False, "head_id": "filter_other"},
               ]
    searches = [
        {'section': '', 'filterlist': [
            {'filter': 'location', 'dbfield': 'name', 'keyS': 'location_ta', 'keyList': 'locationlist', 'infield': 'name' }]}
        ]

    def get_field_value(self, instance, custom):
        sBack = ""
        sTitle = ""
        html = []
        if custom == "manulink":
            # Link to manuscripts in this project
            count = instance.origin_manuscripts.all().count()
            url = reverse('search_manuscript')
            if count > 0:
                html.append("<a href='{}?manu-origin={}'><span class='badge jumbo-3 clickable' title='{} manuscripts with this origin'>{}</span></a>".format(
                    url, instance.id, count, count))
        elif custom == "location":
            sLocation = ""
            if instance.location:
                sLocation = instance.location.name
            html.append(sLocation)
        elif custom == "note":
            sNote = "" if not instance.note else instance.note
            html.append(sNote)
        elif custom == "origin":
            sName = instance.name
            if sName == "": sName = "<i>(unnamed)</i>"
            html.append(sName)
        # Combine the HTML code
        sBack = "\n".join(html)
        return sBack, sTitle


class OriginEdit(BasicDetails):
    """The details of one origin"""

    model = Origin
    mForm = OriginForm
    prefix = "org"
    title = "Origin" 
    rtype = "json"
    basic_name = "origin"
    history_button = True
    mainitems = []
    
    def add_to_context(self, context, instance):
        """Add to the existing context"""

        # Define the main items to show and edit
        context['mainitems'] = [
            {'type': 'plain', 'label': "Name:",             'value': instance.name,         'field_key': "name"},
            {'type': 'line',  'label': "Origin note:",      'value': instance.note,         'field_key': 'note'},
            {'type': 'plain', 'label': "Origin location:",  'value': instance.get_location(),   'field_key': "location"}
            ]

        # Signal that we have select2
        context['has_select2'] = True

        # Return the context we have made
        return context

    def action_add(self, instance, details, actiontype):
        """User can fill this in to his/her liking"""
        passim_action_add(self, instance, details, actiontype)

    def get_history(self, instance):
        return passim_get_history(instance)


class OriginDetails(OriginEdit):
    """Like Origin Edit, but then html output"""
    rtype = "html"
    

class OriginCodEdit(BasicDetails):
    """The details of one 'origin'"""

    model = OriginCod
    mForm = OriginCodForm
    prefix = 'cori'
    title = "CodicoOrigin"
    history_button = False # True
    # rtype = "json"
    # mainitems = []
    
    def add_to_context(self, context, instance):
        """Add to the existing context"""

        # Define the main items to show and edit
        context['mainitems'] = [
            {'type': 'safe',  'label': "Origin:",   'value': instance.get_origin()},
            {'type': 'plain', 'label': "Note:",         'value': instance.note,   'field_key': 'note'     },
            ]

        # Signal that we have select2
        context['has_select2'] = True

        context['listview'] = reverse("codico_details", kwargs={'pk': instance.codico.id})

        # Return the context we have made
        return context

    def action_add(self, instance, details, actiontype):
        """User can fill this in to his/her liking"""
        passim_action_add(self, instance, details, actiontype)

    def get_history(self, instance):
        return passim_get_history(instance)


class OriginCodDetails(OriginCodEdit):
    """Like OriginCod Edit, but then html output"""
    rtype = "html"
        

class SermonEdit(BasicDetails):
    """The editable part of one sermon description (manifestation)"""
    
    model = SermonDescr
    mForm = SermonForm
    prefix = "sermo"
    title = "Manifestation" 
    rtype = "json"
    mainitems = []
    basic_name = "sermon"
    use_team_group = True
    history_button = True
    comment_button = True
    prefix_type = "simple"

    StossgFormSet = inlineformset_factory(SermonDescr, SermonDescrEqual,
                                         form=SermonDescrSuperForm, min_num=0,
                                         fk_name = "sermon",
                                         extra=0, can_delete=True, can_order=False)
    SDkwFormSet = inlineformset_factory(SermonDescr, SermonDescrKeyword,
                                       form=SermonDescrKeywordForm, min_num=0,
                                       fk_name="sermon", extra=0)
    SDcolFormSet = inlineformset_factory(SermonDescr, CollectionSerm,
                                       form=SermonDescrCollectionForm, min_num=0,
                                       fk_name="sermon", extra=0)
    SDsignFormSet = inlineformset_factory(SermonDescr, SermonSignature,
                                         form=SermonDescrSignatureForm, min_num=0,
                                         fk_name = "sermon",
                                         extra=0, can_delete=True, can_order=False)
    SbrefFormSet = inlineformset_factory(SermonDescr, BibRange,
                                         form=BibRangeForm, min_num=0,
                                         fk_name = "sermon",
                                         extra=0, can_delete=True, can_order=False)
    SlinkFormSet = inlineformset_factory(SermonDescr, SermonDescrLink,
                                         form=SermonDescrLinkForm, min_num=0,
                                         fk_name = "src",
                                         extra=0, can_delete=True, can_order=False)

    formset_objects = [
        {'formsetClass': StossgFormSet, 'prefix': 'stossg', 'readonly': False, 'noinit': True, 'linkfield': 'sermon'},
        {'formsetClass': SDkwFormSet,   'prefix': 'sdkw',   'readonly': False, 'noinit': True, 'linkfield': 'sermon'},                       
        {'formsetClass': SDcolFormSet,  'prefix': 'sdcol',  'readonly': False, 'noinit': True, 'linkfield': 'sermo'},
        {'formsetClass': SDsignFormSet, 'prefix': 'sdsig',  'readonly': False, 'noinit': True, 'linkfield': 'sermon'},
        {'formsetClass': SbrefFormSet,  'prefix': 'sbref',  'readonly': False, 'noinit': True, 'linkfield': 'sermon'},
        {'formsetClass': SlinkFormSet,  'prefix': 'slink',  'readonly': False, 'noinit': True, 'initial': [{'linktype': LINK_REL }], 'clean': True},
        ] 

    stype_edi_fields = ['manu', 'locus', 'author', 'sectiontitle', 'title', 'subtitle', 'incipit', 'explicit', 'fulltext',
                        'postscriptum', 'quote', 'bibnotes', 'feast', 'bibleref', 'additional',  
                        #'kwlist', 
                        'altpage', 'altpageslist', 
                        'note', 'notes_altpageslist',
                        'SermonSignature', 'siglist',
                        #'CollectionSerm', 'collist_s',
                        'SermonDescrEqual', 'superlist']

    def custom_init(self, instance):
        method = "nodistance"   # Alternative: "superdist"

        if instance:
            istemplate = (instance.mtype == "tem")
            if istemplate:
                # Need a smaller array of formset objects
                self.formset_objects = [{'formsetClass': self.StossgFormSet, 'prefix': 'stossg', 'readonly': False, 'noinit': True, 'linkfield': 'sermon'}]

            # Indicate where to go to after deleting
            if instance != None and instance.msitem != None and instance.msitem.manu != None:
                self.afterdelurl = reverse('manuscript_details', kwargs={'pk': instance.msitem.manu.id})

            # Then check if all distances have been calculated in SermonEqualDist
            if method == "superdist":
                qs = SermonEqualDist.objects.filter(sermon=instance)
                if qs.count() == 0:
                    # These distances need calculation...
                    instance.do_distance()
        return None

    def get_form_kwargs(self, prefix):
        # Determine the method
        method = "nodistance"   # Alternative: "superdist"

        oBack = None
        if prefix == 'stossg' and method == "superdist":
            if self.object != None:
                # Make sure that the sermon is known
                oBack = dict(sermon_id=self.object.id)
        elif prefix == "slink":
            if self.object != None:
                # Make sure to return the ID of the Manuscript
                oBack = dict(sermon_id=self.object.id)

        return oBack
           
    def add_to_context(self, context, instance):
        """Add to the existing context"""

        oErr = ErrHandle()

        try:
            # Need to know who this user (profile) is
            profile = Profile.get_user_profile(self.request.user.username)

            istemplate = (instance.mtype == "tem")

            # Define the main items to show and edit
            # manu_id = None if instance == None or instance.manu == None else instance.manu.id
            manu_id = None if instance == None else instance.get_manuscript().id
            context['mainitems'] = []
            # Possibly add the Template identifier
            if istemplate:
                context['mainitems'].append(
                    {'type': 'plain', 'label': "Template:", 'value': instance.get_template_link(profile)}
                    )

            # Prepare saveditem handling
            saveditem_button = get_saveditem_html(self.request, instance, profile, sitemtype="serm")
            saveditem_form = get_saveditem_html(self.request, instance, profile, "form", sitemtype="serm")

            # If this user belongs to the ProjectApprover of HUWA, show him the HUWA ID if it is there
            if not istemplate and not instance is None:
                # NOTE: this code should be extended to include other external projects, when the time is right
                bHuwaEditor = profile.is_project_editor("huwa")
                for ext in SermonDescrExternal.objects.filter(sermon=instance):
                    # Get the field values
                    externaltype = ext.externaltype
                    externalid = ext.externalid
                    if externaltype == "huwop" and externalid > 0 and bHuwaEditor:
                        # Okay, the user has the right to see the externalid
                        oItem = dict(type="plain", label="HUWA opera id", value=externalid, field_key="ext_huwop")
                        context['mainitems'].insert(0, oItem)
                    elif externaltype == "huwin" and externalid > 0 and bHuwaEditor:
                        # Okay, the user has the right to see the externalid
                        oItem = dict(type="plain", label="HUWA inhalt id", value=externalid, field_key="ext_huwin")
                        context['mainitems'].insert(0, oItem)

            # Get the main items
            mainitems_main = [
                {'type': 'plain', 'label': "Status:",               'value': instance.get_stype_light(),'field_key': 'stype'}, # TH: hier worden de comments opgepikt
                # -------- HIDDEN field values ---------------
                {'type': 'plain', 'label': "Manuscript id",         'value': manu_id,                   'field_key': "manu",        'empty': 'hide'},
                # --------------------------------------------
                {'type': 'safe',  'label': "Saved item:",           'value': saveditem_button          },
                {'type': 'plain', 'label': "Locus:",                'value': instance.locus,            'field_key': "locus"},                 
                
                {'type': 'safe',  'label': "Attributed author:",    'value': instance.get_author(),     'field_key': 'author'},
                {'type': 'plain', 'label': "Author certainty:",     'value': instance.get_autype(),     'field_key': 'autype', 'editonly': True},
                {'type': 'plain', 'label': "Section title:",        'value': instance.sectiontitle,     'field_key': 'sectiontitle'},
                {'type': 'safe',  'label': "Lectio:",               'value': instance.get_quote_markdown(),'field_key': 'quote'}, 
                {'type': 'plain', 'label': "Title:",                'value': instance.title,            'field_key': 'title'},
                # Issue #237, delete subtitle
                {'type': 'plain', 'label': "Sub title:",            'value': instance.subtitle,         'field_key': 'subtitle', 
                 'editonly': True, 'title': 'The subtitle field is legacy. It is edit-only, non-viewable'},
                {'type': 'safe',  'label': "Incipit:",              'value': instance.get_incipit_markdown(), 
                 'field_key': 'incipit',  'key_ta': 'srmincipit-key'}, 
                {'type': 'safe',  'label': "Explicit:",             'value': instance.get_explicit_markdown(),
                 'field_key': 'explicit', 'key_ta': 'srmexplicit-key'}, 
                {'type': 'safe',  'label': "Postscriptum:",         'value': instance.get_postscriptum_markdown(),
                 'field_key': 'postscriptum'}, 
                {'type': 'safe',  'label': "Transcription:",        'value': self.get_transcription(instance),
                 'field_key': 'fulltext'}, 
                {'type': 'plain', 'label': "Alt page numbering:",   'value': instance.get_altpages(),   'field_list': "altpageslist"},
                 ]
            # issue #715: full-text may not be visible to 'regular' users

            # some more items
            mainitems_add = [
                {'type': 'plain', 'label': "Notes alt page numbering:",   'value': instance.get_notes_altpages(), 'field_list': "notes_altpageslist"},
                # Issue #23: delete bibliographic notes
                {'type': 'plain', 'label': "Bibliographic notes:",  'value': instance.bibnotes,         'field_key': 'bibnotes', 
                 'editonly': True, 'title': 'The bibliographic-notes field is legacy. It is edit-only, non-viewable'},
                {'type': 'plain', 'label': "Feast:",                'value': instance.get_feast(),      'field_key': 'feast'}
                ]
            for oItem in mainitems_add: mainitems_main.append(oItem)
            exclude_field_keys = ['locus']
            for item in mainitems_main: 
                # Make sure to exclude field key 'locus'
                fk = item.get("field_key")
                is_in_exclude_field_keys = False if fk is None else (fk in exclude_field_keys)
                if not istemplate or not is_in_exclude_field_keys:
                # if not istemplate or item['field_key'] not in exclude_field_keys:
                    context['mainitems'].append(item)

            # Bibref and Cod. notes can only be added to non-templates
            if not istemplate:
                mainitems_BibRef ={'type': 'plain', 'label': "Bible reference(s):",   'value': instance.get_bibleref(),        
                 'multiple': True, 'field_list': 'bibreflist', 'fso': self.formset_objects[4]}
                context['mainitems'].append(mainitems_BibRef)
                mainitems_CodNotes ={'type': 'plain', 'label': "Cod. notes:",           'value': instance.additional,       
                 'field_key': 'additional',   'title': 'Codicological notes'}
                context['mainitems'].append(mainitems_CodNotes)

            mainitems_more =[
                {'type': 'line', 'label': "Note:",                 'value': instance.get_note_markdown(),             'field_key': 'note'}
                ]
            for item in mainitems_more: context['mainitems'].append(item)

            if not istemplate:
                username = profile.user.username
                team_group = app_editor
                mainitems_m2m = [
                    {'type': 'line',  'label': "Keywords:",             'value': instance.get_keywords_markdown(), 
                     # 'multiple': True,  'field_list': 'kwlist',         'fso': self.formset_objects[1]},
                     'field_list': 'kwlist',         'fso': self.formset_objects[1]},
                    {'type': 'plain', 'label': "Keywords (user):", 'value': self.get_userkeywords(instance, profile, context),   'field_list': 'ukwlist',
                     'title': 'User-specific keywords. If the moderator accepts these, they move to regular keywords.'},
                    {'type': 'line',  'label': "Keywords (related):",   'value': instance.get_keywords_ssg_markdown(),
                     'title': 'Keywords attached to the Authority file(s)'},
                    {'type': 'line',    'label': "Gryson/Clavis:",'value': instance.get_eqsetsignatures_markdown('combi'),
                     'title': "Gryson/Clavis codes of the Sermons Gold that are part of the same equality set + those manually linked to this manifestation Sermon" }, 
                    {'type': 'line',    'label': "Gryson/Clavis (manual):",'value': instance.get_sermonsignatures_markdown(),
                     'title': "Gryson/Clavis codes manually linked to this manifestation Sermon", 'unique': True, 'editonly': True, 
                     'multiple': True,
                     'field_list': 'siglist_m', 'fso': self.formset_objects[3], 'template_selection': 'ru.passim.sigs_template'},
                    {'type': 'plain',   'label': "Personal datasets:",  'value': instance.get_collections_markdown(username, team_group, settype="pd"), 
                     'multiple': True,  'field_list': 'collist_s',      'fso': self.formset_objects[2] },
                    {'type': 'plain',   'label': "Public datasets (link):",  'value': instance.get_collection_link("pd"), 
                     'title': "Public datasets in which an Authority file is that is linked to this sermon"},
                    {'type': 'plain',   'label': "Historical collections (link):",  'value': instance.get_collection_link("hc"), 
                     'title': "Historical collections in which an Authority file is that is linked to this sermon"},
                    {'type': 'line',    'label': "Editions:",           'value': instance.get_editions_markdown(),
                     'title': "Editions of the Sermons Gold that are part of the same equality set"},
                    {'type': 'line',    'label': "Literature:",         'value': instance.get_litrefs_markdown()},
                    # Project2 HIER
                    {'type': 'plain', 'label': "Project:",     'value': instance.get_project_markdown2(), 'field_list': 'projlist'},
                    # AltPageNumber HIER
                    #{'type': 'plain', 'label': "Alternative page numbering:", 'value': instance.get_altpage_markdown(), 'field_list': 'altpagelist'},
                    {'type': 'line',  'label': "Related Manifestations:",  'value': instance.get_sermolinks_markdown(), 
                        'multiple': True,  'field_list': 'slinklist',   'fso': self.formset_objects[5]},

                    ]
                for item in mainitems_m2m: context['mainitems'].append(item)
            # IN all cases
            mainitems_SSG = {'type': 'line',    'label': "Authority file links:",  'value': self.get_superlinks_markdown(instance), 
                 'multiple': True,  'field_list': 'superlist',       'fso': self.formset_objects[0], 
                 'inline_selection': 'ru.passim.ssglink_template',   'template_selection': 'ru.passim.ssg_template'}
            context['mainitems'].append(mainitems_SSG)
            # Notes:
            # Collections: provide a link to the Sermon-listview, filtering on those Sermons that are part of one particular collection

            # If this user belongs to the ProjectApprover of HUWA, show him the HUWA ID if it is there
            if not instance is None:
                # NOTE: this code should be extended to include other external projects, when the time is right
                ext = SermonDescrExternal.objects.filter(sermon=instance).first()
                if not ext is None:
                    # Get the field values
                    externaltype = ext.externaltype
                    externalid = ext.externalid
                    if externaltype == "huwop" and externalid > 0 and profile.is_project_approver("huwa"):
                        # THis is the HUWA 'opera_id'
                        opera_id = externalid
                        # Now we need to get the handschrift_id from the manuscript
                        manu = instance.get_manuscript()
                        obj = ManuscriptExternal.objects.filter(manu=manu).first()
                        if not obj is None:
                            handschrift_id = obj.externalid
                            # Let's see if there is any editor information
                            lst_edilit = get_huwa_opera_literature(opera_id, handschrift_id)
                            context['edilits'] = lst_edilit
                            value = render_to_string('seeker/huwa_edilit.html', context, self.request)
                            oItemEdi = dict(type="safe", label="HUWA editions", value=value)
                            context['mainitems'].append(oItemEdi)

            # Add a button back to the Manuscript
            topleftlist = []
            if instance.get_manuscript():
                manu = instance.get_manuscript()
                buttonspecs = {'label': "Manuscript", 
                     'title': "Return to manuscript {}".format(manu.idno), 
                     'url': reverse('manuscript_details', kwargs={'pk': manu.id})}
                topleftlist.append(buttonspecs)
                lcity = "" if manu.lcity == None else "{}, ".format(manu.lcity.name)
                lib = "" if manu.library == None else "{}, ".format(manu.library.name)
                idno = "{}{}{}".format(lcity, lib, manu.idno)
            else:
                idno = "(unknown)"
            context['topleftbuttons'] = topleftlist
            # Add something right to the SermonDetails title
            # OLD: context['title_addition'] = instance.get_eqsetsignatures_markdown('first')
            # Changed in issue #241: show the PASSIM code
            context['title_addition'] = instance.get_passimcode_markdown()
            # Add the manuscript's IDNO completely right
            title_right = ["<span class='manuscript-manifes' title='Manuscript'>{}</span>".format(
                idno)]
            #    ... as well as the *title* of the Codico to which I belong
            codico = instance.msitem.codico

            # Old code for [codi_title]: codi_title = "?" if codico == None or codico.name == "" else codico.name
            # Issue #422: change the text of the [codi_title]
            codi_title = "cod. unit. {}".format(codico.order)
            title_right.append("&nbsp;<span class='codico-title-manifes' title='Codicologial unit'>{}</span>".format(codi_title))
            context['title_right'] = "".join(title_right)

            #context['count_comments'] = instance.get_comments(True)

            # Signal that we have select2
            context['has_select2'] = True

            # Start making room for AfterDetails
            lhtml = []

            # These actions are only available for editors
            if user_is_ingroup(self.request, app_editor):
                # Add all needed information for moving a sermon and other sermon actions
                lhtml.append(render_to_string("seeker/sermon_action.html", context, self.request))

            # Add comment modal stuff
            initial = dict(otype="sermo", objid=instance.id, profile=profile)
            context['commentForm'] = CommentForm(initial=initial, prefix="com")
            context['comment_list'] = get_usercomments('sermo', instance, profile)
            lhtml.append(render_to_string("seeker/comment_add.html", context, self.request))
            context['comment_count'] = instance.comments.count()

            # Also add the saved item form
            lhtml.append(saveditem_form)

            # Store the after_details in the context
            context['after_details'] = "\n".join(lhtml)

            # Make stable URI available
            context['stable_uri'] = self.get_abs_uri('sermon_details', instance)

        except:
            msg = oErr.get_error_message()
            oErr.DoError("SermonEdit/add_to_context")

        # Return the context we have made
        return context

    def get_superlinks_markdown(self, instance):
        context = {}
        template_name = 'seeker/sermon_superlinks.html'
        sBack = ""
        if instance:
            # Add to context
            context['superlist'] = instance.sermondescr_super.all().order_by('sermon__author__name', 'sermon__siglist')
            context['is_app_editor'] = user_is_ingroup(self.request, app_editor)
            context['object_id'] = instance.id
            # Calculate with the template
            sBack = render_to_string(template_name, context)
        return sBack

    def get_transcription(self, instance):
        """Make a good visualization of a transcription, which includes a show/hide button"""

        sBack = ""
        oErr = ErrHandle()
        try:
            # Get the text
            if not instance.fulltext is None and instance.fulltext != "":
                sText = instance.get_fulltext_markdown("actual", lowercase=False)
                # Get URL of deleting transcription
                transdelurl = reverse('sermon_transdel', kwargs={'pk': instance.id})
                sInfo = instance.fullinfo
                oInfo = {}
                if not sInfo is None:
                    oInfo = json.loads(sInfo)
                wordcount = oInfo.get("wordcount", 0)
                # Check whether this is a regular user or not
                if user_is_ingroup(self.request, app_editor) or user_is_ingroup(self.request, app_moderator):
                    # Combine with button click + default hidden
                    context = dict(delete_permission=user_is_ingroup(self.request, stemma_editor),
                                   delete_message="",
                                   wordcount=wordcount,
                                   transdelurl=transdelurl,
                                   fulltext=sText)
                    sBack = render_to_string("seeker/ftext_buttons.html", context, None)
                else:
                    # issue #715: only show non-clickable word count
                    sBack = "{}".format(wordcount)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("SermonEdit/get_transcription")
        return sBack

    def get_userkeywords(self, instance, profile, context):
        sBack = ""
        oErr = ErrHandle()
        try:
            # At the very least get the list of user keywords
            sKeywordList = instance.get_keywords_user_markdown(profile)
            # Now determine what interface to show
            if context['is_app_editor']:
                # We simply return the list
                sBack = sKeywordList
            else:
                # We need to construct an interface:
                # (1) Set up the context to include view and edit elements
                context['userkeywords'] = sKeywordList
                context['ukwform'] = context['{}Form'.format(self.prefix)]
                context['ukwupdate'] = reverse('sermon_ukw', kwargs={'pk': instance.id })

                # Combine and provide the code needed
                # sBack = get_userkeyword_interface(self.request, context)
                sBack = render_to_string("seeker/userkeyword_edit.html", context, self.request)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("get_userkeywords")
        return sBack

    def after_new(self, form, instance):
        """Action to be performed after adding a new item"""

        ## Set the 'afternew' URL
        manu = instance.get_manuscript()
        if manu and instance.order < 0:
            # Calculate how many sermons there are
            sermon_count = manu.get_sermon_count()
            # Make sure the new sermon gets changed
            form.instance.order = sermon_count

        if manu:
            # Need to know who is 'talking'...
            username = self.request.user.username
            profile = Profile.get_user_profile(username)
            profile_projects = ProjectApprover.objects.filter(profile=profile, status="incl")

            project_count = instance.projects.count()

            if project_count == 0:
                # How many projects are attached to this manuscript
                manu_project_count = manu.projects.count()
                if manu_project_count == 1:
                    # Issue #546: if the manuscript is for one project, continue checking
                    # Assign this sermon to the project of the manuscript
                    project = manu.projects.first()
                    instance.projects.add(project)
                elif profile_projects.count() == 1:
                    # This editor is only editor for one project
                    # Issue #546: if editor is only ProjectApprover for one project, then assign the sermon to that project
                    project = profile_projects.first().project
                    instance.projects.add(project)

        # Return positively
        return True, "" 

    def process_formset(self, prefix, request, formset):
        """This is for processing *NEWLY* added items (using the '+' sign)"""

        bAllowNewSignatureManually = True   # False
        errors = []
        bResult = True
        oErr = ErrHandle()
        # Determine the method
        method = "nodistance"   # Alternative: "superdist"
        try:
            instance = formset.instance
            for form in formset:
                if form.is_valid():
                    cleaned = form.cleaned_data
                    # Action depends on prefix
                    if prefix == "sdsig" and bAllowNewSignatureManually:
                        # Signature processing
                        # NOTE: this should never be reached, because we do not allow adding *new* signatures manually here
                        editype = ""
                        code = ""
                        if 'newgr' in cleaned and cleaned['newgr'] != "":
                            # Add gryson
                            editype = "gr"
                            code = cleaned['newgr']
                        elif 'newcl' in cleaned and cleaned['newcl'] != "":
                            # Add gryson
                            editype = "cl"
                            code = cleaned['newcl']
                        elif 'newot' in cleaned and cleaned['newot'] != "":
                            # Add gryson
                            editype = "ot"
                            code = cleaned['newot']
                        if editype != "":
                            # Find this item in the Gold Signatures
                            gsig = Signature.objects.filter(editype=editype, code=code).first()
                            if gsig != None:
                                form.instance.gsig = gsig
                            # Set the correct parameters
                            form.instance.code = code
                            form.instance.editype = editype
                            # Note: it will get saved with formset.save()
                    elif prefix == "sdkw":
                        # Keyword processing
                        if 'newkw' in cleaned and cleaned['newkw'] != "":
                            newkw = cleaned['newkw']
                            # Is the KW already existing?
                            obj = Keyword.objects.filter(name=newkw).first()
                            if obj == None:
                                obj = Keyword.objects.create(name=newkw)
                            # Make sure we set the keyword
                            form.instance.keyword = obj
                            # Note: it will get saved with formset.save()
                    elif prefix == "sdcol":
                        # Collection processing
                        if 'newcol' in cleaned and cleaned['newcol'] != "":
                            newcol = cleaned['newcol']
                            # Is the COL already existing?
                            obj = Collection.objects.filter(name=newcol).first()
                            if obj == None:
                                # TODO: add profile here
                                profile = Profile.get_user_profile(request.user.username)
                                obj = Collection.objects.create(name=newcol, type='sermo', owner=profile)
                            # Make sure we set the keyword
                            form.instance.collection = obj
                            # Note: it will get saved with formset.save()
                    elif prefix == "stossg":
                        # SermonDescr-To-EqualGold processing
                        if method == "superdist":
                            # Note: nov/2 went over from 'newsuper' to 'newsuperdist'
                            if 'newsuperdist' in cleaned and cleaned['newsuperdist'] != "":
                                newsuperdist = cleaned['newsuperdist']
                                # Take the default linktype
                                linktype = "uns"

                                # Convert from newsuperdist to actual super (SSG)
                                superdist = SermonEqualDist.objects.filter(id=newsuperdist).first()
                                if superdist != None:
                                    super = superdist.super

                                    # Check existence of link between S-SSG
                                    obj = SermonDescrEqual.objects.filter(sermon=instance, super=super, linktype=linktype).first()
                                    if obj == None:
                                        # Set the right parameters for creation later on
                                        form.instance.linktype = linktype
                                        form.instance.super = super
                        elif method == "nodistance":
                            if 'newsuper' in cleaned and cleaned['newsuper'] != "":
                                newsuper = cleaned['newsuper']
                                # Take the default linktype
                                linktype = "uns"

                                # Check existence
                                obj = SermonDescrEqual.objects.filter(sermon=instance, super=newsuper, linktype=linktype).first()
                                if obj == None:
                                    obj_super = EqualGold.objects.filter(id=newsuper).first()
                                    if obj_super != None:
                                        # Set the right parameters for creation later on
                                        form.instance.linktype = linktype
                                        form.instance.super = obj_super

                        # Note: it will get saved with form.save()
                    elif prefix == "sbref":
                        # Processing one BibRange
                        newintro = cleaned.get('newintro', None)
                        onebook = cleaned.get('onebook', None)
                        newchvs = cleaned.get('newchvs', None)
                        newadded = cleaned.get('newadded', None)

                        # Minimal need is BOOK
                        if onebook != None:
                            # Note: normally it will get saved with formset.save()
                            #       However, 'noinit=False' formsets must arrange their own saving

                            #bNeedSaving = False

                            ## Double check if this one already exists for the current instance
                            #obj = instance.sermonbibranges.filter(book=onebook, chvslist=newchvs, intro=newintro, added=newadded).first()
                            #if obj == None:
                            #    obj = BibRange.objects.create(sermon=instance, book=onebook, chvslist=newchvs)
                            #    bNeedSaving = True
                            #if newintro != None and newintro != "": 
                            #    obj.intro = newintro
                            #    bNeedSaving = True
                            #if newadded != None and newadded != "": 
                            #    obj.added = newadded
                            #    bNeedSaving = True
                            #if bNeedSaving:
                            #    obj.save()
                            #    x = instance.sermonbibranges.all()

                            form.instance.book = onebook
                            if newchvs != None:
                                form.instance.chvslist = newchvs
                            form.instance.intro = newintro
                            form.instance.added = newadded

                    elif prefix == "slink":
                        # Process the link from S to S
                        newsermo = cleaned.get("newsermo")
                        if not newsermo is None:
                            # There also must be a linktype
                            if 'newlinktype' in cleaned and cleaned['newlinktype'] != "":
                                linktype = cleaned['newlinktype']
                                # Get optional parameters
                                note = cleaned.get('note', None)
                                # Check existence
                                obj = SermonDescrLink.objects.filter(src=instance, dst=newsermo, linktype=linktype).first()
                                if obj == None:
                                    sermo = SermonDescr.objects.filter(id=newsermo.id).first()
                                    if sermo != None:

                                        # Set the right parameters for creation later on
                                        form.instance.linktype = linktype
                                        form.instance.dst = sermo
                                        if note != None and note != "": 
                                            form.instance.note = note

                                        # Double check reverse
                                        if linktype in LINK_BIDIR_MANU:
                                            rev_link = SermonDescrLink.objects.filter(src=sermo, dst=instance).first()
                                            if rev_link == None:
                                                # Add it
                                                rev_link = SermonDescrLink.objects.create(src=sermo, dst=instance, linktype=linktype)
                                            else:
                                                # Double check the linktype
                                                if rev_link.linktype != linktype:
                                                    rev_link.linktype = linktype
                                            if note != None and note != "": 
                                                rev_link.note = note
                                            rev_link.save()
                        # Note: it will get saved with form.save()
                            

                else:
                    errors.append(form.errors)
                    bResult = False
        except:
            msg = oErr.get_error_message()
            iStop = 1
        return None

    def before_save(self, form, instance):
        oErr = ErrHandle()
        bBack = True
        msg = ""
        try:
            if hasattr(form, 'cleaned_data'):
                # Make sure the author type is processed correctly
                if 'autype' in form.cleaned_data and form.cleaned_data['autype'] != "":
                    autype = form.cleaned_data['autype']
                    form.instance.autype = autype

                # Issue #421: check how many projects are attached to the manuscript
                if not instance is None and not instance.msitem is None and not instance.msitem.manu is None:
                    # Need to know who is 'talking'...
                    username = self.request.user.username
                    profile = Profile.get_user_profile(username)
                    profile_projects = ProjectApprover.objects.filter(profile=profile, status="incl")

                    # Always get the project list
                    projlist = form.cleaned_data.get("projlist")

                    # There is a sermon and a manuscript
                    manu = instance.msitem.manu
                    # How many projects are attached to this manuscript
                    manu_project_count = manu.projects.count()
                    if manu_project_count > 1:
                        # There are multiple projects attached to the manuscript
                        # This means that the user *must* have specified one project

                        bBack, msg = evaluate_projlist(profile, instance, projlist, "Sermon manifestation")

                        #if len(projlist) == 0:
                        #    # Add a warning that the user must manually provide a project
                        #    msg = "Add a project: A sermon must belong to at least one project"
                        #    bBack = False
                    elif manu_project_count == 1:
                        # Issue #546: if the manuscript is for one project, continue checking
                        if len(projlist) == 0:
                            # Assign this sermon to the project of the manuscript
                            project = manu.projects.first()
                            instance.projects.add(project)
                        else:
                            bBack, msg = evaluate_projlist(profile, instance, projlist, "Sermon manifestation")
                    elif profile_projects.count() == 1:
                        # This editor is only editor for one project
                        # Issue #546: if editor is only ProjectApprover for one project, then assign the sermon to that project
                        project = profile_projects.first().project
                        instance.projects.add(project)
                    else:
                        # Action should not be performed if this is a template
                        if instance.mtype != 'tem':
                            # It would seem that this kind of check is needed anyway...
                            bBack, msg = evaluate_projlist(profile, instance, projlist, "Sermon manifestation")

        except:
            msg = oErr.get_error_message()
            oErr.DoError("SermonEdit/before_save")
            bBack = False
        return bBack, msg

    def after_save(self, form, instance):
        """This is for processing items from the list of available ones"""

        def adapt_external(field, externaltype):
            """Save the externalid"""

            oErr = ErrHandle()
            try:
                externalid = form.cleaned_data.get(field)
                if not externalid is None and externalid != "":
                    # Check if this is the correct id
                    ext = instance.sermonexternals.filter(externaltype=externaltype).first()
                    if ext is None:
                        # Need to make one
                        ext = SermonDescrExternal.objects.create(sermon=instance, externaltype=externaltype, externalid=externalid)
                    else:
                        # Do we need changing?
                        if externalid != ext.externalid:
                            ext.externalid = externalid
                            ext.save()
            except:
                msg = oErr.get_error_message()
                oErr.DoError("adapt_external")
                bResult = False

        msg = ""
        bResult = True
        oErr = ErrHandle()
        method = "nodistance"   # Alternative: "superdist"
        
        try:
            # Process many-to-many changes: Add and remove relations in accordance with the new set passed on by the user
            if getattr(form, 'cleaned_data') != None:

                # External id changes
                adapt_external("ext_huwop", "huwop")
                adapt_external("ext_huwin", "huwin")

                # (1) 'keywords'
                kwlist = form.cleaned_data['kwlist']
                adapt_m2m(SermonDescrKeyword, instance, "sermon", kwlist, "keyword")
            
                # (2) user-specific 'keywords'
                ukwlist = form.cleaned_data['ukwlist']
                profile = Profile.get_user_profile(self.request.user.username)
                adapt_m2m(UserKeyword, instance, "sermo", ukwlist, "keyword", qfilter = {'profile': profile}, extrargs = {'profile': profile, 'type': 'sermo'})

                ## (3) 'Links to Sermon (not 'Gold') Signatures'
                #siglist = form.cleaned_data['siglist']
                #adapt_m2m(SermonSignature, instance, "sermon", siglist, "gsig", extra = ['editype', 'code'])

                # (4) 'Links to Gold Sermons'
                superlist = form.cleaned_data['superlist']
                adapt_m2m(SermonDescrEqual, instance, "sermon", superlist, "super", extra = ['linktype'], related_is_through=True)

                # (5) 'collections'
                collist_s = form.cleaned_data['collist_s']
                adapt_m2m(CollectionSerm, instance, "sermon", collist_s, "collection")

                # (6) Related sermon manifestation links...
                slinklist = form.cleaned_data['slinklist']
                sermo_added = []
                sermo_deleted = []
                adapt_m2m(SermonDescrLink, instance, "src", slinklist, "dst", 
                          extra = ['linktype', 'note'], related_is_through=True,
                          added=sermo_added, deleted=sermo_deleted)
                # Check for partial links in 'deleted'
                for obj in sermo_deleted:
                    # This if-clause is not needed: anything that needs deletion should be deleted
                    # if obj.linktype in LINK_BIDIR:
                    # First find and remove the other link
                    reverse = SermonDescrLink.objects.filter(src=obj.dst, dst=obj.src, linktype=obj.linktype).first()
                    if reverse != None:
                        reverse.delete()
                    # Then remove myself
                    obj.delete()
                # Make sure to add the reverse link in the bidirectionals
                for obj in sermo_added:
                    if obj.linktype in LINK_BIDIR_MANU:
                        # Find the reversal
                        reverse = SermonDescrLink.objects.filter(src=obj.dst, dst=obj.src, linktype=obj.linktype).first()
                        if reverse == None:
                            # Create the reversal 
                            reverse = SermonDescrLink.objects.create(src=obj.dst, dst=obj.src, linktype=obj.linktype)
                            # Other adaptations
                            bNeedSaving = False
                            # Possibly copy note
                            if obj.note != None and obj.note != "":
                              reverse.note = obj.note
                              bNeedSaving = True
                            # Need saving? Then save
                            if bNeedSaving:
                              reverse.save()

                # (7) 'projects'
                projlist = form.cleaned_data['projlist']
                sermo_proj_deleted = []
                adapt_m2m(SermonDescrProject, instance, "sermon", projlist, "project", deleted=sermo_proj_deleted)
                project_dependant_delete(self.request, sermo_proj_deleted)

                # When sermons have been added to the manuscript, the sermons need to be updated 
                # with the existing project names 
                # Issue #412: do *NOT* do automatic adjustment to other sermons or to manuscript
                # instance.adapt_projects() # Gaat direct naar adapt_projects in SermDescr

                # Issue #412: when a sermon doesn't yet have a project, it gets the project of the manuscript
                if instance.projects.count() == 0:
                    manu = instance.msitem.manu
                    # How many projects are attached to this manuscript
                    manu_project_count = manu.projects.count()
                    if manu_project_count == 1:
                        project = manu.projects.first()
                        SermonDescrProject.objects.create(sermon=instance, project=project)

                # Process many-to-ONE changes
                # (1) links from bibrange to sermon
                bibreflist = form.cleaned_data['bibreflist']
                adapt_m2o(BibRange, instance, "sermon", bibreflist)

                # (2) 'sermonsignatures'
                siglist_m = form.cleaned_data['siglist_m']
                adapt_m2o(SermonSignature, instance, "sermon", siglist_m)

                # Possibly read transcription if this is 'new'
                if 'transcription' in form.changed_data:
                    # Try to read the updated transcription
                    oTranscription = read_transcription(instance.transcription)
                    # Are we okay?
                    sStatus = oTranscription.get("status", "")
                    sText = oTranscription.get("text", "")
                    iWordcount = oTranscription.get("wordcount", 0)
                    oFullInfo = dict(wordcount=iWordcount)
                    if sStatus == "ok" and sText != "":
                        instance.fullinfo = json.dumps(oFullInfo)
                        instance.fulltext = sText
                        instance.save()

            ## Make sure the 'verses' field is adapted, if needed
            #bResult, msg = instance.adapt_verses()

            # Check if instances need re-calculation
            if method == "superdist":
                if 'incipit' in form.changed_data or 'explicit' in form.changed_data:
                    instance.do_distance(True)

        except:
            msg = oErr.get_error_message()
            bResult = False
        return bResult, msg

    def action_add(self, instance, details, actiontype):
        """User can fill this in to his/her liking"""
        passim_action_add(self, instance, details, actiontype)

    def get_history(self, instance):
        return passim_get_history(instance)


class SermonDetails(SermonEdit):
    """The details of one sermon manifestation (SermonDescr)"""

    rtype = "html"

    def add_to_context(self, context, instance):
        """Add to the existing context"""

        # Start by executing the standard handling
        context = super(SermonDetails, self).add_to_context(context, instance)

        oErr = ErrHandle()

        try:
            # Are we copying information?? (only allowed if we are the app_editor)
            if 'supercopy' in self.qd and context['is_app_editor']:
                # Get the ID of the SSG from which information is to be copied to the S
                superid = self.qd['supercopy']
                # Get the SSG instance
                equal = EqualGold.objects.filter(id=superid).first()

                if equal != None:
                    # Copy all relevant information to the EqualGold obj (which is a SSG)
                    obj = self.object
                    # (1) copy author
                    if equal.author != None: obj.author = equal.author
                    # (2) copy incipit
                    if equal.incipit != None and equal.incipit != "": obj.incipit = equal.incipit ; obj.srchincipit = equal.srchincipit
                    # (3) copy explicit
                    if equal.explicit != None and equal.explicit != "": obj.explicit = equal.explicit ; obj.srchexplicit = equal.srchexplicit

                    # Now save the adapted EqualGold obj
                    obj.save()

                    # Mark these changes, which are done outside the normal 'form' system
                    actiontype = "save"
                    changes = dict(author=obj.author.id, incipit=obj.incipit, explicit=obj.explicit)
                    details = dict(savetype="change", id=obj.id, changes=changes)
                    passim_action_add(self, obj, details, actiontype)

                # And in all cases: make sure we redirect to the 'clean' GET page
                self.redirectpage = reverse('sermon_details', kwargs={'pk': self.object.id})
            else:
                context['sections'] = []

                # List of post-load objects
                context['postload_objects'] = []

                # Lists of related objects
                context['related_objects'] = []
        except:
            msg = oErr.get_error_message()
            oErr.DoError("SermonDetails/add_to_context")

        # Return the context we have made
        return context

    def before_save(self, form, instance):
        # If needed, create an MsItem
        if instance.msitem == None:
            # Make sure we have the manuscript
            manu = form.cleaned_data.get("manu", None)
            msitem = MsItem.objects.create(manu=manu)
            # Now make sure to set the link from Manuscript to MsItem
            instance.msitem = msitem

            # If possible, also get the mtype
            mtype = self.qd.get("sermo-mtype", None)
            if mtype != None:
                instance.mtype = mtype

        # Double check for the presence of manu and order
        if instance.msitem and instance.msitem.order < 0:
            # Calculate how many MSITEMS (!) there are
            msitem_count = instance.msitem.manu.manuitems.all().count()
            # Adapt the MsItem order
            msitem.order = msitem_count
            msitem.save()
            # Find out which is the one PRECEDING me (if any) at the HIGHEST level
            prec_list = instance.msitem.manu.manuitems.filter(parent__isnull=True, order__gt=msitem.order)
            if prec_list.count() > 0:
                # Get the last item
                prec_item = prec_list.last()
                # Set the 'Next' here correctly
                prec_item.next = msitem
                prec_item.save()

        return True, ""

    def process_formset(self, prefix, request, formset):
        return None

    def after_save(self, form, instance):
        return True, ""


class SermonTransDel(SermonDetails):
    """Remove the fulltrans from this sermon"""

    initRedirect = True

    def custom_init(self, instance):
        oErr = ErrHandle()
        try:
            # Check ... something
            if not instance is None:
                # Make sure to set the correct redirect page
                self.redirectpage = reverse("sermon_details", kwargs={'pk': instance.id})
                # Remove the transcription
                instance.srchfulltext = None
                instance.fulltext = None
                instance.transcription = None
                instance.save()
                # Now it's all been deleted
        except:
            msg = oErr.get_error_message()
            oErr.DoError("SermonTransDel/custom_init")
        return None


class SermonUserKeyword(SermonDetails):
    """This version only looks at one kind of data: the ukwlist"""

    newRedirect = True

    def custom_init(self, instance):
        oErr = ErrHandle()
        try:
            # Make sure to set the correct redirect page
            if instance:
                self.redirectpage = reverse("sermon_details", kwargs={'pk': instance.id})
            # Make sure we are not saving
            self.do_not_save = True

            # Get the value
            qd = self.qd
            if self.prefix_type == "simple":
                sUkwList = "{}-ukwlist".format(self.prefix)
            else:
                sUkwList = "{}-{}-ukwlist".format(self.prefix, instance.id)
            ukwlist_ids = qd.getlist(sUkwList)
            ukwlist = Keyword.objects.filter(id__in=ukwlist_ids)
            
            # Process the contents of the ukwlist, the chosen user-keywords
            profile = Profile.get_user_profile(self.request.user.username)
            adapt_m2m(UserKeyword, instance, "sermo", ukwlist, "keyword", qfilter = {'profile': profile}, extrargs = {'profile': profile, 'type': 'sermo'})

        except:
            msg = oErr.get_error_message()
            oErr.DoError("SermonUserKeyword/custom_init")
    

class SermonMove(SermonDetails):
    # newRedirect = True

    def custom_init(self, instance):
        errHandle = ErrHandle()

        # Note: use [errHandle]
        try:
            # Find the right parameter
            manu_id = self.qd.get("sermo-manuone")
            if not manu_id is None and manu_id != "":
                dst_manu = Manuscript.objects.filter(id=manu_id).first()
                if not dst_manu is None:
                    # Yes, a destination manuscript has been identified
                    # Find out what the last codico-unit is
                    lastcodico = dst_manu.manuscriptcodicounits.all().order_by('-order').first()
                    if not lastcodico is None:
                        # Get the msitem of the sermon
                        msitem = instance.msitem
                        # Move the MsItem to a different Codico
                        msitem.move_codico(lastcodico)

                    # Make sure to set the correct redirect page: the manuscript to which it has been moved
                    if instance:
                        self.redirectpage = reverse("manuscript_details", kwargs={'pk': instance.id})
            # Make sure we are not saving
            self.do_not_save = True

            return True
        except:
            msg = errHandle.get_error_message()
            errHandle.DoError("SermonMove")
            return False


class SermonListView(BasicList):
    """Search and list sermons"""
    
    model = SermonDescr
    listform = SermonForm
    has_select2 = True
    use_team_group = True
    paginate_by = 20
    bUseFilter = True
    prefix = "sermo"
    sel_button = "serm"
    new_button = False      # Don't show the [Add new sermon] button here. It is shown under the Manuscript Details view.
    basketview = False
    plural_name = "Manifestations"
    basic_name = "sermon"    
    sg_name = "Manifestation"

    order_cols = ['author__name;nickname__name', 'siglist', 'srchincipit;srchexplicit', 'manu__idno', 
                  'msitem__codico__manuscript__yearstart;msitem__codico__manuscript__yearfinish', 
                  'srchsectiontitle', 'srchtitle', '','','', 'stype']
    order_default = order_cols
    order_heads = [
        {'name': 'Attr. author', 'order': 'o=1', 'type': 'str', 'custom': 'author', 'linkdetails': True,
         'title': 'Attributed author'}, 
        {'name': 'Gryson/Clavis', 'order': 'o=2', 'type': 'str', 'custom': 'signature', 'allowwrap': True, 'options': '111111',
         'title': 'Gryson/Clavis codes of the Sermons Gold that are part of the same equality set + those manually linked to this manifestation Sermon'}, 
        {'name': 'Incipit ... Explicit', 
                                'order': 'o=3', 'type': 'str',  'custom': 'incexpl', 'main': True, 'linkdetails': True},
        {'name': 'Manuscript',  'order': 'o=4', 'type': 'str',  'custom': 'manuscript'},
        {'name': 'Ms Date',     'order': 'o=5', 'type': 'str',  'custom': 'msdate',
         'title': 'Date of the manuscript'},
        {'name': 'Section',     'order': 'o=6', 'type': 'str',  'custom': 'sectiontitle', 
         'allowwrap': True,    'autohide': "on", 'filter': 'filter_sectiontitle'},
        {'name': 'Title',       'order': 'o=7', 'type': 'str',  'custom': 'title', 
         'allowwrap': True,           'autohide': "on", 'filter': 'filter_title'},        
        {'name': 'Locus',       'order': '',    'type': 'str',  'field':  'locus' },
        {'name': '',            'order': '',    'type': 'str',  'custom':  'saved' },
        {'name': 'Links',       'order': '',    'type': 'str',  'custom': 'links'},
        {'name': 'Status',      'order': 'o=11', 'type': 'str', 'custom': 'status'}]

    filters = [ {"name": "Gryson/Clavis/Other code",    "id": "filter_signature",      "enabled": False},
                {"name": "Attr. author",     "id": "filter_author",         "enabled": False},
                {"name": "Author type",      "id": "filter_autype",         "enabled": False},
                {"name": "Incipit",          "id": "filter_incipit",        "enabled": False},
                {"name": "Explicit",         "id": "filter_explicit",       "enabled": False},                
                {"name": "Section titles",   "id": "filter_sectiontitle",   "enabled": False},
                {"name": "Title",            "id": "filter_title",          "enabled": False},
                {"name": "Keyword",          "id": "filter_keyword",        "enabled": False}, 
                {"name": "Feast",            "id": "filter_feast",          "enabled": False},
                {"name": "Bible reference",  "id": "filter_bibref",         "enabled": False},
                {"name": "Note",             "id": "filter_note",           "enabled": False},
                {"name": "Status",           "id": "filter_stype",          "enabled": False},
                {"name": "Passim code",      "id": "filter_code",           "enabled": False},
                {"name": "Free",             "id": "filter_freetext",       "enabled": False},
                {"name": "Project",          "id": "filter_project",        "enabled": False},
                {"name": "Collection...",    "id": "filter_collection",     "enabled": False, "head_id": "none"},
                {"name": "Manuscript...",    "id": "filter_manuscript",     "enabled": False, "head_id": "none"},
                {"name": "Sermon",           "id": "filter_collsermo",      "enabled": False, "head_id": "filter_collection"},
                # Issue #416: Delete the option to search for a GoldSermon personal dataset
                # {"name": "Sermon Gold",      "id": "filter_collgold",       "enabled": False, "head_id": "filter_collection"},
                {"name": "Authority file",   "id": "filter_collsuper",      "enabled": False, "head_id": "filter_collection"},
                {"name": "Manuscript",       "id": "filter_collmanu",       "enabled": False, "head_id": "filter_collection"},
                {"name": "Historical",       "id": "filter_collhc",         "enabled": False, "head_id": "filter_collection"},
                {"name": "Shelfmark",        "id": "filter_manuid",         "enabled": False, "head_id": "filter_manuscript"},
                {"name": "Country",          "id": "filter_country",        "enabled": False, "head_id": "filter_manuscript"},
                {"name": "City/Location",    "id": "filter_city",           "enabled": False, "head_id": "filter_manuscript"},
                {"name": "Library",          "id": "filter_library",        "enabled": False, "head_id": "filter_manuscript"},
                {"name": "Origin",           "id": "filter_origin",         "enabled": False, "head_id": "filter_manuscript"},
                {"name": "Provenance",       "id": "filter_provenance",     "enabled": False, "head_id": "filter_manuscript"},
                {"name": "Date range",       "id": "filter_daterange",      "enabled": False, "head_id": "filter_manuscript"},
                # {"name": "Date until",       "id": "filter_datefinish",     "enabled": False, "head_id": "filter_manuscript"},
                {"name": "Manuscript type",  "id": "filter_manutype",       "enabled": False, "head_id": "filter_manuscript"},
                ]
    
    searches = [
        {'section': '', 'filterlist': [
            {'filter': 'incipit',       'dbfield': 'srchincipit',       'keyS': 'incipit',  'regex': adapt_regex_incexp},
            {'filter': 'explicit',      'dbfield': 'srchexplicit',      'keyS': 'explicit', 'regex': adapt_regex_incexp},
            {'filter': 'title',         'dbfield': 'srchtitle',         'keyS': 'srch_title'},
            {'filter': 'sectiontitle',  'dbfield': 'srchsectiontitle',  'keyS': 'srch_sectiontitle'},
            {'filter': 'feast',         'fkfield': 'feast',             'keyFk': 'feast', 'keyList': 'feastlist', 'infield': 'id'},
            {'filter': 'note',          'dbfield': 'note',              'keyS': 'note'},
            {'filter': 'bibref',        'dbfield': '$dummy',            'keyS': 'bibrefbk'},
            {'filter': 'bibref',        'dbfield': '$dummy',            'keyS': 'bibrefchvs'},
            {'filter': 'freetext',      'dbfield': '$dummy',            'keyS': 'free_term'},
            {'filter': 'freetext',      'dbfield': '$dummy',            'keyS': 'free_include'},
            {'filter': 'freetext',      'dbfield': '$dummy',            'keyS': 'free_exclude'},
            
            {'filter': 'code',          'fkfield': 'sermondescr_super__super', 'keyS': 'passimcode', 'keyFk': 'code', 'keyList': 'passimlist', 'infield': 'id', 'help': 'passimcode'},
            
            {'filter': 'author',        'fkfield': 'author',            'keyS': 'authorname',
                                        'keyFk': 'name', 'keyList': 'authorlist', 'infield': 'id', 'external': 'sermo-authorname' },
            {'filter': 'autype',                                        'keyS': 'authortype',  'help': 'authorhelp'},
            {'filter': 'signature',     'fkfield': 'signatures|equalgolds__equal_goldsermons__goldsignatures',      'help': 'signature',     
                                        'keyS': 'signature_a', 'keyFk': 'code', 'keyId': 'signatureid', 'keyList': 'siglist_a', 'infield': 'code' },
            #{'filter': 'signature',     'fkfield': 'signatures|goldsermons__goldsignatures', 'help': 'signature',     
            #                            'keyS': 'signature', 'keyFk': 'code', 'keyId': 'signatureid', 'keyList': 'siglist', 'infield': 'code' },
            {'filter': 'keyword',       'fkfield': 'keywords',          'keyFk': 'name', 'keyList': 'kwlist', 'infield': 'id' }, 
            {'filter': 'project',       'fkfield': 'projects',          'keyFk': 'name', 'keyList': 'projectslist', 'infield': 'name'},
            {'filter': 'stype',         'dbfield': 'stype',             'keyList': 'stypelist', 'keyType': 'fieldchoice', 'infield': 'abbr' }
            ]},
        {'section': 'collection', 'filterlist': [
            {'filter': 'collmanu',      'fkfield': 'manu__collections',              'keyFk': 'name', 'keyList': 'collist_m',  'infield': 'id' }, 
            {'filter': 'collsermo',     'fkfield': 'collections',                    'keyFk': 'name', 'keyList': 'collist_s',  'infield': 'id' }, 
            # Issue #416: Delete the option to search for a GoldSermon personal dataset
            # {'filter': 'collgold',      'fkfield': 'goldsermons__collections',       'keyFk': 'name', 'keyList': 'collist_sg', 'infield': 'id' }, 
            {'filter': 'collsuper',     'fkfield': 'equalgolds__collections',        'keyFk': 'name', 'keyList': 'collist_ssg','infield': 'id' }, 
            {'filter': 'collhc',        'fkfield': 'equalgolds__collections',        'keyFk': 'name', 'keyList': 'collist_hist', 'infield': 'id' }
            # {'filter': 'collsuper',     'fkfield': 'goldsermons__equal__collections', 'keyFk': 'name', 'keyList': 'collist_ssg','infield': 'id' }
            ]},
        {'section': 'manuscript', 'filterlist': [
            {'filter': 'manuid',        'fkfield': 'manu',  'keyS': 'manuidno',     'keyList': 'manuidlist', 'keyFk': 'idno', 'infield': 'id'},
            {'filter': 'country',       'fkfield': 'msitem__manu__library__lcountry', 'keyS': 'country_ta',   'keyId': 'country',     'keyFk': "name"},
            {'filter': 'city',          'fkfield': 'msitem__manu__library__lcity',    'keyS': 'city_ta',      'keyId': 'city',        'keyFk': "name"},
            {'filter': 'library',       'fkfield': 'msitem__manu__library',           'keyS': 'libname_ta',   'keyId': 'library',     'keyFk': "name"},
            {'filter': 'origin',        'fkfield': 'msitem__codico__origin',          'keyS': 'origin_ta',    'keyId': 'origin',      'keyFk': "name"},
            {'filter': 'provenance',    'fkfield': 'msitem__codico__provenances|msitem__codico__provenances__location',     
             'keyS': 'prov_ta',      'keyId': 'prov',        'keyFk': "name"},
            {'filter': 'daterange',     'dbfield': 'msitem__codico__codico_dateranges__yearstart__gte',     'keyS': 'date_from'},
            {'filter': 'daterange',     'dbfield': 'msitem__codico__codico_dateranges__yearfinish__lte',    'keyS': 'date_until'},
            {'filter': 'manutype',      'dbfield': 'msitem__manu__mtype',     'keyS': 'manutype',     'keyType': 'fieldchoice', 'infield': 'abbr'},
            ]},
        {'section': 'other', 'filterlist': [
            {'filter': 'mtype',     'dbfield': 'mtype',    'keyS': 'mtype'},
            {'filter': 'sigauto',   'fkfield': 'equalgolds__equal_goldsermons__goldsignatures', 'keyList':  'siglist_a', 'infield': 'id'},
            {'filter': 'sigmanu',   'fkfield': 'sermonsignatures',                              'keyList':  'siglist_m', 'infield': 'id'},
            {'filter': 'atype',     'dbfield': 'sermondescr_super__super__atype',    'keyS': 'atype'}
            #{'filter': 'appr_type', 'fkfield': 'equalgolds__', 'keyList':' ', 'infield': }
            ]}
         ]

    selectbuttons = [
        {'title': 'Add to saved items', 'mode': 'add_saveitem', 'button': 'jumbo-1', 'glyphicon': 'glyphicon-star-empty'},
        {'title': 'Add to basket',      'mode': 'add_basket',   'button': 'jumbo-1', 'glyphicon': 'glyphicon-shopping-cart'},
        # {'title': 'Add to DCT',         'mode': 'add_dct',      'button': 'jumbo-1', 'glyphicon': 'glyphicon-wrench'},
        ]

    def initializations(self):
        oErr = ErrHandle()
        try:
            # ======== One-time adaptations ==============
            listview_adaptations("sermon_list")

            # Check if there are any sermons not connected to a manuscript: remove these
            delete_id = SermonDescr.objects.filter(Q(msitem__isnull=True)|Q(msitem__manu__isnull=True)).values('id')
            if len(delete_id) > 0:
                oErr.Status("Deleting {} sermons that are not connected".format(len(delete_id)))
                SermonDescr.objects.filter(id__in=delete_id).delete()

            # Make sure to set a basic filter
            self.basic_filter = Q(mtype="man")

            # Make the profile available
            self.profile = Profile.get_user_profile(self.request.user.username)

        except:
            msg = oErr.get_error_message()
            oErr.DoError("SermonListiew/initializations")
        return None

    def add_to_context(self, context, initial):
        # Find out who the user is
        profile = Profile.get_user_profile(self.request.user.username)
        context['basketsize'] = 0 if profile == None else profile.basketsize
        context['basket_show'] = reverse('basket_show')
        context['basket_update'] = reverse('basket_update')

        # Count the number of selected items
        iCount = SelectItem.get_selectcount(profile, "serm")
        if iCount > 0:
            context['sel_count'] = str(iCount)

        return context

    def get_basketqueryset(self):
        if self.basketview:
            profile = Profile.get_user_profile(self.request.user.username)
            qs = profile.basketitems.all()
        else:
            qs = SermonDescr.objects.all()
        return qs

    def get_field_value(self, instance, custom):
        sBack = ""
        sTitle = ""
        html = []
        if custom == "author":
            if instance.author:
                html.append("<span style='color: darkgreen; font-size: small;'>{}</span>".format(instance.author.name[:20]))
                sTitle = instance.author.name
            elif instance.nickname:
                html.append("<span style='color: darkred; font-size: small;'>{}</span>".format(instance.nickname.name[:20]))
                sTitle = instance.nickname.name
            else:
                html.append("<span><i>(unknown)</i></span>")
        elif custom == "signature":
            html.append(instance.get_eqsetsignatures_markdown('combi'))
        elif custom == "incexpl":
            html.append("<span>{}</span>".format(instance.get_incipit_markdown()))
            dots = "..." if instance.incipit else ""
            html.append("<span style='color: blue;'>{}</span>".format(dots))
            html.append("<span>{}</span>".format(instance.get_explicit_markdown()))
        elif custom == "manuscript":
            manu = instance.get_manuscript()
            if manu == None:
                html.append("-")
            else:
                if manu.idno == None:
                    value = "-"
                else:
                    # issue #610
                    # value = manu.idno[:20]
                    value = manu.get_full_name()
                html.append("<a href='{}' class='nostyle'><span style='font-size: small;'>{}</span></a>".format(
                    reverse('manuscript_details', kwargs={'pk': manu.id}),
                    value))
                sTitle = manu.idno
        elif custom == "saved":
            # Prepare saveditem handling
            saveditem_button = get_saveditem_html(self.request, instance, self.profile, sitemtype="serm")
            saveditem_form = get_saveditem_html(self.request, instance, self.profile, "form", sitemtype="serm")
            html.append(saveditem_button)
            html.append(saveditem_form)
        elif custom == "msdate":
            manu = instance.get_manuscript()
            # Get the yearstart-yearfinish
            html.append(manu.get_dates(True))
        elif custom == "title":
            sTitle = ""
            if instance.title != None and instance.title != "":
                sTitle = instance.title
            html.append(sTitle)
        elif custom == "sectiontitle":
            sSection = ""
            if instance.sectiontitle != None and instance.sectiontitle != "":
                sSection = instance.sectiontitle
            html.append(sSection)
        elif custom == "links":
            for gold in instance.goldsermons.all():
                for link_def in gold.link_oview():
                    if link_def['count'] > 0:
                        html.append("<span class='badge {}' title='{}'>{}</span>".format(link_def['class'], link_def['title'], link_def['count']))
        elif custom == "status":
            # Provide that status badge
            # html.append("<span class='badge' title='{}'>{}</span>".format(instance.get_stype_light(), instance.stype[:1]))
            html.append(instance.get_stype_light())


            
        # Combine the HTML code
        sBack = "\n".join(html)
        return sBack, sTitle

    def adapt_search(self, fields):
        # Adapt the search to the keywords that *may* be shown
        lstExclude=[]
        qAlternative = None
        oErr = ErrHandle()

        try:
            # Make sure we show MANUSCRIPTS (identifiers) as well as reconstructions
            lstExclude.append(Q(mtype='tem'))
            lstExclude.append(Q(msitem__isnull=True))
            ## Make sure to only show mtype manifestations
            #fields['mtype'] = "man"

            manutype = fields.get('manutype')
            if manutype != None and manutype != "":
                if manutype.abbr == "rec":
                    # Restrict to sermons that are part of a codico that is in table Reconstruction
                    codicolist = [x.codico.id for x in Reconstruction.objects.all()]
                    fields['manutype'] = Q(msitem__codico__id__in=codicolist)

            # Check if a list of keywords is given
            if 'kwlist' in fields and fields['kwlist'] != None and len(fields['kwlist']) > 0:
                # Get the list
                kwlist = fields['kwlist']   # NOTE: the kwlist does not contain id's, but actual items
                # Get the user
                username = self.request.user.username
                user = User.objects.filter(username=username).first()
                # Check on what kind of user I am
                if not user_is_ingroup(self.request, app_editor):
                    # Since I am not an app-editor, I may not filter on keywords that have visibility 'edi'
                    kwlist = Keyword.objects.filter(id__in=kwlist).exclude(Q(visibility="edi")) # Thanks to LILAC, strip: .values('id')
                    fields['kwlist'] = kwlist
                # Create Q expression that includes related ones
                kwlist = ( Q(keywords__in=kwlist) | Q(equalgolds__keywords__in=kwlist) )
                fields['kwlist'] = kwlist
            
            # Check if a list of projects is given
            if 'projlist' in fields and fields['projlist'] != None and len(fields['projlist']) > 0:
                # Get the list
                projlist = fields['projlist']

            # Adapt the bible reference list
            bibrefbk = fields.get("bibrefbk", "")
            if bibrefbk != None and bibrefbk != "":
                bibrefchvs = fields.get("bibrefchvs", "")


                # Get the start and end of this bibref
                start, einde = Reference.get_startend(bibrefchvs, book=bibrefbk)
 
                # Find out which sermons have references in this range
                lstQ = []
                lstQ.append(Q(sermonbibranges__bibrangeverses__bkchvs__gte=start))
                lstQ.append(Q(sermonbibranges__bibrangeverses__bkchvs__lte=einde))
                sermonlist = [x.id for x in SermonDescr.objects.filter(*lstQ).order_by('id').distinct()]

                fields['bibrefbk'] = Q(id__in=sermonlist)

            # Adapt the search for empty authors
            if 'authortype' in fields:
                authortype = fields['authortype']
                if authortype == "non":
                    # lstExclude = []
                    lstExclude.append(Q(author__isnull=False))
                elif authortype == "spe":
                    # lstExclude = []
                    lstExclude.append(Q(author__isnull=True))
                else:
                    # Reset the authortype
                    fields['authortype'] = ""

            # Process the date range stuff
            date_from = fields.get("date_from")
            date_until = fields.get("date_until")
            if not date_from is None and not date_until is None:
                # Both dates are specified: include looking for overlap
                qThis_a = Q(msitem__codico__codico_dateranges__yearstart__gte=date_from)
                qThis_b = Q(msitem__codico__codico_dateranges__yearstart__lte=date_until)
                qThis_c = Q(msitem__codico__codico_dateranges__yearfinish__lte=date_until)
                qThis_d = Q(msitem__codico__codico_dateranges__yearfinish__gte=date_from)
                qThis_1 =  ( qThis_a &  ( qThis_b |  qThis_c ) )
                qThis_2 =  ( qThis_c &  ( qThis_d |  qThis_a ) )
                qThis = ( qThis_1 | qThis_2 )
                fields['date_from'] = qThis
                fields['date_until'] = ""


            # Adapt according to the 'free' fields
            free_term = fields.get("free_term", "")
            if free_term != None and free_term != "":
                all_fields = "all_fields"

                free_include = fields.get("free_include", [])
                free_exclude = fields.get("free_exclude", [])

                # Look for include field(s)
                if len(free_include) == 1 and free_include[0].field == all_fields:
                    # Include all fields from Free
                    s_q_i_lst = ""
                    for obj in Free.objects.exclude(field=all_fields):
                        val = free_term
                        if "*" in val or "#" in val:
                            val = adapt_search(val)
                            s_q = Q(**{"{}__iregex".format(obj.field): val})
                        elif "@" in val:
                            val = val.replace("@", "").strip()
                            s_q = Q(**{"{}__icontains".format(obj.field): val})
                        else:
                            s_q = Q(**{"{}__iexact".format(obj.field): val})
                        if s_q_i_lst == "":
                            s_q_i_lst = s_q
                        else:
                            s_q_i_lst |= s_q
                else:
                    s_q_i_lst = ""
                    for obj in free_include:
                        if obj.field == all_fields:
                            # skip it
                            pass
                        else:
                            val = free_term
                            if "*" in val or "#" in val:
                                val = adapt_search(val)
                                s_q = Q(**{"{}__iregex".format(obj.field): val})
                            elif "@" in val:
                                val = val.replace("@", "").strip()
                                s_q = Q(**{"{}__icontains".format(obj.field): val})
                            else:
                                s_q = Q(**{"{}__iexact".format(obj.field): val})
                            if s_q_i_lst == "":
                                s_q_i_lst = s_q
                            else:
                                s_q_i_lst |= s_q

                # Look for exclude fields
                if len(free_exclude) == 1 and free_exclude[0].field == all_fields:
                    # Include all fields from Free
                    s_q_e_lst = ""
                    for obj in Free.objects.exclude(field=all_fields):
                        val = free_term
                        if "*" in val or "#" in val:
                            val = adapt_search(val)
                            s_q = Q(**{"{}__iregex".format(obj.field): val})
                        elif "@" in val:
                            val = val.replace("@", "").strip()
                            s_q = Q(**{"{}__icontains".format(obj.field): val})
                        else:
                            s_q = Q(**{"{}__iexact".format(obj.field): val})
                        if s_q_e_lst == "":
                            s_q_e_lst = s_q
                        else:
                            s_q_e_lst |= s_q
                else:
                    s_q_e_lst = ""
                    for obj in free_exclude:
                        if obj.field == all_fields:
                            # skip it
                            pass
                        else:
                            val = free_term
                            if "*" in val or "#" in val:
                                val = adapt_search(val)
                                s_q = Q(**{"{}__iregex".format(obj.field): val})
                            elif "@" in val:
                                val = val.replace("@", "").strip()
                                s_q = Q(**{"{}__icontains".format(obj.field): val})
                            else:
                                s_q = Q(**{"{}__iexact".format(obj.field): val})
                            if s_q_e_lst == "":
                                s_q_e_lst = s_q
                            else:
                                s_q_e_lst |= s_q

                if s_q_i_lst != "":
                    qAlternative = s_q_i_lst
                if s_q_e_lst != "":
                    lstExclude.append( s_q_e_lst )

                # CLear the fields
                fields['free_term'] = "yes"
                fields['free_include'] = ""
                fields['free_exclude'] = ""
            # Double check the length of the exclude list
            if len(lstExclude) == 0:
                lstExclude = None
        except:
            msg = oErr.get_error_message()
            oErr.DoError("SermonListView/adapt_search")
        
        # Make sure we only use the Authority Files with accepted modifications
        # This means that atype should be 'acc' (and not: 'mod', 'rej' or 'def') 
        # With this condition we make sure ALL sermons are in de unfiltered listview
        if fields['passimcode'] != '':
            fields['atype'] = 'acc'

        return fields, lstExclude, qAlternative
    
    def view_queryset(self, qs):
        search_id = [x['id'] for x in qs.values('id')]
        profile = Profile.get_user_profile(self.request.user.username)
        profile.search_sermo = json.dumps(search_id)
        profile.save()
        return None

    def get_helptext(self, name):
        """Use the get_helptext function defined in models.py"""
        return get_helptext(name)

    def get_selectitem_info(self, instance, context):
        """Use the get_selectitem_info() defined earlier in this views.py"""
        return get_selectitem_info(self.request, instance, self.profile, self.sel_button, context)


# ============= ONLINESOURCE ==============================


class OnlineSourceEdit(BasicDetails):
    """The details of one online source"""

    model = OnlineSources
    mForm = OnlineSourceForm
    prefix = 'onsc'
    title = "OnlineSourceEdit"
    rtype = "json"
    history_button = True
    mainitems = []
    
    def add_to_context(self, context, instance):
        """Add to the existing context"""

        # Define the main items to show and edit
        context['mainitems'] = [
            {'type': 'plain', 'label': "Name: ",    'value': instance.name, 'field_key': 'name'},     
            {'type': 'plain', 'label': "URL:",  'value': instance.get_onlinesource_url(),  'field_key': 'url'} # --> instance.get_onlinesource_markdown()
            ]
        # Return the context we have made
        return context

    # get_external_markdown
    def action_add(self, instance, details, actiontype):
        """User can fill this in to his/her liking"""
        passim_action_add(self, instance, details, actiontype)

    def get_history(self, instance):
        return passim_get_history(instance)

    def after_save(self, form, instance):
        msg = ""
        bResult = True
        oErr = ErrHandle()
        
        try:
            name = instance.name            
            name_dec = unidecode(name)
            # Store the new version
            instance.sortname = name_dec
            instance.save()
        
        except:
            msg = oErr.get_error_message()
            bResult = False
        return bResult, msg


class OnlineSourceDetails(OnlineSourceEdit):
    """Like OnlineSource Edit, but then html output"""
    rtype = "html"


class OnlineSourceListView(BasicList):
    """Search and list projects"""

    model = OnlineSources
    listform = OnlineSourceForm
    prefix = "onsc"
    has_select2 = True
    paginate_by = 20
    sg_name = "Online source"     # This is the name as it appears e.g. in "Add a new XXX" (in the basic listview)
    plural_name = "Online sources"
    # page_function = "ru.passim.seeker.search_paged_start"
    order_cols = ['name', '']
    order_default = order_cols
    order_heads = [{'name': 'Name','order': 'o=1', 'type': 'str', 'custom': 'name', 'main': True, 'linkdetails': True}, #'field': 'name', 'main': True, 'default': "(unnamed)", 
                   {'name': 'URL', 'order': 'o=2', 'type': 'str', 'custom': 'url'}]
                   
    #filters = [ {"name": "Name",       "id": "filter_name",     "enabled": False},
    #            {"name": "URL",        "id": "filter_url",      "enabled": False},
    #           ]
    searches = [
        {'section': '', 'filterlist': [
            {'filter': 'name',   'dbfield': 'name',   'keyS': '_ta',  'infield': 'name' }, 
            {'filter': 'url',    'dbfield': 'url',    'keyFk': 'name', 'infield': 'url'}
            ]}
        ] 

    def initializations(self):
        # Make sure possible adaptations are executed
        listview_adaptations("onlinesources_list")
        # We are not returning anything        
        return None
     
    def get_field_value(self, instance, custom):
        sBack = ""
        sTitle = ""
        html = []
        oErr = ErrHandle()
        try:
            if custom == "url":
             #   html = []
                # Give the URL a new look, see if css can be modified  
                html.append("<span class='collection'><a href='{}'>{}</a></span>".format(instance.url, instance.url))
        
            elif custom == "name":
                sName = instance.sortname
                # Make sure that the records without name are given a name so that one can select the record.
                if sName == "": sName = "<i>(unnamed)</i>"
                html.append(sName)
        
        except:
            msg = oErr.get_error_message()
            oErr.DoError("OnlineSourcesListView/get_field_value")
        
        # Combine the HTML code
        sBack = "\n".join(html)
        return sBack, sTitle              


# ============= KEYWORD ====================================


class KeywordEdit(BasicDetails):
    """The details of one keyword"""

    model = Keyword
    mForm = KeywordForm
    prefix = 'kw'
    title = "KeywordEdit"
    rtype = "json"
    history_button = True
    mainitems = []
    
    def add_to_context(self, context, instance):
        """Add to the existing context"""

        # Define the main items to show and edit
        context['mainitems'] = [
            {'type': 'plain', 'label': "Name:",       'value': instance.name,                     'field_key': 'name'},
            {'type': 'plain', 'label': "Visibility:", 'value': instance.get_visibility_display(), 'field_key': 'visibility'},
            {'type': 'plain', 'label': "Category:",   'value': instance.get_category_display(),   'field_key': 'category'},
            {'type': 'plain', 'label': "Description:",'value': instance.description,              'field_key': 'description'}
            ]
        # Return the context we have made
        return context

    def action_add(self, instance, details, actiontype):
        """User can fill this in to his/her liking"""
        passim_action_add(self, instance, details, actiontype)

    def get_history(self, instance):
        return passim_get_history(instance)


class KeywordDetails(KeywordEdit):
    """Like Keyword Edit, but then html output"""
    rtype = "html"
    

class KeywordListView(BasicList):
    """Search and list keywords"""

    model = Keyword
    listform = KeywordForm
    prefix = "kw"
    paginate_by = 20
    # template_name = 'seeker/keyword_list.html'
    has_select2 = True
    in_team = False
    sg_name = "Keyword"     
    plural_name = "Keywords"
    page_function = "ru.passim.seeker.search_paged_start"
    order_cols = ['name', 'visibility', 'category', '']
    order_default = order_cols
    order_heads = [
        {'name': 'Keyword',    'order': 'o=1', 'type': 'str', 'field': 'name', 'default': "(unnamed)", 'main': True, 'linkdetails': True},
        {'name': 'Visibility', 'order': 'o=2', 'type': 'str', 'custom': 'visibility'},
        {'name': 'Category',   'order': 'o=3', 'type': 'str', 'custom': 'category'},
        {'name': 'Frequency',  'order': '',    'type': 'str', 'custom': 'links'}]
    filters = [ {"name": "Keyword",         "id": "filter_keyword",     "enabled": False},
                # See issue #628 -- keep the following commented for now 
                # {"name": "Visibility",      "id": "filter_visibility",  "enabled": False}
                ]
    searches = [
        {'section': '', 'filterlist': [
            {'filter': 'keyword',    'dbfield': 'name',         'keyS': 'keyword_ta', 'keyList': 'kwlist', 'infield': 'name' },
            {'filter': 'visibility', 'dbfield': 'visibility',   'keyS': 'visibility' }]}
        ]

    def initializations(self):
        # Make sure possible adaptations are executed
        listview_adaptations("keyword_list")

        # Check out who I am
        in_team = user_is_in_team(self.request)
        self.in_team = in_team
        if in_team:
            self.order_cols = ['name', 'visibility', 'category', '']
            self.order_default = self.order_cols
            self.order_heads = [
                {'name': 'Keyword',    'order': 'o=1', 'type': 'str', 'field': 'name', 'default': "(unnamed)", 'main': True, 'linkdetails': True},
                {'name': 'Visibility', 'order': 'o=2', 'type': 'str', 'custom': 'visibility'},
                {'name': 'Category',   'order': 'o=3', 'type': 'str', 'custom': 'category'},
                {'name': 'Frequency',  'order': '',    'type': 'str', 'custom': 'links'}]
            self.filters = [ 
                {"name": "Keyword",       "id": "filter_keyword",     "enabled": False},
                {"name": "Category",      "id": "filter_category",    "enabled": False}
                # See issue #628 -- keep the following commented for now 
                # {"name": "Visibility",      "id": "filter_visibility",  "enabled": False}
                ]
            self.searches = [
                {'section': '', 'filterlist': [
                    {'filter': 'keyword',    'dbfield': 'name',         'keyS': 'keyword_ta', 'keyList': 'kwlist', 'infield': 'name' },
                    # {'filter': 'category',   'dbfield': 'category',     'keyS': 'category',   'keyList': 'kwcatlist', 'infield': 'english_name'  },
                    {'filter': 'category',   'dbfield': 'category',     'keyList': 'kwcatlist', 'keyType': 'fieldchoice', 'infield': 'abbr'  },
                    {'filter': 'visibility', 'dbfield': 'visibility',   'keyS': 'visibility' }
                    ]}
                ]
            self.bUseFilter = False
        else:
            self.order_cols = ['name', 'category', '']
            self.order_default = self.order_cols
            self.order_heads = [
                {'name': 'Keyword',    'order': 'o=1', 'type': 'str', 'field': 'name', 'default': "(unnamed)", 'main': True, 'linkdetails': True},
                {'name': 'Category',   'order': 'o=2', 'type': 'str', 'custom': 'category'},
                {'name': 'Frequency', 'order': '', 'type': 'str', 'custom': 'links'}]
            self.filters = [ 
                {"name": "Keyword",       "id": "filter_keyword",     "enabled": False},
                {"name": "Category",      "id": "filter_category",    "enabled": False}
                ]
            self.searches = [
                {'section': '', 'filterlist': [
                    {'filter': 'keyword',    'dbfield': 'name',         'keyS': 'keyword_ta', 'keyList': 'kwlist', 'infield': 'name' },
                    {'filter': 'category',   'dbfield': 'category',     'keyList': 'kwcatlist', 'keyType': 'fieldchoice', 'infield': 'abbr'  },
                    ]},
                {'section': 'other', 'filterlist': [
                    {'filter': 'visibility', 'dbfield': 'visibility',   'keyS': 'visibility' }
                    ]}
                ]
            self.bUseFilter = True
        return None

    def get_field_value(self, instance, custom):
        sBack = ""
        sTitle = ""
        oErr = ErrHandle()
        try:
            if custom == "links":
                html = []
                # Get the HTML code for the links of this instance
                number = instance.freqsermo()
                if number > 0:
                    url = reverse('sermon_list')
                    html.append("<a href='{}?sermo-kwlist={}'>".format(url, instance.id))
                    html.append("<span class='badge jumbo-1 clickable' title='Frequency in manifestation sermons'>{}</span></a>".format(number))
                number = instance.freqgold()
                if number > 0:
                    url = reverse('search_gold')
                    html.append("<a href='{}?gold-kwlist={}'>".format(url, instance.id))
                    html.append("<span class='badge jumbo-2 clickable' title='Frequency in gold sermons'>{}</span></a>".format(number))
                number = instance.freqmanu()
                if number > 0:
                    url = reverse('search_manuscript')
                    html.append("<a href='{}?manu-kwlist={}'>".format(url, instance.id))
                    html.append("<span class='badge jumbo-3 clickable' title='Frequency in manuscripts'>{}</span></a>".format(number))
                number = instance.freqsuper()
                if number > 0:
                    url = reverse('equalgold_list')
                    html.append("<a href='{}?ssg-kwlist={}'>".format(url, instance.id))
                    html.append("<span class='badge jumbo-4 clickable' title='Frequency in Authority files'>{}</span></a>".format(number))
                # Combine the HTML code
                sBack = "\n".join(html)
            elif custom == "visibility":
                sBack = instance.get_visibility_display()
            elif custom == "category":
                sBack = instance.get_category_display()
        except:
            msg = oErr.get_error_message()
            oErr.DoError("KeywordListView/get_field_value")
        return sBack, sTitle

    def adapt_search(self, fields):
        lstExclude=None
        qAlternative = None
        # if not self.in_team:
        if not user_is_ingroup(self.request, app_editor):
            # restrict access to "all" marked ons
            fields['visibility'] = "all"

        return fields, lstExclude, qAlternative


# =============== USERKEYWORD ===============================
        

class UserKeywordEdit(BasicDetails):
    """The details of one 'user-keyword': one that has been linked by a user"""

    model = UserKeyword
    mForm = UserKeywordForm
    prefix = 'ukw'
    title = "UserKeywordEdit"
    rtype = "json"
    history_button = True
    mainitems = []
    
    def add_to_context(self, context, instance):
        """Add to the existing context"""

        # Define the main items to show and edit
        context['mainitems'] = [
            {'type': 'plain', 'label': "User:",     'value': instance.get_profile_markdown(),    },
            {'type': 'plain', 'label': "Keyword:",  'value': instance.keyword.name,     },
            {'type': 'plain', 'label': "Type:",     'value': instance.get_type_display()},
            {'type': 'plain', 'label': "Link:",     'value': self.get_link(instance)},
            {'type': 'plain', 'label': "Proposed:", 'value': instance.created.strftime("%d/%b/%Y %H:%M")}
            ]

        if context['is_app_editor']:
            lhtml = []
            lbuttons = [dict(href="{}?approvelist={}".format(reverse('userkeyword_list'), instance.id), 
                             label="Approve Keyword", 
                             title="Approving this keyword attaches it to the target and removes it from the list of user keywords.")]
            lhtml.append("<div class='row'><div class='col-md-12' align='right'>")
            for item in lbuttons:
                lhtml.append("  <a role='button' class='btn btn-xs jumbo-3' title='{}' href='{}'>".format(item['title'], item['href']))
                lhtml.append("     <span class='glyphicon glyphicon-ok'></span>{}</a>".format(item['label']))
            lhtml.append("</div></div>")
            context['after_details'] = "\n".join(lhtml)

        # Return the context we have made
        return context

    def get_link(self, instance):
        details = ""
        value = ""
        url = ""
        sBack = ""
        if instance.type == "manu" and instance.manu: 
            # Manuscript: shelfmark
            url = reverse('manuscript_details', kwargs = {'pk': instance.manu.id})
            # issue #610
            # value = instance.manu.idno
            value = instance.get_full_name()
            sBack = "<span class='badge signature ot'><a href='{}'>{}</a></span>".format(url, value)
        elif instance.type == "sermo" and instance.sermo: 
            # Sermon (manifestation): Gryson/Clavis + shelf mark (of M)
            sig = instance.sermo.get_eqsetsignatures_markdown("first")
            # Sermon identification
            url = reverse('sermon_details', kwargs = {'pk': instance.sermo.id})
            # value = "{}/{}".format(instance.sermo.order, instance.sermo.manu.manusermons.all().count())
            value = "{}/{}".format(instance.sermo.order, instance.sermo.manu.get_sermon_count())
            sermo = "<span><a href='{}'>sermon {}</a></span>".format(url, value)
            # Manuscript shelfmark
            url = reverse('manuscript_details', kwargs = {'pk': instance.sermo.manu.id})
            # issue #610
            # value = instance.sermo.manu.idno
            value = instance.get_full_name()
            manu = "<span class='badge signature ot'><a href='{}'>{}</a></span>".format(url, value)
            # Combine
            sBack = "{} {} {}".format(sermo, manu, sig)
        elif instance.type == "gold" and instance.gold: 
            # Get signatures
            sig = instance.gold.get_signatures_markdown()
            # Get Gold URL
            url = reverse('gold_details', kwargs = {'pk': instance.gold.id})
            # Combine
            sBack = "<span class='badge signature ot'><a href='{}'>{}</a></span> {}".format(url, "gold", sig)
        elif instance.type == "super" and instance.super: 
            # Get signatures
            sig = instance.super.get_goldsigfirst()
            # Get Gold URL
            url = reverse('equalgold_details', kwargs = {'pk': instance.super.id})
            # Combine
            sBack = "<span class='badge signature ot'><a href='{}'>{}</a></span> {}".format(url, "super", sig)
        return sBack

    def action_add(self, instance, details, actiontype):
        """User can fill this in to his/her liking"""
        passim_action_add(self, instance, details, actiontype)

    def get_history(self, instance):
        return passim_get_history(instance)


class UserKeywordDetails(UserKeywordEdit):
    """Like UserKeyword Edit, but then html output"""
    rtype = "html"
    

class UserKeywordSubmit(UserKeywordDetails):
    """Submit an adapted range of user keywords and return to the details page"""

    newRedirect = True

    def custom_init(self, instance):
        """"""

        # Note: use [oErr]
        oErr = ErrHandle()
        try:
            # Make sure to set the correct redirect page
            if instance:
                self.redirectpage = reverse("userkeyword_details", kwargs={'pk': instance.id})
            # Make sure we are not saving
            self.do_not_save = True
  

            return True
        except:
            msg = oErr.get_error_message()
            oErr.DoError("UserKeywordSubmit")
            return False


class UserKeywordListView(BasicList):
    """Search and list keywords"""

    model = UserKeyword
    listform = UserKeywordForm
    prefix = "ukw"
    sg_name = "User Keyword"     
    plural_name = "User Keywords"
    paginate_by = 20
    has_select2 = True
    in_team = False
    new_button = False
    order_cols = ['profile__user__username', 'keyword__name',  'type', '', 'created', '']
    order_default = order_cols
    order_heads = [{'name': 'User',     'order': 'o=1', 'type': 'str', 'custom': 'profile'},
                   {'name': 'Keyword',  'order': 'o=2', 'type': 'str', 'custom': 'keyword', 'main': True, 'linkdetails': True},
                   {'name': 'Type',     'order': 'o=3', 'type': 'str', 'custom': 'itemtype'},
                   {'name': 'Link',     'order': '',    'type': 'str', 'custom': 'link'},
                   {'name': 'Proposed', 'order': 'o=5', 'type': 'str', 'custom': 'date'},
                   {'name': 'Approve',  'order': '',    'type': 'str', 'custom': 'approve', 'align': 'right'}]
    filters = [ {"name": "Keyword",     "id": "filter_keyword", "enabled": False},
                {"name": "User",        "id": "filter_profile", "enabled": False},
                {"name": "Type",        "id": "filter_type",    "enabled": False}]
    searches = [
        {'section': '', 'filterlist': [
            {'filter': 'keyword',   'fkfield': 'keyword', 'keyFk': 'id', 'keyList': 'kwlist',      'infield': 'id' },
            {'filter': 'profile',   'fkfield': 'profile', 'keyFk': 'id', 'keyList': 'profilelist', 'infield': 'id' },
            {'filter': 'type',      'dbfield': 'type',    'keyS': 'type' }]}
        ]
    custombuttons = [{"name": "approve_keywords", "title": "Approve currently filtered keywords", 
                      "icon": "music", "template_name": "seeker/approve_keywords.html" }]

    def initializations(self):
        if self.request.user:
            username = self.request.user.username
            # See if there is a list of approved id's
            qd = self.request.GET if self.request.method == "GET" else self.request.POST
            approvelist = qd.get("approvelist", None)
            if approvelist != None:
                # See if this needs translation
                if approvelist[0] == "[":
                    approvelist = json.loads(approvelist)
                else:
                    approvelist = [ approvelist ]
                # Does this user have the right privilages?
                if user_is_superuser(self.request) or user_is_ingroup(self.request, app_editor):
                    # Get the profile
                    profile = Profile.get_user_profile(username)

                    # Approve the UserKeyword stated here
                    for ukw_id in approvelist:
                        obj = UserKeyword.objects.filter(profile=profile, id=ukw_id).first()
                        if obj != None:
                            obj.moveup()
        return None

    def get_field_value(self, instance, custom):
        sBack = ""
        sTitle = ""
        if custom == "keyword":
            sBack = instance.keyword.name
        elif custom == "profile":
            username = instance.profile.user.username
            url = reverse("profile_details", kwargs = {'pk': instance.profile.id})
            sBack = "<a href='{}'>{}</a>".format(url, username)
        elif custom == "itemtype":
            sBack = instance.get_type_display()
        elif custom == "link":
            sBack = self.get_link(instance)
        elif custom == "date":
            sBack = instance.created.strftime("%d/%b/%Y %H:%M")
        elif custom == "approve":
            url = "{}?approvelist={}".format(reverse("userkeyword_list"), instance.id)
            sBack = "<a class='btn btn-xs jumbo-2' role='button' href='{}' title='Approve this keyword'><span class='glyphicon glyphicon-thumbs-up'></span></a>".format(url)
        return sBack, sTitle

    def get_link(self, instance):
        details = ""
        value = ""
        url = ""
        sBack = ""
        if instance.type == "manu" and instance.manu: 
            # Manuscript: shelfmark
            url = reverse('manuscript_details', kwargs = {'pk': instance.manu.id})
            # issue #610
            # value = instance.manu.idno
            value = instance.get_full_name()
            sBack = "<span class='badge signature ot'><a href='{}'>{}</a></span>".format(url, value)
        elif instance.type == "sermo" and instance.sermo: 
            # Sermon (manifestation): Gryson/Clavis + shelf mark (of M)
            sig = instance.sermo.get_eqsetsignatures_markdown("first")
            # Sermon identification
            url = reverse('sermon_details', kwargs = {'pk': instance.sermo.id})
            manu_obj = instance.sermo.get_manuscript()
            value = "{}/{}".format(instance.sermo.order, manu_obj.get_sermon_count())
            sermo = "<span><a href='{}'>sermon {}</a></span>".format(url, value)
            # Manuscript shelfmark
            url = reverse('manuscript_details', kwargs = {'pk': manu_obj.id})
            # issue #610
            # value = manu_obj.idno
            value = manu_obj.get_full_name()
            manu = "<span class='badge signature ot'><a href='{}'>{}</a></span>".format(url, manu_obj.idno)
            # Combine
            sBack = "{} {} {}".format(sermo, manu, sig)
        elif instance.type == "gold" and instance.gold: 
            # Get signatures
            sig = instance.gold.get_signatures_markdown()
            # Get Gold URL
            url = reverse('gold_details', kwargs = {'pk': instance.gold.id})
            # Combine
            sBack = "<span class='badge signature ot'><a href='{}'>{}</a></span> {}".format(url, "gold", sig)
        elif instance.type == "super" and instance.super: 
            # Get signatures
            sig = instance.super.get_goldsigfirst()
            # Get Gold URL
            url = reverse('equalgold_details', kwargs = {'pk': instance.super.id})
            # Combine
            sBack = "<span class='badge signature ot'><a href='{}'>{}</a></span> {}".format(url, "super", sig)
        return sBack

    def add_to_context(self, context, initial):
        # Make sure to add a list of the currently filtered keywords
        if self.qs != None:
            lst_ukw = [x.id for x in self.qs]
            context['ukw_selection'] = lst_ukw
            context['ukw_list'] = reverse("userkeyword_list")
        return context


# =============== PROVENANCE ===============================
        

class ProvenanceEdit(BasicDetails):
    """The details of one 'provenance'"""

    model = Provenance
    mForm = ProvenanceForm
    prefix = 'prov'
    title = "Provenance"
    rtype = "json"
    history_button = True
    mainitems = []
    
    def add_to_context(self, context, instance):
        """Add to the existing context"""

        # Define the main items to show and edit
        context['mainitems'] = [
            {'type': 'plain', 'label': "Name:",         'value': instance.name,             'field_key': "name"},
            {'type': 'plain', 'label': "Location:",     'value': instance.get_location(),   'field_key': 'location'     },
            # Notes must now appear in the list of related manuscripts
            # {'type': 'line',  'label': "Note:",         'value': instance.note,             'field_key': 'note'},
            # Note issue #289: the manuscripts are now shown in the listview below (see ProvenanceDetails)
            # {'type': 'plain', 'label': "Manuscripts:",  'value': self.get_manuscripts(instance)}
            ]

        # Signal that we have select2
        context['has_select2'] = True

        # Return the context we have made
        return context

    def action_add(self, instance, details, actiontype):
        """User can fill this in to his/her liking"""
        passim_action_add(self, instance, details, actiontype)

    def get_history(self, instance):
        return passim_get_history(instance)

    def get_manuscripts(self, instance):
        # find the shelfmark
        manu = instance.manu
        if manu != None:
            # Get the URL to the manu details
            url = reverse("manuscript_details", kwargs = {'pk': manu.id})
            shelfmark = manu.idno[:20]
            sBack = "<span class='badge signature cl'><a href='{}'>{}</a></span>".format(url, manu.idno)
        return sBack


class ProvenanceDetails(ProvenanceEdit):
    """Like Provenance Edit, but then html output"""
    rtype = "html"

    def add_to_context(self, context, instance):
        # First get the 'standard' context
        context = super(ProvenanceDetails, self).add_to_context(context, instance)

        context['sections'] = []

        # Lists of related objects
        related_objects = []
        resizable = True
        index = 1
        sort_start = '<span class="sortable"><span class="fa fa-sort sortshow"></span>&nbsp;'
        sort_start_int = '<span class="sortable integer"><span class="fa fa-sort sortshow"></span>&nbsp;'
        sort_end = '</span>'

        # List of Manuscripts that use this provenance
        manuscripts = dict(title="Manuscripts with this provenance", prefix="mprov")
        if resizable: manuscripts['gridclass'] = "resizable"

        rel_list =[]
        qs = instance.manuscripts_provenances.all().order_by('manuscript__idno')
        for item in qs:
            manu = item.manuscript
            url = reverse('manuscript_details', kwargs={'pk': manu.id})
            url_pm = reverse('provenanceman_details', kwargs={'pk': item.id})
            rel_item = []

            # S: Order number for this manuscript
            add_rel_item(rel_item, index, False, align="right")
            index += 1

            # Manuscript
            manu_full = "{}, {}, <span class='signature'>{}</span> {}".format(manu.get_city(), manu.get_library(), manu.idno, manu.name)
            add_rel_item(rel_item, manu_full, False, main=False, link=url)

            # Note for this provenance
            note = "(none)" if item.note == None or item.note == "" else item.note
            add_rel_item(rel_item, note, False, nowrap=False, main=True, link=url_pm,
                         title="Note for this provenance-manuscript relation")

            # Add this line to the list
            rel_list.append(dict(id=item.id, cols=rel_item))

        manuscripts['rel_list'] = rel_list

        manuscripts['columns'] = [
            '{}<span>#</span>{}'.format(sort_start_int, sort_end), 
            '{}<span>Manuscript</span>{}'.format(sort_start, sort_end), 
            '{}<span>Note</span>{}'.format(sort_start, sort_end)
            ]
        related_objects.append(manuscripts)

        # List of Codicos that use this provenance
        codicos = dict(title="Codicological units with this provenance", prefix="mcodi")
        if resizable: codicos['gridclass'] = "resizable"

        rel_list =[]
        qs = instance.codico_provenances.all().order_by('codico__manuscript__idno', 'codico__order')
        for item in qs:
            codico = item.codico
            manu = codico.manuscript
            url = reverse('manuscript_details', kwargs={'pk': manu.id})
            url_c = reverse('codico_details', kwargs={'pk': codico.id})
            url_pc = reverse('provenancecod_details', kwargs={'pk': item.id})
            rel_item = []

            # S: Order number for this manuscript
            add_rel_item(rel_item, index, False, align="right")
            index += 1

            # Manuscript
            manu_full = "{}, {}, <span class='signature'>{}</span> {}".format(manu.get_city(), manu.get_library(), manu.idno, manu.name)
            add_rel_item(rel_item, manu_full, False, main=False, link=url)

            # Codico
            codico_full = "<span class='badge signature ot'>{}</span>".format(codico.order)
            add_rel_item(rel_item, codico_full, False, main=False, link=url_c)

            # Note for this provenance
            note = "(none)" if item.note == None or item.note == "" else item.note
            add_rel_item(rel_item, note, False, nowrap=False, main=True, link=url_pc,
                         title="Note for this provenance-codico relation")

            # Add this line to the list
            rel_list.append(dict(id=item.id, cols=rel_item))

        codicos['rel_list'] = rel_list

        codicos['columns'] = [
            '{}<span>#</span>{}'.format(sort_start_int, sort_end), 
            '{}<span>Manuscript</span>{}'.format(sort_start, sort_end), 
            '{}<span>Codicological unit</span>{}'.format(sort_start, sort_end), 
            '{}<span>Note</span>{}'.format(sort_start, sort_end)
            ]
        related_objects.append(codicos)

        # Add all related objects to the context
        context['related_objects'] = related_objects

        # Return the context we have made
        return context
    

class ProvenanceListView(BasicList):
    """Search and list provenances"""

    model = Provenance
    listform = ProvenanceForm
    prefix = "prov"
    plural_name = "Provenances"
    sg_name = "Provenance"
    has_select2 = True
    new_button = True   # Provenances are added in the Manuscript view; each provenance belongs to one manuscript
                        # Issue #289: provenances are to be added *HERE*
    order_cols = ['location__name', 'name']
    order_default = order_cols
    order_heads = [
        {'name': 'Location',    'order': 'o=1', 'type': 'str', 'custom': 'location', 'linkdetails': True},
        {'name': 'Name',        'order': 'o=2', 'type': 'str', 'field':  'name', 'main': True, 'linkdetails': True},
        # Issue #289: remove this note from here
        # {'name': 'Note',        'order': 'o=3', 'type': 'str', 'custom': 'note', 'linkdetails': True},
        {'name': 'Manuscript',  'order': 'o=3', 'type': 'str', 'custom': 'manuscript', 'allowwrap': True}
        ]
    filters = [ {"name": "Name",        "id": "filter_name",    "enabled": False},
                {"name": "Location",    "id": "filter_location","enabled": False},
                {"name": "Manuscript",  "id": "filter_manuid",  "enabled": False},
               ]
    searches = [
        {'section': '', 'filterlist': [
            {'filter': 'name',      'dbfield': 'name', 'keyS': 'name'},
            {'filter': 'location',  'fkfield': 'location', 'keyS': 'location_ta', 'keyId': 'location', 'keyFk': "name", 'keyList': 'locationlist', 'infield': 'id' },
            {'filter': 'manuid',    'fkfield': 'manuscripts_provenances__manuscript', 'keyFk': 'idno', 'keyList': 'manuidlist', 'infield': 'id' }
            # Issue #289: innovation below turned back to the original above
            # {'filter': 'manuid',    'fkfield': 'manu', 'keyFk': 'idno', 'keyList': 'manuidlist', 'infield': 'id' }
            ]}
        ]

    def initializations(self):
        """Perform some initializations"""

        oErr = ErrHandle()
        try:

            # ======== One-time adaptations ==============
            listview_adaptations("provenance_list")

        except:
            msg = oErr.get_error_message()
            oErr.DoError("ProvenanceListView/initializations")

        return None

    def get_field_value(self, instance, custom):
        sBack = ""
        sTitle = ""
        if custom == "manuscript":
            # Multiple connections possible
            # One provenance may be connected to any number of manuscripts!
            lManu = []
            for obj in instance.manuscripts_provenances.all():
                # Add the shelfmark of this one
                manu = obj.manuscript
                url = reverse("manuscript_details", kwargs = {'pk': manu.id})
                # issue #610
                # value = manu.idno
                value = manu.get_full_name()
                lManu.append("<span class='badge signature cl'><a href='{}'>{}</a></span>".format(url, value))
            sBack = ", ".join(lManu)
            # Issue #289: the innovation below is turned back to the original above
            ## find the shelfmark
            #manu = instance.manu
            #if manu != None:
            #    # Get the URL to the manu details
            #    url = reverse("manuscript_details", kwargs = {'pk': manu.id})
            #    shelfmark = manu.idno[:20]
            #    sBack = "<span class='badge signature cl'><a href='{}'>{}</a></span>".format(url, manu.idno)
        elif custom == "location":
            sBack = ""
            if instance.location:
                sBack = instance.location.name
        #elif custom == "note":
        #    sBack = ""
        #    if instance.note:
        #        sBack = instance.note[:40]
        return sBack, sTitle

    def adapt_search(self, fields):
        lstExclude=None
        qAlternative = None
        x = fields
        return fields, lstExclude, qAlternative


class ProvenanceManEdit(BasicDetails):
    """The details of one 'provenance'"""

    model = ProvenanceMan
    mForm = ProvenanceManForm
    prefix = 'mprov'
    title = "ManuscriptProvenance"
    rtype = "json"
    history_button = False # True
    mainitems = []
    
    def add_to_context(self, context, instance):
        """Add to the existing context"""

        # Define the main items to show and edit
        context['mainitems'] = [
            {'type': 'safe',  'label': "Provenance:",   'value': instance.get_provenance()},
            {'type': 'plain', 'label': "Note:",         'value': instance.note,   'field_key': 'note'     },
            ]

        # Signal that we have select2
        context['has_select2'] = True

        context['listview'] = reverse("manuscript_details", kwargs={'pk': instance.manuscript.id})

        # Return the context we have made
        return context

    def action_add(self, instance, details, actiontype):
        """User can fill this in to his/her liking"""
        passim_action_add(self, instance, details, actiontype)

    def get_history(self, instance):
        return passim_get_history(instance)


class ProvenanceManDetails(ProvenanceManEdit):
    """Like ProvenanceMan Edit, but then html output"""
    rtype = "html"
        

class ProvenanceCodEdit(BasicDetails):
    """The details of one 'provenance'"""

    model = ProvenanceCod
    mForm = ProvenanceCodForm
    prefix = 'cprov'
    title = "CodicoProvenance"
    rtype = "json"
    history_button = False # True
    mainitems = []
    
    def add_to_context(self, context, instance):
        """Add to the existing context"""

        # Define the main items to show and edit
        context['mainitems'] = [
            {'type': 'safe',  'label': "Provenance:",   'value': instance.get_provenance()},
            {'type': 'plain', 'label': "Note:",         'value': instance.note,   'field_key': 'note'     },
            ]

        # Signal that we have select2
        context['has_select2'] = True

        context['listview'] = reverse("codico_details", kwargs={'pk': instance.codico.id})

        # Return the context we have made
        return context

    def action_add(self, instance, details, actiontype):
        """User can fill this in to his/her liking"""
        passim_action_add(self, instance, details, actiontype)

    def get_history(self, instance):
        return passim_get_history(instance)


class ProvenanceCodDetails(ProvenanceCodEdit):
    """Like ProvenanceCod Edit, but then html output"""
    rtype = "html"


# =============== BIBRANGE ===============================
        

class BibRangeEdit(BasicDetails):
    """The details of one 'user-keyword': one that has been linked by a user"""

    model = BibRange
    mForm = BibRangeForm
    prefix = 'brng'
    title = "Bible references"
    title_sg = "Bible reference"
    rtype = "json"
    history_button = False # True
    mainitems = []
    
    def add_to_context(self, context, instance):
        """Add to the existing context"""

        # Find out what type of bibref this is: to an SSG or to an S?
        is_sermon_bibref = (instance.equal is None)

        # Define the main items to show and edit
        main_items = [
            {'type': 'plain', 'label': "Book:",         'value': instance.get_book(),   'field_key': 'book', 'key_hide': True },
            {'type': 'plain', 'label': "Abbreviations:",'value': instance.get_abbr()                        },
            {'type': 'plain', 'label': "Chapter/verse:",'value': instance.chvslist,     'field_key': 'chvslist', 'key_hide': True },
            {'type': 'line',  'label': "Intro:",        'value': instance.intro,        'field_key': 'intro'},
            {'type': 'line',  'label': "Extra:",        'value': instance.added,        'field_key': 'added'},
            ]

        if is_sermon_bibref:
            add_links = [
                {'type': 'plain', 'label': "Sermon:",       'value': self.get_sermon(instance)                  },
                {'type': 'plain', 'label': "Manuscript:",   'value': self.get_manuscript(instance)              }
                ]
        else:
            add_links = [
                {'type': 'plain', 'label': "Authority File:",'value': self.get_equal(instance)                  },
                ]

        context['mainitems'] = main_items + add_links

        #context['mainsections'] = [
        #    {'name': 'Experiment', 'id': 'bib_extra', 'fields': [
        #        {'type': 'line',  'label': "Intro:",        'value': instance.intro,        'field_key': 'intro'},
        #        {'type': 'line',  'label': "Extra:",        'value': instance.added,        'field_key': 'added'},
        #        ]}
        #    ]

        # Signal that we have select2
        context['has_select2'] = True

        # Return the context we have made
        return context

    def action_add(self, instance, details, actiontype):
        """User can fill this in to his/her liking"""
        passim_action_add(self, instance, details, actiontype)

    def get_history(self, instance):
        return passim_get_history(instance)

    def get_manuscript(self, instance):
        # find the shelfmark via the sermon
        manu = instance.sermon.msitem.manu
        url = reverse("manuscript_details", kwargs = {'pk': manu.id})
        sBack = "<span class='badge signature cl'><a href='{}'>{}</a></span>".format(url, manu.get_full_name())
        return sBack

    def get_sermon(self, instance):
        # Get the sermon
        sermon = instance.sermon
        url = reverse("sermon_details", kwargs = {'pk': sermon.id})
        title = "{}: {}".format(sermon.msitem.manu.idno, sermon.locus)
        sBack = "<span class='badge signature gr'><a href='{}'>{}</a></span>".format(url, title)
        return sBack

    def get_equal(self, instance):
        # Get the EqualGold instance
        equal = instance.equal
        url = reverse("equalgold_details", kwargs = {'pk': equal.id})
        sBack = equal.get_passimcode_markdown()
        return sBack


class BibRangeDetails(BibRangeEdit):
    """Like BibRange Edit, but then html output"""
    rtype = "html"
    

class BibRangeListView(BasicList):
    """Search and list provenances"""

    model = BibRange
    listform = BibRangeForm
    prefix = "brng"
    has_select2 = True
    sg_name = "Bible reference"
    plural_name = "Bible references"
    new_button = False  # BibRanges are added in the Manuscript view; each provenance belongs to one manuscript
    order_cols = ['book__idno', 'chvslist', 'intro', 'added', 
                  'sermon__msitem__codico__manuscript__idno;sermon__locus', 'equal__code']
    order_default = order_cols
    order_heads = [
        {'name': 'Book',            'order': 'o=1', 'type': 'str', 'custom': 'book', 'linkdetails': True},
        {'name': 'Chapter/verse',   'order': 'o=2', 'type': 'str', 'field': 'chvslist', 'main': True, 'linkdetails': True},
        {'name': 'Intro',           'order': 'o=3', 'type': 'str', 'custom': 'intro', 'linkdetails': True},
        {'name': 'Extra',           'order': 'o=4', 'type': 'str', 'custom': 'added', 'linkdetails': True},
        {'name': 'Manifestation',   'order': 'o=5', 'type': 'str', 'custom': 'sermon', 'allowwrap': True},
        {'name': 'Authority File',  'order': 'o=6', 'type': 'str', 'custom': 'equal'},
        ]
    filters = [ 
        {"name": "Bible reference", "id": "filter_bibref",      "enabled": False},
        {"name": "Intro",           "id": "filter_intro",       "enabled": False},
        {"name": "Extra",           "id": "filter_added",       "enabled": False},
        {"name": "Manuscript...",   "id": "filter_manuscript",  "enabled": False, "head_id": "none"},
        {"name": "Shelfmark",       "id": "filter_manuid",      "enabled": False, "head_id": "filter_manuscript"},
        {"name": "Country",         "id": "filter_country",     "enabled": False, "head_id": "filter_manuscript"},
        {"name": "City",            "id": "filter_city",        "enabled": False, "head_id": "filter_manuscript"},
        {"name": "Library",         "id": "filter_library",     "enabled": False, "head_id": "filter_manuscript"},
        {"name": "Origin",          "id": "filter_origin",      "enabled": False, "head_id": "filter_manuscript"},
        {"name": "Provenance",      "id": "filter_provenance",  "enabled": False, "head_id": "filter_manuscript"},
        {"name": "Date from",       "id": "filter_datestart",   "enabled": False, "head_id": "filter_manuscript"},
        {"name": "Date until",      "id": "filter_datefinish",  "enabled": False, "head_id": "filter_manuscript"},
               ]
    searches = [
        {'section': '', 'filterlist': [
            {'filter': 'bibref',    'dbfield': '$dummy',    'keyS': 'bibrefbk'},
            {'filter': 'bibref',    'dbfield': '$dummy',    'keyS': 'bibrefchvs'},
            {'filter': 'intro',     'dbfield': 'intro',     'keyS': 'intro'},
            {'filter': 'added',     'dbfield': 'added',     'keyS': 'added'}
            ]},
        {'section': 'manuscript', 'filterlist': [
            {'filter': 'manuid',        'fkfield': 'sermon__msitem__manu',                    'keyS': 'manuidno',   'keyList': 'manuidlist', 'keyFk': 'idno', 'infield': 'id'},
            {'filter': 'country',       'fkfield': 'sermon__msitem__manu__library__lcountry', 'keyS': 'country_ta', 'keyId': 'country',     'keyFk': "name"},
            {'filter': 'city',          'fkfield': 'sermon__msitem__manu__library__lcity',    'keyS': 'city_ta',    'keyId': 'city',        'keyFk': "name"},
            {'filter': 'library',       'fkfield': 'sermon__msitem__manu__library',           'keyS': 'libname_ta', 'keyId': 'library',     'keyFk': "name"},
            {'filter': 'origin',        'fkfield': 'sermon__msitem__manu__origin',            'keyS': 'origin_ta',  'keyId': 'origin',      'keyFk': "name"},
            {'filter': 'provenance',    'fkfield': 'sermon__msitem__manu__provenances',       'keyS': 'prov_ta',    'keyId': 'prov',        'keyFk': "name"},
            {'filter': 'datestart',     'dbfield': 'sermon__msitem__codico__codico_dateranges__yearstart__gte',     'keyS': 'date_from'},
            {'filter': 'datefinish',    'dbfield': 'sermon__msitem__codico__codico_dateranges__yearfinish__lte',    'keyS': 'date_until'},
            ]},
        {'section': 'other', 'filterlist': [
            {'filter': 'bibref',     'dbfield': 'id',    'keyS': 'bibref'}
            ]}
        ]

    def get_field_value(self, instance, custom):
        sBack = ""
        sTitle = ""
        if custom == "sermon":
            if not instance.sermon is None:
                sermon = instance.sermon
                # find the shelfmark
                manu = sermon.msitem.manu
                url = reverse("sermon_details", kwargs = {'pk': sermon.id})
                # issue #610
                # value = manu.idno
                value = manu.get_full_name()
                sBack = "<span class='badge signature cl'><a href='{}'>{}: {}</a></span>".format(url, manu.idno, sermon.locus)
        elif custom == "equal":
            if not instance.equal is None:
                sBack = instance.equal.get_passimcode_markdown()
        elif custom == "book":
            sBack = instance.book.name
        elif custom == "intro":
            sBack = " "
            if instance.intro != "":
                sBack = instance.intro
        elif custom == "added":
            sBack = " "
            if instance.added != "":
                sBack = instance.added
        return sBack, sTitle

    def adapt_search(self, fields):
        lstExclude=None
        qAlternative = None

        # Adapt the bible reference list
        bibrefbk = fields.get("bibrefbk", "")
        if bibrefbk != None and bibrefbk != "":
            bibrefchvs = fields.get("bibrefchvs", "")

            # Get the start and end of this bibref
            start, einde = Reference.get_startend(bibrefchvs, book=bibrefbk)
 
            # Find out which sermons have references in this range
            lstQ = []
            lstQ.append(Q(bibrangeverses__bkchvs__gte=start))
            lstQ.append(Q(bibrangeverses__bkchvs__lte=einde))
            # sermonlist = [x.id for x in BibRange.objects.filter(*lstQ).order_by('id').distinct()]

            # fields['bibrefbk'] = Q(id__in=sermonlist)
            fields['bibrefbk'] = Q(bibrangeverses__bkchvs__gte=start) & Q(bibrangeverses__bkchvs__lte=einde)

        return fields, lstExclude, qAlternative


# =============== FEAST ==================================
        

class FeastEdit(BasicDetails):
    """The details of one Christian Feast"""

    model = Feast
    mForm = FeastForm
    prefix = 'fst'
    title = "Feast"
    title_sg = "Feast"
    rtype = "json"
    history_button = False # True
    mainitems = []
    
    def add_to_context(self, context, instance):
        """Add to the existing context"""

        # Define the main items to show and edit
        context['mainitems'] = [
            {'type': 'plain', 'label': "Name:",         'value': instance.name,         'field_key': 'name'     },
            {'type': 'plain', 'label': "Latin name:",   'value': instance.get_latname(),'field_key': 'latname'  },
            {'type': 'plain', 'label': "Feast date:",   'value': instance.get_date(),   'field_key': 'feastdate' },
            #{'type': 'plain', 'label': "Sermon:",       'value': self.get_sermon(instance)                  },
            #{'type': 'plain', 'label': "Manuscript:",   'value': self.get_manuscript(instance)              }
            ]

        # Signal that we have select2
        context['has_select2'] = True

        # Return the context we have made
        return context

    def action_add(self, instance, details, actiontype):
        """User can fill this in to his/her liking"""
        passim_action_add(self, instance, details, actiontype)

    def get_history(self, instance):
        return passim_get_history(instance)

    def get_manuscript(self, instance):
        html = []
        # find the shelfmark via the sermon
        for manu in Manuscript.objects.filter(manuitems__itemsermons__feast=instance).order_by("idno"):
            url = reverse("manuscript_details", kwargs = {'pk': manu.id})
            html.append("<span class='badge signature cl'><a href='{}'>{}</a></span>".format(url, manu.get_full_name()))
        sBack = ", ".join(html)
        return sBack

    def get_sermon(self, instance):
        html = []
        # Get the sermons
        for sermon in SermonDescr.objects.filter(feast=instance).order_by("msitem__manu__idno", "locus"):
            url = reverse("sermon_details", kwargs = {'pk': sermon.id})
            title = "{}: {}".format(sermon.msitem.manu.idno, sermon.locus)
            html.append("<span class='badge signature gr'><a href='{}'>{}</a></span>".format(url, title))
        sBack = ", ".join(html)
        return sBack


class FeastDetails(FeastEdit):
    """Like Feast Edit, but then html output"""
    rtype = "html"

    def add_to_context(self, context, instance):
        # First get the 'standard' context from TestsetEdit
        context = super(FeastDetails, self).add_to_context(context, instance)

        oErr = ErrHandle()
        try:
            context['sections'] = []

            # Lists of related objects
            related_objects = []
            resizable = True
            index = 1
            sort_start = '<span class="sortable"><span class="fa fa-sort sortshow"></span>&nbsp;'
            sort_start_int = '<span class="sortable integer"><span class="fa fa-sort sortshow"></span>&nbsp;'
            sort_end = '</span>'

            # List of Sermons that link to this feast (with an FK)
            sermons = dict(title="Manuscripts with sermons connected to this feast", prefix="tunit")
            if resizable: sermons['gridclass'] = "resizable"

            rel_list =[]
            # Note: default sort order according to issue #632
            qs = instance.feastsermons.all().order_by('msitem__manu__lcity__name', 'msitem__manu__library__name', 'msitem__manu__idno', 'locus')
            for item in qs:
                manu = item.msitem.manu
                url = reverse('sermon_details', kwargs={'pk': item.id})
                url_m = reverse('manuscript_details', kwargs={'pk': manu.id})
                rel_item = []

                # S: Order number for this sermon
                add_rel_item(rel_item, index, False, align="right")
                index += 1

                # Manuscript
                manu_full = manu.get_full_name(plain=False)
                add_rel_item(rel_item, manu_full, False, main=False, nowrap=False, link=url_m)

                # Locus
                locus = "(none)" if item.locus == None or item.locus == "" else item.locus
                add_rel_item(rel_item, locus, False, main=False, nowrap=False, link=url, 
                             title="Locus within the manuscript (links to the manifestation)")

                # Title
                title = item.get_title()
                add_rel_item(rel_item, title, False, main=False, nowrap=False, link=url, 
                             title="Manifestation's title (links to the manifestation)")

                # Section title
                section_title = item.get_sectiontitle()
                add_rel_item(rel_item, section_title, False, main=False, nowrap=False, link=url, 
                             title="Manifestation's section title (links to the manifestation)")

                # Signature (Gryson/Clavis/Other)
                sig_text = item.signature_auto_string()
                add_rel_item(rel_item, sig_text, False, main=False, nowrap=False, link=url, 
                             title="Gryson/Clavis/Other of associated Authority File (links to the manifestation)")

                # Origin/provenance
                or_prov = "{} ({})".format(manu.get_origin(), manu.get_provenance_markdown())
                add_rel_item(rel_item, or_prov, False, main=False, nowrap=False, 
                             title="Origin (if known), followed by provenances (between brackets)")

                # Date
                daterange = "{}-{}".format(manu.yearstart, manu.yearfinish)
                add_rel_item(rel_item, daterange, False, link=url_m, align="right")

                # Add this line to the list
                rel_list.append(dict(id=item.id, cols=rel_item))

            sermons['rel_list'] = rel_list

            sermons['columns'] = [
                '{}<span>#</span>{}'.format(sort_start_int, sort_end), 
                '{}<span>Manuscript</span>{}'.format(sort_start, sort_end), 
                '{}<span>Locus</span>{}'.format(sort_start, sort_end), 
                '{}<span title="Manifestation title (links to the manifestation)">Title</span>{}'.format(sort_start, sort_end), 
                '{}<span title="Manifestation section title (links to the manifestation)">Section</span>{}'.format(sort_start, sort_end), 
                '{}<span title="Gryson/Clavis/Other of associated Authority File (links to the manifestation)">Gr./Cl.</span>{}'.format(
                    sort_start, sort_end), 
                '{}<span title="Origin/Provenance">or./prov.</span>{}'.format(sort_start, sort_end), 
                '{}<span>date</span>{}'.format(sort_start_int, sort_end)
                ]
            related_objects.append(sermons)

            # Add all related objects to the context
            context['related_objects'] = related_objects

        except:
            msg = oErr.get_error_message()
            oErr.DoError("FeastDetails/add_to_context")

        # Return the context we have made
        return context
    

class FeastListView(BasicList):
    """Search and list Christian feasts"""

    model = Feast
    listform = FeastForm
    prefix = "fst"
    has_select2 = True
    sg_name = "Feast"
    plural_name = "Feasts"
    new_button = True  # Feasts can be added from the listview
    order_cols = ['name', 'latname', 'feastdate', '']   # feastsermons__msitem__manu__idno;feastsermons__locus
    order_default = order_cols
    order_heads = [
        {'name': 'Name',    'order': 'o=1', 'type': 'str', 'field': 'name',     'linkdetails': True},
        {'name': 'Latin',   'order': 'o=2', 'type': 'str', 'field': 'latname',  'linkdetails': True},
        {'name': 'Date',    'order': 'o=3', 'type': 'str', 'field': 'feastdate','linkdetails': True, 'main': True},
        {'name': 'Sermons', 'order': '',    'type': 'str', 'custom': 'sermons'}
        ]
    filters = [ 
        {"name": "Name",            "id": "filter_engname",     "enabled": False},
        {"name": "Latin",           "id": "filter_latname",     "enabled": False},
        {"name": "Date",            "id": "filter_feastdate",   "enabled": False},
        {"name": "Manuscript...",   "id": "filter_manuscript",  "enabled": False, "head_id": "none"},
        {"name": "Shelfmark",       "id": "filter_manuid",      "enabled": False, "head_id": "filter_manuscript"},
        {"name": "Country",         "id": "filter_country",     "enabled": False, "head_id": "filter_manuscript"},
        {"name": "City",            "id": "filter_city",        "enabled": False, "head_id": "filter_manuscript"},
        {"name": "Library",         "id": "filter_library",     "enabled": False, "head_id": "filter_manuscript"},
        {"name": "Origin",          "id": "filter_origin",      "enabled": False, "head_id": "filter_manuscript"},
        {"name": "Provenance",      "id": "filter_provenance",  "enabled": False, "head_id": "filter_manuscript"},
        {"name": "Date from",       "id": "filter_datestart",   "enabled": False, "head_id": "filter_manuscript"},
        {"name": "Date until",      "id": "filter_datefinish",  "enabled": False, "head_id": "filter_manuscript"},
               ]
    searches = [
        {'section': '', 'filterlist': [
            {'filter': 'engname',   'dbfield': 'name',      'keyS': 'name'},
            {'filter': 'latname',   'dbfield': 'latname',   'keyS': 'latname'},
            {'filter': 'feastdate', 'dbfield': 'feastdate', 'keyS': 'feastdate'}
            ]},
        {'section': 'manuscript', 'filterlist': [
            {'filter': 'manuid',        'fkfield': 'feastsermons__msitem__manu',                    'keyS': 'manuidno',     'keyList': 'manuidlist', 'keyFk': 'idno', 'infield': 'id'},
            {'filter': 'country',       'fkfield': 'feastsermons__msitem__manu__library__lcountry', 'keyS': 'country_ta',   'keyId': 'country',     'keyFk': "name"},
            {'filter': 'city',          'fkfield': 'feastsermons__msitem__manu__library__lcity',    'keyS': 'city_ta',      'keyId': 'city',        'keyFk': "name"},
            {'filter': 'library',       'fkfield': 'feastsermons__msitem__manu__library',           'keyS': 'libname_ta',   'keyId': 'library',     'keyFk': "name"},
            {'filter': 'origin',        'fkfield': 'feastsermons__msitem__manu__origin',            'keyS': 'origin_ta',    'keyId': 'origin',      'keyFk': "name"},
            {'filter': 'provenance',    'fkfield': 'feastsermons__msitem__manu__provenances',       'keyS': 'prov_ta',      'keyId': 'prov',        'keyFk': "name"},
            {'filter': 'datestart',     'dbfield': 'feastsermons__msitem__codico__codico_dateranges__yearstart__gte',    'keyS': 'date_from'},
            {'filter': 'datefinish',    'dbfield': 'feastsermons__msitem__codico__codico_dateranges__yearfinish__lte',   'keyS': 'date_until'},
            ]}
        ]

    def get_field_value(self, instance, custom):
        sBack = ""
        sTitle = ""
        oErr = ErrHandle()
        try:
            if custom == "sermon":
                # NOTE: this is not accessed!!!
                html = []
                for sermon in instance.feastsermons.all().order_by('feast__name'):
                    # find the shelfmark
                    manu = sermon.msitem.manu
                    url = reverse("sermon_details", kwargs = {'pk': sermon.id})
                    # issue #610
                    # value = sermon.msitem.manu.idno
                    value = sermon.msitem.manu.get_full_name()
                    html.append("<span class='badge signature cl'><a href='{}'>{}: {}</a></span>".format(url, value, sermon.locus))
                sBack = ", ".join(html)
            elif custom == "sermons":
                # sBack = "{}".format(instance.feastsermons.count())
                html = []
                # Get the HTML code for the links of this instance
                number = instance.feastsermons.count()
                if number > 0:
                    url = reverse('sermon_list')
                    html.append("<a href='{}?sermo-feastlist={}'>".format(url, instance.id))
                    html.append("<span class='badge jumbo-1 clickable' title='Frequency in manifestation sermons'>{}</span></a>".format(number))
                sBack = "\n".join(html)

        except:
            msg = oErr.get_error_message()
            oErr.DoError("FeastListView/get_field_value")
        return sBack, sTitle


# ================= TEMPLATE =============================

class TemplateEdit(BasicDetails):
    """The details of one 'user-keyword': one that has been linked by a user"""

    model = Template
    mForm = TemplateForm
    prefix = 'tmpl'
    title = "TemplateEdit"
    rtype = "json"
    history_button = True
    use_team_group = True
    mainitems = []

    stype_edi_fields = ['name', 'description']
        
    def add_to_context(self, context, instance):
        """Add to the existing context"""

        # Define the main items to show and edit
        context['mainitems'] = [
            {'type': 'plain', 'label': "Name:",         'value': instance.name,             'field_key': "name"},
            {'type': 'line',  'label': "Description:",  'value': instance.description,      'field_key': 'description'},
            {'type': 'plain', 'label': "Items:",        'value': instance.get_count()},
            {'type': 'plain', 'label': "Owner:",        'value': instance.get_username()},
            {'type': 'plain', 'label': "Manuscript:",   'value': instance.get_manuscript_link()}
            ]

        # Signal that we have select2
        context['has_select2'] = True

        # Make sure 'authenticated' is adapted to only include EDITORS
        if context['authenticated']:
            if not context['is_app_editor'] and not user_is_superuser(self.request): context['permission'] = False

        # Return the context we have made
        return context

    def action_add(self, instance, details, actiontype):
        """User can fill this in to his/her liking"""
        passim_action_add(self, instance, details, actiontype)

    def get_history(self, instance):
        return passim_get_history(instance)


class TemplateDetails(TemplateEdit):
    """Like Template Edit, but then html output"""
    rtype = "html"

    def before_save(self, form, instance):
        bStatus = True
        msg = ""
        if instance == None or instance.id == None:
            # See if we have the profile id
            profile = Profile.get_user_profile(self.request.user.username)
            form.instance.profile = profile
            # See if we have the 'manubase' item
            manubase = self.qd.get("manubase", None)
            if manubase != None:
                # Find out which manuscript this is
                manu = Manuscript.objects.filter(id=manubase).first()
                if manu != None:
                    # Create a 'template-copy' of this manuscript
                    manutemplate = manu.get_manutemplate_copy()
                    instance.manu = manutemplate
        return bStatus, msg
    

class TemplateApply(TemplateDetails):
    """Create a new manuscript that is based on this template"""

    def custom_init(self, instance):
        # Find out who I am
        profile = Profile.get_user_profile(self.request.user.username)
        # Get the manuscript
        manu_template = instance.manu
        # Create a new manuscript that is based on this one
        manu_new = manu_template.get_manutemplate_copy("man", profile, instance)
        # Re-direct to this manuscript
        self.redirectpage = reverse("manuscript_details", kwargs={'pk': manu_new.id})
        return None


class TemplateImport(TemplateDetails):
    """Import manuscript sermons from a template"""

    initRedirect = True

    def initializations(self, request, pk):
        oErr = ErrHandle()
        try:
            # Find out who I am
            profile = Profile.get_user_profile(request.user.username)

            # Get the parameters
            self.qd = request.POST
            self.object = None
            manu_id = self.qd.get("manu_id", "")
            if manu_id != "":
                instance = Manuscript.objects.filter(id=manu_id).first()

            # The default redirectpage is just this manuscript
            self.redirectpage = reverse("manuscript_details", kwargs = {'pk': instance.id})
            # Get the template to be used as import
            template_id = self.qd.get("template", "")
            if template_id != "":
                template = Template.objects.filter(id=template_id).first()
                if template != None:
                    # Set my own object
                    self.object = template

                    # Import this template into the manuscript
                    instance.import_template(template, profile)
            # Getting here means all went well
        except:
            msg = oErr.get_error_message()
            oErr.DoError("TemplateImport/initializations")
        return None


class TemplateListView(BasicList):
    """Search and list templates"""

    model = Template
    listform = TemplateForm
    prefix = "tmpl"
    sg_name = "Template"     
    plural_name = "Templates"
    has_select2 = True
    new_button = False  # Templates are added in the Manuscript view; each template belongs to one manuscript
    use_team_group = True
    order_cols = ['profile__user__username', 'name', '']
    order_default = order_cols
    order_heads = [
        {'name': 'Owner',       'order': 'o=1', 'type': 'str', 'custom': 'owner'},
        {'name': 'Name',        'order': 'o=2', 'type': 'str', 'field': 'name', 'main': True, 'linkdetails': True},
        {'name': 'Items',       'order': '',    'type': 'str', 'custom': 'items', 'linkdetails': True},
        {'name': 'Manuscript',  'order': '',    'type': 'str', 'custom': 'manuscript'}
        ]
    filters = [
        ]
    searches = [
        ]

    def get_field_value(self, instance, custom):
        sBack = ""
        sTitle = ""
        if custom == "owner":
            # find the owner
            username = instance.get_username()
            sBack = "<span class='template_owner'>{}</span>".format(username)
        elif custom == "items":
            # The number of sermons (items) part of this manuscript
            sBack = "{}".format(instance.get_count())
        elif custom == "manuscript":
            url = reverse('template_apply', kwargs={'pk': instance.id})
            sBack = "<a href='{}' title='Create a new manuscript based on this template'><span class='glyphicon glyphicon-open'></span></a>".format(url)
        return sBack, sTitle

    def add_to_context(self, context, initial):
        # Make sure 'authenticated' is adapted to only include EDITORS
        if context['authenticated']:
            context['permission'] = context['is_app_editor'] or user_is_superuser(self.request)
            # if not context['is_app_editor'] and not user_is_superuser(self.request): context['permission'] = False
        return context


# ================= PROFILE and USER =============================

class UserEdit(BasicDetails):
    """Changing a user's changeable attributes"""

    model = User
    mForm = UserForm
    prefix = "usr"
    title = "UserEdit"
    no_delete = True
    mainitems = []

    def custom_init(self, instance):

        # Set the LISTVIEW to the MyPassim page
        self.listview = reverse('mypassim_details')
        self.listviewtitle = "My Passim"

        # Return nothing
        return None

    def add_to_context(self, context, instance):
        """Add to the existing context"""

        # Define the main items to show and edit
        context['mainitems'] = [
            # Issue #435: user must be able to change: email, pw, name
            {'type': 'plain', 'label': "Username:",     'value': instance.username}, # ,    'field_key': "username"},
            {'type': 'plain', 'label': "Email:",        'value': instance.email,       'field_key': "email"},
            {'type': 'plain', 'label': "First name:",   'value': instance.first_name,  'field_key': "first_name"},
            {'type': 'plain', 'label': "Last name:",    'value': instance.last_name,   'field_key': "last_name"},
            ]

        # Adapt the permission, if this is the actual user that is logged in
        if context['authenticated'] and not context['is_app_moderator']:
            # Check who this is
            user = self.request.user
            # Compare with instance
            if user.id == instance.id:
                self.permission = "write"
            else:
                # This is not the current user
                self.permission = "readonly"
            context['permission'] = self.permission

        # Return the context we have made
        return context


class UserDetails(UserEdit):
    """Like User Edit, but then with HTML output"""
    rtype = "html"


class ProfileEdit(BasicDetails):
    """Details view of profile"""

    model = Profile
    mForm = ProfileForm
    prefix = "prof"
    title = "ProfileEdit"
    rtype = "json"
    has_select2 = True
    history_button = True
    no_delete = True
    mainitems = []

    def add_to_context(self, context, instance):
        """Add to the existing context"""

        not_editable_by_user = ['username', 'status', 'project editing rights', 'project approval rights' ]

        # Define the main items to show and edit
        context['mainitems'] = [
            # The user.id must be passed on, but non-visibly
            {'type': 'plain', 'label': "User",          'value': instance.user.id, 'field_key': "user", 'empty': 'idonly'},
            # Issue #435: user must be able to change: email, pw, name
            {'type': 'plain', 'label': "Username:",     'value': instance.user.username,    'field_key': "newusername"},
            {'type': 'plain', 'label': "Email:",        'value': instance.user.email,       'field_key': "newemail"},
            {'type': 'plain', 'label': "First name:",   'value': instance.user.first_name,  'field_key': "newfirst_name"},
            {'type': 'plain', 'label': "Last name:",    'value': instance.user.last_name,   'field_key': "newlast_name"},
            {'type': 'plain', 'label': "Is staff:",     'value': instance.user.is_staff, },
            {'type': 'plain', 'label': "Is superuser:", 'value': instance.user.is_superuser, },
            {'type': 'plain', 'label': "Date joined:",  'value': instance.user.date_joined.strftime("%d/%b/%Y %H:%M"), },
            {'type': 'plain', 'label': "Last login:",   'value': instance.user.last_login.strftime("%d/%b/%Y %H:%M"), },
            {'type': 'plain', 'label': "Groups:",       'value': instance.get_groups_markdown(), },
            {'type': 'plain', 'label': "Status:",       'value': instance.get_ptype_display(),          'field_key': 'ptype'},
            {'type': 'line',  'label': "Affiliation:",   'value': instance.affiliation,                 'field_key': 'affiliation'},
            {'type': 'line',  'label': "Project editing rights:", 'value': instance.get_editor_projects_markdown(),     'field_list': 'editlist'},
            {'type': 'line',  'label': "Project approval rights:", 'value': instance.get_approver_projects_markdown(),  'field_list': 'projlist'}
            ]

        # If this is a moderator, add a button for editing rights
        if context['is_app_moderator']:
            oItem = dict(type='safe', label='Passim editor:')
            oItem['value'] = "yes" if username_is_ingroup(instance.user, app_editor) else "no"
            oItem['field_key'] = 'editrights'
            context['mainitems'].append(oItem)

        # Adapt the permission, if this is the actual user that is logged in
        if context['authenticated'] and not context['is_app_moderator']:
            # Check who this is
            user = self.request.user
            # Compare with instance
            if user.id == instance.user.id:
                self.permission = "write"
                # BUT: some fields may *NOT* be edited even by the user
                for oItem in context['mainitems']:
                    label_name = oItem.get("label").replace(":", "").lower()
                    if label_name in not_editable_by_user:
                        # Remove any field_key or field_list
                        if 'field_key' in oItem: oItem.pop('field_key')
                        if 'field_list' in oItem: oItem.pop('field_list')
            else:
                # This is not the current user
                self.permission = "readonly"
            context['permission'] = self.permission

        # Return the context we have made
        return context

    def after_save(self, form, instance):
        msg = ""
        bResult = True
        oErr = ErrHandle()
        
        try:
            bSaveUser = False

            # (1) Process newusername
            newusername = form.cleaned_data['newusername']
            if newusername != "":
                # Note: the [newusername] is empty, if the user is not allowed to work on this
                if instance.user.username != newusername:
                    instance.user.username = newusername
                    bSaveUser = True

            # (2) Process newemail
            newemail = form.cleaned_data['newemail']
            if instance.user.email != newemail:
                # Validate this
                try:
                    validate_email(newemail)
                    instance.user.email = newemail
                    bSaveUser = True
                except ValidationError:
                    oErr.Status("ProfileEdit/after_save: invalid email")
                    form.add_error("newemail", "invalid email")

            # (3) Process newfirst_name
            newfirst_name = form.cleaned_data['newfirst_name']
            if instance.user.first_name != newfirst_name:
                newfirst_name = newfirst_name.strip()
                if newfirst_name != "" and len(newfirst_name) > 1:
                    instance.user.first_name = newfirst_name
                    bSaveUser = True
                else:
                    form.add_error("newfirst_name", "Invalid first name")

            # (4) Process newlast_name
            newlast_name = form.cleaned_data['newlast_name']
            if instance.user.last_name != newlast_name:
                newlast_name = newlast_name.strip()
                if newlast_name != "" and len(newlast_name) > 1:
                    instance.user.last_name = newlast_name
                    bSaveUser = True
                else:
                    form.add_error("newlast_name", "Invalid last name")

            # (5) Check affiliation
            affiliation = form.cleaned_data['affiliation']
            affiliation = affiliation.strip()
            if len(affiliation) < 2:
                form.add_error("affiliation", "Invalid affiliation")

            # Possibly save the user
            if bSaveUser:
                instance.user.save()

            # (6) 'projects' - approver
            projlist = form.cleaned_data['projlist']
            adapt_m2m(ProjectApprover, instance, "profile", projlist, "project")

            # (7) 'projects' - editor
            editlist = form.cleaned_data['editlist']
            adapt_m2m(ProjectEditor, instance, "profile", editlist, "project")

            # (8) edit rights
            editrights = form.cleaned_data.get('editrights')
            if not editrights is None and editrights in ['yes', 'no']:
                currentrights = "yes" if username_is_ingroup(instance.user, app_editor) else "no"
                if currentrights != editrights:
                    editor_group = Group.objects.filter(name=app_editor).first()
                    if not editor_group is None:
                        if editrights == "yes":
                            # Grant editing rights
                            editor_group.user_set.add(instance.user)
                        else:
                            # Revoke editing rights
                            editor_group.user_set.remove(instance.user)

        except:
            msg = oErr.get_error_message()
            oErr.DoError("ProfileEdit/after_save")
            bResult = False
        return bResult, msg

    def get_history(self, instance):
        return passim_get_history(instance)


class ProfileDetails(ProfileEdit):
    """Like Profile Edit, but then html output"""
    rtype = "html"


class ProfileListView(BasicList):
    """List user profiles"""

    model = Profile
    listform = ProfileForm
    prefix = "prof"
    sg_name = "Profile"     
    plural_name = "Profiles"
    has_select2 = True
    new_button = False      # Do not allow adding new ones here
    order_cols = ['user__username', '', 'ptype', 'affiliation', '', '']
    order_default = order_cols
    order_heads = [
        {'name': 'Username',    'order': 'o=1', 'type': 'str', 'custom': 'name', 'default': "(unnamed)", 'linkdetails': True},
        {'name': 'Email',       'order': '',    'type': 'str', 'custom': 'email', 'linkdetails': True},
        {'name': 'Status',      'order': 'o=3', 'type': 'str', 'custom': 'status', 'linkdetails': True},
        {'name': 'Affiliation', 'order': 'o=4', 'type': 'str', 'custom': 'affiliation', 'main': True, 'linkdetails': True},
        {'name': 'Project Approver',    'order': '',    'type': 'str', 'custom': 'projects'},
        {'name': 'Groups',      'order': '',    'type': 'str', 'custom': 'groups'}]

    filters = [ {"name": "User name",   "id": "filter_username",    "enabled": False},
                {"name": "Email",       "id": "filter_email",       "enabled": False},
               ]
    searches = [
        {'section': '', 'filterlist': [
            {'filter': 'username',  'fkfield': 'user',  'keyS': 'username_ta', 'keyFk': 'username', 'keyList': 'userlist', 'infield': 'username' },
            {'filter': 'email',     'fkfield': 'user',  'keyS': 'email_ta',    'keyFk': 'email' }
         ]} 
            #{'filter': 'project',   'fkfield': 'projects',    'keyFk': 'name', 'keyList': 'projlist', 'infield': 'name'}]},
        ] 

    def initializations(self):
        # Make sure possible adaptations are executed
        listview_adaptations("profile_list")
        # We are not returning anything        
        return None

    def get_field_value(self, instance, custom):
        sBack = ""
        sTitle = ""
        oErr = ErrHandle()
        try:
            if custom == "name":
                sBack = instance.user.username
            elif custom == "email":
                sBack = instance.user.email
            elif custom == "status":
                sBack = instance.get_ptype_display()
            elif custom == "affiliation":
                sBack = "-" if instance.affiliation == None else instance.affiliation
            elif custom == "projects":
                lHtml = []
                for g in instance.projects.all():
                    name = g.name
                    lHtml.append("<span class='badge signature cl'>{}</span>".format(name))
                sBack = ", ".join(lHtml)
            elif custom == "groups":
                lHtml = []
                for g in instance.user.groups.all():
                    name = g.name.replace("passim_", "")
                    lHtml.append("<span class='badge signature gr'>{}</span>".format(name))
                sBack = ", ".join(lHtml)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("ProfileListView/get_field_value")
        return sBack, sTitle


class DefaultEdit(BasicDetails):
    """User-definable defaults for this user-profile"""

    model = Profile
    mForm = ProfileForm
    prefix = "def"
    title = "DefaultEdit"
    titlesg = "Default"
    basic_name = "default"
    has_select2 = True
    history_button = False
    no_delete = True
    mainitems = []

    def custom_init(self, instance):
        self.listview = reverse('mypassim_details')

    def add_to_context(self, context, instance):
        """Add to the existing context"""

        # Define the main items to show and edit
        context['mainitems'] = [
            {'type': 'plain', 'label': "User",              'value': instance.user.id, 'field_key': "user", 'empty': 'idonly'},
            {'type': 'plain', 'label': "Username:",         'value': instance.user.username,                    },
            {'type': 'line',  'label': "Editing rights:",   'value': instance.get_editor_projects_markdown()    },
            {'type': 'line',  'label': "Approver rights:",  'value': instance.get_approver_projects_markdown()  },
            {'type': 'line',  'label': 'Default projects:', 'value': instance.get_defaults_markdown(), 'field_list': 'deflist'}
            ]
        # Return the context we have made
        return context

    def after_save(self, form, instance):
        msg = ""
        bResult = True
        oErr = ErrHandle()
        
        try:
            # (6) 'default projects'
            deflist = form.cleaned_data['deflist']
            instance.defaults_update(deflist)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("DefaultEdit/after_save")
            bResult = False
        return bResult, msg


class DefaultDetails(DefaultEdit):
    """Like Default Edit, but then html output"""

    rtype = "html"


# ================= PROJECT =============================

class ProjectEdit(BasicDetails):
    """Details and editing of a project (nov 2021 version)"""

    model = Project2
    mForm = ProjectForm
    prefix = 'proj'
    title = "Project"
    # no_delete = True
    history_button = True
    mainitems = []

    # How to handle the app_moderator

    def add_to_context(self, context, instance):
        """Add to the existing context"""

        def get_singles(lst_id, clsThis, field):
            """Get a list of occurrances that have only one project id"""

            # Get all singles
            lst_singles = clsThis.objects.all().values(field).annotate(total=Count("project")).filter(total=1).values(field, "total")
            # Turn them into a dictionary - but only the singular ones
            dict_singles = { x[field]:x['total'] for x in lst_singles}
            # Count the ones that overlap: those are the singular ones
            count = 0
            attr = "{}__id".format(field)
            for oItem in lst_id:
                id = oItem[attr]
                if id in dict_singles:
                    count += 1
            return count

        oErr = ErrHandle()
        try:
            
            # Is the user allowed to edit/delete or not?
            bMayDelete = user_is_ingroup(self.request, app_moderator) or user_is_superuser(self.request)

            # Standard behaviour: no_delete to True
            if not bMayDelete:
                self.no_delete = True
                context['no_delete'] = self.no_delete
                self.permission = "readonly"
                context['permission'] = self.permission

            # Define the main items to show and edit
            context['mainitems'] = [
                {'type': 'plain', 'label': "Name:",             'value': instance.name              },        #, 'field_key': "name"},
                {'type': 'safe',  'label': "Description:",      'value': instance.get_description() },        #, 'field_key': "name"},
                {'type': 'line',  'label': "Approval rights:",  'value': instance.get_editor_markdown(),
                    'title': 'All the current users that have approval rights for this project'     },
                {'type': 'safe',  'label': "Contact:",          'value': instance.get_contact_info(),
                    'title': 'The project editor that can be contacted for details'},
                ]   
        
            if bMayDelete:
                context['mainitems'][0]['field_key'] = "name"
                context['mainitems'][1]['field_key'] = "description"

            # Also add a delete Warning Statistics message (see issue #485)
            lst_proj_m = ManuscriptProject.objects.filter(project=instance).values('manuscript__id')
            lst_proj_hc = CollectionProject.objects.filter(project=instance).values('collection__id')
            lst_proj_s = SermonDescrProject.objects.filter(project=instance).values('sermon__id')
            lst_proj_ssg = EqualGoldProject.objects.filter(project=instance).values('equal__id')

            count_m =  len(lst_proj_m)
            count_hc =  len(lst_proj_hc)
            count_s =  len(lst_proj_s)
            count_ssg = len(lst_proj_ssg)
            single_m = get_singles(lst_proj_m, ManuscriptProject, "manuscript")
            single_hc = get_singles(lst_proj_hc, CollectionProject, "collection")
            single_s = get_singles(lst_proj_s, SermonDescrProject, "sermon")
            single_ssg = get_singles(lst_proj_ssg, EqualGoldProject, "equal")
            
            local_context = dict(
                project=instance, 
                count_m=count_m, count_hc=count_hc, count_s=count_s, count_ssg=count_ssg,
                single_m=single_m, single_hc=single_hc, single_s=single_s, single_ssg=single_ssg,
                )
            context['delete_message'] = render_to_string('seeker/project_statistics.html', local_context, self.request)

        except:
            msg = oErr.get_error_message()
            oErr.DoError("ProjectEdit/add_to_context")

        # Return the context we have made
        return context
    
    def get_history(self, instance):
        return passim_get_history(instance)


class ProjectDetails(ProjectEdit):
    """Just the HTML page"""
    rtype = "html"


class ProjectListView(BasicList):
    """Search and list projects"""

    model = Project2 
    listform = ProjectForm
    prefix = "proj"
    has_select2 = True
    paginate_by = 20
    sg_name = "Project"     
    plural_name = "Projects"
    # page_function = "ru.passim.seeker.search_paged_start"
    order_cols = ['name', '']
    order_default = order_cols
    order_heads = [{'name': 'Project',                'order': 'o=1', 'type': 'str', 'custom': 'project',   'main': True, 'linkdetails': True},
                   {'name': 'Manuscripts',            'order': 'o=2', 'type': 'str', 'custom': 'manulink',  'align': 'right' },
                   {'name': 'Manifestations',         'order': 'o=3', 'type': 'str', 'custom': 'sermolink', 'align': 'right'},
                   {'name': 'Authority files',        'order': 'o=4', 'type': 'str', 'custom': 'ssglink',   'align': 'right'},
                   {'name': 'Historical collections', 'order': 'o=5', 'type': 'str', 'custom': 'hclink',    'align': 'right'}]
                   
    filters = [ {"name": "Project",         "id": "filter_project",     "enabled": False},
                {"name": "Shelfmark",       "id": "filter_manuid",      "enabled": False, "head_id": "filter_other"},
               ]
    searches = [
        {'section': '', 'filterlist': [
            {'filter': 'project',   'dbfield': 'name',         'keyS': 'project_ta', 'keyList': 'projlist', 'infield': 'name' }]} 
            #{'filter': 'project',   'fkfield': 'projects',    'keyFk': 'name', 'keyList': 'projlist', 'infield': 'name'}]},
        ] 

    # hier gaat het nog niet goed
    def get_field_value(self, instance, custom):
        sBack = ""
        sTitle = ""
        html = []
        oErr = ErrHandle()
        try:
            if custom == "manulink":
                # Link to the manuscripts in this project
                count = instance.project2_manuscripts.exclude(mtype="tem").count()
                url = reverse('search_manuscript')
                if count > 0:
                    html.append("<a href='{}?manu-projlist={}'><span class='badge jumbo-3 clickable' title='{} manuscripts in this project'>{}</span></a>".format(
                        url, instance.id, count, count))

            elif custom == "sermolink":
                # Link to the sermons in this project
                count = instance.project2_sermons.count() 
                url = reverse('search_sermon')
                if count > 0:                 
                    html.append("<a href='{}?sermo-projlist={}'><span class='badge jumbo-3 clickable' title='{} sermons in this project'>{}</span></a>".format(
                        url, instance.id, count, count))
            
            elif custom == "ssglink":
                # Link to the authority files in this project
                count = instance.project2_equalgold.count() 
                url = reverse('equalgold_list')
                if count > 0:                 
                    html.append("<a href='{}?ssg-projlist={}'><span class='badge jumbo-3 clickable' title='{} authority files in this project'>{}</span></a>".format(
                        url, instance.id, count, count))

            elif custom == "hclink":
                # Link to the historical collections in this project
                count = instance.project2_collection.exclude(settype="pd").count() # Nog expliciet met HC rekening houden?
                url = reverse('collhist_list')
                if count > 0:                 
                    html.append("<a href='{}?hist-projlist={}'><span class='badge jumbo-3 clickable' title='{} historical collections in this project'>{}</span></a>".format(
                        url, instance.id, count, count))

            elif custom == "project":
                sName = instance.name
                if sName == "": sName = "<i>(unnamed)</i>"
                html.append(sName)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("ProjectListView/get_field_value")
        # Combine the HTML code
        sBack = "\n".join(html)
        return sBack, sTitle


# ================= COLLECTION =============================

class CollAnyEdit(BasicDetails):
    """Manu: Manuscript collections"""

    model = Collection
    mForm = CollectionForm
    prefix = "any"
    basic_name_prefix = "coll"
    rtype = "json"
    settype = "pd"
    title = "Any collection"
    history_button = True
    comment_button = True
    manu = None
    codico = None
    datasettype = ""
    use_team_group = True
    max_items = 500            # Maximum number of items shown in details view
    mainitems = []
    hlistitems = [
        {'type': 'manu',    'clsColl': CollectionMan,   'field': 'manuscript'},
        {'type': 'sermo',   'clsColl': CollectionSerm,  'field': 'sermon'},
        {'type': 'gold',    'clsColl': CollectionGold,  'field': 'gold'},
        {'type': 'super',   'clsColl': CollectionSuper, 'field': 'super'},
        ]

    ClitFormSet = inlineformset_factory(Collection, LitrefCol,
                                         form = CollectionLitrefForm, min_num=0,
                                         fk_name = "collection",
                                         extra=0, can_delete=True, can_order=False)

    ColdrhcFormSet = inlineformset_factory(Collection, DaterangeHistColl,
                                         form=DaterangeHistCollForm, min_num=0,
                                         fk_name = "histcoll",
                                         extra=0, can_delete=True, can_order=False)

    stype_edi_fields = ['name', 'owner', 'readonly', 'type', 'settype', 'descrip', 'url', 'path', 'scope', 
                        'LitrefMan', 'litlist', 
                        'projlist', 
                        'Daterange', 'datelist']
    
    def custom_init(self, instance):
        if instance != None and instance.settype == "hc":
            # Check if the 'clit' and the 'coldrhc' are already in the formset_objects or not
            bFound1 = False
            bFound2 = False
            for oItem in self.formset_objects: 
                if oItem['prefix'] == "clit" or oItem['prefix'] == "coldrhc":
                    bFound1 = True
                    bFound2 = True         
                    break
            if not bFound1:
                self.formset_objects.append(
                    {'formsetClass': self.ClitFormSet,  'prefix': 'clit',  'readonly': False, 'noinit': True, 'linkfield': 'collection'})
            if not bFound2:
                self.formset_objects.append(
                    {'formsetClass': self.ColdrhcFormSet,  'prefix': 'coldrhc',  'readonly': False, 'noinit': True, 'linkfield': 'collection'})       
        if instance != None:
            self.datasettype = instance.type
        return None

    def check_hlist(self, instance):
        """Check if a hlist parameter is given, and hlist saving is called for"""

        oErr = ErrHandle()

        try:
            arg_hlist = instance.type + "-hlist"
            arg_savenew = instance.type + "-savenew"
            if arg_hlist in self.qd and arg_savenew in self.qd:
                # Interpret the list of information that we receive
                hlist = json.loads(self.qd[arg_hlist])
                # Interpret the savenew parameter
                savenew = self.qd[arg_savenew]

                # Make sure we are not saving
                self.do_not_save = True
                # But that we do a new redirect
                self.newRedirect = True

                # Action depends on the particular prefix
                for hlistitem in self.hlistitems:
                    # Is this the one?
                    if hlistitem['type'] == instance.type:
                        # THis is the type
                        clsColl = hlistitem['clsColl']
                        field = hlistitem['field']

                        # First check if this needs to be a *NEW* collection instance
                        if savenew == "true":
                            profile = Profile.get_user_profile(self.request.user.username)
                            # Yes, we need to copy the existing collection to a new one first
                            original = instance
                            instance = original.get_copy(owner=profile)

                        # Change the redirect URL
                        if self.redirectpage == "":
                            this_type = ""
                            if instance.settype == "hc": this_type = "hist"
                            elif instance.scope == "priv": this_type = "priv"
                            else: this_type = "publ" 
                            self.redirectpage = reverse('coll{}_details'.format(this_type), kwargs={'pk': instance.id})
                        else:
                            self.redirectpage = self.redirectpage.replace(original.id, instance.id)

                        # What we have is the ordered list of Manuscript id's that are part of this collection
                        with transaction.atomic():
                            # Make sure the orders are correct
                            for idx, item_id in enumerate(hlist):
                                order = idx + 1
                                lstQ = [Q(collection=instance)]
                                lstQ.append(Q(**{"{}__id".format(field): item_id}))
                                obj = clsColl.objects.filter(*lstQ).first()
                                if obj != None:
                                    if obj.order != order:
                                        obj.order = order
                                        obj.save()
                        # See if any need to be removed
                        existing_item_id = [str(getattr(x, field).id) for x in clsColl.objects.filter(collection=instance)]
                        delete_id = []
                        for item_id in existing_item_id:
                            if not item_id in hlist:
                                delete_id.append(item_id)
                        if len(delete_id)>0:
                            lstQ = [Q(collection=instance)]
                            lstQ.append(Q(**{"{}__id__in".format(field): delete_id}))
                            clsColl.objects.filter(*lstQ).delete()

            return True
        except:
            msg = oErr.get_error_message()
            oErr.DoError("CollAnyEdit/check_hlist")
            return False
    
    def add_to_context(self, context, instance):
        """Add to the existing context"""

        prefix_scope = ['any', 'manu', 'sermo', 'gold', 'super', 'priv', 'publ']
        prefix_type = ['any', 'manu', 'sermo', 'gold', 'super', 'priv', 'publ']
        prefix_readonly = ['any', 'manu', 'sermo', 'gold', 'super']
        prefix_elevate = ['any', 'super', 'priv', 'publ']

        oErr = ErrHandle()
        try:

            # Need to know who this is
            profile = Profile.get_user_profile(self.request.user.username)

            # Prepare saveditem handling
            saveditem_button = get_saveditem_html(self.request, instance, profile, sitemtype=self.settype)
            saveditem_form = get_saveditem_html(self.request, instance, profile, "form", sitemtype=self.settype)

            # Define the main items to show and edit
            context['mainitems'] = [
                # Some items must be passed on
                {'type': 'plain', 'label': "Type:",        'value': instance.get_type_name(),   'field_key': 'type',   'empty': 'hide'},
                {'type': 'plain', 'label': "Owner id:",    'value': instance.owner.id,          'field_key': 'owner',   'empty': 'hide'},

                # Regular visible fields
                {'type': 'plain', 'label': "Name:",        'value': instance.name, 'field_key': 'name'},
                {'type': 'safe',  'label': "Saved item:",  'value': saveditem_button          },
                {'type': 'plain', 'label': "Description:", 'value': instance.descrip, 'field_key': 'descrip'},
                {'type': 'plain', 'label': "URL:",         'value': instance.url, 'field_key': 'url'}, 
                ]

            # Optionally add Scope: but only for the actual *owner* of this one
            # Issue #599: don't offer scope for HC
            if instance.settype != "hc" and self.prefix in prefix_scope and not instance.owner is None \
                and instance.owner.user == self.request.user:
                context['mainitems'].append(
                {'type': 'plain', 'label': "Scope:",       'value': instance.get_scope_display, 'field_key': 'scope'})

            # Always add Type, but its value may not be changed
            context['mainitems'].append(
                {'type': 'plain', 'label': "Type:",        'value': instance.get_type_name()})

            # Always add project label(s)
            #context['mainitems'].append(
            #    {'type': 'plain', 'label': "Project:",     'value': instance.get_project_markdown(), 'field_key': 'project'})

            # Optionally add Readonly
            if self.prefix in prefix_readonly:
                context['mainitems'].append(
                {'type': 'plain', 'label': "Readonly:",    'value': instance.readonly, 'field_key': 'readonly'})

            # This is only for private PDs:
            if self.prefix == "priv" and instance != None and instance.settype == "pd" and instance.id != None:
                name_choice = dict(
                    manu=dict(sg_name="Manuscript", pl_name="Manuscripts"),
                    sermo=dict(sg_name="Manifestation", pl_name="Manifestations"),
                    gold=dict(sg_name="Sermon Gold", pl_name="Sermons Gold"),
                    super=dict(sg_name="Authority file", pl_name="Authority files")
                    )
                # Add a button + text
                context['datasettype'] = instance.type
                context['sg_name'] = name_choice[instance.type]['sg_name']
                context['pl_name'] = name_choice[instance.type]['pl_name']
            
                context['size'] = instance.get_size_markdown()
                size_value = render_to_string("seeker/collpriv.html", context, self.request)
            else:
                size_value = instance.get_size_markdown()
        
            # Always add Created (with name of the creator) and Size                
            context['mainitems'].append( {'type': 'plain', 'label': "Created:", 'value': instance.get_created_user}) 
            context['mainitems'].append( {'type': 'line',  'label': "Size:", 'value': size_value})

            # Add any saved items of this type
            if instance.settype == "pd":
                # The kind of list depends on the kind of dataset that I am
                context['mainitems'].append( {'type': 'plain',  'label': "Saved items:", 'value': instance.get_sitems(profile),
                    'field_list': 'sitemlist_{}'.format(instance.type)})

            # If this is a historical collection,and an app-editor gets here, add a link to a button to create a manuscript
            if instance.settype == "hc" and context['is_app_editor']:
                # If 'manu' is set, then this procedure is called from 'collhist_compare'
                if self.manu == None:
                    context['mainitems'].append({'type': 'safe', 'label': 'Manuscript', 'value': instance.get_manuscript_link()})
            
            # Historical collections may have literature references and dates   
            if instance.settype == "hc" and len(self.formset_objects) > 0 and len(self.formset_objects[0]) > 0:
                oLitref = {'type': 'plain', 'label': "Literature:",   'value': instance.get_litrefs_markdown()}
                oDate = {'type': 'plain', 'label': "Date:",   'value': instance.get_date_markdown()}

                if context['is_app_editor']:
                    # Literature references
                    oLitref['multiple'] = True
                    oLitref['field_list'] = 'litlist'
                    oLitref['fso'] = self.formset_objects[0]
                    oLitref['template_selection'] = 'ru.passim.litref_template'
                    
                    # Dates
                    oDate['multiple'] = True
                    oDate['field_list'] = 'datelist'
                    oDate['fso'] = self.formset_objects[1]

                context['mainitems'].append(oLitref)
                context['mainitems'].append(oDate)  

            # Historical collections have a project assigned to them
            if instance.settype == "hc":
                oProject =  {'type': 'plain', 'label': "Project:",     'value': instance.get_project_markdown2()}
                if may_edit_project(self.request, profile, instance):
                    oProject['field_list'] = 'projlist'
                context['mainitems'].append(oProject)        
                        
            # Add comment modal stuff
            if instance.settype == "hc":
                lhtml = []

                initial = dict(otype="hc", objid=instance.id, profile=profile)
                context['commentForm'] = CommentForm(initial=initial, prefix="com")
                context['comment_list'] = get_usercomments('hc', instance, profile)            
                lhtml.append(render_to_string("seeker/comment_add.html", context, self.request))
                context['comment_count'] = instance.comments.count()

                # Also add the saved item form
                lhtml.append(saveditem_form)

                # Store the after_details in the context
                context['after_details'] = "\n".join(lhtml)    
            else:
                lhtml = []                    
                # Also add the saved item form
                lhtml.append(saveditem_form)

                # Store the after_details in the context
                context['after_details'] = "\n".join(lhtml)    

                # If this is not a HC, then no comments are needed
                self.comment_button = False
                context['comment_button'] = False

            # Possible add URI
            if instance.settype == "hc":
                # Make stable URI available
                context['stable_uri'] = self.get_abs_uri('collhist_details', instance)

            # Any dataset may optionally be elevated to a historical collection
            # BUT: only if a person has permission
            if instance.settype == "pd" and self.prefix in prefix_elevate and instance.type in prefix_elevate and \
                context['authenticated'] and context['is_app_editor']:
                context['mainitems'].append(
                    {'type': 'safe', 'label': "Historical:", 'value': instance.get_elevate()}
                    )
            # Buttons to switch to a listview of M/S/SG/SSG based on this collection
            context['mainitems'].append(
                    {'type': 'safe', 'label': "Listviews:", 'value': self.get_listview_buttons(instance),
                     'title': 'Open a listview that is filtered on this dataset'}
                    )
            # For HC: buttons to switch between related listviews
            if instance.settype == "hc" and context['is_app_editor'] and self.manu == None and self.codico == None:
                context['mainitems'].append(
                        {'type': 'safe', 'label': "Show/hide:", 'value': self.get_hc_buttons(instance),
                         'title': 'Optionally show and edit the Authority files in this collection'}
                        )

            # Signal that we have select2
            context['has_select2'] = True

            # Determine what the permission level is of this collection for the current user
            # (1) Is this user a different one than the one who created the collection?
            profile_owner = instance.owner
            profile_user = Profile.get_user_profile(self.request.user.username)
            # (2) Set default permission
            permission = "readonly"
            if self.may_edit(context):
                if profile_owner.id == profile_user.id:
                    # (3) Any creator of the collection may write it
                    permission = "write"
                else:
                    # (4) permission for different users
                    if context['is_app_editor']:
                        # (5) what if the user is an app_editor?
                        if instance.scope == "publ":
                            # Editors may read/write collections with 'public' scope
                            permission = "write"
                        elif instance.scope == "team":
                            # Editors may read collections with 'team' scope
                            permission = "read"
                    else:
                        # (5) any other users
                        if instance.scope == "publ":
                            # All users may read collections with 'public' scope
                            permission = "read"
            else:
                permission = "readonly"

            context['permission'] = permission

            # adapt visibility of revision history (issue #451)
            if permission == "readonly":
                self.history_button = False
                context['history_button'] = False
        except:
            msg = oErr.get_error_message()
            oErr.DoError("CollAnyEdit/add_to_context")
          
        # Return the context we have made
        return context    

    def get_hc_buttons(self, instance):
        """Get buttons to show/hide the two HC listviews (SSGs and Codicological units)"""

        sBack = ""
        lHtml = []
        abbr = None
        button_list = [
            {'label': 'Authority files', 'id': 'basic_super_set', 'show': False},
            {'label': 'Codicological units','id': 'basic_codi_set',  'show': True},
            ]
        oErr = ErrHandle()
        try:
            for oButton in button_list:
                lHtml.append("<a role='button' class='btn btn-xs jumbo-1' data-toggle='collapse' data-target='#{}' >{}</a>".format(
                    oButton['id'], oButton['label']))
            sBack = "<span>&nbsp;</span>".join(lHtml)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("CollAnyEdit/get_hc_buttons")

        return sBack

    def get_listview_buttons(self, instance):
        """Create an HTML list of buttons for M/S/SG/SSG listviews filtered on this collection"""

        sBack = ""
        context = {}
        abbr = None
        oErr = ErrHandle()
        try:
            url_m = reverse("manuscript_list")
            url_s = reverse("sermon_list")
            url_sg = reverse("gold_list")
            url_ssg = reverse("equalgold_list")
            if instance.type == "manu":
                # collection of manuscripts
                abbr = "m"
            elif instance.type == "sermo":
                # collection of sermons
                abbr = "s"
            elif instance.type == "gold":
                # collection of gold sermons
                abbr = "sg"
            elif instance.type == "super":
                # collection of SSG
                abbr = "ssg"
                if self.settype == "hc": abbr = "hist"
            if url_m != None and url_s != None and url_sg != None and url_ssg != None and abbr != None:
                context['url_manu'] = "{}?manu-collist_{}={}".format(url_m, abbr, instance.id)
                context['url_sermo'] = "{}?sermo-collist_{}={}".format(url_s, abbr, instance.id)
                context['url_gold'] = "{}?gold-collist_{}={}".format(url_sg, abbr, instance.id)
                context['url_super'] = "{}?ssg-collist_{}={}".format(url_ssg, abbr, instance.id) # TH zit het hier?
                #if self.settype == "hc":
                #    context['url_super'] = "{}?ssg-collist_{}={}".format(url_ssg, "hc", instance.id)
                #else:
                #    context['url_super'] = "{}?ssg-collist_{}={}".format(url_ssg, abbr, instance.id)

                sBack = render_to_string('seeker/coll_buttons.html', context, self.request)

        except:
            msg = oErr.get_error_message()
            oErr.DoError("CollAnyEdit/get_listview_buttons")

        return sBack
    
    def process_formset(self, prefix, request, formset): 
        errors = []
        bResult = True
        instance = formset.instance
        for form in formset: # Hier komt het niet verder 
            if form.is_valid():
                cleaned = form.cleaned_data
                # Action depends on prefix
                if prefix == "clit":
                    # Literature reference processing
                    newpages = cleaned.get("newpages")
                    # Also get the litref
                    oneref = cleaned.get("oneref")
                    if not oneref is None:
                        # Double check if this combination [oneref/newpages] is already connected to [instance] or not
                        obj = LitrefCol.objects.filter(reference=oneref, pages=newpages, collection=instance).first()
                        if obj is None:
                            # It is not there yet: add it
                            form.instance.reference = oneref
                            if newpages:
                                form.instance.pages = newpages
                    # Note: it will get saved with form.save()
                elif prefix == 'coldrhc':
                    # Processing one daterange
                    newstart = cleaned.get('newstart', None)
                    newfinish = cleaned.get('newfinish', None)
                    if newstart:
                        # Possibly set newfinish equal to newstart
                        if newfinish == None or newfinish == "":
                            newfinish = newstart
                        # Double check if this one already exists for the current instance
                        obj = instance.histcoll_dateranges.filter(yearstart=newstart, yearfinish=newfinish).first()
                        if obj == None:
                            form.instance.yearstart = int(newstart)
                            form.instance.yearfinish = int(newfinish)                 

            else:
                errors.append(form.errors)
                bResult = False
        return None

    def before_save(self, form, instance):
        oErr = ErrHandle()
        bBack = True
        msg = ""
        try:
            if form != None and instance != None:
                # Search the user profile
                profile = Profile.get_user_profile(self.request.user.username)
                form.instance.owner = profile
                # The collection type is now a parameter
                type = form.cleaned_data.get("type", "")
                if type == "":
                    if self.prefix == "hist":
                        form.instance.type = "super"
                    elif self.prefix == "publ":
                        form.instance.type = self.datasettype
                    elif self.prefix == "priv":
                        type = self.qd.get("datasettype", "")
                        if type == "": type = self.datasettype
                        if type == "": type = "super"
                        form.instance.type = type
                    else:
                        form.instance.type = self.prefix

                # Check out the name, if this is not in use elsewhere
                if instance.id != None:
                    name = form.instance.name
                    if Collection.objects.filter(name__iexact=name).exclude(id=instance.id).exists():
                        # The name is already in use, so refuse it.
                        msg = "The name '{}' is already in use for a dataset. Please chose a different one".format(name)
                        return False, msg
        except:
            msg = oErr.get_error_message()
            oErr.DoError("CollAnyEdit/before_save")
            bBack = False
        return bBack, msg

    def after_save(self, form, instance):
        msg = ""
        bResult = True
        oErr = ErrHandle()
        
        try:
            # Process many-to-many changes: Add and remove relations in accordance with the new set passed on by the user
            # (1) 'literature'
            litlist = form.cleaned_data['litlist']
            adapt_m2m(LitrefCol, instance, "collection", litlist, "reference", extra=['pages'], related_is_through = True)
            
            # (2) 'projects'
            projlist = form.cleaned_data['projlist']
            col_proj_deleted = []
            adapt_m2m(CollectionProject, instance, "collection", projlist, "project", deleted=col_proj_deleted)
            project_dependant_delete(self.request, col_proj_deleted)

            # (3) Process adding items from ['sitemlist_XXX']
            sitemlist_sermo = form.cleaned_data['sitemlist_sermo']
            instance.add_sitems(sitemlist_sermo, "sermo")
            sitemlist_manu = form.cleaned_data['sitemlist_manu']
            instance.add_sitems(sitemlist_manu, "manu")
            sitemlist_super = form.cleaned_data['sitemlist_super']
            instance.add_sitems(sitemlist_super, "super")
                        
            # (4) Process many-to-ONE changes
            # (1) links from Daterange to Codico
            datelist = form.cleaned_data['datelist']
            adapt_m2o(DaterangeHistColl, instance, "histcoll", datelist)

            # Make sure to process changes
            instance.refresh_from_db()     


        except:
            msg = oErr.get_error_message()
            bResult = False
        return bResult, msg

    def action_add(self, instance, details, actiontype):
        """User can fill this in to his/her liking"""
        passim_action_add(self, instance, details, actiontype)

    def get_history(self, instance):
        return passim_get_history(instance)

    def get_histogram_data(self, instance=None, qs=None, listview="collist_hist", method='d3'):
        """Get data to make a histogram"""

        oErr = ErrHandle()
        histogram_data = []
        try:
            # Get the queryset for this view
            if instance != None and qs != None:
                # Get the base url
                baseurl = reverse('equalgold_list')
                # Determine the list
                qs = qs.order_by('scount').values('scount', 'id')
                scount_index = {}
                frequency = None
                for item in qs:
                    scount = item['scount']
                    if frequency == None or frequency != scount:
                        # Initialize the frequency
                        frequency = scount
                        # Add to the histogram data
                        histogram_data.append(dict(scount=scount, freq=1))
                    else:
                        histogram_data[-1]['freq'] += 1

                # Determine the targeturl for each histogram bar
                for item in histogram_data:
                    targeturl = "{}?ssg-{}={}&ssg-soperator=exact&ssg-scount={}".format(baseurl, listview, instance.id, item['scount'])
                    item['targeturl'] = targeturl
                # D3-specific
                if method == "d3":
                    histogram_data = json.dumps(histogram_data)
            
        except:
            msg = oErr.get_error_message()
            oErr.DoError("get_histogram_data")
        return histogram_data


class CollPrivEdit(CollAnyEdit):
    prefix = "priv"
    basic_name = "collpriv"
    title = "My Dataset"


class CollPublEdit(CollAnyEdit):
    prefix = "publ"
    basic_name = "collpubl"
    title = "Public Dataset"


class CollHistEdit(CollAnyEdit):
    prefix = "super"
    settype = "hc"
    basic_name = "collhist"
    title = "Historical Collection"

    #ColdrhcFormSet = inlineformset_factory(Collection, DaterangeHistColl,
    #                                     form=DaterangeHistCollForm, min_num=0,
    #                                     fk_name = "histcoll",
    #                                     extra=0, can_delete=True, can_order=False)

    #formset_objects = [{'formsetClass': ColdrhcFormSet,   'prefix': 'coldrhc',   'readonly': False, 'noinit': True, 'linkfield': 'histcoll'}] # dit goed?

    #stype_edi_fields = ['Daterange', 'datelist']                       


    def before_save(self, form, instance):
        # Make sure the [CollAnyEdit] is executed
        response = super(CollHistEdit, self).before_save(form, instance)

        # Now do the remainder
        oErr = ErrHandle()
        bBack = True
        msg = ""
        try:
            if instance != None:
                # Need to know who is 'talking'...
                username = self.request.user.username
                profile = Profile.get_user_profile(username)

                # Issue #473: automatic assignment of project for particular editor(s)
                projlist = form.cleaned_data.get("projlist")
                bBack, msg = evaluate_projlist(profile, instance, projlist, "Historical collection")
        except:
            msg = oErr.get_error_message()
            oErr.DoError("CollHistEdit/before_save")
            bBack = False
        return bBack, msg

    def after_save(self, form, instance):
        # Make sure the [CollAnyEdit] is executed
        response = super(CollHistEdit, self).after_save(form, instance)

        # Now do the remainder
        msg = ""
        bResult = True
        oErr = ErrHandle()
        
        try:
            # Issue #473: default project assignment
            if instance.projects.count() == 0:
                # Need to know who is 'talking'...
                username = self.request.user.username
                profile = Profile.get_user_profile(username)

                # The user has not selected a project (yet): try default assignment
                # user_projects = profile.projects.all()
                user_projects = profile.get_defaults()
                if user_projects.count() == 1:
                    project = user_projects.first()
                    CollectionProject.objects.create(collection=instance, project=project)
        except:
            msg = oErr.get_error_message()
            bResult = False
        return bResult, msg


class CollManuEdit(CollAnyEdit):
    """Manu: Manuscript collections"""

    prefix = "manu"
    title = "Manuscript collection"


class CollSermoEdit(CollAnyEdit):
    """Sermo: SermonDescr collections """

    prefix = "sermo"
    title = "Sermon collection"


class CollGoldEdit(CollAnyEdit):
    """Gold: SermonGold collections """

    prefix = "gold"
    title = "Gold collection"


class CollSuperEdit(CollAnyEdit):
    """Super: EqualGold collections = super sermon gold """

    prefix = "super"
    title = "Super collection"


class CollAnyDetails(CollAnyEdit):
    """Like CollAnyEdit, but then with html"""
    rtype = "html"


class CollPrivDetails(CollAnyEdit):
    """Like CollPrivEdit, but then with html"""
    
    prefix = "priv"
    basic_name = "collpriv"
    title = ""
    rtype = "html"
    custombuttons = []

    def custom_init(self, instance):
        if instance != None:
            # Check if someone acts as if this is a public dataset, while it is not
            #if instance.settype == "pd":
            #    # Determine what kind of dataset/collection this is          
            #    if instance.scope == "publ":
            #        self.title = "Public Dataset"
            #        if instance.owner != Profile.get_user_profile(self.request.user.username):
            #           # It is a public dataset after all! er is geen owner
            #           self.redirectpage = reverse("collpubl_details", kwargs={'pk': instance.id}) # priv ipv publ
            if instance.settype == "hc":
                # This is a historical collection
                self.redirectpage = reverse("collhist_details", kwargs={'pk': instance.id})

            if instance.type == "super":
                self.custombuttons = [{"name": "scount_histogram", "title": "Sermon Histogram", 
                      "icon": "th-list", "template_name": "seeker/scount_histogram.html" }]
            
            # Give the title correct name using the scope of each collection:
            # For public datasets:
            if instance.scope == "publ":
                self.title = "Public Dataset"
            # For team datasets:
            elif instance.scope == "team":
                self.title = "Team Dataset"
            # For private datasets:
            else:
                self.title = "Private Dataset"
            
            # Check for hlist saving
            self.check_hlist(instance)
        return None

    def add_to_context(self, context, instance):
        # Perform the standard initializations:
        context = super(CollPrivDetails, self).add_to_context(context, instance)

        def add_one_item(rel_item, value, resizable=False, title=None, align=None, link=None, main=None, draggable=None):
            oAdd = dict(value=value)
            if resizable: oAdd['initial'] = 'small'
            if title != None: oAdd['title'] = title
            if align != None: oAdd['align'] = align
            if link != None: oAdd['link'] = link
            if main != None: oAdd['main'] = main
            if draggable != None: oAdd['draggable'] = draggable
            rel_item.append(oAdd)
            return True

        def check_order(qs):
            with transaction.atomic():
                for idx, obj in enumerate(qs):
                    if obj.order < 0:
                        obj.order = idx + 1
                        obj.save()

        # All PDs: show the content
        related_objects = []
        lstQ = []
        rel_list =[]
        resizable = True
        index = 1
        sort_start = '<span class="sortable"><span class="fa fa-sort sortshow"></span>&nbsp;'
        sort_start_int = '<span class="sortable integer"><span class="fa fa-sort sortshow"></span>&nbsp;'
        sort_end = '</span>'

        oErr = ErrHandle()

        try:

            # Action depends on instance.type: M/S/SG/SSG
            if instance.type == "manu":
                # Get all non-template manuscripts part of this PD
                manuscripts = dict(title="Manuscripts within this dataset", prefix="manu")
                if resizable: manuscripts['gridclass'] = "resizable dragdrop"
                manuscripts['savebuttons'] = True
                manuscripts['saveasbutton'] = True

                # Check ordering
                qs_manu = instance.manuscript_col.all().order_by(
                        'order', 'manuscript__lcity__name', 'manuscript__library__name', 'manuscript__idno')
                check_order(qs_manu)

                for idx, obj in enumerate(qs_manu):
                    rel_item = []
                    item = obj.manuscript

                    # Leave if this is too much
                    if idx > self.max_items:
                        break

                    # S: Order in Manuscript
                    #add_one_item(rel_item, index, False, align="right", draggable=True)
                    #index += 1
                    add_one_item(rel_item, obj.order, False, align="right", draggable=True)

                    # Shelfmark = IDNO
                    add_one_item(rel_item,  self.get_field_value("manu", item, "shelfmark"), False, title=item.idno, main=True, 
                                 link=reverse('manuscript_details', kwargs={'pk': item.id}))

                    # Just the name of the manuscript
                    add_one_item(rel_item, self.get_field_value("manu", item, "name"), resizable)

                    # Origin
                    add_one_item(rel_item, self.get_field_value("manu", item, "orgprov"), False, 
                                 title="Origin (if known), followed by provenances (between brackets)")

                    # date range
                    add_one_item(rel_item, self.get_field_value("manu", item, "daterange"), False, align="right")

                    # Number of sermons in this manuscript
                    add_one_item(rel_item, self.get_field_value("manu", item, "sermons"), False, align="right")

                    # Actions that can be performed on this item
                    add_one_item(rel_item, self.get_actions())

                    # Add this line to the list
                    rel_list.append(dict(id=item.id, cols=rel_item))

                manuscripts['rel_list'] = rel_list

                manuscripts['columns'] = [
                    '{}<span title="Default order">Order<span>{}'.format(sort_start_int, sort_end),
                    '{}<span title="City/Library/Shelfmark">Shelfmark</span>{}'.format(sort_start, sort_end), 
                    '{}<span title="Name">Name</span>{}'.format(sort_start, sort_end), 
                    '{}<span title="Origin/Provenance">or./prov.</span>{}'.format(sort_start, sort_end), 
                    '{}<span title="Date range">date</span>{}'.format(sort_start_int, sort_end), 
                    '{}<span title="Sermons in this manuscript">sermons</span>{}'.format(sort_start_int, sort_end),
                    ''
                    ]
                related_objects.append(manuscripts)

            elif instance.type == "sermo":
                # Get all sermons that are part of this PD
                sermons = dict(title="Manifestations within this dataset", prefix="sermo")
                if resizable: sermons['gridclass'] = "resizable dragdrop"
                sermons['savebuttons'] = True
                sermons['saveasbutton'] = True

                qs_sermo = instance.sermondescr_col.all().order_by(
                        'order', 'sermon__author__name', 'sermon__siglist', 'sermon__srchincipit', 'sermon__srchexplicit')
                check_order(qs_sermo)

                # Walk these collection sermons
                for idx, obj in enumerate(qs_sermo):
                    rel_item = []
                    item = obj.sermon

                    # Leave if this is too much
                    if idx > self.max_items:
                        break

                    # S: Order in Sermon
                    #add_one_item(rel_item, index, False, align="right")
                    #index += 1
                    add_one_item(rel_item, obj.order, False, align="right", draggable=True)

                    # S: Author
                    add_one_item(rel_item, self.get_field_value("sermo", item, "author"), False, main=True)

                    # S: Signature
                    add_one_item(rel_item, self.get_field_value("sermo", item, "signature"), False)

                    # S: Inc+Expl
                    add_one_item(rel_item, self.get_field_value("sermo", item, "incexpl"), resizable)

                    # S: Manuscript
                    add_one_item(rel_item, self.get_field_value("sermo", item, "manuscript"), False)

                    # S: Locus
                    add_one_item(rel_item, item.locus, False)

                    # Actions that can be performed on this item
                    add_one_item(rel_item, self.get_actions())

                    # Add this line to the list
                    rel_list.append(dict(id=item.id, cols=rel_item))
            
                sermons['rel_list'] = rel_list
                sermons['columns'] = [
                    '{}<span title="Default order">Order<span>{}'.format(sort_start_int, sort_end),
                    '{}<span title="Attributed author">Author</span>{}'.format(sort_start, sort_end), 
                    '{}<span title="Gryson or Clavis code">Signature</span>{}'.format(sort_start, sort_end), 
                    '{}<span title="Incipit and explicit">inc...expl</span>{}'.format(sort_start, sort_end), 
                    '{}<span title="Manuscript shelfmark">Manuscript</span>{}'.format(sort_start, sort_end), 
                    '{}<span title="Location within the manuscript">Locus</span>{}'.format(sort_start_int, sort_end),
                    ''
                    ]
                related_objects.append(sermons)

            elif instance.type == "gold":
                # Get all sermons that are part of this PD
                goldsermons = dict(title="Gold sermons within this dataset", prefix="gold")   # prefix="sermo") 
                if resizable: goldsermons['gridclass'] = "resizable dragdrop"
                goldsermons['savebuttons'] = True
                goldsermons['saveasbutton'] = True

                qs_sermo = instance.gold_col.all().order_by(
                        'order', 'gold__author__name', 'gold__siglist', 'gold__equal__code', 'gold__srchincipit', 'gold__srchexplicit')
                check_order(qs_sermo)

                # Walk these collection sermons
                for idx, obj in enumerate(qs_sermo):
                    rel_item = []
                    item = obj.gold

                    # Leave if this is too much
                    if idx > self.max_items:
                        break

                    # G: Order in Gold
                    #add_one_item(rel_item, index, False, align="right")
                    #index += 1
                    add_one_item(rel_item, obj.order, False, align="right", draggable=True)

                    # G: Author
                    add_one_item(rel_item, self.get_field_value("gold", item, "author"), False,main=True)

                    # G: Signature
                    add_one_item(rel_item, self.get_field_value("gold", item, "signature"), False)

                    # G: Passim code
                    add_one_item(rel_item, self.get_field_value("gold", item, "code"), False)

                    # G: Inc/Expl
                    add_one_item(rel_item, self.get_field_value("gold", item, "incexpl"), resizable)

                    # G: Editions
                    add_one_item(rel_item, self.get_field_value("gold", item, "edition"), False)

                    # Actions that can be performed on this item
                    add_one_item(rel_item, self.get_actions())

                    # Add this line to the list
                    rel_list.append(dict(id=item.id, cols=rel_item))
            
                goldsermons['rel_list'] = rel_list
                goldsermons['columns'] = [
                    '{}<span title="Default order">Order<span>{}'.format(sort_start_int, sort_end),
                    '{}<span title="Associated author">Author</span>{}'.format(sort_start, sort_end), 
                    '{}<span title="Gryson or Clavis code">Signature</span>{}'.format(sort_start, sort_end), 
                    '{}<span title="PASSIM code">Passim</span>{}'.format(sort_start, sort_end), 
                    '{}<span title="Incipit and explicit">inc...expl</span>{}'.format(sort_start, sort_end), 
                    '{}<span title="Editions where this Gold Sermon is described">Editions</span>{}'.format(sort_start, sort_end), 
                    ''
                    ]
                related_objects.append(goldsermons)

            elif instance.type == "super":
                # Get all sermons that are part of this PD
                supers = dict(title="Authority files within this dataset", prefix="super")   #prefix="sermo")
                if resizable: supers['gridclass'] = "resizable dragdrop"
                supers['savebuttons'] = True
                supers['saveasbutton'] = True

                qs_sermo = instance.super_col.all().order_by(
                        'order', 'super__author__name', 'super__firstsig', 'super__srchincipit', 'super__srchexplicit')
                check_order(qs_sermo)

                # Walk these collection sermons
                for idx, obj in enumerate(qs_sermo):
                    rel_item = []
                    item = obj.super

                    # Leave if this is too much
                    if idx > self.max_items:
                        break

                    # SSG: Order in Manuscript
                    #add_one_item(rel_item, index, False, align="right")
                    #index += 1
                    add_one_item(rel_item, obj.order, False, align="right", draggable=True)

                    # SSG: Author
                    add_one_item(rel_item, self.get_field_value("super", item, "author"), False, main=True)

                    # SSG: Passim code
                    add_one_item(rel_item, self.get_field_value("super", item, "code"), False)

                    # SSG: Gryson/Clavis = signature
                    add_one_item(rel_item, self.get_field_value("super", item, "sig"), False)

                    # SSG: Inc/Expl
                    add_one_item(rel_item, self.get_field_value("super", item, "incexpl"), resizable)

                    # SSG: Size (number of SG in equality set)
                    add_one_item(rel_item, self.get_field_value("super", item, "size"), False)

                    # Actions that can be performed on this item
                    add_one_item(rel_item, self.get_actions())

                    # Add this line to the list
                    rel_list.append(dict(id=item.id, cols=rel_item))
            
                supers['rel_list'] = rel_list
                supers['columns'] = [
                    '{}<span title="Default order">Order<span>{}'.format(sort_start_int, sort_end),
                    '{}<span title="Author">Author</span>{}'.format(sort_start, sort_end), 
                    '{}<span title="PASSIM code">Passim</span>{}'.format(sort_start, sort_end), 
                    '{}<span title="Gryson or Clavis codes of sermons gold in this set">Gryson/Clavis</span>{}'.format(sort_start, sort_end), 
                    '{}<span title="Incipit and explicit">inc...expl</span>{}'.format(sort_start, sort_end), 
                    '{}<span title="Number of Sermons Gold part of this set">Size</span>{}'.format(sort_start_int, sort_end), 
                    ''
                    ]
                related_objects.append(supers)

                context['histogram_data'] = self.get_histogram_data(instance, 
                                                                    instance.collections_super.all(), 
                                                                    'collist_{}'.format(self.prefix), 
                                                                    'd3')

            context['related_objects'] = related_objects
        except:
            msg = oErr.get_error_message()
            oErr.DoError("CollPrivDetails/add_to_context")

        # REturn the total context
        return context

    def get_actions(self):
        html = []
        buttons = ['remove']    # This contains all the button names that need to be added

        # Start the whole spane
        html.append("<div class='blinded'>")
        
        # Add components
        if 'up' in buttons: 
            html.append("<a class='related-up' ><span class='glyphicon glyphicon-arrow-up'></span></a>")
        if 'down' in buttons: 
            html.append("<a class='related-down'><span class='glyphicon glyphicon-arrow-down'></span></a>")
        if 'remove' in buttons: 
            html.append("<a class='related-remove'><span class='glyphicon glyphicon-remove'></span></a>")

        # Finish up the span
        html.append("&nbsp;</span>")

        # COmbine the list into a string
        sHtml = "\n".join(html)
        # Return out HTML string
        return sHtml

    def get_field_value(self, type, instance, custom):
        sBack = ""
        if type == "manu":
            if custom == "shelfmark":
                sBack = "{}, {}, <span class='signature'>{}</span>".format(instance.get_full_name())
            elif custom == "name":
                sBack = instance.name
            elif custom == "origprov":
                sBack = "{} ({})".format(instance.get_origin(), instance.get_provenance_markdown())
            elif custom == "daterange":
                sBack = "{}-{}".format(instance.yearstart, instance.yearfinish)
            elif custom == "sermons":
                sBack = instance.get_sermon_count()
        elif type == "sermo":
            sBack, sTitle = SermonListView.get_field_value(None, instance, custom)
        elif type == "gold":
            sBack, sTitle = SermonGoldListView.get_field_value(None, instance, custom)
        elif type == "super":
            sBack, sTitle = EqualGoldListView.get_field_value(None, instance, custom)
        return sBack


class CollPublDetails(CollPrivDetails):
    """Like CollPrivDetails, but then with public"""

    prefix = "publ"
    basic_name = "collpubl"
    title = "Public Dataset"

    def custom_init(self, instance):
        if instance != None:
            # Check if someone acts as if this is a public dataset, while it is not
            if instance.settype == "pd":
                # Determine what kind of dataset/collection this is
                if instance.owner == Profile.get_user_profile(self.request.user.username):
                    # It is a private dataset after all!
                    self.redirectpage = reverse("collpriv_details", kwargs={'pk': instance.id})
            elif instance.settype == "hc":
                # This is a historical collection
                self.redirectpage = reverse("collhist_details", kwargs={'pk': instance.id})
            if instance.type == "super":
                self.custombuttons = [{"name": "scount_histogram", "title": "Sermon Histogram", 
                      "icon": "th-list", "template_name": "seeker/scount_histogram.html" }]
            # Check for hlist saving
            self.check_hlist(instance)
        return None


class CollHistDetails(CollHistEdit):
    """Like CollHistEdit, but then with html"""
    rtype = "html"
    custombuttons = [{"name": "scount_histogram", "title": "Sermon Histogram", 
                      "icon": "th-list", "template_name": "seeker/scount_histogram.html" }]

    def custom_init(self, instance):
        # First do the original custom init
        response = super(CollHistDetails, self).custom_init(instance)

        if not instance is None:
            # Now continue
            if instance.settype != "hc":
                # Someone does as if this is a historical collection...
                # Determine what kind of dataset/collection this is
                if instance.owner == Profile.get_user_profile(self.request.user.username):
                    # Private dataset
                    self.redirectpage = reverse("collpriv_details", kwargs={'pk': instance.id})
                else:
                    # Public dataset
                    self.redirectpage = reverse("collpubl_details", kwargs={'pk': instance.id})

            # Check for hlist saving
            self.check_hlist(instance)
        return None

    def add_to_context(self, context, instance):
        """Add to the existing context"""

        # Start by executing the standard handling
        super(CollHistDetails, self).add_to_context(context, instance)

        def add_one_item(rel_item, value, resizable=False, title=None, align=None, link=None, main=None, draggable=None):
            oAdd = dict(value=value)
            if resizable: oAdd['initial'] = 'small'
            if title != None: oAdd['title'] = title
            if align != None: oAdd['align'] = align
            if link != None: oAdd['link'] = link
            if main != None: oAdd['main'] = main
            if draggable != None: oAdd['draggable'] = draggable
            rel_item.append(oAdd)
            return True

        def check_order(qs):
            with transaction.atomic():
                for idx, obj in enumerate(qs):
                    if obj.order < 0:
                        obj.order = idx + 1
                        obj.save()

        oErr = ErrHandle()
        try:

            context['sections'] = []

            username = self.request.user.username
            team_group = app_editor

            # Authorization: only app-editors may edit!
            bMayEdit = user_is_ingroup(self.request, team_group)
            
            # Lists of related objects and other initializations
            related_objects = []
            resizable = True
            lstQ = []
            rel_list =[]
            index = 1
            sort_start = ""
            sort_start_int = ""
            sort_end = ""
            show_codico = True  # See issue #363
            show_manu = False   # See issue #363

            if bMayEdit:
                sort_start = '<span class="sortable"><span class="fa fa-sort sortshow"></span>&nbsp;'
                sort_start_int = '<span class="sortable integer"><span class="fa fa-sort sortshow"></span>&nbsp;'
                sort_end = '</span>'

            # In all cases: Get all the SSGs that are part of this historical collection:
            qs_ssg = instance.collections_super.all().values("id")

            # Check what kind of comparison we need to make
            if self.manu == None and self.codico == None:
                # This is the plain historical collection details view

                # Get all SSGs that are part of this PD
                supers = dict(title="Authority files within this historical collection", prefix="super")   #prefix="sermo")
                if resizable: supers['gridclass'] = "resizable dragdrop"
                supers['savebuttons'] = True
                supers['saveasbutton'] = True
                supers['classes'] = 'collapse'

                qs_sermo = instance.super_col.all().order_by(
                        'order', 'super__author__name', 'super__firstsig', 'super__srchincipit', 'super__srchexplicit')
                check_order(qs_sermo)

                # Walk these collection sermons
                for idx, obj in enumerate(qs_sermo):
                    rel_item = []
                    item = obj.super

                    # Leave if this is too much
                    if idx > self.max_items:
                        break

                    # SSG: Order in Manuscript
                    #add_one_item(rel_item, index, False, align="right")
                    #index += 1
                    add_one_item(rel_item, obj.order, False, align="right", draggable=True)

                    # SSG: Author
                    add_one_item(rel_item, self.get_field_value("super", item, "author"), resizable)

                    # SSG: Passim code
                    add_one_item(rel_item, self.get_field_value("super", item, "code"), False)

                    # SSG: Gryson/Clavis = signature
                    add_one_item(rel_item, self.get_field_value("super", item, "sig"), False)

                    # SSG: Inc/Expl
                    add_one_item(rel_item, self.get_field_value("super", item, "incexpl"), False, main=True)

                    # SSG: Size (number of SG in equality set)
                    add_one_item(rel_item, self.get_field_value("super", item, "size"), False)

                    # Actions that can be performed on this item
                    if bMayEdit:
                        add_one_item(rel_item, self.get_actions())

                    # Add this line to the list
                    rel_list.append(dict(id=item.id, cols=rel_item))
            
                supers['rel_list'] = rel_list
                supers['columns'] = [
                    '{}<span title="Default order">Order<span>{}'.format(sort_start_int, sort_end),
                    '{}<span title="Author">Author</span>{}'.format(sort_start, sort_end), 
                    '{}<span title="PASSIM code">Passim</span>{}'.format(sort_start, sort_end), 
                    '{}<span title="Gryson or Clavis codes of sermons gold in this set">Gryson/Clavis</span>{}'.format(sort_start, sort_end), 
                    '{}<span title="Incipit and explicit">inc...expl</span>{}'.format(sort_start, sort_end), 
                    '{}<span title="Number of Sermons Gold part of this set">Size</span>{}'.format(sort_start_int, sort_end), 
                    ''
                    ]
                related_objects.append(supers)

                if show_codico:
                    # NEW (see issue #363): List of codicological units contained in this collection
                    sTitle = "Codicological units with sermons connected to this collection"
                    codicos = dict(title=sTitle, prefix="codi")
                    if resizable:
                        codicos['gridclass'] = "resizable"
                    codicos['classes'] = 'collapse in'

                    # Get the codicos linked to these SSGs (in this historical collection)
                    lstQ = []
                    lstQ.append(Q(codicoitems__itemsermons__equalgolds__collections=instance))
                    lstQ.append(Q(manuscript__mtype="man"))
                    qs_codi = Codico.objects.filter(*lstQ).order_by(
                        'id').distinct().order_by('manuscript__lcity__name', 'manuscript__library__name', 'manuscript__idno')

                    rel_list =[]
                    for idx, item in enumerate(qs_codi):
                        rel_item = []

                        # Leave if this is too much
                        if idx > self.max_items:
                            break

                        # Shelfmark = IDNO
                        add_one_item(rel_item,  self.get_field_value("codicos", item, "manuscript"), False, title=item.manuscript.idno, main=True, 
                                     link=reverse('manuscript_details', kwargs={'pk': item.manuscript.id}))

                        # Origin
                        add_one_item(rel_item, self.get_field_value("codicos", item, "origprov"), resizable, 
                                     title="Origin (if known), followed by provenances (between brackets)")

                        # date range
                        add_one_item(rel_item, self.get_field_value("codicos", item, "daterange"), resizable, align="right")

                        # Number of sermons in this codicological unit
                        add_one_item(rel_item, self.get_field_value("codicos", item, "sermons"), resizable, align="right")

                        # Linked SSG(s)
                        ssg_info = item.get_ssg_count(compare_link=True, collection=instance)
                        add_one_item(rel_item, ssg_info, resizable, align="right")

                        # Add this Manu line to the list
                        rel_list.append(dict(id=item.id, cols=rel_item))

                    codicos['rel_list'] = rel_list

                    codicos['columns'] = [
                        'Manuscript', 
                        '<span title="Origin/Provenance">or./prov.</span>', 
                        '<span title="Date range">date</span>', 
                        '<span title="Sermons in this codicological unit">sermons</span>',
                        '<span title="Authority file links">ssgs.</span>', 
                        ]
                    related_objects.append(codicos)

                if show_manu:
                    # OLD (see issue #363): List of Manuscripts contained in this collection
                    sTitle = "Manuscripts with sermons connected to this collection"
                    manuscripts = dict(title=sTitle, prefix="manu")
                    if resizable:
                        manuscripts['gridclass'] = "resizable"
                    manuscripts['classes'] = 'collapse in'

                    # Get the manuscripts linked to these SSGs (in this historical collection)
                    lstQ = []
                    lstQ.append(Q(manuitems__itemsermons__equalgolds__collections=instance))
                    lstQ.append(Q(mtype="man"))
                    qs_manu = Manuscript.objects.filter(*lstQ).order_by(
                        'id').distinct().order_by('lcity__name', 'library__name', 'idno')

                    rel_list =[]
                    for idx, item in enumerate(qs_manu):
                        rel_item = []

                        # Leave if this is too much
                        if idx > self.max_items:
                            break

                        # Get the codico's for this manuscript
                        codico_lst = item.manuscriptcodicounits.all().order_by('order')

                        # Shelfmark = IDNO
                        add_one_item(rel_item,  self.get_field_value("manu", item, "shelfmark"), False, title=item.idno, main=True, 
                                     link=reverse('manuscript_details', kwargs={'pk': item.id}))

                        # Origin
                        add_one_item(rel_item, self.get_field_value("manucodicos", codico_lst, "origprov"), resizable, 
                                     title="Origin (if known), followed by provenances (between brackets)")

                        # date range
                        add_one_item(rel_item, self.get_field_value("manucodicos", codico_lst, "daterange"), resizable, align="right")

                        # Number of sermons in this manuscript
                        add_one_item(rel_item, self.get_field_value("manu", item, "sermons"), resizable, align="right")

                        # Linked SSG(s)
                        ssg_info = item.get_ssg_count(compare_link=True, collection=instance)
                        add_one_item(rel_item, ssg_info, resizable, align="right")

                        # Add this Manu line to the list
                        rel_list.append(dict(id=item.id, cols=rel_item))

                    manuscripts['rel_list'] = rel_list

                    manuscripts['columns'] = [
                        'Shelfmark', 
                        '<span title="Origin/Provenance">or./prov.</span>', 
                        '<span title="Date range">date</span>', 
                        '<span title="Sermons in this manuscript">sermons</span>',
                        '<span title="Authority file links">ssgs.</span>', 
                        ]
                    related_objects.append(manuscripts)
            elif self.manu != None:
                # This is a comparison between the SSGs in the historical collection and the sermons in the manuscript
                # (1) Start making a comparison table
                title = "Comparison with manuscript [{}]".format(self.manu.get_full_name())
                sermons = dict(title=title, prefix="sermo", gridclass="resizable")
                # (2) Get a list of sermons
                qs_s = SermonDescr.objects.filter(msitem__manu=self.manu).order_by('msitem__order')

                # Build the related list
                rel_list =[]
                equal_list = []
                index = 1
                for item in qs_s:
                    rel_item = []
                    # Determine the matching SSG from the Historical Collection
                    equal = EqualGold.objects.filter(sermondescr_super__super__in=qs_ssg, sermondescr_super__sermon__id=item.id).first()
                    # List of SSGs that have been dealt with already
                    if equal != None: equal_list.append(equal.id)

                    # S: Order in Manuscript
                    rel_item.append({'value': index, 'initial': 'small'})
                    index += 1

                    # S: Locus
                    rel_item.append({'value': item.get_locus()})

                    # S: TItle
                    rel_item.append({'value': item.title, 'initial': 'small'})

                    # SSG: passim code
                    if equal:
                        rel_item.append({'value': equal.get_passimcode_markdown(), 'initial': 'small'})
                    else:
                        rel_item.append({'value': "(none)", 'initial': 'small'})

                    ratio = 0.0
                    if equal:
                        # S: incipit + explicit compared
                        s_equal, ratio_equal = equal.get_incexp_match()
                        comparison, ratio = item.get_incexp_match(s_equal)
                        rel_item.append({'value': comparison, 'initial': 'small'})

                        # SSG: incipit + explicit compared
                        s_sermon, ratio_sermon = item.get_incexp_match()
                        comparison, ratio2 = equal.get_incexp_match(s_sermon)
                        rel_item.append({'value': comparison, 'initial': 'small'})
                    else:
                        # S: incipit + explicit compared
                        s_sermon, ratio_sermon = item.get_incexp_match()
                        rel_item.append({'value': s_sermon, 'initial': 'small'})

                        # SSG: incipit + explicit compared
                        rel_item.append({'value': "", 'initial': 'small'})

                    # Ratio of equalness
                    rel_item.append({'value': "{:.1%}".format(ratio), 'initial': 'small'})

                    rel_list.append(dict(id=item.id, cols=rel_item))

                # Check if there are any SSGs in the collection that have not been dealt with yet
                qs_ssg = instance.collections_super.exclude(id__in=equal_list)
                for idx, item in enumerate(qs_ssg):
                    rel_item = []
                    equal = item

                    # Leave if this is too much
                    if idx > self.max_items:
                        break

                    # S: Order in Manuscript
                    rel_item.append({'value': "-", 'initial': 'small'})

                    # S: Locus
                    rel_item.append({'value': "-"})

                    # S: TItle
                    rel_item.append({'value': "-", 'initial': 'small'})

                    # SSG: passim code
                    rel_item.append({'value': equal.get_passimcode_markdown(), 'initial': 'small'})

                    # S: incipit + explicit compared
                    ratio = 0.0
                    rel_item.append({'value': "", 'initial': 'small'})

                    # SSG: incipit + explicit compared
                    s_equal, ratio_equal = equal.get_incexp_match()
                    rel_item.append({'value': s_equal, 'initial': 'small'})

                    # Ratio of equalness
                    rel_item.append({'value': ratio, 'initial': 'small'})

                    rel_list.append(dict(id=item.id, cols=rel_item))


                # Add the related list
                sermons['rel_list'] = rel_list

                # Set the columns
                sermons['columns'] = ['Order', 'Locus', 'Title', 
                                      '<span title="Authority file">ssg</span>',
                                      '<span title="Incipit + explicit of sermon manifestation">inc/exp. s</span>', 
                                      '<span title="Incipit + explicit of Authority file">inc/exp. ssg</span>',
                                      '<span title="Comparison ratio between inc/exp of S and SSG">ratio</span>']
                # Add to related objects
                related_objects.append(sermons)

            elif self.codico != None:
                # This is a comparison between the SSGs in the historical collection and the sermons in a codicological  unit
                # (1) Start making a comparison table
                title = "Comparison with codicological unit [{}]".format(self.codico.get_full_name())
                sermons = dict(title=title, prefix="sermo", gridclass="resizable")
                # (2) Get a list of sermons
                qs_s = SermonDescr.objects.filter(msitem__codico=self.codico).order_by('msitem__order')

                # Build the related list
                rel_list =[]
                equal_list = []
                index = 1
                for idx, item in enumerate(qs_s):
                    rel_item = []

                    # Leave if this is too much
                    if idx > self.max_items:
                        break

                    # Determine the matching SSG from the Historical Collection
                    equal = EqualGold.objects.filter(sermondescr_super__super__in=qs_ssg, sermondescr_super__sermon__id=item.id).first()
                    # List of SSGs that have been dealt with already
                    if equal != None: equal_list.append(equal.id)

                    # S: Order in Manuscript
                    rel_item.append({'value': index, 'initial': 'small'})
                    index += 1

                    # S: Locus
                    rel_item.append({'value': item.get_locus()})

                    # S: TItle
                    rel_item.append({'value': item.title, 'initial': 'small'})

                    # SSG: passim code
                    if equal:
                        rel_item.append({'value': equal.get_passimcode_markdown(), 'initial': 'small'})
                    else:
                        rel_item.append({'value': "(none)", 'initial': 'small'})

                    ratio = 0.0
                    if equal:
                        # S: incipit + explicit compared
                        s_equal, ratio_equal = equal.get_incexp_match()
                        comparison, ratio = item.get_incexp_match(s_equal)
                        rel_item.append({'value': comparison, 'initial': 'small'})

                        # SSG: incipit + explicit compared
                        s_sermon, ratio_sermon = item.get_incexp_match()
                        comparison, ratio2 = equal.get_incexp_match(s_sermon)
                        rel_item.append({'value': comparison, 'initial': 'small'})
                    else:
                        # S: incipit + explicit compared
                        s_sermon, ratio_sermon = item.get_incexp_match()
                        rel_item.append({'value': s_sermon, 'initial': 'small'})

                        # SSG: incipit + explicit compared
                        rel_item.append({'value': "", 'initial': 'small'})

                    # Ratio of equalness
                    rel_item.append({'value': "{:.1%}".format(ratio), 'initial': 'small'})

                    rel_list.append(dict(id=item.id, cols=rel_item))

                # Check if there are any SSGs in the collection that have not been dealt with yet
                qs_ssg = instance.collections_super.exclude(id__in=equal_list)
                for idx, item in enumerate(qs_ssg):
                    rel_item = []
                    equal = item

                    # Leave if this is too much
                    if idx > self.max_items:
                        break

                    # S: Order in Manuscript
                    rel_item.append({'value': "-", 'initial': 'small'})

                    # S: Locus
                    rel_item.append({'value': "-"})

                    # S: TItle
                    rel_item.append({'value': "-", 'initial': 'small'})

                    # SSG: passim code
                    rel_item.append({'value': equal.get_passimcode_markdown(), 'initial': 'small'})

                    # S: incipit + explicit compared
                    ratio = 0.0
                    rel_item.append({'value': "", 'initial': 'small'})

                    # SSG: incipit + explicit compared
                    s_equal, ratio_equal = equal.get_incexp_match()
                    rel_item.append({'value': s_equal, 'initial': 'small'})

                    # Ratio of equalness
                    rel_item.append({'value': ratio, 'initial': 'small'})

                    rel_list.append(dict(id=item.id, cols=rel_item))


                # Add the related list
                sermons['rel_list'] = rel_list

                # Set the columns
                sermons['columns'] = ['Order', 'Locus', 'Title', 
                                      '<span title="Authority file">ssg</span>',
                                      '<span title="Incipit + explicit of sermon manifestation">inc/exp. s</span>', 
                                      '<span title="Incipit + explicit of Authority file">inc/exp. ssg</span>',
                                      '<span title="Comparison ratio between inc/exp of S and Authority fileSSG">ratio</span>']
                # Add to related objects
                related_objects.append(sermons)

            context['related_objects'] = related_objects
            context['histogram_data'] = self.get_histogram_data(instance, instance.collections_super.all(), 'collist_hist', 'd3')
        except:
            msg = oErr.get_error_message()
            oErr.DoError("CollHistDetails/add_to_context")

        # Return the context we have made
        return context

    def get_actions(self):
        html = []
        buttons = ['remove']    # This contains all the button names that need to be added

        # Start the whole spane
        html.append("<div class='blinded'>")
        
        # Add components
        if 'up' in buttons: 
            html.append("<a class='related-up' ><span class='glyphicon glyphicon-arrow-up'></span></a>")
        if 'down' in buttons: 
            html.append("<a class='related-down'><span class='glyphicon glyphicon-arrow-down'></span></a>")
        if 'remove' in buttons: 
            html.append("<a class='related-remove'><span class='glyphicon glyphicon-remove'></span></a>")

        # Finish up the span
        html.append("&nbsp;</span>")

        # COmbine the list into a string
        sHtml = "\n".join(html)
        # Return out HTML string
        return sHtml

    def get_field_value(self, type, instance, custom):
        sBack = ""
        if type == "manu":
            if custom == "shelfmark":
                sBack = "{}, {}, <span class='signature'>{}</span>".format(instance.get_full_name())
            elif custom == "name":
                sBack = instance.name
            elif custom == "sermons":
                sBack = instance.get_sermon_count()
        elif type == "manucodicos":
            lCombi = []
            if custom == "origprov":
                for obj in instance:
                    lCombi.append( "origin: {} (provenance[s]: {})".format(obj.get_origin(), obj.get_provenance_markdown(table=False)))
                sBack = "; ".join(lCombi)
            elif custom == "daterange":
                for obj in instance:
                    lCombi.append( "{}-{}".format(obj.yearstart, obj.yearfinish))
                sBack = "; ".join(lCombi)
        elif type == "codicos":
            lCombi = []
            if custom == "origprov":
                sBack = "origin: {} (provenance[s]: {})".format(instance.get_origin(), instance.get_provenance_markdown(table=False))
            elif custom == "daterange":
                sBack = "{}-{}".format(instance.yearstart, instance.yearfinish)
            elif custom == "manuscript":
                sBack = "<span class='signature'>{}</span>".format(instance.manuscript.get_full_name())
            elif custom == "sermons":
                sBack = instance.get_sermon_count()

        elif type == "super":
            sBack, sTitle = EqualGoldListView.get_field_value(None, instance, custom)
        return sBack


class CollHistCompare(CollHistDetails):
    """Compare the SSGs in a historical collection with the sermons in a manuscript"""

    def custom_init(self, instance):
        # FIrst perform the standard initialization
        response = super(CollHistCompare, self).custom_init(instance)

        # Make sure to get the Manuscript for comparison
        if user_is_authenticated(self.request) and user_is_ingroup(self.request, app_editor):
            manu_id = self.qd.get('manu')
            codico_id = self.qd.get('codico')
            if manu_id != None:
                manu = Manuscript.objects.filter(id=manu_id).first()
                if manu != None:
                    # We have the manuscript: the comparison can continue
                    self.manu = manu
                    self.codico = None
            elif codico_id != None:
                codico = Codico.objects.filter(id=codico_id).first()
                if codico != None:
                    # We have the codico: the comparison can continue
                    self.codico = codico
                    self.manu = None
        return None


class CollHistElevate(CollHistDetails):
    """ELevate this dataset to be a historical collection"""

    def custom_init(self, instance):
        if user_is_authenticated(self.request):
            # Double check if I have the right to do this...
            if user_is_ingroup(self.request, app_editor):
                # Change the settype to hc
                instance.settype = "hc"
                instance.save()
                self.redirectpage = reverse("collhist_details", kwargs={'pk': instance.id})
            elif instance.settype == "pd":
                # Determine what kind of dataset/collection this is
                if instance.owner == Profile.get_user_profile(self.request.user.username):
                    # Private dataset
                    self.redirectpage = reverse("collpriv_details", kwargs={'pk': instance.id})
                else:
                    # Public dataset
                    self.redirectpage = reverse("collpubl_details", kwargs={'pk': instance.id})
        else:
            self.redirectpage = reverse("home")
        return None


class CollHistApply(CollHistDetails):
    """Apply the historical collection to create a manuscript with sermons from the SSGs"""

    apply_type = ""

    def custom_init(self, instance):
        # Create a new manuscript that is based on this historical collection
        item_new = instance.get_hctemplate_copy(self.request.user.username, self.apply_type)
        if item_new == None:
            # THis wasn't successful: redirect to the details view
            self.redirectpage = reverse("collhist_details", kwargs={'pk': instance.id})
        elif self.apply_type == "tem":
            # A template has been created
            self.redirectpage = reverse("template_details", kwargs={'pk': item_new.id})
        else:
            # Manuscript created: re-direct to this manuscript
            self.redirectpage = reverse("manuscript_details", kwargs={'pk': item_new.id})
        return None


class CollHistManu(CollHistApply):
    """Apply the historical collection to create a manuscript with sermons from the SSGs"""

    apply_type = "man"


class CollHistTemp(CollHistApply):
    """Apply the historical collection to create a manuscript with sermons from the SSGs"""

    apply_type = "tem"


class CollManuDetails(CollManuEdit):
    """Like CollManuEdit, but then with html"""
    rtype = "html"


class CollSermoDetails(CollSermoEdit):
    """Like CollSermoEdit, but then with html"""
    rtype = "html"


class CollGoldDetails(CollGoldEdit):
    """Like CollGoldEdit, but then with html"""
    rtype = "html"


class CollSuperDetails(CollSuperEdit):
    """Like CollSuperEdit, but then with html"""
    rtype = "html"


class CollectionListView(BasicList):
    """Search and list collections"""

    model = Collection
    listform = CollectionForm
    prefix = "any"
    paginate_by = 20
    bUseFilter = True
    has_select2 = True
    basic_name_prefix = "coll"
    settype = "pd"              # Personal Dataset versus Historical Collection
    use_team_group = True
    plural_name = ""
    order_cols = ['scope', 'name', 'created', 'owner__user__username', '']
    order_default = order_cols
    order_heads = [{'name': 'Scope',        'order': 'o=1', 'type': 'str', 'custom': 'scope'},
                   {'name': 'Collection',   'order': 'o=2', 'type': 'str', 'field': 'name', 'linkdetails': True, 'main': True},
                   {'name': 'Created',      'order': 'o=3', 'type': 'str', 'custom': 'created'},
                   {'name': 'Owner',        'order': 'o=4', 'type': 'str', 'custom': 'owner'},
                   {'name': 'Frequency',    'order': '',    'type': 'str', 'custom': 'links'}
                   ]
    filters = [ {"name": "Collection", "id": "filter_collection", "enabled": False},
                {"name": "Owner",      "id": "filter_owner",      "enabled": False}]
    searches = [
        {'section': '', 'filterlist': [
            {'filter': 'owner',     'fkfield': 'owner',  'keyS': 'owner', 'keyFk': 'id', 'keyList': 'ownlist', 'infield': 'id' },
            {'filter': 'collection','dbfield': 'name',   'keyS': 'collection_ta', 'keyList': 'collist', 'infield': 'name'}]},
        {'section': 'other', 'filterlist': [
            {'filter': 'coltype',   'dbfield': 'type',   'keyS': 'type',  'keyList': 'typelist' },
            {'filter': 'settype',   'dbfield': 'settype','keyS': 'settype'},
            ]} # {'filter': 'scope',     'dbfield': 'scope',  'keyS': 'scope'} eruit
        ]

    selectbuttons = [
        {'title': 'Add to saved items', 'mode': 'add_saveitem', 'button': 'jumbo-1', 'glyphicon': 'glyphicon-star-empty'},
        {'title': 'Add to DCT',         'mode': 'show_dct',     'button': 'jumbo-1', 'glyphicon': 'glyphicon-wrench'},
        ]

    def initializations(self):
        if self.prefix == "sermo":
            self.plural_name = "Sermon collections"
            self.sg_name = "Sermon collection"
            self.searches[0]['filterlist'][1]['keyList'] = "collist_s"
        elif self.prefix == "manu":
            self.plural_name = "Manuscript Collections"
            self.sg_name = "Manuscript collection"
            self.searches[0]['filterlist'][1]['keyList'] = "collist_m"
        elif self.prefix == "gold":
            self.plural_name = "Gold sermons Collections"
            self.sg_name = "Sermon gold collection"
            self.searches[0]['filterlist'][1]['keyList'] = "collist_sg"
        elif self.prefix == "super":
            self.plural_name = "Authority file Collections"
            self.sg_name = "Authority file collection"        
            self.searches[0]['filterlist'][1]['keyList'] = "collist_ssg"
        elif self.prefix == "any":
            self.new_button = False
            self.plural_name = "All types Collections"
            self.sg_name = "Collection"  
            self.order_cols = ['typename__full', 'scope', 'name', 'created', 'owner__user__username', '']
            self.order_default = self.order_cols
            self.order_heads  = [
                {'name': 'Type',        'order': 'o=1', 'type': 'str', 'custom': 'type'},
                {'name': 'Scope',       'order': 'o=2', 'type': 'str', 'custom': 'scope'},
                {'name': 'Dataset',     'order': 'o=3', 'type': 'str', 'field': 'name', 'linkdetails': True, 'main': True},
                {'name': 'Created',     'order': 'o=4', 'type': 'str', 'custom': 'created'},
                {'name': 'Owner',       'order': 'o=5', 'type': 'str', 'custom': 'owner'},
                {'name': 'Frequency',   'order': '',    'type': 'str', 'custom': 'links'}
            ]  
        elif self.prefix == "priv":
            self.new_button = False
            self.titlesg = "Personal Dataset"
            self.plural_name = "Datasets"
            self.sg_name = "Dataset"  
            self.order_cols = ['typename__full', 'name', 'scope', 'owner__user__username', '', 'created', '']
            self.order_default = self.order_cols
            self.order_heads  = [
                {'name': 'Type',        'order': 'o=1', 'type': 'str', 'custom': 'type'},
                {'name': 'Dataset',     'order': 'o=2', 'type': 'str', 'field': 'name', 'linkdetails': True, 'main': True},
                {'name': 'Scope',       'order': 'o=3', 'type': 'str', 'custom': 'scope'},
                {'name': 'Owner',       'order': 'o=4', 'type': 'str', 'custom': 'owner'},
                {'name': '',            'order': '',    'type': 'str', 'custom': 'saved',   'align': 'right'},
                {'name': 'Created',     'order': 'o=6', 'type': 'str', 'custom': 'created'},
                {'name': 'Frequency',   'order': '',    'type': 'str', 'custom': 'links'}
            ]  
            self.filters = [ {"name": "My dataset", "id": "filter_collection", "enabled": False},
                             {"name": "Type", "id": "filter_coltype", "enabled": False},
                             {"name": "Scope", "id": "filter_colscope", "enabled": False}]
            self.searches = [
                {'section': '', 'filterlist': [
                    {'filter': 'collection','dbfield': 'name',   'keyS': 'collection_ta', 'keyList': 'collist', 'infield': 'name'},
                    {'filter': 'colscope',  'dbfield': 'scope',  'keyS': 'colscope', 'keyType': 'fieldchoice', 'infield': 'abbr'}
                    ]},
                {'section': 'other', 'filterlist': [
                    {'filter': 'owner',     'fkfield': 'owner',  'keyS': 'owner', 'keyFk': 'id', 'keyList': 'ownlist', 'infield': 'id' },
                    {'filter': 'coltype',   'dbfield': 'type',   'keyS': 'type'},
                    {'filter': 'settype',   'dbfield': 'settype','keyS': 'settype'},
                    # {'filter': 'scope',  'dbfield': 'scope',  'keyS': 'scope'}
                    {'filter': 'defscope',  'dbfield': '$dummy', 'keyS': 'defscope'}
                    ]}
                ]
        elif self.prefix == "publ":
            self.new_button = False
            self.plural_name = "Public Datasets"
            self.sg_name = "Public Dataset"  
            self.order_cols = ['typename__full', 'name', 'created',  'owner__user__username', '', '']
            self.order_default = self.order_cols
            self.order_heads  = [
                {'name': 'Type',        'order': 'o=1', 'type': 'str', 'custom': 'type'},
                {'name': 'Dataset',     'order': 'o=2', 'type': 'str', 'field': 'name', 'linkdetails': True, 'main': True},
                {'name': 'Created',     'order': 'o=3', 'type': 'str', 'custom': 'created'},
                {'name': 'Owner',       'order': 'o=4', 'type': 'str', 'custom': 'owner'},
                {'name': '',            'order': '',    'type': 'str', 'custom': 'saved',   'align': 'right'},
                {'name': 'Frequency',   'order': '',    'type': 'str', 'custom': 'links'}
            ]  
            self.filters = [ {"name": "Public dataset", "id": "filter_collection", "enabled": False}]
            self.searches = [
                {'section': '', 'filterlist': [
                    {'filter': 'collection','dbfield': 'name',   'keyS': 'collection_ta', 'keyList': 'collist', 'infield': 'name'},
                    {'filter': 'owner',     'fkfield': 'owner',  'keyS': 'owner', 'keyFk': 'id', 'keyList': 'ownlist', 'infield': 'id' }]},
                {'section': 'other', 'filterlist': [
                    {'filter': 'coltype',   'dbfield': 'type',   'keyS': 'type',  'keyList': 'typelist' },
                    {'filter': 'settype',   'dbfield': 'settype','keyS': 'settype'},
                   ]}#  {'filter': 'scope',     'dbfield': 'scope',  'keyS': 'scope'}
                ]
        elif self.prefix == "hist":
            self.new_button = False
            self.settype = "hc"
            self.plural_name = "Historical Collections"
            self.sg_name = "Historical Collection"  
            self.order_cols = ['name', '', '', 'ssgauthornum', 'created']
            self.order_default = self.order_cols
            self.order_heads  = [
                {'name': 'Historical Collection',   'order': 'o=1', 'type': 'str', 'field': 'name', 'linkdetails': True},
                {'name': 'Authors',                 'order': '',    'type': 'str', 'custom': 'authors', 'allowwrap': True, 'main': True},
                {'name': '',                        'order': '',    'type': 'str', 'custom': 'saved',   'align': 'right'},
                {'name': 'Author count',            'order': 'o=4', 'type': 'int', 'custom': 'authcount'},
                #{'name': 'Added',                   'order': 'o=5', 'type': 'str', 'custom': 'created'}
            ]  
            # Add if user is app editor
            if user_is_authenticated(self.request) and user_is_ingroup(self.request, app_editor):
                self.order_heads.append({'name': 'Manuscript', 'order': '', 'type': 'str', 'custom': 'manuscript'})
                # Must also add to the order_cols and he order_default
                if len(self.order_default) < len(self.order_heads):
                    self.order_default.append("")
            self.filters = [ 
                {"name": "Collection",             "id": "filter_collection",     "enabled": False},
                {"name": "Project",                "id": "filter_project",        "enabled": False},
                {"name": "Authority file...",      "id": "filter_authority_file", "enabled": False, "head_id": "none"},
                {"name": "Manifestation...",       "id": "filter_sermon",         "enabled": False, "head_id": "none"},
                {"name": "Manuscript...",          "id": "filter_manuscript",     "enabled": False, "head_id": "none"},
                # Section SSG
                {"name": "Author",          "id": "filter_ssgauthor",       "enabled": False, "head_id": "filter_authority_file"},
                {"name": "Incipit",         "id": "filter_ssgincipit",      "enabled": False, "head_id": "filter_authority_file"},
                {"name": "Explicit",        "id": "filter_ssgexplicit",     "enabled": False, "head_id": "filter_authority_file"},
                {"name": "Passim code",     "id": "filter_ssgcode",         "enabled": False, "head_id": "filter_authority_file"}, 
                #{"name": "Number",          "id": "filter_ssgnumber",       "enabled": False, "head_id": "filter_authority_file"},
                {"name": "Gryson/Clavis/Other code","id": "filter_ssgsignature",    "enabled": False, "head_id": "filter_authority_file"},
                {"name": "Keyword",         "id": "filter_ssgkeyword",      "enabled": False, "head_id": "filter_authority_file"},
                {"name": "Status",          "id": "filter_ssgstype",        "enabled": False, "head_id": "filter_authority_file"},
                # Section S
                {"name": "Gryson/Clavis/Other code","id": "filter_sermosignature",  "enabled": False, "head_id": "filter_sermon"},
                {"name": "Author",          "id": "filter_sermoauthor",     "enabled": False, "head_id": "filter_sermon"},
                {"name": "Incipit",         "id": "filter_sermoincipit",    "enabled": False, "head_id": "filter_sermon"},
                {"name": "Explicit",        "id": "filter_sermoexplicit",   "enabled": False, "head_id": "filter_sermon"},
                {"name": "Keyword",         "id": "filter_sermokeyword",    "enabled": False, "head_id": "filter_sermon"}, 
                {"name": "Feast",           "id": "filter_sermofeast",      "enabled": False, "head_id": "filter_sermon"},
                {"name": "Bible reference", "id": "filter_bibref",          "enabled": False, "head_id": "filter_sermon"},
                {"name": "Note",            "id": "filter_sermonote",       "enabled": False, "head_id": "filter_sermon"},
                {"name": "Status",          "id": "filter_sermostype",      "enabled": False, "head_id": "filter_sermon"},
                # Section M
                {"name": "Shelfmark",       "id": "filter_manuid",          "enabled": False, "head_id": "filter_manuscript"},
                {"name": "Country",         "id": "filter_manucountry",     "enabled": False, "head_id": "filter_manuscript"},
                {"name": "City/Location",   "id": "filter_manucity",        "enabled": False, "head_id": "filter_manuscript"},
                {"name": "Library",         "id": "filter_manulibrary",     "enabled": False, "head_id": "filter_manuscript"},
                {"name": "Origin",          "id": "filter_manuorigin",      "enabled": False, "head_id": "filter_manuscript"},
                {"name": "Provenance",      "id": "filter_manuprovenance",  "enabled": False, "head_id": "filter_manuscript"},
                {"name": "Date range",      "id": "filter_manudaterange",   "enabled": False, "head_id": "filter_manuscript"},
                {"name": "Keyword",         "id": "filter_manukeyword",     "enabled": False, "head_id": "filter_manuscript"},
                {"name": "Status",          "id": "filter_manustype",       "enabled": False, "head_id": "filter_manuscript"},
                ]
            self.searches = [
                {'section': '', 'filterlist': [
                    {'filter': 'collection',    'dbfield': 'name',   'keyS': 'collection_ta', 'keyList': 'collist', 'infield': 'name'},
                    {'filter': 'project',       'fkfield': 'projects', 'keyFk': 'name', 'keyList': 'projlist', 'infield': 'name'},
                    ]},
                # Section SSG
                {'section': 'authority_file', 'name': 'Authority File', 'filterlist': [
                    {'filter': 'ssgauthor',    'fkfield': 'super_col__super__author',            
                     'keyS': 'ssgauthorname', 'keyFk': 'name', 'keyList': 'ssgauthorlist', 'infield': 'id', 'external': 'gold-authorname' },
                    {'filter': 'ssgincipit',   'dbfield': 'super_col__super__srchincipit',   'keyS': 'ssgincipit'},
                    {'filter': 'ssgexplicit',  'dbfield': 'super_col__super__srchexplicit',  'keyS': 'ssgexplicit'},                    
                    {'filter': 'ssgcode',      'fkfield': 'super_col__super', 'keyFk': 'code',          
                     'keyS': 'ssgcode', 'keyList': 'ssgpassimlist', 'infield': 'id', 'help': 'passimcode'},
                    {'filter': 'ssgnumber',    'dbfield': 'super_col__super__number',       'keyS': 'ssgnumber'},
                    {'filter': 'ssgsignature', 'fkfield': 'super_col__super__equal_goldsermons__goldsignatures', 
                     'keyS': 'ssgsignature', 'keyFk': 'code', 'keyId': 'signatureid', 'keyList': 'ssgsiglist', 'infield': 'code', 'help': 'signature'},
                    {'filter': 'ssgkeyword',   'fkfield': 'super_col__super__keywords',          
                     'keyFk': 'name', 'keyList': 'ssgkwlist', 'infield': 'id'},
                    {'filter': 'ssgstype',     'dbfield': 'super_col__super__stype',             
                     'keyList': 'ssgstypelist', 'keyType': 'fieldchoice', 'infield': 'abbr' },
                    ]},
                # Section S = Sermon = Manifestation
                {'section': 'sermon', 'name': 'Manifestation', 'filterlist': [
                    {'filter': 'sermoincipit',       'dbfield': 'super_col__super__equalgold_sermons__srchincipit',   'keyS': 'sermoincipit'},
                    {'filter': 'sermoexplicit',      'dbfield': 'super_col__super__equalgold_sermons__srchexplicit',  'keyS': 'sermoexplicit'},
                    {'filter': 'sermotitle',         'dbfield': 'super_col__super__equalgold_sermons__title',         'keyS': 'sermotitle'},                    
                    {'filter': 'sermofeast',         'fkfield': 'super_col__super__equalgold_sermons__feast', 'keyFk': 'name', 'keyList': 'sermofeastlist', 'infield': 'id'},
                    {'filter': 'bibref',             'dbfield': '$dummy',                                             'keyS': 'bibrefbk'},
                    {'filter': 'bibref',             'dbfield': '$dummy',                                             'keyS': 'bibrefchvs'},
                    {'filter': 'sermonote',          'dbfield': 'super_col__super__equalgold_sermons__additional',    'keyS': 'sermonote'},
                    {'filter': 'sermoauthor',        'fkfield': 'super_col__super__equalgold_sermons__author',            
                     'keyS': 'sermoauthorname', 'keyFk': 'name', 'keyList': 'sermoauthorlist', 'infield': 'id', 'external': 'sermo-authorname' },
                    {'filter': 'sermosignature',     
                     'fkfield': 'super_col__super__equalgold_sermons__signatures|super_col__super__equalgold_sermons__goldsermons__goldsignatures',        
                     'keyS': 'sermosignature', 'keyFk': 'code', 'keyId': 'signatureid', 'keyList': 'sermosiglist', 'infield': 'code', 'help': 'signature'},
                    {'filter': 'sermokeyword',       'fkfield': 'super_col__super__equalgold_sermons__keywords',          
                     'keyFk': 'name', 'keyList': 'sermokwlist', 'infield': 'id' }, 
                    {'filter': 'sermostype',         'dbfield': 'super_col__super__equalgold_sermons__stype',             
                     'keyList': 'sermostypelist', 'keyType': 'fieldchoice', 'infield': 'abbr' }                    ]},
                # Section M
                {'section': 'manuscript', 'filterlist': [
                    {'filter': 'manuid',        'fkfield': 'super_col__super__equalgold_sermons__msitem__manu',                   
                     'keyS': 'manuidno',    'keyFk': "idno", 'keyList': 'manuidlist', 'infield': 'id'},
                    {'filter': 'manucountry',       'fkfield': 'super_col__super__equalgold_sermons__msitem__manu__library__lcountry', 'keyS': 'country_ta',   'keyId': 'country',     'keyFk': "name"},
                    {'filter': 'manucity',          'fkfield': 'super_col__super__equalgold_sermons__msitem__manu__library__lcity',    'keyS': 'city_ta',      'keyId': 'city',        'keyFk': "name"},
                    {'filter': 'manulibrary',       'fkfield': 'super_col__super__equalgold_sermons__msitem__manu__library',                
                     'keyS': 'libname_ta',    'keyId': 'library',     'keyFk': "name"},
                    {'filter': 'manukeyword',       'fkfield': 'super_col__super__equalgold_sermons__msitem__manu__keywords',               
                     'keyFk': 'name', 'keyList': 'manukwlist', 'infield': 'name' },
                    {'filter': 'manustype',         'dbfield': 'super_col__super__equalgold_sermons__msitem__manu__stype',                  
                     'keyList': 'manustypelist', 'keyType': 'fieldchoice', 'infield': 'abbr' },
                    {'filter': 'manuprovenance',    'fkfield': 'super_col__super__equalgold_sermons__msitem__codico__provenances__location',  
                     'keyS': 'prov_ta',       'keyId': 'prov',        'keyFk': "name"},
                    {'filter': 'manuorigin',        'fkfield': 'super_col__super__equalgold_sermons__msitem__codico__origin',                 
                     'keyS': 'origin_ta',     'keyId': 'origin',      'keyFk': "name"},
                    {'filter': 'manudaterange',     'dbfield': 'super_col__super__equalgold_sermons__msitem__codico__codico_dateranges__yearstart__gte',         
                     'keyS': 'date_from'},
                    {'filter': 'manudaterange',     'dbfield': 'super_col__super__equalgold_sermons__msitem__codico__codico_dateranges__yearfinish__lte',        
                     'keyS': 'date_until'},
                    ]},             
                # Section Other
                {'section': 'other', 'filterlist': [
                    {'filter': 'owner',     'fkfield': 'owner',  'keyS': 'owner', 'keyFk': 'id', 'keyList': 'ownlist', 'infield': 'id' },
                    {'filter': 'coltype',   'dbfield': 'type',   'keyS': 'type',  'keyList': 'typelist' },
                    {'filter': 'settype',   'dbfield': 'settype','keyS': 'settype'},
                    {'filter': 'atype',    'dbfield': 'super_col__super__atype',    'keyS': 'atype'}, 
                    {'filter': 'scope',     'dbfield': 'scope',  'keyS': 'scope'}]}
                ]
                # ======== One-time adaptations ==============
        self.sel_button = self.settype
        
        listview_adaptations("collhist_list")
        
        # Make the profile available
        self.profile = Profile.get_user_profile(self.request.user.username)

        return None

    def add_to_context(self, context, initial):
        if self.prefix == "priv":
            context['prefix'] = self.prefix
            context['user_button'] = render_to_string('seeker/dataset_add.html', context, self.request)

        # Count the number of selected items
        iCount = SelectItem.get_selectcount(self.profile, self.settype)
        if iCount > 0:
            context['sel_count'] = str(iCount)

        return context

    def get_own_list(self):
        oErr = ErrHandle()
        qs = None
        try:
            # Get the user
            username = self.request.user.username
            user = User.objects.filter(username=username).first()
            # Get to the profile of this user
            if user is None:
                qs = Profile.objects.none()
                oErr.Status("CollectionListView/get_own_list: unknown user is [{}]".format(username))
            else:
                qs = Profile.objects.filter(user=user)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("CollectionListView/get_own_list")
        return qs

    def adapt_search(self, fields):
        lstExclude=None
        qAlternative = None
        if self.prefix == "hist":
            # The settype should be specified
            fields['settype'] = "hc"

            # The collection type is 'super'
            fields['type'] = "super"

            # The scope of a historical collection to be shown should be 'public'
            if user_is_authenticated(self.request) and user_is_ingroup(self.request, app_editor):
                profile = Profile.get_user_profile(self.request.user.username)
                fields['scope'] = ( ( Q(scope="priv") & Q(owner=profile) ) | Q(scope="team") | Q(scope="publ") )
            else:
                fields['scope'] = "publ"

            # Adapt the bible reference list
            bibrefbk = fields.get("bibrefbk", "")
            if bibrefbk != None and bibrefbk != "":
                bibrefchvs = fields.get("bibrefchvs", "")

                # Get the start and end of this bibref
                start, einde = Reference.get_startend(bibrefchvs, book=bibrefbk)

                # Find out which sermons have references in this range
                lstQ = []
                lstQ.append(Q(super_col__super__equalgold_sermons__sermonbibranges__bibrangeverses__bkchvs__gte=start))
                lstQ.append(Q(super_col__super__equalgold_sermons__sermonbibranges__bibrangeverses__bkchvs__lte=einde))
                collectionlist = [x.id for x in Collection.objects.filter(*lstQ).order_by('id').distinct()]

                fields['bibrefbk'] = Q(id__in=collectionlist)
            
            # Make sure we only use the Authority Files with accepted modifications
            # This means that atype should be 'acc' (and not: 'mod', 'rej' or 'def') 
            # With this condition we make sure ALL historical collections are in de unfiltered listview
            if fields['ssgcode'] != '':
                fields['atype'] = 'acc'
        elif self.prefix == "priv":
            # Show private, team and public datasets, provided the person is in the team
            # Bij navigeren naar Datasets is scope leeg                                 
            fields['settype'] = "pd"          
            ownlist = self.get_own_list()
            if user_is_ingroup(self.request, app_editor):
            # When filtering on scope:  
                # When navigating to My Datasets, scope is empty and colscope is None:
                if fields['scope'] == "":
                    # For the complete overview (private datasets of the user, team and public):
                    if fields['colscope'] == None:
                        # issue #446: use [defscope] for default setting
                        # fields['colscope'] = ( Q(scope="priv") & Q(owner__in=ownlist) | Q(scope="team") | Q(scope="publ"))
                        fields['defscope'] = ( Q(scope="priv") & Q(owner__in=ownlist) | Q(scope="team") | Q(scope="publ"))
                    # Filtering on scope Public:
                    elif fields['colscope'].abbr == 'publ':           
                        fields['colscope'] = (Q(scope="publ") )
                    # Filtering on scope Team:
                    elif fields['colscope'].abbr == 'team':                        
                        fields['colscope'] = (Q(scope="team") )
                    # Filtering on scope Private: 
                    elif fields['colscope'].abbr == 'priv':
                        fields['colscope'] = ( ( Q(scope="priv") & Q(owner__in=ownlist)  ))
                # Somehow after the initial first view of all datasets and the first filtering, scope remains "priv" 
                # whether you filter on any of the three options. Only colscope is changed when filtering on Private, Team or Public.        
                elif fields['scope'] == "priv":                                      
                    if fields['colscope'] != None:
                        # Filtering on scope Private:
                        if fields['colscope'].abbr == 'priv':
                            fields['colscope'] = ( ( Q(scope="priv") & Q(owner__in=ownlist)  )) 
                        # Filtering on scope Public:
                        elif fields['colscope'].abbr == 'publ':
                            fields['colscope'] = (Q(scope="publ") )
                        # Filtering on scope Team:
                        elif fields['colscope'].abbr == 'team':
                            fields['colscope'] = (Q(scope="team") )
                    # For the complete overview (private datasets of the user, team and public):
                    elif fields['colscope'] == None:
                        # issue #446: use [defscope] for default setting
                        # fields['colscope'] = ( Q(scope="priv") & Q(owner__in=ownlist) | Q(scope="team") | Q(scope="publ")) 
                        fields['defscope'] = ( Q(scope="priv") & Q(owner__in=ownlist) | Q(scope="team") | Q(scope="publ")) 
             
        
        elif self.prefix == "publ":
            # Show only public datasets
            fields['settype'] = "pd"
            # qAlternative = Q(scope="publ")
            fields['scope'] = "publ"
        else:
            # Check if the collist is identified
            if fields['ownlist'] == None or len(fields['ownlist']) == 0:
                # Get the user
                #username = self.request.user.username
                #user = User.objects.filter(username=username).first()
                ## Get to the profile of this user
                #qs = Profile.objects.filter(user=user)
                #profile = qs[0]
                #fields['ownlist'] = qs
                fields['ownlist'] = self.get_own_list()

                # Check on what kind of user I am
                if user_is_ingroup(self.request, app_editor):
                    # This is an editor: may see collections in the team
                    qAlternative = Q(scope="team") | Q(scope="publ")
                else:
                    # Common user: may only see those with public scope
                    # fields['scope'] = "publ"
                    qAlternative = Q(scope="publ")

            # Also make sure that we add the collection type, which is specified in "prefix"
            if self.prefix != "any":
                fields['type'] = self.prefix
            # The settype should be specified
            fields['settype'] = "pd"
        return fields, lstExclude, qAlternative

    def get_field_value(self, instance, custom):
        sBack = ""
        sTitle = ""
        if custom == "links":
            html = []
            # Get the HTML code for the links of this instance
            number = instance.freqsermo()
            if number > 0:
                url = reverse('sermon_list')
                html.append("<a href='{}?sermo-collist_s={}'>".format(url, instance.id))
                html.append("<span class='badge jumbo-1 clickable' title='Frequency in manifestation sermons'>{}</span></a>".format(number))
            number = instance.freqgold()
            if number > 0:
                url = reverse('search_gold')
                html.append("<a href='{}?gold-collist_sg={}'>".format(url, instance.id))
                html.append("<span class='badge jumbo-2 clickable' title='Frequency in gold sermons'>{}</span></a>".format(number))
            number = instance.freqmanu()
            if number > 0:
                url = reverse('search_manuscript')
                html.append("<a href='{}?manu-collist_m={}'>".format(url, instance.id))
                html.append("<span class='badge jumbo-3 clickable' title='Frequency in manuscripts'>{}</span></a>".format(number))
            number = instance.freqsuper()
            if number > 0:
                url = reverse('equalgold_list')
                html.append("<a href='{}?ssg-collist_ssg={}'>".format(url, instance.id))
                html.append("<span class='badge jumbo-3 clickable' title='Frequency in Authority files'>{}</span></a>".format(number))
            # Combine the HTML code
            sBack = "\n".join(html)
        elif custom == "type":
            sBack = instance.get_type_name()
        elif custom == "scope":
            sBack = instance.get_scope_display()
        elif custom == "created":
            sBack = get_crpp_date(instance.created, True)
        elif custom == "owner":
            if instance.owner is None:
                sBack = "(no user)"
            else:
                sBack = instance.owner.user.username
        elif custom == "authors":
            sBack = instance.get_authors_markdown()
        elif custom == "authcount":
            sBack = "{}".format(instance.ssgauthornum)
        elif custom == "manuscript":
            html = []
            url = reverse('collhist_manu', kwargs={'pk': instance.id})
            html.append("<a href='{}' title='Create a manuscript based on this historical collection'><span class='glyphicon glyphicon-open'></span></a>".format(url))
            url = reverse('collhist_temp', kwargs={'pk': instance.id})
            html.append("<a href='{}' title='Create a template based on this historical collection'><span class='glyphicon glyphicon-open' style='color: darkblue;'></span></a>".format(url))
            sBack = "\n".join(html)
        elif custom == "saved":
            # Prepare saveditem handling
            html = []
            saveditem_button = get_saveditem_html(self.request, instance, self.profile, sitemtype=self.settype)
            saveditem_form = get_saveditem_html(self.request, instance, self.profile, "form", sitemtype=self.settype)
            html.append(saveditem_button)
            html.append(saveditem_form)
            sBack = "".join(html)
        return sBack, sTitle

    def get_selectitem_info(self, instance, context):
        """Use the get_selectitem_info() defined earlier in this views.py"""
        return get_selectitem_info(self.request, instance, self.profile, self.settype, context)
    
    def get_helptext(self, name):
        """Use the get_helptext function defined in models.py"""
        print(name)
        return get_helptext(name)

# ================= COMMENT =============================

class CommentSend(BasicPart):
    """Receive a comment from a user"""

    MainModel = Comment
    template_name = 'seeker/comment_add.html'

    def add_to_context(self, context):

        url_names = {"manu": "manuscript_details", "sermo": "sermon_details",
                     "gold": "sermongold_details", "super": "equalgold_details",
                     "codi": "codico_details", "hc": "collhist_details"}
        obj_names = {"manu": "Manuscript", "sermo": "Sermon",
                     "gold": "Sermon Gold", "super": "Authority file",
                     "codi": "Codicological unit", "hc": "Historical collection"}
        def get_object(otype, objid):
            obj = None
            if otype == "manu":
                obj = Manuscript.objects.filter(id=objid).first()
            elif otype == "sermo":
                obj = SermonDescr.objects.filter(id=objid).first()
            elif otype == "gold":
                obj = SermonGold.objects.filter(id=objid).first()
            elif otype == "super":
                obj = EqualGold.objects.filter(id=objid).first()
            elif otype == "codi":
                obj = Codico.objects.filter(id=objid).first()
            elif otype == "hc":
                obj = Collection.objects.filter(id=objid).first()
            return obj

        oErr = ErrHandle()
        try:
            # Check validity
            if not self.userpermissions("w"):
                # Don't do anything
                return context

            if self.add:
                # Get the form
                form = CommentForm(self.qd, prefix="com")
                if form.is_valid():
                    cleaned = form.cleaned_data
                    # Yes, we are adding something new - check what we have
                    profile = cleaned.get("profile")
                    otype = cleaned.get("otype")
                    objid = cleaned.get("objid")
                    content = cleaned.get("content")
                    if content != None and content != "":
                        # Yes, there is a remark
                        comment = Comment.objects.create(profile=profile, content=content, otype=otype)
                        obj = get_object(otype, objid)
                        # Add a new object for this user 
                        obj.comments.add(comment) 

                        # Send this comment by email
                        objurl = reverse(url_names[otype], kwargs={'pk': obj.id}) # HIER GAAT HET MIS
                        context['objurl'] = self.request.build_absolute_uri(objurl)
                        context['objname'] = obj_names[otype]
                        context['comdate'] = comment.get_created()
                        context['user'] = profile.user
                        context['objcontent'] = content
                        contents = render_to_string('seeker/comment_mail.html', context, self.request)
                        comment.send_by_email(contents)

                        # Get a list of comments by this user for this item
                        context['comment_list'] = get_usercomments(otype, obj, profile)
                        # Translate this list into a valid string
                        comment_list = render_to_string('seeker/comment_list.html', context, self.request)
                        # And then pass on this string in the 'data' part of the POST response
                        #  (this is done using the BasicPart POST handling)
                        context['data'] = dict(comment_list=comment_list)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("CommentSend")

        # Send the result
        return context


class CommentEdit(BasicDetails):
    """The details of one comment"""

    model = Comment        
    mForm = None        # We are not using a form here!
    prefix = 'com'
    title = "User Comment"
    new_button = False
    # no_delete = True
    permission = "readonly"
    mainitems = []

    def check_hlist(self, instance):
        """Check if a hlist parameter is given, and hlist saving is called for"""

        oErr = ErrHandle()
        bChanges = False
        bDebug = True

        try:
            arg_hlist = "cresp-hlist"
            arg_savenew = "cresp-savenew"
            if arg_hlist in self.qd and arg_savenew in self.qd:
                # Interpret the list of information that we receive
                hlist = json.loads(self.qd[arg_hlist])
                # Interpret the savenew parameter
                savenew = self.qd[arg_savenew]

                # Make sure we are not saving
                self.do_not_save = True
                # But that we do a new redirect
                self.newRedirect = True

                # Change the redirect URL
                if self.redirectpage == "":
                    self.redirectpage = reverse('comment_details', kwargs={'pk': instance.id})

                # See if any need to be removed
                existing_item_id = [str(x.id) for x in CommentResponse.objects.filter(comment=instance)]
                delete_id = []
                for item_id in existing_item_id:
                    if not item_id in hlist:
                        delete_id.append(item_id)
                if len(delete_id)>0:
                    lstQ = [Q(comment=instance)]
                    lstQ.append(Q(**{"id__in": delete_id}))
                    CommentResponse.objects.filter(*lstQ).delete()

            return True
        except:
            msg = oErr.get_error_message()
            oErr.DoError("Comment/check_hlist")
            return False

    def add_to_context(self, context, instance):
        """Add to the existing context"""

        oErr = ErrHandle()
        try:
            # Define the main items to show and edit
            context['mainitems'] = [
                {'type': 'plain', 'label': "Timestamp:",    'value': instance.get_created(),    },
                {'type': 'plain', 'label': "User name:",    'value': instance.profile.user.username,     },
                {'type': 'plain', 'label': "Comment:",      'value': instance.content,     },
                {'type': 'plain', 'label': "Item type:",    'value': instance.get_otype()},
                {'type': 'safe',  'label': "Link:",         'value': self.get_link(instance)}
                ]

            # Check if we need to indicate that the comment has been read
            thisprofile = self.request.user.user_profiles.first()
            qs = CommentRead.objects.filter(comment=instance, profile=thisprofile)
            if qs.count() == 0:
                # Add that we have read it
                obj = CommentRead.objects.create(comment=instance, profile=thisprofile)

            # Figure out who I am
            profile = self.request.user.user_profiles.first()
            context['profile'] = profile

            # provide a button for the user to add a response to the comment
            context['after_details'] = render_to_string("seeker/comment_response.html", context, self.request)

        except:
            msg = oErr.get_error_message()
            oErr.DoError("CommentEdit/add_to_context")

        # Return the context we have made 
        return context

    def get_link(self, instance):
        url = ""
        label = ""
        sBack = ""
        if instance.otype == "manu":
            obj = instance.comments_manuscript.first()
            url = reverse("manuscript_details", kwargs={'pk': obj.id})
            label = "manu_{}".format(obj.id)
        elif instance.otype == "sermo":
            obj = instance.comments_sermon.first()
            url = reverse("sermon_details", kwargs={'pk': obj.id})
            label = "sermo_{}".format(obj.id)
        elif instance.otype == "gold":
            obj = instance.comments_gold.first()
            url = reverse("sermongold_details", kwargs={'pk': obj.id})
            label = "gold_{}".format(obj.id)
        elif instance.otype == "super":
            obj = instance.comments_super.first()
            url = reverse("equalgold_details", kwargs={'pk': obj.id})
            label = "super_{}".format(obj.id)
        if url != "":
            sBack = "<span class='badge signature gr'><a href='{}'>{}</a></span>".format(url, label)
 
        return sBack


class CommentDetails(CommentEdit):
    """Like Comment Edit, but then html output"""
    rtype = "html"

    def custom_init(self, instance):
        if not instance is None:
            self.check_hlist(instance)
        return None

    def add_to_context(self, context, instance):
        """Add to the existing context"""

        # Start by executing the standard handling
        super(CommentDetails, self).add_to_context(context, instance)

        def add_one_item(rel_item, value, resizable=False, title=None, align=None, link=None, main=None, draggable=None, editable=None, 
                         colspan=None, html=None):
            oAdd = dict(value=value, editable=False)
            if resizable: oAdd['initial'] = 'small'
            if title != None: oAdd['title'] = title
            if align != None: oAdd['align'] = align
            if link != None: oAdd['link'] = link
            if main != None: oAdd['main'] = main
            if draggable != None: oAdd['draggable'] = draggable
            if not colspan is None: oAdd['colspan'] = colspan
            if not editable is None: oAdd['editable'] = editable
            if not html is None: oAdd['html'] = html
            rel_item.append(oAdd)
            return True

        username = self.request.user.username
        team_group = app_editor

        # Authorization: only app-editors may edit!
        bMayEdit = user_is_ingroup(self.request, team_group)
        bEditableResponse = user_is_superuser(self.request) # True

        # On top of that (see issue #601): only editors may see the comment responses!!
            
        # All PDs: show the content
        related_objects = []
        lstQ = []
        rel_list =[]
        resizable = True
        index = 1
        sort_start = ""
        sort_start_int = ""
        sort_end = ""
        if bMayEdit:
            sort_start = '<span class="sortable"><span class="fa fa-sort sortshow"></span>&nbsp;'
            sort_start_int = '<span class="sortable integer"><span class="fa fa-sort sortshow"></span>&nbsp;'
            sort_end = '</span>'

        oErr = ErrHandle()

        try:

            # Additional sections
            context['sections'] = []

            # (see issue #601): only editors may see the comment responses!!
            if bMayEdit:

                # The choice is between an editable and a non-editable response
                if not bEditableResponse:

                    # [1] =============================================================
                    # Get all 'CommentResponse' objects that are part of this 'Comment'
                    responses = dict(title="Response(s) to this comment", prefix="cresp")
                    if resizable: responses['gridclass'] = "resizable"
                    responses['savebuttons'] = bMayEdit
                    responses['saveasbutton'] = True
                    responses['classes'] = ''
                
                    rel_list =[]
                    # Get a queryset with CommentResponse objects, newest first
                    qs_resp = instance.comment_cresponses.all().order_by('-created')

                    # Walk these collection sermons
                    for idx, obj in enumerate(qs_resp):
                        rel_item = []
                        # The item is the actual response
                        item = obj
                        order = idx + 1

                        # Get the URL to this response
                        url = reverse('commentresponse_details', kwargs={'pk': obj.id})

                        # Response: Order from enumeration
                        add_one_item(rel_item, order, False, align="right", link=url, draggable=True)

                        # Response: person who responded
                        add_one_item(rel_item, self.get_field_value("cresp", item, "person"), resizable, link=url) #, False)

                        # Response: The response itself
                        add_one_item(rel_item, self.get_field_value("cresp", item, "content"), resizable, link=url, main=True) #, False)

                        # Response: Date of the response
                        add_one_item(rel_item, self.get_field_value("cresp", item, "created"), resizable, link=url) #, False)

                        # Actions that can be performed on this item
                        if bMayEdit:
                            add_one_item(rel_item, self.get_actions())

                        # Add this line to the list
                        rel_list.append(dict(id=obj.id, cols=rel_item))
            
                    responses['rel_list'] = rel_list
                    responses['columns'] = [
                        '{}<span title="Default order">Order<span>{}'.format(sort_start_int, sort_end),
                        '{}<span title="Person">Person</span>{}'.format(sort_start, sort_end), 
                        '{}<span title="Response">Response</span>{}'.format(sort_start, sort_end), 
                        '{}<span title="Date of the response">Date</span>{}'.format(sort_start, sort_end)
                        ]
                    if bMayEdit:
                        responses['columns'].append("")
                    related_objects.append(responses)

                else:
                    # [2] =============================================================
                    # EXPERIMENT: editable individual responses
                    #
                    # Get all 'CommentResponse' objects that are part of this 'Comment'
                    ediresps = dict(title="Response(s) to this comment", prefix="cresp")
                    if resizable: ediresps['gridclass'] = "resizable"
                    ediresps['savebuttons'] = bMayEdit
                    ediresps['saveasbutton'] = True
                    ediresps['classes'] = ''
                    ediresps['editable'] = True

                    rel_list =[]
                    # Get a queryset with CommentResponse objects, newest first
                    qs_resp = instance.comment_cresponses.all().order_by('-created')

                    # Walk these collection sermons
                    for idx, obj in enumerate(qs_resp):
                        rel_item = []
                        # The item is the actual response
                        item = obj
                        order = idx + 1

                        # Get the URL to this response
                        url = reverse('commentresponse_details', kwargs={'pk': obj.id})

                        # Response: Order from enumeration
                        add_one_item(rel_item, order, False, align="right", link=url, draggable=True)

                        # Response: person who responded
                        sHtml = self.get_editable_form(obj)
                        add_one_item(rel_item, self.get_field_value("cresp", item, "person"), resizable, link=url, editable=True, colspan=3, html=sHtml) 

                        # Response: The response itself
                        add_one_item(rel_item, self.get_field_value("cresp", item, "content"), resizable, link=url, editable=True,  main=True) 

                        # Response: Date of the response
                        add_one_item(rel_item, self.get_field_value("cresp", item, "created"), resizable, link=url, editable=True, ) 

                        # Actions that can be performed on this item
                        if bMayEdit:
                            add_one_item(rel_item, self.get_actions())

                        # Add this line to the list
                        rel_list.append(dict(id=obj.id, cols=rel_item))
            
                    ediresps['rel_list'] = rel_list
                    ediresps['columns'] = [
                        '{}<span title="Default order">Order<span>{}'.format(sort_start_int, sort_end),
                        '{}<span title="Person">Person</span>{}'.format(sort_start, sort_end), 
                        '{}<span title="Response">Response</span>{}'.format(sort_start, sort_end), 
                        '{}<span title="Date of the response">Date</span>{}'.format(sort_start, sort_end)
                        ]
                    if bMayEdit:
                        ediresps['columns'].append("")
                    related_objects.append(ediresps)

            # Lists of related objects
            context['related_objects'] = related_objects

        except:
            msg = oErr.get_error_message()
            oErr.DoError("CommentDetails/add_to_context")

        # Return the context we have made
        return context

    def get_actions(self):
        html = []
        buttons = ['remove']    # This contains all the button names that need to be added

        # Start the whole div
        html.append("<div class='blinded'>")
        
        # Add components
        if 'remove' in buttons: 
            html.append("<a class='related-remove'><span class='glyphicon glyphicon-remove'></span></a>")

        # Finish up the div
        html.append("&nbsp;</div>")

        # COmbine the list into a string
        sHtml = "\n".join(html)
        # Return out HTML string
        return sHtml

    def get_editable_form(self, obj):
        """Get HTML with a <form> to define the editable fields' values"""

        template_name ="seeker/comment_edit.html"
        edifields = ['profile', 'content' ]
        form = CommentResponseForm(instance=obj, prefix="cresp")
        context = dict(form=form, instance=obj, edifields=edifields)
        sBack = render_to_string(template_name, context, self.request)
        return sBack

    def get_field_value(self, type, instance, custom, kwargs=None):
        sBack = ""

        oErr = ErrHandle()
        try:

            if type == "cresp":
                sBack, sTitle = "", ""
                if custom == "person":
                    sBack = instance.profile.user.username
                elif custom == "content":
                    sBack = instance.content
                elif custom == "created":
                    sBack = instance.get_created()
        except:
            msg = oErr.get_error_message()
            oErr.DoError("CommentDetails/get_field_value")

        return sBack
    

class CommentListView(BasicList):
    """Search and list comments"""

    model = Comment
    listform = CommentForm
    prefix = "com"
    sg_name = "User Comment"     
    plural_name = "User Comments"
    paginate_by = 20
    has_select2 = True
    order_cols = ['created', 'profile__user__username', '', 'otype', '']
    order_default = ['-created', 'profile__user__username', 'otype']
    order_heads = [
        {'name': 'Timestamp',   'order': 'o=1', 'type': 'str', 'custom': 'created', 'main': True, 'linkdetails': True},
        {'name': 'User name',   'order': 'o=2', 'type': 'str', 'custom': 'username'},
        {'name': 'Response',    'order': '',    'type': 'str', 'custom': 'response'},
        {'name': 'Item Type',   'order': 'o=4', 'type': 'str', 'custom': 'otype'},
        {'name': 'Link',        'order': '',    'type': 'str', 'custom': 'link'},
        ]
    filters = [ {"name": "Item type",   "id": "filter_otype",       "enabled": False},
                {"name": "User name",   "id": "filter_username",    "enabled": False}]
    searches = [
        {'section': '', 'filterlist': [
            {'filter': 'otype',    'dbfield': 'otype',   'keyS': 'otype',           'keyList': 'otypelist' },
            {'filter': 'username', 'fkfield': 'profile', 'keyFk': 'user__username', 'keyList': 'profilelist', 'infield': 'id'}
            ]
         }
        ]
    
    def initializations(self):
        """Perform some initializations"""

        # Check if otype has already been taken over
        comment_otype = Information.get_kvalue("comment_otype")
        if comment_otype == None or comment_otype != "done":
            # Get all the comments that have no o-type filled in yet
            qs = Comment.objects.filter(otype="-")
            with transaction.atomic():
                for obj in qs:
                    # Figure out where it belongs to
                    if obj.comments_manuscript.count() > 0:
                        obj.otype = "manu"
                    elif obj.comments_sermon.count() > 0:
                        obj.otype = "sermo"
                    elif obj.comments_gold.count() > 0:
                        obj.otype = "gold"
                    elif obj.comments_super.count() > 0:
                        obj.otype = "super"
                    elif obj.comments_codi.count() > 0:
                        obj.otype = "codi"
                    obj.save()
            # Success
            Information.set_kvalue("comment_otype", "done")

        return None

    def get_field_value(self, instance, custom):
        sBack = ""
        sTitle = ""
        oErr = ErrHandle()
        try:
            if custom == "username":
                sBack = instance.profile.user.username
            elif custom == "created":
                sBack = instance.get_created()
                # Check if it has not been read
                thisprofile = self.request.user.user_profiles.first()
                if instance.profile.id != thisprofile.id:
                    # This is not the same user, so check if it has been read
                    bRead = (CommentRead.objects.filter(comment=instance, profile=thisprofile).count() > 0)
                    if not bRead:
                        # Not read: so bolden it
                        sBack = "<b>{}</b>".format(sBack)
            elif custom == "response":
                sBack = "-"
                last = instance.comment_cresponses.last()
                if not last is None:
                    sBack = last.get_created()
            elif custom == "otype":
                sBack = instance.get_otype()
            elif custom == "link":
                url = ""
                label = ""
                if instance.otype == "manu":
                    obj = instance.comments_manuscript.first()
                    if not obj is None:
                        url = reverse("manuscript_details", kwargs={'pk': obj.id})
                        label = "manu_{}".format(obj.id)
                    else:
                        iStop = 1
                elif instance.otype == "sermo":
                    obj = instance.comments_sermon.first()
                    if obj is None:
                        iStop = 1
                    else:
                        url = reverse("sermon_details", kwargs={'pk': obj.id})
                        label = "sermo_{}".format(obj.id)
                elif instance.otype == "gold":
                    obj = instance.comments_gold.first()
                    if obj is None:
                        iStop = 1
                    else:
                        url = reverse("sermongold_details", kwargs={'pk': obj.id})
                        label = "gold_{}".format(obj.id)
                elif instance.otype == "super":
                    obj = instance.comments_super.first()
                    if obj is None:
                        iStop = 1
                    else:
                        url = reverse("equalgold_details", kwargs={'pk': obj.id})
                        label = "super_{}".format(obj.id)
                elif instance.otype == "codi":
                    obj = instance.comments_codi.first()
                    if obj is None:
                        iStop = 1
                    else:
                        url = reverse("codico_details", kwargs={'pk': obj.id})
                        label = "codi_{}".format(obj.id)
                if url != "":
                    sBack = "<span class='badge signature gr'><a href='{}'>{}</a></span>".format(url, label)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("CommentListView/get_field_value")
        return sBack, sTitle


# ================= COMMENT RESPONSE =============================

class CommentResponseEdit(BasicDetails):
    """The details of one comment"""

    model = CommentResponse        
    mForm = CommentResponseForm
    prefix = 'comr'
    prefix_type = "simple"
    title = "Comment Response"
    new_button = False
    listview = None     # Initially no listview, but that listview will become the comment itself
    listviewtitle = "User comment"
    # no_delete = True
    mainitems = []

    def custom_init(self, instance):
        # If this is an instance, then provide a link to the listview
        if not instance is None and not instance.id is None:
            self.listview = reverse('comment_details', kwargs={'pk': instance.comment.id})
        return None

    def add_to_context(self, context, instance):
        """Add to the existing context"""

        oErr = ErrHandle()
        try:
            # Figure out who I am
            profile = self.request.user.user_profiles.first()

            # If I have editing rights, no readonly
            if not context['is_app_editor']:
                self.permission = "readonly"

            # print("CommentResponseEdit: add_to_context status #1 = {}".format(instance.status))

            # Define the main items to show and edit
            comment_id = None if instance.comment is None else instance.comment.id
            profile_id = profile.id
            # Define the main items to show and edit
            context['mainitems'] = [
                # -------- HIDDEN field values ---------------
                {'type': 'plain', 'label': "Comment id",    'value': comment_id,    'field_key': "comment", 'empty': 'hide'},
                {'type': 'plain', 'label': "Profile id",    'value': profile_id,    'field_key': "profile", 'empty': 'hide'},
                # --------------------------------------------
                {'type': 'safe',  'label': "Comment:",      'value': instance.get_comment(),    },
                {'type': 'plain', 'label': "Timestamp:",    'value': instance.get_saved(),      },
                {'type': 'plain', 'label': "User name:",    'value': instance.profile.user.username,     },
                {'type': 'safe',  'label': "Response:",     'value': instance.get_content(),   'field_key': "content"       },
                {'type': 'plain', 'label': "Visible:",      'value': instance.get_visible(),   'field_key': "visible",
                 'title': 'Indicate whether the response may be visible to other users in the Overview' },
                {'type': 'plain', 'label': "Status:",       'value': instance.get_status(),     },
                {'type': 'safe',  'label': "",              'value': self.get_button(instance), },
                ]

            #print("CommentResponseEdit: add_to_context status #2 = {}".format(instance.status))

            # Note: this must always be here, to be ready
            # provide a button for the user to add a response to the comment
            context['response_id'] = instance.id
            # Only show if there is content
            if not instance.content is None and not instance.content == "":
                context['after_details'] = render_to_string("seeker/commentresponse_send.html", context, self.request)
            # print("CommentResponseEdit: add_to_context status #3 = {}".format(instance.status))

        except:
            msg = oErr.get_error_message()
            oErr.DoError("CommentResponseEdit/add_to_context")

        # Return the context we have made 
        return context

    def get_button(self, instance):
        """Get a button to send this response, if needed
        
        NOTE: the BUTTON <a>...</a> is here, and gets displayed in the 'Edit' section
              the FORM <form>...</form> that is called 'create_new_response' is loaded in 'after_details'
        """

        sBack = ""
        oErr = ErrHandle()
        try:
            if instance.status == "changed":
                # The user should send it
                lHtml = []
                lHtml.append('<a class="btn btn-xs jumbo-3" role="button" ')
                lHtml.append('   onclick="document.getElementById(\'send_response\').submit();">')
                lHtml.append('<span class="glyphicon glyphicon-plus"></span>')
                lHtml.append('Email this response</a>')
                sBack = "\n".join(lHtml)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("")
        return sBack


class CommentResponseDetails(CommentResponseEdit):
    """Like CommentResponse Edit, but then html output"""
    rtype = "html"


class CommentResponseSend(CommentResponseDetails):
    """Actually send the response to a comment"""

    initRedirect = True

    def custom_init(self, instance):
        data = dict(status="ok")
        url_names = {"manu": "manuscript_details", "sermo": "sermon_details",
                     "gold": "sermongold_details", "super": "equalgold_details",
                     "codi": "codico_details", "hc": "collhist_details"}
        obj_names = {"manu": "Manuscript", "sermo": "Sermon",
                     "gold": "Sermon Gold", "super": "Authority file",
                     "codi": "Codicological unit", "hc": "Historical collection"}
       
        oErr = ErrHandle()
        try:
            # Figure out where to go to after processing this
            self.redirectpage = reverse('commentresponse_details', kwargs={'pk': instance.id})

            if not instance is None and user_is_ingroup(self.request, app_editor):

                comment = instance.comment
                if not comment is None and not comment.content is None:
                    # Whenever this commentresponses is saved, it should be sent to:
                    recipient = comment.profile

                    # Get the object from the comment
                    obj = comment.get_object()
                    otype = comment.otype
                    objurl = reverse(url_names[otype], kwargs={'pk': obj.id})

                    # Determine the context for the response email
                    context = {}
                    context['objurl'] = self.request.build_absolute_uri(objurl)
                    context['objname'] = obj_names[otype]
                    context['comdate'] = comment.get_created()
                    context['user'] = comment.profile.user
                    context['objcontent'] = comment.content
                    context['respdate'] = instance.get_saved()
                    context['respcontent'] = instance.get_content()

                    # Now create the appropriate response email
                    contents = render_to_string('seeker/commentresponse_mail.html', context, self.request)

                    # And here is to the method that sends it
                    instance.send_by_email(recipient, contents)

                    print("CommentResponseSend: done")
        except:
            msg = oErr.get_error_message()
            oErr.DoError("CommentResponseSend")

        return None


# ================= MANUSCRIPT =============================

class ManuscriptEdit(BasicDetails):
    """The details of one manuscript"""

    model = Manuscript  
    mForm = ManuscriptForm
    prefix = 'manu'
    titlesg = "Manuscript identifier"
    rtype = "json"
    new_button = True
    mainitems = []
    use_team_group = True
    history_button = True
    comment_button = True
    
    McolFormSet = inlineformset_factory(Manuscript, CollectionMan,
                                       form=ManuscriptCollectionForm, min_num=0,
                                       fk_name="manuscript", extra=0)
    MlitFormSet = inlineformset_factory(Manuscript, LitrefMan,
                                         form = ManuscriptLitrefForm, min_num=0,
                                         fk_name = "manuscript",
                                         extra=0, can_delete=True, can_order=False)
    MprovFormSet = inlineformset_factory(Manuscript, ProvenanceMan,
                                         form=ManuscriptProvForm, min_num=0,
                                         fk_name = "manuscript",
                                         extra=0, can_delete=True, can_order=False)
    MextFormSet = inlineformset_factory(Manuscript, ManuscriptExt,
                                         form=ManuscriptExtForm, min_num=0,
                                         fk_name = "manuscript",
                                         extra=0, can_delete=True, can_order=False)    
    MlinkFormSet = inlineformset_factory(Manuscript, ManuscriptLink,
                                         form=ManuscriptLinkForm, min_num=0,
                                         fk_name = "src",
                                         extra=0, can_delete=True, can_order=False)

    formset_objects = [
        {'formsetClass': McolFormSet,  'prefix': 'mcol',  'readonly': False, 'noinit': True, 'linkfield': 'manuscript'},
        {'formsetClass': MlitFormSet,  'prefix': 'mlit',  'readonly': False, 'noinit': True, 'linkfield': 'manuscript'},
        {'formsetClass': MprovFormSet, 'prefix': 'mprov', 'readonly': False, 'noinit': True, 'linkfield': 'manuscript'},
        {'formsetClass': MextFormSet,  'prefix': 'mext',  'readonly': False, 'noinit': True, 'linkfield': 'manuscript'},
        {'formsetClass': MlinkFormSet, 'prefix': 'mlink', 'readonly': False, 'noinit': True, 'initial': [{'linktype': LINK_PARTIAL }], 'clean': True},
        #{'formsetClass': MprojFormSet, 'prefix': 'mproj', 'readonly': False, 'noinit': True, 'linkfield': 'manuscript'}
        ]
    
    form_objects = [{'form': ManuReconForm, 'prefix': 'mrec', 'readonly': False}]

    stype_edi_fields = ['library', 'lcountry', 'lcity', 'idno', 'origin', 'source', #'project', # PROJECT_MOD_HERE
                        'hierarchy',
                        'LitrefMan', 'litlist',
                        'ManuscriptExt', 'extlist']

    def custom_init(self, instance):
        if instance != None and instance.mtype == "rec":
            # Also adapt the title
            self.titlesg = "Reconstructed manuscript"
        return None

    def add_to_context(self, context, instance):
        """Add to the existing context"""

        oErr = ErrHandle()
        try:
            # Need to know who this user (profile) is
            profile = Profile.get_user_profile(self.request.user.username)

            # Check if this is creating a reconstructed manuscript
            if instance != None and "manu-codicostart" in self.qd:
                # Get the codicostart
                codicostart = self.qd.get("manu-codicostart")
                codico = Codico.objects.filter(id=codicostart).first()
                if codico != None:
                    # Set the mtype to Reconstruction
                    instance.mtype = "rec"
                    instance.save()
                    # Create a Reconstruction object
                    obj = Reconstruction.objects.filter(codico=codico, manuscript=instance).first()
                    if obj == None:
                        obj = Reconstruction.objects.create(codico=codico, manuscript=instance, order=1)

            istemplate = (instance.mtype == "tem")
            # Make sure we mark reconstructed in the context
            context['reconstructed'] = (instance.mtype == "rec")

            # Define the main items to show and edit
            context['mainitems'] = []
            # Possibly add the Template identifier
            if istemplate:
                context['mainitems'].append(
                    {'type': 'plain', 'label': "Template:", 'value': instance.get_template_link(profile)}
                    )

            # Prepare saveditem handling
            saveditem_button = get_saveditem_html(self.request, instance, profile, sitemtype="manu")
            saveditem_form = get_saveditem_html(self.request, instance, profile, "form", sitemtype="manu")

            # If this user belongs to the ProjectApprover of HUWA, show him the HUWA ID if it is there
            if not istemplate and not instance is None:
                # NOTE: this code should be extended to include other external projects, when the time is right
                ext = ManuscriptExternal.objects.filter(manu=instance).first()
                if not ext is None:
                    # Get the field values
                    externaltype = ext.externaltype
                    externalid = ext.externalid
                    # Issue #730.3
                    # bEditHuwa = profile.is_project_approver("huwa")
                    bEditHuwa = profile.is_project_editor("huwa")
                    if externaltype == "huwop" and externalid > 0 and bEditHuwa:
                        # Okay, the user has the right to see the externalid
                        oItem = dict(type="plain", label="HUWA id", value=externalid, field_key="externalid")
                        context['mainitems'].insert(0, oItem)

            # Get the main items
            mainitems_main = [
                {'type': 'plain', 'label': "Status:",       'value': instance.get_stype_light(),      'field_key': 'stype'},
                {'type': 'safe',  'label': "Saved item:",   'value': saveditem_button          },
                {'type': 'plain', 'label': "Country:",      'value': instance.get_country(),          'field_key': 'lcountry'},
                {'type': 'plain', 'label': "City:",         'value': instance.get_city(),             'field_key': 'lcity',
                 'title': 'City, village or abbey (monastery) of the library'},
                {'type': 'safe',  'label': "Library:",      'value': instance.get_library_markdown(), 'field_key': 'library'},
                {'type': 'plain', 'label': "Shelfmark:",    'value': instance.idno,                   'field_key': 'idno'},
                # Project assignment: see below
                # {'type': 'plain', 'label': "Project:",      'value': instance.get_project_markdown(),       'field_key': 'project'},
                # {'type': 'plain', 'label': "Project2:",      'value': instance.get_project_markdown2(),       'field_key': 'project2'},
                ]
            for item in mainitems_main: context['mainitems'].append(item)
            if not istemplate:
                username = profile.user.username
                team_group = app_editor
                mainitems_m2m = [
                    {'type': 'plain', 'label': "Keywords:",     'value': instance.get_keywords_markdown(),      'field_list': 'kwlist'},
                    {'type': 'line', 'label': "Keywords (user):", 'value': self.get_userkeywords(instance, profile, context),   
                     'field_list': 'ukwlist', 'title': 'User-specific keywords. If the moderator accepts these, they move to regular keywords.'},
                    #{'type': 'plain', 'label': "Keywords (user):", 'value': instance.get_keywords_user_markdown(profile),   'field_list': 'ukwlist',
                    # 'title': 'User-specific keywords. If the moderator accepts these, they move to regular keywords.'},
                    {'type': 'plain', 'label': "Personal Datasets:",  'value': instance.get_collections_markdown(username, team_group, settype="pd"), 
                        'multiple': True, 'field_list': 'collist', 'fso': self.formset_objects[0] },
                    {'type': 'plain', 'label': "Literature:",   'value': instance.get_litrefs_markdown(), 
                        'multiple': True, 'field_list': 'litlist', 'fso': self.formset_objects[1], 'template_selection': 'ru.passim.litref_template' },

                    # Project2 HIER
                    {'type': 'plain', 'label': "Project:", 'value': instance.get_project_markdown2()},

                    {'type': 'plain', 'label': "Provenances:",          'value': self.get_provenance_markdown(instance), 
                        'multiple': True, 'field_list': 'mprovlist',    'fso': self.formset_objects[2] },
                    {'type': 'line',  'label': "Related manuscripts:",  'value': instance.get_manulinks_markdown(), 
                        'multiple': True,  'field_list': 'mlinklist',   'fso': self.formset_objects[4]},
                    ]
                for item in mainitems_m2m: context['mainitems'].append(item)

                # Possibly append notes view
                if user_is_ingroup(self.request, app_editor):
                    context['mainitems'].append(
                        {'type': 'plain', 'label': "Notes:",       'value': instance.get_notes_markdown(),  'field_key': 'notes'}  )
                    # Also add a view on the editornotes, if available
                    editornotes = instance.editornotes
                    if not editornotes is None:
                        context['mainitems'].append(
                            {'type': 'safe', 'label': "Editor notes (Dutch):", 'value': markdown(editornotes),  'field_key': 'editornotes'}  )

                # Always append external links and the buttons for codicological units
                context['mainitems'].append({'type': 'plain', 'label': "External links:",   'value': instance.get_external_markdown(), 
                        'multiple': True, 'field_list': 'extlist', 'fso': self.formset_objects[3] })
                context['mainitems'].append(
                    {'type': 'safe', 'label': 'Codicological:', 'value': self.get_codico_buttons(instance, context)}
                    )

                # Check if this is an editor with permission for this project
                if may_edit_project(self.request, profile, instance):
                    for oItem in context['mainitems']:
                        if oItem['label'] == "Project:":
                            # Add the list
                            oItem['field_list'] = "projlist"

            # Signal that we have select2
            context['has_select2'] = True

            # Specify that the manuscript info should appear at the right
            title_right = '<span style="font-size: xx-small">{}</span>'.format(instance.get_full_name())
            context['title_right'] = title_right

            # Note: non-app editors may still add a comment
            lhtml = []
            if context['is_app_editor']:
                lbuttons = []
                template_import_button = "import_template_button"
                has_sermons = (instance.manuitems.count() > 0)

                # Action depends on template/not
                if not istemplate:
                    # Also add the manuscript download code
                    local_context = dict(
                        ajaxurl     = reverse("manuscript_download", kwargs={'pk': instance.id}),
                        is_superuser=user_is_superuser(self.request),
                        csrf        = '<input type="hidden" name="csrfmiddlewaretoken" value="{}" />'.format(
                                                get_token(self.request)))
                    lbuttons.append(dict(html=render_to_string('seeker/manuscript_download.html', local_context, self.request)))

                    if instance.mtype != "rec":
                        # Add a button so that the user can import sermons + hierarchy from an existing template
                        if not has_sermons:
                            lbuttons.append(dict(title="Import Manifestations from a template", 
                                        id=template_import_button, open="import_from_template", label="Import from template..."))

                        # Add a button so that the user can turn this manuscript into a `Template`
                        lbuttons.append(dict(title="Create template from this manuscript", 
                                     submit="create_new_template", label="Create template..."))
                # Some buttons are needed anyway...
                lbuttons.append(dict(title="Open a list of origins", href=reverse('origin_list'), label="Origins..."))
                lbuttons.append(dict(title="Open a list of locations", href=reverse('location_list'), label="Locations..."))

                # Build the HTML on the basis of the above
                lhtml.append("<div class='row'><div class='col-md-12 container-small' align='right'><form method='post'>")
                for item in lbuttons:
                    if 'html' in item:
                        lhtml.append(item['html'])
                    else:
                        idfield = ""
                        if 'click' in item:
                            ref = " onclick='document.getElementById(\"{}\").click();'".format(item['click'])
                        elif 'submit' in item:
                            ref = " onclick='document.getElementById(\"{}\").submit();'".format(item['submit'])
                        elif 'open' in item:
                            ref = " data-toggle='collapse' data-target='#{}'".format(item['open'])
                        else:
                            ref = " href='{}'".format(item['href'])
                        if 'id' in item:
                            idfield = " id='{}'".format(item['id'])
                        lhtml.append("  <a role='button' class='btn btn-xs jumbo-3' title='{}' {} {}>".format(item['title'], ref, idfield))
                        lhtml.append("     <span class='glyphicon glyphicon-chevron-right'></span>{}</a>".format(item['label']))
                lhtml.append("</form></div></div>")

                if not istemplate:
                    # Add HTML to allow for the *creation* of a template from this manuscript
                    local_context = dict(manubase=instance.id)
                    lhtml.append(render_to_string('seeker/template_create.html', local_context, self.request))

                    # Add HTML to allow the user to choose sermons from a template
                    local_context['frmImport'] = TemplateImportForm({'manu_id': instance.id})
                    local_context['import_button'] = template_import_button
                    lhtml.append(render_to_string('seeker/template_import.html', local_context, self.request))

            #if instance.mtype in ["rec", "man"]:
            # Add Codico items - depending on reconstructed or not
            if instance.mtype == "rec":
                # Note: we need to go through Reconstruction, 
                #       since that table has the correct 'order' values for the reconstruction
                codicos = [x.codico for x in instance.manuscriptreconstructions.all().order_by('order')]
            else:
                # Note: we need to go directly to Codico, since the order values are there
                codicos = instance.manuscriptcodicounits.all().order_by('order')
            codico_list = []
            for codico in codicos:
                # Get the codico details URL
                url = reverse("codico_details", kwargs={'pk': codico.id})
                url_manu = reverse("manuscript_details", kwargs={'pk': codico.manuscript.id})
                # Get the list of manuscripts that use this codico
                manu_recons = []
                for recon in codico.codicoreconstructions.filter(manuscript__mtype='rec'):
                    url_recon = reverse("manuscript_details", kwargs={'pk': recon.manuscript.id})
                    manu_recons.append(url_recon)
                # Add the information to the codico list
                codico_list.append( dict(url=url, url_manu=url_manu, kvlist=self.get_kvlist(codico, instance), 
                                         codico_id=codico.id, recons=manu_recons) )
            context['codico_list'] = codico_list

            # Make sure to add the mtype to the context
            context['mtype'] = instance.mtype
            lhtml.append(render_to_string("seeker/codico_list.html", context, self.request))

            # Add comment modal stuff
            initial = dict(otype="manu", objid=instance.id, profile=profile)
            context['commentForm'] = CommentForm(initial=initial, prefix="com")
            context['comment_list'] = get_usercomments('manu', instance, profile)            
            
            lhtml.append(render_to_string("seeker/comment_add.html", context, self.request))
            context['comment_count'] = instance.comments.count()

            # Also add the saved item form
            lhtml.append(saveditem_form)

            # Store the after_details in the context
            context['after_details'] = "\n".join(lhtml)

            # Make stable URI available
            context['stable_uri'] = self.get_abs_uri('manuscript_details', instance)

        except:
            msg = oErr.get_error_message()
            oErr.DoError("ManuscriptEdit/add_to_context")

        # Return the context we have made
        return context

    def get_codico_buttons(self, instance, context):
        sBack = ""
        # The number of codico's depends on the mtype of the manuscript
        if instance.mtype == "rec":
            context['codico_count'] = instance.manuscriptreconstructions.count()
        else:
            context['codico_count'] = instance.manuscriptcodicounits.count()
        lhtml = []
        lhtml.append(render_to_string("seeker/codico_buttons.html", context, self.request))
        sBack = "\n".join(lhtml)
        return sBack

    def get_kvlist(self, codico, manu):
        """Get a list of fields and values"""

        # Get a list of sermon information for this codico
        sermon_list = []
        for msitem in codico.codicoitems.all().order_by('order'):
            for sermon in msitem.itemsermons.all():
                # Add information of this sermon to the list
                sermon_url = reverse('sermon_details', kwargs={'pk': sermon.id})
                sermon_html = "<span class='badge signature ot'><a href='{}'>{}</a></span>".format(sermon_url, sermon.locus)
                sermon_list.append(sermon_html)
        # Action depends on the size of the list
        if len(sermon_list) == 0:
            sermons = "(none)"
        elif len(sermon_list) == 1:
            sermons = sermon_list[0]
        else:
            sermons = "{}...{}".format(sermon_list[0], sermon_list[-1])
        # OLD:
        # sermons = ", ".join(sermon_list)
        lkv = []
        if codico.manuscript.id == manu.id:
            lkv.append(dict(label="Order", value=codico.order))
        else:
            # Get the 'reconstruction' element
            reconstruction = Reconstruction.objects.filter(manuscript=manu, codico=codico).first()
            if reconstruction != None:
                # sOrder = "{} (in identifier: {})".format(reconstruction.order, codico.order)
                sOrder = "{}".format(reconstruction.order)
                lkv.append(dict(label="Order", value=sOrder))
        lkv.append(dict(label="Sermons", value=sermons))
        lkv.append(dict(label="Title", value=codico.name))
        lkv.append(dict(label="Date", value=codico.get_date_markdown()))
        lkv.append(dict(label="Support", value=codico.support))
        lkv.append(dict(label="Extent", value=codico.extent))
        lkv.append(dict(label="Format", value=codico.format))
        lkv.append(dict(label="Keywords", value=codico.get_keywords_markdown()))
        lkv.append(dict(label="Origin", value=self.get_codiorigin_markdown(codico)))
        lkv.append(dict(label="Provenances", value=self.get_codiprovenance_markdown(codico)))
        lkv.append(dict(label="Notes", value=codico.get_notes_markdown()))
        return lkv

    def get_codiorigin_markdown(self, codico):
        """Calculate a collapsable table view of the origins for this codico, for Codico details view"""

        context = dict(codi=codico)
        sBack = render_to_string("seeker/codi_origins.html", context, self.request)
        return sBack

    def get_codiprovenance_markdown(self, codico):
        """Calculate a collapsable table view of the provenances for this codico, for Codico details view"""

        context = dict(codi=codico)
        sBack = render_to_string("seeker/codi_provs.html", context, self.request)
        return sBack

    def get_form_kwargs(self, prefix):
        # This is for manulink

        oBack = None
        if prefix == "mlink":
            if self.object != None:
                # Make sure to return the ID of the Manuscript
                oBack = dict(manu_id=self.object.id)

        return oBack

    def get_provenance_markdown(self, instance):
        """Calculate a collapsible table view of the provenances for this manuscript, for Manu details view"""

        context = dict(manu=instance)
        sBack = render_to_string("seeker/manu_provs.html", context, self.request)
        return sBack

    def get_userkeywords(self, instance, profile, context):
        sBack = ""
        oErr = ErrHandle()
        try:
            # At the very least get the list of user keywords
            sKeywordList = instance.get_keywords_user_markdown(profile)
            # Now determine what interface to show
            if context['is_app_editor']:
                # We simply return the list
                sBack = sKeywordList
            else:
                # We need to construct an interface:
                # (1) Set up the context to include view and edit elements
                context['userkeywords'] = sKeywordList
                context['ukwform'] = context['manuForm']
                context['ukwupdate'] = reverse('manuscript_ukw', kwargs={'pk': instance.id })

                # Combine and provide the code needed
                # sBack = get_userkeyword_interface(self.request, context)
                sBack = render_to_string("seeker/userkeyword_edit.html", context, self.request)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("get_userkeywords")
        return sBack

    def process_formset(self, prefix, request, formset):
        errors = []
        bResult = True
        instance = formset.instance
        for form in formset:
            if form.is_valid():
                cleaned = form.cleaned_data
                # Action depends on prefix

                #if prefix == "mdr":
                #    # Processing one daterange
                #    newstart = cleaned.get('newstart', None)
                #    newfinish = cleaned.get('newfinish', None)
                #    oneref = cleaned.get('oneref', None)
                #    newpages = cleaned.get('newpages', None)

                #    if newstart:
                #        # Possibly set newfinish equal to newstart
                #        if newfinish == None or newfinish == "":
                #            newfinish = newstart
                #        # Double check if this one already exists for the current instance
                #        obj = instance.manuscript_dateranges.filter(yearstart=newstart, yearfinish=newfinish).first()
                #        if obj == None:
                #            form.instance.yearstart = int(newstart)
                #            form.instance.yearfinish = int(newfinish)
                #        # Do we have a reference?
                #        if oneref != None:
                #            form.instance.reference = oneref
                #            if newpages != None:
                #                form.instance.pages = newpages
                #        # Note: it will get saved with formset.save()
                if prefix == "mcol":
                    # Collection processing
                    newcol = cleaned.get('newcol', None)
                    if newcol != None:
                        # Is the COL already existing?
                        obj = Collection.objects.filter(name=newcol).first()
                        if obj == None:
                            # TODO: add profile here
                            profile = Profile.get_user_profile(request.user.username)
                            obj = Collection.objects.create(name=newcol, type='manu', owner=profile)
                        # Make sure we set the keyword
                        form.instance.collection = obj
                        # Note: it will get saved with formset.save()
                elif prefix == "mlit":
                    # Literature reference processing
                    newpages = cleaned.get("newpages")
                    # Also get the litref
                    oneref = cleaned.get("oneref")
                    if oneref:
                        litref = cleaned['oneref']
                        # Check if all is in order
                        if litref:
                            form.instance.reference = litref
                            if newpages:
                                form.instance.pages = newpages
                    # Note: it will get saved with form.save()
                elif prefix == "mprov":
                    # ========= OLD (issue #289) =======
                    #name = cleaned.get("name")
                    #note = cleaned.get("note")
                    #location = cleaned.get("location")
                    #prov_new = cleaned.get("prov_new")proccess
                    #if name:
                    #    obj = Provenance.objects.filter(name=name, note=note, location=location).first()
                    #    if obj == None:
                    #        obj = Provenance.objects.create(name=name)
                    #        if note: obj.note = note
                    #        if location: obj.location = location
                    #        obj.save()
                    #    if obj:
                    #        form.instance.provenance = obj
                    # New method, issue #289 (last part)
                    note = cleaned.get("note")
                    prov_new = cleaned.get("prov_new")
                    if prov_new != None:
                        form.instance.provenance = prov_new
                        form.instance.note = note

                    # Note: it will get saved with form.save()
                elif prefix == "mext":
                    newurl = cleaned.get('newurl')
                    if newurl:
                        form.instance.url = newurl
                #elif prefix == "mproj":
                #    proj_new = cleaned.get("proj_new")
                #    if proj_new != None:
                #        form.instance.project = proj_new

                    # Note: it will get saved with [sub]form.save()

                elif prefix == "mlink":
                    # Process the link from M to M
                    newmanu = cleaned.get("newmanu")
                    if not newmanu is None:
                        # There also must be a linktype
                        if 'newlinktype' in cleaned and cleaned['newlinktype'] != "":
                            linktype = cleaned['newlinktype']
                            # Get optional parameters
                            note = cleaned.get('note', None)
                            # Check existence
                            obj = ManuscriptLink.objects.filter(src=instance, dst=newmanu, linktype=linktype).first()
                            if obj == None:
                                manu = Manuscript.objects.filter(id=newmanu.id).first()
                                if manu != None:

                                    # Set the right parameters for creation later on
                                    form.instance.linktype = linktype
                                    form.instance.dst = manu
                                    if note != None and note != "": 
                                        form.instance.note = note

                                    # Double check reverse
                                    if linktype in LINK_BIDIR_MANU:
                                        rev_link = ManuscriptLink.objects.filter(src=manu, dst=instance).first()
                                        if rev_link == None:
                                            # Add it
                                            rev_link = ManuscriptLink.objects.create(src=manu, dst=instance, linktype=linktype)
                                        else:
                                            # Double check the linktype
                                            if rev_link.linktype != linktype:
                                                rev_link.linktype = linktype
                                        if note != None and note != "": 
                                            rev_link.note = note
                                        rev_link.save()
                    # Note: it will get saved with form.save()

            else:
                errors.append(form.errors)
                bResult = False
        return None

    def before_save(self, form, instance):
        oErr = ErrHandle()
        bBack = True
        msg = ""
        try:
            if instance != None:
                # Need to know who is 'talking'...
                username = self.request.user.username
                profile = Profile.get_user_profile(username)

                # If there is no source, then create a source for this one
                if instance.source == None:
                    source = SourceInfo.objects.create(
                        code="Manually created",
                        collector=username, 
                        profile=profile)
                    instance.source = source

                # Issue #473: automatic assignment of project for particular editor(s)
                projlist = form.cleaned_data.get("projlist")
                bBack, msg = evaluate_projlist(profile, instance, projlist, "Manuscript")
        except:
            msg = oErr.get_error_message()
            oErr.DoError("ManuscriptEdit/before_save")
            bBack = False
        return bBack, msg

    def after_save(self, form, instance):

        def adapt_external(field, externaltype):
            """Save the externalid"""

            oErr = ErrHandle()
            try:
                externalid = form.cleaned_data.get(field)
                if not externalid is None and externalid != "":
                    # Check if this is the correct id
                    ext = instance.manuexternals.filter(externaltype=externaltype).first()
                    if ext is None:
                        # Need to make one
                        ext = ManuscriptExternal.objects.create(manu=instance, externaltype=externaltype, externalid=externalid)
                    else:
                        # Do we need changing?
                        if externalid != ext.externalid:
                            ext.externalid = externalid
                            ext.save()
            except:
                msg = oErr.get_error_message()
                oErr.DoError("adapt_external")
                bResult = False

        msg = ""
        bResult = True
        oErr = ErrHandle()
        
        try:
            # External id changes
            adapt_external("externalid", "huwop")

            # Process many-to-many changes: Add and remove relations in accordance with the new set passed on by the user
            # (1) 'collections'
            collist_m = form.cleaned_data['collist']
            adapt_m2m(CollectionMan, instance, "manuscript", collist_m, "collection")

            # (2) 'keywords'
            kwlist = form.cleaned_data['kwlist']
            adapt_m2m(ManuscriptKeyword, instance, "manuscript", kwlist, "keyword")

            # (3) user-specific 'keywords'
            ukwlist = form.cleaned_data['ukwlist']
            profile = Profile.get_user_profile(self.request.user.username)
            adapt_m2m(UserKeyword, instance, "manu", ukwlist, "keyword", qfilter = {'profile': profile}, extrargs = {'profile': profile, 'type': 'manu'})

            # (4) 'literature'
            litlist = form.cleaned_data['litlist']
            adapt_m2m(LitrefMan, instance, "manuscript", litlist, "reference", extra=['pages'], related_is_through = True)

            # (5) 'provenances'Select a provenance...
            mprovlist = form.cleaned_data['mprovlist']
            adapt_m2m(ProvenanceMan, instance, "manuscript", mprovlist, "provenance", extra=['note'], related_is_through = True)

            # (6) Related manuscript links...
            mlinklist = form.cleaned_data['mlinklist']
            manu_added = []
            manu_deleted = []
            adapt_m2m(ManuscriptLink, instance, "src", mlinklist, "dst", 
                      extra = ['linktype', 'note'], related_is_through=True,
                      added=manu_added, deleted=manu_deleted)
            # Check for partial links in 'deleted'
            for obj in manu_deleted:
                # This if-clause is not needed: anything that needs deletion should be deleted
                # if obj.linktype in LINK_BIDIR:
                # First find and remove the other link
                reverse = ManuscriptLink.objects.filter(src=obj.dst, dst=obj.src, linktype=obj.linktype).first()
                if reverse != None:
                    reverse.delete()
                # Then remove myself
                obj.delete()
            # Make sure to add the reverse link in the bidirectionals
            for obj in manu_added:
                if obj.linktype in LINK_BIDIR_MANU:
                    # Find the reversal
                    reverse = ManuscriptLink.objects.filter(src=obj.dst, dst=obj.src, linktype=obj.linktype).first()
                    if reverse == None:
                        # Create the reversal 
                        reverse = ManuscriptLink.objects.create(src=obj.dst, dst=obj.src, linktype=obj.linktype)
                        # Other adaptations
                        bNeedSaving = False
                        # Possibly copy note
                        if obj.note != None and obj.note != "":
                          reverse.note = obj.note
                          bNeedSaving = True
                        # Need saving? Then save
                        if bNeedSaving:
                          reverse.save()


            # (6) 'projects'
            projlist = form.cleaned_data['projlist']
            manu_proj_deleted = []
            adapt_m2m(ManuscriptProject, instance, "manuscript", projlist, "project", deleted=manu_proj_deleted)
            project_dependant_delete(self.request, manu_proj_deleted)

            # When projects have been added to the manuscript, the sermons need to be updated too 
            # or vice versa
            # Issue #412: do *NOT* call this any more
            #             when the project of a manuscript changes, underlying sermons are *not* automatically affected
            # instance.adapt_projects() 

            # Issue #412 + #473: default project assignment
            if instance.projects.count() == 0:
                # Need to know who is 'talking'...
                username = self.request.user.username
                profile = Profile.get_user_profile(username)

                # The user has not selected a project (yet): try default assignment
                # user_projects = profile.projects.all()
                user_projects = profile.get_defaults()
                if user_projects.count() == 1:
                    project = user_projects.first()
                    ManuscriptProject.objects.create(manuscript=instance, project=project)
            
            # Process many-to-ONE changes

            # Issue #426: the following is OBSOLETE, since daterange is now coupled to codico
            ## (1) links from Daterange to Manuscript
            #datelist = form.cleaned_data['datelist']
            #adapt_m2o(Daterange, instance, "manuscript", datelist)

            # (2) external URLs
            extlist = form.cleaned_data['extlist']
            adapt_m2o(ManuscriptExt, instance, "manuscript", extlist)

        except:
            msg = oErr.get_error_message()
            bResult = False
        return bResult, msg

    def after_new(self, form, instance):
        """When a Manuscript has been created, it needs to get a Codico"""

        bResult = True
        msg = ""
        oErr = ErrHandle()
        try:
            bResult, msg = add_codico_to_manuscript(instance)
        except:
            msg = oErr.get_error_message()
            bResult = False
        return bResult, msg 

    def action_add(self, instance, details, actiontype):
        """User can fill this in to his/her liking"""
        passim_action_add(self, instance, details, actiontype)

    def get_history(self, instance):
        lst_key_exclude = []
        # Double check if this person is an editor or not
        if not user_is_ingroup(self.request, app_editor):
            lst_key_exclude.append("editornotes")
        return passim_get_history(instance, lst_key_exclude)


class ManuscriptDetails(ManuscriptEdit):
    rtype = "html"

    def add_to_context(self, context, instance):
        """Add to the existing context"""

        template_sermon = 'seeker/sermon_view.html'

        def sermon_object(obj ):
            # Initialisations
            html = ""
            label = ""
            # Check if this points to a sermon
            sermon = obj.itemsermons.first()
            if sermon != None:
                # Calculate the HTML for this sermon
                context = dict(msitem=sermon)
                html = treat_bom( render_to_string(template_sermon, context))
                # Determine what the label is going to be
                label = obj.locus
            else:
                # Determine what the label is going to be
                head = obj.itemheads.first()
                if head != None:
                    label = head.locus
            # Determine the parent, if any
            parent = 1 if obj.parent == None else obj.parent.order + 1
            id = obj.order + 1
            # Create a sermon representation
            oSermon = dict(id=id, parent=parent, pos=label, child=[], f = dict(order=obj.order, html=html))
            return oSermon

        # Start by executing the standard handling
        super(ManuscriptDetails, self).add_to_context(context, instance)

        oErr = ErrHandle()
        try:
            # Additional sections
            context['sections'] = []

            # Lists of related objects
            context['related_objects'] = []

            # Need to know who this user (profile) is
            username = self.request.user.username
            team_group = app_editor

            # Construct the hierarchical list
            sermon_list = instance.get_sermon_list(username, team_group)
 
            # The following only goes for the correct mtype
            if instance.mtype in ["man", "tem"]:
                # Add instances to the list, noting their childof and order
                context['sermon_list'] = sermon_list
                context['sermon_count'] = len(sermon_list)
                # List of codicological units that are not yet linked to data
                codi_empty_list = []
                for codi in instance.manuscriptcodicounits.all().order_by('order'):
                    if codi.codicoitems.count() == 0:
                        codi_empty_list.append(codi)
                context['codi_empty_list'] = codi_empty_list

                # Add the list of sermons and the comment button
                context['add_to_details'] = render_to_string("seeker/manuscript_sermons.html", context, self.request)
            elif instance.mtype == "rec":
                # THis is a reconstruction: show hierarchy view-only
                context['sermon_list'] = sermon_list
                context['sermon_count'] = len(sermon_list)

                # Note: do *NOT* give this list for reconstructions!!
                context['codi_empty_list'] = []

                # Add the list of sermons and the comment button
                context_special = copy.copy(context)
                context_special['is_app_editor'] = False
                context['add_to_details'] = render_to_string("seeker/manuscript_sermons.html", context_special, self.request)

        except:
            msg = oErr.get_error_message()
            oErr.DoError("ManuscriptDetails/add_to_context")

        # Return the context we have made
        return context

    def before_save(self, form, instance):
        return True, ""

    def process_formset(self, prefix, request, formset):
        return None

    def after_save(self, form, instance):
        if instance != None and instance.id != None:
            # If no project has been selected, then select the default project(s) - see issue #479
            count = instance.projects.count()
            if count == 0:
                # Set the default projects
                profile = Profile.get_user_profile(self.request.user.username)
                projects = profile.get_defaults()
                instance.set_projects(projects)
        return True, ""


class ManuscriptUserKeyword(ManuscriptDetails):
    """This version only looks at one kind of data: the ukwlist"""

    newRedirect = True

    def custom_init(self, instance):
        oErr = ErrHandle()
        try:
            # Make sure to set the correct redirect page
            if instance:
                self.redirectpage = reverse("manuscript_details", kwargs={'pk': instance.id})
            # Make sure we are not saving
            self.do_not_save = True

            # Get the value
            qd = self.qd
            if self.prefix_type == "simple":
                sUkwList = "{}-ukwlist".format(self.prefix)
            else:
                sUkwList = "{}-{}-ukwlist".format(self.prefix, instance.id)
            ukwlist_ids = qd.getlist(sUkwList)
            ukwlist = Keyword.objects.filter(id__in=ukwlist_ids)
            
            # Process the contents of the ukwlist, the chosen user-keywords
            profile = Profile.get_user_profile(self.request.user.username)
            adapt_m2m(UserKeyword, instance, "manu", ukwlist, "keyword", qfilter = {'profile': profile}, extrargs = {'profile': profile, 'type': 'manu'})

        except:
            msg = oErr.get_error_message()
            oErr.DoError("ManuscriptUserKeyword/custom_init")
    

class ManuscriptHierarchy(ManuscriptDetails):
    newRedirect = True

    def custom_init(self, instance):
        errHandle = ErrHandle()

        def getid(item, key, mapping):
            id_value = item[key]
            if "new" in id_value:
                id_value = mapping[id_value]
            obj = MsItem.objects.filter(id=id_value).first()
            id = None if obj == None else obj.id
            return id

        # Note: use [errHandle]
        try:
            # Make sure to set the correct redirect page
            if instance:
                self.redirectpage = reverse("manuscript_details", kwargs={'pk': instance.id})
            # Make sure we are not saving
            self.do_not_save = True
            # Get the [hlist] value
            if 'manu-hlist' in self.qd:
                # Interpret the list of information that we receive
                hlist = json.loads(self.qd['manu-hlist'])
                # Debugging:
                str_hlist = json.dumps(hlist, indent=2)

                # Step 0: make sure that the very *FIRST* element contains a 'codi' item
                if not 'codi' in hlist[0]:
                    codi_id = None
                    for item in hlist:
                        if 'codi' in item:
                            codi_id = item['codi']
                            break
                    if codi_id is None:
                        # We are in *DEEP* trouble. Not one single CODI
                        errHandle.Status("ManuscriptHierarchy: Cannot find first CODI. Manu={}".format(instance.id))
                    else:
                        hlist[0]['codi'] = codi_id


                # Step 1: Convert any new hierarchical elements into [MsItem] with SermonHead
                head_to_id = {}
                deletables = []
                with transaction.atomic():
                    for idx, item in enumerate(hlist):
                        if 'new' in item['id']:
                            # This is a new structural element - Create an MsItem
                            msitem = MsItem.objects.create(manu=instance)
                            # Create a new MsHead item
                            shead = SermonHead.objects.create(msitem=msitem,title=item['title'])
                            # Make a mapping
                            head_to_id[item['id']] = msitem.id
                            # Testing
                            id=getid(item, "id", head_to_id)
                        elif 'action' in item and item['action'] == "delete":
                            # This one must be deleted
                            deletables.append(item['id'])

                # Step 1b: adapt the list with deletables
                hlist[:] = [x for x in hlist if x.get("action") != "delete"]

                # Step 1c: delete those that need it
                MsItem.objects.filter(id__in=deletables).delete()

                # Step 2: store the hierarchy
                changes = {}
                hierarchy = []
                codi = None
                codi_order = 1
                with transaction.atomic():
                    for idx, item in enumerate(hlist):
                        bNeedSaving = False
                        # Get the msitem of this item
                        msitem = MsItem.objects.filter(id=getid(item, "id", head_to_id)).first()
                        # if this doesn't exist (anymore), just skip it
                        if not msitem is None:
                            # Get the next if any
                            next = None if item['nextid'] == "" else MsItem.objects.filter(id=getid(item, "nextid", head_to_id)).first()
                            # Get the first child
                            firstchild = None if item['firstchild'] == "" else MsItem.objects.filter(id=getid(item, "firstchild", head_to_id)).first()
                            # Get the parent
                            parent = None if item['parent'] == "" else MsItem.objects.filter(id=getid(item, "parent", head_to_id)).first()

                            # Get a possible codi id
                            codi_id = item.get("codi")
                            if codi_id != None:
                                if codi == None or codi.id != codi_id:
                                    codi = Codico.objects.filter(id=codi_id).first()
                                    # Possibly reset the codi order
                                    if codi_order != codi.order:
                                        codi.order = codi_order
                                        codi.save()
                                    codi_order += 1

                            # Safe guarding
                            if codi is None:
                                errHandle.Status("ManuscriptHierarchy: codi is none. Manu={}, msitem={}".format(instance.id, msitem.id))
                                x = msitem.itemsermons.first()

                            # Possibly set the msitem codi
                            if msitem.codico != codi:
                                msitem.codico = codi
                                bNeedSaving = True
                            elif codi == None and msitem.codico == None:
                                # This MsItem is inserted before something that may already have a codico
                                codi = instance.manuscriptcodicounits.order_by('order').first()
                                if codi != None:
                                    msitem.codico = codi
                                    bNeedSaving = True

                            # Possibly adapt the [shead] title and locus
                            itemhead = msitem.itemheads.first()
                            if itemhead and 'title' in item and 'locus' in item:
                                title= item['title'].strip()
                                locus = item['locus']
                                if itemhead.title != title or itemhead.locus != locus:
                                    itemhead.title = title.strip()
                                    itemhead.locus = locus
                                    # Save the itemhead
                                    itemhead.save()
                            
                            order = idx + 1

                            sermon_id = "none"
                            sermonlog = dict(msitem=msitem.id)
                            if msitem.itemsermons.count() > 0:
                                sermon_id = msitem.itemsermons.first().id
                                sermonlog['sermon'] =sermon_id
                            elif not itemhead is None:
                                srmhead_id = itemhead.id
                                sermonlog['srmhead']=srmhead_id
                            else:
                                sermonlog['srmwarn']="no SD or SH".format(msitem.id)
                            bAddSermonLog = False

                            # Check if anytyhing changed
                            if msitem.order != order:
                                # Implement the change
                                msitem.order = order
                                bNeedSaving =True
                            if msitem.parent is not parent:
                                # Track the change
                                old_parent_id = "none" if msitem.parent == None else msitem.parent.id
                                new_parent_id = "none" if parent == None else parent.id
                                if old_parent_id != new_parent_id:
                                    # Track the change
                                    sermonlog['parent_new'] = new_parent_id
                                    sermonlog['parent_old'] = old_parent_id
                                    bAddSermonLog = True

                                    # Implement the change
                                    msitem.parent = parent
                                    bNeedSaving = True
                                else:
                                    no_change = 1

                            if msitem.firstchild != firstchild:
                                # Implement the change
                                msitem.firstchild = firstchild
                                bNeedSaving =True
                            if msitem.next != next:
                                # Track the change
                                old_next_id = "none" if msitem.next == None else msitem.next.id
                                new_next_id = "none" if next == None else next.id
                                sermonlog['next_new'] = new_next_id
                                sermonlog['next_old'] = old_next_id
                                bAddSermonLog = True

                                # Implement the change
                                msitem.next = next
                                bNeedSaving =True
                            # Do we need to save this one?
                            if bNeedSaving:
                                msitem.save()
                                if bAddSermonLog:
                                    # Store the changes
                                    hierarchy.append(sermonlog)

                details = dict(id=instance.id, savetype="change", changes=dict(hierarchy=hierarchy))
                passim_action_add(self, instance, details, "save")

            return True
        except:
            msg = errHandle.get_error_message()
            errHandle.DoError("ManuscriptHierarchy")
            return False


class ManuscriptCodico(ManuscriptDetails):
    """Link a codico to a manuscript"""
    
    initRedirect = True

    def custom_init(self, instance):
        errHandle = ErrHandle()

        try:
            # Check if the right parameters have been passed on
            if "mrec-rmanu" in self.qd and "mrec-codicostart" in self.qd:
                manu_id = self.qd.get("mrec-rmanu")
                codico_id = self.qd.get("mrec-codicostart")
                if manu_id == None or codico_id == None:
                    # Open another place
                    self.redirectpage = reverse("manuscript_list")
                else:

                    # Check if this thing is already existing
                    obj = Reconstruction.objects.filter(manuscript=manu_id, codico=codico_id).first()
                    if obj == None:
                        # Doesn't exist (yet), so create it
                        order = Reconstruction.objects.filter(manuscript=manu_id).count() + 1
                        obj = Reconstruction.objects.create(manuscript_id=manu_id, codico_id=codico_id, order=order)

                    # Make sure to set the correct redirect page
                    self.redirectpage = reverse("manuscript_details", kwargs={'pk': manu_id})
                    # Make sure we set the object to the reconstruction manuscript
                    self.object = obj.manuscript
            elif "mrec-rcodico" in self.qd and "mrec-manuscript" in self.qd:
                manu_id = self.qd.get("mrec-manuscript")
                codico_id = self.qd.get("mrec-rcodico")
                if manu_id == None or codico_id == None:
                    # Open another place
                    self.redirectpage = reverse("manuscript_list")
                else:
                    # Check if this thing is already existing
                    obj = Reconstruction.objects.filter(manuscript=manu_id, codico=codico_id).first()
                    if obj == None:
                        # Doesn't exist (yet), so create it
                        order = Reconstruction.objects.filter(manuscript=manu_id).count() + 1
                        obj = Reconstruction.objects.create(manuscript_id=manu_id, codico_id=codico_id, order=order)

                    # Make sure to set the correct redirect page
                    self.redirectpage = reverse("manuscript_details", kwargs={'pk': manu_id})
                    # Make sure we set the object to the reconstruction manuscript
                    self.object = obj.manuscript
            elif "mrec-codicolist" in self.qd and "mrec-manuscript" in self.qd:
                manu_id = self.qd.get("mrec-manuscript")
                codico_str = self.qd.get("mrec-codicolist")
                if manu_id == None:
                    # Open another place
                    self.redirectpage = reverse("manuscript_list")
                elif codico_str == None or codico_str == "[]":
                    # Make sure to set the correct redirect page
                    self.redirectpage = reverse("manuscript_details", kwargs={'pk': manu_id})
                else:
                    # Get the actual manuscript
                    manu = Manuscript.objects.filter(id=manu_id).first()
                    # Get the list of codico id's (in their proper order)
                    codico_lst = json.loads(codico_str)
                    # Action depends on the manuscript type
                    if manu.mtype == "rec":
                        # This is a reconstructed manuscript
                        delete_lst = []
                        current_lst = Reconstruction.objects.filter(manuscript=manu_id).order_by("order")
                        for obj in current_lst:
                            if obj.codico.id not in codico_lst:
                                delete_lst.append(obj.id)
                        # Remove those that need deletion
                        if len(delete_lst) > 0:
                            Reconstruction.objects.filter(id__in=delete_lst).delete()
                        # Add and re-order
                        order = 1
                        with transaction.atomic():
                            for id in codico_lst:
                                # Check if this one is there
                                obj = Reconstruction.objects.filter(manuscript=manu_id, codico=id).first()
                                if obj == None:
                                    # Add it
                                    obj = Reconstruction.objects.create(manuscript_id=manu_id, codico_id=id)
                                obj.order = order
                                obj.save()
                                order += 1
                        # Make sure to set the correct redirect page
                        self.redirectpage = reverse("manuscript_details", kwargs={'pk': manu_id})
                    else:
                        # This is a common manuscript (or a template, but I'm not sure that should be allowed)
                        delete_lst = []
                        preceding_lst = []
                        current_lst = [x.id for x in manu.manuscriptcodicounits.all().order_by('order')]
                        prev_id = None
                        for id in current_lst:
                            if id not in codico_lst:
                                delete_lst.append(id)
                                preceding_lst.append(prev_id)
                            prev_id = id

                        # Remove the codico's that need deletion
                        if len(delete_lst) > 0:
                            # First: re-attach the MsItems in this list to the 'previous' codico
                            for idx, stale in enumerate(delete_lst):
                                # Check out if there is a Codico preceding this stale one
                                prev_id = preceding_lst[idx]
                                if not prev_id is None:
                                    prev = Codico.objects.filter(id=prev_id).first()
                                    # Attach all 'previous' MsItems to the new codico
                                    for item in MsItem.objects.filter(codico_id=stale):
                                        item.codico = prev
                                        item.save()

                            # Second: remove the stale Codico's in one go
                            Codico.objects.filter(id__in=delete_lst).delete()

                        # Double check the order of the items
                        # (1) Put the codicological units in the correct order
                        #     (the order in which they were presented by the user in hList)
                        order = 1
                        with transaction.atomic():
                            for id in codico_lst:
                                # Get the codico
                                codi = Codico.objects.filter(id=id).first()
                                # Set the correct order
                                if codi.order != order:
                                    codi.order = order
                                    codi.save()
                                # Go to the next order count
                                order += 1

                        # (2) Put the MsItem-s in the correct order: first codico-order then their own
                        order = 1
                        do_order = []
                        for msitem in MsItem.objects.filter(manu=manu).order_by('codico__order', 'order'):
                            if msitem.order != order:
                                do_order.append(dict(obj=msitem, order=order))
                            order += 1

                        with transaction.atomic():
                            for oItemOrder in do_order:
                                obj = oItemOrder['obj']
                                obj.order = oItemOrder['order']
                                obj.save()

                        # Make sure to set the correct redirect page
                        self.redirectpage = reverse("manuscript_details", kwargs={'pk': manu_id})
                    
                    # FOr debugging purposes
                    x = manu.manuscriptcodicounits.all()
            # Return positively
            return True
        except:
            msg = errHandle.get_error_message()
            errHandle.DoError("ManuscriptCodico")
            return False


class ManuscriptListView(BasicList):
    """Search and list manuscripts"""
    
    model = Manuscript
    listform = SearchManuForm
    has_select2 = True
    use_team_group = True
    paginate_by = 20
    bUseFilter = True
    profile = None
    sel_button = "manu"
    prefix = "manu"
    sg_name = "Manuscript"     # This is the name as it appears e.g. in "Add a new XXX" (in the basic listview)
    plural_name = "Manuscripts"
    basketview = False
    
    order_cols = ['library__lcity__name;library__location__name', 'library__name', 'idno;name', '', '', 'yearstart','yearfinish', 'stype','']
    order_default = order_cols
    order_heads = [
        {'name': 'City/Location',    'order': 'o=1', 'type': 'str', 'custom': 'city',
                    'title': 'City or other location, such as monastery'},
        {'name': 'Library',  'order': 'o=2', 'type': 'str', 'custom': 'library'},
        {'name': 'Name',     'order': 'o=3', 'type': 'str', 'custom': 'name', 'main': True, 'linkdetails': True},
        {'name': '',         'order': '',    'type': 'str', 'custom': 'saved',   'align': 'right'},
        {'name': 'Items',    'order': '',    'type': 'int', 'custom': 'count',   'align': 'right'},
        {'name': 'From',     'order': 'o=6', 'type': 'int', 'custom': 'from',    'align': 'right'},
        {'name': 'Until',    'order': 'o=7', 'type': 'int', 'custom': 'until',   'align': 'right'},
        {'name': 'Status',   'order': 'o=8', 'type': 'str', 'custom': 'status'},
        {'name': '',         'order': '',    'type': 'str', 'custom': 'links'}]
    filters = [         
        {"name": "Country",         "id": "filter_country",          "enabled": False}, 
        {"name": "City/Location",   "id": "filter_city",             "enabled": False},
        {"name": "Library",         "id": "filter_library",          "enabled": False}, 
        {"name": "Shelfmark",       "id": "filter_manuid",           "enabled": False},
        {"name": "Origin",          "id": "filter_origin",           "enabled": False},        
        {"name": "Provenance",      "id": "filter_provenance",       "enabled": False},
        {"name": "Date range",      "id": "filter_daterange",        "enabled": False},
        {"name": "Title",           "id": "filter_title",            "enabled": False},
        {"name": "Keyword",         "id": "filter_keyword",          "enabled": False},        
        {"name": "Manuscript type", "id": "filter_manutype",         "enabled": False},        
        {"name": "Status",          "id": "filter_stype",            "enabled": False},
        {"name": "Project",         "id": "filter_project",          "enabled": False},
        
        # Issue #717: rename into "Comparative search..."
        # {"name": "Collection/Dataset...",  "id": "filter_collection",     "enabled": False, "head_id": "none"},
        {"name": "Comparative search...",  "id": "filter_collection",     "enabled": False, "head_id": "none"},
        {"name": "Manifestation...",       "id": "filter_sermon",         "enabled": False, "head_id": "none"},       
        {"name": "Authority file...",      "id": "filter_authority_file", "enabled": False, "head_id": "none"},        

        # issue #718: remove options to search for Gryson/Clavis/other code - for sermons
        # {"name": "Gryson/Clavis/Other code: manual",    "id": "filter_signature_m",     "enabled": False, "head_id": "filter_sermon"},
        # {"name": "Gryson/Clavis/Other code: automatic", "id": "filter_signature_a",     "enabled": False, "head_id": "filter_sermon"},

        {"name": "Attr. author",            "id": "filter_sermo_author",        "enabled": False, "head_id": "filter_sermon"},
        {"name": "Author type",             "id": "filter_sermo_authortype",    "enabled": False, "head_id": "filter_sermon"},
        {"name": "Incipit",                 "id": "filter_sermo_incipit",       "enabled": False, "head_id": "filter_sermon"},
        {"name": "Explicit",                "id": "filter_sermo_explicit",      "enabled": False, "head_id": "filter_sermon"},
        {"name": "Bible reference",         "id": "filter_sermo_bibref",        "enabled": False, "head_id": "filter_sermon"},
        {"name": "Feast",                   "id": "filter_sermo_feast",         "enabled": False, "head_id": "filter_sermon"},
               
        {"name": "Manuscript comparison",   "id": "filter_collection_manuidno", "enabled": False, "include_id": "filter_collection_hcptc", "head_id": "filter_collection"},
        {"name": "Historical Collection",   "id": "filter_collection_hc",       "enabled": False, "include_id": "filter_collection_hcptc", "head_id": "filter_collection"},
        {"name": "PD: Authority file",      "id": "filter_collection_super",    "enabled": False, "include_id": "filter_collection_hcptc", "head_id": "filter_collection"},
        {"name": "HC/Manu overlap",         "id": "filter_collection_hcptc",    "enabled": False, "head_id": "filter_collection", "hide_button": True},
        # issue #717: delete the PD:Manuscript and PD:Sermon options
        #{"name": "PD: Manuscript",          "id": "filter_collection_manu",     "enabled": False, "head_id": "filter_collection"},
        #{"name": "PD: Sermon",              "id": "filter_collection_sermo",    "enabled": False, "head_id": "filter_collection"},
        # Issue #416: Delete the option to search for a GoldSermon dataset 
        # {"name": "PD: Sermon Gold",         "idco": "filter_collection_gold", "enabled": False, "head_id": "filter_collection"},
        
        {"name": "Author",                  "id": "filter_authority_file_author",       "enabled": False, "head_id": "filter_authority_file"},
        {"name": "Incipit",                 "id": "filter_authority_file_incipit",      "enabled": False, "head_id": "filter_authority_file"},
        {"name": "Explicit",                "id": "filter_authority_file_explicit",     "enabled": False, "head_id": "filter_authority_file"},
        # issue #718: add *ONE* option to search for Gryson/Clavis/other code - for AFs
        {"name": "Gryson/Clavis/Other code","id": "filter_authority_file_signature",    "enabled": False, "head_id": "filter_authority_file"},
    ]

    searches = [
        {'section': '', 'filterlist': [
            {'filter': 'manuid',        'dbfield': 'idno',                   'keyS': 'idno',          'keyList': 'manuidlist', 'infield': 'id'},
            {'filter': 'country',       'fkfield': 'library__lcountry',      'keyS': 'country_ta',    'keyId': 'country',     'keyFk': "name"},
            {'filter': 'city',          'fkfield': 'library__lcity|library__location',         
                                                                             'keyS': 'cityloc_ta',    'keyId': 'city',        'keyFk': "name"},
            {'filter': 'library',       'fkfield': 'library',                'keyS': 'libname_ta',    'keyId': 'library',     'keyFk': "name"},
            {'filter': 'provenance',    'fkfield': 'manuscriptcodicounits__provenances__location|manuscriptcodicounits__provenances',  
             'keyS': 'prov_ta',       'keyId': 'prov',        'keyFk': "name"},  
            {'filter': 'title',         'fkfield': 'manuscriptcodicounits', 'keyS': 'title_ta',               'keyFk': 'name'},            
            {'filter': 'origin',        'fkfield': 'manuscriptcodicounits__origins__location|manuscriptcodicounits__origins', 
             # Issue #427. This was: 'manuscriptcodicounits__origin',                 
             'keyS': 'origin_ta',     'keyId': 'origin',      'keyFk': "name"},
            {'filter': 'keyword',       'fkfield': 'keywords',               'keyFk': 'name', 'keyList': 'kwlist', 'infield': 'name' },
            {'filter': 'project',       'fkfield': 'projects',               'keyFk': 'name', 'keyList': 'projlist', 'infield': 'name'},
            {'filter': 'daterange',     'dbfield': 'manuscriptcodicounits__codico_dateranges__yearstart__gte', 'keyS': 'date_from'},
            {'filter': 'daterange',     'dbfield': 'manuscriptcodicounits__codico_dateranges__yearfinish__lte', 'keyS': 'date_until'},            
            {'filter': 'manutype',      'dbfield': 'mtype',                  'keyS': 'manutype', 'keyType': 'fieldchoice', 'infield': 'abbr'},
            {'filter': 'stype',         'dbfield': 'stype',                  'keyList': 'stypelist', 'keyType': 'fieldchoice', 'infield': 'abbr'}
            ]},

        {'section': 'collection', 'filterlist': [
            # === Overlap with a specific manuscript ===
            {'filter': 'collection_manuidno',  'keyS': 'cmpmanu', 'dbfield': 'dbcmpmanu', 'keyList': 'cmpmanuidlist', 'infield': 'id', 'help': 'overlap_manu'},            
            {'filter': 'collection_hcptc', 'keyS': 'overlap', 'dbfield': 'hcptc', 
             'title': 'Percentage overlap between the Historical Collection SSGs and the SSGs referred to in the manuscripts'},             
            #{'filter': 'collection_manuptc', 'keyS': 'overlap', 'dbfield': 'hcptc',
            # 'title': 'Percentage overlap between the "Comparison manuscript" SSGs and the SSGs referred to in other manuscripts'},

            # === Historical Collection ===
            {'filter': 'collection_hc',  'fkfield': 'manuitems__itemsermons__equalgolds__collections', 'help': 'overlap_hc',                         
             'keyS': 'collection',    'keyFk': 'name', 'keyList': 'collist_hist', 'infield': 'name' },             

            # === Personal Dataset ===
            {'filter': 'collection_manu',  'fkfield': 'collections',                            
             'keyS': 'collection',    'keyFk': 'name', 'keyList': 'collist_m', 'infield': 'name' },
            {'filter': 'collection_sermo', 'fkfield': 'manuitems__itemsermons__collections',               
             'keyS': 'collection_s',  'keyFk': 'name', 'keyList': 'collist_s', 'infield': 'name' },
            # Issue #416: Delete the option to search for a GoldSermon dataset 
            #{'filter': 'collection_gold',  'fkfield': 'manuitems__itemsermons__goldsermons__collections',  
            # 'keyS': 'collection_sg', 'keyFk': 'name', 'keyList': 'collist_sg', 'infield': 'name' },
            {'filter': 'collection_super', 'fkfield': 'manuitems__itemsermons__equalgolds__collections', 'help': 'overlap_af',
             'keyS': 'collection_ssg','keyFk': 'name', 'keyList': 'collist_ssg', 'infield': 'name' },
            # ===================
            ]},
        {'section': 'sermon', 'name': 'manifestation', 'filterlist': [
            # issue #717: delete the PD:Manuscript and PD:Sermon options
            #{'filter': 'signature_m', 'fkfield': 'manuitems__itemsermons__sermonsignatures',     'help': 'signature',
            # 'keyS': 'signature', 'keyFk': 'code', 'keyId': 'signatureid', 'keyList': 'siglist', 'infield': 'code' },
            #{'filter': 'signature_a', 'fkfield': 'manuitems__itemsermons__equalgolds__equal_goldsermons__goldsignatures',     'help': 'signature',
            # 'keyS': 'signaturea', 'keyFk': 'code', 'keyId': 'signatureaid', 'keyList': 'siglist_a', 'infield': 'code' }, # KAN WEG
            {'filter': 'sermo_authortype', 'keyS': 'authortype',  'help': 'authorhelp'}, 
            {'filter': 'sermo_author',        'fkfield': 'manuitems__itemsermons__author',            
                     'keyS': 'sermo_authorname', 'keyFk': 'name', 'keyList': 'sermo_authorlist', 'infield': 'id', 'external': 'sermo-authorname'},            
            {'filter': 'sermo_incipit',       'dbfield': 'manuitems__itemsermons__incipit',   'keyS': 'sermo_incipit'},
            {'filter': 'sermo_explicit',      'dbfield': 'manuitems__itemsermons__explicit',  'keyS': 'sermo_explicit'},
            {'filter': 'sermo_feast',         'fkfield': 'manuitems__itemsermons__feast', 'keyFk': 'name', 'keyList': 'sermo_feastlist', 'infield': 'id'},            
            {'filter': 'sermo_bibref',    'dbfield': '$dummy', 'keyS': 'bibrefbk'}, #toevoegen sermo_?
            {'filter': 'sermo_bibref',    'dbfield': '$dummy', 'keyS': 'bibrefchvs'}
            ]},        
        {'section': 'authority_file', 'name': 'Authority File', 'filterlist': [
            {'filter': 'authority_file_code', 'fkfield': 'manuitems__itemsermons__sermondescr_super__super', 'help': 'passimcode',
             'keyS': 'passimcode', 'keyFk': 'code', 'keyList': 'passimlist', 'infield': 'id'},
            
            # Signature NEW signature
            {'filter': 'authority_file_signature', 'fkfield': 'manuitems__itemsermons__sermonsignatures',     'help': 'signature', # HOE MOET DIE fkfield worden genoemd? Filteren op twee links
             'keyS': 'signature', 'keyFk': 'code', 'keyId': 'signatureid', 'keyList': 'siglist', 'infield': 'code' }, # Q expressie toevoegen

            ## Sign manual SERMON KAN LATER WEG
            #{'filter': 'authority_file_signature_m', 'fkfield': 'manuitems__itemsermons__sermonsignatures',     'help': 'signature',
            # 'keyS': 'signature', 'keyFk': 'code', 'keyId': 'signatureid', 'keyList': 'siglist', 'infield': 'code' },
            
            ## Sign auto SSG KAN LATER WEG
            #{'filter': 'authority_file_signature_a', 'fkfield': 'manuitems__itemsermons__equalgolds__equal_goldsermons__goldsignatures',     'help': 'signature',
            # 'keyS': 'signaturea', 'keyFk': 'code', 'keyId': 'signatureaid', 'keyList': 'siglist_a', 'infield': 'code' },                   
                         
            {'filter': 'authority_file_author', 'fkfield': 'manuitems__itemsermons__sermondescr_super__super__author',            
                     'keyS': 'ssg_authorname', 'keyFk': 'name', 'keyList': 'ssg_authorlist', 'infield': 'id', 'external': 'gold-authorname' },
            {'filter': 'authority_file_incipit',   'dbfield': 'manuitems__itemsermons__sermondescr_super__super__srchincipit',   'keyS': 'ssg_incipit'},             
            {'filter': 'authority_file_explicit',  'dbfield': 'manuitems__itemsermons__sermondescr_super__super__srchexplicit',  'keyS': 'ssg_explicit'}, 

            ]},
        {'section': 'other', 'filterlist': [
            #{'filter': 'other_project',   'fkfield': 'project',  'keyS': 'project', 'keyFk': 'id', 'keyList': 'prjlist', 'infield': 'name' },
            {'filter': 'source',    'fkfield': 'source',   'keyS': 'source',  'keyFk': 'id', 'keyList': 'srclist', 'infield': 'id' },
            {'filter': 'atype',     'dbfield': 'manuitems__itemsermons__sermondescr_super__super__atype',    'keyS': 'atype'},
            {'filter': 'mtype', 'dbfield': 'mtype', 'keyS': 'mtype'}
            ]}
         ]
    uploads = reader_uploads
    downloads = [
        {"label": "Ead:Excel", "dtype": "xlsx", "url": 'ead_results'},
        {"label": "Ead:csv (tab-separated)", "dtype": "csv", "url": 'ead_results'},
        {"label": None},
        {"label": "Ead:json", "dtype": "json", "url": 'ead_results'},
        {"label": None},
        {"label": "Huwa Manuscripts: json", "dtype": "json", "url": 'manuscript_huwajson'},
        #{"label": "Huwa Manuscripts: json",  "dtype": "json_manu", "url": 'equalgold_huwajson'},                 
                 ]
    custombuttons = [{"name": "search_ecodex", "title": "Convert e-codices search results into a list", 
                      "icon": "music", "template_name": "seeker/search_ecodices.html" }]

    selectbuttons = [
        {'title': 'Add to saved items', 'mode': 'add_saveitem', 'button': 'jumbo-1', 'glyphicon': 'glyphicon-star-empty'},
        {'title': 'Add to basket',      'mode': 'add_basket',   'button': 'jumbo-1', 'glyphicon': 'glyphicon-shopping-cart'},
        {'title': 'Add to DCT',         'mode': 'show_dct',     'button': 'jumbo-1', 'glyphicon': 'glyphicon-wrench'},
        ]

    def initializations(self):
        # Possibly add to 'uploads'
        bHasExcel = False
        bHasGalway = False
        bHasJson = False
        for item in self.uploads:
            if item['title'] == "excel":
                bHasExcel = True
            elif item['title'] == "galway":
                bHasGalway = True
            elif item['title'] == "json":
                bHasJson = True

        # Should galway be added?
        if not bHasGalway:
            # Add a reference to the Excel upload method
            html = []
            html.append("Import manuscripts from Galway using one or more CSV files.")
            html.append("<b>Note 1:</b> this OVERWRITES a manuscript/sermon if it exists!")
            html.append("<b>Note 2:</b> default PROJECT assignment according to MyPassim!")
            msg = "<br />".join(html)
            oGalway = dict(title="galway", label="Galway",
                          url=reverse('manuscript_upload_galway'),
                          type="multiple",msg=msg)
            self.uploads.append(oGalway)

        # Should excel be added?
        if not bHasExcel:
            # Add a reference to the Excel upload method
            html = []
            html.append("Import manuscripts from one or more Excel files.")
            html.append("<b>Note 1:</b> this OVERWRITES a manuscript/sermon if it exists!")
            html.append("<b>Note 2:</b> default PROJECT assignment according to MyPassim!")
            msg = "<br />".join(html)
            oExcel = dict(title="excel", label="Excel",
                          url=reverse('manuscript_upload_excel'),
                          type="multiple", msg=msg)
            self.uploads.append(oExcel)

        # Should json be added?
        if not bHasJson:
            # Add a reference to the Json upload method
            html = []
            html.append("Import manuscripts from one or more JSON files.")
            html.append("<b>Note 1:</b> this OVERWRITES a manuscript/sermon if it exists!")
            html.append("<b>Note 2:</b> default PROJECT assignment according to MyPassim!")
            msg = "<br />".join(html)
            oJson = dict(title="json", label="Json",
                          url=reverse('manuscript_upload_json'),
                          type="multiple", msg=msg)
            self.uploads.append(oJson)

        # Possibly *NOT* show the downloads
        if not user_is_ingroup(self.request, app_developer):
            self.downloads = []
        if not user_is_authenticated(self.request) or not (user_is_superuser(self.request) or user_is_ingroup(self.request, app_moderator)):
            # Do *not* unnecessarily show the custombuttons
            self.custombuttons = []

        # ======== One-time adaptations ==============
        listview_adaptations("manuscript_list")

        self.profile = Profile.get_user_profile(self.request.user.username)

        return None

    def add_to_context(self, context, initial):
        # Add a files upload form
        context['uploadform'] = UploadFilesForm()

        # Add a form to enter a URL
        context['searchurlform'] = SearchUrlForm()
        
        # Find out who the user is
        profile = Profile.get_user_profile(self.request.user.username)
        context['basketsize'] = 0 if profile == None else profile.basketsize_manu
        context['basket_show'] = reverse('basket_show_manu')
        context['basket_update'] = reverse('basket_update_manu')

        context['colltype'] = "manu"

        # Count the number of selected items
        iCount = SelectItem.get_selectcount(profile, "manu")
        if iCount > 0:
            context['sel_count'] = str(iCount)

        return context

    def get_basketqueryset(self):
        if self.basketview:
            profile = Profile.get_user_profile(self.request.user.username)
            qs = profile.basketitems_manu.all()
        else:
            qs = Manuscript.objects.all()
        return qs
    
    def get_field_value(self, instance, custom):
        sBack = ""
        sTitle = ""
        html = []
        if custom == "city":
            if not instance.library is None:
                city = None
                if instance.library.lcity:
                    city = instance.library.lcity.name
                elif instance.library.location:
                    city = instance.library.location.name
                if city == None:
                    html.append("??")
                    sTitle = "City or location unclear"
                else:
                    html.append("<span>{}</span>".format(city[:12]))        
                    sTitle = city
            elif not instance.lcity is None:
                city = instance.lcity.name
                html.append("<span>{}</span>".format(city[:12]))        
                sTitle = city

        elif custom == "library":
            if not instance.library is None:
                lib = instance.library.name
                html.append("<span>{}</span>".format(lib[:12]))  
                sTitle = lib      
        elif custom == "name":
            html.append("<span class='manuscript-idno'>{}</span>".format(instance.idno))
            # THe name should come from the codico unit!!!
            codico = Codico.objects.filter(manuscript=instance).first()
            if codico != None and codico.name != None:
                html.append("<span class='manuscript-title'>| {}</span>".format(codico.name[:100]))
                sTitle = codico.name
        elif custom == "saved":
            # Prepare saveditem handling
            saveditem_button = get_saveditem_html(self.request, instance, self.profile, sitemtype="manu")
            saveditem_form = get_saveditem_html(self.request, instance, self.profile, "form", sitemtype="manu")
            html.append(saveditem_button)
            html.append(saveditem_form)
        elif custom == "count":
            # html.append("{}".format(instance.manusermons.count()))
            html.append("{}".format(instance.get_sermon_count()))
        elif custom == "from":
            # for item in instance.manuscript_dateranges.all():
            # Walk all codico's
            for item in Daterange.objects.filter(codico__manuscript=instance).order_by('yearstart'):
                html.append("<div>{}</div>".format(item.yearstart))
        elif custom == "until":
            # for item in instance.manuscript_dateranges.all():
            for item in Daterange.objects.filter(codico__manuscript=instance).order_by('yearfinish'):
                html.append("<div>{}</div>".format(item.yearfinish))
        elif custom == "status":
            # html.append("<span class='badge'>{}</span>".format(instance.stype[:1]))
            html.append(instance.get_stype_light())
            sTitle = instance.get_stype_display()
        elif custom == "links":
            sLinks = ""
            if instance.url:
                sLinks = "<a role='button' class='btn btn-xs jumbo-1' href='{}'><span class='glyphicon glyphicon-link'><span></a>".format(instance.url)
                sTitle = "External link"
            html.append(sLinks)
        # Combine the HTML code
        sBack = "\n".join(html)
        return sBack, sTitle

    def adapt_search(self, fields):                      

        def get_overlap_ptc(base_ssgs, comp_ssgs):
            """Calculate the overlap percentage between base and comp"""

            total = len(base_ssgs)
            count = 0
            for ssg_id in comp_ssgs:
                if ssg_id in base_ssgs: count += 1
            result = 100 * count / total
            return result

        # Adapt the search to the keywords that *may* be shown
        lstExclude=None
        qAlternative = None

        #prjlist = None # old
        projlist = None

        oErr = ErrHandle()
        try:

            # Check if a list of keywords is given
            if 'kwlist' in fields and fields['kwlist'] != None and len(fields['kwlist']) > 0:
                # Get the list
                kwlist = fields['kwlist']
                # Get the user
                username = self.request.user.username
                user = User.objects.filter(username=username).first()
                # Check on what kind of user I am
                if not user_is_ingroup(self.request, app_editor):
                    # Since I am not an app-editor, I may not filter on keywords that have visibility 'edi'
                    kwlist = Keyword.objects.filter(id__in=kwlist).exclude(Q(visibility="edi")).values('id')
                    fields['kwlist'] = kwlist
       
            # Check if a list of projects is given
            if 'projlist' in fields and fields['projlist'] != None and len(fields['projlist']) > 0:
                # Get the list
                projlist = fields['projlist']

            # Check if an overlap percentage is specified
            if 'overlap' in fields and fields['overlap'] != None:
                # Get the overlap
                overlap = fields.get('overlap', "0")
                # Use an overt truth 
                fields['overlap'] = Q(mtype="man")
                if 'collist_hist' in fields and fields['collist_hist'] != None:
                    coll_list = fields['collist_hist']
                    if len(coll_list) > 0:
                        # Yes, overlap specified
                        if isinstance(overlap, int):
                            # We also need to have the profile
                            profile = Profile.get_user_profile(self.request.user.username)

                            # Make sure the string is interpreted as an integer
                            overlap = int(overlap)
                            # Now add a Q expression that includes:
                            # 1 - Greater than or equal the overlap percentage
                            # 2 - The list of historical collections
                            # 3 - This particular user
                            fields['overlap'] = Q(manu_colloverlaps__overlap__gte=overlap) & \
                               Q(manu_colloverlaps__collection__in=coll_list) & \
                               Q(manu_colloverlaps__profile=profile)

                            # Make sure to actually *calculate* the overlap between the different collections and manuscripts
                
                            # (1) Possible manuscripts only filter on: mtype=man, prjlist
                            lstQ = []
                            # if prjlist != None: lstQ.append(Q(project__in=prjlist))
                            lstQ.append(Q(mtype="man"))
                            lstQ.append(Q(manuitems__itemsermons__equalgolds__collections__in=coll_list))
                            manu_list = Manuscript.objects.filter(*lstQ)

                            # Now calculate the overlap for all
                            with transaction.atomic():
                                for coll in coll_list:
                                    for manu in manu_list:
                                        ptc = CollOverlap.get_overlap(profile, coll, manu)
                    if 'cmpmanuidlist' in fields and fields['cmpmanuidlist'] != None:
                        # The base manuscripts with which the comparison goes
                        base_manu_list = fields['cmpmanuidlist']
                        if len(base_manu_list) > 0:
                            # Yes, overlap specified
                            if isinstance(overlap, int):
                                # Make sure the string is interpreted as an integer
                                overlap = int(overlap)
                                # Now add a Q expression
                                # fields['overlap'] = Q(manu_colloverlaps__overlap__gte=overlap)
                                # Make sure to actually *calculate* the overlap between the different collections and manuscripts

                                # (1) Get a list of SSGs associated with these manuscripts
                                base_ssg_list = EqualGold.objects.filter(sermondescr_super__sermon__msitem__manu__in=base_manu_list).values('id')
                                base_ssg_list = [x['id'] for x in base_ssg_list]
                                base_count = len(base_ssg_list)
                
                                # (2) Possible overlapping manuscripts only filter on: mtype=man, prjlist and the SSG list
                                lstQ = []
                                # if prjlist != None: lstQ.append(Q(project__in=prjlist))
                                lstQ.append(Q(mtype="man"))
                                lstQ.append(Q(manuitems__itemsermons__equalgolds__id__in=base_ssg_list))
                                manu_list = Manuscript.objects.filter(*lstQ)

                                # We also need to have the profile
                                profile = Profile.get_user_profile(self.request.user.username)
                                # Now calculate the overlap for all
                                manu_include = []
                                with transaction.atomic():
                                    for manu in manu_list:
                                        # Get a list of SSG id's associated with this particular manuscript
                                        manu_ssg_list = [x['id'] for x in EqualGold.objects.filter(sermondescr_super__sermon__msitem__manu__id=manu.id).values('id')]
                                        if get_overlap_ptc(base_ssg_list, manu_ssg_list) >= overlap:
                                            # Add this manuscript to the list 
                                            if not manu.id in manu_include:
                                                manu_include.append(manu.id)
                                fields['cmpmanuidlist'] = None
                                fields['cmpmanu'] = Q(id__in=manu_include)

            # Process the date range stuff
            date_from = fields.get("date_from")
            date_until = fields.get("date_until")
            if not date_from is None and not date_until is None:
                # Both dates are specified: include looking for overlap
                qThis_a = Q(manuscriptcodicounits__codico_dateranges__yearstart__gte=date_from)
                qThis_b = Q(manuscriptcodicounits__codico_dateranges__yearstart__lte=date_until)
                qThis_c = Q(manuscriptcodicounits__codico_dateranges__yearfinish__lte=date_until)
                qThis_d = Q(manuscriptcodicounits__codico_dateranges__yearfinish__gte=date_from)
                qThis_1 =  ( qThis_a &  ( qThis_b |  qThis_c ) )
                qThis_2 =  ( qThis_c &  ( qThis_d |  qThis_a ) )
                qThis = ( qThis_1 | qThis_2 )
                fields['date_from'] = qThis
                fields['date_until'] = ""

            # Issue #718: combine search for Signature *automatic* and *manual*
            af_siglist = fields.get("siglist")
            if not af_siglist is None and af_siglist.count() > 0:
                af_siglist_list = [dict(code=x.code, editype=x.editype) for x in af_siglist]
                lst_code = [x['code'] for x in af_siglist_list]
                qs_sermonsig = SermonSignature.objects.filter(code__in=lst_code).distinct()
                exclude_id = []
                for oItem in af_siglist_list:
                    # Check if the combination of code and editype is in here
                    res_list = qs_sermonsig.filter(code=oItem['code'])
                    if res_list.count() > 1:
                        # Check which to exclude
                        for obj in res_list:
                            if obj.editype != oItem['editype']:
                                # This one must be excluded
                                exclude_id.append(obj.id)
                # Adapt the code list if needed
                if len(exclude_id) > 0:
                    qs_sermonsig = qs_sermonsig.exclude(id__in=exclude_id)


                # Build a Q-expression to facilitate this
                q1 = Q(manuitems__itemsermons__sermonsignatures__in=qs_sermonsig)
                q2 = Q(manuitems__itemsermons__equalgolds__equal_goldsermons__goldsignatures__in=af_siglist)
                fields['siglist'] = ( q1 | q2 )

                # Debugging: 
                # qs = Manuscript.objects.filter(q1)

            # Adapt the bible reference list
            bibrefbk = fields.get("bibrefbk", "")
            if bibrefbk != None and bibrefbk != "":
                bibrefchvs = fields.get("bibrefchvs", "")

                # Get the start and end of this bibref
                start, einde = Reference.get_startend(bibrefchvs, book=bibrefbk)

                # Find out which manuscripts have sermons having references in this range
                lstQ = []
                lstQ.append(Q(manuitems__itemsermons__sermonbibranges__bibrangeverses__bkchvs__gte=start))
                lstQ.append(Q(manuitems__itemsermons__sermonbibranges__bibrangeverses__bkchvs__lte=einde))
                manulist = [x.id for x in Manuscript.objects.filter(*lstQ).order_by('id').distinct()]

                fields['bibrefbk'] = Q(id__in=manulist)

            # Make sure we only show manifestations
            # fields['mtype'] = 'man'
            # Make sure we show MANUSCRIPTS (identifiers) as well as reconstructions

            # Make sure we only use the Authority Files with accepted modifications
            # This means that atype should be 'acc' (and not: 'mod', 'rej' or 'def')        
            # With this condition we make sure ALL manuscripts are in de unfiltered listview
            print (fields['passimcode'])
            if fields['passimcode'] != '':
                fields['atype'] = 'acc'
       
            lstExclude = [ Q(mtype='tem') ]
        
            # Adapt the search for empty authors
            # Hoe komen we aan dat authortype? Waar komt die bij sermon list view vandaan? EK vragen
            if 'authortype' in fields:
                authortype = fields['authortype']
                if authortype == "non":
                    # lstExclude = []
                    lstExclude.append(Q(manuitems__itemsermons__author__isnull=False))
                elif authortype == "spe":
                    # lstExclude = []
                    lstExclude.append(Q(manuitems__itemsermons__author__isnull=True))
                else:
                    # Reset the authortype
                    fields['authortype'] = ""   
        except:
            msg = oErr.get_error_message()
            oErr.DoError("ManuscriptListView/adapt_search")

        return fields, lstExclude, qAlternative

    def view_queryset(self, qs):
        search_id = [x['id'] for x in qs.values('id')]
        profile = Profile.get_user_profile(self.request.user.username)
        profile.search_manu = json.dumps(search_id)
        profile.save()
        return None

    def get_helptext(self, name):
        """Use the get_helptext function defined in models.py"""
        return get_helptext(name)

    def get_selectitem_info(self, instance, context):
        """Use the get_selectitem_info() defined earlier in this views.py"""
        return get_selectitem_info(self.request, instance, self.profile, self.sel_button, context)
  

class ManuscriptDownload(BasicPart):
    MainModel = Manuscript
    template_name = "seeker/download_status.html"
    action = "download"
    dtype = "excel"       # downloadtype

    def custom_init(self):
        """Calculate stuff"""
        
        dt = self.qd.get('downloadtype', "")
        if dt != None and dt != '':
            self.dtype = dt

    def get_func(self, instance, path, profile, username, team_group):
        sBack = ""
        if path == "dateranges":
            # qs = instance.manuscript_dateranges.all().order_by('yearstart')
            qs = Daterange.objects.filter(codico__manuscript=instance).order_by('yearstart')
            dates = []
            for obj in qs:
                dates.append(obj.__str__())
            sBack = ", ".join(dates)
        elif path == "keywords":
            sBack = instance.get_keywords_markdown(plain=True)
        elif path == "keywordsU":
            sBack =  instance.get_keywords_user_markdown(profile, plain=True)
        elif path == "datasets":
            sBack = instance.get_collections_markdown(username, team_group, settype="pd", plain=True)
        elif path == "literature":
            sBack = instance.get_litrefs_markdown(plain=True)
        elif path == "origin":
            sBack = instance.get_origin()
        elif path == "provenances":
            sBack = instance.get_provenance_markdown(plain=True)
        elif path == "external":
            sBack = instance.get_external_markdown(plain=True)
        elif path == "brefs":
            sBack = instance.get_bibleref(plain=True)
        elif path == "signaturesM":
            sBack = instance.get_sermonsignatures_markdown(plain=True)
        elif path == "signaturesA":
            sBack = instance.get_eqsetsignatures_markdown(plain=True)
        elif path == "ssglinks":
            sBack = instance.get_eqset()
        return sBack

    def get_data(self, prefix, dtype, response=None):
        """Gather the data as CSV, including a header line and comma-separated"""

        # Initialize
        lData = []
        sData = ""
        manu_fields = []
        oErr = ErrHandle()

        try:
            # Need to know who this user (profile) is
            profile = Profile.get_user_profile(self.request.user.username)
            username = profile.user.username
            team_group = app_editor

            # Make sure we only look at lower-case Dtype
            dtype = dtype.lower()

            # Is this Excel?
            if dtype == "excel" or dtype == "xlsx":
                # Start workbook
                wb = openpyxl.Workbook()

                # First worksheet: MANUSCRIPT itself
                ws = wb.get_active_sheet()
                ws.title="Manuscript"

                # Read the header cells and make a header row in the MANUSCRIPT worksheet
                headers = ["Field", "Value"]
                for col_num in range(len(headers)):
                    c = ws.cell(row=1, column=col_num+1)
                    c.value = headers[col_num]
                    c.font = openpyxl.styles.Font(bold=True)
                    # Set width to a fixed size
                    ws.column_dimensions[get_column_letter(col_num+1)].width = 5.0        

                # Walk the mainitems
                row_num = 2
                kwargs = {'profile': profile, 'username': username, 'team_group': team_group}
                for item in Manuscript.specification:
                    key, value = self.obj.custom_getkv(item, kwargs=kwargs)
                    # Add the K/V row
                    ws.cell(row=row_num, column = 1).value = key
                    ws.cell(row=row_num, column = 2).value = value
                    row_num += 1

                # Second worksheet: ALL CODICOLOGICAL units in this manuscript
                ws = wb.create_sheet("Codicos")

                # Read the header cells and make a header row in the CODICO worksheet
                headers = [x['name'] for x in Codico.specification ]
                for col_num in range(len(headers)):
                    c = ws.cell(row=1, column=col_num+1)
                    c.value = headers[col_num]
                    c.font = openpyxl.styles.Font(bold=True)
                    # Set width to a fixed size
                    ws.column_dimensions[get_column_letter(col_num+1)].width = 5.0        

                row_num = 1
                # Walk all codico's of this manuscript
                for codico in self.obj.manuscriptcodicounits.all().order_by('order'):
                    row_num += 1
                    col_num = 1
                    # Show the order number of this Codico
                    # ws.cell(row=row_num, column=col_num).value = msitem.order

                    #for item in Codico.specification:
                    #    key, value = codico.custom_getkv(item, kwargs=kwargs)
                    #    # Add the K/V row
                    #    ws.cell(row=row_num, column = 1).value = key
                    #    ws.cell(row=row_num, column = 2).value = value
                    #    row_num += 1
                    # Walk the items
                    for item in Codico.specification:
                        if item['type'] != "":
                            key, value = codico.custom_getkv(item, kwargs=kwargs)
                            ws.cell(row=row_num, column=col_num).value = value
                            col_num += 1


                # Third worksheet: ALL SERMONS in the manuscript
                ws = wb.create_sheet("Sermons")

                # Read the header cells and make a header row in the SERMON worksheet
                headers = [x['name'] for x in SermonDescr.specification ]
                for col_num in range(len(headers)):
                    c = ws.cell(row=1, column=col_num+1)
                    c.value = headers[col_num]
                    c.font = openpyxl.styles.Font(bold=True)
                    # Set width to a fixed size
                    ws.column_dimensions[get_column_letter(col_num+1)].width = 5.0        

                row_num = 1
                # Walk all msitems of this manuscript
                for msitem in self.obj.manuitems.all().order_by('order'):
                    row_num += 1
                    col_num = 1
                    # Show the order number of this MsItem
                    ws.cell(row=row_num, column=col_num).value = msitem.order
                    # Get other stuff
                    parent = "" if msitem.parent == None else msitem.parent.order
                    firstchild = "" if msitem.firstchild == None else msitem.firstchild.order
                    next = "" if msitem.next == None else msitem.next.order
                    codico = "" if msitem.codico is None else msitem.codico.order

                    # Process the structural elements
                    col_num += 1
                    ws.cell(row=row_num, column=col_num).value = parent
                    col_num += 1
                    ws.cell(row=row_num, column=col_num).value = firstchild
                    col_num += 1
                    ws.cell(row=row_num, column=col_num).value = next
                    col_num += 1
                    ws.cell(row=row_num, column=col_num).value = codico

                    # What kind of item is this?
                    col_num += 1
                    if msitem.itemheads.count() > 0:
                        sermonhead = msitem.itemheads.first()
                        # This is a SermonHead
                        ws.cell(row=row_num, column=col_num).value = "Structural"
                        # issue #609: needs adjustment
                        col_num = get_spec_col_num(SermonDescr, 'locus')
                        ws.cell(row=row_num, column=col_num).value = sermonhead.locus
                        col_num = get_spec_col_num(SermonDescr, 'title')
                        ws.cell(row=row_num, column=col_num).value = sermonhead.title.strip()
                    else:
                        # This is a SermonDescr
                        ws.cell(row=row_num, column=col_num).value = "Plain"
                        col_num += 1
                        sermon = msitem.itemsermons.first()
                        # Walk the items
                        for item in SermonDescr.specification:
                            if item['type'] != "":
                                key, value = sermon.custom_getkv(item, kwargs=kwargs)
                                ws.cell(row=row_num, column=col_num).value = value
                                col_num += 1
                

                # Save it
                wb.save(response)
                sData = response
            elif dtype == "json":
                # Start a *list* of manuscripts
                #  (so that we have one generic format for both a single as well as a number of manuscripts)
                lst_manu = []

                # Start one object for this particular manuscript
                oManu = dict(msitems=[])

                # Walk the mainitems
                kwargs = {'profile': profile, 'username': username, 'team_group': team_group, 'keyfield': 'path'}
                for item in Manuscript.specification:
                    # Only skip key_id items
                    if item['type'] != "fk_id":
                        key, value = self.obj.custom_getkv(item, **kwargs)
                        # Add the K/V row
                        oManu[key] = value

                # Walk all msitems of this manuscript
                for msitem in self.obj.manuitems.all().order_by('order'):
                    # Create an object for this sermon
                    oMsItem = {}

                    # Add the order of this item as well as he parent, firstchild, next
                    oMsItem['order'] = msitem.order
                    oMsItem['parent'] = "" if msitem.parent == None else msitem.parent.order
                    oMsItem['firstchild'] = "" if msitem.firstchild == None else msitem.firstchild.order
                    oMsItem['next'] = "" if msitem.next == None else msitem.next.order

                    # Create an object for this sermon
                    oSermon = {}

                    # What kind of item is this?
                    if msitem.itemheads.count() > 0:
                        sermonhead = msitem.itemheads.first()
                        # This is a SermonHead
                        oSermon['type'] = "Structural"
                        oSermon['locus'] = sermonhead.locus
                        oSermon['title'] = sermonhead.title.strip()
                    else:
                        # This is a SermonDescr
                        oSermon['type'] = "Plain"

                        # Get the actual sermon
                        sermon = msitem.itemsermons.first()
                        # Walk the items of this sermon (defined in specification)
                        for item in SermonDescr.specification:
                            if item['type'] != "" and item['type'] != "fk_id":
                                key, value = sermon.custom_getkv(item, **kwargs)
                                oSermon[key] = value
                    # Add sermon to msitem
                    oMsItem['sermon'] = oSermon
                    # Add this sermon to the list of sermons within the manuscript
                    oManu['msitems'].append(oMsItem)

                # Add object to the list
                lst_manu.append(oManu)
                # Make sure to return this list
                sData = json.dumps( lst_manu, indent=2)
            elif dtype == "tei" or dtype== "xml-tei":
                # Prepare a context for the XML creation
                context = dict(details_id=self.obj.id, download_person=username)
                context['details_url'] = 'https://passim.rich.ru.nl{}'.format(reverse('manuscript_details', kwargs={'pk': self.obj.id}))
                context['download_date_ymd'] = get_current_datetime().strftime("%Y-%m-%d")
                context['download_date'] = get_current_datetime().strftime("%d/%b/%Y")
                context['manu'] = self.obj

                # Convert into string
                sData = render_to_string("seeker/tei-template.xml", context, self.request)

                # Perform pretty printing
                tree = ET.fromstring(sData, parser=ET.XMLParser(encoding='utf-8', remove_blank_text=True))
                pretty = ET.tostring(tree, encoding="utf-8", pretty_print=True, xml_declaration=True)
                sData = pretty
        except:
            msg = oErr.get_error_message()
            oErr.DoError("ManuscriptDownload/get_data")


        return sData


# ================= CODICO =============================

class CodicoEdit(BasicDetails):
    """The details of one codicological unit"""

    model = Codico  
    mForm = CodicoForm
    prefix = 'codi'
    title = "Codicological Unit"
    rtype = "json"
    prefix_type = "simple"
    new_button = True
    backbutton = False
    mainitems = []
    use_team_group = True
    history_button = True
    
    CdrFormSet = inlineformset_factory(Codico, Daterange,
                                         form=DaterangeForm, min_num=0,
                                         fk_name = "codico",
                                         extra=0, can_delete=True, can_order=False)
    CprovFormSet = inlineformset_factory(Codico, ProvenanceCod,
                                         form=CodicoProvForm, min_num=0,
                                         fk_name = "codico",
                                         extra=0, can_delete=True, can_order=False)
    CoriFormSet = inlineformset_factory(Codico, OriginCod,
                                         form=CodicoOriginForm, min_num=0, max_num=1,
                                         fk_name = "codico",
                                         extra=0, can_delete=True, can_order=False)

    formset_objects = [{'formsetClass': CdrFormSet,   'prefix': 'cdr',   'readonly': False, 'noinit': True, 'linkfield': 'codico'},
                       {'formsetClass': CprovFormSet, 'prefix': 'cprov', 'readonly': False, 'noinit': True, 'linkfield': 'codico'},
                       {'formsetClass': CoriFormSet,  'prefix': 'cori',  'readonly': False, 'noinit': True, 'linkfield': 'codico'}]

    stype_edi_fields = ['name', 'order', 'origin', 'support', 'extent', 'format', 
                        'Daterange', 'datelist',
                        'ProvenanceCod', 'cprovlist',
                        'OriginCod', 'corilist']
    
    def custom_init(self, instance):
        if instance != None:
            manu_id = instance.manuscript.id
            # Also make sure to change the afterdelurl
            self.afterdelurl = reverse("manuscript_details", kwargs={'pk': manu_id})

        return None

    def add_to_context(self, context, instance):
        """Add to the existing context"""

        oErr = ErrHandle()
        try:
            # Need to know who this user (profile) is
            profile = Profile.get_user_profile(self.request.user.username)

            # Define the main items to show and edit
            context['mainitems'] = []

            manu_id = None if instance == None else instance.manuscript.id

            # Add a button back to the Manuscript
            topleftlist = []
            if manu_id != None:
                buttonspecs = {'label': "M", 
                     'title': "Go to manuscript {}".format(instance.manuscript.idno), 
                     'url': reverse('manuscript_details', kwargs={'pk': manu_id})}
                topleftlist.append(buttonspecs)

                ## Also make sure to change the afterdelurl
                #context['afterdelurl'] = reverse("manuscript_details", kwargs={'pk': manu_id})

                # Check if this is the *first* codico of the manuscript
                if instance.manuscript.manuscriptcodicounits.all().order_by("order").first().id == instance.id:
                    # Make sure deleting is not allowed
                    context['no_delete'] = True
            context['topleftbuttons'] = topleftlist

            # Get the main items
            mainitems_main = [
                {'type': 'plain', 'label': "Status:",       'value': instance.get_stype_light(),    'field_key': 'stype'},
                # -------- HIDDEN field values ---------------
                {'type': 'plain', 'label': "Manuscript id", 'value': manu_id,   'field_key': "manuscript",  'empty': 'hide'},
                # --------------------------------------------
                {'type': 'plain', 'label': "Manuscript:",   'value': instance.get_manu_markdown()},
                {'type': 'plain', 'label': "Title:",        'value': instance.name,                     'field_key': 'name'},
                {'type': 'safe',  'label': "Order:",        'value': instance.order},
                {'type': 'line',  'label': "Date:",         'value': instance.get_date_markdown(), 
                 'multiple': True, 'field_list': 'datelist', 'fso': self.formset_objects[0]},   #, 'template_selection': 'ru.passim.litref_template' },
                {'type': 'plain', 'label': "Support:",      'value': instance.support,                  'field_key': 'support'},
                {'type': 'plain', 'label': "Extent:",       'value': instance.extent,                   'field_key': 'extent'},
                {'type': 'plain', 'label': "Format:",       'value': instance.format,                   'field_key': 'format'},
                {'type': 'plain', 'label': "Project:",      'value': instance.get_project_markdown2()}
                ]
            for item in mainitems_main: context['mainitems'].append(item)
            username = profile.user.username
            team_group = app_editor
            mainitems_m2m = [
                {'type': 'plain', 'label': "Keywords:",     'value': instance.get_keywords_markdown(),  'field_list': 'kwlist'},
                # Was: (see issue #427)
                #      {'type': 'safe',  'label': "Origin:",       'value': instance.get_origin_markdown(),    'field_key': 'origin'},
                {'type': 'plain', 'label': "Origin:",       'value': self.get_origin_markdown(instance),    
                 'multiple': True, 'field_list': 'corilist', 'fso': self.formset_objects[2]},
                {'type': 'plain', 'label': "Provenances:",  'value': self.get_provenance_markdown(instance), 
                 'multiple': True, 'field_list': 'cprovlist', 'fso': self.formset_objects[1] }
                ]
            for item in mainitems_m2m: context['mainitems'].append(item)

            # Possibly append notes view
            if user_is_ingroup(self.request, app_editor):
                context['mainitems'].append(
                    {'type': 'plain', 'label': "Notes:",       'value': instance.get_notes_markdown(),  'field_key': 'notes'}  )

            # Signal that we have select2
            context['has_select2'] = True

            # Specify that the manuscript info should appear at the right
            title_right = '<span style="font-size: xx-small">{}</span>'.format(instance.get_full_name())
            context['title_right'] = title_right

            # Note: non-app editors may still add a comment
            lhtml = []
            if context['is_app_editor']:
                lbuttons = []

                # Some buttons are needed anyway...
                lbuttons.append(dict(title="Open a list of origins", href=reverse('origin_list'), label="Origins..."))
                lbuttons.append(dict(title="Open a list of locations", href=reverse('location_list'), label="Locations..."))

                # Build the HTML on the basis of the above
                lhtml.append("<div class='row'><div class='col-md-12' align='right'>")
                for item in lbuttons:
                    idfield = ""
                    if 'click' in item:
                        ref = " onclick='document.getElementById(\"{}\").click();'".format(item['click'])
                    elif 'submit' in item:
                        ref = " onclick='document.getElementById(\"{}\").submit();'".format(item['submit'])
                    elif 'open' in item:
                        ref = " data-toggle='collapse' data-target='#{}'".format(item['open'])
                    else:
                        ref = " href='{}'".format(item['href'])
                    if 'id' in item:
                        idfield = " id='{}'".format(item['id'])
                    lhtml.append("  <a role='button' class='btn btn-xs jumbo-3' title='{}' {} {}>".format(item['title'], ref, idfield))
                    lhtml.append("     <span class='glyphicon glyphicon-chevron-right'></span>{}</a>".format(item['label']))
                lhtml.append("</div></div>")

            # Add comment modal stuff
            initial = dict(otype="codi", objid=instance.id, profile=profile)
            context['commentForm'] = CommentForm(initial=initial, prefix="com")

            context['comment_list'] = get_usercomments('codi', instance, profile)
            lhtml.append(render_to_string("seeker/comment_add.html", context, self.request))

            # Store the after_details in the context
            context['after_details'] = "\n".join(lhtml)

        except:
            msg = oErr.get_error_message()
            oErr.DoError("CodicoEdit/add_to_context")

        # Return the context we have made
        return context

    def get_origin_markdown(self, instance):
        """Calculate a collapsable table view of the origins for this codico, for Codico details view"""

        context = dict(codi=instance)
        sBack = render_to_string("seeker/codi_origins.html", context, self.request)
        return sBack

    def get_provenance_markdown(self, instance):
        """Calculate a collapsable table view of the provenances for this codico, for Codico details view"""

        context = dict(codi=instance)
        sBack = render_to_string("seeker/codi_provs.html", context, self.request)
        return sBack

    def process_formset(self, prefix, request, formset):
        errors = []
        bResult = True
        instance = formset.instance
        for form in formset:
            if form.is_valid():
                cleaned = form.cleaned_data
                # Action depends on prefix

                if prefix == "cdr":
                    # Processing one daterange
                    newstart = cleaned.get('newstart', None)
                    newfinish = cleaned.get('newfinish', None)
                    oneref = cleaned.get('oneref', None)
                    newpages = cleaned.get('newpages', None)

                    if newstart:
                        # Possibly set newfinish equal to newstart
                        if newfinish == None or newfinish == "":
                            newfinish = newstart
                        # Double check if this one already exists for the current instance
                        obj = instance.codico_dateranges.filter(yearstart=newstart, yearfinish=newfinish).first()
                        if obj == None:
                            form.instance.yearstart = int(newstart)
                            form.instance.yearfinish = int(newfinish)
                        # Do we have a reference?
                        if oneref != None:
                            form.instance.reference = oneref
                            if newpages != None:
                                form.instance.pages = newpages
                        # Note: it will get saved with formset.save()
                elif prefix == "cprov":
                    # New method, issue #289 (last part)
                    note = cleaned.get("note")
                    prov_new = cleaned.get("prov_new")
                    if prov_new != None:
                        form.instance.provenance = prov_new
                        form.instance.note = note
                elif prefix == "cori":
                    # Don't allow more than one origin
                    count = instance.origins.count()
                    if count < 1:
                        # See issue #427
                        note = cleaned.get("note")
                        origin_new = cleaned.get("origin_new")
                        if origin_new != None:
                            form.instance.origin = origin_new
                            form.instance.note = note
                    else:
                        errors.append("A codicological unit may not have more than one Origin")
            else:
                errors.append(form.errors)
                bResult = False
        return None

    def before_save(self, form, instance):
        if instance != None:
            # Double check for the correct 'order'
            if instance.order <= 0:
                # Calculate how many CODICOs (!) there are
                codico_count = instance.manuscript.manuscriptcodicounits.count()
                # Adapt the order of this codico
                instance.order = codico_count + 1
                # The number will be automatically saved
        return True, ""

    def after_save(self, form, instance):
        msg = ""
        bResult = True
        oErr = ErrHandle()
        
        try:
            # Process many-to-many changes: Add and remove relations in accordance with the new set passed on by the user
            # (1) 'keywords'
            kwlist = form.cleaned_data['kwlist']
            adapt_m2m(CodicoKeyword, instance, "codico", kwlist, "keyword")

            # (2) 'provenances'
            cprovlist = form.cleaned_data['cprovlist']
            adapt_m2m(ProvenanceCod, instance, "codico", cprovlist, "provenance", extra=['note'], related_is_through = True)

            # (3) 'origins'
            corilist = form.cleaned_data['corilist']
            adapt_m2m(OriginCod, instance, "codico", corilist, "origin", extra=['note'], related_is_through = True)

            # Process many-to-ONE changes
            # (1) links from Daterange to Codico
            datelist = form.cleaned_data['datelist']
            adapt_m2o(Daterange, instance, "codico", datelist)

            # Make sure to process changes
            instance.refresh_from_db()
        except:
            msg = oErr.get_error_message()
            bResult = False
        return bResult, msg

    def action_add(self, instance, details, actiontype):
        """User can fill this in to his/her liking"""
        passim_action_add(self, instance, details, actiontype)

    def get_history(self, instance):
        return passim_get_history(instance)


class CodicoDetails(CodicoEdit):
    rtype = "html"

    def add_to_context(self, context, instance):
        """Add to the existing context"""

        # Start by executing the standard handling
        super(CodicoDetails, self).add_to_context(context, instance)

        oErr = ErrHandle()
        try:
            # Additional sections
            context['sections'] = []

            # Lists of related objects
            context['related_objects'] = []

        except:
            msg = oErr.get_error_message()
            oErr.DoError("CodicoDetails/add_to_context")

        # Return the context we have made
        return context

    def process_formset(self, prefix, request, formset):
        return None

    def after_save(self, form, instance):
        return True, ""


class CodicoListView(BasicList):
    """Search and list manuscripts"""
    
    model = Codico
    listform = CodicoForm
    has_select2 = True
    use_team_group = True
    paginate_by = 20
    bUseFilter = True
    prefix = "codi"
    order_cols = ['manuscript__idno', 'name', 'order', 'yearstart','yearfinish', 'stype']
    order_default = order_cols
    order_heads = [{'name': 'Manuscript', 'order': 'o=1', 'type': 'str', 'custom': 'manu'},
                   {'name': 'Title',     'order': 'o=2', 'type': 'str', 'custom': 'name', 'main': True, 'linkdetails': True},
                   {'name': 'Unit',     'order': 'o=3', 'type': 'int', 'custom': 'order',   'align': 'right'},
                   {'name': 'From',     'order': 'o=4', 'type': 'int', 'custom': 'from',    'align': 'right'},
                   {'name': 'Until',    'order': 'o=5', 'type': 'int', 'custom': 'until',   'align': 'right'},
                   {'name': 'Status',   'order': 'o=6', 'type': 'str', 'custom': 'status'}]
    filters = [ 
        {"name": "Shelfmark",       "id": "filter_manuid",           "enabled": False},
        {"name": "Title",           "id": "filter_title",            "enabled": False},
        {"name": "Origin",          "id": "filter_origin",           "enabled": False},
        {"name": "Provenance",      "id": "filter_provenance",       "enabled": False},
        {"name": "Date range",      "id": "filter_daterange",        "enabled": False},
        {"name": "Keyword",         "id": "filter_keyword",          "enabled": False},
        {"name": "Status",          "id": "filter_stype",            "enabled": False},
        {"name": "Project",         "id": "filter_project",          "enabled": False, "head_id": "filter_other"},
      ]

    searches = [
        {'section': '', 'filterlist': [
            {'filter': 'manuid',        'fkfield': 'manuscript',       'keyS': 'manuidno',      
             'keyFk': 'idno', 'keyList': 'manuidlist', 'infield': 'id'},
            {'filter': 'provenance',    'fkfield': 'provenances__location',  'keyS': 'prov_ta',       'keyId': 'prov',        'keyFk': "name"},
            {'filter': 'title',         'dbfield': 'name',                   'keyS': 'name_ta'},
            {'filter': 'origin',        'fkfield': 'origin',                 'keyS': 'origin_ta',     'keyId': 'origin',      'keyFk': "name"},
            {'filter': 'keyword',       'fkfield': 'keywords',               'keyFk': 'name', 'keyList': 'kwlist', 'infield': 'name' },
            {'filter': 'daterange',     'dbfield': 'codico_dateranges__yearstart__gte',         'keyS': 'date_from'},
            {'filter': 'daterange',     'dbfield': 'codico_dateranges__yearfinish__lte',        'keyS': 'date_until'},
            {'filter': 'stype',         'dbfield': 'stype',                  'keyList': 'stypelist', 'keyType': 'fieldchoice', 'infield': 'abbr' }
            ]},
        {'section': 'other', 'filterlist': [
            #{'filter': 'project',   'fkfield': 'manuscript__project',  'keyS': 'project', 'keyFk': 'id', 'keyList': 'prjlist', 'infield': 'name' }
            ]}
         ]

    def add_to_context(self, context, initial):

        # Add a form to enter a URL
        context['searchurlform'] = SearchUrlForm()
        
        return context

    def get_field_value(self, instance, custom):
        sBack = ""
        sTitle = ""
        html = []
        if custom == "manu":
            if instance.manuscript != None:
                url = reverse("manuscript_details", kwargs={'pk': instance.manuscript.id})
                html.append("<span class='manuscript-idno'><a href='{}'>{}</a></span>".format(url,instance.manuscript.idno))
        elif custom == "name":
            if instance.name:
                html.append("<span class='manuscript-title'>{}</span>".format(instance.name[:100]))
                sTitle = instance.name
        elif custom == "order":
            html.append("{}".format(instance.order))
        elif custom == "from":
            for item in instance.codico_dateranges.all():
                html.append("<div>{}</div>".format(item.yearstart))
        elif custom == "until":
            for item in instance.codico_dateranges.all():
                html.append("<div>{}</div>".format(item.yearfinish))
        elif custom == "status":
            html.append(instance.get_stype_light())
            sTitle = instance.get_stype_display()
        # Combine the HTML code
        sBack = "\n".join(html)
        return sBack, sTitle

    def adapt_search(self, fields):
        # Adapt the search to the keywords that *may* be shown
        lstExclude=None
        qAlternative = None

        oErr = ErrHandle()
        try:

            # prjlist = None
            # Check if a list of keywords is given
            if 'kwlist' in fields and fields['kwlist'] != None and len(fields['kwlist']) > 0:
                # Get the list
                kwlist = fields['kwlist']
                # Get the user
                username = self.request.user.username
                user = User.objects.filter(username=username).first()
                # Check on what kind of user I am
                if not user_is_ingroup(self.request, app_editor):
                    # Since I am not an app-editor, I may not filter on keywords that have visibility 'edi'
                    kwlist = Keyword.objects.filter(id__in=kwlist).exclude(Q(visibility="edi")).values('id')
                    fields['kwlist'] = kwlist

            # Process the date range stuff
            date_from = fields.get("date_from")
            date_until = fields.get("date_until")
            if not date_from is None and not date_until is None:
                # Both dates are specified: include looking for overlap
                qThis_a = Q(codico_dateranges__yearstart__gte=date_from)
                qThis_b = Q(codico_dateranges__yearstart__lte=date_until)
                qThis_c = Q(codico_dateranges__yearfinish__lte=date_until)
                qThis_d = Q(codico_dateranges__yearfinish__gte=date_from)
                qThis_1 =  ( qThis_a &  ( qThis_b |  qThis_c ) )
                qThis_2 =  ( qThis_c &  ( qThis_d |  qThis_a ) )
                qThis = ( qThis_1 | qThis_2 )
                fields['date_from'] = qThis
                fields['date_until'] = ""

        except:
            msg = oErr.get_error_message()
            oErr.DoError("CodicoListView/adapt_search")

        return fields, lstExclude, qAlternative

    def get_helptext(self, name):
        """Use the get_helptext function defined in models.py"""
        return get_helptext(name)


# ================= SERMONGOLD =============================

class SermonGoldListView(BasicList):
    """Search and list manuscripts"""
    
    model = SermonGold
    listform = SermonGoldForm
    prefix = "gold"
    basic_name = "gold"
    plural_name = "Gold sermons"
    sg_name = "Gold sermon"    
    new_button = False      # Don't show the [Add a new Gold Sermon] button here. 
                            # Issue #173: creating Gold Sermons may only happen from SuperSermonGold list view
    has_select2 = True
    use_team_group = True
    bUseFilter = True 
    paginate_by = 20
    order_default = ['author__name', 'siglist', 'equal__code', 'srchincipit;srchexplicit', '', '', 'stype']
    order_cols = order_default
    order_heads = [{'name': 'Author', 'order': 'o=1', 'type': 'str', 'custom': 'author'}, 
                   {'name': 'Signature', 'order': 'o=2', 'type': 'str', 'custom': 'signature'}, 
                   {'name': 'Passim Code', 'order': 'o=3', 'type': 'str', 'custom': 'code'}, 
                   {'name': 'Incipit ... Explicit', 'order': 'o=4', 'type': 'str', 'custom': 'incexpl', 'main': True, 'linkdetails': True},
                   {'name': 'Editions', 'order': '', 'type': 'str', 'custom': 'edition'},
                   {'name': 'Links', 'order': '', 'type': 'str', 'custom': 'links'},
                   {'name': 'Status', 'order': '', 'type': 'str', 'custom': 'status'}]
    filters = [
        {"name": "Gryson or Clavis","id": "filter_signature",   "enabled": False},
        {"name": "Passim code",     "id": "filter_code",        "enabled": False},
        {"name": "Author",          "id": "filter_author",      "enabled": False},
        {"name": "Incipit",         "id": "filter_incipit",     "enabled": False},
        {"name": "Explicit",        "id": "filter_explicit",    "enabled": False},
        {"name": "Keyword",         "id": "filter_keyword",     "enabled": False},
        {"name": "Status",          "id": "filter_stype",       "enabled": False},
        {"name": "Project",         "id": "filter_project",     "enabled": False},
        {"name": "Collection...",   "id": "filter_collection",  "enabled": False, "head_id": "none"},
        {"name": "Manuscript",      "id": "filter_collmanu",    "enabled": False, "head_id": "filter_collection"},
        {"name": "Sermon",          "id": "filter_collsermo",   "enabled": False, "head_id": "filter_collection"},
        {"name": "Sermon Gold",     "id": "filter_collgold",    "enabled": False, "head_id": "filter_collection"},
        {"name": "Authority file",  "id": "filter_collsuper",  "enabled": False, "head_id": "filter_collection"},
        {"name": "Historical",      "id": "filter_collhist",    "enabled": False, "head_id": "filter_collection"},
        ]       
    searches = [
        {'section': '', 'filterlist': [
            {'filter': 'incipit',   'dbfield': 'srchincipit',       'keyS': 'incipit',  'regex': adapt_regex_incexp},
            {'filter': 'explicit',  'dbfield': 'srchexplicit',      'keyS': 'explicit', 'regex': adapt_regex_incexp},
            {'filter': 'author',    'fkfield': 'author',            'keyS': 'authorname', 
             'keyFk': 'name',       'keyList': 'authorlist', 'infield': 'id', 'external': 'gold-authorname' },
            {'filter': 'signature', 'fkfield': 'goldsignatures',    'keyS': 'signature',    'help': 'signature',
             'keyFk': 'code',       'keyList': 'siglist',   'keyId': 'signatureid', 'infield': 'code' },
            
            {'filter': 'code',      'fkfield': 'equal',             'keyS': 'codetype',     'help': 'passimcode',      
             'keyFk': 'code',       'keyList': 'passimlist',  'infield': 'code'}, # passimlist
            
             {'filter': 'keyword',   'fkfield': 'keywords',    'keyFk': 'name',  'keyList': 'kwlist',   'infield': 'name' },
            {'filter': 'project',   'fkfield': 'equal__projects',    'keyFk': 'name',  'keyList': 'projlist', 'infield': 'name'}, # view keyword
            {'filter': 'stype',     'dbfield': 'stype',             'keyList': 'stypelist', 'keyType': 'fieldchoice', 'infield': 'abbr' } 
            ]},
        {'section': 'collection', 'filterlist': [
            {'filter': 'collmanu',  'fkfield': 'sermondescr__manu__collections','keyS': 'collection','keyFk': 'name', 'keyList': 'collist_m', 'infield': 'name' }, 
            {'filter': 'collsermo', 'fkfield': 'sermondescr__collections',      'keyS': 'collection','keyFk': 'name', 'keyList': 'collist_s', 'infield': 'name' }, 
            {'filter': 'collgold',  'fkfield': 'collections',                   'keyS': 'collection','keyFk': 'name', 'keyList': 'collist_sg', 'infield': 'name' }, 
            {'filter': 'collsuper', 'fkfield': 'equal__collections',            'keyS': 'collection','keyFk': 'name', 'keyList': 'collist_ssg', 'infield': 'name' }, 
            {'filter': 'collhist',  'fkfield': 'equal__collections',                            
             'keyS': 'collection',  'keyFk': 'name', 'keyList': 'collist_hist', 'infield': 'name' },
            {'filter': 'atype',     'fkfield': 'equal', 'keyS': 'atype', 'keyFk': 'atype', 'keyList': 'atypelist',  'infield': 'atype'},
            ]},
        ]
    uploads = [{"title": "gold", "label": "Gold", "url": "import_gold", "msg": "Upload Excel files"}]

    def add_to_context(self, context, initial):
        # Find out who the user is
        profile = Profile.get_user_profile(self.request.user.username)
        context['basketsize'] = 0 if profile == None else profile.basketsize_gold
        context['basket_show'] = reverse('basket_show_gold')
        context['basket_update'] = reverse('basket_update_gold')
        return context

    def get_basketqueryset(self):
        if self.basketview:
            profile = Profile.get_user_profile(self.request.user.username)
            # TODO: chck the below -- is that SG??
            qs = profile.basketitems_gold.all()
        else:
            qs = SermonGold.objects.all()
        return qs

    def get_field_value(self, instance, custom):
        sBack = ""
        sTitle = ""
        html = []
        if custom == "author":
            if instance.author:
                html.append("<span style='color: darkgreen; font-size: small;'>{}</span>".format(instance.author.name[:20]))
                sTitle = instance.author.name
            else:
                html.append("<span><i>(unknown)</i></span>")
        elif custom == "signature": 
            html.append(instance.get_signatures_markdown())            
        elif custom == "code":
            equal = instance.equal
            if equal:
                code = "(undetermined)" if equal.code == None else equal.code
                url = reverse('equalgold_details', kwargs={'pk': equal.id})
            else:
                code = "(not specified)"
                url = "#"
            html.append("<span class='passimcode'><a class='nostyle' href='{}'>{}</a></span>".format(url, code))
        elif custom == "edition":
            html.append("<span>{}</span>".format(instance.editions()[:20]))
            sTitle = instance.editions()
        elif custom == "incexpl":
            if instance.incipit == "" and instance.explicit == "":
                url = reverse('gold_details', kwargs={'pk': instance.id})
                html.append("<span><a href='{}'><i>(not specified)</i></a></span>".format(url))
            else:
                html.append("<span>{}</span>".format(instance.get_incipit_markdown()))
                dots = "..." if instance.incipit else ""
                html.append("<span style='color: blue;'>{}</span>".format(dots))
                html.append("<span>{}</span>".format(instance.get_explicit_markdown()))
        elif custom == "links":
            for link_def in instance.link_oview():
                if link_def['count'] > 0:
                    html.append("<span class='badge {}' title='{}'>{}</span>".format(link_def['class'], link_def['title'], link_def['count']))
        elif custom == "status":
            # Provide that status badge
            # html.append("<span class='badge' title='{}'>{}</span>".format(instance.get_stype_light(), instance.stype[:1]))
            html.append(instance.get_stype_light())
            sTitle = instance.get_stype_display()
        # Combine the HTML code
        sBack = "\n".join(html)
        return sBack, sTitle

    def initializations(self):
        # ======== One-time adaptations ==============
        listview_adaptations("sermongold_list")

        # Should json be added?
        if user_is_superuser(self.request):
            # Possibly add to 'uploads'
            bHasJson = False
            for item in self.uploads:
                if item['title'] == "huwajson":
                    bHasJson = True

            # Should json be added?
            if not bHasJson:
                # Add a reference to the Json upload method
                msg = "Upload Opera Editions from a HUWA json specification."
                oJson = dict(title="huwajson", label="HUWA opera editions json",
                              url=reverse('gold_upload_json'),
                              type="multiple", msg=msg)
                self.uploads.append(oJson)

        return None

    def adapt_search(self, fields):
        # Adapt the search to the keywords that *may* be shown
        lstExclude=None
        qAlternative = None

        # Check if a list of keywords is given
        if 'kwlist' in fields and fields['kwlist'] != None and len(fields['kwlist']) > 0:
            # Get the list
            kwlist = fields['kwlist']
            # Get the user
            username = self.request.user.username
            user = User.objects.filter(username=username).first()
            # Check on what kind of user I am
            if not user_is_ingroup(self.request, app_editor):
                # Since I am not an app-editor, I may not filter on keywords that have visibility 'edi'
                kwlist = Keyword.objects.filter(id__in=kwlist).exclude(Q(visibility="edi")).values('id')
                fields['kwlist'] = kwlist

        # Adapt the search for empty passim codes
        if 'codetype' in fields:
            codetype = fields['codetype']
            if codetype == "non":
                lstExclude = []
                lstExclude.append(Q(equal__isnull=False))
            elif codetype == "spe":
                lstExclude = []
                lstExclude.append(Q(equal__isnull=True))
            # Reset the codetype
            fields['codetype'] = ""
                
        # Return the adapted stuff
        return fields, lstExclude, qAlternative

    def view_queryset(self, qs):
        search_id = [x['id'] for x in qs.values('id')]
        profile = Profile.get_user_profile(self.request.user.username)
        profile.search_gold = json.dumps(search_id)
        profile.save()
        return None

    def get_helptext(self, name):
        """Use the get_helptext function defined in models.py"""
        return get_helptext(name)


class SermonGoldEdit(BasicDetails):
    """The details of one sermon"""

    model = SermonGold
    mForm = SermonGoldForm
    prefix = 'gold'
    title = "Sermon Gold"
    rtype = "json"
    mainitems = []
    basic_name = "gold"
    use_team_group = True
    history_button = True

    GkwFormSet = inlineformset_factory(SermonGold, SermonGoldKeyword,
                                       form=SermonGoldKeywordForm, min_num=0,
                                       fk_name="gold", extra=0)
    GsignFormSet = inlineformset_factory(SermonGold, Signature,
                                         form=SermonGoldSignatureForm, min_num=0,
                                         fk_name = "gold",
                                         extra=0, can_delete=True, can_order=False)
    GediFormSet = inlineformset_factory(SermonGold, EdirefSG,
                                         form = SermonGoldEditionForm, min_num=0,
                                         fk_name = "sermon_gold",
                                         extra=0, can_delete=True, can_order=False)
    GcolFormSet = inlineformset_factory(SermonGold, CollectionGold,
                                       form=SermonGoldCollectionForm, min_num=0,
                                       fk_name="gold", extra=0)
    GlitFormSet = inlineformset_factory(SermonGold, LitrefSG,
                                         form = SermonGoldLitrefForm, min_num=0,
                                         fk_name = "sermon_gold",
                                         extra=0, can_delete=True, can_order=False)
    GftxtFormSet = inlineformset_factory(SermonGold, Ftextlink,
                                         form=SermonGoldFtextlinkForm, min_num=0,
                                         fk_name = "gold",
                                         extra=0, can_delete=True, can_order=False)
    
    formset_objects = [{'formsetClass': GsignFormSet, 'prefix': 'gsign', 'readonly': False, 'noinit': True, 'linkfield': 'gold'},
                       {'formsetClass': GkwFormSet,   'prefix': 'gkw',   'readonly': False, 'noinit': True, 'linkfield': 'gold'},
                       {'formsetClass': GediFormSet,  'prefix': 'gedi',  'readonly': False, 'noinit': True, 'linkfield': 'sermon_gold'}, 
                       {'formsetClass': GcolFormSet,  'prefix': 'gcol',  'readonly': False, 'noinit': True, 'linkfield': 'gold'},
                       {'formsetClass': GlitFormSet,  'prefix': 'glit',  'readonly': False, 'noinit': True, 'linkfield': 'sermon_gold'},
                       {'formsetClass': GftxtFormSet, 'prefix': 'gftxt', 'readonly': False, 'noinit': True, 'linkfield': 'gold'}]

    # Note: do *NOT* include 'authorname', 
    stype_edi_fields = ['author', 'incipit', 'explicit', 'bibliography', 'equal',
                        #'kwlist', 
                        'Signature', 'siglist',
                        #'CollectionGold', 'collist_sg',
                        'EdirefSG', 'edilist',
                        'LitrefSG', 'litlist',
                        'Ftextlink', 'ftxtlist']

    def add_to_context(self, context, instance):
        """Add to the existing context"""

        oErr = ErrHandle()
        try:
            # Need to know who this user (profile) is
            profile = Profile.get_user_profile(self.request.user.username)
            username = profile.user.username
            team_group = app_editor

            # Define the main items to show and edit
            context['mainitems'] = [
                {'type': 'safe', 'label': "Belongs to:",            'value': instance.get_ssg_markdown,   
                 'title': 'Belongs to the equality set of Authority file...', 'field_key': "equal"}, 
                {'type': 'safe',  'label': "Together with:",        'value': instance.get_eqset,
                 'title': 'Other Sermons Gold members of the same equality set'},
                {'type': 'plain', 'label': "Status:",               'value': instance.get_stype_light(),    'field_key': 'stype', 'hidenew': True},
                {'type': 'plain', 'label': "Associated author:",    'value': instance.get_author,           'field_key': 'author'},
                {'type': 'safe',  'label': "Incipit:",              'value': instance.get_incipit_markdown, 
                 'field_key': 'incipit',  'key_ta': 'gldincipit-key'}, 
                {'type': 'safe',  'label': "Explicit:",             'value': instance.get_explicit_markdown,
                 'field_key': 'explicit', 'key_ta': 'gldexplicit-key'}, 
                {'type': 'plain', 'label': "Retractationes:",       'value': instance.get_retr(),           'field_key': 'retractationes'},
                {'type': 'plain', 'label': "Bibliography:",         'value': instance.get_bibliography_markdown(),  'field_key': 'bibliography'},
                {'type': 'line',  'label': "Keywords:",             'value': instance.get_keywords_markdown(), 
                 'field_list': 'kwlist', 'fso': self.formset_objects[1], 'maywrite': True},
                {'type': 'plain', 'label': "Keywords (user):", 'value': self.get_userkeywords(instance, profile, context),   'field_list': 'ukwlist',
                 'title': 'User-specific keywords. If the moderator accepts these, they move to regular keywords.'},
                {'type': 'line',  'label': "Keywords (related):",   'value': instance.get_keywords_ssg_markdown(),
                 'title': 'Keywords attached to the Authority file of which this Sermon Gold is part'},
                {'type': 'line', 'label': "Gryson/Clavis codes:",   'value': instance.get_signatures_markdown(),  'unique': True, 
                 'multiple': True, 'field_list': 'siglist', 'fso': self.formset_objects[0]},
                {'type': 'plain', 'label': "Personal datasets:",    'value': instance.get_collections_markdown(username, team_group, settype="pd"), 
                 'multiple': True, 'field_list': 'collist_sg', 'fso': self.formset_objects[3] },
                {'type': 'line', 'label': "Editions:",              'value': instance.get_editions_markdown(), 
                 'multiple': True, 'field_list': 'edilist', 'fso': self.formset_objects[2], 'template_selection': 'ru.passim.litref_template'},
                {'type': 'line', 'label': "Literature:",            'value': instance.get_litrefs_markdown(), 
                 'multiple': True, 'field_list': 'litlist', 'fso': self.formset_objects[4], 'template_selection': 'ru.passim.litref_template'},
                {'type': 'line', 'label': "Full text links:",       'value': instance.get_ftxtlinks_markdown(), 
                 'multiple': True, 'field_list': 'ftxtlist', 'fso': self.formset_objects[5]},
                  # Project2 HIER, aanpassen want mag geen optie zijn om te editen, wel het geval, of het lijkt iig
                {'type': 'plain', 'label': "Project (SSG):",     'value': instance.get_project_ssg_markdown2()},
                ]
            # Notes:
            # Collections: provide a link to the SSG-listview, filtering on those SSGs that are part of one particular collection

            # If this user belongs to the ProjectApprover of HUWA, show him the HUWA ID if it is there
            if not instance is None:
                # NOTE: this code should be extended to include other external projects, when the time is right
                ext = SermonGoldExternal.objects.filter(gold=instance).first()
                if not ext is None:
                    # Get the field values
                    externaltype = ext.externaltype
                    externalid = ext.externalid
                    if externaltype == "huwop" and externalid > 0 and profile.is_project_approver("huwa"):
                        # Okay, the user has the right to see the externalid
                        oItem = dict(type="plain", label="HUWA id", value=externalid)
                        context['mainitems'].insert(0, oItem)
                        # The user also has the right to see EDITION information
                        if not instance.edinote is None and instance.edinote[0] == "[":
                            # Read the list of literature references
                            lst_edilit = json.loads(instance.edinote)
                            context['edilits'] = lst_edilit
                            value = render_to_string('seeker/huwa_edilit.html', context, self.request)
                            oItemEdi = dict(type="safe", label="HUWA editions", value=value)
                            context['mainitems'].append(oItemEdi)

            # Add comment modal stuff
            lhtml = []
            initial = dict(otype="gold", objid=instance.id, profile=profile)
            context['commentForm'] = CommentForm(initial=initial, prefix="com")
            context['comment_list'] = get_usercomments('gold', instance, profile)
            lhtml.append(render_to_string("seeker/comment_add.html", context, self.request))
            context['after_details'] = "\n".join(lhtml)


            # Signal that we have select2
            context['has_select2'] = True
        except:
            msg = oErr.get_error_message()
            oErr.DoError("SermonGoldEdit/add_to_context")

        # Return the context we have made
        return context

    def after_new(self, form, instance):
        """Action to be performed after adding a new item"""

        # Set the 'afternew' URL
        self.afternewurl = reverse('search_gold')

        # Create a new equality set to which we add this Gold sermon
        if instance.equal == None:
            geq = EqualGold.create_empty()
            instance.equal = geq
            instance.save()

        # Return positively
        return True, "" 

    def process_formset(self, prefix, request, formset):

        errors = []
        bResult = True
        instance = formset.instance
        for form in formset:
            if form.is_valid():
                cleaned = form.cleaned_data
                # Action depends on prefix
                if prefix == "gsign":
                    # Signature processing
                    editype = ""
                    code = ""
                    if 'newgr' in cleaned and cleaned['newgr'] != "":
                        # Add gryson
                        editype = "gr"
                        code = cleaned['newgr']
                    elif 'newcl' in cleaned and cleaned['newcl'] != "":
                        # Add gryson
                        editype = "cl"
                        code = cleaned['newcl']
                    elif 'newot' in cleaned and cleaned['newot'] != "":
                        # Add gryson
                        editype = "ot"
                        code = cleaned['newot']
                    if editype != "":
                        # Set the correct parameters
                        form.instance.code = code
                        form.instance.editype = editype
                        # Note: it will get saved with formset.save()
                elif prefix == "gkw":
                    # Keyword processing
                    if 'newkw' in cleaned and cleaned['newkw'] != "":
                        newkw = cleaned['newkw']
                        # Is the KW already existing?
                        obj = Keyword.objects.filter(name=newkw).first()
                        if obj == None:
                            obj = Keyword.objects.create(name=newkw)
                        # Make sure we set the keyword
                        form.instance.keyword = obj
                        # Note: it will get saved with formset.save()
                elif prefix == "gcol":
                    # Collection processing
                    if 'newcol' in cleaned and cleaned['newcol'] != "":
                        newcol = cleaned['newcol']
                        # Is the COL already existing?
                        obj = Collection.objects.filter(name=newcol).first()
                        if obj == None:
                            # TODO: add profile here
                            profile = Profile.get_user_profile(request.user.username)
                            obj = Collection.objects.create(name=newcol, type='gold', owner=profile)
                        # Make sure we set the keyword
                        form.instance.collection = obj
                        # Note: it will get saved with formset.save()
                elif prefix == "gedi":
                    # Edition processing
                    newpages = ""
                    if 'newpages' in cleaned and cleaned['newpages'] != "":
                        newpages = cleaned['newpages']
                    # Also get the litref
                    if 'oneref' in cleaned:
                        litref = cleaned['oneref']
                        # Check if all is in order
                        if litref:
                            form.instance.reference = litref
                            if newpages:
                                form.instance.pages = newpages
                    # Note: it will get saved with form.save()
                elif prefix == "glit":
                    # Literature reference processing
                    newpages = ""
                    if 'newpages' in cleaned and cleaned['newpages'] != "":
                        newpages = cleaned['newpages']
                    # Also get the litref
                    if 'oneref' in cleaned:
                        litref = cleaned['oneref']
                        # Check if all is in order
                        if litref:
                            form.instance.reference = litref
                            if newpages:
                                form.instance.pages = newpages
                    # Note: it will get saved with form.save()
                elif prefix == "gftxt":
                    # Process many-to-ONE full-text links
                    if 'newurl' in cleaned and cleaned['newurl'] != "":
                        form.instance.url = cleaned['newurl']
                        # Note: it will get saved with formset.save()
            else:
                errors.append(form.errors)
                bResult = False
        return None

    def before_save(self, form, instance):
        # Get the old equal
        equal_old = SermonGold.objects.filter(id=instance.id).first().equal
        # Get the new equal
        equal = instance.equal
        # Normal behaviour
        response = super(SermonGoldEdit, self).before_save(form, instance)
        # If old differs from new
        if equal_old != equal:
            # Adapt the SG count value
            if equal_old != None: equal_old.set_sgcount()
            if equal != None: equal.set_sgcount()
            # Adapt the 'firstsig' value
            if equal_old != None: equal_old.set_firstsig()
            if equal != None: equal.set_firstsig()
        # Return result
        return response

    def after_save(self, form, instance):
        msg = ""
        bResult = True
        oErr = ErrHandle()
        
        try:
            # Need to know if the user has special priviledges
            userplus = None
            if user_is_ingroup(self.request, app_userplus): userplus = True

            # Process many-to-many changes: Add and remove relations in accordance with the new set passed on by the user
            # (1) 'keywords'
            kwlist = form.cleaned_data['kwlist']
            adapt_m2m(SermonGoldKeyword, instance, "gold", kwlist, "keyword", userplus=userplus)
            
            # (2) user-specific 'keywords'
            ukwlist = form.cleaned_data['ukwlist']
            profile = Profile.get_user_profile(self.request.user.username)
            adapt_m2m(UserKeyword, instance, "gold", ukwlist, "keyword", qfilter = {'profile': profile}, extrargs = {'profile': profile, 'type': 'gold'})

            # (3) 'editions'
            edilist = form.cleaned_data['edilist']
            adapt_m2m(EdirefSG, instance, "sermon_gold", edilist, "reference", extra=['pages'], related_is_through = True)

            # (4) 'collections'
            collist_sg = form.cleaned_data['collist_sg']
            adapt_m2m(CollectionGold, instance, "gold", collist_sg, "collection")

            # (5) 'literature'
            litlist = form.cleaned_data['litlist']
            adapt_m2m(LitrefSG, instance, "sermon_gold", litlist, "reference", extra=['pages'], related_is_through = True)

            # Process many-to-ONE changes
            # (1) 'goldsignatures'
            siglist = form.cleaned_data['siglist']
            adapt_m2o(Signature, instance, "gold", siglist)

            # (2) 'full text links'
            ftxtlist = form.cleaned_data['ftxtlist']
            adapt_m2o(Ftextlink, instance, "gold", ftxtlist)
        except:
            msg = oErr.get_error_message()
            bResult = False
        return bResult, msg

    def action_add(self, instance, details, actiontype):
        """User can fill this in to his/her liking"""
        passim_action_add(self, instance, details, actiontype)

    def get_userkeywords(self, instance, profile, context):
        sBack = ""
        oErr = ErrHandle()
        try:
            # At the very least get the list of user keywords
            sKeywordList = instance.get_keywords_user_markdown(profile)
            # Now determine what interface to show
            if context['is_app_editor']:
                # We simply return the list
                sBack = sKeywordList
            else:
                # We need to construct an interface:
                # (1) Set up the context to include view and edit elements
                context['userkeywords'] = sKeywordList
                context['ukwform'] = context['{}Form'.format(self.prefix)]
                context['ukwupdate'] = reverse('gold_ukw', kwargs={'pk': instance.id })

                # Combine and provide the code needed
                # sBack = get_userkeyword_interface(self.request, context)
                sBack = render_to_string("seeker/userkeyword_edit.html", context, self.request)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("get_userkeywords")
        return sBack

    def get_history(self, instance):
        return passim_get_history(instance)


class SermonGoldDetails(SermonGoldEdit):
    """The details of one sermon"""

    rtype = "html"

    def add_to_context(self, context, instance):
        """Add to the existing context"""

        # Start by executing the standard handling
        super(SermonGoldDetails, self).add_to_context(context, instance)

        # Are we copying information?? (only allowed if we are the app_editor)
        if 'goldcopy' in self.qd and context['is_app_editor']:
            # Get the ID of the gold sermon from which information is to be copied to the SSG
            goldid = self.qd['goldcopy']
        else:
            context['sections'] = []

            # List of post-load objects
            postload_objects = []
            context['postload_objects'] = postload_objects

            # Lists of related objects
            related_objects = []
            context['related_objects'] = related_objects

        # Return the context we have made
        return context

    def before_save(self, form, instance):
        return True, ""

    def process_formset(self, prefix, request, formset):
        return None

    def after_save(self, form, instance):
        return True, ""


class SermonGoldUserKeyword(SermonGoldDetails):
    """This version only looks at one kind of data: the ukwlist"""

    newRedirect = True

    def custom_init(self, instance):
        oErr = ErrHandle()
        try:
            # Make sure to set the correct redirect page
            if instance:
                self.redirectpage = reverse("gold_details", kwargs={'pk': instance.id})
            # Make sure we are not saving
            self.do_not_save = True

            # Get the value
            qd = self.qd
            if self.prefix_type == "simple":
                sUkwList = "{}-ukwlist".format(self.prefix)
            else:
                sUkwList = "{}-{}-ukwlist".format(self.prefix, instance.id)
            ukwlist_ids = qd.getlist(sUkwList)
            ukwlist = Keyword.objects.filter(id__in=ukwlist_ids)
            
            # Process the contents of the ukwlist, the chosen user-keywords
            profile = Profile.get_user_profile(self.request.user.username)
            adapt_m2m(UserKeyword, instance, "gold", ukwlist, "keyword", qfilter = {'profile': profile}, extrargs = {'profile': profile, 'type': 'gold'})

        except:
            msg = oErr.get_error_message()
            oErr.DoError("SermonGoldUserKeyword/custom_init")
    

# ================= EQUALGOLD =============================

class EqualGoldEdit(BasicDetails):
    model = EqualGold
    mForm = SuperSermonGoldForm
    prefix = 'ssg'
    title = "Authority file"
    rtype = "json"
    new_button = True
    mainitems = []
    use_team_group = True
    history_button = True
    comment_button = True
        
    EqgcolFormSet = inlineformset_factory(EqualGold, CollectionSuper,
                                       form=SuperSermonGoldCollectionForm, min_num=0,
                                       fk_name="super", extra=0)
    SsgLinkFormSet = inlineformset_factory(EqualGold, EqualGoldLink,
                                         form=EqualGoldLinkForm, min_num=0,
                                         fk_name = "src",
                                         extra=0, can_delete=True, can_order=False)
        
    # This one is not in use right now
    # If used: double check the proper working of the EqualGoldForm
    GeqFormSet = inlineformset_factory(EqualGold, SermonGold, 
                                         form=EqualGoldForm, min_num=0,
                                         fk_name = "equal",
                                         extra=0, can_delete=True, can_order=False)

    EqgBrefFormSet = inlineformset_factory(EqualGold, BibRange,
                                         form=BibRangeForm, min_num=0,
                                         fk_name = "equal",
                                         extra=0, can_delete=True, can_order=False)
    
    formset_objects = [
        {'formsetClass': EqgcolFormSet,  'prefix': 'eqgcol',  'readonly': False, 'noinit': True, 'linkfield': 'super'},
        {'formsetClass': SsgLinkFormSet, 'prefix': 'ssglink', 'readonly': False, 'noinit': True, 'initial': [{'linktype': LINK_PARTIAL }], 'clean': True},     
        # {'formsetClass': GeqFormSet,    'prefix': 'geq',    'readonly': False, 'noinit': True, 'linkfield': 'equal'}
        {'formsetClass': EqgBrefFormSet, 'prefix': 'eqgbref', 'readonly': False, 'noinit': True, 'linkfield': 'equal'},
        ]

    # Note: do not include [code] in here
    stype_edi_fields = ['author', 'number', 'incipit', 'explicit','bibleref', 'fulltext',
                        #'kwlist', 
                        #'CollectionSuper', 'collist_ssg',
                        'EqualGoldLink', 'superlist',
                        'LitrefSG', 'litlist',
                        'goldlist', 'projlist']

    def add_to_context(self, context, instance):
        """Add to the existing context"""

        oErr = ErrHandle()
        try:
            # One general item is the 'help-popup' to be shown when the user clicks on 'Author'
            info = render_to_string('seeker/author_info.html')

            # Need to know who this user (profile) is
            profile = Profile.get_user_profile(self.request.user.username)
            username = profile.user.username
            team_group = app_editor

            # Prepare saveditem handling
            saveditem_button = get_saveditem_html(self.request, instance, profile, sitemtype="ssg")
            saveditem_form = get_saveditem_html(self.request, instance, profile, "form", sitemtype="ssg")

            # Determine the passim code and possibly a warning if this is a moved item
            str_code = instance.get_code()
            if not instance.moved is None:
                message = "this AF is an old copy of one that has moved. See Details."
                str_code = '{} <span><b>NOTE</b>:</span>&nbsp;<span style="color: red;">{}</span>'.format(str_code, message)

            # Define the main items to show and edit
            author_id = None if instance.author is None else instance.author.id
            context['mainitems'] = [
                {'type': 'plain', 'label': "Status:",         'value': instance.get_stype_light(),'field_key': 'stype'},
                {'type': 'safe',  'label': "Saved item:",     'value': saveditem_button          },
                {'type': 'plain', 'label': "Gryson/Clavis :", 'value': instance.get_signatures_markdown_equal},               
                # Issue #295: the [number] (number within author) must be there, though hidden, not editable
                {'type': 'plain', 'label': "Number:",        'value': instance.number,    'field_key': 'number',   'empty': 'hide'},
                {'type': 'plain', 'label': "Author id:",     'value': author_id,          'field_key': 'author',   'empty': 'hide'},
                {'type': 'plain', 'label': "Incipit:",       'value': instance.incipit,   'field_key': 'incipit',  'empty': 'hide'},
                {'type': 'plain', 'label': "Explicit:",      'value': instance.explicit,  'field_key': 'explicit', 'empty': 'hide'},
                {'type': 'plain', 'label': "Transcription:", 'value': instance.fulltext,  'field_key': 'fulltext', 'empty': 'hide'},
                # Issue #212: remove this sermon number
                # {'type': 'plain', 'label': "Sermon number:", 'value': instance.number, 'field_view': 'number',             
                {'type': 'plain', 'label': "Passim Code:",   'value': str_code,   'title': 'The Passim Code is automatically determined'},   
                ]

            # some more items need to be added
            mainitems_add = []

            for oItem in mainitems_add: 
                context['mainitems'].append(oItem)
            
            # Notes:
            # Collections: provide a link to the SSG-listview, filtering on those SSGs that are part of one particular collection           
            context['mainsections'] = [
            {'name': 'Details', 'id': 'equalgold_details', 'fields': [                
                {'type': 'line', 'label': "Author:", 'value': instance.author_help(info), 'field_key': 'newauthor', 'order': 1},                
                {'type': 'plain', 'label': "Incipit:", 'value': instance.incipit,   'field_key': 'incipit',  'empty': 'hide', 'order': 2},
                {'type': 'plain', 'label': "Explicit:", 'value': instance.explicit,  'field_key': 'explicit', 'empty': 'hide', 'order': 3},                 
                {'type': 'safe',  'label': "Incipit:", 'value': instance.get_incipit_markdown("search"), 
                 'field_key': 'newincipit',  'key_ta': 'gldincipit-key', 'title': instance.get_incipit_markdown("actual"), 'order': 2}, 
                {'type': 'safe',  'label': "Explicit:", 'value': instance.get_explicit_markdown("search"),
                 'field_key': 'newexplicit', 'key_ta': 'gldexplicit-key', 'title': instance.get_explicit_markdown("actual"), 'order': 3}, 
                 {'type': 'line', 'label': "Editions:", 'value': instance.get_editions_markdown(),
                 'title': 'All the editions associated with the Gold Sermons in this equality set', 'order': 4},
                {'type': 'line', 'label': "Literature:", 'value': instance.get_litrefs_markdown(), 
                 'title': 'All the literature references associated with the Gold Sermons in this equality set', 'order': 6},
                {'type': 'line',  'label': "Keywords:", 'value': instance.get_keywords_markdown(), 'field_list': 'kwlist', 'order': 7},
                {'type': 'plain', 'label': "Bible reference(s):", 'value': instance.get_bibleref(),        
                'multiple': True, 'field_list': 'bibreflist', 'fso': self.formset_objects[2], 'order': 8},                
                {'type': 'safe',  'label': "Transcription:", 'value': self.get_transcription(instance),
                 'field_key': 'newfulltext', 'order':9},
                {'type': 'bold',  'label': "Moved to:",   'value': instance.get_moved_code(), 'empty': 'hidenone', 'link': instance.get_moved_url(),'order': 12},
                {'type': 'bold',  'label': "Previous:",   'value': instance.get_previous_code(), 'empty': 'hidenone', 'link': instance.get_previous_url(),'order': 13},
                {'type': 'line', 'label': "Project:",     'value': instance.get_project_markdown2(), 'order':11},                
                 ]},            
            
            {'name': 'User contributions', 'id': 'equalgold_usercontributions', 'fields': [
                {'type': 'plain', 'label': "Keywords (user): ", 'value': self.get_userkeywords(instance, profile, context), 
                 'field_list': 'ukwlist', 'title': 'User-specific keywords. If the moderator accepts these, they move to regular keywords.'}, # plain?
                {'type': 'line',  'label': "Personal datasets:", 'value': instance.get_collections_markdown(username, team_group, settype="pd"), 
                    'multiple': True, 'field_list': 'collist_ssg', 'fso': self.formset_objects[0] },
                ]},
            
            {'name': 'Connections', 'id': 'equalgold_connections', 'fields': [                
                {'type': 'line',  'label': "Equality set:", 'title': 'The gold sermons in this equality set',  'value': self.get_goldset_markdown(instance), 
                    'field_list': 'goldlist', 'inline_selection': 'ru.passim.sg_template' },                
                {'type': 'line',  'label': "Historical collections:",   'value': instance.get_collections_markdown(username, team_group, settype="hc"), 
                    'field_list': 'collist_hist', 'fso': self.formset_objects[0] },
                {'type': 'line',    'label': "Links:",  'title': "Authority file links:",  'value': instance.get_superlinks_markdown(), 
                    'multiple': True,  'field_list': 'superlist',       'fso': self.formset_objects[1], 
                    'inline_selection': 'ru.passim.ssg2ssg_template',   'template_selection': 'ru.passim.ssg_template'},
                  ]},           
                 ]

            context['mainsections'] += [            
                {'name': 'Networks', 'id': 'equalgold_networks', 'button': True, 'fields': [ 
                    ]},
                {'name': 'Graphs', 'id': 'equalgold_graphs',  'button': True,'fields': [ 
                    ]},
                {'name': 'Manifestations', 'id': 'equalgold_manifestations', 'button': True, 'fields': [ 
                    ]},
                      ] 
                     
            # Some tests can only be performed if this is *not* a new instance
            if not instance is None and not instance.id is None:
                pending = approval_pending(instance)
                if user_is_ingroup(self.request, app_editor) and pending.count() > 0:
                    context['approval_pending'] = pending
                    context['approval_pending_list'] = approval_pending_list(instance)
                    context['mainitems'].append(dict(
                        type='safe', label='', value=render_to_string('seeker/pending_changes.html', context, self.request)))
                
                # Special processing for those with editing rights
                if may_edit_project(self.request, profile, instance):
                    for oMainSection in context['mainsections']: 
                    # Of course the items belowd need to be added only to the 'Details' section
                        if oMainSection['name'] == "Details": 
                            oMainItems = oMainSection['fields']
                                
                            # NOTE: this code should be extended to include other external projects, when the time is right
                            ext = EqualGoldExternal.objects.filter(equal=instance).first()
                            if not ext is None:
                                # Get the field values
                                externaltype = ext.externaltype
                                externalid = ext.externalid
                                # Issue #730.3
                                # bEditHuwa = profile.is_project_approver("huwa")
                                bEditHuwa = profile.is_project_editor("huwa")
                                if externaltype == "huwop" and externalid > 0 and bEditHuwa:
                                    # Okay, the user has the right to see the externalid                                    
                                    oItem = dict(type="plain", 
                                            label="HUWA id", 
                                            value=externalid,
                                            field_key="externalid"
                                            )
                                    context['mainitems'].insert(0, oItem) 
                                  
                                    # The user also has the right to see EDITION information
                                    if not instance.edinote is None and instance.edinote[0] == "[":
                                        # Read the list of literature references
                                        lst_edilit = json.loads(instance.edinote)
                                        context['edilits'] = lst_edilit
                                        value = render_to_string('seeker/huwa_edilit.html', context, self.request)
                                        oItem = dict(type="safe", 
                                                        label="HUWA editions:", 
                                                        value=value, 
                                                        order=5)                                            
                                        oMainSection['fields'].append(oItem)

                            # Any editor can add a Transcription file
                            oItem = dict(type="plain", 
                                            label="Transcription file:",
                                            title="Add a transcription file to this Authority file",
                                            value=instance.get_trans_file(),
                                            field_key='transcription',
                                            order=10)                             
                            oMainSection['fields'].append(oItem)

                            # Make sure the spectypes are checked
                            instance.check_links()  
                        
                            # Adapt the PROJECT line in the mainitems list
                            for oItem in oMainItems:
                                if oItem['label'] == "Project:":
                                    # Add the list
                                    oItem['field_list'] = "projlist"
                                    # We can now leave from here
                                    break

                            # Any editor may suggest that an SSG be added to other project(s)
                            oItem = dict(type="plain", 
                                            label="Add to project:",
                                            title="Submit a request to add this SSG to the following project(s)",
                                            value=self.get_prj_submitted(instance, "other", profile), 
                                            order=12)
                            oItem['field_list']="addprojlist"                                
                            oMainSection['fields'].append(oItem)

                            # Any editor may suggest that an SSG be deleted from particular project(s)
                            oItem = dict(type="plain", 
                                            label="Remove from project:",
                                            title="Submit a request to remove this SSG from the following project(s)",
                                            value=self.get_prj_submitted(instance, 'current'),
                                            order=13)
                            oItem['field_list']="delprojlist"                             
                            oMainSection['fields'].append(oItem)                                 
                        else:
                            pass
                
                # Now the correct sequence of items can be realized by sorting on the 'order' key in each of the 
                # dictionaries.
                for oMainSection in context['mainsections']: 
                    # Of course only for the 'Details' section...
                    if oMainSection['name'] == "Details":
                        # Sort all items
                        oMainSection['fields'] = sorted(oMainSection['fields'], key=itemgetter('order')) 
                    # All other sections can be disregarded.
                    else: 
                        pass

                # The SSG items that have a value in *moved* may not be editable
                editable = (instance.moved == None)
                if not editable:
                    self.permission = "readonly"
                    context['permission'] = self.permission

                # Add comment modal stuff
                lhtml = []
                initial = dict(otype="super", objid=instance.id, profile=profile)
                context['commentForm'] = CommentForm(initial=initial, prefix="com")
                context['comment_list'] = get_usercomments('super', instance, profile)                
                lhtml.append(render_to_string("seeker/comment_add.html", context, self.request))
                context['comment_count'] = instance.comments.count()

                # Also add the saved item form
                lhtml.append(saveditem_form)

                # Store the after_details in the context
                context['after_details'] = "\n".join(lhtml)

            # Signal that we have select2
            context['has_select2'] = True

            # Make stable URI available
            context['stable_uri'] = self.get_abs_uri('equalgold_details', instance)

            # Test if the code to "add a new sermon gold" may be safely added or not
            if instance.moved is None:
                self.new_button = True
                context['new_button'] = True
                # SPecification of the new button
                context['new_button_title'] = "Sermon Gold"
                context['new_button_name'] = "gold"
                context['new_button_url'] = reverse("gold_details")
                context['new_button_params'] = [
                    {'name': 'gold-n-equal', 'value': instance.id}
                    ]
            else:
                self.new_button = False
                context['new_button'] = False
        except:
            msg = oErr.get_error_message()
            oErr.DoError("EqualGoldEdit/add_to_context")

        # Return the context we have made
        return context

    def get_transcription(self, instance):
        """Make a good visualization of a transcription, which includes a show/hide button"""

        sBack = ""
        oErr = ErrHandle()
        try:
            # Get the text
            if not instance.fulltext is None and instance.fulltext != "":
                sText = instance.get_fulltext_markdown("actual", lowercase=False)
                # Get URL of deleting transcription
                transdelurl = reverse('equalgold_transdel', kwargs={'pk': instance.id})
                sInfo = instance.fullinfo
                oInfo = {}
                if not sInfo is None:
                    oInfo = json.loads(sInfo)
                wordcount = oInfo.get("wordcount", 0)
                # Combine with button click + default hidden
                context = dict(delete_permission=user_is_ingroup(self.request, stemma_editor),
                               delete_message="",
                               wordcount=wordcount,
                               transdelurl=transdelurl,
                               fulltext=sText)
                sBack = render_to_string("seeker/ftext_buttons.html", context, None)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("EqualGoldEdit/get_transcription")
        return sBack

    def get_goldset_html(goldlist):
        context = {}
        template_name = 'seeker/super_goldset.html'
        sBack = ""
        if goldlist != None:
            # Add to context
            context['goldlist'] = SermonGold.objects.filter(id__in=goldlist).order_by('siglist')
            context['is_app_editor'] = False
            context['object_id'] = None
            # Calculate with the template
            sBack = render_to_string(template_name, context)
        return sBack

    def get_goldset_markdown(self, instance):
        context = {}
        template_name = 'seeker/super_goldset.html'
        sBack = ""
        if instance:
            # Add to context
            context['goldlist'] = instance.equal_goldsermons.all().order_by('siglist')
            context['is_app_editor'] = user_is_ingroup(self.request, app_editor)
            context['object_id'] = instance.id
            # Calculate with the template
            sBack = render_to_string(template_name, context)
        return sBack

    def process_formset(self, prefix, request, formset):
        errors = []
        bResult = True
        instance = formset.instance
        # Need to know who is 'talking'...
        username = self.request.user.username
        profile = Profile.get_user_profile(username)
        for form in formset:
            if form.is_valid():
                oErr = ErrHandle()
                try:
                    cleaned = form.cleaned_data
                    # Action depends on prefix
                
                    # Note: eqgcol can be either settype 'pd' or 'hc'
                    if prefix == "eqgcol":
                        # Keyword processing
                        if 'newcol' in cleaned and cleaned['newcol'] != "":
                            newcol = cleaned['newcol']
                            # Is the COL already existing?
                            obj = Collection.objects.filter(name=newcol).first()
                            if obj == None:
                                # TODO: add profile here
                                profile = Profile.get_user_profile(request.user.username)
                                obj = Collection.objects.create(name=newcol, type='super', owner=profile)
                            # Make sure we set the keyword
                            form.instance.collection = obj
                            # Note: it will get saved with formset.save()
                    elif prefix == "ssglink":
                        # SermonDescr-To-EqualGold processing
                        newsuper = cleaned.get("newsuper")
                        if not newsuper is None:
                            # There also must be a linktype
                            if 'newlinktype' in cleaned and cleaned['newlinktype'] != "":
                                linktype = cleaned['newlinktype']
                                # Get optional parameters
                                note = cleaned.get('note', None)
                                spectype = cleaned.get('newspectype', None)
                                # Alternatives: this is true if it is in there, and false otherwise
                                alternatives = cleaned.get("newalt", None)
                                # Check existence
                                obj = EqualGoldLink.objects.filter(src=instance, dst=newsuper, linktype=linktype).first()
                                if obj == None:
                                    super = EqualGold.objects.filter(id=newsuper.id).first()
                                    if super != None:
                                        # See if this can be accepted right away or needs waiting
                                        new_data = dict(linktype=linktype, note=note, spectype=spectype, alternatives=alternatives, dst=super.id)
                                        iCount = approval_parse_formset(profile, prefix, new_data, instance)

                                        # Only proceed if changes don't need to be reviewed by others
                                        if iCount == 0:

                                            # Set the right parameters for creation later on
                                            form.instance.linktype = linktype
                                            form.instance.dst = super
                                            if note != None and note != "": 
                                                form.instance.note = note
                                            if spectype != None and len(spectype) > 1:
                                                form.instance.spectype = spectype
                                            form.instance.alternatives = alternatives

                                            # Double check reverse
                                            if linktype in LINK_BIDIR:
                                                rev_link = EqualGoldLink.objects.filter(src=super, dst=instance).first()
                                                if rev_link == None:
                                                    # Add it
                                                    rev_link = EqualGoldLink.objects.create(src=super, dst=instance, linktype=linktype)
                                                else:
                                                    # Double check the linktype
                                                    if rev_link.linktype != linktype:
                                                        rev_link.linktype = linktype
                                                if note != None and note != "": 
                                                    rev_link.note = note
                                                if spectype != None and len(spectype) > 1:
                                                    rev_link.spectype = get_reverse_spec(spectype)
                                                rev_link.alternatives = alternatives
                                                rev_link.save()
                                        else:
                                            # Make sure this one does not get saved!
                                            setattr(form, 'do_not_save', True)
                        # Note: it will get saved with form.save()
                    elif prefix == "eqgbref":
                        # Processing one BibRange
                        newintro = cleaned.get('newintro', None)
                        onebook = cleaned.get('onebook', None)
                        newchvs = cleaned.get('newchvs', None)
                        newadded = cleaned.get('newadded', None)

                        # Minimal need is BOOK
                        if onebook != None:
                            # Note: normally it will get saved with formset.save()
                            #       However, 'noinit=False' formsets must arrange their own saving
                            form.instance.book = onebook
                            if newchvs != None:
                                form.instance.chvslist = newchvs
                            form.instance.intro = newintro
                            form.instance.added = newadded

                except:
                    msg = oErr.get_error_message()
                    oErr.DoError("EqualGoldEdit/process_formset")
            else:
                errors.append(form.errors)
                bResult = False
        return None

    def get_form_kwargs(self, prefix):
        # This is for ssglink

        oBack = None
        if prefix == "ssglink":
            if self.object != None:
                # Make sure to return the ID of the EqualGold
                oBack = dict(super_id=self.object.id)

        return oBack

    def get_prj_submitted(self, instance, type=None, profile=None):
        """Get an HTML list of projects to which this SSG has already been submitted"""

        oErr = ErrHandle()
        sBack = ""
        try:
            # Determine which projects should be shown
            if type is None:
                # Get the list of EqualAdd objects (with atype ['def', 'mod'], i.e. not yet accepted)
                qs = addapproval_pending(instance)
            elif type == "current":
                # Show the projects currently connected to this AF
                qs = instance.projects.all()
            elif type == "other":
                # Show the list of projects to which I am not an approver
                #      and to which this AF has not been attached yet
                approval_project_ids = [x['id'] for x in profile.projects.all().values('id')]
                current_project_ids = [x['id'] for x in instance.projects.all().values('id')]
                qs = Project2.objects.exclude(id__in=approval_project_ids).exclude(id__in=current_project_ids)

            lHtml = []
            # Visit all project items
            for obj in qs:
                if type is None:
                    project = obj.project
                else:
                    project = obj
                # Determine where clicking should lead to
                url = "{}?ssg-projlist={}".format(reverse('equalgold_list'), project.id) 
                # Create a display for this topic
                lHtml.append("<span class='project'><a href='{}'>{}</a></span>".format(url, project.name))
            sBack = ", ".join(lHtml)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("EqualGoldEdit/get_prj_submitted")
            bBack = ""

        return sBack

    def get_userkeywords(self, instance, profile, context):
        sBack = ""
        oErr = ErrHandle()
        try:
            # At the very least get the list of user keywords
            sKeywordList = instance.get_keywords_user_markdown(profile)
            # Now determine what interface to show
            if context['is_app_editor']:
                # We simply return the list
                sBack = sKeywordList
            else:
                # We need to construct an interface:
                # (1) Set up the context to include view and edit elements
                context['userkeywords'] = sKeywordList
                context['ukwform'] = context['{}Form'.format(self.prefix)]
                context['ukwupdate'] = reverse('equalgold_ukw', kwargs={'pk': instance.id })

                # Combine and provide the code needed
                # sBack = get_userkeyword_interface(self.request, context)
                sBack = render_to_string("seeker/userkeyword_edit.html", context, self.request)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("get_userkeywords")
        return sBack

    def before_delete(self, instance):
        bResult = True
        msg = ""
        oErr = ErrHandle()
        try:
            # Who is this?
            profile = Profile.get_user_profile(self.request.user.username)
            # Check if deleting may take place
            qs = instance.projects.all()
            if qs.count() > 1:
                # Get the list of projects
                projlist = qs
                iCountNeeded = approval_parse_deleting(profile, projlist, instance) 
                if iCountNeeded == 0:
                    # The SSG may be deleted right away
                    pass
                else:
                    # It may not be delete
                    bResult = False
                    msg = "ok"
            else:
                # There is no problem deleting this SSG, since it is only attached to one project
                pass
        except:
            msg = oErr.get_error_message()
            bResult = False
        return bResult, msg
           
    def before_save(self, form, instance):
        oErr = ErrHandle()
        bBack = True
        msg = ""
        transfer_changes = [
            {'src': 'newincipit',  'dst': 'incipit',  'type': 'text'},
            {'src': 'newexplicit', 'dst': 'explicit', 'type': 'text'},
            {'src': 'newfulltext', 'dst': 'fulltext', 'type': 'text'},
            {'src': 'newauthor',   'dst': 'author',   'type': 'fk'},
            ]
        try:
            if instance != None:
                # Need to know who is 'talking'...
                username = self.request.user.username
                profile = Profile.get_user_profile(username)

                ## Check for author
                #if instance.author == None:
                #    # Set to "undecided" author if possible
                #    author = Author.get_undecided()
                #    instance.author = author

                # Get the cleaned data: this is the new stuff
                cleaned_data = form.cleaned_data

                # See if and how many changes are suggested
                iCount, bNeedReload = approval_parse_changes(profile, cleaned_data, instance)
                if bNeedReload:
                    # Signal that we need to have a re-load
                    self.bNeedReload = True

                # Only proceed if changes don't need to be reviewed by others
                if iCount == 0:

                    # This means that any changes may be implemented right away
                    for oTransfer in transfer_changes:
                        type = oTransfer.get("type")
                        src_field = oTransfer.get("src")
                        dst_field = oTransfer.get("dst")
                        src_value = cleaned_data.get(src_field)
                        
                        # Transfer the value
                        if type == "fk" or type == "text":
                            # Is there any change?
                            prev_value = getattr(instance, dst_field)
                            if src_value != prev_value:
                                # Special cases
                                if dst_field == "author":
                                    authornameLC = instance.author.name.lower()
                                    # Determine what to do in terms of 'moved'.
                                    if authornameLC != "undecided":
                                        # Create a copy of the object I used to be
                                        moved = EqualGold.create_moved(instance)

                                # Perform the actual change
                                setattr(form.instance, dst_field, src_value)

                    # Check for author
                    if instance.author == None:
                        # Set to "undecided" author if possible
                        author = Author.get_undecided()
                        instance.author = author

                    # Issue #473: automatic assignment of project for particular editor(s)
                    projlist = form.cleaned_data.get("projlist")
                    bBack, msg = evaluate_projlist(profile, instance, projlist, "Authority File")

                else:
                    # The changes may *NOT* be committed
                    msg = None   # "The suggested changes will be reviewed by the other projects' editors"
                    bBack = False

                    # Changes may not be commited: reset the changes in the transfer_changes formfields
                    for oTransfer in transfer_changes:
                        type = oTransfer.get("type")
                        src_field = oTransfer.get("src")
                        dst_field = oTransfer.get("dst")
                        key_reset = "{}-{}".format(form.prefix, src_field)
                        value_reset = None
                        if type == "text":
                            value_reset = getattr(instance, dst_field)
                        elif dst_field == "author":
                            value_reset = str(instance.author.id)
                        if value_reset != None:
                            form.data[key_reset] = value_reset

                    ## The author gets a special treatment: [newauthor] should equal [author]
                    #key_newauthor = '{}-newauthor'.format(form.prefix)
                    #form.data[key_newauthor] = str(instance.author.id)

                    # NOTE (EK): the following (redirection) is no longer needed, since all the changes are shown in the EDIT view
                    ## Make sure redirection takes place
                    #self.redirect_to = reverse('equalgold_details', kwargs={'pk': instance.id})
        except:
            msg = oErr.get_error_message()
            oErr.DoError("EqualGoldEdit/before_save")
            bBack = False
        return bBack, msg

    def after_save(self, form, instance):
        """Actions to be undertaken after the EqualGold (=SSG/AF) proper has been saved"""

        # Definition of functions that are used inside [after_save()]
        def process_proj_adding(profile, projlist):
            oErr = ErrHandle()

            try:
                allow_adding = []
                iCountAdd = approval_parse_adding(profile, projlist, instance, allow_adding) 
                if len(allow_adding) > 0:
                    # Some combinations of Project-SSG may be added right away
                    with transaction.atomic():
                        for oItem in allow_adding:
                            equal = oItem.get("super")
                            project = oItem.get("project")
                            if not equal is None and not project is None:
                                obj = EqualGoldProject.objects.create(equal=equal, project=project)
            except:
                msg = oErr.get_error_message()
                oErr.DoError("process_proj_adding")
            return True

        def process_proj_removing(profile, projlist):
            oErr = ErrHandle()

            try:
                if instance.projects.count() > 1:
                    allow_removing = []
                    iCountAddB = approval_parse_removing(profile, projlist, instance, allow_removing) 
                    if len(allow_removing) > 0:
                        # There are some project-SSG associations that may be removed right away
                        delete_id = []
                        for oItem in allow_removing:
                            equal = oItem.get("super")
                            project = oItem.get("project")
                            if not equal is None and not project is None:
                                obj = EqualGoldProject.objects.filter(equal=equal, project=project).first()
                                delete_id.append(obj.id)
                        # Now delete them
                        if len(delete_id) > 0:
                            EqualGoldProject.objects.delete(delete_id)
            except:
                msg = oErr.get_error_message()
                oErr.DoError("process_proj_removing")
            return True

        def get_proj_add_remove(projlist):
            """Split the list of projects into those being added and those being removed"""
            oErr = ErrHandle()
            addprojlist = []
            delprojlist = []
            try:
                projcurrent = instance.projects.all()
                for obj in projlist:
                    # Check for addition
                    if obj in projlist and not obj in projcurrent:
                        addprojlist.append(obj)
                    elif obj in projcurrent and not obj in projlist:
                        delprojlist.append(obj)
            except:
                msg = oErr.get_error_message()
                oErr.DoError("get_proj_add_remove")
            return addprojlist, delprojlist

        def adapt_external(field, externaltype):
            """Save the externalid"""

            oErr = ErrHandle()
            try:
                externalid = form.cleaned_data.get(field)
                if not externalid is None and externalid != "":
                    # Check if this is the correct id
                    ext = instance.equalexternals.filter(externaltype=externaltype).first()
                    if ext is None:
                        # Need to make one
                        ext = EqualGoldExternal.objects.create(equal=instance, externaltype=externaltype, externalid=externalid)
                    else:
                        # Do we need changing?
                        if externalid != ext.externalid:
                            ext.externalid = externalid
                            ext.save()
            except:
                msg = oErr.get_error_message()
                oErr.DoError("adapt_external")
                bResult = False

        # Initialisations
        msg = ""
        bResult = True
        oErr = ErrHandle()
                
        try:
            # Need to know who I am for some operations
            profile = Profile.get_user_profile(self.request.user.username)

            # External id changes
            adapt_external("externalid", "huwop")

            # ====== Process many-to-many changes: Add and remove relations in accordance with the new set passed on by the user
            # (1) 'Personal Datasets' and 'Historical Collections'
            collist_ssg_id = [x['id'] for x in form.cleaned_data['collist_ssg'].values('id') ]
            collist_hist_id = [x['id'] for x in form.cleaned_data['collist_hist'].values('id')]
            collist_others_id = [x['id'] for x in instance.collections.filter(scope="priv", type="super").exclude(owner=profile).values('id')]
            collist_id = collist_ssg_id + collist_hist_id + collist_others_id
            collist_ssg = Collection.objects.filter(Q(id__in=collist_id))
            adapt_m2m(CollectionSuper, instance, "super", collist_ssg, "collection")

            # (2) links from one SSG to another SSG
            superlist = form.cleaned_data['superlist']
            super_added = []
            super_deleted = []
            adapt_m2m(EqualGoldLink, instance, "src", superlist, "dst", 
                      extra = ['linktype', 'alternatives', 'spectype', 'note'], related_is_through=True,
                      added=super_added, deleted=super_deleted)
            # Check for partial links in 'deleted'
            for obj in super_deleted:
                # This if-clause is not needed: anything that needs deletion should be deleted
                # if obj.linktype in LINK_BIDIR:
                # First find and remove the other link
                reverse = EqualGoldLink.objects.filter(src=obj.dst, dst=obj.src, linktype=obj.linktype).first()
                if reverse != None:
                    reverse.delete()
                # Then remove myself
                obj.delete()
            # Make sure to add the reverse link in the bidirectionals
            for obj in super_added:
                if obj.linktype in LINK_BIDIR:
                    # Find the reversal
                    reverse = EqualGoldLink.objects.filter(src=obj.dst, dst=obj.src, linktype=obj.linktype).first()
                    if reverse == None:
                        # Create the reversal 
                        reverse = EqualGoldLink.objects.create(src=obj.dst, dst=obj.src, linktype=obj.linktype)
                        # Other adaptations
                        bNeedSaving = False
                        # Set the correct 'reverse' spec type
                        if obj.spectype != None and obj.spectype != "":
                          reverse.spectype = get_reverse_spec(obj.spectype)
                          bNeedSaving = True
                        # Possibly copy note
                        if obj.note != None and obj.note != "":
                          reverse.note = obj.note
                          bNeedSaving = True
                        # Need saving? Then save
                        if bNeedSaving:
                          reverse.save()

            # (3) 'keywords'
            kwlist = form.cleaned_data['kwlist']
            adapt_m2m(EqualGoldKeyword, instance, "equal", kwlist, "keyword")

            # (4) user-specific 'keywords'
            ukwlist = form.cleaned_data['ukwlist']
            adapt_m2m(UserKeyword, instance, "super", ukwlist, "keyword", qfilter = {'profile': profile}, 
                      extrargs = {'profile': profile, 'type': 'super'})

            # (6) 'projects'
            projlist = form.cleaned_data['projlist']
            # === EK: CODE PRIOR TO #517 =====
            #equal_proj_deleted = []
            #adapt_m2m(EqualGoldProject, instance, "equal", projlist, "project", deleted=equal_proj_deleted)
            #project_dependant_delete(self.request, equal_proj_deleted)
            # ================================
            # Figure out what are the proposed additions, and what are the proposed deletions
            addprojlist, delprojlist = get_proj_add_remove(projlist)
            process_proj_adding(profile, addprojlist)
            process_proj_removing(profile, delprojlist)

            # Issue #517: submit request to add this SSG to indicated project(s)
            # Process the line "Add to a project"
            addprojlist = form.cleaned_data.get("addprojlist")
            process_proj_adding(profile, addprojlist)

            # Process the line "Remove from a project"
            delprojlist = form.cleaned_data.get("delprojlist")
            process_proj_removing(profile, delprojlist)

            # Issue #473: default project assignment
            if instance.projects.count() == 0:
                # Need to know who is 'talking'...
                username = self.request.user.username
                profile = Profile.get_user_profile(username)

                # The user has not selected a project (yet): try default assignment
                user_projects = profile.project_approver.filter(status="incl")
                if user_projects.count() == 1:
                    obj = user_projects.first()
                    project = obj.project
                    # project = user_projects.first()
                    EqualGoldProject.objects.create(equal=instance, project=project)
                    # Make sure the atype is set correctly
                    instance.atype = "acc"
                    instance.save()
                else:
                    # Look for projects where I am editor of
                    editor_projects = profile.get_editor_projects()
                    if editor_projects.count() == 1:
                        # Get this project
                        project = editor_projects.first()
                        # Assign the project to this SSG/AF
                        EqualGoldProject.objects.create(equal=instance, project=project)
                        # Make sure the atype is set correctly
                        instance.atype = "acc"
                        instance.save()

            # ====== Process many-to-ONE changes
            # (1) links from SG to SSG
            goldlist = form.cleaned_data['goldlist']
            ssglist = [x.equal for x in goldlist]
            adapt_m2o(SermonGold, instance, "equal", goldlist)
            # Adapt the SSGs needed
            for ssg in ssglist:
                # Adapt the SG count value
                ssg.set_sgcount()
                # Adapt the 'firstsig' value
                ssg.set_firstsig()

            # (2) links from BibRange to SSG
            bibreflist = form.cleaned_data['bibreflist']
            adapt_m2o(BibRange, instance, "equal", bibreflist)
            
            # Possibly read transcription if this is 'new'
            if 'transcription' in form.changed_data:
                # Try to read the updated transcription
                oTranscription = read_transcription(instance.transcription)
                # Are we okay?
                sStatus = oTranscription.get("status", "")
                sText = oTranscription.get("text", "")
                iWordcount = oTranscription.get("wordcount", 0)
                oFullInfo = dict(wordcount=iWordcount)
                if sStatus == "ok" and sText != "":
                    instance.fullinfo = json.dumps(oFullInfo)
                    instance.fulltext = sText
                    instance.save()


            # ADDED Take over any data from [instance] to [frm.data]
            #       Provided these fields are in the form's [initial_fields]
            if instance != None:

                # Walk the fields that need to be taken from the instance
                for key in form.initial_fields:
                    value = getattr(instance, key)

                    key_prf = '{}-{}'.format(form.prefix, key)
                    if isinstance(value, str) or isinstance(value, int):
                        form.data[key_prf] = value
                    elif isinstance(value, object):
                        form.data[key_prf] = str(value.id)

        except:
            msg = oErr.get_error_message()
            bResult = False
        return bResult, msg

    def action_add(self, instance, details, actiontype):
        """User can fill this in to his/her liking"""
        passim_action_add(self, instance, details, actiontype)

    def get_history(self, instance):
        return passim_get_history(instance)


class EqualGoldDetails(EqualGoldEdit):
    rtype = "html"

    def add_to_context(self, context, instance):
        """Add to the existing context"""

        # One general item is the 'help-popup' to be shown when the user clicks on 'Author'
        info = render_to_string('seeker/author_info.html')
        # Need to know who this user (profile) is
        profile = Profile.get_user_profile(self.request.user.username)
        username = profile.user.username
        team_group = app_editor

        # Start by executing the standard handling
        context = super(EqualGoldDetails, self).add_to_context(context, instance)

        # Sections: Networks / Graphs / Manifestations

        context['sections'] = [
                    
            {'name': 'Networks', 'id': 'equalgold_networks', 'nobutton': True, 'fields': [ 
                ], 'template': 'seeker/af_networks.html'},
            {'name': 'Graphs', 'id': 'equalgold_graphs',  'nobutton': True,'fields': [ 
                ], 'template': 'seeker/af_graphs.html'},
            {'name': 'Manifestations', 'id': 'equalgold_manifestations', 'nobutton': True, 'fields': [ 
                ], 'template': 'seeker/af_manifestations.html'},
                  ] 
       
        # Use the 'graph' function or not?
        use_network_graph = True

        oErr = ErrHandle()
        try:

            # One general item is the 'help-popup' to be shown when the user clicks on 'Author'
            info = render_to_string('seeker/author_info.html')

            # Are we copying information?? (only allowed if we are the app_editor)
            if 'goldcopy' in self.qd and context['is_app_editor']:
                # Get the ID of the gold sermon from which information is to be copied to the SSG
                goldid = self.qd['goldcopy']
                # Also get the simple value
                simple = self.qd.get("simple", "f")
                # Get the GOLD SERMON instance
                gold = SermonGold.objects.filter(id=goldid).first()

                if gold != None:
                    # Copy all relevant information to the EqualGold obj (which as a SSG)
                    obj = self.object
                    # Issue #313: <Anonymus> *may* be copied, but not <Undecided>
                    exclude_authors = ['Undecided']  # ['Anonymus','Undecided']
                    if simple == "f" and gold.author != None and not gold.author.name in exclude_authors:
                        # (1) copy author - only if not simple
                        if gold.author != None: obj.author = gold.author
                    # (2) copy incipit
                    if gold.incipit != None and gold.incipit != "": obj.incipit = gold.incipit ; obj.srchincipit = gold.srchincipit
                    # (3) copy explicit
                    if gold.explicit != None and gold.explicit != "": obj.explicit = gold.explicit ; obj.srchexplicit = gold.srchexplicit

                    # Now save the adapted EqualGold obj
                    obj.save()

                    # Mark these changes, which are done outside the normal 'form' system
                    actiontype = "save"
                    changes = dict(incipit=obj.incipit, explicit=obj.explicit)
                    if simple == "f" and obj.author != None:
                        changes['author'] = obj.author.id
                    details = dict(savetype="change", id=obj.id, changes=changes)
                    passim_action_add(self, obj, details, actiontype)

                # And in all cases: make sure we redirect to the 'clean' GET page
                self.redirectpage = reverse('equalgold_details', kwargs={'pk': self.object.id})
            
            elif instance != None and instance.id != None and instance.moved is None:
                
                # Lists of related objects
                related_objects = []

                resizable = True
                index = 1 
                sort_start = '<span class="sortable"><span class="fa fa-sort sortshow"></span>&nbsp;'
                sort_start_int = '<span class="sortable integer"><span class="fa fa-sort sortshow"></span>&nbsp;'
                sort_end = '</span>'

                username = self.request.user.username
                team_group = app_editor
                profile = Profile.get_user_profile(username=username)

                # Be sure to have the 'object' containing a link to the EqualGold instance
                if context['object'] == None:
                    context['object'] = instance

                # New: Get all the SermonDescr instances linked with equality to SSG:
                # But make sure the EXCLUDE those with `mtype` = `tem`
                qs_s = SermonDescrEqual.objects.filter(super=instance).exclude(sermon__mtype="tem").order_by('sermon__msitem__manu__idno', 'sermon__locus')
                
                # Count the number of records in the qs, make a string out of it
                count = str(qs_s.count())              

                # Create title with count            
                title_count = "Total: " + count              
                
                # List of manuscripts related to the SSG via sermon descriptions
                manuscripts = dict(title= title_count, prefix="manu", gridclass="resizable") 

                # Get all the SermonDescr instances linked with equality to SSG:
                # But make sure the EXCLUDE those with `mtype` = `tem`
                qs_s = SermonDescrEqual.objects.filter(super=instance).exclude(sermon__mtype="tem").order_by('sermon__msitem__manu__idno', 'sermon__locus')
                rel_list =[]
                method = "FourColumns"
                method = "Issue216"
                for sermonlink in qs_s:
                    sermon = sermonlink.sermon
                    # Get the 'item': the manuscript
                    # OLD: item = sermon.manu
                    item = sermon.msitem.manu
                    codico = sermon.msitem.codico
                    rel_item = []
                
                    if method == "FourColumns":
                        # Name as CITY - LIBRARY - IDNO + Name
                        manu_name = "{}, {}, <span class='signature'>{}</span> {}".format(item.library.lcity.name, item.library.name, item.idno, item.name)
                        rel_item.append({'value': manu_name, 'title': item.idno, 'main': True,
                                         'link': reverse('manuscript_details', kwargs={'pk': item.id})})

                        # Location number and link to the correct point in the manuscript details view...
                        itemloc = "{}/{}".format(sermon.order, item.get_sermon_count())
                        rel_item.append({'value': itemloc, 'align': "right", 'title': 'Jump to the sermon in the manuscript',
                                         'link': "{}#sermon_{}".format(reverse('manuscript_details', kwargs={'pk': item.id}), sermon.id)  })

                        # date range
                        daterange = "{}-{}".format(item.yearstart, item.yearfinish)
                        rel_item.append({'value': daterange, 'align': "right"})

                        # Sermon Locus + Title + link
                        sermo_name = "<span class='signature'>{}</span>".format(sermon.locus)
                        rel_item.append({'value': sermon.locus, 'title': sermo_name, 
                                         'link': reverse('sermon_details', kwargs={'pk': sermon.id})})
                    elif method == "Issue216":
                        # Shelfmark = IDNO
                        # manu_full = "{}, {}, {}".format(item.get_city(), item.get_library(), item.idno)
                        manu_full = item.get_full_name()
                        # issue #610
                        # value = item.idno
                        value = manu_full
                        manu_name = "<span class='signature' title='{}'>{}</span>".format(manu_full, value)
                        # Name as CITY - LIBRARY - IDNO + Name
                        manu_name = "{}, {}, <span class='signature'>{}</span> {}".format(item.get_city(), item.get_library(), item.idno, item.name)
                        rel_item.append({'value': manu_name, 'title': item.idno, 'main': False, 'myclasses': 'shelfm',
                                         'link': reverse('manuscript_details', kwargs={'pk': item.id})}) # 'initial': 'small',

                        # Origin: this should actually be the origin of the CODICO in which the sermon is
                        #         and the provenance should also be based on the CODICO
                        or_prov = "{} ({})".format(codico.get_origin(), codico.get_provenance_markdown(table=False))
                        rel_item.append({'value': or_prov, 
                                         'title': "Origin (if known), followed by provenances (between brackets)",  'myclasses': 'orprov'}) #, 'initial': 'small'})

                        # Date range
                        daterange = "{}-{}".format(item.yearstart, item.yearfinish)
                        rel_item.append({'value': daterange, 'align': "right", 'myclasses': 'date'}) #, 'initial': 'small'})

                        # Collection(s)
                        #coll_info = item.get_collections_markdown(username, team_group)
                        #rel_item.append({'value': coll_info, 'myclasses': 'coll'}) # , 'initial': 'small'  / 

                        # Location number and link to the correct point in the manuscript details view...
                        itemloc = "{}/{}".format(sermon.msitem.order, item.get_sermon_count())
                        link_on_manu_page = "{}#sermon_{}".format(reverse('manuscript_details', kwargs={'pk': item.id}), sermon.id)
                        link_to_sermon = reverse('sermon_details', kwargs={'pk': sermon.id})
                        rel_item.append({'value': itemloc, 'myclasses': 'item', 'align': "right", 'title': 'Jump to the sermon in the manuscript', 'link': link_to_sermon }) #'initial': 'small'

                        # Folio number of the item
                        rel_item.append({'value': sermon.locus, 'myclasses': 'ff'}) # , 'initial': 'small'

                        # Attributed author
                        rel_item.append({'value': sermon.get_author(), 'myclasses': 'attraut'}) # , 'initial': 'small'

                        # Title 
                        rel_item.append({'value': sermon.title, 'myclasses': 'title'}) # Goed plaatsen!

                        # Incipit
                        rel_item.append({'value': sermon.get_incipit_markdown(), 'myclasses': 'inc' }) #, 'initial': 'small'})

                        # Explicit
                        rel_item.append({'value': sermon.get_explicit_markdown(), 'myclasses': 'expl'}) #, 'initial': 'small'})

                        # Postscriptum
                        rel_item.append({'value': sermon.get_postscriptum_markdown(), 'myclasses': 'post_scriptum'}) # Goed plaatsen!

                        # Bible reference 
                        rel_item.append({'value': sermon.get_bibleref(), 'myclasses': 'bible_ref '}) # Goed plaatsen!

                        # Feast 
                        rel_item.append({'value': sermon.get_feast(), 'myclasses': 'feast_manif'}) # Goed plaatsen!

                        # Keywords
                        rel_item.append({'value': sermon.get_keywords_markdown(), 'myclasses': 'keyw'}) # , 'initial': 'small'
                        
                    # Add this Manu/Sermon line to the list
                    rel_list.append(dict(id=item.id, cols=rel_item))
                manuscripts['rel_list'] = rel_list

                # Make the number of Sermons available (confusingly called 'manuscripts')
                context['manuscripts'] = qs_s.count()

                if method == "FourColumns":
                    manuscripts['columns'] = ['Manuscript', 'Items', 'Date range', 'Sermon manifestation']
                elif method == "Issue216":
                    manuscripts['columns'] = [                                              
                        '{}<span title="Shelfmark">Shelfmark.</span>{}'.format(sort_start, sort_end),
                        '{}<span title="Origin/Provenance">or./prov.</span>{}'.format(sort_start, sort_end),                         
                        '{}<span title="Date range">date</span>{}'.format(sort_start, sort_end),                        
                       #'{}<span title="Collection name">coll.</span>{}'.format(sort_start, sort_end),
                        '{}<span title="Item">item</span>{}'.format(sort_start, sort_end),
                        '{}<span title="Folio number">ff</span>{}'.format(sort_start, sort_end),
                        '{}<span title="Attributed author">attr. auth.</span>{}'.format(sort_start, sort_end),
                        '{}<span title="Title">title</span>{}'.format(sort_start, sort_end),
                        '{}<span title="Incipit">inc.</span>{}'.format(sort_start, sort_end),
                        '{}<span title="Explicit">expl.</span>{}'.format(sort_start, sort_end),
                        '{}<span title="Postscriptum">postscr.</span>{}'.format(sort_start, sort_end),
                        '{}<span title="Bible reference">bib. ref.</span>{}'.format(sort_start, sort_end),
                        '{}<span title="Feast">feast</span>{}'.format(sort_start, sort_end),
                        '{}<span title="Keywords of the Sermon manifestation">keyw.</span>{}'.format(sort_start, sort_end),
                        ]

                # Add the manuscript to the related objects
                related_objects.append(manuscripts)               
                                
                context['related_objects_af'] = related_objects

                # Old version, without this the table is not shown automatically, only when clicking on a button
                #context['related_objects'] = related_objects

                # ======== VISUALIZATION PREPARATION ===============================
                # (see seeker/visualizations)

                # Make sure to delete any previous corpora of mine
                EqualGoldCorpus.objects.filter(profile=profile, ssg=instance).delete()

                # Old, extinct
                ManuscriptCorpus.objects.filter(super=instance).delete()
                ManuscriptCorpusLock.objects.filter(profile=profile, super=instance).delete()

                # THe graph also needs room in after details
                if use_network_graph:
                    # Prepare the CONTEXT for network-related graphs
                    context['equalgold_graph'] = reverse("equalgold_graph", kwargs={'pk': instance.id})
                    context['equalgold_trans'] = reverse("equalgold_trans", kwargs={'pk': instance.id})
                    context['equalgold_overlap'] = reverse("equalgold_overlap", kwargs={'pk': instance.id})

                # Graph that is always there: PCA
                context['equalgold_pca'] = reverse("equalgold_pca", kwargs={'pk': instance.id})

                # Prepare context for attributed author pie chart
                context['equalgold_attr'] = reverse("equalgold_attr", kwargs={'pk': instance.id})

                # Prepare context for origin pie chart
                context['equalgold_origin'] = reverse("equalgold_origin", kwargs={'pk': instance.id})

                # Prepare context for chronological codico date line chart
                context['equalgold_chrono'] = reverse("equalgold_chrono", kwargs={'pk': instance.id})

                lHtml = []
                if 'after_details' in context:
                    lHtml.append(context['after_details'])

                # Note (EK): this must be here, see issue #508 OLD
                #lHtml.append(render_to_string('seeker/super_graph.html', context, self.request))

                context['after_details'] = "\n".join(lHtml)

                # ===================================================================

        except:
            msg = oErr.get_error_message()
            oErr.DoError("EqualGoldDetails/add_to_context")

        # Return the context we have made
        return context

    def before_save(self, form, instance):
        oErr = ErrHandle()
        bBack = True
        msg = ""
        try:
            self.isnew = False
            if not instance is None:
                if instance.id is None:
                    # This is a new SSG being created.
                    # Provide standard stuff:
                    instance.author = Author.get_undecided()
                    instance.stype = STYPE_MANUAL
                    self.isnew = True
        except:
            msg = oErr.get_error_message()
            oErr.DoError("EqualGoldDetails/before_save")
            bBack = False
        return bBack, msg

    def process_formset(self, prefix, request, formset):
        return None

    def after_save(self, form, instance):
        oErr = ErrHandle()
        bBack = True
        msg = ""
        method = "before_issue742"
        method = "after_issue742"
        try:
            if self.isnew:
                # Try default project assignment
                profile = Profile.get_user_profile(self.request.user.username)

                if method == "before_issue742":
                    qs = profile.project_approver.filter(status="incl")
                    for obj in qs:
                        EqualGoldProject.objects.create(project=obj.project, equal=instance)
                    # is it just one project?
                    if qs.count() == 1:
                        # Make sure the atype is set correctly
                        instance.atype = "acc"
                        instance.save()
                else:
                    # Method is "after_issue742"
                    # Determine what the projects are to which the user has default editing rights.
                    qs = profile.project_editor.filter(status="incl")
                    for obj in qs:
                        EqualGoldProject.objects.create(project=obj.project, equal=instance)
                    # Make sure the atype is set correctly
                    instance.atype = "acc"
                    instance.save()
        except:
            msg = oErr.get_error_message()
            oErr.DoError("EqualGoldDetails/before_save")
            bBack = False
        return bBack, msg


class EqualGoldTransDel(EqualGoldDetails):
    """Remove the fulltrans from this SSG"""

    initRedirect = True

    def custom_init(self, instance):
        oErr = ErrHandle()
        try:
            # Check ... something
            if not instance is None:
                # Make sure to set the correct redirect page
                self.redirectpage = reverse("equalgold_details", kwargs={'pk': instance.id})
                # Remove the transcription
                instance.srchfulltext = None
                instance.fulltext = None
                instance.transcription = None
                instance.save()
                # Now it's all been deleted
        except:
            msg = oErr.get_error_message()
            oErr.DoError("EqualGoldTransDel/custom_init")
        return None


class EqualGoldUserKeyword(EqualGoldDetails):
    """This version only looks at one kind of data: the ukwlist"""

    newRedirect = True

    def custom_init(self, instance):
        oErr = ErrHandle()
        try:
            # Make sure to set the correct redirect page
            if instance:
                self.redirectpage = reverse("equalgold_details", kwargs={'pk': instance.id})
            # Make sure we are not saving
            self.do_not_save = True

            # Get the value
            qd = self.qd
            if self.prefix_type == "simple":
                sUkwList = "{}-ukwlist".format(self.prefix)
            else:
                sUkwList = "{}-{}-ukwlist".format(self.prefix, instance.id)
            ukwlist_ids = qd.getlist(sUkwList)
            ukwlist = Keyword.objects.filter(id__in=ukwlist_ids)
            
            # Process the contents of the ukwlist, the chosen user-keywords
            profile = Profile.get_user_profile(self.request.user.username)
            adapt_m2m(UserKeyword, instance, "super", ukwlist, "keyword", qfilter = {'profile': profile}, extrargs = {'profile': profile, 'type': 'super'})

        except:
            msg = oErr.get_error_message()
            oErr.DoError("EqualGoldUserKeyword/custom_init")
    

class EqualGoldListView(BasicList):
    """List super sermon gold instances"""

    model = EqualGold
    listform = SuperSermonGoldForm
    has_select2 = True  # Check
    use_team_group = True    
    prefix = "ssg"
    sel_button = "ssg"
    bUseFilter = True  
    plural_name = "Authority Files"
    sg_name = "Authority File"
    order_cols = ['author__name', 'code', 'firstsig', 'srchincipit', '', 'scount', 'sgcount', 'ssgcount', 'hccount', 'stype'] 
    order_default= ['code', 'author__name', 'firstsig', 'srchincipit', '', 'scount', 'sgcount', 'ssgcount', 'hccount', 'stype']
    order_heads = [
        {'name': 'Author',                  'order': 'o=1', 'type': 'str', 'custom': 'author', 'linkdetails': True},

        # Issue #212: remove sermon number from listview
        # {'name': 'Number',                  'order': 'o=2', 'type': 'int', 'custom': 'number', 'linkdetails': True},

        {'name': 'Code',                    'order': 'o=2', 'type': 'str', 'custom': 'code',   'linkdetails': True},
        {'name': 'Gryson/Clavis',           'order': 'o=3', 'type': 'str', 'custom': 'sig', 'allowwrap': True, 'options': "abcd",
         'title': "The Gryson/Clavis codes of all the Sermons Gold in this equality set"},
        {'name': 'Incipit ... Explicit',    'order': 'o=4', 'type': 'str', 'custom': 'incexpl', 'main': True, 'linkdetails': True,
         'title': "The incipit...explicit that has been chosen for this Authority file"},        
        {'name': '',                        'order': '',    'type': 'str', 'custom': 'saved',   'align': 'right'},
        {'name': 'Manifestations',          'order': 'o=6'   , 'type': 'int', 'custom': 'scount',
         'title': "Number of Sermon (manifestation)s that are connected with this Authority file"},
        {'name': 'Contains',                'order': 'o=7'   , 'type': 'int', 'custom': 'size',
         'title': "Number of Sermons Gold that are part of the equality set of this Authority file"},
        {'name': 'Linked AFs',              'order': 'o=8'   , 'type': 'int', 'custom': 'ssgcount',
         'title': "Number of other Authority files this Authority file links to"},
        {'name': 'Hist. Coll.',             'order': 'o=9'   , 'type': 'int', 'custom': 'hccount',
         'title': "Number of historical collections associated with this Authority file"},
        {'name': 'Status',                  'order': 'o=10',   'type': 'str', 'custom': 'status'}        
        ]
    filters = [
        {"name": "Author",          "id": "filter_author",            "enabled": False},
        {"name": "Incipit",         "id": "filter_incipit",           "enabled": False},
        {"name": "Explicit",        "id": "filter_explicit",          "enabled": False},
        {"name": "Passim code",     "id": "filter_code",              "enabled": False},
        #{"name": "Number",         "id": "filter_number",            "enabled": False},
        {"name": "Gryson/Clavis/Other code","id": "filter_signature", "enabled": False},
        {"name": "Keyword",         "id": "filter_keyword",           "enabled": False},
        {"name": "Status",          "id": "filter_stype",             "enabled": False},
        {"name": "Manifestation count", "id": "filter_scount",        "enabled": False},
        {"name": "AF relation count","id": "filter_ssgcount",         "enabled": False},
        {"name": "Project",         "id": "filter_project",           "enabled": False},        
        {"name": "Transcription",   "id": "filter_transcr",           "enabled": False},        
        {"name": "Linked via manuscript","id": "filter_manuscript",   "enabled": False},        
        {"name": "Collection/Dataset...","id": "filter_collection",   "enabled": False, "head_id": "none"},
        {"name": "Manuscript",      "id": "filter_collmanu",          "enabled": False, "head_id": "filter_collection"},
        {"name": "Sermon",          "id": "filter_collsermo",         "enabled": False, "head_id": "filter_collection"},
        {"name": "Sermon Gold",     "id": "filter_collgold",          "enabled": False, "head_id": "filter_collection"},
        {"name": "Authority file",  "id": "filter_collsuper",         "enabled": False, "head_id": "filter_collection"},
        {"name": "Historical",      "id": "filter_collhist",          "enabled": False, "head_id": "filter_collection"},
               ]
    searches = [
        {'section': '', 'filterlist': [
            {'filter': 'incipit',   'dbfield': 'srchincipit',       'keyS': 'incipit',  'regex': adapt_regex_incexp},
            {'filter': 'explicit',  'dbfield': 'srchexplicit',      'keyS': 'explicit', 'regex': adapt_regex_incexp},
            
            {'filter': 'code',      'dbfield': 'code',              'keyS': 'code',     'help': 'passimcode',
             'keyList': 'passimlist', 'infield': 'id'},
           
           # {'filter': 'number',    'dbfield': 'number',            'keyS': 'number',
           #  'title': 'The per-author-sermon-number (these numbers are assigned automatically and have no significance)'},
            {'filter': 'scount',    'dbfield': 'soperator',         'keyS': 'soperator'         },
            {'filter': 'scount',    'dbfield': 'scount',            'keyS': 'scount',
             'title': 'The number of sermons (manifestations) belonging to this Authority file' },
            {'filter': 'transcr',   'dbfield': 'foperator',         'keyS': 'foperator'         },
            {'filter': 'transcr',   'dbfield': 'srchfulltext',      'keyS': 'srchfulltext'      },
            {'filter': 'ssgcount',  'dbfield': 'ssgoperator',       'keyS': 'ssgoperator'       },
            {'filter': 'ssgcount',  'dbfield': 'ssgcount',          'keyS': 'ssgcount',
             'title': 'The number of links an Authority file has to other Authority files'      },
            {'filter': 'manuscript','dbfield': '$dummy',            'keyS': 'manulink'          }, 
            {'filter': 'keyword',   'fkfield': 'keywords',          'keyFk': 'name', 'keyList': 'kwlist', 'infield': 'id'},
            {'filter': 'author',    'fkfield': 'author',            
             'keyS': 'authorname', 'keyFk': 'name', 'keyList': 'authorlist', 'infield': 'id', 'external': 'gold-authorname' },
            {'filter': 'stype',     'dbfield': 'stype',             'keyList': 'stypelist', 'keyType': 'fieldchoice', 'infield': 'abbr' },
            {'filter': 'signature', 'fkfield': 'equal_goldsermons__goldsignatures',    'help': 'signature',
             'keyS': 'signature', 'keyFk': 'code', 'keyId': 'signatureid', 'keyList': 'siglist', 'infield': 'code'},
            {'filter': 'project',   'fkfield': 'projects',   'keyFk': 'name', 'keyList': 'lstprojlist', 'infield': 'name'},
            ]},
        {'section': 'collection', 'filterlist': [
            {'filter': 'collmanu',  'fkfield': 'equal_goldsermons__sermondescr__manu__collections',  
             'keyS': 'collection','keyFk': 'name', 'keyList': 'collist_m', 'infield': 'name' }, 
            {'filter': 'collsermo', 'fkfield': 'equalgold_sermons__sermondescr_col__collection',        
            # issue #466: fkfield was 'equal_goldsermons__sermondescr__collections'
            #             changed into 'equalgold_sermons__sermondescr_col__collection'
             'keyS': 'collection','keyFk': 'name', 'keyList': 'collist_s', 'infield': 'name' }, 
            {'filter': 'collgold',  'fkfield': 'equal_goldsermons__collections',                     
             'keyS': 'collection','keyFk': 'name', 'keyList': 'collist_sg', 'infield': 'name' }, 
            {'filter': 'collsuper', 'fkfield': 'collections',                                        
             'keyS': 'collection','keyFk': 'name', 'keyList': 'collist_ssg', 'infield': 'name' }, 
            {'filter': 'collhist', 'fkfield': 'collections',                                        
             'keyS': 'collection','keyFk': 'name', 'keyList': 'collist_hist', 'infield': 'name' }
            ]},
        {'section': 'other', 'filterlist': [
            {'filter': 'atype', 'dbfield': 'atype', 'keyS': 'atype'}
            ]}
        ]
    custombuttons = [{"name": "scount_histogram", "title": "Sermon Histogram", 
                      "icon": "th-list", "template_name": "seeker/scount_histogram.html" }]

    selectbuttons = [
        {'title': 'Add to saved items', 'mode': 'add_saveitem', 'button': 'jumbo-1', 'glyphicon': 'glyphicon-star-empty'},
        {'title': 'Add to basket',      'mode': 'add_basket',   'button': 'jumbo-1', 'glyphicon': 'glyphicon-shopping-cart'},
        # {'title': 'Add to DCT',         'mode': 'add_dct',      'button': 'jumbo-1', 'glyphicon': 'glyphicon-wrench'},
        ]

    def initializations(self):

        # ======== One-time adaptations ==============
        listview_adaptations("equalgold_list")

        # Should json be added?
        if user_is_superuser(self.request):
            self.downloads = [
                {"label": "Huwa AFs: json",         "dtype": "json", "url": 'equalgold_huwajson'},
                {"label": "Huwa AFs: csv",          "dtype": "csv",  "url": 'equalgold_huwajson'},
                {"label": "Huwa AFs: Excel",        "dtype": "xlsx", "url": 'equalgold_huwajson'},
                {"label": "Huwa Literature: json",  "dtype": "json", "url": 'equalgold_huwalitjson'},
                {"label": "Huwa BHL etc: json",     "dtype": "json", "url": 'equalgold_huwaopera'},
                {"label": "Huwa retractationes: json",  "dtype": "json", "url": 'equalgold_huwaretr'},
                {"label": "Huwa indiculum: json",   "dtype": "json", "url": 'equalgold_huwaindiculum'},
                {"label": "Huwa nebenwerk: json",   "dtype": "json", "url": 'equalgold_huwanebenwerk'}
                ]
            # Possibly add to 'uploads'
            bHasJson = False
            bHasEdilit = False
            bHasTrans = False
            for item in self.uploads:
                if item['title'] == "huwajson":
                    bHasJson = True
                elif item['title'] == "huwaedilit":
                    bHasEdilit = True

            # Should json be added?
            if not bHasJson:
                # Add a reference to the Json upload method
                html = []
                html.append("Upload SSGs/AFs from a HUWA json specification.")
                html.append("<b>Note 1:</b> this may OVERWRITE existing information!")
                html.append("<b>Note 2:</b> PROJECT assignment is according to issue #534")
                msg = "<br />".join(html)
                oJson = dict(title="huwajson", label="HUWA json",
                              url=reverse('import_huwa'),
                              type="multiple", msg=msg)
                self.uploads.append(oJson)
            if not bHasEdilit:
                # Add a reference to the Edilit JSON upload method
                msg = "Upload EdiLit from a HUWA edilit json specification."
                oJson = dict(title="huwaedilit", label="HUWA edilit",
                              url=reverse('equalgold_upload_edilit'),
                              type="multiple", msg=msg)
                self.uploads.append(oJson)

            if not bHasTrans:
                # Add a reference to the XML transcription upload method
                html = []
                msg = "Upload sermon transcription(s) from TEI-P5 xml file(s)"
                oJson = dict(title="xtranseqg", label="XML transcription", 
                             url = reverse('import_trans_eqg'),
                             type='multiple', msg=msg)
                self.uploads.append(oJson)


        # Make the profile available
        self.profile = Profile.get_user_profile(self.request.user.username)

        return None
    
    def add_to_context(self, context, initial):
        # Find out who the user is
        profile = Profile.get_user_profile(self.request.user.username)
        context['basketsize'] = 0 if profile == None else profile.basketsize_super
        context['basket_show'] = reverse('basket_show_super')
        context['basket_update'] = reverse('basket_update_super')
        context['histogram_data'] = self.get_histogram_data('d3')

        # Count the number of selected items
        iCount = SelectItem.get_selectcount(profile, "ssg")
        if iCount > 0:
            context['sel_count'] = str(iCount)

        return context

    def get_histogram_data(self, method='d3'):
        """Get data to make a histogram"""

        oErr = ErrHandle()
        histogram_data = []
        b_chart = None
        try:
            # Get the base url
            baseurl = reverse('equalgold_list')
            # Get the queryset for this view
            qs = self.get_queryset().order_by('scount').values('scount', 'id')
            scount_index = {}
            frequency = None
            for item in qs:
                scount = item['scount']
                if frequency == None or frequency != scount:
                    frequency = scount
                    histogram_data.append(dict(scount=scount, freq=1))
                else:
                    histogram_data[-1]['freq'] += 1

            # Determine the targeturl for each histogram bar
            other_list = []
            for item in self.param_list:
                if "-soperator" not in item and "-scount" not in item:
                    other_list.append(item)
            other_filters = "&".join(other_list)
            for item in histogram_data:
                targeturl = "{}?ssg-soperator=exact&ssg-scount={}".format(baseurl, item['scount'], other_filters)
                item['targeturl'] = targeturl

            if method == "d3":
                histogram_data = json.dumps(histogram_data)
            
        except:
            msg = oErr.get_error_message()
            oErr.DoError("get_histogram_data")
        return histogram_data
        
    def get_basketqueryset(self):
        if self.basketview:
            profile = Profile.get_user_profile(self.request.user.username)
            qs = profile.basketitems_super.all()
        else:
            qs = EqualGold.objects.all()
        return qs

    def get_field_value(self, instance, custom):
        sBack = ""
        sTitle = ""
        html = []
        if custom == "author": 
            # Get a good name for the author
            if instance.author:
                html.append(instance.author.name)
            else:
                html.append("<i>(not specified)</i>")
        elif custom == "size":
            iSize = instance.sgcount
            html.append("{}".format(iSize))
        elif custom == "scount":
            sCount = instance.scount
            if sCount == None: sCount = 0
            html.append("{}".format(sCount))
        elif custom == "ssgcount":
            sCount = instance.ssgcount
            if sCount == None: sCount = 0
            html.append("{}".format(sCount))
        elif custom == "hccount": 
            hCount = instance.hccount
            if hCount == None: hCount = 0            
            html.append("{}".format(hCount)) 
        elif custom == "hclist":
            html.append(instance.get_hclist_markdown())
        elif custom == "code":
            sCode = "-" if instance.code  == None else instance.code
            html.append("{}".format(sCode))
        elif custom == "incexpl":
            html.append("<span>{}</span>".format(instance.get_incipit_markdown()))
            dots = "..." if instance.incipit else ""
            html.append("<span style='color: blue;'>{}</span>".format(dots))
            html.append("<span>{}</span>".format(instance.get_explicit_markdown()))
        elif custom == "sig":           
            # The prefered sequence of codes (Gryson, Clavis, Other)
            editype_pref_seq = ['gr', 'cl', 'ot'] 
            # Use the prefered sequence of codes 
            for editype in editype_pref_seq: 
            # Visit all signatures for each Authority File               
                qs = Signature.objects.filter(gold__equal=instance, editype=editype).order_by('code')
                # Iterate over all signatures (for each editype)
                for sig in qs:
                    editype = sig.editype
                    url = "{}?gold-siglist={}".format(reverse("gold_list"), sig.id)
                    short = sig.short()
                    html.append("<span class='badge signature {}' title='{}'><a class='nostyle' href='{}'>{}</a></span>".format(editype, short, url, short[:20]))           
        elif custom == "saved":
            # Prepare saveditem handling
            saveditem_button = get_saveditem_html(self.request, instance, self.profile, sitemtype="ssg")
            saveditem_form = get_saveditem_html(self.request, instance, self.profile, "form", sitemtype="ssg")
            html.append(saveditem_button)
            html.append(saveditem_form)
        elif custom == "status":
            # Provide the status traffic light
            html.append(instance.get_stype_light())

        sBack = "\n".join(html) 
        return sBack, sTitle

    def adapt_search(self, fields):
        # Adapt the search to the keywords that *may* be shown
        lstExclude= None
        qAlternative = None
        oErr = ErrHandle()

        try:
            # Check if a list of keywords is given
            if 'kwlist' in fields and fields['kwlist'] != None and len(fields['kwlist']) > 0:
                # Get the list
                kwlist = fields['kwlist']
                # Get the user
                username = self.request.user.username
                user = User.objects.filter(username=username).first()
                # Check on what kind of user I am
                if not user_is_ingroup(self.request, app_editor):
                    # Since I am not an app-editor, I may not filter on keywords that have visibility 'edi'
                    kwlist = Keyword.objects.filter(id__in=kwlist).exclude(Q(visibility="edi")).values('id')
                    fields['kwlist'] = kwlist

            scount = fields.get('scount', -1)
            soperator = fields.pop('soperator', None)
            if scount != None and scount >= 0 and soperator != None:
                # Action depends on the operator
                fields['scount'] = Q(**{"scount__{}".format(soperator): scount})

            ssgcount = fields.get('ssgcount', -1)
            ssgoperator = fields.pop('ssgoperator', None)
            if ssgcount != None and ssgcount >= 0 and ssgoperator != None:
                # Action depends on the operator
                fields['ssgcount'] = Q(**{"ssgcount__{}".format(ssgoperator): ssgcount})

            # Look for manulink
            manulink = fields.get('manulink')
            if not manulink is None:
                # We have one Manuscript that serves as a basis to search for AFs
                qThis = Q(sermondescr_super__manu=manulink)
                fields['manulink'] = qThis

            foperator = fields.get('foperator')
            srchfulltext = fields.get('srchfulltext')
            if not foperator is None and foperator != "":
                # Choices are: "obl" and "txt"
                if foperator == "obl":
                    # Just check whether a text is there or not
                    fields['srchfulltext'] = Q(transcription__isnull=False) & ~ Q(transcription="")
                else:
                    # Well, if we are really looking for the full text, we don't have to do anything
                    pass
            elif not srchfulltext is None and srchfulltext != "":
                # This just means that we are really looking for the full text
                pass
        
            # Make sure we only show the SSG/AF's that have accepted modifications
            # (fields['atype'] = 'acc'), so exclude the others:
            lstExclude = [ Q(atype__in=['mod', 'def', 'rej']) | Q(moved__isnull=False) ]     
        except:
            msg = oErr.get_error_message()
            oErr.DoError("EqualGoldListView/adapt_search")
       
        return fields, lstExclude, qAlternative        

    def view_queryset(self, qs):
        search_id = [x['id'] for x in qs.values('id')]
        profile = Profile.get_user_profile(self.request.user.username)
        profile.search_super = json.dumps(search_id)
        profile.save()
        return None

    def get_helptext(self, name):
        """Use the get_helptext function defined in models.py"""
        return get_helptext(name)

    def get_selectitem_info(self, instance, context):
        """Use the get_selectitem_info() defined earlier in this views.py"""
        return get_selectitem_info(self.request, instance, self.profile, self.sel_button, context)
        

class EqualGoldScountDownload(BasicPart):
    MainModel = EqualGold
    template_name = "seeker/download_status.html"
    action = "download"
    dtype = "csv"       # downloadtype

    def custom_init(self):
        """Calculate stuff"""
        
        dt = self.qd.get('downloadtype', "")
        if dt != None and dt != '':
            self.dtype = dt

    def get_queryset(self, prefix):

        # Construct the QS
        #qs = TestsetUnit.objects.all().order_by('testset__round', 'testset__number').values(
        #    'testset__round', 'testset__number', 'testunit__speaker__name', 'testunit__fname',
        #    'testunit__sentence__name', 'testunit__ntype', 'testunit__speaker__gender')
        qs = EqualGold.objects.none()
        #TODO: this should be corrected

        return qs

    def get_data(self, prefix, dtype, response=None):
        """Gather the data as CSV, including a header line and comma-separated"""

        # Initialize
        lData = []
        sData = ""

        if dtype == "json":
            # Loop over all round/number combinations (testsets)
            for obj in self.get_queryset(prefix):
                round = obj.get('testset__round')               # obj.testset.round
                number = obj.get('testset__number')             # obj.testset.number
                speaker = obj.get('testunit__speaker__name')    # obj.testunit.speaker.name
                gender = obj.get('testunit__speaker__gender')   # obj.testunit.speaker.gender
                sentence = obj.get('testunit__sentence__name')  # obj.testunit.sentence.name
                ntype = obj.get('testunit__ntype')              # obj.testunit.ntype
                fname = obj.get('testunit__fname')              # Pre-calculated filename
                row = dict(round=round, testset=number, speaker=speaker, gender=gender,
                    filename=fname, sentence=sentence, ntype=ntype)
                lData.append(row)
            # convert to string
            sData = json.dumps(lData, indent=2)
        elif dtype == "csv" or dtype == "xlsx":
            # Create CSV string writer
            output = StringIO()
            delimiter = "\t" if dtype == "csv" else ","
            csvwriter = csv.writer(output, delimiter=delimiter, quotechar='"')
            # Headers
            headers = ['round', 'testset', 'speaker', 'gender', 'filename', 'sentence', 'ntype']
            csvwriter.writerow(headers)
            for obj in self.get_queryset(prefix):
                round = obj.get('testset__round')                # obj.testset.round
                number = obj.get('testset__number')             # obj.testset.number
                speaker = obj.get('testunit__speaker__name')    # obj.testunit.speaker.name
                gender = obj.get('testunit__speaker__gender')   # obj.testunit.speaker.gender
                sentence = obj.get('testunit__sentence__name')  # obj.testunit.sentence.name
                fname = obj.get('testunit__fname')              # Pre-calculated filename
                ntype = obj.get('testunit__ntype')              # obj.testunit.ntype
                row = [round, number, speaker, gender, fname, sentence, ntype]
                csvwriter.writerow(row)

            # Convert to string
            sData = output.getvalue()
            output.close()

        return sData


class EqualGoldVisDownload(BasicPart):
    """Generic treatment of Visualization downloads for SSGs"""

    MainModel = EqualGold
    template_name = "seeker/download_status.html"
    action = "download"
    dtype = "hist-svg"
    vistype = ""

    def custom_init(self):
        """Calculate stuff"""
        
        dt = self.qd.get('downloadtype', "")
        if dt != None and dt != '':
            self.dtype = dt

    def get_data(self, prefix, dtype, response=None):
        """Gather the data as CSV, including a header line and comma-separated"""

        # Initialize
        lData = []
        sData = ""
        oErr = ErrHandle()

        try:
            if dtype == "json":
                # Retrieve the actual data from self.data
                oData = dict(legend=self.data['legend'],
                             link_list=self.data['link_list'],
                             node_list=self.data['node_list'])
                sData = json.dumps(oData, indent=2)
            elif dtype == "hist-svg":
                pass
            elif dtype == "hist-png":
                pass
            elif dtype == "csv":
                # THIS IS NOT YET IMPLEMENTED

                # Create CSV string writer
                output = StringIO()
                delimiter = "\t" if dtype == "csv" else ","
                csvwriter = csv.writer(output, delimiter=delimiter, quotechar='"')
                # Headers
                headers = ['round', 'testset', 'speaker', 'gender', 'filename', 'sentence', 'ntype']
                csvwriter.writerow(headers)

                # Convert to string
                sData = output.getvalue()
                output.close()
            elif dtype == "excel":

                if self.vistype in ['trans', 'overlap']:
                    # Start workbook
                    wb = openpyxl.Workbook()

                    # First worksheet: Name of the legend
                    ws = wb.get_active_sheet()
                    ws.title="Data"
                    c = ws.cell(row=1, column=1)
                    c.value = "Legend"
                    c = ws.cell(row=1, column=2)
                    c.value = self.data['legend']

                    # Second sheet: node data
                    ws = wb.create_sheet("Nodes")
                    node_list=self.data['node_list']

                    # (2): headers
                    if self.vistype == "overlap":
                        headers = ["group", "HCs", "id", "label", "passim", "scount"]
                    elif self.vistype == "trans":
                        headers = ["id", "scount", "label", "category", "rating", "sig"]
                    for col_num in range(len(headers)):
                        c = ws.cell(row=1, column=col_num+1)
                        c.value = headers[col_num]
                        c.font = openpyxl.styles.Font(bold=True)
                        # Set width to a fixed size
                        ws.column_dimensions[get_column_letter(col_num+1)].width = 5.0    
                    
                    # (2): items   
                    row_num = 2 
                    for item in node_list:
                        if self.vistype == "overlap":
                            ws.cell(row=row_num, column=1).value = item.get("group")
                            ws.cell(row=row_num, column=2).value = json.dumps(item.get("hcs"))
                            ws.cell(row=row_num, column=3).value = item.get("id")
                            ws.cell(row=row_num, column=4).value = item.get("label")
                            ws.cell(row=row_num, column=5).value = item.get("passim")
                            ws.cell(row=row_num, column=6).value = item.get("scount")
                        elif self.vistype == "trans":
                            ws.cell(row=row_num, column=1).value = item.get("id")
                            ws.cell(row=row_num, column=2).value = item.get("scount")
                            ws.cell(row=row_num, column=3).value = item.get("label")
                            ws.cell(row=row_num, column=4).value = item.get("category")
                            ws.cell(row=row_num, column=5).value = item.get("rating")
                            ws.cell(row=row_num, column=6).value = item.get("sig")

                        row_num += 1

                    # Third sheet: link data
                    ws = wb.create_sheet("Links")
                    link_list=self.data['link_list']
                
                    # (3): headers
                    if self.vistype == "overlap":
                        headers = ["value", "spectype", "spec", "linktype", "source", "target", "alternatives", "link", "note"]
                    elif self.vistype == "trans":
                        headers = ["source", "target", "value"]
                    for col_num in range(len(headers)):
                        c = ws.cell(row=1, column=col_num+1)
                        c.value = headers[col_num]
                        c.font = openpyxl.styles.Font(bold=True)
                        # Set width to a fixed size
                        ws.column_dimensions[get_column_letter(col_num+1)].width = 5.0        

                    # (3): items   
                    row_num = 2 
                    for item in link_list:
                        if self.vistype == "overlap":
                            ws.cell(row=row_num, column=1).value = item.get("value")
                            ws.cell(row=row_num, column=2).value = item.get("spectype")
                            ws.cell(row=row_num, column=3).value = item.get("spec")
                            ws.cell(row=row_num, column=4).value = item.get("linktype")
                            ws.cell(row=row_num, column=5).value = item.get("source")
                            ws.cell(row=row_num, column=6).value = item.get("target")
                            ws.cell(row=row_num, column=7).value = item.get("alternatives")
                            ws.cell(row=row_num, column=8).value = item.get("link")
                            ws.cell(row=row_num, column=9).value = item.get("note")
                        elif self.vistype == "trans":
                            ws.cell(row=row_num, column=1).value = item.get("source")
                            ws.cell(row=row_num, column=2).value = item.get("target")
                            ws.cell(row=row_num, column=3).value = item.get("value")
                        row_num += 1

                # Save it
                wb.save(response)
                sData = response
        except:
            msg = oErr.get_error_message()
            oErr.DoError("EqualGoldVisDownload/get_data")

        return sData


class EqualGoldGraphDownload(EqualGoldVisDownload):
    """Network graph"""
    vistype = "graph"


class EqualGoldTransDownload(EqualGoldVisDownload):
    """Transmission graph"""
    vistype = "trans"


class EqualGoldOverlapDownload(EqualGoldVisDownload):
    """Overlap graph"""
    vistype = "overlap"


# ================= MANUSCRIPTLINK ========================

class ManuscriptLinkEdit(BasicDetails):
    model = ManuscriptLink
    mForm = ManuscriptLinkForm
    prefix = 'mlink'
    title = "Manuscript link"
    new_button = False
    mainitems = []
    use_team_group = False
    history_button = False

    def custom_init(self, instance):
        # Is this an instance?
        if not instance is None and not instance.src is None:
            # Figure out what the SRC is
            src = instance.src
            # Figure out what the details form for this M is
            self.listview = reverse('manuscript_details', kwargs={'pk': src.id})
            self.listviewtitle = "Manuscript"
        return None

    def add_to_context(self, context, instance):
        """Add to the existing context"""

        # Define the main items to show and edit
        src_id = None if instance.src is None else instance.src.id
        dst_id = None if instance.dst is None else instance.dst.id
        context['mainitems'] = [
            # -------- HIDDEN field values ---------------
            {'type': 'plain', 'label': "Source id",     'value': src_id,        'field_key': "src", 'empty': 'hide'},
            {'type': 'plain', 'label': "Target id",     'value': dst_id,        'field_key': "dst", 'empty': 'hide'},
            # --------------------------------------------
            {'type': 'safe',  'label': "Source manuscript:",    'value': instance.src.get_view(True) },
            {'type': 'safe',  'label': "Target manuscript:",    'value': instance.dst.get_view(True) },
            {'type': 'plain', 'label': "Link type:",            'value': instance.get_linktype_display(),   'field_key': 'linktype'},
            {'type': 'plain', 'label': "Note:",                 "value": instance.get_note(),               'field_key': 'note'},
            ]

        # Signal that we have select2
        context['has_select2'] = True

        # Return the context we have made
        return context


class ManuscriptLinkDetails(ManuscriptLinkEdit):
    """The Details variant of Edit"""

    rtype = "html"


class ManuscriptLinkListView(BasicList):
    """List manuscript link instances"""

    model = ManuscriptLink
    listform = ManuscriptLinkForm
    has_select2 = True  # Check
    prefix = "mlink"
    bUseFilter = True  
    plural_name = "Manuscript links"
    sg_name = "Manuscript Link"
    order_cols = ['src__idno', 'dst__idno', 'linktype'] 
    order_default= order_cols
    order_heads = [
        {'name': 'Source AF',       'order': 'o=1', 'type': 'str', 'custom': 'src',         'linkdetails': True},
        {'name': 'Target AF',       'order': 'o=2', 'type': 'str', 'custom': 'dst',         'linkdetails': True, 'main': True},
        {'name': 'Link type',       'order': 'o=3', 'type': 'str', 'custom': 'linktype',    'linkdetails': True},
        ]
    filters = [
        {"name": "Source",      "id": "filter_source",      "enabled": False},
        {"name": "Target",      "id": "filter_target",      "enabled": False},
        {"name": "Link type",   "id": "filter_linktype",    "enabled": False},
               ]
    searches = [
        {'section': '', 'filterlist': [
            {'filter': 'source',      'fkfield': 'src',   'keyS': 'source',   'keyFk': 'id', 'keyList': 'srclist', 'infield': 'id'},
            {'filter': 'target',      'fkfield': 'dst',   'keyS': 'target',   'keyFk': 'id', 'keyList': 'dstlist', 'infield': 'id'},
            {'filter': 'linktype',    'dbfield': 'linktype',    'keyList': 'linktypelist', 'keyType': 'fieldchoice', 'infield': 'abbr' },
            ]},
        ]

    def get_field_value(self, instance, custom):
        sBack = ""
        sTitle = ""
        if custom == "src":
            sBack = instance.src.get_passimcode()
        elif custom == "dst":
            sBack = instance.dst.get_passimcode()
        elif custom == "linktype":
            sBack = instance.get_linktype_display()

        return sBack, sTitle


# ================= SERMONDESCRLINK ======================

class SermonDescrLinkEdit(BasicDetails):
    model = SermonDescrLink
    mForm = SermonDescrLinkForm
    prefix = 'slink'
    title = "Manifestation link"
    new_button = False
    mainitems = []
    use_team_group = False
    history_button = False
    listview = None
    listviewtitle = None

    def custom_init(self, instance):
        # Is this an instance?
        if not instance is None and not instance.src is None:
            # Figure out what the SRC is
            src = instance.src
            # Figure out what the details form for this S is
            self.listview = reverse('sermon_details', kwargs={'pk': src.id})
            self.listviewtitle = "Manifestation"
        return None

    def add_to_context(self, context, instance):
        """Add to the existing context"""

        # Define the main items to show and edit
        src_id = None if instance.src is None else instance.src.id
        dst_id = None if instance.dst is None else instance.dst.id
        context['mainitems'] = [
            # -------- HIDDEN field values ---------------
            {'type': 'plain', 'label': "Source id",     'value': src_id,        'field_key': "src", 'empty': 'hide'},
            {'type': 'plain', 'label': "Target id",     'value': dst_id,        'field_key': "dst", 'empty': 'hide'},
            # --------------------------------------------
            {'type': 'safe',  'label': "Source sermon manifestation:",  'value': instance.src.get_view(True) },
            {'type': 'safe',  'label': "Target sermon manifestation:",  'value': instance.dst.get_view(True) },
            {'type': 'plain', 'label': "Link type:",                    'value': instance.get_linktype_display(),   'field_key': 'linktype'},
            {'type': 'plain', 'label': "Note:",                         "value": instance.get_note(),               'field_key': 'note'},
            ]

        # Signal that we have select2
        context['has_select2'] = True

        # Return the context we have made
        return context


class SermonDescrLinkDetails(SermonDescrLinkEdit):
    """The Details variant of Edit"""

    rtype = "html"


class SermonDescrLinkListView(BasicList):
    """List manuscript link instances"""

    model = SermonDescrLink
    listform = SermonDescrLinkForm
    has_select2 = True  # Check
    prefix = "slink"
    bUseFilter = True  
    plural_name = "Sermon links"
    sg_name = "Manifestation Link"
    order_cols = ['src__idno', 'dst__idno', 'linktype'] 
    order_default= order_cols
    order_heads = [
        {'name': 'Source AF',       'order': 'o=1', 'type': 'str', 'custom': 'src',         'linkdetails': True},
        {'name': 'Target AF',       'order': 'o=2', 'type': 'str', 'custom': 'dst',         'linkdetails': True, 'main': True},
        {'name': 'Link type',       'order': 'o=3', 'type': 'str', 'custom': 'linktype',    'linkdetails': True},
        ]
    filters = [
        {"name": "Source",      "id": "filter_source",      "enabled": False},
        {"name": "Target",      "id": "filter_target",      "enabled": False},
        {"name": "Link type",   "id": "filter_linktype",    "enabled": False},
               ]
    searches = [
        {'section': '', 'filterlist': [
            {'filter': 'source',      'fkfield': 'src',   'keyS': 'source',   'keyFk': 'id', 'keyList': 'srclist', 'infield': 'id'},
            {'filter': 'target',      'fkfield': 'dst',   'keyS': 'target',   'keyFk': 'id', 'keyList': 'dstlist', 'infield': 'id'},
            {'filter': 'linktype',    'dbfield': 'linktype',    'keyList': 'linktypelist', 'keyType': 'fieldchoice', 'infield': 'abbr' },
            ]},
        ]

    def get_field_value(self, instance, custom):
        sBack = ""
        sTitle = ""
        if custom == "src":
            sBack = instance.src.get_passimcode()
        elif custom == "dst":
            sBack = instance.dst.get_passimcode()
        elif custom == "linktype":
            sBack = instance.get_linktype_display()

        return sBack, sTitle


# ================= EQUALGOLDLINK ======================

class EqualGoldLinkEdit(BasicDetails):
    model = EqualGoldLink
    mForm = EqualGoldLinkForm
    prefix = 'ssglink'
    title = "Authority file link"
    new_button = False
    mainitems = []
    use_team_group = False
    history_button = False

    def custom_init(self, instance):
        # Is this an instance?
        if not instance is None and not instance.src is None:
            # Figure out what the SRC is
            src = instance.src
            # Figure out what the details form for this SSG is
            self.listview = reverse('equalgold_details', kwargs={'pk': src.id})
        return None

    def add_to_context(self, context, instance):
        """Add to the existing context"""

        # Define the main items to show and edit
        src_id = None if instance.src is None else instance.src.id
        dst_id = None if instance.dst is None else instance.dst.id
        context['mainitems'] = [
            # -------- HIDDEN field values ---------------
            {'type': 'plain', 'label': "Source id",     'value': src_id,                       'field_key': "src", 'empty': 'hide'},
            {'type': 'plain', 'label': "Target id",     'value': dst_id,                       'field_key': "dst", 'empty': 'hide'},
            # --------------------------------------------
            {'type': 'safe',  'label': "Source AF:",    'value': instance.src.get_view(True) },
            {'type': 'safe',  'label': "Target AF:",    'value': instance.dst.get_view(True) },
            {'type': 'plain', 'label': "Link type:",    'value': instance.get_linktype_display(),       'field_key': 'linktype'},
            {'type': 'plain', 'label': "Spec type:",    'value': instance.get_spectype_display(),       'field_key': 'spectype'},
            {'type': 'plain', 'label': "Alternatives:", "value": instance.get_alternatives(),           'field_key': 'alternatives'},
            {'type': 'plain', 'label': "Note:",         "value": instance.get_note(),                   'field_key': 'note'},
            ]

        # Signal that we have select2
        context['has_select2'] = True

        # Return the context we have made
        return context


class EqualGoldLinkDetails(EqualGoldLinkEdit):
    """The Details variant of Edit"""

    rtype = "html"


class EqualGoldLinkListView(BasicList):
    """List super sermon gold instances"""

    model = EqualGoldLink
    listform = EqualGoldLinkForm
    has_select2 = True  # Check
    prefix = "ssglink"
    bUseFilter = True  
    plural_name = "Authority File links"
    sg_name = "Authority File Link"
    order_cols = ['src__code', 'dst__code', 'linktype', 'spectype', 'alternatives'] 
    order_default= order_cols
    order_heads = [
        {'name': 'Source AF',       'order': 'o=1', 'type': 'str', 'custom': 'src',         'linkdetails': True},
        {'name': 'Target AF',       'order': 'o=2', 'type': 'str', 'custom': 'dst',         'linkdetails': True},
        {'name': 'Link type',       'order': 'o=3', 'type': 'str', 'custom': 'linktype',    'linkdetails': True},
        {'name': 'Spec type',       'order': 'o=4', 'type': 'str', 'custom': 'spectype',    'linkdetails': True, 'main': True},        
        {'name': 'Alternatives',    'order': 'o=5', 'type': 'str', 'custom': 'alternatives','linkdetails': True},
        ]
    filters = [
        {"name": "Source",      "id": "filter_source",      "enabled": False},
        {"name": "Target",      "id": "filter_target",      "enabled": False},
        {"name": "Link type",   "id": "filter_linktype",    "enabled": False},
        {"name": "Spec type",   "id": "filter_spectype",    "enabled": False}
               ]
    searches = [
        {'section': '', 'filterlist': [
            {'filter': 'source',      'fkfield': 'src',   'keyS': 'source',   'keyFk': 'code', 'keyList': 'srclist', 'infield': 'code'},
            {'filter': 'target',      'fkfield': 'dst',   'keyS': 'target',   'keyFk': 'code', 'keyList': 'dstlist', 'infield': 'code'},
            {'filter': 'linktype',    'dbfield': 'linktype',    'keyList': 'linktypelist', 'keyType': 'fieldchoice', 'infield': 'abbr' },
            {'filter': 'spectype',    'dbfield': 'spectype',    'keyList': 'spectypelist', 'keyType': 'fieldchoice', 'infield': 'abbr' },
            ]},
        ]

    def get_field_value(self, instance, custom):
        sBack = ""
        sTitle = ""
        if custom == "src":
            sBack = instance.src.get_passimcode()
        elif custom == "dst":
            sBack = instance.dst.get_passimcode()
        elif custom == "linktype":
            sBack = instance.get_linktype_display()
        elif custom == "spectype":
            sBack = instance.get_spectype_display()
        elif custom == "alternatives":
            sBack = instance.get_alternatives_display()

        return sBack, sTitle


# ================= AUTHOR =============================

class AuthorEdit(BasicDetails):
    """The details of one author"""

    model = Author
    mForm = AuthorEditForm
    prefix = 'author'
    title = "People"
    mainitems = []

    def add_to_context(self, context, instance):
        """Add to the existing context"""

        # Define the main items to show and edit
        context['mainitems'] = [
            {'type': 'plain', 'label': "Name:",         'value': instance.name,     'field_key': "name" },
            {'type': 'plain', 'label': "Abbreviation:", 'value': instance.abbr,     'field_key': 'abbr' },
            {'type': 'plain', 'label': "Number:",       'value': instance.number, 
             'title': "The author number is automatically assigned" },
            {'type': 'plain', 'label': "Editable:",     "value": instance.get_editable()                }
            ]

        # Signal that we have select2
        context['has_select2'] = True

        # Return the context we have made
        return context


class AuthorDetails(AuthorEdit):
    """Html variant of AuthorEdit"""

    rtype = "html"


class AuthorListView(BasicList):
    """Search and list authors"""

    model = Author
    listform = AuthorSearchForm
    has_select2 = True
    prefix = "auth"
    paginate_by = 20
    delete_line = True
    sg_name = "People"     # This is the name as it appears e.g. in "Add a new XXX" (in the basic listview)
    plural_name = "People"
    page_function = "ru.passim.seeker.search_paged_start"
    order_cols = ['abbr', 'number', 'name', '', '']
    order_default = ['name', 'abbr', 'number', '', '']
    order_heads = [{'name': 'Abbr',        'order': 'o=1', 'type': 'str', 
                    'title': 'Abbreviation of this name (used in standard literature)', 'field': 'abbr', 'default': ""},
                   {'name': 'Number',      'order': 'o=2', 'type': 'int', 'title': 'Passim author number', 'field': 'number', 'default': 10000, 'align': 'right'},
                   {'name': 'Name', 'order': 'o=3', 'type': 'str', 'field': "name", "default": "", 'main': True, 'linkdetails': True},
                   {'name': 'Links',       'order': '',    'type': 'str', 'title': 'Number of links from Sermon Descriptions and Gold Sermons', 'custom': 'links' },
                   {'name': '',            'order': '',    'type': 'str', 'options': ['delete']}]
    filters = [ {"name": "People",  "id": "filter_people",  "enabled": False}]
    searches = [
        {'section': '', 'filterlist': [
            {'filter': 'people', 'dbfield': 'name', 'keyS': 'author_ta', 'keyList': 'authlist', 'infield': 'name' }
            ]}
        ]
    downloads = [{"label": "Excel", "dtype": "xlsx", "url": 'author_results'},
                 {"label": "csv (tab-separated)", "dtype": "csv", "url": 'author_results'},
                 {"label": None},
                 {"label": "json", "dtype": "json", "url": 'author_results'}]
    uploads = [{"url": "import_authors", "label": "Authors (csv/json)", "msg": "Specify the CSV file (or the JSON file) that contains the PASSIM authors"}]

    def get_field_value(self, instance, custom):
        sBack = ""
        sTitle = ""
        if custom == "links":
            html = []
            # Get the HTML code for the links of this instance
            number = instance.author_goldsermons.count()
            if number > 0:
                url = reverse('search_gold')
                html.append("<span class='badge jumbo-2' title='linked gold sermons'>")
                html.append("<a class='nostyle' href='{}?gold-author={}'>{}</a></span>".format(url, instance.id, number))
            number = instance.author_sermons.count()
            if number > 0:
                url = reverse('sermon_list')
                html.append("<span class='badge jumbo-1' title='linked sermon descriptions'>")
                html.append("<a href='{}?sermo-author={}'>{}</a></span>".format(url, instance.id, number))           
            number = instance.author_equalgolds.count()
            if number > 0:
                url = reverse('equalgold_list')
                html.append("<span class='badge jumbo-4' title='linked authority file descriptions'>")
                html.append("<a href='{}?ssg-author={}'>{}</a></span>".format(url, instance.id, number))               

            # Combine the HTML code
            sBack = "\n".join(html)
        return sBack, sTitle


class AuthorListDownload(BasicPart):
    MainModel = Author
    template_name = "seeker/download_status.html"
    action = "download"
    dtype = "csv"       # downloadtype

    def custom_init(self):
        """Calculate stuff"""
        
        dt = self.qd.get('downloadtype', "")
        if dt != None and dt != '':
            self.dtype = dt

    def add_to_context(self, context):
        # Provide search URL and search name
        #context['search_edit_url'] = reverse("seeker_edit", kwargs={"object_id": self.basket.research.id})
        #context['search_name'] = self.basket.research.name
        return context

    def get_queryset(self, prefix):

        # Get parameters
        name = self.qd.get("name", "")

        # Construct the QS
        lstQ = []
        if name != "": lstQ.append(Q(name__iregex=adapt_search(name)))
        qs = Author.objects.filter(*lstQ).order_by('name')

        return qs

    def get_data(self, prefix, dtype, response=None):
        """Gather the data as CSV, including a header line and comma-separated"""

        # Initialize
        lData = []
        sData = ""

        if dtype == "json":
            # Loop
            for author in self.get_queryset(prefix):
                row = {"id": author.id, "name": author.name}
                lData.append(row)
            # convert to string
            sData = json.dumps(lData)
        else:
            # Create CSV string writer
            output = StringIO()
            delimiter = "\t" if dtype == "csv" else ","
            csvwriter = csv.writer(output, delimiter=delimiter, quotechar='"')
            # Headers
            headers = ['id', 'name']
            csvwriter.writerow(headers)
            # Loop
            for author in self.get_queryset(prefix):
                row = [author.id, author.name]
                csvwriter.writerow(row)

            # Convert to string
            sData = output.getvalue()
            output.close()

        return sData


# ================= LIBRARY =============================

class LibraryListView(BasicList):
    """Listview of libraries in countries/cities"""

    model = Library
    listform = LibrarySearchForm
    has_select2 = True
    prefix = "lib"
    plural_name = "Libraries"
    sg_name = "Library"
    order_cols = ['lcountry__name', 'lcity__name', 'name', 'idLibrEtab', 'mcount', '']
    order_default = order_cols
    order_heads = [
        {'name': 'Country',     'order': 'o=1', 'type': 'str', 'custom': 'country', 'default': "", 'linkdetails': True},
        {'name': 'City',        'order': 'o=2', 'type': 'str', 'custom': 'city',    'default': "", 'linkdetails': True},
        {'name': 'Library',     'order': 'o=3', 'type': 'str', 'field':  "name",    "default": "", 'main': True, 'linkdetails': True},
        {'name': 'CNRS',        'order': 'o=4', 'type': 'int', 'custom': 'cnrs',    'align': 'right' },
        {'name': 'Manuscripts', 'order': 'o=5', 'type': 'int', 'custom': 'manuscripts'},
        {'name': 'type',        'order': '',    'type': 'str', 'custom': 'libtype'},
        # {'name': '',            'order': '',    'type': 'str', 'custom': 'links'}
        ]
    filters = [ 
        {"name": "Country", "id": "filter_country", "enabled": False},
        {"name": "City",    "id": "filter_city",    "enabled": False},
        {"name": "Library", "id": "filter_library", "enabled": False}
        ]
    searches = [
        {'section': '', 'filterlist': [
            {'filter': 'country',   'fkfield': 'lcountry',  'keyList': 'countrylist',    'infield': 'name' },
            {'filter': 'city',      'fkfield': 'lcity',     'keyList': 'citylist',       'infield': 'name' },
            {'filter': 'library',   'dbfield': 'name',      'keyList': 'librarylist',    'infield': 'name', 'keyS': 'library_ta' }
            ]
         }
        ]
    downloads = [{"label": "Excel", "dtype": "xlsx", "url": 'library_results'},
                 {"label": "csv (tab-separated)", "dtype": "csv", "url": 'library_results'},
                 {"label": None},
                 {"label": "json", "dtype": "json", "url": 'library_results'}]
    uploads = [
        {"title": "huwa", "label": "Huwa (Excel)", "url": "library_upload_excel", "type": "multiple", 
                "msg": "Upload Huwa library Excel files"}]

    def initializations(self):
        oErr = ErrHandle()
        try:
            # Check if signature adaptation is needed
            mcounting = Information.get_kvalue("mcounting")
            if mcounting == None or mcounting != "done":
                # Perform adaptations
                with transaction.atomic():
                    for lib in Library.objects.all():
                        mcount = lib.library_manuscripts.count()
                        if lib.mcount != mcount:
                            lib.mcount = mcount
                            lib.save()
                # Success
                Information.set_kvalue("mcounting", "done")

        except:
            msg = oErr.get_error_message()
            oErr.DoError("LibraryListView/initializations")
        return None

    def get_field_value(self, instance, custom):
        sBack = ""
        sTitle = ""
        if custom == "country":
            if instance.lcountry != None:
                sBack = instance.lcountry.name
        elif custom == "city":
            if instance.lcity != None:
                sBack = instance.lcity.name
        elif custom == "cnrs":
            if instance.idLibrEtab >= 0:
                sBack = instance.idLibrEtab
        elif custom == "manuscripts":
            count = instance.num_manuscripts()
            if count == 0:
                sBack = "-"
            else:
                html = []
                html.append("<span>{}</span>".format(count))
                # Create the URL
                url = "{}?manu-library={}".format(reverse('search_manuscript'), instance.id)
                # Add a link to them
                html.append('<a role="button" class="btn btn-xs jumbo-3" title="Go to these manuscripts" ')
                html.append(' href="{}"><span class="glyphicon glyphicon-chevron-right"></span></a>'.format(url))
                sBack = "\n".join(html)
        elif custom == "libtype":
            if instance.libtype != "":
                sTitle = instance.get_libtype_display()
                sBack = instance.libtype
        #elif custom == "links":
        #    html = []
        #    # Get the HTML code for the links of this instance
        #    number = instance.author_goldsermons.count()
        #    if number > 0:
        #        url = reverse('search_gold')
        #        html.append("<span class='badge jumbo-2' title='linked gold sermons'>")
        #        html.append(" <a class='nostyle' href='{}?gold-author={}'>{}</a></span>".format(url, instance.id, number))
        #    number = instance.author_sermons.count()
        #    if number > 0:
        #        url = reverse('sermon_list')
        #        html.append("<span class='badge jumbo-1' title='linked sermon descriptions'>")
        #        html.append(" <a href='{}?sermo-author={}'>{}</a></span>".format(url, instance.id, number))
        #    # Combine the HTML code
        #    sBack = "\n".join(html)
        return sBack, sTitle


class LibraryListDownload(BasicPart):
    MainModel = Library
    template_name = "seeker/download_status.html"
    action = "download"
    dtype = "csv"       # downloadtype

    def custom_init(self):
        """Calculate stuff"""
        
        dt = self.qd.get('downloadtype', "")
        if dt != None and dt != '':
            self.dtype = dt

    def add_to_context(self, context):
        # Provide search URL and search name
        #context['search_edit_url'] = reverse("seeker_edit", kwargs={"object_id": self.basket.research.id})
        #context['search_name'] = self.basket.research.name
        return context

    def get_queryset(self, prefix):

        # Get parameters
        country = self.qd.get("country", "")
        city = self.qd.get("city", "")
        library = self.qd.get("library", "")

        # Construct the QS
        lstQ = []
        loc_qs = None

        # if country != "": lstQ.append(Q(country__name__iregex=adapt_search(country)))
        # if city != "": lstQ.append(Q(city__name__iregex=adapt_search(city)))

        if country != "":
            lstQ = []
            lstQ.append(Q(name__iregex=adapt_search(country)))
            lstQ.append(Q(loctype__name="country"))
            country_qs = Location.objects.filter(*lstQ)
            if city == "":
                loc_qs = country_qs
            else:
                lstQ = []
                lstQ.append(Q(name__iregex=adapt_search(city)))
                lstQ.append(Q(loctype__name="city"))
                loc_qs = Location.objects.filter(*lstQ)
        elif city != "":
            lstQ = []
            lstQ.append(Q(name__iregex=adapt_search(city)))
            lstQ.append(Q(loctype__name="city"))
            loc_qs = Location.objects.filter(*lstQ)

        lstQ = []
        if library != "": lstQ.append(Q(name__iregex=adapt_search(library)))
        if loc_qs != None: lstQ.append(Q(location__in=loc_qs))

        qs = Library.objects.filter(*lstQ).order_by('country__name', 'city__name', 'name')

        return qs

    def get_data(self, prefix, dtype, response=None):
        """Gather the data as CSV, including a header line and comma-separated"""

        # Initialize
        lData = []
        sData = ""

        if dtype == "json":
            # Loop
            with transaction.atomic():
                for lib in self.get_queryset(prefix):
                    country = ""
                    city = ""
                    if lib.country: country = lib.country.name
                    if lib.city: city = lib.city.name
                    row = {"id": lib.id, "country": lib.get_country_name(), "city": lib.get_city_name(), "library": lib.name, "libtype": lib.libtype}
                    lData.append(row)

            ## Loop
            #for oLib in self.get_queryset(prefix).values('id', 'lcountry__name', 'lcity__name', 'name', 'libtype'):
            #    country = ""
            #    city = ""
            #    if lib.country: country = lib.country.name
            #    if lib.city: city = lib.city.name
            #    row = {"id": lib.id, "country": lib.get_country_name(), "city": lib.get_city_name(), "library": lib.name, "libtype": lib.libtype}
            #    lData.append(row)
            # convert to string
            sData = json.dumps(lData)
        else:
            # Create CSV string writer
            output = StringIO()
            delimiter = "\t" if dtype == "csv" else ","
            csvwriter = csv.writer(output, delimiter=delimiter, quotechar='"')
            # Headers
            headers = ['id', 'country', 'city', 'library', 'libtype']
            csvwriter.writerow(headers)
            qs = self.get_queryset(prefix)
            if qs.count() > 0:
                # Loop
                with transaction.atomic():
                    for lib in qs:
                        row = [lib.id, lib.get_country_name(), lib.get_city_name(), lib.name, lib.libtype]
                        csvwriter.writerow(row)

            # Convert to string
            sData = output.getvalue()
            output.close()

        return sData


class LibraryEdit(BasicDetails):
    model = Library
    mForm = LibraryForm
    prefix = 'lib'
    prefix_type = "simple"
    title = "LibraryDetails"
    rtype = "json"
    history_button = True
    mainitems = []
    stype_edi_fields = ['idLibrEtab', 'name', 'libtype', 'location', 'lcity', 'lcountry']
    
    def add_to_context(self, context, instance):
        """Add to the existing context"""

        # Define the main items to show and edit
        context['mainitems'] = [
            {'type': 'plain', 'label': "Name:",                 'value': instance.name,                     'field_key': "name"},
            {'type': 'plain', 'label': "Library type:",         'value': instance.get_libtype_display(),    'field_key': 'libtype'},
            {'type': 'plain', 'label': "CNRS library id:",      'value': instance.idLibrEtab,               'field_key': "idLibrEtab"},
            {'type': 'plain', 'label': "Library location:",     "value": instance.get_location_markdown(),  'field_key': "location"},
            {'type': 'plain', 'label': "City of library:",      "value": instance.get_city_name()},
            {'type': 'plain', 'label': "Country of library: ",  "value": instance.get_country_name()}
            ]

        # Signal that we have select2
        context['has_select2'] = True

        # Return the context we have made
        return context

    def before_save(self, form, instance):
        bNeedSaving = False
        # Check whether the location has changed
        if 'location' in form.changed_data:
            # Get the new location
            location = form.cleaned_data['location']
            if location != None:
                # Get the hierarchy including myself
                hierarchy = location.hierarchy()
                for item in hierarchy:
                    if item.loctype.name == "city":
                        instance.lcity = item
                        bNeedSaving = True
                    elif item.loctype.name == "country":
                        instance.lcountry = item
                        bNeedSaving = True

        return True, ""

    def action_add(self, instance, details, actiontype):
        """User can fill this in to his/her liking"""
        passim_action_add(self, instance, details, actiontype)

    def get_history(self, instance):
        return passim_get_history(instance)
    

class LibraryDetails(LibraryEdit):
    """The full HTML version of the edit, together with French library contents"""

    rtype = "html"

    def add_to_context(self, context, instance):
        # First make sure we have the 'default' context from LibraryEdit
        context = super(LibraryDetails, self).add_to_context(context, instance)

        # Do we have a city and a library?
        city = instance.lcity
        library = instance.name
        if city != None and library != None:
            # Go and find out if there are any French connections
            sLibList = get_cnrs_manuscripts(city, library)
            sLibList = sLibList.strip()
            # Add this to 'after details'
            if sLibList != "":
                lhtml = []
                lhtml.append("<h4>Available in the CNRS library</h4>")
                lhtml.append("<div>{}</div>".format(sLibList))
                context['after_details'] = "\n".join(lhtml)

        # Return the adapted context
        return context


# ================= REPORT =============================

class ReportListView(BasicList):
    """Listview of reports"""

    model = Report
    listform = ReportEditForm
    has_select2 = True
    bUseFilter = True
    new_button = False
    basic_name = "report"
    order_cols = ['created', 'user', 'reptype', '']
    order_default = ['-created', 'user', 'reptype']
    order_heads = [{'name': 'Date', 'order': 'o=1', 'type': 'str', 'custom': 'date', 'align': 'right', 'linkdetails': True},
                   {'name': 'User', 'order': 'o=2', 'type': 'str', 'custom': 'user', 'linkdetails': True},
                   {'name': 'Type', 'order': 'o=3', 'type': 'str', 'custom': 'reptype', 'main': True, 'linkdetails': True},
                   {'name': 'Size', 'order': '',    'type': 'str', 'custom': 'size'}]
    filters = [ {"name": "User",       "id": "filter_user",      "enabled": False} ]
    searches = [
        {'section': '', 'filterlist': [
            {'filter': 'user', 'fkfield': 'user', 'keyFk': 'username', 'keyList': 'userlist', 'infield': 'id'}
            ]}
         ]

    def get_field_value(self, instance, custom):
        sBack = ""
        sTitle = ""
        if custom == "date":
            sBack = instance.created.strftime("%d/%b/%Y %H:%M")
        elif custom == "user":
            sBack = instance.user.username
        elif custom == "reptype":
            sBack = instance.get_reptype_display()
        elif custom == "size":
            # Get the total number of downloaded elements
            iSize = 0
            rep = instance.contents
            if rep != None and rep != "" and rep[0] == "{":
                oRep = json.loads(rep)
                if 'list' in oRep:
                    iSize = len(oRep['list'])
            sBack = "{}".format(iSize)
        return sBack, sTitle


class ReportEdit(BasicDetails):
    model = Report
    mForm = ReportEditForm
    prefix = "rpt"
    title = "ReportDetails"
    no_delete = True            # Don't allow users to remove a report
    mainitems = []

    def add_to_context(self, context, instance):
        """Add to the existing context"""

        # Define the main items to show and edit
        context['mainitems'] = [
            {'type': 'plain', 'label': "Created:",      'value': instance.get_created()         },
            {'type': 'line',  'label': "User:",         'value': instance.user.username         },
            {'type': 'line',  'label': "Report type:",  'value': instance.get_reptype_display() },
            # {'type': 'safe',  'label': "Download:",     'value': self.get_download_html(instance)},
            {'type': 'safe',  'label': "Raw data:",     'value': self.get_raw(instance)}
            ]

        # Signal that we do have select2
        context['has_select2'] = True

        # Return the context we have made
        return context

    def get_download_html(self, instance):
        """Get HTML representation of the report download buttons"""

        sBack = ""
        template_name = "seeker/report_download.html"
        oErr = ErrHandle()
        try:
            context = dict(report=instance)
            sBack = render_to_string(template_name, context, self.request)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("ReportEdit/get_download_html")
        return sBack

    def get_raw(self, instance):
        """Get HTML representation of the report details"""

        sBack = ""
        if instance.contents != None and instance.contents != "" and instance.contents[0] == "{":
            # There is a real report
            sContents = "-" if instance.contents == None else instance.contents
            sBack = "<textarea rows='1' style='width: 100%;'>{}</textarea>".format(sContents)
        return sBack


class ReportDetails(ReportEdit):
    """HTML output for a Report"""

    rtype = "html"

    def add_to_context(self, context, instance):
        context = super(ReportDetails, self).add_to_context(context, instance)

        context['after_details'] = self.get_download_html(instance)

        return context


class ReportDownload(BasicPart):
    MainModel = Report
    template_name = "seeker/download_status.html"
    action = "download"
    dtype = "csv"       # Download Type

    def custom_init(self):
        """Calculate stuff"""
        
        dt = self.qd.get('downloadtype', "")
        if dt != None and dt != '':
            self.dtype = dt

    def get_data(self, prefix, dtype, response=None):
        """Gather the data as CSV, including a header line and comma-separated"""

        # Initialize
        lData = []
        sData = ""
        oErr = ErrHandle()

        try:
            # Unpack the report contents
            sData = self.obj.contents

            if dtype == "json":
                # no need to do anything: the information is already in sData
                pass
            else:
                # Convert the JSON to a Python object
                oContents = json.loads(sData)
                # Get the headers and the list
                headers = oContents['headers']

                # Create CSV string writer
                output = StringIO()
                delimiter = "\t" if dtype == "csv" else ","
                csvwriter = csv.writer(output, delimiter=delimiter, quotechar='"')

                # Write Headers
                csvwriter.writerow(headers)

                # Two lists
                todo = [oContents['list'], oContents['read'] ]
                for lst_report in todo:

                    # Loop
                    for item in lst_report:
                        row = []
                        for key in headers:
                            if key in item:
                                element = item[key]
                                if isinstance(element, str):
                                    element = element.replace("\r", " ").replace("\n", " ")
                                row.append(element)
                            else:
                                row.append("")
                        csvwriter.writerow(row)

                # Convert to string
                sData = output.getvalue()
                output.close()
        except:
            msg = oErr.get_error_message()
            oErr.DoError("ReportDownload/get_data")

        return sData


# ================= SOURCE =============================

class SourceListView(BasicList):
    """Listview of sources"""

    model = SourceInfo
    listform = SourceEditForm
    has_select2 = True
    bUseFilter = True
    prefix = "src"
    new_button = False
    basic_name = "source"
    order_cols = ['created', 'collector', 'url', '']
    order_default = ['-created', 'collector', 'url']
    order_heads = [{'name': 'Date',           'order': 'o=1','type': 'str', 'custom': 'date', 'align': 'right', 'linkdetails': True},
                   {'name': 'Collector',      'order': 'o=2', 'type': 'str', 'field': 'collector', 'linkdetails': True},
                   {'name': 'Collected from', 'order': 'o=3', 'type': 'str', 'custom': 'from', 'main': True},
                   {'name': 'Manuscripts',    'order': '',    'type': 'int', 'custom': 'manucount'}]
    filters = [ {"name": "Collector",       "id": "filter_collector",      "enabled": False} ]
    searches = [
        {'section': '', 'filterlist': [
            {'filter': 'collector', 'fkfield': 'profile', 'keyS': 'profile_ta', 'keyFk': 'user__username', 'keyList': 'profilelist', 'infield': 'id'}
            ]}
         ]

    def initializations(self):
        # Remove SourceInfo's that are not tied to a Manuscript anymore
        remove_id = []
        for obj in SourceInfo.objects.all():
            m_count = obj.sourcemanuscripts.count()
            if m_count == 0:
                remove_id.append(obj.id)
        # Remove them
        if len(remove_id) > 0:
            SourceInfo.objects.filter(id__in=remove_id).delete()
        # Find out if any manuscripts need source info
        with transaction.atomic():
            for obj in Manuscript.objects.filter(source__isnull=True):
                # Get the snote info
                snote = obj.snote
                if snote != None and snote != "" and snote[0] == "[":
                    snote_lst = json.loads(snote)
                    if len(snote_lst)>0:
                        snote_first = snote_lst[0]
                        if 'username' in snote_first:
                            username = snote_first['username']
                            profile = Profile.get_user_profile(username)
                            created = obj.created
                            source = SourceInfo.objects.create(
                                code="Manually created",
                                created=created,
                                collector=username, 
                                profile=profile)
                            obj.source = source
                            obj.save()
        # Are there still manuscripts without source?
        if Manuscript.objects.filter(source__isnull=True).count() > 0:
            # Make the NEW button available
            self.new_button = True
        return None

    def get_field_value(self, instance, custom):
        sBack = ""
        sTitle = ""
        if custom == "date":
            sBack = instance.created.strftime("%d/%b/%Y %H:%M")
        elif custom == "from":
            if instance.url != None:
                sBack = instance.url
            elif instance.code != None:
                sBack = instance.code
        elif custom == "manucount":
            count_m = instance.sourcemanuscripts.filter(mtype='man').count()
            qs_t = instance.sourcemanuscripts.filter(mtype='tem')
            count_t = qs_t.count()
            if count_m == 0:
                if count_t == 0:
                    sBack = "&nbsp;"
                    sTitle = "No manuscripts are left from this source"
                elif count_t == 1:
                    # Get the id of the manuscript
                    manu = qs_t.first()
                    # Get the ID of the template with this manuscript
                    obj_t = Template.objects.filter(manu=manu).first()
                    if not obj_t is None:
                        url = reverse("template_details", kwargs={'pk': obj_t.id})
                        sBack = "<a href='{}' title='One template manuscript'><span class='badge jumbo-2 clickable'>{}</span></a>".format(
                            url, count_t)
                else:
                    url = reverse('template_list')
                    sBack = "<a href='{}?tmp-srclist={}' title='Template manuscripts'><span class='badge jumbo-1 clickable'>{}</span></a>".format(
                        url, instance.id, count_t)
            else:
                url = reverse('manuscript_list')
                sBack = "<a href='{}?manu-srclist={}' title='Manuscripts'><span class='badge jumbo-3 clickable'>{}</span></a>".format(
                    url, instance.id, count_m)
        return sBack, sTitle

    def add_to_context(self, context, initial):
        SourceInfo.init_profile()
        return context


class SourceEdit(BasicDetails):
    model = SourceInfo
    mForm = SourceEditForm
    prefix = 'source'
    prefix_type = "simple"
    basic_name = 'source'
    title = "SourceInfo"
    mainitems = []

    def add_to_context(self, context, instance):
        """Add to the existing context"""

        # Define the main items to show and edit
        context['mainitems'] = [
            {'type': 'plain', 'label': "Created:",      'value': instance.get_created()     },
            {'type': 'line',  'label': "Collector:",    'value': instance.get_username()    },
            {'type': 'line',  'label': "URL:",          'value': instance.url,              'field_key': 'url'  },
            {'type': 'line',  'label': "Code:",         'value': instance.get_code_html(),  'field_key': 'code' },
            {'type': 'safe',  'label': "Manuscript:",   'value': instance.get_manu_html(),  
             'field_list': 'manulist' }
            ]

        # Signal that we do have select2
        context['has_select2'] = True

        # Return the context we have made
        return context

    def before_save(self, form, instance):
        # Determine the user
        if self.request.user != None:
            profile = Profile.get_user_profile(self.request.user.username)
            form.instance.profile = profile
            # Check if a manuscript has been given
            manulist = form.cleaned_data.get('manulist', None)
            if manulist != None:
                #  manuscript has been added
                manu = Manuscript.objects.filter(id=manulist.id).first()
                if manu != None:
                    manu.source = instance
                    manu.save()
        return True, ""


class SourceDetails(SourceEdit):
    """The HTML variant of [SourceEdit]"""

    rtype = "html"


# =======================================================

class LitRefListView(ListView):
    """Listview of edition and literature references"""
       
    model = Litref
    paginate_by = 2000
    template_name = 'seeker/literature_list.html'
    entrycount = 0    
    # EK: nee dus, dit zijn geen projecten. plural_name = "Projects"
    
    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super(LitRefListView, self).get_context_data(**kwargs)

        # Get parameters
        initial = self.request.GET

        # Determine the count for literature references
        context['entrycount'] = self.entrycount # self.get_queryset().count()
        
        # Set the prefix
        context['app_prefix'] = APP_PREFIX

        # Make sure the paginate-values are available
        context['paginateValues'] = paginateValues

        if 'paginate_by' in initial:
            context['paginateSize'] = int(initial['paginate_by'])
        else:
            context['paginateSize'] = paginateSize

        # Set the title of the application
        context['title'] = "Passim literature info"

        # Change name of the qs for edition references 
        context['edition_list'] = self.get_editionset() 

        # Determine the count for edition references
        context['entrycount_edition'] = self.entrycount_edition

        # Change name of the qs for online references 
        context['online_list'] = self.create_onlineset() 

        # Determine the count for online references
        context['entrycount_online'] = self.entrycount_online

        # Check if user may upload
        context['is_authenticated'] = user_is_authenticated(self.request)
        context['is_app_uploader'] = user_is_ingroup(self.request, app_uploader)
        context['is_app_editor'] = user_is_ingroup(self.request, app_editor)

        # Process this visit and get the new breadcrumbs object
        prevpage = reverse('home')
        context['prevpage'] = prevpage
        context['breadcrumbs'] = get_breadcrumbs(self.request, "Literature references", True)

        # Return the calculated context
        return context

    def get_paginate_by(self, queryset):
        """
        Paginate by specified value in default class property value.
        """
        return self.paginate_by
    
    # Queryset literature references
    def get_queryset(self):
        # Get the parameters passed on with the GET or the POST request
        get = self.request.GET if self.request.method == "GET" else self.request.POST
        get = get.copy()
        self.get = get

        # Calculate the final qs for the manuscript litrefs
        litref_ids_man = [x['reference'] for x in LitrefMan.objects.all().values('reference')]
        # Calculate the final qs for the Gold sermon litrefs
        litref_ids_sg = [x['reference'] for x in LitrefSG.objects.all().values('reference')]

        # Combine the two qs into one and filter
        litref_ids = chain(litref_ids_man, litref_ids_sg)
        qs = Litref.objects.filter(id__in=litref_ids).order_by('sortref')

        # Determine the length
        self.entrycount = len(qs)

        # Return the resulting filtered and sorted queryset
        return qs
    
    def create_onlineset(self):
        #Get the parameters passed on with the GET or the POST request
        get = self.request.GET if self.request.method == "GET" else self.request.POST
        get = get.copy()
        self.get = get

        online_list=[]

        online_list.append
        
        # Calculate the final qs for the manuscript litrefs MODIFY!
        #ediref_ids = [x['reference'] for x in EdirefSG.objects.all().values('reference')]
       
        # Sort and filter all editions MODIFY
        #qs = Litref.objects.filter(id__in=ediref_ids).order_by('short')
        qs = OnlineSources.objects.order_by('name').order_by('sortname')

        # Determine the length
        self.entrycount_online = len(qs)

        # Return the resulting filtered and sorted queryset
        return qs

    # Queryset edition references
    def get_editionset(self):
        # Get the parameters passed on with the GET or the POST request
        get = self.request.GET if self.request.method == "GET" else self.request.POST
        get = get.copy()
        self.get = get

        # Calculate the final qs for the manuscript litrefs
        ediref_ids = [x['reference'] for x in EdirefSG.objects.all().values('reference')]
       
        # Sort and filter all editions
        qs = Litref.objects.filter(id__in=ediref_ids).order_by('short')

        # Determine the length
        self.entrycount_edition = len(qs)

        # Return the resulting filtered and sorted queryset
        return qs   
    

class BasketView(SermonListView):
    """Like SermonListView, but then with the basket set true"""
    basketview = True


class BasketViewManu(ManuscriptListView):
    """Like ManuscriptListView, but then with the basket set true"""
    basketview = True


class BasketViewGold(SermonGoldListView):
    """Like SermonGoldListView, but then with the basket set true"""
    basketview = True


class BasketViewSuper(EqualGoldListView):
    """Like EqualGoldListView, but then with the basket set true"""
    basketview = True


class BasketUpdate(BasicPart):
    """Update contents of the sermondescr basket"""

    MainModel = SermonDescr
    clsBasket = Basket
    template_name = "seeker/basket_choices.html"
    entrycount = 0
    bFilter = False
    s_view = SermonListView
    s_form = SermonForm
    s_field = "sermon"
    colltype = "sermo"
    form_objects = [{'form': CollectionForm, 'prefix': colltype, 'readonly': True}]

    def add_to_context(self, context):

        oErr = ErrHandle()
        try:
            # Check validity
            if not self.userpermissions("r"):
                # Don't do anything
                context['basketsize'] = 0
                context['basketuser'] = "no_user"
                context['basketwarning'] = "Sorry, only Passim users are allowed to use the baskets"
                return context

            # Reset the redirect page
            self.redirectpage = ""

            method = "use_profile_search_id_list"

            # Get the operation
            if 'operation' in self.qd:
                operation = self.qd['operation']
            else:
                return context

            username=self.request.user.username
            team_group=app_editor

            # Note: only operations in either of these two lists will be executed
            lst_basket_target = ["create", "add", "remove", "shared", "missing", "reset"]
            lst_basket_source = ["collcreate", "colladd", "rsetcreate", "rsetadd", "dctlaunch"]

            # Get our profile
            profile = Profile.get_user_profile(self.request.user.username)
            if profile != None:

                # Obligatory initialization
                rset = None

                # Get the queryset
                self.filters, self.bFilter, qs, ini, oFields = search_generic(self.s_view, self.MainModel, self.s_form, self.qd, username, team_group)

                # Check if *any* filter has been set

                # Action depends on operations
                if operation in lst_basket_target:
                    if method == "use_profile_search_id_list":
                        # Get the latest search results
                        search_s = getattr(profile, "search_{}".format(self.colltype))
                        search_id = []
                        if search_s != None and search_s != "" and search_s[0] == "[":
                            search_id = json.loads(search_s)
                        search_count = len(search_id)

                        kwargs = {'profile': profile}

                        # NOTE PROBLEM - we don't have the [oFields] at this point...

                        # Action depends on the operation specified
                        if search_count > 0 and operation == "create":
                            # Remove anything there
                            self.clsBasket.objects.filter(profile=profile).delete()
                            # Add
                            with transaction.atomic():
                                for item in search_id:
                                    kwargs["{}_id".format(self.s_field)] = item
                                    self.clsBasket.objects.create(**kwargs)
                            # Process history
                            profile.history(operation, self.colltype, oFields)
                        elif search_count > 0  and operation == "add":
                            # Add
                            with transaction.atomic():
                                for item in search_id:
                                    kwargs["{}_id".format(self.s_field)] = item
                                    obj = self.clsBasket.objects.filter(**kwargs).first()
                                    if obj == None:
                                        self.clsBasket.objects.create(**kwargs)
                            # Process history
                            profile.history(operation, self.colltype, oFields)
                        elif search_count > 0  and operation == "remove":
                            # Add
                            with transaction.atomic():
                                for item in search_id:
                                    kwargs["{}_id".format(self.s_field)] = item
                                    self.clsBasket.objects.filter(**kwargs).delete()
                            # Process history
                            profile.history(operation, self.colltype, oFields)
                        elif search_count > 0  and operation == "shared":
                            # What are the current id's?
                            current_ids = [x['id'] for x in self.clsBasket.objects.filter(**kwargs).values('id')]
                            # Filter out those who are allowed to stay...
                            kwargs["{}_id__in".format(self.s_field)] = search_id
                            # Find out the id's of basket items that may stay
                            maystay_ids = [x['id'] for x in self.clsBasket.objects.filter(**kwargs).values('id')]
                            # Remove all the id's from the basket, whose s_field is not in search_id
                            delete_id = []
                            for this_id in current_ids:
                                if not this_id in maystay_ids:
                                    delete_id.append(this_id)
                            # Possibly remove some
                            if len(delete_id) > 0:
                                self.clsBasket.objects.filter(id__in=delete_id).delete()

                            # Process history
                            profile.history(operation, self.colltype, oFields)
                        elif operation == "reset":
                            # Remove everything from our basket
                            self.clsBasket.objects.filter(profile=profile).delete()
                            # Reset the history for this one
                            profile.history(operation, self.colltype)

                    else:
                    
                        kwargs = {'profile': profile}

                        # Action depends on the operation specified
                        if qs and operation == "create":
                            # Remove anything there
                            self.clsBasket.objects.filter(profile=profile).delete()
                            # Add
                            with transaction.atomic():
                                for item in qs:
                                    kwargs[self.s_field] = item
                                    self.clsBasket.objects.create(**kwargs)
                            # Process history
                            profile.history(operation, self.colltype, oFields)
                        elif qs and operation == "add":
                            # Add
                            with transaction.atomic():
                                for item in qs:
                                    kwargs[self.s_field] = item
                                    self.clsBasket.objects.create(**kwargs)
                            # Process history
                            profile.history(operation, self.colltype, oFields)
                        elif qs and operation == "remove":
                            # Add
                            with transaction.atomic():
                                for item in qs:
                                    kwargs[self.s_field] = item
                                    self.clsBasket.objects.filter(**kwargs).delete()
                            # Process history
                            profile.history(operation, self.colltype, oFields)
                        elif operation == "reset":
                            # Remove everything from our basket
                            self.clsBasket.objects.filter(profile=profile).delete()
                            # Reset the history for this one
                            profile.history(operation, self.colltype)

                elif operation in lst_basket_source:
                    # Queryset: the basket contents
                    qs = self.clsBasket.objects.filter(profile=profile)

                    # Get the history string
                    history = getattr(profile, "history{}".format(self.colltype))

                    # New collection or existing one?
                    coll = None
                    bChanged = False
                    if operation == "collcreate":
                        # Save the current basket as a collection that needs to receive a name
                        # Note: this assumes [scope='priv'] default
                        if self.colltype != "super":
                            # Need to adapt the wording
                            adapted_word = dict(sermo="Manifestation", manu="Manuscript", gold="GoldSermon")
                            sWord = adapted_word.get(self.colltype, "Unknown")
                            coll = Collection.objects.create(path=history, settype="pd",
                                    descrip="Created from a {} listview basket".format(sWord), 
                                    owner=profile, type=self.colltype)
                        else:
                            # So this is only super = ssg = authority file
                            coll = Collection.objects.create(path=history, settype="pd",
                                    descrip="Created from an Authority file listview basket", 
                                    owner=profile, type=self.colltype)
                        # Assign it a name based on its ID number and the owner
                        if self.colltype != "super":
                            name = "{}_{}_{}".format(profile.user.username, coll.id, self.colltype)                         
                        else:
                            name = "{}_{}_{}".format(profile.user.username, coll.id, "af")
                        coll.name = name
                        coll.save()
                    elif operation == "rsetcreate":
                        # Need to adapt the wording
                        adapted_word = dict(sermo="Manifestation", manu="Manuscript", gold="GoldSermon", super="Authority File")
                        sWord = adapted_word.get(self.colltype, "Unknown")
                        # Save the current basket as a research-set that needs to receive a name
                        rset = ResearchSet.objects.create(
                            name="tijdelijk",
                            notes="Created from a {} listview basket".format(sWord),
                            profile=profile)
                        # Assign it a name based on its ID number and the owner
                        name = "{}_{}_{}".format(profile.user.username, rset.id, self.colltype)
                        rset.name = name
                        rset.save()
                    elif operation == "rsetadd":
                        # Get the rsetone parameter
                        if 'manuForm' in context:
                            manuForm = context['manuForm']
                            # Get the value: the ID
                            rsetone_id = manuForm['rsetone'].value()
                            if not rsetone_id is None or rsetone_id == "":
                                # Get the researchset
                                rset = ResearchSet.objects.filter(id=rsetone_id).first()
                                # Set parameters to show upon return
                                rseturl = reverse('researchset_details', kwargs={'pk': rset.id})
                                rsetname = rset.name
                                context['data'] = dict(collurl=rseturl, collname=rsetname)
                    elif operation == "dctlaunch":
                        # Need to adapt the wording
                        adapted_word = dict(sermo="Manifestation", manu="Manuscript", gold="GoldSermon", super="Authority File")
                        sWord = adapted_word.get(self.colltype, "Unknown")
                        # Save the current basket as a research-set that needs to receive a name
                        rset = ResearchSet.objects.create(
                            name="tijdelijk",
                            notes="Created from a {} listview basket for direct DCT launching".format(sWord),
                            profile=profile)
                        # Assign it a name based on its ID number and the owner
                        name = "{}_{}_{}".format(profile.user.username, rset.id, self.colltype)
                        rset.name = name
                        rset.save()
                    elif oFields['collone']:
                        coll = oFields['collone']

                    # Process the basket elements into the ResearchSet or into the Collection
                    if rset != None:
                        with transaction.atomic():
                            for idx, item in enumerate(qs):
                                # Check if it doesn't exist yet
                                obj = SetList.objects.filter(researchset=rset, manuscript=item.manu).first()
                                if obj == None:
                                    # Create this
                                    order = idx + 1
                                    SetList.objects.create(researchset=rset, 
                                                           order = order,
                                                           setlisttype="manu",
                                                           manuscript=item.manu)

                        # Make sure to redirect to this instance -- but only for RSETCREATE, RSETADD and DCTLAUNCH
                        if operation in ["rsetcreate"]:
                            self.redirectpage = reverse('researchset_details', kwargs={'pk': rset.id})
                        elif operation == "dctlaunch":
                            # Get the default DCT for this ad-hoc ResearchSet
                            dct = rset.researchset_setdefs.first()
                            self.redirectpage = reverse('setdef_details', kwargs={'pk': dct.id})
                    elif coll == None:
                        # TODO: provide some kind of error??
                        pass
                    else:
                        # Link the collection with the correct model
                        kwargs = {'collection': coll}
                        if self.colltype == "sermo":
                            clsColl = CollectionSerm
                            field = "sermon"
                        elif self.colltype == "gold":
                            clsColl = CollectionGold
                            field = "gold"
                        elif self.colltype == "manu":
                            clsColl = CollectionMan
                            field = "manuscript"
                        elif self.colltype == "super":
                            clsColl = CollectionSuper
                            field = "super"

                        # THis is only needed for collections
                        with transaction.atomic():
                            for item in qs:
                                kwargs[field] = getattr( item, self.s_field)
                                # Check if it doesn't exist yet
                                obj = clsColl.objects.filter(**kwargs).first()
                                if obj == None:
                                    clsColl.objects.create(**kwargs)
                                    # Note that some changes have been made
                                    bChanged = True

                        # Make sure to redirect to this instance -- but only for COLLCREATE
                        if operation == "collcreate":
                            self.redirectpage = reverse('collpriv_details', kwargs={'pk': coll.id})
                        else:
                            # We are adding to an existing Collecion that is either public or private (or 'team' in scope)
                            if coll.settype == "pd":
                                if coll.scope == "publ":
                                    # Public dataset
                                    urltype = "publ"
                                else:
                                    # Team or Priv
                                    urltype = "priv"
                            elif coll.settype == "hc":
                                urltype = "hist"
                            collurl = reverse('coll{}_details'.format(urltype), kwargs={'pk': coll.id})
                            collname = coll.name
                            context['data'] = dict(collurl=collurl, collname=collname)
                            # Have changes been made?
                            if bChanged:
                                # Add the current basket history to the collection's path
                                lst_history_basket = json.loads(history)
                                lst_history_coll = json.loads(coll.path)
                                for item in lst_history_basket:
                                    lst_history_coll.append(item)
                                coll.path = json.dumps(lst_history_coll)
                                coll.save()

                # Adapt the basket size
                context['basketsize'] = self.get_basketsize(profile)

                # Set the other context parameters
                if self.colltype == "sermo":
                    context['basket_show'] = reverse('basket_show' )
                    context['basket_update'] = reverse('basket_update')
                else:
                    context['basket_show'] = reverse('basket_show_{}'.format(self.colltype))
                    context['basket_update'] = reverse('basket_update_{}'.format(self.colltype))
                context['colltype'] = self.colltype

        except:
            msg = oErr.get_error_message()
            oErr.DoError("BasketUpdate")

        # Return the updated context
        return context

    def get_basketsize(self, profile):
        # Adapt the basket size
        basketsize = profile.basketitems.count()
        profile.basketsize = basketsize
        profile.save()
        # Return the basketsize
        return basketsize

 
class BasketUpdateManu(BasketUpdate):
    """Update contents of the manuscript basket"""

    MainModel = Manuscript
    clsBasket = BasketMan 
    s_view = ManuscriptListView
    s_form = SearchManuForm
    s_field = "manu"
    colltype = "manu"
    form_objects = [{'form': CollectionForm, 'prefix': colltype, 'readonly': True}]

    def get_basketsize(self, profile):
        # Adapt the basket size
        basketsize = profile.basketitems_manu.count()
        profile.basketsize_manu = basketsize
        profile.save()
        # Return the basketsize
        return basketsize


class BasketUpdateGold(BasketUpdate):
    """Update contents of the sermondescr basket"""

    MainModel = SermonGold
    clsBasket = BasketGold
    s_view = SermonGoldListView
    s_form = SermonGoldForm
    s_field = "gold"
    colltype = "gold"
    form_objects = [{'form': CollectionForm, 'prefix': colltype, 'readonly': True}]

    def get_basketsize(self, profile):
        # Adapt the basket size
        basketsize = profile.basketitems_gold.count()
        profile.basketsize_gold = basketsize
        profile.save()
        # Return the basketsize
        return basketsize
    

class BasketUpdateSuper(BasketUpdate):
    """Update contents of the EqualGold basket"""

    MainModel = EqualGold
    clsBasket = BasketSuper
    s_view = EqualGoldListView
    s_form = SuperSermonGoldForm
    s_field = "super"
    colltype = "super"
    form_objects = [{'form': CollectionForm, 'prefix': colltype, 'readonly': True}]

    def get_basketsize(self, profile):
        # Adapt the basket size
        basketsize = profile.basketitems_super.count()
        profile.basketsize_super = basketsize
        profile.save()
        # Return the basketsize
        return basketsize
