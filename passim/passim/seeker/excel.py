from passim.utils import ErrHandle
from passim.settings import WRITABLE_DIR, MEDIA_DIR

import io, sys, os
import openpyxl
from openpyxl.utils.cell import get_column_letter
from openpyxl.cell import Cell
from openpyxl import Workbook
from io import StringIO

def excel_to_list(data, filename, lExpected = None, lField = None):
    """Read an excel file into a list of objects

    This assumes that the first row contains column headers
    """

    oErr = ErrHandle()
    bResult = True
    lData = []
    msg = ""
    try:
        # Write data temporarily to the WRITABLE dir, but with a temporary filename
        tmp_path = os.path.abspath(os.path.join( MEDIA_DIR, filename))
        with io.open(tmp_path, "wb") as f:
            sData = data.read()
            f.write(sData)

        # Read string file
        wb = openpyxl.load_workbook(tmp_path, read_only=False)
        ws = wb.active

        # Iterate through rows
        bFirst = True
        
        lHeader = []
        if lExpected == None or lField == None:
            # Cannot handle this
            msg = "Please specify lExpected and lField"
            return False, [], msg

        # Iterate
        for row in ws.iter_rows(min_row=1, min_col=1):
            if bFirst:
                # Expect header
                for cell in row:
                    sValue = cell.value.strip().lower()                    
                    sKey = ""
                    for idx, item in enumerate(lExpected):
                        if item in sValue:
                            sKey = lField[idx]
                            break
                    # Check if it's okay
                    if sKey == "":
                        # Cannot read this
                        msg = "Don't understand column header [{}]".format(sValue)
                        return False, [], msg
                    lHeader.append(sKey)
                bFirst = False
            else:
                oRow = {}
                for idx, key in enumerate(lHeader):
                    cell = row[idx]
                    # Get the value
                    oRow[key] = cell.value.strip()
                lData.append(oRow)
        # Close the workbook
        wb.close()

        # Remove the file
        os.remove(tmp_path)
        # Return positively
        bResult = True
    except:
        # Note the error here
        msg = oErr.get_error_message()
        bResult = False
        oErr.DoError("excel_to_list")

    # Return what we have found
    return bResult, lData, msg
