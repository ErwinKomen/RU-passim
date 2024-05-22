
from django.apps import apps
from django.contrib import admin
from django.contrib.auth import login, authenticate
from django.contrib.auth.models import Group, User
from django.urls import reverse
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db import transaction
from django.db.models import Q, Prefetch, Count, F
from django.db.models.functions import Lower
from django.db.models.query import QuerySet 
from django.forms import formset_factory, modelformset_factory, inlineformset_factory, ValidationError
from django.forms.models import model_to_dict
from django.http import HttpRequest, HttpResponse, HttpResponseRedirect, JsonResponse, FileResponse
from django.shortcuts import get_object_or_404, render, redirect
from django.template.loader import render_to_string
from django.template import Context
from django.views.generic.detail import DetailView
from django.views.generic.base import RedirectView
from django.views.generic import ListView, View
from django.views.decorators.csrf import csrf_exempt

# General imports
import io, sys, os, re
import openpyxl, json
import copy
from openpyxl.utils.cell import get_column_letter
from openpyxl.cell import Cell
from openpyxl import Workbook
from io import StringIO
import csv
import requests

# ======= imports from my own application ======
from passim.settings import APP_PREFIX, MEDIA_DIR
from passim.utils import ErrHandle

from passim.seeker.models import Manuscript, SermonDescr, Profile, Report, Codico, Location, LocationType, Library, \
    SermonGoldExternal, EqualGoldExternal
from passim.seeker.views import app_editor
from passim.reader.views import ReaderImport
from passim.reader.forms import UploadFileForm



class ManuscriptUploadExcel(ReaderImport):
    """Specific parameters for importing manuscripts from Excel"""

    import_type = "excel"
    sourceinfo_url = "https://www.ru.nl/english/people/boodts-s/"
    # Note: set manucreate to [True], if a new manuscript must be created
    #       instead of adding to the existing manuscript
    manucreate = False # True

    def process_files(self, request, source, lResults, lHeader):
        file_list = []
        oErr = ErrHandle()
        bOkay = True
        code = ""
        oStatus = self.oStatus
        try:
            # Make sure we have the username
            username = self.username
            profile = Profile.get_user_profile(username)
            team_group = app_editor
            kwargs = {'profile': profile, 'username': username, 'team_group': team_group}

            # Get the contents of the imported file
            files = request.FILES.getlist('files_field')
            if files != None:
                lst_err = []
                for data_file in files:
                    filename = data_file.name
                    file_list.append(filename)

                    # Set the status
                    oStatus.set("reading", msg="file={}".format(filename))

                    # Get the source file
                    if data_file == None or data_file == "":
                        self.arErr.append("No source file specified for the selected project")
                    else:
                        # Check the extension
                        arFile = filename.split(".")
                        extension = arFile[len(arFile)-1]

                        lst_manual = []
                        lst_read = []

                        # Further processing depends on the extension
                        oResult = {'status': 'ok', 'count': 0, 'sermons': 0, 'msg': "", 'user': username}

                        if extension == "xlsx":
                            # This is an Excel file: read the file using openpyxl
                            # Write data temporarily to the WRITABLE dir, but with a temporary filename
                            tmp_path = os.path.abspath(os.path.join( MEDIA_DIR, filename))
                            with io.open(tmp_path, "wb") as f:
                                sData = data_file.read()
                                f.write(sData)

                            # Read string file
                            wb = openpyxl.load_workbook(tmp_path, read_only=True)
                            sheetnames = wb.sheetnames
                            ws_manu = None
                            ws_sermo = None
                            ws_codico = None
                            lst_ws = []
                            for sname in sheetnames:
                                if "manu" in sname.lower():
                                    ws_manu = wb[sname]
                                    lst_ws.append(ws_manu)
                                elif "codico" in sname.lower():
                                    ws_codico = wb[sname]
                                    lst_ws.append(ws_codico)
                                elif "sermo" in sname.lower():
                                    ws_sermo = wb[sname]
                                    lst_ws.append(ws_sermo)

                            # Check the contents of JSON coded cells
                            err_json = []
                            for ws in lst_ws:
                                row_num = 1
                                for value_tuple in ws.iter_rows():
                                    # ------------ DEBUGGING -----------------------
                                    print("File={}, sheet={}, row={}".format(filename, ws.title, row_num))
                                    # ----------------------------------------------
                                    col_num = 1
                                    for cell in value_tuple:
                                        value = cell.value
                                        if not value is None and isinstance(value, str) and value[0] == "[":
                                            # This should be a json list
                                            try:
                                                lst_this = json.loads(value)
                                            except:
                                                # This one is bad...
                                                oBad = dict(sheet=ws.title, row=row_num, column=col_num, coordinate=cell.coordinate, value=value)
                                                err_json.append(oBad)
                                        col_num += 1

                                    row_num += 1

                            if len(err_json) > 0:
                                # We cannot process this one
                                html = []
                                html.append("There were errors in the Excel {}:".format(filename))
                                html.append("<table><thead><tr><th>Sheet</th><th>Cell</th><th>Value</th></tr></thead><tbody>")
                                for oBad in err_json:
                                    html.append("<tr><td>{}</td><td>{}</td><td>{}</td></tr>".format(
                                        oBad['sheet'], oBad['coordinate'], oBad['value']))
                                html.append("</tbody></table>")
                                code = "\n".join(html)
                                lst_err.append(code)

                            else:
                                # Do we have a manuscript worksheet?
                                if ws_manu != None:
                                    # Process the manuscript-proper details: columns Name and Value
                                    oManu = {}
                                    row_num = 1
                                    if ws_manu.cell(row=row_num, column=1).value.lower() == "field" and \
                                        ws_manu.cell(row=row_num, column=2).value.lower() == "value":
                                        # we can skip the first row
                                        row_num += 1
                                    bStop = False
                                    while not bStop:
                                        k = ws_manu.cell(row=row_num, column=1).value
                                        v = ws_manu.cell(row=row_num, column=2).value
                                        if k == "" or k == None:
                                            bStop = True
                                        else:
                                            row_num += 1
                                            k = k.lower()
                                            oManu[k] = v

                                    params = {}

                                    # We have an object with key/value pairs: process it
                                    kwargs['manucreate'] = self.manucreate
                                    manu = Manuscript.custom_add(oManu, params, **kwargs)

                                    # Process codicological unit - if it has been provided in the Excel
                                    dict_codico = {}
                                    if ws_codico is None:
                                        # No codicological info has been provided...

                                        # Now get the codicological unit that has been automatically created and adapt it
                                        codico = manu.manuscriptcodicounits.first()
                                        if codico != None:
                                            oManu['manuscript'] = manu
                                            codico = Codico.custom_add(oManu, **kwargs)
                                            # Make sure there is a link between the codico order and the codico object
                                            dict_codico['1'] = codico

                                    else:
                                        # We have codicological information!!!
                                        # Get the column names
                                        row_num = 1
                                        column = 1
                                        header = []
                                        v = ws_codico.cell(row=row_num, column=column).value
                                        while v != None and v != "" and v != "-":                                        
                                            header.append(v.lower())
                                            column += 1
                                            v = ws_codico.cell(row=row_num, column=column).value
                                        # Process the codico units in this sheet
                                        codico_list = []
                                        column = 1

                                        # Iterate over the rows
                                        for row in ws_codico.iter_rows():
                                            row_values = [x.value for x in row]
                                            v = row_values[0]
                                            if row_num > 1 and not v is None and v!= "":
                                                # Create a new codico object
                                                oCodico = {}
                                                # Process this row
                                                for idx, col_name in enumerate(header):
                                                    column = idx + 1
                                                    oCodico[col_name] = row_values[idx]
                                                # Process this sermon
                                                order = oCodico['order']
                                                oCodico['manuscript'] = manu
                                                codico = Codico.custom_add(oCodico, **kwargs)

                                                # Make sure there is a link between the codico order and the codico object
                                                dict_codico[str(order)] = codico
                                            # GO to the next row for the next sermon
                                            row_num += 1

                                    oResult['count'] += 1
                                    oResult['obj'] = manu
                                    oResult['name'] = manu.idno

                                    # Check if there is a "Sermon" worksheet
                                    if ws_sermo != None:
                                        # Get the column names
                                        row_num = 1
                                        column = 1
                                        header = []
                                        v = ws_sermo.cell(row=row_num, column=column).value
                                        while v != None and v != "" and v != "-":                                        
                                            header.append(v.lower())
                                            column += 1
                                            v = ws_sermo.cell(row=row_num, column=column).value
                                        # Process the sermons in this sheet
                                        sermon_list = []
                                        column = 1

                                        # Iterate over the rows
                                        for row in ws_sermo.iter_rows():
                                            row_values = [x.value for x in row]
                                            v = None if len(row_values) ==0 else row_values[0]
                                            if row_num > 1 and not v is None and v!= "":
                                                # ==== DEBUG ====
                                                oErr.Status("Upload excel row_num={}".format(row_num))
                                                # ===============
                                                
                                                # Create a new sermon object
                                                oSermon = {}
                                                # Process this row
                                                for idx, col_name in enumerate(header):
                                                    column = idx + 1
                                                    oSermon[col_name] = row_values[idx]

                                                # Get the codico that should be used
                                                codi_order = str(oSermon.get("codico", 1))
                                                codico = dict_codico[codi_order]
                                                # Process this sermon
                                                order = oSermon['order']
                                                sermon = SermonDescr.custom_add(oSermon, manu, codico, order, **kwargs)

                                                oResult['sermons'] += 1

                                                # Get parent, firstchild, next
                                                parent = oSermon['parent']
                                                firstchild = oSermon['firstchild']
                                                nextone = oSermon['next']
                                                # Add to list
                                                sermon_list.append({'order': order, 'parent': parent, 'firstchild': firstchild,
                                                                    'next': nextone, 'sermon': sermon})
                                            # GO to the next row for the next sermon
                                            row_num += 1

                                        # Now process the parent/firstchild/next items
                                        with transaction.atomic():
                                            for oSermo in sermon_list:
                                                # Get the p/f/n numbers
                                                parent_id = oSermo['parent']
                                                firstchild_id = oSermo['firstchild']
                                                next_id = oSermo['next']
                                                # Process parent
                                                if parent_id != '' and parent_id != None:
                                                    # parent_id = str(parent_id)
                                                    parent = next((obj['sermon'] for obj in sermon_list if obj['order'] == parent_id), None)
                                                    if not parent is None:
                                                        oSermo['sermon'].msitem.parent = parent.msitem
                                                        oSermo['sermon'].msitem.save()
                                                # Process firstchild
                                                if firstchild_id != '' and firstchild_id != None:
                                                    # firstchild_id = str(firstchild_id)
                                                    firstchild = next((obj['sermon'] for obj in sermon_list if obj['order'] == firstchild_id), None)
                                                    if not firstchild is None:
                                                        oSermo['sermon'].msitem.firstchild = firstchild.msitem
                                                        oSermo['sermon'].msitem.save()
                                                # Process next
                                                if next_id != '' and next_id != None:
                                                    # next_id = str(next_id)
                                                    nextone = next((obj['sermon'] for obj in sermon_list if obj['order'] == next_id), None)
                                                    if not nextone is None:
                                                        oSermo['sermon'].msitem.next = nextone.msitem
                                                        oSermo['sermon'].msitem.save()


                        # Create a report and add it to what we return
                        oContents = {'headers': lHeader, 'list': lst_manual, 'read': lst_read}
                        oReport = Report.make(username, "ixlsx", json.dumps(oContents))
                                
                        # Determine a status code
                        statuscode = "error" if oResult == None or oResult['status'] == "error" else "completed"
                        if oResult == None:
                            self.arErr.append("There was an error. No manuscripts have been added")
                        else:
                            lResults.append(oResult)

                if len(lst_err) > 0:
                    # code = json.dumps(lst_err, indent=2)
                    code = "\n".join(lst_err)
                    bOkay = False
            if code == "":
                code = "Imported using the [import_excel] function on these files: {}".format(", ".join(file_list))
        except:
            bOkay = False
            code = oErr.get_error_message()
        return bOkay, code


class ManuscriptUploadJson(ReaderImport):
    import_type = "json"
    sourceinfo_url = "https://www.ru.nl/passim/upload_manu_json"

    def process_files(self, request, source, lResults, lHeader):

        def sermones_fixes(oManu):
            """Several fixes for importing sermones - but they are generic
            
            1 - If there is literature on the Sermon level, copy that to the manu level
            2 - If there is a manuscript identifier like [ms.381], then add a space after the period
            3 - Change the project from "Passim" into None, so that the default projects are assigned
            """

            oErr = ErrHandle()
            bResult = True
            re_manuidno = re.compile( r'^[Mm][Ss]\.[0-9].*')
            try:
                msitems = oManu['msitems']
                if not msitems is None and len(msitems) > 0:
                    msitem = msitems[0]

                    # Fix 1: literature
                    oSermon = msitem.get("sermon")
                    if not oSermon is None:
                        # Does this sermon have any literature
                        sLiterature = oSermon.get("literature")
                        if not sLiterature is None:
                            # Check out the manuscript level
                            if "literature" in oManu:
                                literatures = oManu['literature']
                                iCount = len(literatures)
                                if iCount == 1:
                                    if len(literatures[0]) == 0:
                                        # Replace
                                        literatures[0] = sLiterature
                                    else:
                                        # Add it
                                        literatures.append(sLiterature)
                                else:
                                    # Just add it
                                    literatures.append(sLiterature)

                # Fix 2: space after ms.
                idno = oManu.get("idno")
                if re_manuidno.match(idno):
                    # Repair the first part of the idno
                    idno = "Ms. {}".format(idno[3:])
                    oManu['idno'] = idno

                # Fix 3: Remove project specification
                project = oManu.get("project")
                if not project is None:
                    oManu['project'] = None
                # Okay, no problems
            except:
                msg = oErr.get_error_message()
                oErr.DoError("sermones_fixes")
                bResult = False
            return bResult

        file_list = []
        oErr = ErrHandle()
        bOkay = True
        code = ""
        sourcetype = ""
        oStatus = self.oStatus

        try:
            # Make sure we have the username
            username = self.username
            profile = Profile.get_user_profile(username)
            team_group = app_editor
            kwargs = {'profile': profile, 'username': username, 'team_group': team_group, 'keyfield': 'path', 'source': source}

            # Get the contents of the imported file
            files = request.FILES.getlist('files_field')
            if files != None:
                for data_file in files:
                    filename = data_file.name
                    file_list.append(filename)

                    # Set the status
                    oStatus.set("reading", msg="file={}".format(filename))

                    # Get the source file
                    if data_file == None or data_file == "":
                        self.arErr.append("No source file specified for the selected project")
                    else:
                        # Check the extension
                        arFile = filename.split(".")
                        extension = arFile[len(arFile)-1]

                        lst_manual = []
                        lst_read = []

                        # Further processing depends on the extension
                        oResult = {'status': 'ok', 'count': 0, 'sermons': 0, 'msg': "", 'user': username}

                        lst_msg = []

                        if extension == "json":
                            # This is a JSON file: Load the file into a variable
                            sData = data_file.read()
                            lst_manu = json.loads( sData.decode(encoding="utf8"))

                            # Check if this is a dictionary or a list
                            if isinstance(lst_manu, dict):
                                # It is a dictionary: turn it into a list
                                oManuList = lst_manu
                                sorted_keys = sorted(oManuList.keys(), key=lambda x: int(re.search(r'\d+', x).group()))
                                lst_manu = [oManuList[x] for x in sorted_keys]
                                # lst_manu = [v for k,v in sorted(oManuList.items(), key=lambda key: int(re.search(r'\d+', key).group()))]

                            # Look at the first manuscript to determine any specific source type
                            count_manu = len(lst_manu)
                            if count_manu > 0:
                                oManuFirst = lst_manu[0]
                                datasets = oManuFirst.get("datasets")
                                if not datasets is None and len(datasets) > 0:
                                    sDataset = datasets[0].lower()
                                    if "sermones" in sDataset:
                                        sourcetype = "sermones"
                                    elif "huwa" in sDataset:
                                        sourcetype = "huwa"

                            # Make sure we pass the sourcetype on to Manuscript.custom_add()
                            kwargs['sourcetype'] = sourcetype

                            # Walk through the manuscripts
                            for idx, oManu in enumerate(lst_manu):
                                # There is one result per manuscript
                                oResult = {'status': 'ok', 'count': 0, 'sermons': 0, 'msg': "", 'user': username}


                                # Issue #509: literature fix
                                if sourcetype == "sermones":
                                    sermones_fixes(oManu)
                                # Make sure the stype is set to "imported"
                                oManu['stype'] = "imp"

                                # Each manuscript has some stuff of its own
                                # We have an object with key/value pairs: process it
                                params = dict(overwriting = False)
                                manu = Manuscript.custom_add(oManu, params, **kwargs)

                                # Show where we are
                                if params['overwriting']:
                                    # Skip, do not overwrite
                                    oErr.Status("{}/{}: {} - NOT OVERWRITING".format(idx+1, count_manu, manu.idno))

                                    # The only things we look at are: editornotes, date, and list of sermons

                                    # (1) date
                                    daterange = oManu.get("date")
                                    codico = manu.manuscriptcodicounits.first()
                                    if not daterange is None or daterange == "":
                                        if codico.codico_dateranges.count() == 0:
                                            # Get the date
                                            codico.add_one_daterange(daterange)

                                    # (2) list of sermons
                                    html_en = []
                                    editornotes = manu.editornotes
                                    if not editornotes is None: 
                                        html_en.append(editornotes)
                                    msitems = oManu.get("msitems")
                                    if not msitems is None and len(msitems) > 0:
                                        sermon_siglist = []
                                        for idx, msitem in enumerate(msitems):
                                            oSermon = msitem.get("sermon")
                                            sig = oSermon.get("signaturesA")
                                            if not sig is None:
                                                sermon_siglist.append("sermon {}: {}".format(idx+1, json.dumps(sig)))
                                        sSiglist = ";\n".join(sermon_siglist)
                                        html_en.append("Sermons: {}".format(sSiglist))

                                    # (3) Editornotes
                                    en = oManu.get("editornotes")
                                    if not en is None:
                                        html_en.append(en)

                                    # COmbine the editor notes
                                    manu.editornotes = "\n\nHUWA sermon information:".join(html_en)
                                    manu.save()

                                    oResult['count'] += 1
                                    oResult['obj'] = manu
                                    oResult['name'] = manu.idno
                                    oResult['msg'] = "NOT OVERWRITING this manuscript"

                                    # Append this result
                                    lResults.append(oResult)
                                else:
                                    oErr.Status("{}/{}: {}".format(idx+1, count_manu, manu.idno))

                                    if 'msg' in params:
                                        lst_msg.append(params['msg'])

                                    # Now get the codicological unit that has been automatically created and adapt it
                                    codico = manu.manuscriptcodicounits.first()
                                    if codico != None:
                                        oManu['manuscript'] = manu
                                        codico = Codico.custom_add(oManu, **kwargs)

                                    oResult['count'] += 1
                                    oResult['obj'] = manu
                                    oResult['name'] = manu.idno

                                    # Process all the MsItems into a list of sermons
                                    sermon_list = []
                                    for oMsItem in oManu['msitems']:
                                        # Get the sermon object
                                        oSermon = oMsItem['sermon']
                                        order = oMsItem['order']

                                        # Make sure the stype is set to "imported"
                                        oSermon['stype'] = "imp"

                                        sermon = SermonDescr.custom_add(oSermon, manu, codico, order, **kwargs)

                                        # Keep track of the number of sermons read
                                        oResult['sermons'] += 1

                                        # Some sourcetype specific processing of SERMON


                                        # Get parent, firstchild, next
                                        parent = oMsItem['parent']
                                        firstchild = oMsItem['firstchild']
                                        nextone = oMsItem['next']

                                        # Add to list
                                        sermon_list.append({'order': order, 'parent': parent, 'firstchild': firstchild,
                                                            'next': nextone, 'sermon': sermon})

                                    # Now process the parent/firstchild/next items
                                    with transaction.atomic():
                                        for oSermo in sermon_list:
                                            # Get the p/f/n numbers
                                            parent_id = oSermo['parent']
                                            firstchild_id = oSermo['firstchild']
                                            next_id = oSermo['next']
                                            # Process parent
                                            if parent_id != '' and parent_id != None:
                                                # parent_id = str(parent_id)
                                                parent = next((obj['sermon'] for obj in sermon_list if obj['order'] == parent_id), None)
                                                oSermo['sermon'].msitem.parent = parent.msitem
                                                oSermo['sermon'].msitem.save()
                                            # Process firstchild
                                            if firstchild_id != '' and firstchild_id != None:
                                                # firstchild_id = str(firstchild_id)
                                                firstchild = next((obj['sermon'] for obj in sermon_list if obj['order'] == firstchild_id), None)
                                                oSermo['sermon'].msitem.firstchild = firstchild.msitem
                                                oSermo['sermon'].msitem.save()
                                            # Process next
                                            if next_id != '' and next_id != None:
                                                # next_id = str(next_id)
                                                nextone = next((obj['sermon'] for obj in sermon_list if obj['order'] == next_id), None)
                                                oSermo['sermon'].msitem.next = nextone.msitem
                                                oSermo['sermon'].msitem.save()

                                    # Append this result
                                    lResults.append(oResult)

                                # Prepare a 'read' item
                                # Fields: ['status', 'msg', 'name', 'yearstart', 'yearfinish', 'library', 'idno', 'filename', 'url']
                                yearstart = -1
                                yearfinish = -1
                                daterange = codico.codico_dateranges.first()
                                if not daterange is None:
                                    yearstart = daterange.yearstart
                                    yearfinish = daterange.yearfinish
                                library = manu.get_library()
                                idno = manu.idno
                                msg = "read" if not params['overwriting'] else "overwriting"
                                oRead = dict(status="ok", msg=msg, name="-", 
                                                yearstart=yearstart, yearfinish=yearfinish,
                                                library=library, idno=idno, filename="-", url="-")
                                lst_read.append(oRead)

                            # Possibly show messages
                            if len(lst_msg) > 0:
                                oErr.Status("Messages:\n{}".format(json.dumps(lst_msg, indent=2)))
                                # And write it to a dump file
                                sFile = os.path.abspath(os.path.join(MEDIA_DIR, "passim", "manuscriptuploadjson.json"))
                                with open(sFile, "w", encoding="utf-8") as f:
                                    json.dump(lst_msg, f, indent=2)
                                oErr.Status("   See file: {}".format(sFile))

                        # Create a report and add it to what we return                        
                        oContents = {'headers': lHeader, 'list': lst_manual, 'read': lst_read, 'msg': lst_msg}
                        oReport = Report.make(username, "ijson", json.dumps(oContents))
                                
                        # Determine a status code
                        statuscode = "error" if oResult == None or oResult['status'] == "error" else "completed"
                        if oResult == None:
                            self.arErr.append("There was an error. No manuscripts have been added")
                        else:
                            lResults.append(oResult)

            code = "Imported using the [import_json] function on this file list: {}".format(", ".join(file_list))
        except:
            bOkay = False
            code = oErr.get_error_message()
        return bOkay, code


class ManuscriptUploadGalway(ReaderImport):
    """Specific parameters for importing manuscripts from https://elmss.nuigalway.ie/"""

    import_type = "nuigalway"
    sourceinfo_url = "https://elmss.nuigalway.ie/"

    def process_files(self, request, source, lResults, lHeader):
        file_list = []
        oErr = ErrHandle()
        bOkay = True
        code = ""
        oStatus = self.oStatus
        # The list of headers to be shown
        lHeader = ['status', 'msg', 'name', 'daterange', 'library', 'idno', 'url']

        def add_manu(lst_manual, lst_read, status="", msg="", user="", name="", url="", daterange="", 
                     library="", filename="", sermons="", idno=""):
            oInfo = {}
            oInfo['status'] = status
            oInfo['msg'] = msg
            oInfo['user'] = user
            oInfo['name'] = name
            oInfo['url'] = url
            oInfo['daterange'] = daterange
            oInfo['library'] = library
            oInfo['idno'] = idno
            oInfo['filename'] = filename
            oInfo['sermons'] = sermons
            if status == "error":
                lst_manual.append(oInfo)
            else:
                lst_read.append(oInfo)
            return True

        try:
            # Make sure we have the username
            username = self.username
            profile = Profile.get_user_profile(username)
            team_group = app_editor
            kwargs = {'profile': profile, 'username': username, 'team_group': team_group, 'source': source}

            # Get the contents of the imported file
            files = request.FILES.getlist('files_field')
            if files != None:
                for data_file in files:
                    filename = data_file.name
                    file_list.append(filename)

                    # Set the status
                    oStatus.set("reading", msg="file={}".format(filename))

                    # Get the source file
                    if data_file == None or data_file == "":
                        self.arErr.append("No source file specified for the selected project")
                    else:
                        # Check the extension
                        arFile = filename.split(".")
                        extension = arFile[len(arFile)-1]

                        lst_manual = []
                        lst_read = []

                        # Further processing depends on the extension
                        oResult = {'status': 'ok', 'count': 0, 'sermons': 0, 'msg': "", 'user': username, 'filename': filename}

                        if extension == "csv":
                            # This is a CSV file. We expect the catalogue id's to be in the leftmost column

                            # Write data temporarily to the WRITABLE dir, but with a temporary filename
                            tmp_path = os.path.abspath(os.path.join( MEDIA_DIR, filename))
                            with io.open(tmp_path, "wb") as f:
                                sData = data_file.read()
                                f.write(sData)

                            # Read the CSV file with a reader
                            with open(tmp_path, "r", encoding="utf-8") as f:
                                reader = csv.reader(f, delimiter=",", dialect='excel')
                                # Read the header cells and make a header row in the worksheet
                                headers = next(reader)
                                row_num = 1
                                column = 1
                                lCsv = []
                                for row in reader:
                                    # Keep track of the EXCEL row we are in
                                    row_num += 1
                                    # Get the ID
                                    cell_value = row[0]
                                    if cell_value != None and cell_value != "":
                                        # Get the catalogue id
                                        catalogue_id = int(cell_value)

                                        # Set the status
                                        oStatus.set("reading", msg="catalogue id={}".format(catalogue_id))

                                        # Clear the galway and codico objects
                                        oGalway = None
                                        oCodico = None

                                        # Read the manuscript object from the Galway site
                                        oGalway = self.get_galway(catalogue_id)
                                        # Extract Manuscript information and Codico information from [oGalway]
                                        oManu, oCodico = self.get_manucodico(oGalway)

                                        if oManu != None and oCodico != None:
                                            libname = "{}, {}, {}".format(oManu['country_name'], oManu['city_name'], oManu['library_name'])

                                            # Add manuscript (if not yet there)
                                            params = {}
                                            manu = Manuscript.custom_add(oManu, params, **kwargs)

                                            if manu.library == None:
                                                # Log that the library is not recognized
                                                oErr.Status("Library not recognized: {}".format(libname))
                                                # Also add this in the notes
                                                notes = "" if manu.notes == None else manu.notes
                                                manu.notes = "Library not found: {}  \n{}".format(libname, notes)

                                            # Make sure to add the source and the RAW data
                                            manu.source = source
                                            manu.raw = json.dumps(oGalway, indent=2)
                                            manu.save()

                                            # Now get the codicological unit that has been automatically created and adapt it
                                            codico = manu.manuscriptcodicounits.first()
                                            if codico != None:
                                                oCodico['manuscript'] = manu
                                                codico = Codico.custom_add(oCodico, **kwargs)

                                            # Process results
                                            add_manu(lst_manual, lst_read, status=oResult['status'], user=oResult['user'],
                                                            name=codico.name, daterange=oCodico['date ranges'],
                                                            library=libname, filename=manu.idno, sermons=0,
                                                            idno=manu.idno)

                                            oResult['count'] += 1
                                            #oResult['obj'] = manu
                                            #oResult['name'] = manu.idno

                                            oResultManu = dict(name=manu.idno, filename=oManu['url'], sermons=0)
                                            lResults.append(oResultManu)


                        # Create a report and add it to what we return
                        oContents = {'headers': lHeader, 'list': lst_manual, 'read': lst_read}
                        oReport = Report.make(username, "xlsx", json.dumps(oContents))
                                
                        # Determine a status code
                        statuscode = "error" if oResult == None or oResult['status'] == "error" else "completed"
                        if oResult == None:
                            self.arErr.append("There was an error. No manuscripts have been added")
                        else:
                            lResults.append(oResult)
            
            # Make sure we have a success message available
            code = "Imported using the [import_galway] function on this file: {}".format(", ".join(file_list))

            # Indicate we are ready
            oStatus.set("ready")
        except:
            bOkay = False
            code = oErr.get_error_message()
        return bOkay, code

    def get_galway(self, catalogue_id):
        # Read the manuscript with the indicated id

        oBack = None
        bResult = False
        oErr = ErrHandle()
        try:
            url = "https://elmss.nuigalway.ie/api/v1/catalogue/{}".format(catalogue_id)
            try:
                r = requests.get(url)
            except:
                sMsg = oErr.get_error_message()
                oErr.DoError("Request problem")
                return False, sMsg
            if r.status_code == 200:
                # Read the response
                sText = r.text
                oBack = json.loads(sText)
                bResult = True
            else:
                bResult = False
                sResult = "download_file received status {} for {}".format(r.status_code, url)
                oErr.Status("get_galway reading error: {}".format(sResult))
        except:
            msg = oErr.get_error_message()
            oErr.DoError("get_galway")
            oBack = None

        return oBack

    def get_manucodico(self, oGalway):
        # Convert a Galway object into a processable Manu and Codico object

        oManu = None
        oCodico = None
        oErr = ErrHandle()
        try:
            # Initialisations
            oManu = {}
            oCodico = {}
            oGalwayItem = oGalway['item']
            city = None
            country = None
            library = None

            loctype_country = LocationType.objects.filter(name="country").first()
            loctype_city= LocationType.objects.filter(name="city").first()

            # Start with admin notes
            oManu['notes'] = oGalwayItem['admin_notes']

            # Check for shelf marks
            shelf_history = ""
            shelfmark = None
            oLibrary = None
            for oShelfmark in oGalwayItem['shelfmarks']:
                shelfmark = oShelfmark['shelfmark']
                oLibrary = oShelfmark['library']
                if oLibrary != "" and oLibrary != None:
                    country = oLibrary['country']
                    city = oLibrary['city']
                    library = oLibrary['library']
                    shelfmark = "{}, {}, {}, {}".format(country, city, library, shelfmark)
                shelf_history = "{}  \nshelfmark: {}".format(shelf_history, shelfmark)
            shelfcount = len(oGalwayItem['shelfmarks'])
            if shelfcount > 1:
                # Add the history to notes
                oManu['notes'] = "{}  \n{}".format(oManu['notes'], shelf_history)

            # Load the values of the *LAST* shelfmark
            if shelfmark != None and oLibrary != None:
                # Set the final shelf mark
                oManu['shelf mark'] = shelfmark

                # derive country/city/library
                oManu['country_name'] = oLibrary['country']
                oManu['city_name'] = oLibrary['city']
                oManu['library_name'] = oLibrary['library']

                # Try get the country id
                country = Location.objects.filter(loctype=loctype_country, name__iexact=oManu['country_name']).first()
                if country != None: 
                    oManu['country id'] = country.id

                    # Try to get the city, restricted to the country
                    city = Location.objects.filter(loctype=loctype_city,
                                                   name__iexact=oManu['city_name'],
                                                   lcountry=country).first()
                    if city != None:
                        oManu['city id'] = city.id

                        # Try to get the library, restricted to city
                        library = Library.objects.filter(name__iexact = oManu['library_name'],
                                                          lcountry=country, lcity=city).first()
                        if library != None:
                            oManu['library id'] = library.id

            oManu['notes'] = "{}  \nScript commentary: {}".format(oManu['notes'],
               oGalwayItem['script_commentary'])
            oManu['url'] = oGalwayItem['public_url']
            oManu['external id'] = oGalwayItem['id']
            lExternal = []
            dataurl = oManu.get('url')
            facsimile = oGalwayItem.get('facsimile_url')
            if dataurl != None and dataurl != "": lExternal.append(dataurl)
            if facsimile != None and facsimile != "": lExternal.append(facsimile)
            oManu['external links'] = lExternal

            # Load the codico values
            oCodico['title'] = oGalwayItem['contents']
            oCodico['date ranges'] = "{}-{}".format(oGalwayItem['numerical_date_start'], oGalwayItem['numerical_date_end'])
            oCodico['support'] = oGalwayItem['support']
            oCodico['provenances'] = oGalwayItem['provenance']

            # Note: extent and format are not provided:
            #oCodico['extent'] = oGalwayItem['']
            #oCodico['format'] = oGalwayItem['']

            # Debugging
            provenances_txt = oCodico['provenances']
            # oErr.Status("document [{}] provenances: {}".format(oManu['shelf mark'], oCodico['provenances']))
        except:
            msg = oErr.get_error_message()
            oErr.DoError("get_manucodico")
            oManu = None

        return oManu, oCodico


class LibraryUploadExcel(ReaderImport):
    """Specific parameters for importing manuscripts from Excel"""

    import_type = "excel"
    sourceinfo_url = "https://www.ru.nl/passim/upload_library_excel"
    template_name = "reader/import_libraries.html"

    def process_files(self, request, source, lResults, lHeader):

        file_list = []
        oErr = ErrHandle()
        bOkay = True
        code = ""
        oStatus = self.oStatus
        try:
            # Make sure we have the username
            username = self.username
            profile = Profile.get_user_profile(username)
            team_group = app_editor
            kwargs = {'profile': profile, 'username': username, 'team_group': team_group, 'keyfield': 'path'}

            # Get the contents of the imported file
            files = request.FILES.getlist('files_field')
            if files != None:
                for data_file in files:
                    filename = data_file.name
                    file_list.append(filename)

                    # Set the status
                    oStatus.set("reading", msg="file={}".format(filename))

                    # Get the source file
                    if data_file == None or data_file == "":
                        self.arErr.append("No source file specified for the selected project")
                    else:
                        # Check the extension
                        arFile = filename.split(".")
                        extension = arFile[len(arFile)-1]

                        lst_manual = []
                        lst_read = []
                        lst_added = []

                        # Further processing depends on the extension
                        oResult = {'status': 'ok', 'count': 0, 'sermons': 0, 'msg': "", 'user': username}

                        lHeader.clear()
                        lHeader.append("Country")
                        lHeader.append("City")
                        lHeader.append("Library")

                        if extension == "xlsx":
                            # This is an Excel file: read the file using openpyxl
                            # Write data temporarily to the WRITABLE dir, but with a temporary filename
                            tmp_path = os.path.abspath(os.path.join( MEDIA_DIR, filename))
                            with io.open(tmp_path, "wb") as f:
                                sData = data_file.read()
                                f.write(sData)

                            # Read string file
                            wb = openpyxl.load_workbook(tmp_path, read_only=True)
                            sheetnames = wb.sheetnames
                            ws_data = None
                            for sname in sheetnames:
                                if "data" in sname.lower():
                                    ws_data = wb[sname]
                                    break
                            # Do we have a 'Data' worksheet?
                            if not ws_data is None:
                                # Process the header details
                                oHeader = dict(lcountry=0, lcity=0, library=0, toevoegen=0, opmerkingen=0, approved=0)
                                row_num = 1
                                col_num = 1
                                while not ws_data.cell(row=row_num, column=col_num).value is None:
                                    sKey = ws_data.cell(row=row_num, column=col_num).value
                                    if sKey in oHeader:
                                        oHeader[sKey] = col_num
                                    col_num += 1

                                # Double check to see if we now have all the columns needed
                                bMissing = False
                                for k,v in oHeader.items():
                                    if v is None or v == 0:
                                        bMissing = True
                                        break

                                # Can we continue?
                                if bMissing:
                                    # We cannot continue: informatino is missing
                                    bOkay = False
                                    code = "Missing header"
                                else:
                                    # Yes, find out how many libraries there are right now
                                    lib_count = Library.objects.count()
                                    loc_count = Location.objects.count()

                                    # Yes, we can continue: walk through all rows
                                    row_num = 2
                                    while not ws_data.cell(row=row_num, column=1).value is None:
                                        # Process this row
                                        lcountry = ws_data.cell(row=row_num, column=oHeader['lcountry']).value
                                        lcity = ws_data.cell(row=row_num, column=oHeader['lcity']).value
                                        library = ws_data.cell(row=row_num, column=oHeader['library']).value
                                        toevoegen = ws_data.cell(row=row_num, column=oHeader['toevoegen']).value
                                        opmerkingen = ws_data.cell(row=row_num, column=oHeader['opmerkingen']).value
                                        approved = ws_data.cell(row=row_num, column=oHeader['approved']).value

                                        if not approved is None and approved == "j":
                                            # Add library
                                            sNote = "Menna approved LibraryUploadExcel"
                                            added, lib_id, lib_city, lib_country = Location.get_or_create_library(library, lcity, lcountry, sNote)

                                            # Has this been added?
                                            if added:
                                                # Then put it into a list
                                                oAdded = dict(country=lcountry, city=lcity, library=library)
                                                lResults.append(oAdded)

                                        # Go to the next row
                                        row_num += 1
 
                                    # Check what has been added
                                    lib_added = Library.objects.count() - lib_count
                                    loc_added = Location.objects.count() - loc_count



                        # Create a report and add it to what we return
                        
                        oContents = {'headers': lHeader, 'libraries': lib_added, 'locations': loc_added,
                                     'added': lResults}
                        oReport = Report.make(username, "ixlsx", json.dumps(oContents, indent=2))
                                
                        # Determine a status code
                        statuscode = "error" if oResult == None or oResult['status'] == "error" else "completed"

            code = "Imported using the [library_upload_excel] function on this file list: {}".format(", ".join(file_list))
        except:
            bOkay = False
            code = oErr.get_error_message()
        return bOkay, code


class SermonGoldUploadJson(ReaderImport):
    """Specifically upload EDILIT Json table into SG"""

    import_type = "json"
    sourceinfo_url = "https://www.ru.nl/passim/upload_sg_json"

    def process_files(self, request, source, lResults, lHeader):

        def get_edi_string(oEdi):
            """Turn the edition object into a string in a unified way"""

            html = []
            for key in sorted(oEdi):
                # Do *NOT* take the key 'edition' into account
                if key == "edition":
                    continue
                # Do look at other k/v pairs
                v = oEdi[key]
                if isinstance(v,int) or isinstance(v, str):
                    html.append("{}:{}".format(key,v))
                elif key == "author":
                    html.append("{}:{}".format(key,v['full']))
                # NOTE: do not consider lists and other keys
            sBack = ",".join(html)
            return sBack

        file_list = []
        oErr = ErrHandle()
        bOkay = True
        code = ""
        sourcetype = ""
        oStatus = self.oStatus

        try:
            # Make sure we have the username
            username = self.username
            profile = Profile.get_user_profile(username)
            team_group = app_editor
            kwargs = {'profile': profile, 'username': username, 'team_group': team_group, 'keyfield': 'path', 'source': source}

            # Get the contents of the imported file
            files = request.FILES.getlist('files_field')
            if files != None:
                for data_file in files:
                    filename = data_file.name
                    file_list.append(filename)

                    # Set the status
                    oStatus.set("reading", msg="file={}".format(filename))

                    # Get the source file
                    if data_file == None or data_file == "":
                        self.arErr.append("No source file specified for the selected project")
                    else:
                        # Check the extension
                        arFile = filename.split(".")
                        extension = arFile[len(arFile)-1]

                        lst_manual = []
                        lst_read = []

                        # Further processing depends on the extension
                        oResult = {'status': 'ok', 'count': 0, 'sermons': 0, 'msg': "", 'user': username}


                        if extension == "json":
                            # This is a JSON file: Load the file into a variable
                            sData = data_file.read()
                            lst_edilit = json.loads( sData.decode(encoding="utf8"))

                            # Check if this is a dictionary or a list
                            if isinstance(lst_edilit, dict):
                                # It is a dictionary: turn it into a list
                                oEdiList = lst_edilit
                                sorted_keys = sorted(oEdiList.keys(), key=lambda x: int(re.search(r'\d+', x).group()))
                                lst_edilit = [oEdiList[x] for x in sorted_keys]

                            # Look at the first manuscript to determine any specific source type
                            count_edilit = len(lst_edilit)

                            # Make sure we pass the sourcetype on to Manuscript.custom_add()
                            kwargs['sourcetype'] = sourcetype

                            # Walk through the Edition elements
                            for idx, oEdiLit in enumerate(lst_edilit):
                                # Show where we are
                                if idx % 100 == 0:
                                    oErr.Status("SermonGoldUploadJson editions: {}/{}".format(idx+1, count_edilit))

                                # There is one result per manuscript
                                oResult = {'status': 'ok', 'count': 0, 'edition': 0, 'msg': "", 'user': username}

                                edition_id = oEdiLit.get("edition")
                                opera_id = oEdiLit.get("opera")
                                sEdiLit = get_edi_string(oEdiLit)
                                # Find the correct element, the gold
                                qs = SermonGoldExternal.objects.filter(externalid=opera_id)
                                # Process all the gold items in it
                                for obj in qs:
                                    gold = obj.gold
                                    # Get the current edition stuff
                                    lst_edi = []
                                    if not gold.edinote is None:
                                        lst_edi = json.loads(gold.edinote)
                                    # Check if the [edition_id] is already in there
                                    bFound = False
                                    for oItem in lst_edi:
                                        #if edition_id == oItem.get("edition"):
                                        #    # The editions are the same, but are the objects really equal??
                                        #    sItem = get_edi_string(oItem)
                                        #    if sItem == sEdiLit:
                                        #        bFound = True
                                        sItem = get_edi_string(oItem)
                                        if sItem == sEdiLit:
                                            bFound = True

                                    if not bFound:
                                        # This edition needs to be added
                                        lst_edi.append(copy.copy(oEdiLit))
                                        # Save the object
                                        gold.edinote = json.dumps(lst_edi, indent=2)
                                        gold.save()

                                        msg = "-"
                                        oRead = dict(status="ok", msg=msg, name="-", 
                                                        edition=edition_id, opera=opera_id,
                                                        gold=gold.id, filename="-", url="-")
                                        lst_read.append(oRead)
    
    
                        # Create a report and add it to what we return
                        
                        oContents = {'headers': lHeader, 'list': lst_manual, 'read': lst_read}
                        oReport = Report.make(username, "ijson", json.dumps(oContents))
                                
                        # Determine a status code
                        statuscode = "error" if oResult == None or oResult['status'] == "error" else "completed"
                        if oResult == None:
                            self.arErr.append("There was an error. No manuscripts have been added")
                        else:
                            lResults.append(oResult)


            code = "Imported using the [import_json] function on this file list: {}".format(", ".join(file_list))
        except:
            bOkay = False
            code = oErr.get_error_message()
        return bOkay, code


class EqualGoldUploadEdilit(ReaderImport):
    """Specifically upload EDILIT Json table into SSG"""

    import_type = "json"
    sourceinfo_url = "https://www.ru.nl/passim/upload_ssgedilit_json"

    def process_files(self, request, source, lResults, lHeader):

        def get_edi_string(oEdi):
            """Turn the edition object into a string in a unified way"""

            html = []
            for key in sorted(oEdi):
                # Do *NOT* take the key 'edition' into account
                if key == "edition":
                    continue
                # Do look at other k/v pairs
                v = oEdi[key]
                if isinstance(v,int) or isinstance(v, str):
                    html.append("{}:{}".format(key,v))
                elif key == "author":
                    html.append("{}:{}".format(key,v['full']))
                # NOTE: do not consider lists and other keys
            sBack = ",".join(html)
            return sBack

        file_list = []
        oErr = ErrHandle()
        bOkay = True
        code = ""
        sourcetype = ""
        oStatus = self.oStatus

        try:
            # Make sure we have the username
            username = self.username
            profile = Profile.get_user_profile(username)
            team_group = app_editor
            kwargs = {'profile': profile, 'username': username, 'team_group': team_group, 'keyfield': 'path', 'source': source}

            # Get the contents of the imported file
            files = request.FILES.getlist('files_field')
            if files != None:
                for data_file in files:
                    filename = data_file.name
                    file_list.append(filename)

                    # Set the status
                    oStatus.set("reading", msg="file={}".format(filename))

                    # Get the source file
                    if data_file == None or data_file == "":
                        self.arErr.append("No source file specified for the selected project")
                    else:
                        # Check the extension
                        arFile = filename.split(".")
                        extension = arFile[len(arFile)-1]

                        lst_manual = []
                        lst_read = []

                        # Further processing depends on the extension
                        oResult = {'status': 'ok', 'count': 0, 'sermons': 0, 'msg': "", 'user': username}


                        if extension == "json":
                            # This is a JSON file: Load the file into a variable
                            sData = data_file.read()
                            lst_edilit = json.loads( sData.decode(encoding="utf8"))

                            # Check if this is a dictionary or a list
                            if isinstance(lst_edilit, dict):
                                # It is a dictionary: turn it into a list
                                oEdiList = lst_edilit
                                sorted_keys = sorted(oEdiList.keys(), key=lambda x: int(re.search(r'\d+', x).group()))
                                lst_edilit = [oEdiList[x] for x in sorted_keys]

                            # Look at the first manuscript to determine any specific source type
                            count_edilit = len(lst_edilit)

                            # Make sure we pass the sourcetype on to Manuscript.custom_add()
                            kwargs['sourcetype'] = sourcetype

                            # Walk through the Edition elements
                            for idx, oEdiLit in enumerate(lst_edilit):
                                # Show where we are
                                if idx % 100 == 0:
                                    oErr.Status("EqualGoldUploadEdilit editions: {}/{}".format(idx+1, count_edilit))

                                # There is one result per manuscript
                                oResult = {'status': 'ok', 'count': 0, 'edition': 0, 'msg': "", 'user': username}

                                edition_id = oEdiLit.get("edition")
                                opera_id = oEdiLit.get("opera")
                                sEdiLit = get_edi_string(oEdiLit)

                                if edition_id > 500 and edition_id < 507:
                                    iStop = 1

                                # Find the correct element, the SSG
                                qs = EqualGoldExternal.objects.filter(externalid=opera_id)
                                # Process all the SSG items in it
                                for obj in qs:
                                    equal = obj.equal
                                    # Get the current edition stuff
                                    lst_edi = []
                                    if not equal.edinote is None:
                                        lst_edi = json.loads(equal.edinote)
                                    # Check if the [edition_id] is already in there
                                    bFound = False
                                    for oItem in lst_edi:
                                        #if edition_id == oItem.get("edition"):
                                        #    # The editions are the same, but are the objects really equal??
                                        #    sItem = get_edi_string(oItem)
                                        #    if sItem == sEdiLit:
                                        #        bFound = True
                                        sItem = get_edi_string(oItem)
                                        if sItem == sEdiLit:
                                            bFound = True

                                    if not bFound:
                                        # This edition needs to be added
                                        lst_edi.append(copy.copy(oEdiLit))
                                        # Save the object
                                        equal.edinote = json.dumps(lst_edi, indent=2)
                                        equal.save()

                                        msg = "-"
                                        oRead = dict(status="ok", msg=msg, name="-", 
                                                        edition=edition_id, opera=opera_id,
                                                        equal=equal.id, filename="-", url="-")
                                        lst_read.append(oRead)
    
    
                        # Create a report and add it to what we return
                        
                        oContents = {'headers': lHeader, 'list': lst_manual, 'read': lst_read}
                        oReport = Report.make(username, "ijson", json.dumps(oContents))
                                
                        # Determine a status code
                        statuscode = "error" if oResult == None or oResult['status'] == "error" else "completed"
                        if oResult == None:
                            self.arErr.append("There was an error. No manuscripts have been added")
                        else:
                            lResults.append(oResult)


            code = "Imported using the [import_json] function on this file list: {}".format(", ".join(file_list))
        except:
            bOkay = False
            code = oErr.get_error_message()
        return bOkay, code

